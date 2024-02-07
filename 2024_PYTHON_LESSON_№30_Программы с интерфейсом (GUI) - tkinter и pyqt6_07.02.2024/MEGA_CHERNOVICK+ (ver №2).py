# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
class Fruit:
    __color = ''
    __weight = 50

    def __init__(self, a_weight):
        self.__weight = a_weight

    def setWeight(self, a_weight):
        if a_weight <= 0:
            self.__weight = 0
        else:
            self.__weight = a_weight

    def getWeight(self):
        # if self.__weight < 100: return 0
        return self.__weight

    def __del__(self):
        print('Object is about to be destoyed')

apple = Fruit(150)
pineapple = Fruit(470)
melon = Fruit(5000)

apple.setWeight(-120)
print(apple.getWeight())

# конец алгоритма
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #

Lesson18 - 22.12.2023

# задача: система учета зоомагазина
# животные: кошки и собаки, * рыбки, птички
# сотрудники: менеджер, * ветеринар, бухгалтер, управленец
# вольеры: место временного пребывания животных

class Cat:
    # name, age, breed, is_vacinated
    # public, protected _, private __
    def __init__(self, a_name, a_age, a_breed, a_isvac):
        self.name = a_name
        self.__age = a_age
        self.breed = a_breed
        self.is_vacinated = a_isvac

    def setAge(self, a_age):
        if 250 > a_age > 0:
            self.__age = a_age

    def getAge(self):
        return self.__age

    def incAge(self, n=1):
        self.__age += n

    def makeSound(self):
        print('meow')

class Dog:
    def __init__(self, a_name, a_age, a_breed, a_isvac, a_size, a_cage=None):
        self.name = a_name
        self.__age = a_age
        self.breed = a_breed
        self.is_vacinated = a_isvac
        self.size = a_size
        self.__cage = a_cage

    def makeSound(self):
        print('woof woof')

    def setCage(self, a_cage):
        self.__cage = a_cage

    def getCage(self):
        return self.__cage

    def getAge(self):
        if self.__age > 1:
            return self.__age * 10

class Cage:
    def __init__(self, a_id, a_size, a_ismovable, a_animal=None):
        self.id = a_id
        self.size = a_size
        self.is_movable = a_ismovable
        self.__animal = a_animal

    def setAnimal(self, a_animal):
        self.__animal = a_animal

    def getAnimal(self):
        return self.__animal


barky = Dog('ГавГав', 2, 'Дворняжка', True, 'M')
cage_1 = Cage(1, 'S', True)

barky.setCage(cage_1)
cage_1.setAnimal(barky)

print(cage_1.getAnimal())
print(cage_1.getAnimal().name)
print(cage_1.getAnimal().getCage())
print(cage_1.getAnimal().getCage().getAnimal())
print(cage_1.getAnimal().getCage().getAnimal().getCage())
print(barky.getCage())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
""""
Дата выполнения практической работы: 22-23 - ДЕКАБРЯ 2023
""""
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 22.12.2023
Практическая работа №18: ООП. Полиморфизм
'''
'''
Выполните следующие задания: # ЕСЛИ ХОТИТЕ ТО СРАЗУ ПЕРЕХОДИТЕ НА 1120 СТРОЧКУ
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''' # ЕСЛИ ХОТИТЕ ТО СРАЗУ ПЕРЕХОДИТЕ НА 1120 СТРОЧКУ
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задание № 1

В качестве практической работы попробуйте самостоятельно перегрузить оператор сложения.

Для его перегрузки используется метод __add__(). 
Он вызывается, когда объекты класса, имеющего данный метод, фигурируют в операции сложения, причем с левой стороны.
Это значит, что в выражении a + b у объекта a должен быть метод __add__().
Объект b может быть чем угодно, но чаще всего он бывает объектом того же класса. 
Объект b будет автоматически передаваться в метод __add__() в качестве второго аргумента (первый – self).

Отметим, в Python также есть правосторонний метод перегрузки сложения - __radd__().
Согласно полиморфизму ООП, возвращать метод __add__() может что угодно. 
Может вообще ничего не возвращать, а "молча" вносить изменения в какие-то уже существующие объекты. 
Допустим, в вашей программе метод перегрузки сложения будет возвращать новый объект того же класса.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №2

Пересмотрите алгоритм решения прошлой практической работы, с использованием инкапсуляции.
Реализуйте старый алгоритм с использованием полиморфизма.

Написать программу, в которой есть главный класс Games со статическим полем Year, опишите конструктор присваивающий
значение полю Year, также опишите метод getName, который возвращает имя игры. На основе главного класса путем 
наследования опишите четыре класса PCGames, PS4Games, XboxGames, MobileGames. Добавьте каждому классу дополнительные 
поля и переопределите у всех классов метод getName.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №3

Пересмотрите алгоритм решения прошлой практической работы, с использованием инкапсуляции.
Реализуйте старый алгоритм с использованием полиморфизма.

Напишите программу с пустым классом Country. Опишите наследуемые от класса Country классы Russia, Canada, Germany.
Добавьте каждому классу поле population и опишите метод setPopulation и getPopulation.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 1 - Вариант № 1
'''
class MyClass:
    def __init__(self, value):
        self.value = value

    def __add__(self, other):
        if isinstance(other, MyClass):
            # Если other того же класса, создаем новый объект с суммой значений
            new_value = self.value + other.value
            return MyClass(new_value)
        else:
            # Иначе выводим сообщение об ошибке
            print("Нельзя сложить объекты разных классов")


# Создаем два объекта
obj1 = MyClass(5)
obj2 = MyClass(10)

# Пробуем сложить их
result = obj1 + obj2

# Выводим результат
print(result.value)  # Выведет: 15
'''
Пояснение:

В данном примере при сложении объектов класса MyClass создается новый объект с атрибутом value, равным сумме 
значений атрибута value слагаемых объектов. Обратите внимание на использование метода isinstance, чтобы убедиться,
что второй операнд того же класса.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1. Определение класса MyClass и его конструктора (__init__):
'''

class MyClass:
    def __init__(self, value):
        self.value = value
'''
Описание:

MyClass - это определение класса, который мы создаем.
__init__ - это конструктор класса, который вызывается при создании нового объекта. Он принимает два аргумента: 
self (представляющий создаваемый объект) и value (значение, которое будет установлено в атрибут value объекта).
'''
'''
Шаг №2. Перегрузка оператора сложения (__add__):
'''
def __add__(self, other):
    if isinstance(other, MyClass):
        new_value = self.value + other.value
        return MyClass(new_value)
    else:
        print("Нельзя сложить объекты разных классов")
'''
Описание:

__add__ - это метод, который перегружает оператор сложения (+). 
Он принимает два аргумента: self (представляющий вызываемый объект) и other (другой объект, с которым выполняется
сложение).
Внутри метода проверяется, является ли other объектом того же класса (MyClass).
Если да, то создается новый объект класса MyClass с атрибутом value, равным сумме значений атрибута value текущего
и переданного объектов. Этот новый объект затем возвращается.
Если other не принадлежит к классу MyClass, выводится сообщение об ошибке.
'''
'''
Шаг №3. Создание объектов и использование перегруженного оператора сложения:
'''
# Создаем два объекта
obj1 = MyClass(5)
obj2 = MyClass(10)
# Пробуем сложить их
result = obj1 + obj2
# Выводим результат
print(result.value)  # Выведет: 15
'''
Описание:

Создаются два объекта класса MyClass с атрибутами value равными 5 и 10 соответственно.
Оператор сложения + применяется к объектам obj1 и obj2. Так как мы перегрузили оператор сложения для класса MyClass,
вызывается метод __add__.
Метод __add__ выполняет сложение значений атрибута value двух объектов и создает новый объект с результатом.
Результат выводится на экран, и он равен 15, что является суммой значений атрибутов объектов obj1 и obj2.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 1 - Вариант № 2
'''
class Vector:
    def __init__(self, x, y):
        self.x = x
        self.y = y

    def __add__(self, other):
        if isinstance(other, Vector):
            # Если other тоже объект класса Vector, создаем новый объект Vector
            new_x = self.x + other.x
            new_y = self.y + other.y
            return Vector(new_x, new_y)
        else:
            # Иначе выводим сообщение об ошибке
            print("Нельзя сложить объекты разных классов")

    def __str__(self):
        return f"Vector({self.x}, {self.y})"


# Пример использования
v1 = Vector(1, 2)
v2 = Vector(3, 4)

# Перегрузка оператора сложения
result_vector = v1 + v2

# Вывод результата
print(result_vector)  # Выведет: Vector(4, 6)
'''
Пояснение:

В этом примере создается класс Vector для представления двумерных векторов.
Метод __add__ перегружает оператор сложения так, что он создает новый объект Vector,
содержащий сумму соответствующих координат. Метод __str__ используется для определения строкового представления
объекта Vector при его выводе.

Этот код так-же демонстрирует, как перегрузить оператор сложения для конкретного класса
(в данном случае, для векторов), что позволяет более естественно и интуитивно использовать операцию сложения
для объектов этого класса.
'''
'''
Шаг №1. Определение класса Vector и его конструктора (__init__):
'''
class Vector:
    def __init__(self, x, y):
        self.x = x
        self.y = y


'''
Описание:

Vector - это класс, представляющий двумерные векторы.
__init__ - конструктор класса, который вызывается при создании нового объекта. 
Принимает три аргумента: self (представляющий создаваемый объект), x и y (координаты вектора).
'''
'''
Шаг №2. Перегрузка оператора сложения (__add__):
'''
def __add__(self, other):
    if isinstance(other, Vector):
        new_x = self.x + other.x
        new_y = self.y + other.y
        return Vector(new_x, new_y)
    else:
        print("Нельзя сложить объекты разных классов")

'''
Описание:

__add__ - метод, который перегружает оператор сложения (+). 
Принимает два аргумента: self (текущий объект) и other (другой объект, с которым выполняется сложение).
Проверяет, является ли other объектом того же класса (Vector). Если да, то создает новый объект Vector с координатами,
равными сумме соответствующих координат текущего и переданного объектов. Этот новый объект затем возвращается.
Если other не принадлежит к классу Vector, выводится сообщение об ошибке.
'''
'''
Шаг №3. Метод для строкового представления объекта (__str__):
'''


def __str__(self):
    return f"Vector({self.x}, {self.y})"

'''
Описание:

__str__ - метод, который возвращает строковое представление объекта при вызове функции str().
В данном случае, он возвращает строку, представляющую объект Vector в виде "Vector(x, y)".
'''
'''
Шаг №4. Пример использования класса:
'''
# Пример использования
v1 = Vector(1, 2)
v2 = Vector(3, 4)

# Перегрузка оператора сложения
result_vector = v1 + v2

# Вывод результата
print(result_vector)  # Выведет: Vector(4, 6)
'''
Описание:

Создаются два объекта класса Vector с координатами (1, 2) и (3, 4).
Оператор сложения + применяется к объектам v1 и v2. Так как мы перегрузили оператор сложения для класса Vector,
вызывается метод __add__.
Метод __add__ создает новый объект Vector с координатами (4, 6) (сумма соответствующих координат v1 и v2).
Результат выводится на экран с использованием метода __str__.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 1 - Вариант № 3
'''
class MyClass:
    """Пример класса с перегруженным оператором сложения."""

    def __init__(self, value):
        """Инициализация объекта."""
        self.value = value

    def __add__(self, other):
        """Перегрузка оператора сложения."""
        try:
            if isinstance(other, MyClass):
                new_value = self.value + other.value
                return MyClass(new_value)
        except AttributeError:
            pass
        return None

    def __str__(self):
        """Строковое представление объекта."""
        return f"MyClass({self.value})"

'''
Пояснение:

В этом коде try блок используется для выполнения операции сложения, а except блок будет выполнен в случае,
если произойдет исключение (например, AttributeError, если other не имеет атрибута value).

Однако, как говорит GOOGLE стоит помнить, что использование try и except для контроля логики программы,
а не для обработки исключений, может сделать код менее предсказуемым и сложным для отладки.
Поэтому лучше использовать такие конструкции с осторожностью и в ситуациях,
где они действительно улучшают читаемость и логику кода.
'''
'''
Шаг №1. Инициализация объекта:
'''
def __init__(self, value):
    self.value = value

'''
Описание:

Название: __init__ - это конструктор класса, который вызывается при создании нового объекта.
Выполнение функции: Метод инициализирует атрибут value объекта значением, переданным при создании объекта.
'''
'''
Шаг №2. Перегрузка оператора сложения:
'''
def __add__(self, other):
    try:
        if isinstance(other, MyClass):
            new_value = self.value + other.value
            return MyClass(new_value)
    except AttributeError:
        pass
    return None

'''
Описание:

Название: __add__ - перегружает оператор сложения (+).

Выполнение функции:
В блоке try, проверяется, является ли other объектом класса MyClass.
Если это так, то создается новый объект MyClass с атрибутом value, равным сумме значений атрибутов value текущего и
переданного объектов.
В блоке except AttributeError, обрабатывается случай, если у объекта other нет атрибута value.
Возвращается новый объект MyClass или None, если other не является объектом MyClass.
'''
'''
Шаг №3. Строковое представление объекта:
'''

def __str__(self):
    return f"MyClass({self.value})"

'''
Описание:

Название: __str__ - возвращает строковое представление объекта.
Выполнение функции: Возвращает строку, представляющую объект MyClass в формате "MyClass(value)".

Эти методы совместно обеспечивают функциональность класса MyClass для инициализации, сложения с другими объектами
того же класса и предоставления строкового представления объекта.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 1 - Вариант № - "уже пошел далеко в дебри" (наверное)
'''
class MyNumber:
    """Класс, представляющий числовое значение."""

    def __init__(self, value):
        """Инициализация объекта."""
        self.value = value

    def __add__(self, other):
        """Перегрузка оператора сложения."""
        if isinstance(other, MyNumber):
            # Создаем новый объект MyNumber с суммой значений
            return MyNumber(self.value + other.value)
        else:
            # Если объект other не является MyNumber, возвращаем None
            return None

    def __radd__(self, other):
        """Правосторонняя перегрузка оператора сложения."""
        # В данном примере правосторонняя перегрузка идентична левосторонней
        return self.__add__(other)

    def __str__(self):
        """Строковое представление объекта."""
        return f"MyNumber({self.value})"

# Пример использования
a = MyNumber(5)
b = MyNumber(10)

# Левосторонняя перегрузка оператора сложения (a + b)
result1 = a + b
print(result1)  # Выведет: MyNumber(15)

# Правосторонняя перегрузка оператора сложения (10 + a)
result2 = 10 + a
print(result2)  # Выведет: MyNumber(15)
'''
Пояснение:

В этом коде класс MyNumber представляет числовое значение.
Метод __add__ перегружает оператор сложения, создавая новый объект MyNumber с суммой значений.
Метод __radd__ обеспечивает правостороннюю перегрузку оператора сложения.

Пример использования демонстрирует, как объекты класса могут использоваться в операции сложения,
как с левой, так и с правой стороны.
'''
'''
Шаг №1. Инициализация объекта:

Пример кода:
'''
def __init__(self, value):
    self.value = value

'''
Описание:

Название: __init__ - это конструктор класса, который вызывается при создании нового объекта.
Выполнение функции: Метод инициализирует атрибут value объекта значением, переданным при создании объекта.
'''
'''
Шаг №2. Перегрузка оператора сложения:

Пример кода:
'''
def __add__(self, other):
    if isinstance(other, MyNumber):
        return MyNumber(self.value + other.value)
    else:
        return None

'''
Описание:

Название: __add__ - перегружает оператор сложения (+).
Выполнение функции:
Проверяется, является ли other объектом класса MyNumber.
Если это так, то создается новый объект MyNumber с атрибутом value, равным сумме значений атрибутов value текущего
и переданного объектов.
Если other не является объектом MyNumber, возвращается None.
'''
'''
Шаг №3. Правосторонняя перегрузка оператора сложения:

Пример кода:
'''
def __radd__(self, other):
    return self.__add__(other)

'''
Описание:

Название: __radd__ - перегружает правосторонний оператор сложения (+).
Выполнение функции:
В данном примере правосторонняя перегрузка просто вызывает метод __add__, обеспечивая симметричность операции сложения.
'''
'''
Шаг №4. Строковое представление объекта:

Пример кода:
'''
def __str__(self):
    return f"MyNumber({self.value})"

'''
Описание:

Название: __str__ - возвращает строковое представление объекта.
Выполнение функции: Возвращает строку, представляющую объект MyNumber в формате "MyNumber(value)".
'''
'''
Эти методы совместно обеспечивают функциональность класса MyNumber для инициализации, сложения с другими 
объектами того же класса и предоставления строкового представления объекта.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 2 - Вариант № 1
'''
class Games:
    Year = 0  # Статическое поле

    def __init__(self, year, name, rating=0, players=1):
        if players < 1:
            raise ValueError("Players must be at least 1")
        self.__year = year  # Приватное поле
        self.__name = name
        self.__rating = rating
        self.__players = players

    def get_name(self):
        return self.__name

    def get_year(self):
        return self.__year

    def get_rating(self):
        return self.__rating

    def get_players(self):
        return self.__players

class PCGames(Games):
    def __init__(self, year, name, platform, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__platform = platform

    def get_name(self):
        return f"{self.__name} (PC)"

    def get_platform(self):
        return self.__platform

class PS4Games(Games):
    def __init__(self, year, name, exclusive_title, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__exclusive_title = exclusive_title

    def get_name(self):
        return f"{self.__name} (PS4 Exclusive)"

class XboxGames(Games):
    def __init__(self, year, name, online_multiplayer, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__online_multiplayer = online_multiplayer

    def get_name(self):
        return f"{self.__name} (Xbox)"

class MobileGames(Games):
    def __init__(self, year, name, genre, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__genre = genre

    def get_name(self):
        return f"{self.__name} (Mobile)"

class NintendoSwitchGames(Games):
    def __init__(self, year, name, exclusive_title, portable_mode=False, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__exclusive_title = exclusive_title
        self.__portable_mode = portable_mode

    def get_name(self):
        return f"{self.__name} (Nintendo Switch)"

    def is_portable(self):
        return self.__portable_mode


# Добавим метод для сортировки списка игр по году выпуска
def sort_games_by_year(games_list):
    return sorted(games_list, key=lambda x: x.get_year())

# Пример использования с играми 2022 и 2023 годов, онлайн-шутерами и Diablo IV
games_list = [
    PCGames(2022, "Cyberpunk 2077", "Windows", 9.0, 1),
    PS4Games(2022, "Horizon Forbidden West", True),
    XboxGames(2022, "Starfield", True),
    MobileGames(2023, "Elden Ring", "Action RPG", 9.5, 1),
    PCGames(2023, "Diablo IV", "Windows", 9.8, 1),
    PCGames(2022, "Battlefield 2042", "Windows", 8.5, 64),
    PS4Games(2022, "Call of Duty: Vanguard", True),
    XboxGames(2023, "Halo Infinite", True),
    MobileGames(2023, "PUBG: New State", "Battle Royale", 8.0, 100),
    NintendoSwitchGames(2022, "Metroid Prime 4", True, True),
]
# Обработка ошибок при создании объектов
try:
    invalid_game = PCGames(2023, "Invalid Game", "Windows", 9.5, -5)
except ValueError as e:
    print(f"Error: {e}")

# Сортировка и вывод списка игр
sorted_games = sort_games_by_year(games_list)
for game in sorted_games:
    print(
        f"Year: {game.get_year()}, Name: {game.get_name()}, Rating: {game.get_rating()}, Players: {game.get_players()},"
        f" Platform: {getattr(game, 'get_platform', lambda: '')()}, Portable: {getattr(game,
                                                                                       'is_portable', lambda: '')()}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Создание базового класса Games
'''
class Games:
    Year = 0  # Статическое поле

    def __init__(self, year, name, rating=0, players=1):
        if players < 1:
            raise ValueError("Players must be at least 1")
        self.__year = year  # Приватное поле
        self.__name = name
        self.__rating = rating
        self.__players = players

    def get_name(self):
        return self.__name

    def get_year(self):
        return self.__year

    def get_rating(self):
        return self.__rating

    def get_players(self):
        return self.__players
'''
Объяснение:

class Games - это базовый класс для всех игр. В нем определены основные свойства и методы, которые будут общими для
всех подклассов.
Year - это статическое поле класса, которое общее для всех экземпляров класса и доступное через класс,
а не через объект.
__init__ - конструктор класса, инициализирующий основные свойства (год, имя, рейтинг, количество игроков).
Проверка if players < 1 используется для гарантии, что количество игроков не меньше 1.
Методы get_name, get_year, get_rating, get_players используются для получения соответствующих свойств.
'''
'''
Шаг 2: Создание подклассов
'''
class PCGames(Games):
    def __init__(self, year, name, platform, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__platform = platform

    def get_name(self):
        return f"{self.__name} (PC)"

    def get_platform(self):
        return self.__platform


'''
Объяснение:

class PCGames(Games) - это подкласс, который наследует все свойства и методы от базового класса Games.
super().__init__ - вызов конструктора родительского класса (Games) для инициализации общих свойств.
get_name - переопределенный метод, который добавляет к имени игры информацию о платформе.
Аналогично создаются и остальные подклассы (PS4Games, XboxGames, MobileGames, NintendoSwitchGames),
каждый из которых имеет свои дополнительные поля и переопределение метода get_name.
'''
'''
Шаг 3: Добавление метода для сортировки
'''
def sort_games_by_year(games_list):
    return sorted(games_list, key=lambda x: x.get_year())

'''
Объяснение:

sort_games_by_year - это функция, которая принимает список игр и возвращает его, отсортированный по году выпуска с
использованием метода get_year. Здесь используется встроенная функция sorted, которая сортирует элементы на основе
заданного ключа (в данном случае, метода get_year).
'''
'''
Шаг 4: Пример использования
'''
games_list = [
    PCGames(2022, "Cyberpunk 2077", "Windows", 9.0, 1),
    PS4Games(2022, "Horizon Forbidden West", True),
    XboxGames(2022, "Starfield", True),
    MobileGames(2023, "Elden Ring", "Action RPG", 9.5, 1),
    PCGames(2023, "Diablo IV", "Windows", 9.8, 1),
    PCGames(2022, "Battlefield 2042", "Windows", 8.5, 64),
    PS4Games(2022, "Call of Duty: Vanguard", True),
    XboxGames(2023, "Halo Infinite", True),
    MobileGames(2023, "PUBG: New State", "Battle Royale", 8.0, 100),
    NintendoSwitchGames(2022, "Metroid Prime 4", True, True),
]
# Обработка ошибок при создании объектов
try:
    invalid_game = PCGames(2023, "Invalid Game", "Windows", 9.5, -5)
except ValueError as e:
    print(f"Error: {e}")

# Сортировка и вывод списка игр
sorted_games = sort_games_by_year(games_list)
for game in sorted_games:
    print(
        f"Year: {game.get_year()}, Name: {game.get_name()}, Rating: {game.get_rating()}, Players: {game.get_players()},"
        f" Platform: {getattr(game, 'get_platform', lambda: '')()}")
'''
Объяснение:

games_list - это список объектов различных типов игр, созданных с использованием подклассов.
Обработка ошибок демонстрирует, как можно обрабатывать исключения при создании объектов с недопустимыми значениями.
Список игр сортируется по году выпуска и выводится информация о каждой игре с использованием методов объектов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 2 - Вариант № 2
'''
class Games:
    Year = 0

    def __init__(self, year, name, rating=0, players=1):
        if players < 1:
            raise ValueError("Players must be at least 1")
        self.__year = year
        self.__name = name
        self.__rating = rating
        self.__players = players

    def get_name(self):
        return self.__name

    def get_year(self):
        return self.__year

    def get_rating(self):
        return self.__rating

    def get_players(self):
        return self.__players

class PS5Games(Games):
    def __init__(self, year, name, exclusive_title, ray_tracing_enabled=False, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__exclusive_title = exclusive_title
        self.__ray_tracing_enabled = ray_tracing_enabled

    def get_name(self):
        return f"{super().get_name()} (PS5 Exclusive)"

    def is_ray_tracing_enabled(self):
        return self.__ray_tracing_enabled

class XboxSeriesXGames(Games):
    def __init__(self, year, name, online_multiplayer, game_pass_available=False,
                 smart_delivery_supported=False, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__online_multiplayer = online_multiplayer
        self.__game_pass_available = game_pass_available
        self.__smart_delivery_supported = smart_delivery_supported

        if (not isinstance(online_multiplayer, bool) or not isinstance(game_pass_available, bool) or
                not isinstance(smart_delivery_supported, bool)):
            raise ValueError("Invalid input for boolean fields")

    def get_name(self):
        return f"{super().get_name()} (Xbox Series X)"

    def is_game_pass_available(self):
        return self.__game_pass_available

    def is_smart_delivery_supported(self):
        return self.__smart_delivery_supported

class ActionRPGGames(Games):
    def __init__(self, year, name, single_player=True, open_world=True, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__single_player = single_player
        self.__open_world = open_world

    def get_name(self):
        return f"{super().get_name()} (Action RPG)"

    def is_single_player(self):
        return self.__single_player

    def is_open_world(self):
        return self.__open_world

class StrategyGames(Games):
    def __init__(self, year, name, single_player=True, multiplayer=True, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__single_player = single_player
        self.__multiplayer = multiplayer

    def get_name(self):
        return f"{super().get_name()} (Strategy)"

    def is_single_player(self):
        return self.__single_player

    def is_multiplayer(self):
        return self.__multiplayer

class PCGames(Games):
    def __init__(self, year, name, operating_system, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__operating_system = operating_system

    def get_name(self):
        return f"{super().get_name()} (PC)"

    def get_platform(self):
        return "PC"

    def is_portable(self):
        return False

class PS4Games(Games):
    def __init__(self, year, name, exclusive_title, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__exclusive_title = exclusive_title

    def get_name(self):
        return f"{super().get_name()} (PS4 Exclusive)"

    def get_platform(self):
        return "PS4"

    def is_portable(self):
        return False

class XboxGames(Games):
    def __init__(self, year, name, online_multiplayer, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__online_multiplayer = online_multiplayer

    def get_name(self):
        return f"{super().get_name()} (Xbox)"

    def get_platform(self):
        return "Xbox"

    def is_portable(self):
        return False

class MobileGames(Games):
    def __init__(self, year, name, genre, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__genre = genre

    def get_name(self):
        return f"{super().get_name()} (Mobile)"

    def get_platform(self):
        return "Mobile"

    def is_portable(self):
        return True

class NintendoSwitchGames(Games):
    def __init__(self, year, name, exclusive_title, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__exclusive_title = exclusive_title

    def get_name(self):
        return f"{super().get_name()} (Nintendo Switch Exclusive)"

    def get_platform(self):
        return "Nintendo Switch"

    def is_portable(self):
        return True

# Создание экземпляров игр для демонстрации
games_list = [
    PCGames(2022, "Cyberpunk 2077", "Windows", 9.0, 1),
    PS4Games(2022, "Horizon Forbidden West", True),
    XboxGames(2022, "Starfield", True),
    MobileGames(2023, "Elden Ring", "Action RPG", 9.5, 1),
    PCGames(2023, "Diablo IV", "Windows", 9.8, 1),
    PS5Games(2022, "Ratchet & Clank: Rift Apart", True, True),
    XboxSeriesXGames(2023, "Fable", True, True),
    MobileGames(2023, "PUBG: New State", "Battle Royale", 8.0, 100),
    NintendoSwitchGames(2022, "Metroid Prime 4", True, True),
    ActionRPGGames(2022, "The Witcher 4", True, True),
    StrategyGames(2023, "Civilization VI", True, True),
]
# Вывод информации о каждой игре
for game in games_list:
    print(
        f"Year: {game.get_year()}, Name: {game.get_name()}, Rating: {game.get_rating()}, Players: {game.get_players()}")
    if isinstance(game, PS5Games):
        print(f"Ray Tracing Enabled: {game.is_ray_tracing_enabled()}")
    elif isinstance(game, XboxSeriesXGames):
        print(
            f"Game Pass Available: {game.is_game_pass_available()}, Smart Delivery Supported: {game.is_smart_delivery_supported()}")
    elif isinstance(game, ActionRPGGames):
        print(f"Single Player: {game.is_single_player()}, Open World: {game.is_open_world()}")
    elif isinstance(game, StrategyGames):
        print(f"Single Player: {game.is_single_player()}, Multiplayer: {game.is_multiplayer()}")
    print()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Создание базового класса Games
'''
class Games:
    Year = 0

    def __init__(self, year, name, rating=0, players=1):
        if players < 1:
            raise ValueError("Players must be at least 1")
        self.__year = year
        self.__name = name
        self.__rating = rating
        self.__players = players

    def get_name(self):
        return self.__name

    def get_year(self):
        return self.__year

    def get_rating(self):
        return self.__rating

    def get_players(self):
        return self.__players
'''
Описание:

Создается базовый класс Games, который содержит статическое поле Year.
Конструктор принимает год, название игры, рейтинг и количество игроков.
Реализованы методы get_name, get_year, get_rating, get_players для получения соответствующих свойств.
'''
'''
Шаг №2: Создание подклассов для различных типов игр и платформ
'''
# Пример подкласса
class PS5Games(Games):
    def __init__(self, year, name, exclusive_title, ray_tracing_enabled=False, rating=0, players=1):
        super().__init__(year, name, rating, players)
        self.__exclusive_title = exclusive_title
        self.__ray_tracing_enabled = ray_tracing_enabled

    def get_name(self):
        return f"{super().get_name()} (PS5 Exclusive)"

    def is_ray_tracing_enabled(self):
        return self.__ray_tracing_enabled

'''
Описание:

Создаются подклассы, представляющие различные платформы и типы игр (PS5Games, XboxSeriesXGames, ActionRPGGames,
StrategyGames, и др.).
В каждом подклассе добавляются дополнительные поля, специфичные для этого типа игры или платформы.
Переопределяется метод get_name для возвращения уникального названия игры, основанного на базовом названии и типе.
'''
'''
Шаг №3: Создание экземпляров игр и вывод информации
'''
# Создание экземпляров игр для демонстрации
games_list = [
    PCGames(2022, "Cyberpunk 2077", "Windows", 9.0, 1),
    PS4Games(2022, "Horizon Forbidden West", True),
    XboxGames(2022, "Starfield", True),
    MobileGames(2023, "Elden Ring", "Action RPG", 9.5, 1),
    PCGames(2023, "Diablo IV", "Windows", 9.8, 1),
    PS5Games(2022, "Ratchet & Clank: Rift Apart", True, True),
    XboxSeriesXGames(2023, "Fable", True, True),
    MobileGames(2023, "PUBG: New State", "Battle Royale", 8.0, 100),
    NintendoSwitchGames(2022, "Metroid Prime 4", True, True),
    ActionRPGGames(2022, "The Witcher 4", True, True),
    StrategyGames(2023, "Civilization VI", True, True),
]
# Вывод информации о каждой игре
for game in games_list:
    print(
        f"Year: {game.get_year()}, Name: {game.get_name()}, Rating: {game.get_rating()}, Players: {game.get_players()}")
    if isinstance(game, PS5Games):
        print(f"Ray Tracing Enabled: {game.is_ray_tracing_enabled()}")
    elif isinstance(game, XboxSeriesXGames):
        print(
            f"Game Pass Available: {game.is_game_pass_available()}, Smart Delivery Supported: {game.is_smart_delivery_supported()}")
    elif isinstance(game, ActionRPGGames):
        print(f"Single Player: {game.is_single_player()}, Open World: {game.is_open_world()}")
    elif isinstance(game, StrategyGames):
        print(f"Single Player: {game.is_single_player()}, Multiplayer: {game.is_multiplayer()}")
    print()
'''
Описание:

Создаются экземпляры различных игр с использованием созданных классов.
Выводится информация о каждой игре, используя методы классов.
Используется проверка типов для вывода дополнительной информации, специфичной для каждого типа игры или платформы.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
'''
Задание № 3 - Вариант № 1
'''
class Country:
    def __init__(self):
        self._population = None

    def set_population(self, population):
        if population >= 0:
            self._population = population
        else:
            print("Население не может быть отрицательным.")

    def get_population(self):
        return self._population
class Russia(Country):
    def __init__(self):
        super().__init__()  # Вызывает конструктор родительского класса
        self._name = "Russia"  # Добавляет поле _name для хранения имени страны

    # Добавляет геттер и сеттер для имени страны
    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name
class Canada(Country):
    def __init__(self):
        super().__init__()
        self._name = "Canada"

    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name
class Germany(Country):
    def __init__(self):
        super().__init__()
        self._name = "Germany"

    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name
# Пример использования
russia = Russia()
russia.set_population(145000000000000)
print(f"{russia.get_name()} population: {russia.get_population()}")
# Вывод: Russia population: 145000000000000

canada = Canada()
canada.set_population(380000)
print(f"{canada.get_name()} population: {canada.get_population()}")
# Вывод: Canada population: 380000

germany = Germany()
germany.set_population(830000)
print(f"{germany.get_name()} population: {germany.get_population()}")
# Вывод: Germany population: 830000
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Определение базового класса Country:
'''
class Country:
    def __init__(self):
        self._population = None

    def set_population(self, population):
        if population >= 0:
            self._population = population
        else:
            print("Население не может быть отрицательным.")

    def get_population(self):
        return self._population
'''
Разбор базового класса Country:
__init__ метод: Инициализирует объект страны. В этом примере, устанавливается пустое значение для _population.

set_population метод: Позволяет установить значение населения для страны. Проверяет, что население неотрицательно.
Если передано отрицательное значение, выводит сообщение об ошибке.

get_population метод: Возвращает текущее значение населения страны.
'''
'''
Определение класса Russia:
'''
class Russia(Country):
    def __init__(self):
        super().__init__()  # Вызывает конструктор родительского класса
        self._name = "Russia"  # Добавляет поле _name для хранения имени страны

    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name
'''
Разбор класса Russia:
__init__ метод: Инициализирует объект класса Russia. Вызывает конструктор родительского класса Country.
Добавляет поле _name для хранения имени страны и устанавливает его значение по умолчанию как "Russia".

get_name метод: Возвращает текущее значение имени страны.

set_name метод: Позволяет установить новое значение для имени страны.
'''
'''
Определение класса Canada (аналогично классу Russia):
'''
class Canada(Country):
    def __init__(self):
        super().__init__()
        self._name = "Canada"

    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name
'''
Определение класса Germany (аналогично классу Russia):
'''
class Germany(Country):
    def __init__(self):
        super().__init__()
        self._name = "Germany"

    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name

'''
Пример использования:
'''
# Создание объекта Russia
russia = Russia()
russia.set_population(145000000000000)
print(f"{russia.get_name()} population: {russia.get_population()}")
# Вывод: Russia population: 145000000000000

# Создание объекта Canada
canada = Canada()
canada.set_population(380000)
print(f"{canada.get_name()} population: {canada.get_population()}")
# Вывод: Canada population: 380000

# Создание объекта Germany
germany = Germany()
germany.set_population(830000)
print(f"{germany.get_name()} population: {germany.get_population()}")
# Вывод: Germany population: 830000
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''ВСЕ СРАЗУ, СКОПОМ'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# Задание №1: Перегрузка оператора сложения

class Number:
    def __init__(self, value):
        self.value = value

    def __add__(self, other):
        # Перегрузка оператора сложения
        if isinstance(other, Number):
            return Number(self.value + other.value)
        else:
            return NotImplemented

    def __str__(self):
        return str(self.value)

a = Number(5)
b = Number(10)
result = a + b
print(result)  # Вывод: 15

# Задание №2: Инкапсуляция и полиморфизм для классов игр
class Games:
    Year = 2023

    def __init__(self, name):
        self._name = name

    def get_name(self):
        # Метод для получения имени игры
        return self._name

class PCGames(Games):
    def __init__(self, name, genre):
        super().__init__(name)
        self._genre = genre

    def get_name(self):
        # Переопределение метода для PCGames
        return f"{self._name} (PC)"

class PS4Games(Games):
    def __init__(self, name, exclusive_title):
        super().__init__(name)
        self._exclusive_title = exclusive_title

    def get_name(self):
        # Переопределение метода для PS4Games
        return f"{self._name} (PS4)"

class XboxGames(Games):
    def __init__(self, name, backward_compatible):
        super().__init__(name)
        self._backward_compatible = backward_compatible

    def get_name(self):
        # Переопределение метода для XboxGames
        return f"{self._name} (Xbox)"

class MobileGames(Games):
    def __init__(self, name, platform):
        super().__init__(name)
        self._platform = platform

    def get_name(self):
        # Переопределение метода для MobileGames
        return f"{self._name} (Mobile)"
# Пример использования
pc_game = PCGames("Cyberpunk 2077", "RPG")
ps4_game = PS4Games("The Last of Us Part II", True)
xbox_game = XboxGames("Halo Infinite", False)
mobile_game = MobileGames("Angry Birds", "iOS")

print(pc_game.get_name())  # Вывод: Cyberpunk 2077 (PC)
print(ps4_game.get_name())  # Вывод: The Last of Us Part II (PS4)
print(xbox_game.get_name())  # Вывод: Halo Infinite (Xbox)
print(mobile_game.get_name())  # Вывод: Angry Birds (Mobile)

# Задание №3: Инкапсуляция и полиморфизм для классов стран
class Country:
    def __init__(self):
        self._population = None

    def set_population(self, population):
        # Метод для установки населения
        if population >= 0:
            self._population = population
        else:
            print("Население не может быть отрицательным.")

    def get_population(self):
        # Метод для получения населения
        return self._population
class Russia(Country):
    def __init__(self):
        super().__init__()
        self._name = "Russia"

    def get_name(self):
        # Переопределение метода для Russia
        return self._name
class Canada(Country):
    def __init__(self):
        super().__init__()
        self._name = "Canada"

    def get_name(self):
        # Переопределение метода для Canada
        return self._name


class Germany(Country):
    def __init__(self):
        super().__init__()
        self._name = "Germany"

    def get_name(self):
        # Переопределение метода для Germany
        return self._name
    
# Пример использования
russia = Russia()
russia.set_population(145000000000000)
print(f"{russia.get_name()} population: {russia.get_population()}")
# Вывод: Russia population: 145000000000000

canada = Canada()
canada.set_population(380000)
print(f"{canada.get_name()} population: {canada.get_population()}")
# Вывод: Canada population: 380000

germany = Germany()
germany.set_population(830000)
print(f"{germany.get_name()} population: {germany.get_population()}")
# Вывод: Germany population: 830000
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение класса Number для перегрузки оператора сложения
'''
class Number:
    def __init__(self, value):
        self.value = value

    def __add__(self, other):
        if isinstance(other, Number):
            return Number(self.value + other.value)
        else:
            return NotImplemented

    def __str__(self):
        return str(self.value)

'''
Описание:

Создается класс Number, представляющий объекты с числовыми значениями.
Метод __add__ перегружает оператор сложения (+). Если второй операнд (other) является объектом класса Number,
возвращается новый объект с суммой значений.
Метод __str__ используется для представления объекта в виде строки.
'''
'''
Шаг №2: Пример использования оператора сложения для объектов Number
'''
a = Number(5)
b = Number(10)
result = a + b
print(result)  # Вывод: 15
'''
Описание:

Создаются два объекта класса Number с числовыми значениями 5 и 10.
Используется оператор сложения для объектов a и b.
Результат (объект с числовым значением 15) выводится на экран.
'''
'''
Шаг №3: Определение базового класса Games
'''
class Games:
    Year = 2023

    def __init__(self, name):
        self._name = name

    def get_name(self):
        return self._name

'''
Описание:

Создается базовый класс Games, содержащий статическое поле Year и конструктор для установки имени игры.
Метод get_name возвращает имя игры.
'''
'''
Шаг №4: Определение подклассов PCGames, PS4Games, XboxGames, MobileGames
'''
class PCGames(Games):
    def __init__(self, name, genre):
        super().__init__(name)
        self._genre = genre

    def get_name(self):
        return f"{self._name} (PC)"

'''
Описание:

Создается подкласс PCGames, наследующийся от Games.
Добавляется поле _genre и конструктор для установки имени и жанра игры.
Метод get_name переопределен для возвращения имени игры с указанием платформы (PC).
Аналогичные шаги выполняются для подклассов PS4Games, XboxGames и MobileGames.
'''
'''
Шаг №5: Пример использования классов игр
'''
pc_game = PCGames("Cyberpunk 2077", "RPG")
ps4_game = PS4Games("The Last of Us Part II", True)
xbox_game = XboxGames("Halo Infinite", False)
mobile_game = MobileGames("Angry Birds", "iOS")

print(pc_game.get_name())  # Вывод: Cyberpunk 2077 (PC)
print(ps4_game.get_name())  # Вывод: The Last of Us Part II (PS4)
print(xbox_game.get_name())  # Вывод: Halo Infinite (Xbox)
print(mobile_game.get_name())  # Вывод: Angry Birds (Mobile)
'''
Описание:

Создаются объекты различных классов игр.
Вызываются методы get_name для каждого объекта, выводящие информацию о каждой игре на экран.
'''
'''
Шаг №6: Определение класса Country и его подклассов Russia, Canada, Germany
'''
class Country:
    def __init__(self):
        self._population = None

    def set_population(self, population):
        if population >= 0:
            self._population = population
        else:
            print("Население не может быть отрицательным.")

    def get_population(self):
        return self._population

'''
Описание:

Создается базовый класс Country для представления страны с полем _population и методами для установки и 
получения населения.
Аналогичные шаги выполняются для подклассов Russia, Canada и Germany, которые добавляют поле _name и
переопределяют метод get_name.
'''
'''
Шаг №7: Пример использования классов стран
'''
russia = Russia()
russia.set_population(145000000)
print(f"{russia.get_name()} population: {russia.get_population()}")
# Вывод: Russia population: 145000000

canada = Canada()
canada.set_population(38000000)
print(f"{canada.get_name()} population: {canada.get_population()}")
# Вывод: Canada population: 38000000

germany = Germany()
germany.set_population(83000000)
print(f"{germany.get_name()} population: {germany.get_population()}")
# Вывод: Germany population: 83000000
'''
Описание:

Создаются объекты различных классов стран.
Устанавливается население для каждой страны с использованием метода set_population.
Выводится информация о каждой стране, включая имя и население.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
Дата выполнения ДОМАШНЕЙ РАБОТЫ: 22-23 - ДЕКАБРЯ 2023
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 22.12.2023
Домашняя работа №18: ООП. Полиморфизм
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:   
'''
'''
Пересмотрите алгоритм решения прошлой практической работы, с использованием инкапсуляции.
Реализуйте старый алгоритм с использованием полиморфизма.

Напишите программу, в которой есть главный класс с текстовым полем.
В главное классе должен быть метод для присваивания значения полю: без аргументов и с одним текстовым аргументом.
Объект главного класса создаётся передачей одного текстового аргумента конструктору. 
На основе главного класса создается класса потомок. 
В классе-потомке нужно добавить числовое поле. 
У конструктора класса-потомка два аргумента.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''''Вариант № 0'''''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Этот пример демонстрирует использование инкапсуляции (сокрытия данных внутри класса), 
наследования (переиспользование кода из родительского класса) и полиморфизма (переопределение метода display_info).
'''
class MainClass:
    def __init__(self, text=None):
        self._text = text

    def set_text(self, text):
        self._text = text

    def display_info(self):
        print(f"MainClass: Text - {self._text}")
class SubClass(MainClass):
    def __init__(self, text=None, number=None):
        super().__init__(text)
        self._number = number

    def set_number(self, number):
        self._number = number

    def display_info(self):
        super().display_info()
        print(f"SubClass: Number - {self._number}")
# Пример использования
main_obj = MainClass("Hello, world!")
main_obj.display_info()

sub_obj = SubClass("Hi there!", 42)
sub_obj.display_info()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Определение MainClass
'''
class MainClass:
    def __init__(self, text=None):
        self._text = text

    def set_text(self, text):
        self._text = text

    def display_info(self):
        print(f"MainClass: Text - {self._text}")

'''
MainClass - это основной класс. Он имеет конструктор __init__, который принимает один аргумент text (по умолчанию None).
Этот аргумент инициализирует защищенное (по соглашению с одним подчеркиванием _) текстовое поле _text.
set_text - метод для установки значения _text.
display_info - метод для отображения информации о текущем объекте MainClass, включая значение _text.
'''
'''
Шаг 2: Определение SubClass
'''
class SubClass(MainClass):
    def __init__(self, text=None, number=None):
        super().__init__(text)
        self._number = number

    def set_number(self, number):
        self._number = number

    def display_info(self):
        super().display_info()
        print(f"SubClass: Number - {self._number}")

'''
SubClass - это класс-потомок MainClass. Он наследует все свойства и методы MainClass.
Конструктор __init__ класса SubClass принимает два аргумента, text и number. super().__init__(text) 
вызывает конструктор родительского класса, чтобы инициализировать текстовое поле. Затем self._number инициализируется 
значением number.
set_number - метод для установки значения _number.
display_info - переопределенный метод отображения информации. 
Сначала вызывается метод display_info родительского класса (super().display_info()), затем выводится информация 
о числовом поле _number.
'''
'''
Шаг 3: Пример использования
'''
# Пример использования
main_obj = MainClass("Hello, world!")
main_obj.display_info()

sub_obj = SubClass("Hi there!", 42)
sub_obj.display_info()
'''
Создается объект main_obj типа MainClass с текстовым значением "Hello, world!".
Выводится информация о main_obj с помощью display_info.
Создается объект sub_obj типа SubClass с текстовым значением "Hi there!" и числовым значением 42.
Выводится информация о sub_obj с помощью display_info, который теперь включает и текст, и число.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''''Вариант № 1'''''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
class MainClass:
    def __init__(self, text_value):
        self._text_value = text_value

    def set_text_value(self, text_value=None):
        if text_value:
            self._text_value = text_value

    def display_info(self):
        print(f"Text Value: {self._text_value}")


class SubClass(MainClass):
    def __init__(self, text_value, numeric_value):
        super().__init__(text_value)
        self._numeric_value = numeric_value

    def set_numeric_value(self, numeric_value):
        self._numeric_value = numeric_value

    def display_info(self):
        super().display_info()
        print(f"Numeric Value: {self._numeric_value}")

# Пример использования
main_object = MainClass("Hello")
sub_object = SubClass("Polymorphism", 42)

# Вывод информации
main_object.display_info()
sub_object.display_info()

# Изменение значений полей
main_object.set_text_value("New Text Value")
sub_object.set_numeric_value(99)

# Повторный вывод информации
main_object.display_info()
sub_object.display_info()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Описание:
Создается класс MainClass с текстовым полем и методами для работы с этим полем.
Создается класс SubClass, который наследуется от MainClass и добавляет числовое поле.
Используется полиморфизм: метод display_info вызывается у объектов разных классов.
Демонстрируется возможность изменения значений полей с использованием соответствующих методов у разных объектов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Класс MainClass:
'''
class MainClass:
    def __init__(self, text_value):
        self._text_value = text_value

    def set_text_value(self, text_value=None):
        if text_value:
            self._text_value = text_value

    def display_info(self):
        print(f"Text Value: {self._text_value}")

'''
Описание:
__init__ метод:

Это конструктор класса. Он вызывается при создании нового объекта класса.
Принимает два аргумента: self (ссылка на сам объект) и text_value (текстовое значение, которое будет присвоено полю).
Инициализирует поле _text_value переданным значением.

set_text_value метод:

Этот метод используется для изменения значения текстового поля.
Принимает self и text_value (по умолчанию None).
Если передан аргумент text_value, то он присваивается полю _text_value.

display_info метод:

Этот метод выводит информацию о текущем значении текстового поля.
Просто использует функцию print для вывода значения поля.
'''
'''
Класс SubClass (наследуется от MainClass):
'''
class SubClass(MainClass):
    def __init__(self, text_value, numeric_value):
        super().__init__(text_value)
        self._numeric_value = numeric_value

    def set_numeric_value(self, numeric_value):
        self._numeric_value = numeric_value

    def display_info(self):
        super().display_info()
        print(f"Numeric Value: {self._numeric_value}")

'''
Описание:
__init__ метод:

Вызывает конструктор родительского класса (MainClass) с передачей ему текстового значения.
Инициализирует свое числовое поле (_numeric_value) переданным числовым значением.

set_numeric_value метод:

Этот метод используется для изменения значения числового поля.
Принимает self и numeric_value.
Присваивает значение атрибуту _numeric_value.

display_info метод:

Переопределяет метод display_info из родительского класса.
Первым делом вызывает родительский display_info для вывода текстового значения.
Затем выводит информацию о числовом поле.
'''
'''
Пример использования:
'''
main_object = MainClass("Hello")
sub_object = SubClass("Polymorphism", 42)

main_object.display_info()
sub_object.display_info()

main_object.set_text_value("New Text Value")
sub_object.set_numeric_value(99)

main_object.display_info()
sub_object.display_info()
'''
Описание:
Создаются объекты классов MainClass и SubClass.
Вызываются методы для вывода информации о полях объектов.
Значения полей изменяются с использованием соответствующих методов.
Снова выводится информация о полях объектов, чтобы продемонстрировать изменения.
'''
'''
Общие замечания:
Применение инкапсуляции: Поля классов начинаются с символа подчеркивания (_), что указывает на их "защищенный" статус.
Такие поля не предназначены для прямого доступа извне класса, и их значения следует изменять только через методы класса.

Полиморфизм: 
Оба класса имеют метод с одинаковым именем (display_info), который используется в примере использования.
Однако, класс SubClass переопределяет этот метод, добавляя свой функционал к функционалу родительского класса.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''''Вариант № Бонус (хотя все одно и тоже)'''''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
class MainClass:
    def __init__(self, text_field=""):
        self._text_field = text_field

    def set_text_field(self, text):
        self._text_field = text


class ChildClass(MainClass):
    def __init__(self, text_field, numeric_field):
        super().__init__(text_field)
        self._numeric_field = numeric_field

    def set_numeric_field(self, numeric_value):
        if numeric_value > 0:
            self._numeric_field = numeric_value
        else:
            print("Значение числового поля должно быть положительным.")

    def display_info(self):
        print(f"Текстовое поле: {self._text_field}")
        print(f"Числовое поле: {self._numeric_field}")


# Пример использования
main_obj = MainClass("Привет, это главный класс.")
print(main_obj._text_field)  # Выведет: Привет, это главный класс.

child_obj = ChildClass("Привет, это дочерний класс.", 42)
child_obj.display_info()
# Выведет:
# Текстовое поле: Привет, это дочерний класс.
# Числовое поле: 42

child_obj.set_text_field("Новый текст для дочернего класса.")
child_obj.set_numeric_field(100)

child_obj.display_info()
# Выведет:
# Текстовое поле: Новый текст для дочернего класса.
# Числовое поле: 100
'''
Здесь код включает в себя классы MainClass и ChildClass, 
а также метод display_info(), который выводит информацию о значениях
текстового и числового полей. Вы можете использовать этот код 
как отправную точку и настроить его под свои конкретные требования.
'''
'''
Теперь давайте рассмотрим, как он соответствует каждому из пунктов:

Инкапсуляция: 
В коде используется инкапсуляция с использованием приватных полей (поле, 
начинающееся с символа _). Например, _text_field и _numeric_field - это 
приватные поля, к которым можно обращаться только изнутри класса.

Полиморфизм: 
В примере есть два класса - MainClass и ChildClass. ChildClass является 
потомком MainClass и переопределяет метод display_info(). 
Это пример полиморфизма, где метод display_info() ведет себя по-разному 
для объектов MainClass и ChildClass.

Главный класс с текстовым полем и методом для присваивания значения:
Класс MainClass соответствует этим требованиям. У него есть текстовое 
поле _text_field и метод set_text_field для установки значения этого поля.

Объект главного класса создается с передачей текстового 
аргумента конструктору: 
Это выполняется в строке main_obj = MainClass("Привет, 
это главный класс.").

Создание класса-потомка с добавлением числового поля: 
Класс ChildClass является классом-потомком MainClass и добавляет 
числовое поле _numeric_field.

У конструктора класса-потомка два аргумента: 
Конструктор ChildClass имеет два аргумента: text_field и numeric_field.
'''
'''
Таким образом, написанный код соответствует всем требованиям 
домашнего задания задания. 
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:   
'''
'''
Пересмотрите алгоритм решения прошлой практической работы, с использованием инкапсуляции.
Реализуйте старый алгоритм с использованием полиморфизма.

Напишите программу, в которой есть главный класс с текстовым полем.
В главное классе должен быть метод для присваивания значения полю: без аргументов и с одним текстовым аргументом.
Объект главного класса создаётся передачей одного текстового аргумента конструктору. 
На основе главного класса создается класса потомок. 
В классе-потомке нужно добавить числовое поле. 
У конструктора класса-потомка два аргумента.
'''
'''
Итак, у нас есть три задачи:

1. Пересмотрите алгоритм решения прошлой практической работы, с использованием инкапсуляции.

Это предполагает, что у вас уже есть какой-то алгоритм, и вам нужно его пересмотреть, добавив инкапсуляцию.
Инкапсуляция - это принцип объектно-ориентированного программирования, который предполагает упаковку данных и методов,
работающих с этими данными,в одном классе. Это делается для контроля доступа к данным и изоляции их от внешнего вмешательства.

Пример:
'''
class MyClass:
    def __init__(self):
        self._my_private_variable = 0  # приватная переменная с использованием _

    def get_private_variable(self):
        return self._my_private_variable

    def set_private_variable(self, value):
        if value > 0:
            self._my_private_variable = value
        else:
            print("Значение должно быть положительным.")
'''
2. Реализуйте старый алгоритм с использованием полиморфизма.

Полиморфизм - это еще один принцип ООП, который позволяет объектам одного класса использовать методы другого класса.
В данном контексте, возможно, нужно создать базовый класс и производные классы, используя полиморфизм.

Пример:
'''
class BaseClass:
    def my_method(self):
        pass  # базовый метод


class DerivedClass1(BaseClass):
    def my_method(self):
        print("Метод из класса DerivedClass1")

class DerivedClass2(BaseClass):
    def my_method(self):
        print("Метод из класса DerivedClass2")

'''
3. Напишите программу с главным классом с текстовым полем и методом для присваивания значения полю.
Объект главного класса создаётся передачей одного текстового аргумента конструктору.
В классе-потомке нужно добавить числовое поле. У конструктора класса-потомка два аргумента.

Пример:
'''
class MainClass:
    def __init__(self, text_field):
        self.text_field = text_field

    def set_text_field(self, text):
        self.text_field = text


class ChildClass(MainClass):
    def __init__(self, text_field, numeric_field):
        super().__init__(text_field)
        self.numeric_field = numeric_field
'''
В данном примере ChildClass наследуется от MainClass, добавляя при этом новое числовое поле numeric_field.
Конструктор ChildClass принимает два аргумента: text_field и numeric_field. 
С помощью super().__init__(text_field) вызывается конструктор родительского 
класса, чтобы проинициализировать текстовое поле.
'''
'''

'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
print(Практическая работа №19: ООП.)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
""""
Дата выполнения практической работы: 26 - ДЕКАБРЯ 2023
""""
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 25.12.2023
Практическая работа №19: ООП.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:
'''
'''
В некой игре-стратегии есть солдаты и герои. 
У всех есть свойство, содержащее уникальный номер объекта, и свойство, в котором хранится принадлежность команде.
У солдат есть метод "иду за героем", который в качестве аргумента принимает объект типа "герой". 
У героев есть метод увеличения собственного уровня.

В основной ветке программы создается по одному герою для каждой команды.
В цикле генерируются объекты-солдаты. Их принадлежность команде определяется случайно.
Солдаты разных команд добавляются в разные списки.

Измеряется длина списков солдат противоборствующих команд и выводится на экран. 
У героя, принадлежащего команде с более длинным списком, увеличивается уровень.

Отправьте одного из солдат первого героя следовать за ним. Выведите на экран идентификационные номера этих двух юнитов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Решение этого Задания ↑   ↑   ↑   ↑   ↑   ↑   ↑
'''
'''
Вариант №1
'''
import random

# Базовый класс для всех юнитов (героев и солдат)
class Unit:
    def __init__(self, unit_id, team):
        self.unit_id = unit_id  # Уникальный номер юнита
        self.team = team  # Принадлежность команде


# Класс солдата, наследуется от класса Unit
class Soldier(Unit):
    def follow_hero(self, hero):
        print(f"Солдат {self.unit_id} идет за героем {hero.unit_id}")


# Класс героя, также наследуется от класса Unit
class Hero(Unit):
    def increase_level(self):
        print(f"Уровень героя {self.unit_id} увеличен!")

    def follow_hero(self, soldier_to_follow):
        print(f"Герой {self.unit_id} ведет с собой солдата {soldier_to_follow.unit_id}")


# Создаем по одному герою для каждой команды
hero_team1 = Hero(unit_id=1, team=1)
hero_team2 = Hero(unit_id=2, team=2)

# Генерируем объекты-солдаты и добавляем их в соответствующие списки команд
soldiers_team1 = [Soldier(unit_id=i, team=1) for i in range(10)]
soldiers_team2 = [Soldier(unit_id=i, team=2) for i in range(8)]

# Измеряем длину списков солдат и выводим на экран
length_team1 = len(soldiers_team1)
length_team2 = len(soldiers_team2)
print(f"Длина списка солдат команды 1: {length_team1}")
print(f"Длина списка солдат команды 2: {length_team2}")

# Увеличиваем уровень героя той команды, у которой больше солдат
if length_team1 > length_team2:
    hero_team1.increase_level()
else:
    hero_team2.increase_level()

# Отправляем первого солдата первого героя следовать за ним
if length_team1 > 0:
    soldier_to_follow = soldiers_team1[0]
    hero_team1.follow_hero(soldier_to_follow)
    # Выводим идентификационные номера этих двух юнитов
    print(f"Идентификационный номер героя: {hero_team1.unit_id}")
    print(f"Идентификационный номер солдата: {soldier_to_follow.unit_id}")
else:
    print("Нет солдат для следования за героем команды 1")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Далее я немножко распишу свои "умозаключения"....
'''
'''
Шаг №1:
Название кода: Инициализация классов и создание объектов

Сам код:
'''
import random

# Базовый класс для всех юнитов (героев и солдат)
class Unit:
    def __init__(self, unit_id, team):
        self.unit_id = unit_id  # Уникальный номер юнита
        self.team = team  # Принадлежность команде

# Класс солдата, наследуется от класса Unit
class Soldier(Unit):
    def follow_hero(self, hero):
        print(f"Солдат {self.unit_id} идет за героем {hero.unit_id}")

# Класс героя, также наследуется от класса Unit
class Hero(Unit):
    def increase_level(self):
        print(f"Уровень героя {self.unit_id} увеличен!")

    def follow_hero(self, soldier_to_follow):
        print(f"Герой {self.unit_id} ведет с собой солдата {soldier_to_follow.unit_id}")

# Создаем по одному герою для каждой команды
hero_team1 = Hero(unit_id=1, team=1)
hero_team2 = Hero(unit_id=2, team=2)

# Генерируем объекты-солдаты и добавляем их в соответствующие списки команд
soldiers_team1 = [Soldier(unit_id=i, team=1) for i in range(10)]
soldiers_team2 = [Soldier(unit_id=i, team=2) for i in range(8)]
'''
Полное и подробное описание:
Этот код создает базовый класс Unit для всех юнитов (героев и солдат). 
Затем определяются классы Soldier и Hero, наследующиеся от Unit. У каждого юнита есть уникальный номер и 
принадлежность к команде.

Создаются по одному герою для каждой из двух команд. 
Затем генерируются списки солдат для каждой команды с использованием списковых включений.
'''
'''
Шаг №2:
Название кода: Измерение длины списков солдат и вывод на экран

Сам код:
'''
# Измеряем длину списков солдат и выводим на экран
length_team1 = len(soldiers_team1)
length_team2 = len(soldiers_team2)
print(f"Длина списка солдат команды 1: {length_team1}")
print(f"Длина списка солдат команды 2: {length_team2}")
'''
Полное и подробное описание:
Этот участок кода измеряет длину списков солдат для каждой команды и выводит результат на экран
с использованием функции len и функции print. Таким образом, мы получаем информацию о количестве солдат
в каждой команде.
'''
'''
Шаг №3:
Название кода: Увеличение уровня героя для команды с большим числом солдат

Сам код:
'''
# Увеличиваем уровень героя той команды, у которой больше солдат
if length_team1 > length_team2:
    hero_team1.increase_level()
else:
    hero_team2.increase_level()
'''
Полное и подробное описание:
В этом фрагменте кода сравниваются длины списков солдат для команд 1 и 2.
Если длина списка солдат команды 1 больше, чем у команды 2, увеличивается уровень героя команды 1, вызывая
метод increase_level(). В противном случае увеличивается уровень героя команды 2.
'''
'''
Шаг №4:
Название кода: Отправка солдата следовать за героем и вывод их идентификационных номеров

Сам код:
'''
# Отправляем первого солдата первого героя следовать за ним
if length_team1 > 0:
    soldier_to_follow = soldiers_team1[0]
    hero_team1.follow_hero(soldier_to_follow)
    # Выводим идентификационные номера этих двух юнитов
    print(f"Идентификационный номер героя: {hero_team1.unit_id}")
    print(f"Идентификационный номер солдата: {soldier_to_follow.unit_id}")
else:
    print("Нет солдат для следования за героем команды 1")
'''
Полное и подробное описание:
Этот фрагмент кода проверяет, есть ли солдаты в списке команды 1. 
Если да, то выбирается первый солдат из списка, и метод follow_hero вызывается для первого героя, 
отправляя этого солдата следовать за ним. Затем выводятся идентификационные номера героя и солдата на экран. 
Если в команде 1 нет солдат, выводится сообщение о том, что нет солдат для следования за героем команды 1.
'''
'''
Рекомендации по улучшению кода:

1. Добавьте проверку в методе follow_hero для класса Hero, чтобы удостовериться, что объект,
переданный в качестве аргумента, действительно является экземпляром класса Soldier. 
Это может предотвратить ошибки в случае передачи некорректных объектов.

2. Добавьте комментарии в метод follow_hero для класса Hero, чтобы объяснить его назначение и то, 
что ожидается в качестве аргумента.

3. Можно добавить функциональность для генерации уникальных идентификационных номеров юнитов,
чтобы избежать возможных конфликтов.

4. Рассмотрите возможность создания отдельных методов или классов для различных этапов выполнения задачи,
чтобы сделать код более модульным и легко читаемым.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Код с рекомендациями
'''
import random

class Unit:
    def __init__(self, unit_id, team):
        self.unit_id = unit_id  # Уникальный номер юнита
        self.team = team  # Принадлежность команде


class Soldier(Unit):
    def follow_hero(self, hero):
        if isinstance(hero, Hero):  # Проверка, что объект - экземпляр класса Hero
            print(f"Солдат {self.unit_id} идет за героем {hero.unit_id}")
        else:
            print("Ошибка: Переданный объект не является героем!")


class Hero(Unit):
    def increase_level(self):
        print(f"Уровень героя {self.unit_id} увеличен!")

    def follow_hero(self, soldier_to_follow):
        if isinstance(soldier_to_follow, Soldier):  # Проверка, что объект - экземпляр класса Soldier
            print(f"Герой {self.unit_id} ведет с собой солдата {soldier_to_follow.unit_id}")
        else:
            print("Ошибка: Переданный объект не является солдатом!")


# Создаем по одному герою для каждой команды
hero_team1 = Hero(unit_id=1, team=1)
hero_team2 = Hero(unit_id=2, team=2)

# Генерируем объекты-солдаты и добавляем их в соответствующие списки команд
soldiers_team1 = [Soldier(unit_id=i, team=1) for i in range(10)]
soldiers_team2 = [Soldier(unit_id=i, team=2) for i in range(8)]

# Измеряем длину списков солдат и выводим на экран
length_team1 = len(soldiers_team1)
length_team2 = len(soldiers_team2)
print(f"Длина списка солдат команды 1: {length_team1}")
print(f"Длина списка солдат команды 2: {length_team2}")

# Увеличиваем уровень героя той команды, у которой больше солдат
if length_team1 > length_team2:
    hero_team1.increase_level()
else:
    hero_team2.increase_level()

# Отправляем первого солдата первого героя следовать за ним
if length_team1 > 0:
    soldier_to_follow = soldiers_team1[0]
    hero_team1.follow_hero(soldier_to_follow)
    # Выводим идентификационные номера этих двух юнитов
    print(f"Идентификационный номер героя: {hero_team1.unit_id}")
    print(f"Идентификационный номер солдата: {soldier_to_follow.unit_id}")
else:
    print("Нет солдат для следования за героем команды 1")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Для себя (другого себя через пол года =))) )
'''
'''
Краткое описание:

Данный код представляет собой пример объектно-ориентированного программирования на языке Python.
В нем создаются классы Unit, Soldier, и Hero, представляющие базовый класс юнита, класс солдата и класс героя
 соответственно. Объекты этих классов имеют уникальный идентификационный номер и принадлежат к какой-то команде.

В основной части программы создаются по одному герою для каждой команды, генерируются списки солдат для каждой команды,
измеряется длина этих списков и выводится на экран. 
Уровень героя увеличивается для команды с более длинным списком солдат.

Один из солдат первого героя отправляется следовать за ним, 
и идентификационные номера героя и солдата выводятся на экран.
'''
'''
Подробное описание:

Класс Unit:
__init__: Инициализирует объект класса с уникальным номером (unit_id) и принадлежностью к команде (team).

Класс Soldier (наследуется от Unit):
follow_hero(hero): Метод, который позволяет солдату следовать за героем.
Внутри метода добавлена проверка, что объект, переданный в качестве аргумента, является экземпляром класса Hero.
В случае успеха выводится сообщение о следовании.

Класс Hero (наследуется от Unit):
increase_level(): Увеличивает уровень героя и выводит сообщение.
follow_hero(soldier_to_follow): Метод, который позволяет герою вести с собой солдата. Внутри метода добавлена проверка,
что объект, переданный в качестве аргумента, является экземпляром класса Soldier. 
В случае успеха выводится сообщение о ведении солдата.

Создание героев и солдат:
Создаются по одному герою для каждой команды (например, hero_team1 и hero_team2).
Генерируются списки солдат для каждой команды (например, soldiers_team1 и soldiers_team2) с использованием списковых
включений.

Измерение длины списков солдат и вывод на экран:
Считается длина списков солдат для каждой команды и выводится на экран.

Увеличение уровня героя для команды с большим числом солдат:
Проверяется, у какой команды больше солдат, и соответствующему герою увеличивается уровень.

Отправка солдата следовать за героем и вывод их идентификационных номеров:
Если в команде есть солдаты, выбирается первый солдат из списка, и герой отправляет его следовать за собой.
Идентификационные номера героя и солдата выводятся на экран.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Рекомендации по улучшению кода:

1. Добавлена проверка в методе follow_hero для класса Hero, чтобы удостовериться, что объект,
переданный в качестве аргумента, действительно является экземпляром класса Soldier.
2. Добавлены комментарии, объясняющие назначение и ожидаемые аргументы для методов follow_hero.
3. Внесены изменения в код согласно рекомендациям.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''

'''
В коде используется конструкция if-else для определения, кому из двух героев увеличить уровень. 
Давайте разберем, почему это нужно:
'''
'''
1. Создание героев:
'''
hero_team1 = Hero(unit_id=1, team=1)
hero_team2 = Hero(unit_id=2, team=2)
'''
2. Генерация солдат:
'''
soldiers_team1 = [Soldier(unit_id=i, team=1) for i in range(10)]
soldiers_team2 = [Soldier(unit_id=i, team=2) for i in range(8)]
'''
3. Измерение длины списков солдат:
'''
length_team1 = len(soldiers_team1)
length_team2 = len(soldiers_team2)
print(f"Длина списка солдат команды 1: {length_team1}")
print(f"Длина списка солдат команды 2: {length_team2}")
'''
4. Увеличение уровня героя той команды, у которой больше солдат:
'''
if length_team1 > length_team2:
    hero_team1.increase_level()
else:
    hero_team2.increase_level()
'''
Здесь проверяется условие: если длина списка солдат команды 1 (length_team1) больше длины списка солдат
команды 2 (length_team2), то вызывается метод increase_level для hero_team1, иначе - для hero_team2. 
Это позволяет увеличить уровень того героя, у которого больше солдат.
'''
'''
5. Отправление одного из солдат первого героя следовать за ним:
'''
if length_team1 > 0:
    soldier_to_follow = random.choice(soldiers_team1)
    hero_team1.follow_hero(soldier_to_follow)
    # Выводим идентификационные номера этих двух юнитов
    print(f"Идентификационный номер героя: {hero_team1.unit_id}")
    print(f"Идентификационный номер солдата: {soldier_to_follow.unit_id}")
else:
    print("Нет солдат для следования за героем команды 1")
'''
Здесь проверяется условие: если длина списка солдат команды 1 больше 0, то выбирается 
случайный солдат из списка soldiers_team1 и вызывается метод follow_hero для героя hero_team1. 
Иначе выводится сообщение о том, что нет солдат для следования за героем команды 1.
'''
'''
Таким образом, конструкция if-else используется для принятия решения в 
зависимости от условия (количество солдат в каждой из команд) в ходе выполнения программы.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №2
'''
import random

# Базовый класс для всех юнитов
class Unit:
    def __init__(self, id, team):
        self.id = id        # Уникальный номер объекта
        self.team = team    # Принадлежность к команде

# Класс для солдат, наследуется от базового класса Unit
class Soldier(Unit):
    def follow_hero(self, hero):
        print(f"Солдат {self.id} следует за героем {hero.id}.")

# Класс для героев, также наследуется от базового класса Unit
class Hero(Unit):
    def increase_level(self):
        print(f"Уровень героя {self.id} увеличен.")

# Создаем героев
hero_team1 = Hero(id=1, team=1)
hero_team2 = Hero(id=2, team=2)

# Создаем солдат
soldiers_team1 = [Soldier(id=i, team=1) for i in range(1, 6)]
soldiers_team2 = [Soldier(id=i, team=2) for i in range(6, 11)]

# Измеряем длину списков солдат
len_team1 = len(soldiers_team1)
len_team2 = len(soldiers_team2)

# Выводим длину списков на экран
print(f"Длина списка солдат команды 1: {len_team1}")
print(f"Длина списка солдат команды 2: {len_team2}")

# Увеличиваем уровень героя команды с более длинным списком
if len_team1 > len_team2:
    hero_team1.increase_level()
else:
    hero_team2.increase_level()

# Отправляем одного из солдат следовать за героем
random_soldier = random.choice(soldiers_team1 + soldiers_team2)
random_soldier.follow_hero(hero_team1 if random_soldier.team == 1 else hero_team2)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1
'''
class Unit:
''''
Пример кода:
''''
def __init__(self, id, team):
    self.id = id  # Уникальный номер объекта
    self.team = team  # Принадлежность к команде
'''
Описание кода или функции:
Этот код определяет базовый класс Unit для всех юнитов в игре. У него есть конструктор __init__, 
который инициализирует уникальный номер объекта (id) и принадлежность к команде (team).
'''
'''
Шаг №2
'''
class Soldier(Unit):
''''
Пример кода:
''''
def follow_hero(self, hero):
    print(f"Солдат {self.id} следует за героем {hero.id}.")
'''
Описание кода или функции:
Этот код определяет класс Soldier, который наследуется от класса Unit. У него есть метод follow_hero,
который принимает объект типа "герой" и выводит сообщение о том, что солдат следует за героем.
'''
'''
Шаг №3
'''
class Hero(Unit):
''''
Пример кода:
''''
def increase_level(self):
    print(f"Уровень героя {self.id} увеличен.")
'''
Описание кода или функции:
Этот код определяет класс Hero, также наследуемый от класса Unit.
У него есть метод increase_level, который выводит сообщение о том, что уровень героя увеличен.
'''
'''
Шаг №4
'''
hero_team1 = Hero(id=1, team=1)
hero_team2 = Hero(id=2, team=2)
'''
Описание кода:
Создаются два объекта героев, представляющих команды 1 и 2.
'''
'''
Шаг №5
'''
soldiers_team1 = [Soldier(id=i, team=1) for i in range(1, 6)]
soldiers_team2 = [Soldier(id=i, team=2) for i in range(6, 11)]
'''
Описание кода:
Создаются списки солдат для каждой команды. В цикле создаются объекты солдат с уникальными номерами и 
принадлежностью к команде.
'''
'''
Шаг №6
'''
len_team1 = len(soldiers_team1)
len_team2 = len(soldiers_team2)
'''
Описание кода:
Измеряется длина списков солдат для обеих команд.
'''
'''
Шаг №7
'''
print(f"Длина списка солдат команды 1: {len_team1}")
print(f"Длина списка солдат команды 2: {len_team2}")
'''
Описание кода:
Выводится на экран длина списков солдат для обеих команд.
'''
'''
Шаг №8
'''
if len_team1 > len_team2:
    hero_team1.increase_level()
else:
    hero_team2.increase_level()
'''
Описание кода:
Увеличивается уровень героя той команды, у которой больше длина списка солдат.
'''
'''
Шаг №9
'''
random_soldier = random.choice(soldiers_team1 + soldiers_team2)
random_soldier.follow_hero(hero_team1 if random_soldier.team == 1 else hero_team2)
'''
Описание кода:
Случайным образом выбирается один солдат из общего списка, и он отправляется следовать
за героем соответствующей команды. Выводится сообщение о следовании.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Рекомендации к рассмотрению:
В коде уже используется наследование и полиморфизм, что хорошо соответствует принципам ООП.

Можно добавить проверки на валидность входных данных при создании объектов 
(например, что id и team являются корректными значениями).

Для более строгой инкапсуляции можно было бы объявить атрибуты id и team в классе Unit как приватные с помощью
двойного подчеркивания (__id, __team).

Рассмотрите возможность использования стандартного модуля enum для удобного представления команд.

Комментарии в коде облегчают его понимание, но их можно расширить, если у вас есть более сложная логика или
особенности, которые стоит выделить.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import random
from enum import Enum

class Team(Enum):
    TEAM1 = 1
    TEAM2 = 2

class Unit:
    def __init__(self, id, team):
        """
        Конструктор базового класса для всех юнитов.

        :param id: Уникальный номер объекта
        :param team: Принадлежность к команде (используется enum Team)
        """
        self.__id = id
        self.__team = team

class Soldier(Unit):
    def follow_hero(self, hero):
        """
        Метод для солдата, который следует за героем.

        :param hero: Объект типа Hero
        """
        print(f"Солдат {self._Unit__id} следует за героем {hero._Unit__id}.")

class Hero(Unit):
    def increase_level(self):
        """
        Метод для увеличения уровня героя.
        """
        print(f"Уровень героя {self._Unit__id} увеличен.")

# Создаем героев
hero_team1 = Hero(id=1, team=Team.TEAM1)
hero_team2 = Hero(id=2, team=Team.TEAM2)

# Создаем солдат
soldiers_team1 = [Soldier(id=i, team=Team.TEAM1) for i in range(1, 6)]
soldiers_team2 = [Soldier(id=i, team=Team.TEAM2) for i in range(6, 11)]

# Измеряем длину списков солдат
len_team1 = len(soldiers_team1)
len_team2 = len(soldiers_team2)

# Выводим длину списков на экран
print(f"Длина списка солдат команды 1: {len_team1}")
print(f"Длина списка солдат команды 2: {len_team2}")

# Увеличиваем уровень героя команды с более длинным списком
if len_team1 > len_team2:
    hero_team1.increase_level()
else:
    hero_team2.increase_level()

# Отправляем одного из солдат следовать за героем
random_soldier = random.choice(soldiers_team1 + soldiers_team2)
random_soldier.follow_hero(hero_team1 if random_soldier._Unit__team == Team.TEAM1 else hero_team2)
'''
Изменения включают в себя:

1. Использование enum Team для более читаемого представления команд.
2. Добавление комментариев к конструктору и методам классов.
3. Приватизация атрибутов id и team с использованием двойного подчеркивания.
4. Обновление вывода сообщений в методах для использования приватных атрибутов.


Эти изменения делают код более читаемым и обеспечивают более строгую инкапсуляцию.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
                              '''     ООП. Множественное наследование     '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 1
'''
'''
Задан класс Point, описывающий точку с координатами x, y на координатной плоскости.

Используя механизм наследования, нужно расширить возможности класса Point путем добавления нового атрибута цвета.
Для этого реализовать подкласс PointColor.

В классе Point реализовать следующие атрибуты:
− координаты точки;
− метод инициализации, который получает 2 параметра — координаты точки x, y;
− метод вычисления расстояния от точки до начала координат;
− метод getPoint(), который возвращает точку в виде списка.

В подклассе PointColor реализовать следующие атрибуты:
− цвет точки color;
− метод начальной инициализации, который получает 3 параметра: координаты точки и цвет;
− метод доступа к цвету color с именем getColor().
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 2
'''
'''
Создать базовый класс «Домашнее животное» и производные классы «Собака», «Кошка», «Попугай», «Хомяк».
С помощью конструктора установить имя каждого животного и его характеристики.

Реализуйте для каждого из классов методы:
• Sound — издает звук животного (пишем текстом в консоль);
• Show — отображает имя животного;
• Type — отображает название его подвида.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 3
'''
'''
Создать базовый класс Employer (служащий) с функцией Print().

Она должна выводить информацию о служащем. 
В случае базового класса это может быть строка с надписью This is Employer class.

Создайте от него три производных класса: President, Manager, Worker.
Переопределите функцию Print() для вывода информации, соответствующей каждому типу служащего.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 1: Классы Point и PointColor
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Этот код моделирует два класса, представляющих точки на координатной плоскости, 
иллюстрируя концепцию наследования в объектно-ориентированном программировании (ООП).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import math

class Point:
    def __init__(self, x, y):
        """
        Конструктор класса Point.

        :param x: Координата x
        :param y: Координата y
        """
        self.x = x
        self.y = y

    def distance_to_origin(self):
        """
        Метод для вычисления расстояния от точки до начала координат.

        :return: Расстояние от точки до начала координат
        """
        return math.sqrt(self.x**2 + self.y**2)

    def get_point(self):
        """
        Метод, возвращающий точку в виде списка.

        :return: Список [x, y]
        """
        return [self.x, self.y]

class PointColor(Point):
    def __init__(self, x, y, color):
        """
        Конструктор класса PointColor.

        :param x: Координата x
        :param y: Координата y
        :param color: Цвет точки
        """
        super().__init__(x, y)
        self.color = color

    def get_color(self):
        """
        Метод доступа к цвету точки.

        :return: Цвет точки
        """
        return self.color

# Пример использования классов
point1 = Point(3, 4)
print("Point1 - Distance to origin:", point1.distance_to_origin())
print("Point1 - Point:", point1.get_point())

point_color1 = PointColor(1, 2, "red")
print("PointColor1 - Distance to origin:", point_color1.distance_to_origin())
print("PointColor1 - Point:", point_color1.get_point())
print("PointColor1 - Color:", point_color1.get_color())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1. Описание класса Point:
'''
import math

class Point:
    def __init__(self, x, y):
        """
        Конструктор класса Point.

        :param x: Координата x
        :param y: Координата y
        """
        self.x = x
        self.y = y

    def distance_to_origin(self):
        """
        Метод для вычисления расстояния от точки до начала координат.

        :return: Расстояние от точки до начала координат
        """
        return math.sqrt(self.x**2 + self.y**2)

    def get_point(self):
        """
        Метод, возвращающий точку в виде списка.

        :return: Список [x, y]
        """
        return [self.x, self.y]
'''
Описание:

Конструктор __init__: Инициализирует объект класса Point с переданными координатами x и y.
Метод distance_to_origin: Вычисляет расстояние от точки до начала координат по формуле sqrt(x^2 + y^2).
Метод get_point: Возвращает координаты точки в виде списка [x, y].
'''
'''
Шаг №2. Описание класса PointColor (подкласс Point):
'''
class PointColor(Point):
    def __init__(self, x, y, color):
        """
        Конструктор класса PointColor.

        :param x: Координата x
        :param y: Координата y
        :param color: Цвет точки
        """
        super().__init__(x, y)
        self.color = color

    def get_color(self):
        """
        Метод доступа к цвету точки.

        :return: Цвет точки
        """
        return self.color
'''
Описание:

Конструктор __init__: Инициализирует объект класса PointColor, вызывая конструктор родительского класса Point и 
добавляя атрибут цвета.
Метод get_color: Возвращает цвет точки.
'''
'''
Шаг №3. Пример использования классов:
'''
# Пример использования классов
point1 = Point(3, 4)
print("Point1 - Distance to origin:", point1.distance_to_origin())
print("Point1 - Point:", point1.get_point())

point_color1 = PointColor(1, 2, "red")
print("PointColor1 - Distance to origin:", point_color1.distance_to_origin())
print("PointColor1 - Point:", point_color1.get_point())
print("PointColor1 - Color:", point_color1.get_color())
'''
Описание:

Создаются объекты point1 типа Point и point_color1 типа PointColor.
Выводятся результаты вызова методов: distance_to_origin(), get_point(), get_color() для обоих объектов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 2: Классы "Домашнее животное", "Собака", "Кошка", "Попугай", "Хомяк"
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
class Pet:
    def __init__(self, name, age, breed, is_vaccinated):
        # Конструктор класса
        self.name = name
        self.__age = age  # Приватное поле для возраста
        self.breed = breed
        self.is_vaccinated = is_vaccinated

    def set_age(self, age):
        # Метод для установки возраста
        if 250 > age > 0:
            self.__age = age

    def get_age(self):
        # Метод для получения возраста
        return self.__age

    def inc_age(self, n=1):
        # Метод для увеличения возраста на n лет
        self.__age += n

    def make_sound(self):
        # Общий метод для издания звука животного
        print('Generic pet sound')

    def show(self):
        # Метод для отображения имени животного
        print(f"Name: {self.name}")

    def animal_type(self):
        # Метод для отображения общего типа животного
        print("Generic animal type")


class Dog(Pet):
    def make_sound(self):
        # Переопределенный метод для издания звука собаки
        print('Woof woof')

    def animal_type(self):
        # Переопределенный метод для отображения типа собаки
        print('Dog')


class Cat(Pet):
    def make_sound(self):
        # Переопределенный метод для издания звука кошки
        print('Meow')

    def animal_type(self):
        # Переопределенный метод для отображения типа кошки
        print('Cat')


class Parrot(Pet):
    def make_sound(self):
        # Переопределенный метод для издания звука попугая
        print('Squawk')

    def animal_type(self):
        # Переопределенный метод для отображения типа попугая
        print('Parrot')


class Hamster(Pet):
    def make_sound(self):
        # Переопределенный метод для издания звука хомяка
        print('Squeak')

    def animal_type(self):
        # Переопределенный метод для отображения типа хомяка
        print('Hamster')


# Пример использования

barky = Dog('Barky', 2, 'Mixed Breed', True)
meowster = Cat('Meowster', 3, 'Siamese', True)
polly = Parrot('Polly', 1, 'Ara', True)
nibbles = Hamster('Nibbles', 1, 'Dwarf', False)

pets = [barky, meowster, polly, nibbles]

for pet in pets:
    # В цикле вызываем методы каждого объекта
    pet.show()
    pet.make_sound()
    pet.animal_type()
    print("\n")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение базового класса "Домашнее животное" (Pet)
'''
# Пример кода:
class Pet:
    def __init__(self, name, age, breed, is_vaccinated):
        self.name = name
        self.__age = age
        self.breed = breed
        self.is_vaccinated = is_vaccinated

    def set_age(self, age):
        if 250 > age > 0:
            self.__age = age

    def get_age(self):
        return self.__age

    def inc_age(self, n=1):
        self.__age += n

    def make_sound(self):
        print('Generic pet sound')

    def show(self):
        print(f"Name: {self.name}")

    def animal_type(self):
        print("Generic animal type")
'''
Описание:

Создается класс Pet с конструктором для инициализации характеристик животного (имя, возраст, порода, привито ли).
Используется приватное поле __age для возраста, чтобы ограничить доступ извне.
Реализованы методы для установки возраста (set_age), получения возраста (get_age), увеличения возраста (inc_age).
Метод make_sound выводит общий звук животного, а метод show отображает имя животного.
animal_type выводит общий тип животного.
'''
'''
Шаг №2: Создание производных классов для каждого вида животного (Dog, Cat, Parrot, Hamster)
'''
# Пример кода:
class Dog(Pet):
    def make_sound(self):
        print('Woof woof')

    def animal_type(self):
        print('Dog')


class Cat(Pet):
    def make_sound(self):
        print('Meow')

    def animal_type(self):
        print('Cat')


class Parrot(Pet):
    def make_sound(self):
        print('Squawk')

    def animal_type(self):
        print('Parrot')


class Hamster(Pet):
    def make_sound(self):
        print('Squeak')

    def animal_type(self):
        print('Hamster')
'''
Описание:

Создаются отдельные классы для каждого вида домашнего животного (собака, кошка, попугай, хомяк).
Каждый класс наследует характеристики и методы базового класса Pet.
Метод make_sound переопределяется для издания специфичного звука каждого вида животного, а метод animal_type для
отображения конкретного типа.
'''
'''
Шаг №3: Пример использования классов
'''
# Пример кода:
barky = Dog('Barky', 2, 'Mixed Breed', True)
meowster = Cat('Meowster', 3, 'Siamese', True)
polly = Parrot('Polly', 1, 'Ara', True)
nibbles = Hamster('Nibbles', 1, 'Dwarf', False)

pets = [barky, meowster, polly, nibbles]

for pet in pets:
    pet.show()
    pet.make_sound()
    pet.animal_type()
    print("\n")
'''
Описание:

Создаются объекты различных классов, представляющих собой разные виды домашних животных.
Объекты добавляются в список pets.
В цикле выводятся информация о каждом животном с использованием методов show, make_sound, animal_type.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Еще один вариант для Задания № 2: Классы "Домашнее животное", "Собака", "Кошка", "Попугай", "Хомяк"
'''
class Pet:
    def __init__(self, name, breed):
        """
        Конструктор базового класса "Домашнее животное".

        :param name: Имя животного
        :param breed: Подвид животного
        """
        self.name = name
        self.breed = breed

    def sound(self):
        """
        Метод для издания звука животного.

        :return: Звук животного
        """
        print("Generic pet sound")

    def show(self):
        """
        Метод для отображения имени животного.

        :return: Имя животного
        """
        print(f"Name: {self.name}")

    def animal_type(self):
        """
        Метод для отображения подвида животного.

        :return: Подвид животного
        """
        print("Generic animal type")


class Dog(Pet):
    def sound(self):
        print("Woof woof")

    def animal_type(self):
        print("Dog")


class Cat(Pet):
    def sound(self):
        print("Meow")

    def animal_type(self):
        print("Cat")


class Parrot(Pet):
    def sound(self):
        print("Squawk")

    def animal_type(self):
        print("Parrot")


class Hamster(Pet):
    def sound(self):
        print("Squeak")

    def animal_type(self):
        print("Hamster")


# Пример использования

dog = Dog("Buddy", "Golden Retriever")
cat = Cat("Whiskers", "Siamese")
parrot = Parrot("Polly", "Ara")
hamster = Hamster("Nibbles", "Dwarf")

pets = [dog, cat, parrot, hamster]

for pet in pets:
    pet.show()
    pet.sound()
    pet.animal_type()
    print("\n")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Создание базового класса "Домашнее животное" (Pet):
'''
class Pet:
    def __init__(self, name, breed):
        """
        Конструктор базового класса "Домашнее животное".

        :param name: Имя животного
        :param breed: Подвид животного
        """
        self.name = name
        self.breed = breed

    def sound(self):
        """
        Метод для издания звука животного.

        :return: Звук животного
        """
        print("Generic pet sound")

    def show(self):
        """
        Метод для отображения имени животного.

        :return: Имя животного
        """
        print(f"Name: {self.name}")

    def animal_type(self):
        """
        Метод для отображения подвида животного.

        :return: Подвид животного
        """
        print("Generic animal type")
'''
Описание:

Создается базовый класс Pet с конструктором для инициализации имени и подвида животного.
Метод sound возвращает общий звук животного, show отображает имя, а animal_type отображает подвид.
'''
'''
Шаг №2: Создание производных классов для каждого вида животного (Dog, Cat, Parrot, Hamster):
'''
class Dog(Pet):
    def sound(self):
        print("Woof woof")

    def animal_type(self):
        print("Dog")


class Cat(Pet):
    def sound(self):
        print("Meow")

    def animal_type(self):
        print("Cat")


class Parrot(Pet):
    def sound(self):
        print("Squawk")

    def animal_type(self):
        print("Parrot")


class Hamster(Pet):
    def sound(self):
        print("Squeak")

    def animal_type(self):
        print("Hamster")
'''
Описание:

Создаются отдельные классы для каждого вида домашнего животного (собака, кошка, попугай, хомяк).
Каждый класс наследует характеристики и методы базового класса Pet.
Метод sound переопределяется для издания специфичного звука каждого вида животного, а метод animal_type для
отображения конкретного типа.
'''
'''
Шаг №3: Пример использования классов:
'''
# Пример использования

dog = Dog("Buddy", "Golden Retriever")
cat = Cat("Whiskers", "Siamese")
parrot = Parrot("Polly", "Ara")
hamster = Hamster("Nibbles", "Dwarf")

pets = [dog, cat, parrot, hamster]

for pet in pets:
    pet.show()
    pet.sound()
    pet.animal_type()
    print("\n")
'''
Описание:

Создаются объекты различных классов, представляющих собой разные виды домашних животных.
Объекты добавляются в список pets.
В цикле выводятся информация о каждом животном с использованием методов show, sound, animal_type.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание № 3: Классы Классы Employer, President, Manager, Worker
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# Определение базового класса Employer
class Employer:
    def print_info(self):
        """
        Функция для вывода информации о служащем.
        """
        print("This is Employer class.")

# Определение производного класса President, наследующего от Employer
class President(Employer):
    def print_info(self):
        # Переопределение метода print_info для класса President
        print("This is President class.")

# Определение производного класса Manager, наследующего от Employer
class Manager(Employer):
    def print_info(self):
        # Переопределение метода print_info для класса Manager
        print("This is Manager class.")

# Определение производного класса Worker, наследующего от Employer
class Worker(Employer):
    def print_info(self):
        # Переопределение метода print_info для класса Worker
        print("This is Worker class.")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Так как у нас есть комментарии, Мы можем попытаться разъяснить каждую часть кода:
'''
'''
1. Базовый класс Employer:

class Employer:: Определяет базовый класс Employer.
def print_info(self):: Определяет метод print_info для вывода информации о служащем.

2. Производные классы (President, Manager, Worker):

class President(Employer):, class Manager(Employer):, class Worker(Employer):: Определяют три производных класса, 
каждый из которых наследует от базового класса Employer.
def print_info(self):: Каждый производный класс переопределяет метод print_info базового класса для вывода информации, 
соответствующей своему типу служащего.

3. Пример использования:

employer = Employer()
president = President()
manager = Manager()
worker = Worker()

employers = [employer, president, manager, worker]

for emp in employers:
    emp.print_info()


Созданы объекты каждого класса.
Объекты добавлены в список employers.
В цикле вызывается метод print_info() для каждого объекта, что приводит к выводу соответствующей информации
о типе служащего.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
print(Домашняя работа №19: ООП.)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
Дата выполнения ДОМАШНЕЙ РАБОТЫ: 26 - ДЕКАБРЯ 2023
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 25.12.2023
Домашняя работа №19: ООП.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:   
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №1

Создайте класс Device, который содержит информацию об устройстве.

С помощью механизма наследования, реализуйте класс CoffeeMachine (содержит информацию о кофемашине),
класс Blender (содержит информацию о блендере), класс MeatGrinder (содержит информацию о мясорубке).

Каждый из классов должен содержать необходимые для работы методы.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №2

Создайте класс Ship, который содержит информацию о корабле.

С помощью механизма наследования, реализуйте класс Frigate (содержит информацию о фрегате),
класс Destroyer (содержит информацию об эсминце), класс Cruiser (содержит информацию о крейсере).
Каждый из классов должен содержать необходимые для работы методы.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ООП. Множественное следование
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Выполните следующее задания:
'''
'''
Задание №1

Используя понятие множественного наследования, разработайте класс «Окружность, вписанная в квадрат».
'''
'''
Задание №2

Используя механизм множественного наследования разработайте класс “Автомобиль”.
Должны быть классы «Колеса», «Двигатель», «Двери».
'''
'''
Задание №3

Создайте базовый класс Shape для рисования плоских фигур.

Определите методы:
Show() — вывод на экран информации о фигуре;
Save() — сохранение фигуры в файл;
Load() — считывание фигуры из файла.

Определите производные классы:

Square — квадрат, который характеризуется координатами левого верхнего угла и длиной стороны;
Rectangle — прямоугольник с заданными координатами верхнего левого угла и размерами;
Circle — окружность с заданными координатами центра и радиусом;
Ellipse — эллипс с заданными координатами верхнего угла, описанного вокруг него прямоугольника со сторонами,
параллельными осям координат, и размерами этого прямоугольника.

Создайте список фигур, сохраните фигуры в файл, загрузите в другой список и отобразите информацию о каждой из фигур.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Задание №1 - Создайте класс Device, который содержит информацию об устройстве.
'''
class Device:
    def __init__(self, brand, model):
        self.brand = brand
        self.model = model

    def start(self):
        print(f"{self.brand} {self.model} is starting")

    def stop(self):
        print(f"{self.brand} {self.model} is stopping")


class CoffeeMachine(Device):
    def __init__(self, brand, model, coffee_type):
        super().__init__(brand, model)
        self.coffee_type = coffee_type

    def make_coffee(self):
        print(f"{self.brand} {self.model} is making {self.coffee_type} coffee")


class Blender(Device):
    def __init__(self, brand, model, speed_levels):
        super().__init__(brand, model)
        self.speed_levels = speed_levels

    def blend(self):
        print(f"{self.brand} {self.model} is blending at {self.speed_levels} speed levels")


class MeatGrinder(Device):
    def __init__(self, brand, model, grind_type):
        super().__init__(brand, model)
        self.grind_type = grind_type

    def grind_meat(self):
        print(f"{self.brand} {self.model} is grinding meat with {self.grind_type} grind")


# Пример использования

coffee_machine = CoffeeMachine("КОФЕВАРИТЕЛЬ", "ЗЯК-КРЯК9000", "Espresso")
blender = Blender("КРУЧУВЕРЧУВКАШУПРЕВРАЩУ", "ХРЯСЬ-БАЦ9000", 5)
meat_grinder = MeatGrinder("ФАРШАДЕЛКА", "МУ-МЯУ-ГАФ9000", "Fine")

devices = [coffee_machine, blender, meat_grinder]

for device in devices:
    device.start()
    device.stop()
    if isinstance(device, CoffeeMachine):
        device.make_coffee()
    elif isinstance(device, Blender):
        device.blend()
    elif isinstance(device, MeatGrinder):
        device.grind_meat()
    print()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение базового класса Device:
'''
class Device:
    def __init__(self, brand, model):
        self.brand = brand
        self.model = model

    def start(self):
        print(f"{self.brand} {self.model} is starting")

    def stop(self):
        print(f"{self.brand} {self.model} is stopping")
'''
Описание:

Создан базовый класс Device, представляющий общие характеристики устройства (бренд и модель).
Конструктор __init__ принимает бренд и модель, и инициализирует соответствующие атрибуты.
Методы start и stop выводят сообщения о запуске и остановке устройства.
'''
'''
Шаг №2: Создание производных классов CoffeeMachine, Blender, MeatGrinder:
'''
class CoffeeMachine(Device):
    def __init__(self, brand, model, coffee_type):
        super().__init__(brand, model)
        self.coffee_type = coffee_type

    def make_coffee(self):
        print(f"{self.brand} {self.model} is making {self.coffee_type} coffee")


class Blender(Device):
    def __init__(self, brand, model, speed_levels):
        super().__init__(brand, model)
        self.speed_levels = speed_levels

    def blend(self):
        print(f"{self.brand} {self.model} is blending at {self.speed_levels} speed levels")


class MeatGrinder(Device):
    def __init__(self, brand, model, grind_type):
        super().__init__(brand, model)
        self.grind_type = grind_type

    def grind_meat(self):
        print(f"{self.brand} {self.model} is grinding meat with {self.grind_type} grind")
'''
Описание:

Созданы три производных класса: CoffeeMachine, Blender, MeatGrinder,
каждый из которых наследует характеристики и методы базового класса Device.
Каждый производный класс имеет свои уникальные атрибуты (например, coffee_type, speed_levels, grind_type) и методы
(make_coffee, blend, grind_meat), соответствующие функциональности устройства.
'''
'''
Шаг №3: Пример использования классов:
'''
coffee_machine = CoffeeMachine("КОФЕВАРИТЕЛЬ", "ЗЯК-КРЯК9000", "Espresso")
blender = Blender("КРУЧУВЕРЧУВКАШУПРЕВРАЩУ", "ХРЯСЬ-БАЦ9000", 5)
meat_grinder = MeatGrinder("ФАРШАДЕЛКА", "МУ-МЯУ-ГАФ9000", "Fine")

devices = [coffee_machine, blender, meat_grinder]

for device in devices:
    device.start()
    device.stop()
    if isinstance(device, CoffeeMachine):
        device.make_coffee()
    elif isinstance(device, Blender):
        device.blend()
    elif isinstance(device, MeatGrinder):
        device.grind_meat()
    print()
'''
Описание:

Создаются объекты каждого класса, представляющие конкретные устройства (кофеварку, блендер, мясорубку).
Объекты добавляются в список devices.
В цикле вызываются методы start и stop для каждого устройства.
Для каждого устройства проверяется его тип, и в зависимости от типа вызывается соответствующий метод 
(make_coffee, blend, grind_meat).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №2 - Создайте класс Ship, который содержит информацию о корабле.
'''
class Ship:
    def __init__(self, name, displacement):
        self.name = name
        self.displacement = displacement

    def start_engine(self):
        print(f"{self.name} для несения боевой службы, запускает свой двигатель, во имя РОДИНЫ!")

    def stop_engine(self):
        print(f"{self.name} обнаружив врага отечества и готовясь к бою, при необходимости, останавливает свой двигатель")

    def perform_special_action(self):
        pass


class Frigate(Ship):
    def __init__(self, name, displacement, missile_count):
        super().__init__(name, displacement)
        self.missile_count = missile_count

    def perform_special_action(self):
        print(f"{self.name} (Фрегат) запускает {self.missile_count} ракеты что-бы уничтожить врага!")


class Destroyer(Ship):
    def __init__(self, name, displacement, torpedo_tubes):
        super().__init__(name, displacement)
        self.torpedo_tubes = torpedo_tubes

    def perform_special_action(self):
        print(f"{self.name} (Эсминец) выпускает {self.torpedo_tubes} торпед с пусковой установки что-бы уничтожить врага!")


class Cruiser(Ship):
    def __init__(self, name, displacement, gun_caliber):
        super().__init__(name, displacement)
        self.gun_caliber = gun_caliber

    def perform_special_action(self):
        print(f"{self.name} (Крейсер) стреляет из своего главного {self.gun_caliber} -го калибра что-бы уничтожить врага!")


# Пример использования

admiral_golovko = Frigate("Адмирал Головко", 4500, 32)
admiral_ushakov = Destroyer("Адмирал Ушаков", 8000, 16)
peter_the_great = Cruiser("Петр Великий", 25000, 130)

ships = [admiral_golovko, admiral_ushakov, peter_the_great]

for ship in ships:
    ship.start_engine()
    ship.stop_engine()
    ship.perform_special_action()
    print()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение базового класса Ship:
'''
class Ship:
    def __init__(self, name, displacement):
        self.name = name
        self.displacement = displacement

    def start_engine(self):
        print(f"{self.name} для несения боевой службы, запускает свой двигатель, во имя РОДИНЫ!")

    def stop_engine(self):
        print(f"{self.name} обнаружив врага отечества и готовясь к бою, при необходимости, останавливает свой двигатель")

    def perform_special_action(self):
        pass
'''
Описание:

Создан базовый класс Ship, представляющий корабль с атрибутами name (имя корабля) и displacement (водоизмещение).
Конструктор __init__ инициализирует атрибуты name и displacement.
Методы start_engine и stop_engine выводят сообщения о запуске и остановке двигателя корабля.
Метод perform_special_action оставлен абстрактным (пустым), так как действие специального действия может 
различаться для каждого типа корабля.
'''
'''
Шаг №2: Создание производных классов Frigate, Destroyer, Cruiser:
'''
class Frigate(Ship):
    def __init__(self, name, displacement, missile_count):
        super().__init__(name, displacement)
        self.missile_count = missile_count

    def perform_special_action(self):
        print(f"{self.name} (Фрегат) запускает {self.missile_count} ракеты что-бы уничтожить врага!")


class Destroyer(Ship):
    def __init__(self, name, displacement, torpedo_tubes):
        super().__init__(name, displacement)
        self.torpedo_tubes = torpedo_tubes

    def perform_special_action(self):
        print(f"{self.name} (Эсминец) выпускает {self.torpedo_tubes} торпед с пусковой установки что-бы уничтожить врага!")


class Cruiser(Ship):
    def __init__(self, name, displacement, gun_caliber):
        super().__init__(name, displacement)
        self.gun_caliber = gun_caliber

    def perform_special_action(self):
        print(f"{self.name} (Крейсер) стреляет из своего главного {self.gun_caliber} -го калибра что-бы уничтожить врага!")
'''
Описание:

Созданы три производных класса: Frigate, Destroyer, Cruiser, каждый из которых наследует характеристики и 
методы базового класса Ship.
Каждый производный класс имеет свои уникальные атрибуты (missile_count, torpedo_tubes, gun_caliber) и 
переопределенный метод perform_special_action для каждого типа корабля.
'''
'''
Шаг №3: Пример использования классов:
'''
admiral_golovko = Frigate("Адмирал Головко", 4500, 32)
admiral_ushakov = Destroyer("Адмирал Ушаков", 8000, 16)
peter_the_great = Cruiser("Петр Великий", 25000, 130)

ships = [admiral_golovko, admiral_ushakov, peter_the_great]

for ship in ships:
    ship.start_engine()
    ship.stop_engine()
    ship.perform_special_action()
    print()
'''
Описание:

Создаются объекты каждого класса, представляющие конкретные типы кораблей (фрегат, эсминец, крейсер).
Объекты добавляются в список ships.
В цикле вызываются методы start_engine и stop_engine для каждого корабля.
Для каждого корабля вызывается метод perform_special_action, который выводит сообщение о выполнении специального
действия (запуск ракет, выпуск торпед, стрельба из орудий).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ООП. Множественное следование
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №1: Класс "Окружность, вписанная в квадрат"
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
class Square:
    def __init__(self, side_length):
        self.side_length = side_length

    def area(self):
        return self.side_length ** 2


class CircleInSquare(Square):
    def __init__(self, side_length):
        super().__init__(side_length)

    def radius(self):
        return self.side_length / 2

    def area(self):
        return 3.14 * self.radius() ** 2


def get_positive_float(prompt):
    while True:
        try:
            value = float(input(prompt))
            if value <= 0:
                raise ValueError("Значение должно быть больше 0")
            return value
        except ValueError:
            print("Неверный ввод. Пожалуйста, введите положительное число.")


# Ввод данных с клавиатуры
try:
    side_length = get_positive_float("Введите длину стороны квадрата: ")
except KeyboardInterrupt:
    print("\nПользователь прервал ввод.")
    exit()

# Создание объекта
try:
    circle_in_square = CircleInSquare(side_length)
except ValueError as e:
    print(f"Ошибка: {e}")
    exit()

# Вывод результатов
print(f"Площадь квадрата: {circle_in_square.area()}")
print(f"Радиус вписанной окружности: {circle_in_square.radius()}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение класса Square:
'''
class Square:
    def __init__(self, side_length):
        self.side_length = side_length

    def area(self):
        return self.side_length ** 2
'''
Описание:

Создан класс Square, представляющий квадрат.
Конструктор __init__ инициализирует атрибут side_length (длина стороны квадрата).
Метод area возвращает площадь квадрата, вычисляемую как квадрат длины стороны.
'''
'''
Шаг №2: Определение класса CircleInSquare (наследуется от Square):
'''
class CircleInSquare(Square):
    def __init__(self, side_length):
        super().__init__(side_length)

    def radius(self):
        return self.side_length / 2

    def area(self):
        return 3.14 * self.radius() ** 2
'''
Описание:

Создан класс CircleInSquare, который наследует от класса Square.
Конструктор __init__ вызывает конструктор родительского класса для инициализации длины стороны квадрата.
Метод radius возвращает радиус вписанной окружности (половина длины стороны квадрата).
Метод area переопределен и возвращает площадь окружности вписанной в квадрат.
'''
'''
Шаг №3: Определение функции get_positive_float:
'''
def get_positive_float(prompt):
    while True:
        try:
            value = float(input(prompt))
            if value <= 0:
                raise ValueError("Значение должно быть больше 0")
            return value
        except ValueError:
            print("Неверный ввод. Пожалуйста, введите положительное число.")
'''
Описание:

Создана функция get_positive_float, которая запрашивает у пользователя ввод положительного числа.
Используется цикл для обработки некорректного ввода.
Если ввод не является положительным числом, возбуждается исключение ValueError.
'''
'''
Шаг №4: Ввод данных с клавиатуры:
'''
try:
    side_length = get_positive_float("Введите длину стороны квадрата: ")
except KeyboardInterrupt:
    print("\nПользователь прервал ввод.")
    exit()
'''
Описание:

Пользователь вводит длину стороны квадрата с клавиатуры, используя функцию get_positive_float.
Обработан случай прерывания пользователем (KeyboardInterrupt), в этом случае программа завершается.
'''
'''
Шаг №5: Создание объекта CircleInSquare:
'''
try:
    circle_in_square = CircleInSquare(side_length)
except ValueError as e:
    print(f"Ошибка: {e}")
    exit()
'''
Описание:

Создается объект circle_in_square класса CircleInSquare с заданной длиной стороны.
Обработан случай возникновения исключения ValueError, если произошла ошибка при создании объекта.
'''
'''
Шаг №6: Вывод результатов:
'''
print(f"Площадь квадрата: {circle_in_square.area()}")
print(f"Радиус вписанной окружности: {circle_in_square.radius()}")
'''
Описание:

Выводится площадь квадрата и радиус вписанной окружности, используя методы объекта circle_in_square.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №2: Класс "Автомобиль" с подклассами "Колеса", "Двигатель", "Двери"
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
class Wheels:
    def __init__(self, wheel_count):
        self.wheel_count = wheel_count


class Engine:
    def __init__(self, fuel_type):
        self.fuel_type = fuel_type

    def start(self):
        print("Двигатель завелся и мы поехали!! СЧАСТЛИВОГО ПУТИ - кричали нам РИПТИЛОЙДЫ")


class Doors:
    def __init__(self, door_count):
        self.door_count = door_count


class Car(Wheels, Engine, Doors):
    def __init__(self, wheel_count, fuel_type, door_count):
        Wheels.__init__(self, wheel_count)
        Engine.__init__(self, fuel_type)
        Doors.__init__(self, door_count)


# Пример использования

my_car = Car(wheel_count=4, fuel_type="Gasoline", door_count=4)
print(f"У моей машины {my_car.wheel_count} колеса, она работает на {my_car.fuel_type}, и имеет {my_car.door_count} двери.")
my_car.start()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение класса Wheels:
'''
class Wheels:
    def __init__(self, wheel_count):
        self.wheel_count = wheel_count
'''
Описание:

Создан класс Wheels, представляющий количество колес.
Конструктор __init__ инициализирует атрибут wheel_count (количество колес).
'''
'''
Шаг №2: Определение класса Engine:
'''
class Engine:
    def __init__(self, fuel_type):
        self.fuel_type = fuel_type

    def start(self):
        print("Двигатель завелся и мы поехали!! СЧАСТЛИВОГО ПУТИ - кричали нам РИПТИЛОЙДЫ")
'''
Описание:

Создан класс Engine, представляющий тип топлива и содержащий метод start.
Конструктор __init__ инициализирует атрибут fuel_type (тип топлива).
Метод start выводит сообщение о том, что двигатель завелся и началось путешествие.
'''
'''
Шаг №3: Определение класса Doors:
'''
class Doors:
    def __init__(self, door_count):
        self.door_count = door_count
'''
Описание:

Создан класс Doors, представляющий количество дверей.
Конструктор __init__ инициализирует атрибут door_count (количество дверей).
'''
'''
Шаг №4: Определение класса Car (множественное наследование):
'''
class Car(Wheels, Engine, Doors):
    def __init__(self, wheel_count, fuel_type, door_count):
        Wheels.__init__(self, wheel_count)
        Engine.__init__(self, fuel_type)
        Doors.__init__(self, door_count)
'''
Описание:

Создан класс Car, который наследует от классов Wheels, Engine, и Doors.
Конструктор __init__ инициализирует атрибуты wheel_count, fuel_type, и door_count, вызывая конструкторы родительских классов.
'''
'''
Шаг №5: Пример использования класса Car:
'''
# Пример использования
my_car = Car(wheel_count=4, fuel_type="Gasoline", door_count=4)
print(f"У моей машины {my_car.wheel_count} колеса, она работает на {my_car.fuel_type}, и имеет {my_car.door_count} двери.")
my_car.start()
'''
Описание:

Создается объект my_car класса Car с указанными параметрами.
Выводится информация о количестве колес, типе топлива и количестве дверей.
Вызывается метод start объекта my_car, выводящий сообщение о запуске двигателя.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №3: Классы для рисования фигур
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# Определение базового класса Shape
class Shape:
    def __init__(self, x=0, y=0):
        """
        Инициализация базового класса для рисования плоских фигур.

        :param x: Координата x
        :param y: Координата y
        """
        self.x = x
        self.y = y

    def show(self):
        """
        Вывод информации о фигуре на экран.
        """
        print(f"Местоположение: ({self.x}, {self.y})")

    def save(self, filename):
        """
        Сохранение информации о фигуре в файл.

        :param filename: Имя файла для сохранения
        """
        with open(filename, 'w') as file:
            file.write(f"{self.__class__.__name__} {self.x} {self.y}\n")

    def load(self, filename):
        """
        Загрузка информации о фигуре из файла.

        :param filename: Имя файла для загрузки
        """
        with open(filename, 'r') as file:
            data = file.read().split()
            self.x, self.y = int(data[1]), int(data[2])


# Определение производных классов

class Square(Shape):
    def __init__(self, x=0, y=0, side_length=0):
        """
        Инициализация класса для рисования квадрата.

        :param x: Координата x верхнего левого угла квадрата
        :param y: Координата y верхнего левого угла квадрата
        :param side_length: Длина стороны квадрата
        """
        super().__init__(x, y)
        self.side_length = side_length

    def show(self):
        """
        Вывод информации о квадрате на экран.
        """
        super().show()
        print(f"Тип: Квадрат, Длина стороны: {self.side_length}")


class Rectangle(Shape):
    def __init__(self, x=0, y=0, width=0, height=0):
        """
        Инициализация класса для рисования прямоугольника.

        :param x: Координата x верхнего левого угла прямоугольника
        :param y: Координата y верхнего левого угла прямоугольника
        :param width: Ширина прямоугольника
        :param height: Высота прямоугольника
        """
        super().__init__(x, y)
        self.width = width
        self.height = height

    def show(self):
        """
        Вывод информации о прямоугольнике на экран.
        """
        super().show()
        print(f"Тип: Прямоугольник, Ширина: {self.width}, Высота: {self.height}")


class Circle(Shape):
    def __init__(self, x=0, y=0, radius=0):
        """
        Инициализация класса для рисования окружности.

        :param x: Координата x центра окружности
        :param y: Координата y центра окружности
        :param radius: Радиус окружности
        """
        super().__init__(x, y)
        self.radius = radius

    def show(self):
        """
        Вывод информации о окружности на экран.
        """
        super().show()
        print(f"Тип: Окружность, Радиус: {self.radius}")


class Ellipse(Shape):
    def __init__(self, x=0, y=0, major_axis=0, minor_axis=0):
        """
        Инициализация класса для рисования эллипса.

        :param x: Координата x верхнего угла описанного прямоугольника
        :param y: Координата y верхнего угла описанного прямоугольника
        :param major_axis: Большая полуось эллипса
        :param minor_axis: Малая полуось эллипса
        """
        super().__init__(x, y)
        self.major_axis = major_axis
        self.minor_axis = minor_axis

    def show(self):
        """
        Вывод информации об эллипсе на экран.
        """
        super().show()
        print(f"Тип: Эллипс, Большая полуось: {self.major_axis}, Малая полуось: {self.minor_axis}")


# Пример использования

shapes = []

try:
    # Ввод данных с клавиатуры и обработка ошибок
    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    side_length = int(input("Введите длину стороны квадрата: "))
    shapes.append(Square(x, y, side_length))

    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    width = int(input("Введите ширину прямоугольника: "))
    height = int(input("Введите высоту прямоугольника: "))
    shapes.append(Rectangle(x, y, width, height))

    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    radius = int(input("Введите радиус окружности: "))
    shapes.append(Circle(x, y, radius))

    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    major_axis = int(input("Введите большую полуось эллипса: "))
    minor_axis = int(input("Введите малую полуось эллипса: "))
    shapes.append(Ellipse(x, y, major_axis, minor_axis))
except ValueError:
    print("Ошибка ввода данных. Введите целые числа.")

# Сохранение в файлы
for i, shape in enumerate(shapes):
    shape.save(f'shape_{i}.txt')

# Загрузка из файлов
loaded_shapes = [Square(), Rectangle(), Circle(), Ellipse()]

for i, shape in enumerate(loaded_shapes):
    shape.load(f'shape_{i}.txt')
    shape.show()
    print()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение базового класса Shape:
'''
class Shape:
    def __init__(self, x=0, y=0):
        """
        Инициализация базового класса для рисования плоских фигур.

        :param x: Координата x
        :param y: Координата y
        """
        self.x = x
        self.y = y

    def show(self):
        """
        Вывод информации о фигуре на экран.
        """
        print(f"Местоположение: ({self.x}, {self.y})")

    def save(self, filename):
        """
        Сохранение информации о фигуре в файл.

        :param filename: Имя файла для сохранения
        """
        with open(filename, 'w') as file:
            file.write(f"{self.__class__.__name__} {self.x} {self.y}\n")

    def load(self, filename):
        """
        Загрузка информации о фигуре из файла.

        :param filename: Имя файла для загрузки
        """
        with open(filename, 'r') as file:
            data = file.read().split()
            self.x, self.y = int(data[1]), int(data[2])
'''
Описание:

Создан базовый класс Shape, содержащий координаты x и y.
Конструктор __init__ инициализирует координаты. По умолчанию они равны 0.
Метод show выводит информацию о местоположении фигуры.
Метод save сохраняет информацию о фигуре в файл, используя имя класса, x и y.
Метод load загружает информацию о фигуре из файла и устанавливает соответствующие координаты.
'''
'''
Шаг №2: Определение производных классов (Square, Rectangle, Circle, Ellipse):
'''
class Square(Shape):
    def __init__(self, x=0, y=0, side_length=0):
        """
        Инициализация класса для рисования квадрата.

        :param x: Координата x верхнего левого угла квадрата
        :param y: Координата y верхнего левого угла квадрата
        :param side_length: Длина стороны квадрата
        """
        super().__init__(x, y)
        self.side_length = side_length

    def show(self):
        """
        Вывод информации о квадрате на экран.
        """
        super().show()
        print(f"Тип: Квадрат, Длина стороны: {self.side_length}")


class Rectangle(Shape):
    def __init__(self, x=0, y=0, width=0, height=0):
        """
        Инициализация класса для рисования прямоугольника.

        :param x: Координата x верхнего левого угла прямоугольника
        :param y: Координата y верхнего левого угла прямоугольника
        :param width: Ширина прямоугольника
        :param height: Высота прямоугольника
        """
        super().__init__(x, y)
        self.width = width
        self.height = height

    def show(self):
        """
        Вывод информации о прямоугольнике на экран.
        """
        super().show()
        print(f"Тип: Прямоугольник, Ширина: {self.width}, Высота: {self.height}")


class Circle(Shape):
    def __init__(self, x=0, y=0, radius=0):
        """
        Инициализация класса для рисования окружности.

        :param x: Координата x центра окружности
        :param y: Координата y центра окружности
        :param radius: Радиус окружности
        """
        super().__init__(x, y)
        self.radius = radius

    def show(self):
        """
        Вывод информации о окружности на экран.
        """
        super().show()
        print(f"Тип: Окружность, Радиус: {self.radius}")


class Ellipse(Shape):
    def __init__(self, x=0, y=0, major_axis=0, minor_axis=0):
        """
        Инициализация класса для рисования эллипса.

        :param x: Координата x верхнего угла описанного прямоугольника
        :param y: Координата y верхнего угла описанного прямоугольника
        :param major_axis: Большая полуось эллипса
        :param minor_axis: Малая полуось эллипса
        """
        super().__init__(x, y)
        self.major_axis = major_axis
        self.minor_axis = minor_axis

    def show(self):
        """
        Вывод информации об эллипсе на экран.
        """
        super().show()
        print(f"Тип: Эллипс, Большая полуось: {self.major_axis}, Малая полуось: {self.minor_axis}")
'''
Описание:

Созданы производные классы Square, Rectangle, Circle, и Ellipse, наследующие от базового класса Shape.
Каждый из производных классов имеет свои собственные атрибуты и метод show,
который переопределяет метод из базового класса.
'''
'''
Шаг №3: Пример использования:
'''
# Пример использования
shapes = []

try:
    # Ввод данных с клавиатуры и обработка ошибок
    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    side_length = int(input("Введите длину стороны квадрата: "))
    shapes.append(Square(x, y, side_length))

    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    width = int(input("Введите ширину прямоугольника: "))
    height = int(input("Введите высоту прямоугольника: "))
    shapes.append(Rectangle(x, y, width, height))

    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    radius = int(input("Введите радиус окружности: "))
    shapes.append(Circle(x, y, radius))

    x = int(input("Введите координату x: "))
    y = int(input("Введите координату y: "))
    major_axis = int(input("Введите большую полуось эллипса: "))
    minor_axis = int(input("Введите малую полуось эллипса: "))
    shapes.append(Ellipse(x, y, major_axis, minor_axis))
except ValueError:
    print("Ошибка ввода данных. Введите целые числа.")

# Сохранение в файлы
for i, shape in enumerate(shapes):
    shape.save(f'shape_{i}.txt')

# Загрузка из файлов
loaded_shapes = [Square(), Rectangle(), Circle(), Ellipse()]

for i, shape in enumerate(loaded_shapes):
    shape.load(f'shape_{i}.txt')
    shape.show()
    print()
'''
Описание:

Создается список shapes, в который добавляются объекты разных фигур (в данном случае, только квадрат).
Происходит ввод данных с клавиатуры, создание объектов и добавление их в список shapes.
Фигуры сохраняются в файлы (с именами shape_0.txt, shape_1.txt, и т.д.).
Затем создается список loaded_shapes для загрузки фигур из файлов.
Происходит загрузка фигур из файлов и вывод их информации с помощью метода show.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# # множественное наследование
# class LivingBeing:
#     # общие свойства и методы
#     # def move(self, destination, speed = 0, origin = ' '):
#     pass
#
# class CanMove:
#     def move(self, destination, speed = 0, origin = ''):
#         pass
#
# class CanRun:
#     def move(self, destination, speed = 50, origin = ''):
#         pass
#
# class Human(LivingBeing, CanMove, CanRun):
#     def move(self, destination, type):
#
#
# # проблемы множественного наследования
# class SuperUser(User, Admin): pass

class my_int:
    def __init__(self, a_value):
        self.value = a_value
    def plus(self, second_int):
        return self.value + second_int.value

    def __add__(self, second_int):
        return self.value + second_int.value

    def __str__(self):
        return str(self.value)
one = my_int(1)
two = my_int(2)
three = my_int(3)

print(one.plus(two))
print(one)

# print(type(1+1))
# print(type(1.5+1.5))
# print(type('a'+'b'))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
""""
Дата выполнения практической работы: 27-28 - ДЕКАБРЯ 2023
""""
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 27.12.2023
Практическая работа №20: ООП.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №1
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс Число (или используйте уже ранее созданный вами).
Класс число хранит внутри одно значение. 
Используя перегрузку операторов реализуйте для него арифметические операции для работы с числом (операции +, -, *, /).

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №2
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс Дробь (или используйте уже ранее созданный вами). 
Используя перегрузку операторов реализуйте для него арифметические операции для работы с дробями (операции +, -, *, /).

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №3
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс Библиотека. 
Класс предназначен для хранения информации о библиотеке (название, адрес, количество книг). 
Реализуйте необходимые для класса методы. 
Используя перегрузку операторов реализуйте для него следующие арифметические операции:
+ добавляет к количеству книг указанное значение;
- вычитает из количества книг указанное значение;
+= добавляет к количеству книг указанное значение;
-= вычитает из количества книг указанное значение; Используя перегрузку операторов реализуйте 
(сравнение по количеству книг):
<;
>;
<=;
>=;
==;
!=.

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №4
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс Date, который будет содержать информацию о дате (день, месяц, год).
С помощью механизма перегрузки операторов, определите операцию разности двух дат 
(результат в виде количества дней между датами), а также операцию увеличения даты на определенное количество дней.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Решение этих заданий ↑   ↑   ↑   ↑
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №1: Класс Число
'''
'''
Этот код согласно заданию №1 создает класс Number, 
который может выполнять арифметические операции с объектами этого класса.
При этом он обеспечивает контроль типов и обрабатывает ситуации, такие как деление на ноль.
'''
class Number:
    def __init__(self, value):
        self.value = value

    def __add__(self, other):
        return Number(self.value + other.value)

    def __sub__(self, other):
        return Number(self.value - other.value)

    def __mul__(self, other):
        return Number(self.value * other.value)

    def __truediv__(self, other):
        return Number(self.value / other.value)

# Пример использования
num1 = Number(5)
num2 = Number(3)

result_sum = num1 + num2
result_diff = num1 - num2
result_mul = num1 * num2
result_div = num1 / num2

print(result_sum.value)  # 8
print(result_diff.value)  # 2
print(result_mul.value)  # 15
print(result_div.value)  # 1.6666666666666667
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1 - Создание класса Number
'''
class Number:
    def __init__(self, value):
        self.value = value

    # Перегрузка оператора сложения (+)
    def __add__(self, other):
        if isinstance(other, Number):
            return Number(self.value + other.value)
        else:
            raise TypeError("Unsupported operand type for +: {}".format(type(other)))

    # Перегрузка оператора вычитания (-)
    def __sub__(self, other):
        if isinstance(other, Number):
            return Number(self.value - other.value)
        else:
            raise TypeError("Unsupported operand type for -: {}".format(type(other)))

    # Перегрузка оператора умножения (*)
    def __mul__(self, other):
        if isinstance(other, Number):
            return Number(self.value * other.value)
        else:
            raise TypeError("Unsupported operand type for *: {}".format(type(other)))

    # Перегрузка оператора деления (/)
    def __truediv__(self, other):
        if isinstance(other, Number):
            if other.value != 0:
                return Number(self.value / other.value)
            else:
                raise ValueError("Division by zero is not allowed")
        else:
            raise TypeError("Unsupported operand type for /: {}".format(type(other)))

    def __str__(self):
        return str(self.value)
'''
Подробное описание (шаг №1):

1. Определение класса Number: Класс Number содержит конструктор __init__, который принимает значение и 
инициализирует атрибут value.

2. Перегрузка операторов:
__add__: Перегружает оператор сложения (+). Если второй операнд тоже объект класса Number, 
создается новый объект с суммой значений. В противном случае, выбрасывается исключение типа данных.
__sub__: Перегружает оператор вычитания (-). Аналогично, создается новый объект с разностью значений.
__mul__: Перегружает оператор умножения (*). Создает новый объект с произведением значений.
__truediv__: Перегружает оператор деления (/). Создает новый объект с результатом деления значений. 
Проверяет, что значение делителя не равно нулю, чтобы избежать ошибки деления на ноль.
__str__: Возвращает строковое представление объекта, чтобы можно было вывести его с использованием print.
'''
'''
Шаг №2 - Пример использования
'''
# Пример использования
num1 = Number(5)
num2 = Number(3)

result_sum = num1 + num2
result_diff = num1 - num2
result_prod = num1 * num2
result_div = num1 / num2

print("Сложение:", result_sum)  # Вывод: Сложение: 8
print("Вычитание:", result_diff)  # Вывод: Вычитание: 2
print("Умножение:", result_prod)  # Вывод: Умножение: 15
print("Деление:", result_div)  # Вывод: Деление: 1.6666666666666667
'''
Подробное описание (шаг №2):

1. Создание объектов: Создаются два объекта класса Number: num1 и num2 с значениями 5 и 3 соответственно.

2. Арифметические операции:
Производятся арифметические операции с использованием перегруженных операторов (+, -, *, /).
Результаты операций присваиваются переменным result_sum, result_diff, result_prod, result_div.
3. Вывод результатов: Выводятся результаты операций с использованием print.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №2: Класс Дробь - Вариант 1
'''
class Fraction:
    def __init__(self, numerator, denominator):
        self.numerator = numerator
        self.denominator = denominator

    def __add__(self, other):
        return Fraction(
            self.numerator * other.denominator + other.numerator * self.denominator,
            self.denominator * other.denominator
        )

    def __sub__(self, other):
        return Fraction(
            self.numerator * other.denominator - other.numerator * self.denominator,
            self.denominator * other.denominator
        )

    def __mul__(self, other):
        return Fraction(
            self.numerator * other.numerator,
            self.denominator * other.denominator
        )

    def __truediv__(self, other):
        return Fraction(
            self.numerator * other.denominator,
            self.denominator * other.numerator
        )

# Пример использования
frac1 = Fraction(1, 2)
frac2 = Fraction(3, 4)

result_sum = frac1 + frac2
result_diff = frac1 - frac2
result_mul = frac1 * frac2
result_div = frac1 / frac2

print(result_sum.numerator, '/', result_sum.denominator)  # 5 / 4
print(result_diff.numerator, '/', result_diff.denominator)  # -1 / 4
print(result_mul.numerator, '/', result_mul.denominator)  # 3 / 8
print(result_div.numerator, '/', result_div.denominator)  # 2 / 3
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1 - Создание класса Fraction
'''
class Fraction:
    def __init__(self, numerator, denominator):
        self.numerator = numerator
        self.denominator = denominator

    def __add__(self, other):
        return Fraction(
            self.numerator * other.denominator + other.numerator * self.denominator,
            self.denominator * other.denominator
        )

    def __sub__(self, other):
        return Fraction(
            self.numerator * other.denominator - other.numerator * self.denominator,
            self.denominator * other.denominator
        )

    def __mul__(self, other):
        return Fraction(
            self.numerator * other.numerator,
            self.denominator * other.denominator
        )

    def __truediv__(self, other):
        return Fraction(
            self.numerator * other.denominator,
            self.denominator * other.numerator
        )
'''
Подробное описание (шаг №1):

Определение класса Fraction: Создается класс Fraction, представляющий дробь. 
Конструктор __init__ принимает числитель (numerator) и знаменатель (denominator) дроби и инициализирует 
соответствующие атрибуты.

Перегрузка операторов:

__add__: Перегружает оператор сложения (+). Создает новый объект класса Fraction, представляющий сумму двух дробей.
__sub__: Перегружает оператор вычитания (-). Создает новый объект с разностью двух дробей.
__mul__: Перегружает оператор умножения (*). Создает новый объект с произведением двух дробей.
__truediv__: Перегружает оператор деления (/). Создает новый объект с результатом деления двух дробей.
'''
'''
Шаг №2 - Пример использования
'''
# Пример использования
frac1 = Fraction(1, 2)
frac2 = Fraction(3, 4)

result_sum = frac1 + frac2
result_diff = frac1 - frac2
result_mul = frac1 * frac2
result_div = frac1 / frac2

print(result_sum.numerator, '/', result_sum.denominator)  # 5 / 4
print(result_diff.numerator, '/', result_diff.denominator)  # -1 / 4
print(result_mul.numerator, '/', result_mul.denominator)  # 3 / 8
print(result_div.numerator, '/', result_div.denominator)  # 2 / 3
'''
Подробное описание (шаг №2):

1. Создание объектов: Создаются два объекта класса Fraction: frac1 с числителем 1 и знаменателем 2,
и frac2 с числителем 3 и знаменателем 4.

2. Арифметические операции:
Выполняются арифметические операции с использованием перегруженных операторов (+, -, *, /).
Результаты операций присваиваются переменным result_sum, result_diff, result_mul, result_div.

3. Вывод результатов: Выводятся числители и знаменатели результатов операций в виде дроби.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №2: Класс Дробь - Вариант 2
'''
class Fraction:
    def __init__(self, numerator=0, denominator=1):
        self.numerator = numerator
        self.denominator = denominator
        self.simplify()

    def simplify(self):
        common_factor = self.gcd(self.numerator, self.denominator)
        self.numerator //= common_factor
        self.denominator //= common_factor

    def gcd(self, a, b):
        while b:
            a, b = b, a % b
        return a

    def input_fraction(self):
        try:
            numerator = int(input("Введите числитель: "))
            denominator = int(input("Введите знаменатель: "))
            if denominator == 0:
                raise ValueError("Знаменатель не может быть равен 0.")
            return Fraction(numerator, denominator)
        except ValueError as e:
            print(f"Ошибка: {e}. Пожалуйста, введите корректные целочисленные значения.")
            return None

    def __add__(self, other):
        if isinstance(other, Fraction):
            return Fraction(
                self.numerator * other.denominator + other.numerator * self.denominator,
                self.denominator * other.denominator
            )
        else:
            raise TypeError("Unsupported operand type for +: {}".format(type(other)))

    def __sub__(self, other):
        if isinstance(other, Fraction):
            return Fraction(
                self.numerator * other.denominator - other.numerator * self.denominator,
                self.denominator * other.denominator
            )
        else:
            raise TypeError("Unsupported operand type for -: {}".format(type(other)))

    def __mul__(self, other):
        if isinstance(other, Fraction):
            return Fraction(
                self.numerator * other.numerator,
                self.denominator * other.denominator
            )
        else:
            raise TypeError("Unsupported operand type for *: {}".format(type(other)))

    def __truediv__(self, other):
        if isinstance(other, Fraction):
            if other.numerator == 0:
                raise ValueError("Деление на ноль не допустимо.")
            return Fraction(
                self.numerator * other.denominator,
                self.denominator * other.numerator
            )
        else:
            raise TypeError("Unsupported operand type for /: {}".format(type(other)))

    def __str__(self):
        return f"{self.numerator}/{self.denominator}"


# Пример использования с вводом от пользователя
fraction1 = Fraction().input_fraction()
fraction2 = Fraction().input_fraction()

if fraction1 and fraction2:
    result_sum = fraction1 + fraction2
    result_diff = fraction1 - fraction2
    result_mul = fraction1 * fraction2
    result_div = fraction1 / fraction2

    print("Сложение:", result_sum)
    print("Вычитание:", result_diff)
    print("Умножение:", result_mul)
    print("Деление:", result_div)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
1. Конструктор __init__:
'''
def __init__(self, numerator=0, denominator=1):
    self.numerator = numerator
    self.denominator = denominator
    self.simplify()
'''
Описание: 
Конструктор класса инициализирует объект дроби заданными значениями числителя (numerator) и знаменателя (denominator).
Если значения не переданы, используются значения по умолчанию (0 для числителя и 1 для знаменателя). 
После инициализации вызывается метод self.simplify(), который упрощает дробь.
'''
'''
2. Метод simplify:
'''
def simplify(self):
    common_factor = self.gcd(self.numerator, self.denominator)
    self.numerator //= common_factor
    self.denominator //= common_factor
'''
Описание: 
Метод упрощает дробь, выделяя общий делитель между числителем и знаменателем. 
Внутри метода используется функция gcd для нахождения наибольшего общего делителя. 
Затем числитель и знаменатель дроби делятся на этот общий делитель, что приводит к упрощенной дроби.
'''
'''
3. Метод gcd:
'''
def gcd(self, a, b):
    while b:
        a, b = b, a % b
    return a
'''
Описание: 
Метод вычисляет наибольший общий делитель двух чисел a и b с использованием алгоритма Евклида.
Метод возвращает найденный наибольший общий делитель.
'''
'''
4. Метод input_fraction:
'''
def input_fraction(self):
    try:
        numerator = int(input("Введите числитель: "))
        denominator = int(input("Введите знаменатель: "))
        if denominator == 0:
            raise ValueError("Знаменатель не может быть равен 0.")
        return Fraction(numerator, denominator)
    except ValueError as e:
        print(f"Ошибка: {e}. Пожалуйста, введите корректные целочисленные значения.")
        return None
'''
Описание: 
Метод запрашивает у пользователя ввод числителя и знаменателя и возвращает новый объект класса Fraction
с введенными значениями. При возникновении ошибки (например, если знаменатель равен 0), выводится сообщение об
ошибке, и метод возвращает None.
'''
'''
5. Магические методы для арифметических операций (__add__, __sub__, __mul__, __truediv__):
'''
def __add__(self, other):
    # ...

def __sub__(self, other):
    # ...

def __mul__(self, other):
    # ...

def __truediv__(self, other):
    # ...
''''
Описание: 
Эти методы реализуют арифметические операции сложения, вычитания, умножения и деления для объектов класса Fraction.
Проверяется, является ли other объектом класса Fraction, и выполняются соответствующие операции.
''''
'''
6. Метод __str__:
'''
def __str__(self):
    return f"{self.numerator}/{self.denominator}"
'''
Описание: Метод возвращает строковое представление дроби в виде "числитель/знаменатель".
'''
'''
7. Пример использования:
'''
fraction1 = Fraction().input_fraction()
fraction2 = Fraction().input_fraction()

if fraction1 and fraction2:
    result_sum = fraction1 + fraction2
    result_diff = fraction1 - fraction2
    result_mul = fraction1 * fraction2
    result_div = fraction1 / fraction2

    print("Сложение:", result_sum)
    print("Вычитание:", result_diff)
    print("Умножение:", result_mul)
    print("Деление:", result_div)
'''
Описание: 
Пример использования класса Fraction с вводом от пользователя. 
Создаются две дроби (fraction1 и fraction2), затем выполняются арифметические 
операции (+, -, *, /), и результаты выводятся на экран.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №3: Класс Библиотека
'''
class Library:
    def __init__(self, name, address, num_books):
        self.name = name
        self.address = address
        self.num_books = num_books

    def __add__(self, other):
        return Library(self.name, self.address, self.num_books + other)

    def __sub__(self, other):
        return Library(self.name, self.address, self.num_books - other)

    def __iadd__(self, other):
        self.num_books += other
        return self

    def __isub__(self, other):
        self.num_books -= other
        return self

    def __lt__(self, other):
        return self.num_books < other.num_books

    def __gt__(self, other):
        return self.num_books > other.num_books

    def __le__(self, other):
        return self.num_books <= other.num_books

    def __ge__(self, other):
        return self.num_books >= other.num_books

    def __eq__(self, other):
        return self.num_books == other.num_books

    def __ne__(self, other):
        return self.num_books != other.num_books

# Пример использования
library1 = Library("City Library", "123 Main St", 100)
library2 = Library("University Library", "456 College Ave", 150)

library3 = library1 + 50
library4 = library2 - 20

print(library3.num_books)  # 150
print(library4.num_books)  # 130

library1 += 30
library2 -= 10

print(library1.num_books)  # 130
print(library2.num_books)  # 140

print(library1 > library2)  # False
print(library1 < library2)  # True
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение класса Library
'''
class Library:
    def __init__(self, name, address, num_books):
        self.name = name
        self.address = address
        self.num_books = num_books
'''
Подробное описание (шаг №1):

1. Определение класса Library: 
Создается класс Library с тремя атрибутами: name (имя библиотеки), address (адрес библиотеки) и num_books (количество
книг в библиотеке).
2. Конструктор __init__: 
Конструктор принимает три параметра (name, address, num_books) и инициализирует соответствующие атрибуты объекта.
'''
'''
Шаг №2: Перегрузка операторов сложения и вычитания
'''
def __add__(self, other):
    return Library(self.name, self.address, self.num_books + other)

def __sub__(self, other):
    return Library(self.name, self.address, self.num_books - other)
'''
Подробное описание (шаг №2):

1. Перегрузка оператора сложения (__add__): 
Создает новый объект Library с тем же именем, адресом и суммой количества книг текущей библиотеки
и переданного значения (other).
2. Перегрузка оператора вычитания (__sub__): 
Создает новый объект Library с тем же именем, адресом и разностью количества книг текущей библиотеки и 
переданного значения (other).
'''
'''
Шаг №3: In-Place Операторы (+= и -=)
'''
def __iadd__(self, other):
    self.num_books += other
    return self

def __isub__(self, other):
    self.num_books -= other
    return self
'''
Подробное описание (шаг №3):

1. In-Place Оператор +=: 
Увеличивает количество книг текущей библиотеки на значение, переданное в other, и возвращает измененный объект.
2. In-Place Оператор -=: 
Уменьшает количество книг текущей библиотеки на значение, переданное в other, и возвращает измененный объект.
'''
'''
Шаг №4: Сравнительные Операторы (<, >, <=, >=, ==, !=)
'''
def __lt__(self, other):
    return self.num_books < other.num_books

def __gt__(self, other):
    return self.num_books > other.num_books

def __le__(self, other):
    return self.num_books <= other.num_books

def __ge__(self, other):
    return self.num_books >= other.num_books

def __eq__(self, other):
    return self.num_books == other.num_books

def __ne__(self, other):
    return self.num_books != other.num_books
'''
Подробное описание (шаг №4):

1. Оператор <: 
Возвращает True, если количество книг текущей библиотеки меньше количества книг библиотеки, 
переданной в other, иначе возвращает False.

2. Оператор >: 
Возвращает True, если количество книг текущей библиотеки больше количества книг библиотеки, 
переданной в other, иначе возвращает False.

3. Оператор <=: 
Возвращает True, если количество книг текущей библиотеки меньше или равно количеству книг библиотеки,
переданной в other, иначе возвращает False.

4. Оператор >=: 
Возвращает True, если количество книг текущей библиотеки больше или равно количеству книг библиотеки,
переданной в other, иначе возвращает False.

5. Оператор ==: 
Возвращает True, если количество книг текущей библиотеки равно количеству книг библиотеки, переданной в other,
иначе возвращает False.

6. Оператор !=: 
Возвращает True, если количество книг текущей библиотеки не равно количеству книг библиотеки, переданной в other,
иначе возвращает False.
'''
'''
Шаг №5: Пример использования
'''
# Пример использования
library1 = Library("City Library", "123 Main St", 100)
library2 = Library("University Library", "456 College Ave", 150)

library3 = library1 + 50
library4 = library2 - 20

print(library3.num_books)  # 150
print(library4.num_books)  # 130

library1 += 30
library2 -= 10

print(library1.num_books)  # 130
print(library2.num_books)  # 140

print(library1 > library2)  # False
print(library1 < library2)  # True
'''
Подробное описание (шаг №5):

1. Создание объектов: Создаются два объекта Library: 
library1 и library2.

2. Использование операторов сложения и вычитания: 
Создаются новые библиотеки (library3 и library4) с использованием операторов + и -.

3. Использование in-place операторов: 
Изменяются текущие библиотеки (library1 и library2) с использованием in-place операторов += и -=.

4. Вывод результатов: 
Выводится количество книг в новых и измененных библиотеках.

5. Сравнение библиотек: 
Выполняются сравнения количества книг в library1 и library2, и результаты выводятся.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №4: Класс Date
'''
from datetime import datetime


class Date:
    def __init__(self, day=1, month=1, year=2000):
        self.day = day
        self.month = month
        self.year = year

    @classmethod
    def from_input(cls):
        try:
            day = int(input("Введите день: "))
            month = int(input("Введите месяц: "))
            year = int(input("Введите год: "))

            if not (1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100):
                raise ValueError("Некорректная дата.")

            return cls(day, month, year)
        except ValueError as e:
            print(f"Ошибка: {e}. Пожалуйста, введите корректные значения.")
            return None

    @classmethod
    def from_string(cls, date_string, date_format="%d.%m.%Y"):
        parsed_date = datetime.strptime(date_string, date_format)
        return cls(day=parsed_date.day, month=parsed_date.month, year=parsed_date.year)

    def __sub__(self, other):
        if isinstance(other, Date):
            days_in_self = self.to_days()
            days_in_other = other.to_days()
            difference_in_days = days_in_self - days_in_other

            return difference_in_days
        else:
            raise TypeError("Unsupported operand type for -: {}".format(type(other)))

    def add_days(self, days):
        if days < 0:
            raise ValueError("Количество дней должно быть положительным числом.")

        days_in_self = self.to_days()
        new_days = days_in_self + days

        return self.from_days(new_days)

    def subtract_months(self, months):
        # Реализация вычитания месяцев
        pass

    def subtract_years(self, years):
        # Реализация вычитания лет
        pass

    @classmethod
    def current_date(cls):
        today = datetime.today()
        return cls(day=today.day, month=today.month, year=today.year)

    def __eq__(self, other):
        if isinstance(other, Date):
            return (self.day, self.month, self.year) == (other.day, other.month, other.year)
        return False

    def to_days(self):
        days = self.day
        for m in range(1, self.month):
            days += self.days_in_month(m, self.year)
        for y in range(1900, self.year):
            days += self.days_in_year(y)
        return days

    def from_days(self, days):
        year = 1900
        while days > self.days_in_year(year):
            days -= self.days_in_year(year)
            year += 1

        month = 1
        while days > self.days_in_month(month, year):
            days -= self.days_in_month(month, year)
            month += 1

        return Date(day=days, month=month, year=year)

    def days_in_month(self, month, year):
        if 1 <= month <= 12:
            if month in {1, 3, 5, 7, 8, 10, 12}:
                return 31
            elif month in {4, 6, 9, 11}:
                return 30
            elif month == 2 and self.is_leap_year(year):
                return 29
            elif month == 2:
                return 28
        raise ValueError("Некорректный месяц.")

    def days_in_year(self, year):
        return 366 if self.is_leap_year(year) else 365

    def is_leap_year(self, year):
        return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)

    def __str__(self):
        return f"{self.day:02d}.{self.month:02d}.{self.year}"


# Пример использования с вводом от пользователя
date1 = Date.from_input()
date2 = Date.from_input()

if date1 and date2:
    days_difference = date1 - date2
    print("Разница в днях:", days_difference)

    try:
        days_to_add = int(input("Введите количество дней для добавления: "))
        new_date = date1.add_days(days_to_add)
        print("Новая дата:", new_date)
    except ValueError as e:
        print(f"Ошибка: {e}. Пожалуйста, введите корректное количество дней.")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение класса Date
'''
class Date:
    def __init__(self, day=1, month=1, year=2000):
        self.day = day
        self.month = month
        self.year = year
'''
Подробное описание (шаг №1):

1. Определение класса Date: Создается класс Date, представляющий информацию о дате. 
Конструктор __init__ принимает три параметра (день, месяц, год) и инициализирует соответствующие атрибуты объекта.
'''
'''
Шаг №2: Два метода класса Date с использованием декоратора @classmethod
Метод from_input
'''


@classmethod
def from_input(cls):
    try:
        day = int(input("Введите день: "))
        month = int(input("Введите месяц: "))
        year = int(input("Введите год: "))

        if not (1 <= day <= 31 and 1 <= month <= 12 and 1900 <= year <= 2100):
            raise ValueError("Некорректная дата.")

        return cls(day, month, year)
    except ValueError as e:
        print(f"Ошибка: {e}. Пожалуйста, введите корректные значения.")
        return None
'''
Метод from_string
'''


@classmethod
def from_string(cls, date_string, date_format="%d.%m.%Y"):
    parsed_date = datetime.strptime(date_string, date_format)
    return cls(day=parsed_date.day, month=parsed_date.month, year=parsed_date.year)
'''
Подробное описание (шаг №2):

1. Метод from_input: 
Этот метод использует декоратор @classmethod и предназначен для создания объекта Date на основе ввода пользователя.
Пользователь вводит день, месяц и год, и метод проверяет корректность введенных данных.
2. Метод from_string: 
Этот метод также использует декоратор @classmethod и создает объект Date на основе строки с датой (date_string).
Строка парсится с использованием библиотеки datetime, а затем извлекаются день, месяц и год для создания объекта Date.
'''
'''
Шаг №3: Перегрузка оператора вычитания (__sub__)
'''


def __sub__(self, other):
    if isinstance(other, Date):
        days_in_self = self.to_days()
        days_in_other = other.to_days()
        difference_in_days = days_in_self - days_in_other

        return difference_in_days
    else:
        raise TypeError("Unsupported operand type for -: {}".format(type(other)))
'''
Подробное описание (шаг №3):

1. Перегрузка оператора вычитания (__sub__): 
Если операнд (other) является объектом класса Date, то выполняется вычисление разницы в днях между двумя датами с
использованием методов to_days. Возвращается результат.
2. Если операнд не является объектом класса Date, генерируется исключение TypeError.
'''
'''
Шаг №4: Метод для увеличения даты на определенное количество дней (add_days)
'''
def add_days(self, days):
    if days < 0:
        raise ValueError("Количество дней должно быть положительным числом.")

    days_in_self = self.to_days()
    new_days = days_in_self + days

    return self.from_days(new_days)
'''
Подробное описание (шаг №4):

Метод add_days: 
1. Принимает количество дней (days) для увеличения текущей даты. 
Проверяет, что количество дней положительно.
2. Вычисляет количество дней, представляющих текущую дату (days_in_self), 
добавляет к ним переданное количество дней и создает новый объект Date с использованием метода from_days.
'''
'''
Шаг №5: Два метода для вычитания месяцев и лет (subtract_months и subtract_years)
'''
def subtract_months(self, months):
    # Реализация вычитания месяцев
    pass


def subtract_years(self, years):
    # Реализация вычитания лет
    pass
'''
Подробное описание (шаг №5):

1. Метод subtract_months: 
Метод, предположительно предназначенный для вычитания месяцев из текущей даты. 
В данный момент не имеет конкретной реализации (помечен pass).
2. Метод subtract_years: 
Метод, предположительно предназначенный для вычитания лет из текущей даты. 
В данный момент не имеет конкретной реализации (помечен pass).
'''
'''
Шаг №6: Метод класса для получения текущей даты (current_date)
'''
@classmethod
def current_date(cls):
    today = datetime.today()
    return cls(day=today.day, month=today.month, year=today.year)
'''
Подробное описание (шаг №6):

1. Метод current_date: 
Этот метод использует декоратор @classmethod и возвращает объект Date, представляющий текущую дату.
'''
'''
Шаг №7: Перегрузка оператора сравнения (__eq__)
'''
def __eq__(self, other):
    if isinstance(other, Date):
        return (self.day, self.month, self.year) == (other.day, other.month, other.year)
    return False
'''
Подробное описание (шаг №7):

1. Перегрузка оператора сравнения (__eq__): 
Если операнд (other) является объектом класса Date, то выполняется сравнение по дням, месяцам и годам. 
Возвращается булево значение.
'''
'''
Шаг №8: Три метода для работы с днями, месяцами и годами
'''


def to_days(self):
    days = self.day
    for m in range(1, self.month):
        days += self.days_in_month(m, self.year)
    for y in range(1900, self.year):
        days += self.days_in_year(y)
    return days


def from_days(self, days):
    year = 1900
    while days > self.days_in_year(year):
        days -= self.days_in_year(year)
        year += 1

    month = 1
    while days > self.days_in_month(month, year):
        days -= self.days_in_month(month, year)
        month += 1

    return Date(day=days, month=month, year=year)
'''
Подробное описание (шаг №8):

1. Метод to_days: Вычисляет и возвращает количество дней, представляющих текущую дату, учитывая дни, месяцы и годы.
2. Метод from_days: Создает новый объект Date на основе переданного количества дней.
3. Методы days_in_month и days_in_year: Определяют количество дней в месяце и году соответственно.
'''
'''
Шаг №9: Два метода для проверки високосного года и строкового представления объекта (__str__)
'''
def is_leap_year(self, year):
    return (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)
def __str__(self):
    return f"{self.day:02d}.{self.month:02d}.{self.year}"
'''
Подробное описание (шаг №9):

1. Метод is_leap_year: Проверяет, является ли год високосным.
2. Метод __str__: Возвращает строковое представление объекта Date в формате "день.месяц.год".
'''
'''
Шаг №10: Пример использования с вводом от пользователя
'''
# Пример использования с вводом от пользователя
date1 = Date.from_input()
date2 = Date.from_input()

if date1 and date2:
    days_difference = date1 - date2
    print("Разница в днях:", days_difference)

    try:
        days_to_add = int(input("Введите количество дней для добавления: "))
        new_date = date1.add_days(days_to_add)
        print("Новая дата:", new_date)
    except ValueError as e:
        print(f"Ошибка: {e}. Пожалуйста, введите корректное количество дней.")
'''
Подробное описание (шаг №10):

1. Пример использования с вводом от пользователя: 
Создаются два объекта Date (date1 и date2) с использованием метода from_input, 
который запрашивает у пользователя ввод дня, месяца и года.
2. Вычисление разницы в днях и добавление дней: 
Вычисляется разница в днях между двумя датами. 
Пользователь вводит количество дней для добавления к первой дате, и выводится новая дата после добавления.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
Дата выполнения ДОМАШНЕЙ РАБОТЫ: 27-28 - ДЕКАБРЯ 2023
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 27.12.2023
Домашняя работа №20: ООП.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №1
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс Circle (окружность). Для данного класса реализуйте ряд перегруженных операторов:
Проверка на равенство радиусов двух окружностей (операция = =);
Сравнения длин двух окружностей (операции >, <, <=,>=);
Пропорциональное изменение размеров окружности, путем изменения ее радиуса (операции + - += -=).
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задание №2
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс Complex (комплексное число).
Создайте перегруженные операторы для реализации арифметических операций для по работе с комплексными числами
(операции +, -, *, /).
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задание №3
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Вам необходимо создать класс Airplane (самолет).
С помощью перегрузки операторов реализовать:
Проверка на равенство типов самолетов (операция = =);
Увеличение и уменьшение пассажиров в салоне самолета (операции + - += -=);
Сравнение двух самолетов по максимально возможному количеству пассажиров на борту (операции > < <= >=).
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задание №4
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создать класс Flat (квартира). Реализовать перегруженные операторы:
Проверка на равенство площадей квартир (операция ==);
Проверка на неравенство площадей квартир (операция !=);
Сравнение двух квартир по цене (операции > < <= >=).
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Далее решение этих заданий ↑   ↑   ↑   ↑
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №1: Класс Circle
'''
class Circle:
    def __init__(self, radius):
        self.radius = radius

    def __eq__(self, other):
        if isinstance(other, Circle):
            return self.radius == other.radius
        return False

    def __gt__(self, other):
        if isinstance(other, Circle):
            return self.radius > other.radius
        return NotImplemented

    def __lt__(self, other):
        if isinstance(other, Circle):
            return self.radius < other.radius
        return NotImplemented

    def __le__(self, other):
        if isinstance(other, Circle):
            return self.radius <= other.radius
        return NotImplemented

    def __ge__(self, other):
        if isinstance(other, Circle):
            return self.radius >= other.radius
        return NotImplemented

    def __add__(self, value):
        if isinstance(value, (int, float)):
            return Circle(self.radius + value)
        return NotImplemented

    def __sub__(self, value):
        if isinstance(value, (int, float)):
            return Circle(self.radius - value)
        return NotImplemented

    def __iadd__(self, value):
        if isinstance(value, (int, float)):
            self.radius += value
            return self
        return NotImplemented

    def __isub__(self, value):
        if isinstance(value, (int, float)):
            self.radius -= value
            return self
        return NotImplemented
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение класса Circle
'''
class Circle:
    def __init__(self, radius):
        self.radius = radius
'''
Подробное описание (шаг №1):

1. Определение класса Circle: Создается класс Circle, представляющий круг.
Конструктор __init__ принимает радиус круга и инициализирует соответствующий атрибут объекта.
'''
'''
Шаг №2: Перегрузка операторов сравнения (__eq__, __gt__, __lt__, __le__, __ge__)
'''
def __eq__(self, other):
    if isinstance(other, Circle):
        return self.radius == other.radius
    return False

def __gt__(self, other):
    if isinstance(other, Circle):
        return self.radius > other.radius
    return NotImplemented

def __lt__(self, other):
    if isinstance(other, Circle):
        return self.radius < other.radius
    return NotImplemented

def __le__(self, other):
    if isinstance(other, Circle):
        return self.radius <= other.radius
    return NotImplemented

def __ge__(self, other):
    if isinstance(other, Circle):
        return self.radius >= other.radius
    return NotImplemented
'''
Подробное описание (шаг №2):

1. Перегрузка оператора равенства (__eq__): 
Сравнивает два объекта Circle по радиусу. Возвращает True, если радиусы равны, и False в противном случае.
2. Перегрузка операторов сравнения (__gt__, __lt__, __le__, __ge__): 
Сравнивают два объекта Circle по радиусу и возвращают соответствующие булевы значения (True/False). 
Если операнд не является объектом Circle, возвращается NotImplemented.
'''
'''
Шаг №3: Перегрузка операторов сложения и вычитания (__add__, __sub__)
'''
def __add__(self, value):
    if isinstance(value, (int, float)):
        return Circle(self.radius + value)
    return NotImplemented

def __sub__(self, value):
    if isinstance(value, (int, float)):
        return Circle(self.radius - value)
    return NotImplemented
'''
Подробное описание (шаг №3):

1. Перегрузка оператора сложения (__add__): 
Позволяет сложить объект Circle с числовым значением (int или float). 
Возвращает новый объект Circle с радиусом, увеличенным на значение операнда.
2. Перегрузка оператора вычитания (__sub__): 
Позволяет вычесть из объекта Circle числовое значение (int или float). 
Возвращает новый объект Circle с радиусом, уменьшенным на значение операнда.
'''
'''
Шаг №4: Перегрузка операторов присваивания с добавлением и вычитанием (__iadd__, __isub__)
'''
def __iadd__(self, value):
    if isinstance(value, (int, float)):
        self.radius += value
        return self
    return NotImplemented

def __isub__(self, value):
    if isinstance(value, (int, float)):
        self.radius -= value
        return self
    return NotImplemented
'''
Подробное описание (шаг №4):

1. Перегрузка оператора присваивания с добавлением (__iadd__):
Позволяет добавить к радиусу объекта Circle числовое значение (int или float). 
Изменяет текущий объект и возвращает его.
2. Перегрузка оператора присваивания с вычитанием (__isub__):
Позволяет вычесть из радиуса объекта Circle числовое значение (int или float). 
Изменяет текущий объект и возвращает его.
'''
'''
Общее использование:
'''
# Пример использования
circle1 = Circle(5)
circle2 = Circle(3)

result_eq = circle1 == circle2
result_gt = circle1 > circle2
result_lt = circle1 < circle2
result_add = circle1 + 2
result_sub = circle1 - 1

print(result_eq)  # False
print(result_gt)  # True
print(result_lt)  # False
print(result_add.radius)  # 7
print(result_sub.radius)  # 4

circle1 += 3
circle2 -= 1

print(circle1.radius)  # 8
print(circle2.radius)  # 2
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №2: Класс Complex
'''
class Complex:
    def __init__(self, real, imag):
        self.real = real
        self.imag = imag

    def __add__(self, other):
        if isinstance(other, Complex):
            return Complex(self.real + other.real, self.imag + other.imag)
        return NotImplemented

    def __sub__(self, other):
        if isinstance(other, Complex):
            return Complex(self.real - other.real, self.imag - other.imag)
        return NotImplemented

    def __mul__(self, other):
        if isinstance(other, Complex):
            real_part = self.real * other.real - self.imag * other.imag
            imag_part = self.real * other.imag + self.imag * other.real
            return Complex(real_part, imag_part)
        return NotImplemented

    def __truediv__(self, other):
        if isinstance(other, Complex):
            denominator = other.real**2 + other.imag**2
            real_part = (self.real * other.real + self.imag * other.imag) / denominator
            imag_part = (self.imag * other.real - self.real * other.imag) / denominator
            return Complex(real_part, imag_part)
        return NotImplemented
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Определение класса Complex
'''
class Complex:
    def __init__(self, real, imag):
        self.real = real
        self.imag = imag
'''
Подробное описание (шаг №1):

1. Определение класса Complex: 
Создается класс Complex, представляющий комплексное число. 
Конструктор __init__ принимает два аргумента: real (действительная часть) и imag (мнимая часть), и инициализирует
соответствующие атрибуты объекта.
'''
'''
Шаг №2: Перегрузка оператора сложения (__add__)
'''
def __add__(self, other):
    if isinstance(other, Complex):
        return Complex(self.real + other.real, self.imag + other.imag)
    return NotImplemented
'''
Подробное описание (шаг №2):

1. Перегрузка оператора сложения (__add__): 
Позволяет складывать два комплексных числа. Если второй операнд является объектом Complex, возвращается новый объект
Complex с суммой соответствующих действительных и мнимых частей. В случае, если второй операнд не является 
объектом Complex, возвращается NotImplemented.
'''
'''
Шаг №3: Перегрузка оператора вычитания (__sub__)
'''
def __sub__(self, other):
    if isinstance(other, Complex):
        return Complex(self.real - other.real, self.imag - other.imag)
    return NotImplemented
'''
Подробное описание (шаг №3):

1. Перегрузка оператора вычитания (__sub__): 
Позволяет вычитать из одного комплексного числа другое. 
Если второй операнд является объектом Complex, возвращается новый объект Complex с разностью соответствующих 
действительных и мнимых частей. В случае, если второй операнд не является объектом Complex, 
возвращается NotImplemented.
'''
'''
Шаг №4: Перегрузка оператора умножения (__mul__)
'''
def __mul__(self, other):
    if isinstance(other, Complex):
        real_part = self.real * other.real - self.imag * other.imag
        imag_part = self.real * other.imag + self.imag * other.real
        return Complex(real_part, imag_part)
    return NotImplemented
'''
Подробное описание (шаг №4):

1. Перегрузка оператора умножения (__mul__): 
Позволяет умножать два комплексных числа. Если второй операнд является объектом Complex, вычисляются действительная
и мнимая части произведения, и возвращается новый объект Complex с этими значениями. В случае, если второй операнд не
является объектом Complex, возвращается NotImplemented.
'''
'''
Шаг №5: Перегрузка оператора деления (__truediv__)
'''
def __truediv__(self, other):
    if isinstance(other, Complex):
        denominator = other.real ** 2 + other.imag ** 2
        real_part = (self.real * other.real + self.imag * other.imag) / denominator
        imag_part = (self.imag * other.real - self.real * other.imag) / denominator
        return Complex(real_part, imag_part)
    return NotImplemented
'''
Подробное описание (шаг №5):

1. Перегрузка оператора деления (__truediv__): 
Позволяет делить одно комплексное число на другое. Если второй операнд является объектом Complex, вычисляются
действительная и мнимая части частного, и возвращается новый объект Complex с этими значениями. 
В случае, если второй операнд не является объектом Complex, возвращается NotImplemented.
'''
'''
Общее использование:
'''
# Пример использования
complex1 = Complex(2, 3)
complex2 = Complex(1, 2)

result_add = complex1 + complex2
result_sub = complex1 - complex2
result_mul = complex1 * complex2
result_div = complex1 / complex2

print(result_add.real, "+", result_add.imag, "i")  # 3 + 5i
print(result_sub.real, "-", result_sub.imag, "i")  # 1 + 1i
print(result_mul.real, "+", result_mul.imag, "i")  # -4 + 7i
print(result_div.real, "+", result_div.imag, "i")  # 1.6 - 0.2i
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №3: Класс Airplane
'''
class Airplane:
    def __init__(self, passengers, max_passengers):
        self.passengers = passengers
        self.max_passengers = max_passengers

    def __eq__(self, other):
        if isinstance(other, Airplane):
            return self.max_passengers == other.max_passengers
        return False

    def __gt__(self, other):
        if isinstance(other, Airplane):
            return self.max_passengers > other.max_passengers
        return NotImplemented

    def __lt__(self, other):
        if isinstance(other, Airplane):
            return self.max_passengers < other.max_passengers
        return NotImplemented

    def __le__(self, other):
        if isinstance(other, Airplane):
            return self.max_passengers <= other.max_passengers
        return NotImplemented

    def __ge__(self, other):
        if isinstance(other, Airplane):
            return self.max_passengers >= other.max_passengers
        return NotImplemented

    def __add__(self, value):
        if isinstance(value, int):
            return Airplane(self.passengers + value, self.max_passengers)
        return NotImplemented

    def __sub__(self, value):
        if isinstance(value, int):
            return Airplane(self.passengers - value, self.max_passengers)
        return NotImplemented

    def __iadd__(self, value):
        if isinstance(value, int):
            self.passengers += value
            return self
        return NotImplemented

    def __isub__(self, value):
        if isinstance(value, int):
            self.passengers -= value
            return self
        return NotImplemented

# Пример использования
airplane1 = Airplane(50, 100)
airplane2 = Airplane(75, 150)

# Примеры операций
print(airplane1 == airplane2)  # False
print(airplane1 > airplane2)   # False
print(airplane1 < airplane2)   # True
print(airplane1 + 20)          # Airplane(passengers=70, max_passengers=100)
print(airplane2 - 10)          # Airplane(passengers=65, max_passengers=150)

# Примеры операций с присваиванием
airplane1 += 30
airplane2 -= 15

print(airplane1.passengers)    # 80
print(airplane2.passengers)    # 60
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1:
'''
# Определение класса Airplane
class Airplane:
    # Инициализация объекта класса
    def __init__(self, passengers, max_passengers):
        self.passengers = passengers  # Количество пассажиров
        self.max_passengers = max_passengers  # Максимальное количество пассажиров
'''
Описание: 
В этом шаге создается класс Airplane, представляющий объект "самолет". 
Класс имеет два атрибута: passengers (количество пассажиров) и max_passengers (максимальное количество пассажиров).
'''
'''
Шаг 2:
'''
# Перегрузка оператора равенства
def __eq__(self, other):
    if isinstance(other, Airplane):
        return self.max_passengers == other.max_passengers
    return False
'''
Описание: 
В этом шаге добавляется перегрузка оператора равенства (__eq__). 
Метод сравнивает максимальное количество пассажиров текущего самолета с максимальным количеством пассажиров 
другого самолета и возвращает True, если они равны, иначе False.
'''
'''
Шаг 3:
'''
# Перегрузка оператора больше
def __gt__(self, other):
    if isinstance(other, Airplane):
        return self.max_passengers > other.max_passengers
    return NotImplemented
'''
Описание:
В этом шаге добавляется перегрузка оператора больше (__gt__). 
Метод сравнивает максимальное количество пассажиров текущего самолета с максимальным количеством пассажиров другого
самолета и возвращает True, если у текущего самолета больше мест, иначе False.
'''
'''
Шаг 4:
'''
# Перегрузка оператора меньше
def __lt__(self, other):
    if isinstance(other, Airplane):
        return self.max_passengers < other.max_passengers
    return NotImplemented
'''
Описание: 
В этом шаге добавляется перегрузка оператора меньше (__lt__). 
Метод сравнивает максимальное количество пассажиров текущего самолета с максимальным количеством пассажиров другого
самолета и возвращает True, если у текущего самолета меньше мест, иначе False.
'''
'''
Шаг 5:
'''
# Перегрузка оператора меньше или равно
def __le__(self, other):
    if isinstance(other, Airplane):
        return self.max_passengers <= other.max_passengers
    return NotImplemented
'''
Описание: 
В этом шаге добавляется перегрузка оператора меньше или равно (__le__). 
Метод сравнивает максимальное количество пассажиров текущего самолета с максимальным количеством пассажиров другого
самолета и возвращает True, если у текущего самолета меньше или равно мест, иначе False.
'''
'''
Шаг 6:
'''
# Перегрузка оператора больше или равно
def __ge__(self, other):
    if isinstance(other, Airplane):
        return self.max_passengers >= other.max_passengers
    return NotImplemented
'''
Описание: 
В этом шаге добавляется перегрузка оператора больше или равно (__ge__). 
Метод сравнивает максимальное количество пассажиров текущего самолета с максимальным количеством пассажиров другого
самолета и возвращает True, если у текущего самолета больше или равно мест, иначе False.
'''
'''
Шаг 7:
'''
# Перегрузка оператора сложения
def __add__(self, value):
    if isinstance(value, int):
        return Airplane(self.passengers + value, self.max_passengers)
    return NotImplemented
'''
Описание: 
В этом шаге добавляется перегрузка оператора сложения (__add__). 
Метод позволяет добавлять целочисленное значение к количеству пассажиров текущего самолета, создавая новый самолет
с обновленным количеством пассажиров.
'''
'''
Шаг 8:
'''
# Перегрузка оператора вычитания
def __sub__(self, value):
    if isinstance(value, int):
        return Airplane(self.passengers - value, self.max_passengers)
    return NotImplemented
'''
Описание: 
В этом шаге добавляется перегрузка оператора вычитания (__sub__). Метод позволяет вычитать целочисленное 
значение из количества пассажиров текущего самолета, создавая новый самолет с обновленным количеством пассажиров.
'''
'''
Шаг 9:
'''
# Перегрузка оператора присваивания суммы
def __iadd__(self, value):
    if isinstance(value, int):
        self.passengers += value
        return self
    return NotImplemented
'''
Описание: 
В этом шаге добавляется перегрузка оператора присваивания суммы (__iadd__). Метод позволяет присваивать
сумму целочисленного значения к количеству пассажиров текущего самолета, обновляя его в процессе.
'''
'''
Шаг 10:
'''
# Перегрузка оператора присваивания разности
def __isub__(self, value):
    if isinstance(value, int):
        self.passengers -= value
        return self
    return NotImplemented
'''
Описание: 
В этом шаге добавляется перегрузка оператора присваивания разности (__isub__).
Метод позволяет присваивать разность целочисленного значения из количества пассажиров текущего самолета,
обновляя его в процессе.
'''
'''
Шаг 11:
'''
# Пример использования
airplane1 = Airplane(50, 100)
airplane2 = Airplane(75, 150)

# Примеры операций
print(airplane1 == airplane2)  # False
print(airplane1 > airplane2)   # False
print(airplane1 < airplane2)   # True
print(airplane1 + 20)          # Airplane(passengers=70, max_passengers=100)
print(airplane2 - 10)          # Airplane(passengers=65, max_passengers=150)

# Примеры операций с присваиванием
airplane1 += 30
airplane2 -= 15

print(airplane1.passengers)    # 80
print(airplane2.passengers)    # 60
'''
Описание:
Приведен пример использования класса Airplane с различными операторами и операциями с присваиванием.
Создаются два объекта Airplane и выполняются операции равенства, сравнения, сложения и вычитания, а также операции
с присваиванием для изменения количества пассажиров.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №4: Класс Flat
'''
class Flat:
    def __init__(self, area, price):
        self.area = area
        self.price = price

    def __eq__(self, other):
        if isinstance(other, Flat):
            return self.area == other.area
        return False

    def __ne__(self, other):
        if isinstance(other, Flat):
            return self.area != other.area
        return NotImplemented

    def __gt__(self, other):
        if isinstance(other, Flat):
            return self.price > other.price
        return NotImplemented

    def __lt__(self, other):
        if isinstance(other, Flat):
            return self.price < other.price
        return NotImplemented

    def __le__(self, other):
        if isinstance(other, Flat):
            return self.price <= other.price
        return NotImplemented

    def __ge__(self, other):
        if isinstance(other, Flat):
            return self.price >= other.price
        return NotImplemented
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1:
'''
class Flat:
''''
Описание: Создается класс Flat, представляющий объект "квартира".
'''''
'''
Шаг 2:
'''
def __init__(self, area, price):
    self.area = area
    self.price = price
'''
Описание: 
В конструкторе класса инициализируются два атрибута area (площадь квартиры) и price (цена квартиры). 
Эти атрибуты будут храниться в каждом созданном объекте класса.
'''
'''
Шаг 3:
'''
def __eq__(self, other):
    if isinstance(other, Flat):
        return self.area == other.area
    return False
'''
Описание: 
Перегружается оператор равенства (==). 
Метод сравнивает объект с другим объектом типа Flat по значению атрибута area. 
Возвращает True, если атрибуты area обоих объектов равны, и False в противном случае.
'''
'''
Шаг 4:
'''
def __ne__(self, other):
    if isinstance(other, Flat):
        return self.area != other.area
    return NotImplemented
'''
Описание: 
Перегружается оператор неравенства (!=). 
Метод сравнивает объект с другим объектом типа Flat по значению атрибута area. 
Возвращает True, если атрибуты area обоих объектов не равны, и False в противном случае.
'''
'''
Шаг 5:
'''
def __gt__(self, other):
    if isinstance(other, Flat):
        return self.price > other.price
    return NotImplemented
'''
Описание:
Перегружается оператор больше (>). 
Метод сравнивает объект с другим объектом типа Flat по значению атрибута price. 
Возвращает True, если атрибут price текущего объекта больше, чем у другого объекта, и False в противном случае.
'''
'''
Шаг 6:
'''
def __lt__(self, other):
    if isinstance(other, Flat):
        return self.price < other.price
    return NotImplemented
'''
Описание: 
Перегружается оператор меньше (<). 
Метод сравнивает объект с другим объектом типа Flat по значению атрибута price. 
Возвращает True, если атрибут price текущего объекта меньше, чем у другого объекта, и False в противном случае.
'''
'''
Шаг 7:
'''
def __le__(self, other):
    if isinstance(other, Flat):
        return self.price <= other.price
    return NotImplemented
'''
Описание: 
Перегружается оператор меньше или равно (<=). 
Метод сравнивает объект с другим объектом типа Flat по значению атрибута price. 
Возвращает True, если атрибут price текущего объекта меньше или равен, чем у другого объекта, и False в противном случае.
'''
'''
Шаг 8:
'''
def __ge__(self, other):
    if isinstance(other, Flat):
        return self.price >= other.price
    return NotImplemented
'''
Описание: 
Перегружается оператор больше или равно (>=). 
Метод сравнивает объект с другим объектом типа Flat по значению атрибута price. 
Возвращает True, если атрибут price текущего объекта больше или равен, чем у другого объекта, и False в противном случае.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
Классная работа от 29 декабря 2023 года.

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения практической работы: 29-30 - ДЕКАБРЯ 2023
''''
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 29.12.2023
Практическая работа №21: ООП.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующие задания:
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №1
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Работаем с примером, приведенным на практике. 
Мы можем менять значения полей dia и h объекта за пределами класса простым присваиванием (например, a.dia = 10).
При этом площадь никак не будет пересчитываться. Также мы можем назначить новое значение для площади, 
как простым присваиванием, так и вызовом функции make_area() с последующим присваиванием. 
Например, a.area = a.make_area(2, 3). При этом не меняются высота и диаметр.

Защитите код от возможных логических ошибок следующим образом:
Свойствам dia и h объекта по-прежнему можно выполнять присваивание за пределами класса. 
Однако при этом "за кулисами" происходит пересчет площади, т. е. изменение значения area.
Свойству area нельзя присваивать за пределами класса. Можно только получать его значение.

Подсказка: 
вспомните про метод __setattr__(), упомянутый в уроке про инкапсуляцию.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №2
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс для подсчета площади геометрических фигур.
Класс должен предоставлять функциональность для подсчета площади треугольника по разным формулам, 
площади прямоугольника, площади квадрата, площади ромба. Методы класса для подсчета площади должны быть реализованы
с помощью статических методов. Также класс должен считать количество подсчетов площади и возвращать это значение 
с помощью статического метода.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Задания №3
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс для подсчета максимума из четырех аргументов, минимума из четырех аргументов,
среднеарифметического из четырех аргументов, факториала аргумента. 
Функциональность необходимо реализовать в виде статических методов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Решение этих заданий ↑   ↑   ↑   ↑
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №1:
'''


class Cylinder:
    def __init__(self, dia, h):
        # Инициализация объекта с заданными значениями диаметра и высоты
        self._dia = dia
        self._h = h
        # Вычисление и сохранение площади поверхности
        self._area = self.make_area()

    def make_area(self):
        # Метод для вычисления площади поверхности цилиндра
        return 3.14 * (self._dia / 2) ** 2 * self._h

    def __setattr__(self, name, value):
        # Переопределение метода для контроля изменений атрибутов
        if name in ('dia', 'h'):
            # Если изменяются атрибуты dia или h, обновляем их значения
            object.__setattr__(self, name, value)
            # Пересчитываем площадь поверхности при изменении dia или h
            self._area = self.make_area()
        elif name == 'area':
            # Если пытаются изменить атрибут area, возбуждаем исключение
            raise AttributeError("Can't set attribute 'area'")
        else:
            # Для других атрибутов вызываем базовый метод __setattr__
            object.__setattr__(self, name, value)

    @property
    def area(self):
        # Свойство для доступа к значению площади поверхности (только чтение)
        return self._area

    @property
    def dia(self):
        # Свойство для доступа к значению диаметра (только чтение)
        return self._dia

    @dia.setter
    def dia(self, value):
        # Сеттер для свойства dia, обновляет dia и пересчитывает площадь
        self._dia = value
        self._area = self.make_area()

    @property
    def h(self):
        # Свойство для доступа к значению высоты (только чтение)
        return self._h

    @h.setter
    def h(self, value):
        # Сеттер для свойства h, обновляет h и пересчитывает площадь
        self._h = value
        self._area = self.make_area()


# Функция для ввода числа с проверкой на корректность
def input_float(prompt):
    while True:
        try:
            return float(input(prompt))
        except ValueError:
            print("Пожалуйста, введите корректное число.")


# Пример использования с вводом данных с клавиатуры
dia_input = input_float("Введите диаметр цилиндра: ")
h_input = input_float("Введите высоту цилиндра: ")

# Создание объекта цилиндра с введенными значениями
c = Cylinder(dia=dia_input, h=h_input)
# Вывод площади поверхности цилиндра с двумя знаками после запятой
print(f"Площадь поверхности цилиндра: {c.area:.2f}")

# Изменение диаметра с клавиатуры
new_dia = input_float("Введите новый диаметр цилиндра: ")
# Изменение диаметра и автоматический пересчет площади
c.dia = new_dia
# Вывод обновленной площади поверхности цилиндра с двумя знаками после запятой
print(f"Обновленная площадь поверхности цилиндра: {c.area:.2f}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Инициализация класса Cylinder
'''


class Cylinder:
    def __init__(self, dia, h):
        # Инициализация объекта с заданными значениями диаметра и высоты
        self._dia = dia
        self._h = h
        # Вычисление и сохранение площади поверхности
        self._area = self.make_area()


'''
В этом шаге определен конструктор __init__, который инициализирует объект класса Cylinder заданными значениями диаметра
(dia) и высоты (h). Затем вычисляется и сохраняется площадь поверхности с использованием метода make_area().
'''
'''
Шаг №2: Метод make_area()
'''


def make_area(self):
    # Метод для вычисления площади поверхности цилиндра
    return 3.14 * (self._dia / 2) ** 2 * self._h


'''
Метод make_area() вычисляет и возвращает площадь поверхности цилиндра на основе текущих значений диаметра и высоты.
'''
'''
Шаг №3: Переопределение __setattr__
'''


def __setattr__(self, name, value):
    # Переопределение метода для контроля изменений атрибутов
    if name in ('dia', 'h'):
        # Если изменяются атрибуты dia или h, обновляем их значения
        object.__setattr__(self, name, value)
        # Пересчитываем площадь поверхности при изменении dia или h
        self._area = self.make_area()
    elif name == 'area':
        # Если пытаются изменить атрибут area, возбуждаем исключение
        raise AttributeError("Can't set attribute 'area'")
    else:
        # Для других атрибутов вызываем базовый метод __setattr__
        object.__setattr__(self, name, value)


'''
Этот шаг переопределяет метод __setattr__, который вызывается при попытке изменения атрибутов объекта.
Если изменяются атрибуты dia или h, их значения обновляются, и затем пересчитывается площадь поверхности.
Если пытаются изменить атрибут area, возбуждается исключение.
'''
'''
Шаг №4-6: Свойства для доступа к атрибутам
'''


@property
def area(self):
    # Свойство для доступа к значению площади поверхности (только чтение)
    return self._area


@property
def dia(self):
    # Свойство для доступа к значению диаметра (только чтение)
    return self._dia


@dia.setter
def dia(self, value):
    # Сеттер для свойства dia, обновляет dia и пересчитывает площадь
    self._dia = value
    self._area = self.make_area()


@property
def h(self):
    # Свойство для доступа к значению высоты (только чтение)
    return self._h


@h.setter
def h(self, value):
    # Сеттер для свойства h, обновляет h и пересчитывает площадь
    self._h = value
    self._area = self.make_area()


'''
Эти шаги определяют свойства (area, dia, h) с использованием декораторов @property, которые обеспечивают доступ
только для чтения. Для атрибутов dia и h также определены сеттеры (@dia.setter и @h.setter), которые обновляют
соответствующий атрибут и пересчитывают площадь.
'''
'''
Пример использования:
'''
# Пример использования с вводом данных с клавиатуры
dia_input = input_float("Введите диаметр цилиндра: ")
h_input = input_float("Введите высоту цилиндра: ")

# Создание объекта цилиндра с введенными значениями
c = Cylinder(dia=dia_input, h=h_input)
# Вывод площади поверхности цилиндра с двумя знаками после запятой
print(f"Площадь поверхности цилиндра: {c.area:.2f}")

# Изменение диаметра с клавиатуры
new_dia = input_float("Введите новый диаметр цилиндра: ")
# Изменение диаметра и автоматический пересчет площади
c.dia = new_dia
# Вывод обновленной площади поверхности цилиндра с двумя знаками после запятой
print(f"Обновленная площадь поверхности цилиндра: {c.area:.2f}")
'''
Этот пример иллюстрирует использование созданного класса Cylinder и подтверждает,
что изменения диаметра автоматически пересчитывают площадь поверхности цилиндра, как и было задумано в задании.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №2:
'''


class GeometryCalculator:
    """
    Класс для подсчета площади геометрических фигур.
    Предоставляет функциональность для подсчета площади треугольника по разным формулам,
    площади прямоугольника, площади квадрата, площади ромба.
    Методы класса для подсчета площади реализованы с использованием статических методов.
    Класс также считает количество подсчетов площади и возвращает это значение с помощью статического метода.
    """

    _calculation_count = 0

    @staticmethod
    def triangle_area(base, height):
        """
        Статический метод для подсчета площади треугольника.

        :param base: Длина основания треугольника.
        :param height: Высота треугольника.
        :return: Площадь треугольника.
        """
        if base <= 0 or height <= 0:
            raise ValueError("Base and height must be positive numbers.")
        return 0.5 * base * height

    @staticmethod
    def rectangle_area(length, width):
        """
        Статический метод для подсчета площади прямоугольника.

        :param length: Длина прямоугольника.
        :param width: Ширина прямоугольника.
        :return: Площадь прямоугольника.
        """
        if length <= 0 or width <= 0:
            raise ValueError("Length and width must be positive numbers.")
        return length * width

    @staticmethod
    def square_area(side):
        """
        Статический метод для подсчета площади квадрата.

        :param side: Длина стороны квадрата.
        :return: Площадь квадрата.
        """
        if side <= 0:
            raise ValueError("Side length must be a positive number.")
        return side ** 2

    @staticmethod
    def rhombus_area(diagonal1, diagonal2):
        """
        Статический метод для подсчета площади ромба.

        :param diagonal1: Длина первой диагонали ромба.
        :param diagonal2: Длина второй диагонали ромба.
        :return: Площадь ромба.
        """
        if diagonal1 <= 0 or diagonal2 <= 0:
            raise ValueError("Diagonals must be positive numbers.")
        return 0.5 * diagonal1 * diagonal2

    @staticmethod
    def count_calculations():
        """
        Статический метод для возврата количества подсчетов площади.

        :return: Количество подсчетов площади.
        """
        return GeometryCalculator._calculation_count

    def __init__(self):
        """
        Конструктор класса, увеличивает количество подсчетов при создании объекта.
        """
        GeometryCalculator._calculation_count += 1


# Функция для ввода числа с проверкой на корректность
def input_float(prompt):
    while True:
        try:
            # Заменяем запятую на точку для корректного преобразования в float
            input_str = input(prompt).replace(',', '.')
            return float(input_str)
        except ValueError:
            print("Пожалуйста, введите корректное число.")


# Пример использования с вводом данных с клавиатуры
gc = GeometryCalculator()

# Ввод данных для расчета площади треугольника
base_input = input_float("Введите длину основания треугольника: ")
height_input = input_float("Введите высоту треугольника: ")
area_triangle = GeometryCalculator.triangle_area(base_input, height_input)
print(f"Площадь треугольника: {area_triangle}")

# Ввод данных для расчета площади прямоугольника
length_input = input_float("Введите длину прямоугольника: ")
width_input = input_float("Введите ширину прямоугольника: ")
area_rectangle = GeometryCalculator.rectangle_area(length_input, width_input)
print(f"Площадь прямоугольника: {area_rectangle}")

# Ввод данных для расчета площади квадрата
side_input = input_float("Введите длину стороны квадрата: ")
area_square = GeometryCalculator.square_area(side_input)
print(f"Площадь квадрата: {area_square}")

# Ввод данных для расчета площади ромба
diagonal1_input = input_float("Введите длину первой диагонали ромба: ")
diagonal2_input = input_float("Введите длину второй диагонали ромба: ")
area_rhombus = GeometryCalculator.rhombus_area(diagonal1_input, diagonal2_input)
print(f"Площадь ромба: {area_rhombus}")

# Вывод общего количества подсчетов
print(f"Общее количество подсчетов: {GeometryCalculator.count_calculations()}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Описание класса GeometryCalculator
'''


class GeometryCalculator:
    """
    Класс для подсчета площади геометрических фигур.
    Предоставляет функциональность для подсчета площади треугольника по разным формулам,
    площади прямоугольника, площади квадрата, площади ромба.
    Методы класса для подсчета площади реализованы с использованием статических методов.
    Класс также считает количество подсчетов площади и возвращает это значение с помощью статического метода.
    """
    _calculation_count = 0

    @staticmethod
    def triangle_area(base, height):
        """
        Статический метод для подсчета площади треугольника.

        :param base: Длина основания треугольника.
        :param height: Высота треугольника.
        :return: Площадь треугольника.
        """
        if base <= 0 or height <= 0:
            raise ValueError("Base and height must be positive numbers.")
        return 0.5 * base * height

    @staticmethod
    def rectangle_area(length, width):
        """
        Статический метод для подсчета площади прямоугольника.

        :param length: Длина прямоугольника.
        :param width: Ширина прямоугольника.
        :return: Площадь прямоугольника.
        """
        if length <= 0 or width <= 0:
            raise ValueError("Length and width must be positive numbers.")
        return length * width

    @staticmethod
    def square_area(side):
        """
        Статический метод для подсчета площади квадрата.

        :param side: Длина стороны квадрата.
        :return: Площадь квадрата.
        """
        if side <= 0:
            raise ValueError("Side length must be a positive number.")
        return side ** 2

    @staticmethod
    def rhombus_area(diagonal1, diagonal2):
        """
        Статический метод для подсчета площади ромба.

        :param diagonal1: Длина первой диагонали ромба.
        :param diagonal2: Длина второй диагонали ромба.
        :return: Площадь ромба.
        """
        if diagonal1 <= 0 or diagonal2 <= 0:
            raise ValueError("Diagonals must be positive numbers.")
        return 0.5 * diagonal1 * diagonal2

    @staticmethod
    def count_calculations():
        """
        Статический метод для возврата количества подсчетов площади.

        :return: Количество подсчетов площади.
        """
        return GeometryCalculator._calculation_count

    def __init__(self):
        """
        Конструктор класса, увеличивает количество подсчетов при создании объекта.
        """
        GeometryCalculator._calculation_count += 1


'''
Этот шаг описывает класс GeometryCalculator, который предоставляет функциональность для подсчета площади различных
геометрических фигур. Методы для подсчета площади (треугольника, прямоугольника, квадрата, ромба) реализованы как
статические методы. Также в классе есть статический метод для подсчета количества подсчетов площади и конструктор,
увеличивающий это количество при создании объекта.
'''
'''
Шаг №2: Пример использования класса
'''


# Функция для ввода числа с проверкой на корректность
def input_float(prompt):
    while True:
        try:
            # Заменяем запятую на точку для корректного преобразования в float
            input_str = input(prompt).replace(',', '.')
            return float(input_str)
        except ValueError:
            print("Пожалуйста, введите корректное число.")


# Пример использования с вводом данных с клавиатуры
gc = GeometryCalculator()

# Ввод данных для расчета площади треугольника
base_input = input_float("Введите длину основания треугольника: ")
height_input = input_float("Введите высоту треугольника: ")
area_triangle = GeometryCalculator.triangle_area(base_input, height_input)
print(f"Площадь треугольника: {area_triangle}")

# Ввод данных для расчета площади прямоугольника
length_input = input_float("Введите длину прямоугольника: ")
width_input = input_float("Введите ширину прямоугольника: ")
area_rectangle = GeometryCalculator.rectangle_area(length_input, width_input)
print(f"Площадь прямоугольника: {area_rectangle}")

# Ввод данных для расчета площади квадрата
side_input = input_float("Введите длину стороны квадрата: ")
area_square = GeometryCalculator.square_area(side_input)
print(f"Площадь квадрата: {area_square}")

# Ввод данных для расчета площади ромба
diagonal1_input = input_float("Введите длину первой диагонали ромба: ")
diagonal2_input = input_float("Введите длину второй диагонали ромба: ")
area_rhombus = GeometryCalculator.rhombus_area(diagonal1_input, diagonal2_input)
print(f"Площадь ромба: {area_rhombus}")

# Вывод общего количества подсчетов
print(f"Общее количество подсчетов: {GeometryCalculator.count_calculations()}")
'''
Этот шаг показывает пример использования класса с вводом данных с клавиатуры и расчетом площади для различных 
геометрических фигур. Создается объект gc класса GeometryCalculator, и затем пользователь вводит данные для расчета
площади треугольника, прямоугольника, квадрата и ромба. Результаты выводятся на экран, и также выводится общее
количество подсчетов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Задание №3:
'''
import math


class MathOperations:
    """
    Класс для выполнения математических операций над четырьмя числами.
    Предоставляет функциональность для подсчета максимума, минимума,
    среднеарифметического и факториала из четырех аргументов.
    """

    @staticmethod
    def max_of_four(a, b, c, d):
        """
        Статический метод для подсчета максимума из четырех аргументов.

        :param a: Первый аргумент.
        :param b: Второй аргумент.
        :param c: Третий аргумент.
        :param d: Четвертый аргумент.
        :return: Максимальное значение.
        """
        return max(a, b, c, d)

    @staticmethod
    def min_of_four(a, b, c, d):
        """
        Статический метод для подсчета минимума из четырех аргументов.

        :param a: Первый аргумент.
        :param b: Второй аргумент.
        :param c: Третий аргумент.
        :param d: Четвертый аргумент.
        :return: Минимальное значение.
        """
        return min(a, b, c, d)

    @staticmethod
    def average_of_four(a, b, c, d):
        """
        Статический метод для подсчета среднеарифметического из четырех аргументов.

        :param a: Первый аргумент.
        :param b: Второй аргумент.
        :param c: Третий аргумент.
        :param d: Четвертый аргумент.
        :return: Среднеарифметическое значение.
        """
        return (a + b + c + d) / 4

    @staticmethod
    def factorial(n):
        """
        Статический метод для подсчета факториала числа.

        :param n: Целое число.
        :return: Факториал числа n.
        """
        if not isinstance(n, int) or n < 0:
            raise ValueError("Факториал определен только для неотрицательных целых чисел.")
        return math.factorial(n)


# Функция для ввода числа с проверкой на корректность
def input_integer(prompt):
    while True:
        try:
            return int(input(prompt))
        except ValueError:
            print("Пожалуйста, введите целое число.")


# Пример использования с вводом данных с клавиатуры
try:
    a = input_integer("Введите первое число: ")
    b = input_integer("Введите второе число: ")
    c = input_integer("Введите третье число: ")
    d = input_integer("Введите четвертое число: ")

    result_max = MathOperations.max_of_four(a, b, c, d)
    result_min = MathOperations.min_of_four(a, b, c, d)
    result_avg = MathOperations.average_of_four(a, b, c, d)
    result_fact = MathOperations.factorial(a)

    print(f"Максимальное из четырех чисел: {result_max}")
    print(f"Минимальное из четырех чисел: {result_min}")
    print(f"Среднее арифметическое из четырех чисел: {result_avg}")
    print(f"Факториал первого числа: {result_fact}")

except ValueError as e:
    print(f"Ошибка: {e}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Описание класса MathOperations
'''
import math


class MathOperations:
    """
    Класс для выполнения математических операций над четырьмя числами.
    Предоставляет функциональность для подсчета максимума, минимума,
    среднеарифметического и факториала из четырех аргументов.
    """


'''
Описание:

Импортируется модуль math, который предоставляет математические функции, такие как факториал.
Создается класс MathOperations для выполнения математических операций.
'''
'''
Шаг 2: Статические методы класса
'''


@staticmethod
def max_of_four(a, b, c, d):
    """
    Статический метод для подсчета максимума из четырех аргументов.
    :param a: Первый аргумент.
    :param b: Второй аргумент.
    :param c: Третий аргумент.
    :param d: Четвертый аргумент.
    :return: Максимальное значение.
    """
    return max(a, b, c, d)


@staticmethod
def min_of_four(a, b, c, d):
    """
    Статический метод для подсчета минимума из четырех аргументов.
    :param a: Первый аргумент.
    :param b: Второй аргумент.
    :param c: Третий аргумент.
    :param d: Четвертый аргумент.
    :return: Минимальное значение.
    """
    return min(a, b, c, d)


@staticmethod
def average_of_four(a, b, c, d):
    """
    Статический метод для подсчета среднеарифметического из четырех аргументов.
    :param a: Первый аргумент.
    :param b: Второй аргумент.
    :param c: Третий аргумент.
    :param d: Четвертый аргумент.
    :return: Среднеарифметическое значение.
    """
    return (a + b + c + d) / 4


@staticmethod
def factorial(n):
    """
    Статический метод для подсчета факториала числа.
    :param n: Целое число.
    :return: Факториал числа n.
    """
    if not isinstance(n, int) or n < 0:
        raise ValueError("Факториал определен только для неотрицательных целых чисел.")
    return math.factorial(n)


'''
Описание:

max_of_four: Возвращает максимальное значение из четырех аргументов, используя встроенную функцию max.
min_of_four: Возвращает минимальное значение из четырех аргументов, используя встроенную функцию min.
average_of_four: Возвращает среднеарифметическое значение из четырех аргументов.
factorial: Возвращает факториал числа n, используя функцию math.factorial.
'''
'''
Шаг 3: Функция ввода целого числа с проверкой
'''


# Функция для ввода числа с проверкой на корректность
def input_integer(prompt):
    while True:
        try:
            return int(input(prompt))
        except ValueError:
            print("Пожалуйста, введите целое число.")


'''
Описание:

input_integer: Бесконечный цикл для ввода целого числа с обработкой исключений. 
Если ввод не является целым числом, выводится сообщение об ошибке.
'''
'''
Шаг 4: Пример использования с вводом данных
'''
# Пример использования с вводом данных с клавиатуры
try:
    a = input_integer("Введите первое число: ")
    b = input_integer("Введите второе число: ")
    c = input_integer("Введите третье число: ")
    d = input_integer("Введите четвертое число: ")

    result_max = MathOperations.max_of_four(a, b, c, d)
    result_min = MathOperations.min_of_four(a, b, c, d)
    result_avg = MathOperations.average_of_four(a, b, c, d)
    result_fact = MathOperations.factorial(a)

    print(f"Максимальное из четырех чисел: {result_max}")
    print(f"Минимальное из четырех чисел: {result_min}")
    print(f"Среднее арифметическое из четырех чисел: {result_avg}")
    print(f"Факториал первого числа: {result_fact}")

except ValueError as e:
    print(f"Ошибка: {e}")
'''
Описание:

Используется функция input_integer для ввода четырех целых чисел с клавиатуры.
Затем вызываются статические методы класса MathOperations для подсчета максимума, минимума, 
среднеарифметического и факториала.

Результаты выводятся на экран.
Если происходит ошибка (например, ввод не является целым числом), выводится сообщение об ошибке.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import roman

class Roman:
    word_to_number = {
        'one': 1, 'first': 1, 'two': 2, 'second': 2, 'three': 3, 'third': 3,
        # Добавьте остальные числа по мере необходимости
    }

    roman_numerals = {
        'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
        'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
        'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
    }

    def __init__(self, value):
        self.value = self.parse_input(value)

    @staticmethod
    def parse_input(input_str):
        if isinstance(input_str, int):
            return input_str
        try:
            return roman.fromRoman(input_str)
        except roman.InvalidRomanNumeralError:
            try:
                return int(input_str)
            except ValueError:
                raise ValueError(f"Invalid input: {input_str}")

    @staticmethod
    def int_to_roman(num):
        if not 0 < num < 10000000000:
            raise ValueError("Number out of range (1 to 9999999999)")

        result = ''
        for numeral, integer in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
            while num >= integer:
                result += numeral
                num -= integer

        return result

    def __add__(self, other):
        if isinstance(other, Roman):
            result = self.value + other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for +")

    def __sub__(self, other):
        if isinstance(other, Roman):
            result = self.value - other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for -")

    def __mul__(self, other):
        if isinstance(other, Roman):
            result = self.value * other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for *")

    def __truediv__(self, other):
        if isinstance(other, Roman):
            result = self.value / other.value
            return Roman(int(result))
        else:
            raise TypeError("Unsupported operand type for /")

    def __str__(self):
        return str(self.value)


class RomanNumberConverter:
    roman_numerals = {
        1: 'I', 4: 'IV', 5: 'V', 9: 'IX', 10: 'X',
        40: 'XL', 50: 'L', 90: 'XC', 100: 'C',
        400: 'CD', 500: 'D', 900: 'CM', 1000: 'M'
    }

    word_to_number = {
        'one': 1, 'first': 1, 'two': 2, 'second': 2, 'three': 3, 'third': 3,
        # Добавьте остальные числа по мере необходимости
    }

    @staticmethod
    def parse_input(input_str):
        input_str_lower = input_str.lower()
        if input_str_lower in RomanNumberConverter.word_to_number:
            return RomanNumberConverter.word_to_number[input_str_lower]
        else:
            try:
                return int(input_str)
            except ValueError:
                raise ValueError(f"Invalid input: {input_str}")

    @staticmethod
    def int_to_roman(num):
        if not 0 < num < 10000000000:
            raise ValueError("Number out of range (1 to 9999999999)")

        result = ''
        for value, numeral in sorted(RomanNumberConverter.roman_numerals.items(), key=lambda x: x[0], reverse=True):
            while num >= value:
                result += numeral
                num -= value

        return result

# Пример использования
try:
    a = Roman(input("Введите первое число: "))
    b = Roman(input("Введите второе число: "))

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a} + {b} = {result_add}")
    print(f"{a} - {b} = {result_sub}")
    print(f"{a} * {b} = {result_mul}")
    print(f"{a} / {b} = {result_div}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
близко
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import roman

class Roman:
    word_to_number = {
        'one': 1, 'first': 1, 'один': 1,
        'two': 2, 'second': 2, 'два': 2,
        'three': 3, 'third': 3, 'три': 3,
        'four': 4, 'четыре': 4,
        'five': 5, 'пять': 5,
        'six': 6, 'шесть': 6,
        'seven': 7, 'семь': 7,
        'eight': 8, 'восемь': 8,
        'nine': 9, 'девять': 9,
        'ten': 10, 'десять': 10,
        # Добавьте остальные числа по мере необходимости
    }

    roman_numerals = {
        'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
        'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
        'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
    }

    def __init__(self, value):
        self.value = self.parse_input(value)

    @staticmethod
    def parse_input(input_str):
        if isinstance(input_str, int):
            return input_str
        input_str_lower = input_str.lower()
        if input_str_lower in Roman.word_to_number:
            return Roman.word_to_number[input_str_lower]
        else:
            try:
                return roman.fromRoman(input_str)
            except roman.InvalidRomanNumeralError:
                try:
                    return int(input_str)
                except ValueError:
                    raise ValueError(f"Invalid input: {input_str}")

    @staticmethod
    def int_to_roman(num):
        if not 0 < num < 10000000000:
            raise ValueError("Number out of range (1 to 9999999999)")

        result = ''
        for numeral, integer in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
            while num >= integer:
                result += numeral
                num -= integer

        return result

    def __add__(self, other):
        if isinstance(other, Roman):
            result = self.value + other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for +")

    def __sub__(self, other):
        if isinstance(other, Roman):
            result = self.value - other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for -")

    def __mul__(self, other):
        if isinstance(other, Roman):
            result = self.value * other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for *")

    def __truediv__(self, other):
        if isinstance(other, Roman):
            result = self.value / other.value
            return Roman(int(result))
        else:
            raise TypeError("Unsupported operand type for /")

    def __str__(self):
        if self.value in Roman.roman_numerals.values():
            # Если значение уже в римском формате, просто вернем его
            return roman.toRoman(self.value)
        elif self.value in Roman.word_to_number.values():
            # Если значение в числительной форме, преобразуем в римское число
            return Roman.int_to_roman(self.value)
        else:
            return str(self.value)

# Пример использования
try:
    a = Roman(input("Введите первое число: "))
    b = Roman(input("Введите второе число: "))

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a} + {b} = {result_add}")
    print(f"{a} - {b} = {result_sub}")
    print(f"{a} * {b} = {result_mul}")
    print(f"{a} / {b} = {result_div}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
Почти рядом
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
import roman

class Roman:
    word_to_number = {
        'one': 1, 'first': 1, 'один': 1,
        'two': 2, 'second': 2, 'два': 2,
        'three': 3, 'third': 3, 'три': 3,
        'four': 4, 'четыре': 4,
        'five': 5, 'пять': 5,
        'six': 6, 'шесть': 6,
        'seven': 7, 'семь': 7,
        'eight': 8, 'восемь': 8,
        'nine': 9, 'девять': 9,
        'ten': 10, 'десять': 10,
        'eleven': 11, 'одиннадцать': 11,
        'twelve': 12, 'двенадцать': 12,
        'thirteen': 13, 'тринадцать': 13,
        'fourteen': 14, 'четырнадцать': 14,
        'fifteen': 15, 'пятнадцать': 15,
        'sixteen': 16, 'шестнадцать': 16,
        'seventeen': 17, 'семнадцать': 17,
        'eighteen': 18, 'восемнадцать': 18,
        'nineteen': 19, 'девятнадцать': 19,
        'twenty': 20, 'двадцать': 20,
        'thirty': 30, 'тридцать': 30,
        'forty': 40, 'сорок': 40,
        'fifty': 50, 'пятьдесят': 50,
        'sixty': 60, 'шестьдесят': 60,
        'seventy': 70, 'семьдесят': 70,
        'eighty': 80, 'восемьдесят': 80,
        'ninety': 90, 'девяносто': 90,
        'hundred': 100, 'сто': 100,
    }

    roman_numerals = {
        'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
        'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
        'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
    }

    def __init__(self, value):
        self.value = self.parse_input(value)

    @staticmethod
    def parse_input(input_str):
        if isinstance(input_str, int):
            return input_str
        input_str_lower = input_str.lower()
        if input_str_lower in Roman.word_to_number:
            return Roman.word_to_number[input_str_lower]
        else:
            try:
                return roman.fromRoman(input_str)
            except roman.InvalidRomanNumeralError:
                try:
                    return int(input_str)
                except ValueError:
                    raise ValueError(f"Invalid input: {input_str}")

    @staticmethod
    def int_to_roman(num):
        if not 0 < num < 10000000000:
            raise ValueError("Number out of range (1 to 9999999999)")

        result = ''
        for numeral, integer in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
            while num >= integer:
                result += numeral
                num -= integer

        return result

    def to_roman_string(self):
        result = ''
        remaining = self.value

        for numeral, integer in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
            while remaining >= integer:
                result += numeral
                remaining -= integer

        return result

    def to_word_number(self):
        for word, number in Roman.word_to_number.items():
            if number == self.value:
                return word.capitalize()
        return str(self.value)

    def __add__(self, other):
        if isinstance(other, Roman):
            result = self.value + other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for +")

    def __sub__(self, other):
        if isinstance(other, Roman):
            result = self.value - other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for -")

    def __mul__(self, other):
        if isinstance(other, Roman):
            result = self.value * other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for *")

    def __truediv__(self, other):
        if isinstance(other, Roman):
            result = self.value / other.value
            return Roman(int(result))
        else:
            raise TypeError("Unsupported operand type for /")

    def __str__(self):
        if self.value in Roman.roman_numerals.values():
            # Если значение уже в римском формате, просто вернем его
            return roman.toRoman(self.value)
        elif self.value in Roman.word_to_number.values():
            # Если значение в числительной форме, преобразуем в римское число
            return Roman.int_to_roman(self.value)
        else:
            return str(self.value)

# Пример использования
try:
    a = Roman(input("Введите первое число: "))
    b = Roman(input("Введите второе число: "))

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a.to_roman_string()} + {b.to_roman_string()} = {result_add.to_roman_string()}")
    print(f"{a.to_roman_string()} - {b.to_roman_string()} = {result_sub.to_roman_string()}")
    print(f"{a.to_roman_string()} * {b.to_roman_string()} = {result_mul.to_roman_string()}")
    print(f"{a.to_roman_string()} / {b.to_roman_string()} = {result_div.to_roman_string()}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
from datetime import date

class Person:
    def __init__(self, a_name, a_age):
        self.name = a_name
        self.age = a_age

    ADULT_AGE = 18

    @staticmethod
    def isAdult(age):
        return age > 18

    @classmethod
    def getAdultAge(cls):
        return cls.ADULT_AGE

# altynbek = Person('Altynbek', 24)
# almas = Person('Almas', 16)
# aliya = Person('Aliya', 22)

print(adil.name,)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
КЛАССНАЯ РАБОТА!!!
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
class Цифра:
    def __init__(self, a_value):
        self.value = a_value

    def на_экран(self):
        print(self.value)

    def __str__(self):
        return str(Цифра.toЦифра(self.value))

    def плюс(self, вторая_цифра):
        if type(вторая_цифра) is Цифра:
            return Цифра(self.value + вторая_цифра.value)
        if type(вторая_цифра) is int:
            return Цифра(self.value + вторая_цифра)

    def __add__(self, вторая_цифра):
        if type(вторая_цифра) is Цифра:
            return Цифра(self.value + вторая_цифра.value)
        return Цифра(self.value + вторая_цифра)

    def __sub__(self, вторая_цифра):
        return Цифра(self.value - вторая_цифра.value)

    def __eq__(self, other):
        if type(other) is Цифра:
            return self.value == other.value
        if type(other) is int:
            return self.value == other

    @classmethod
    def toInt(cls, цифра):
        return int(цифра.value)

    @staticmethod
    def toЦифра(value):
        int_to_cifra = ['ноль', 'один', 'два', 'три', 'четыре', 'пять', 'шесть']
        return int_to_cifra[value]

    # @classmethod
    # def fromInt(cls, value):
    #     return cls(value)


# class Person: pass
# class Person(Person): pass

class int(int):
    def __init__(self, a_value):
        if type(a_value) is int:
            self.value = a_value.value
        self.value = a_value

    def __add__(self, other):
        print('operand')
        # if type(other) is Цифра:
        # return int(self + other.value)
        # else:
        #     super().__add__(self, other)


один = int(1)
два = int(2)
три = int(3)
шесть = int(6)

# один.на_экран()
# print(один.плюс(два))
# print(один)
# print(Цифра.toЦифра(2))
# print(три + два + один == шесть)
# print(три - два + один == 6)
# # print(1 + 2 + 3)
print(три + 1)
print(1 + три)
print(1 + 2)


# print(1 + 2)

# x.x = 10
# y.x = 3

# print(один + три)

class Number:
    def __init__(self, a_value):
        self.x = a_value

    def __add__(self, y):
        return Number(self.x + y.x)


x = Number(10)
y = Number(3)

print(y + x)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
Дата выполнения ДОМАШНЕЙ РАБОТЫ: 29-30 - ДЕКАБРЯ 2023
''''  ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓   ↓
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
'''
'''
Урок от 29.12.2023
Домашняя работа №21: ООП.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполните следующее задания:
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
Создайте класс Roman (РимскоеЧисло), представляющий римское число и поддерживающий операции +, -, *, /.

При реализации класса:

операции +, -, *, / реализуйте как специальные методы
методы преобразования как статические методы.
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Далее решение этого задания ↑
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №1.
'''
class Roman:
    roman_numerals = {
        'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
        'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
        'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
    }

    def __init__(self, value):
        self.roman_str = value.upper()

    @staticmethod
    def parse_input(input_str):
        try:
            return Roman.roman_numerals[input_str]
        except KeyError:
            raise ValueError(f"Invalid Roman numeral: {input_str}")

    @staticmethod
    def int_to_roman(num):
        if not 0 < num < 4000:
            raise ValueError("Number out of range (1 to 3999)")

        result = ''
        for roman, value in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
            while num >= value:
                result += roman
                num -= value
        return result

    def to_int(self):
        result = 0
        i = 0
        while i < len(self.roman_str):
            if i + 1 < len(self.roman_str) and self.roman_str[i:i + 2] in Roman.roman_numerals:
                result += Roman.roman_numerals[self.roman_str[i:i + 2]]
                i += 2
            else:
                result += Roman.roman_numerals[self.roman_str[i]]
                i += 1
        return result

    def __add__(self, other):
        if isinstance(other, Roman):
            result = self.to_int() + other.to_int()
            return Roman(self.int_to_roman(result))
        else:
            raise TypeError("Unsupported operand type for +")

    def __sub__(self, other):
        if isinstance(other, Roman):
            result = self.to_int() - other.to_int()
            return Roman(self.int_to_roman(result))
        else:
            raise TypeError("Unsupported operand type for -")

    def __mul__(self, other):
        if isinstance(other, Roman):
            result = self.to_int() * other.to_int()
            return Roman(self.int_to_roman(result))
        else:
            raise TypeError("Unsupported operand type for *")

    def __truediv__(self, other):
        if isinstance(other, Roman):
            result = self.to_int() / other.to_int()
            return Roman(self.int_to_roman(int(result)))
        else:
            raise TypeError("Unsupported operand type for /")

    def __str__(self):
        return self.roman_str

# Пример использования
try:
    a = Roman("XII")
    b = Roman("V")

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a} + {b} = {result_add}")
    print(f"{a} - {b} = {result_sub}")
    print(f"{a} * {b} = {result_mul}")
    print(f"{a} / {b} = {result_div}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1. Инициализация класса Roman
'''
class Roman:
    roman_numerals = {
        'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
        'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
        'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
    }

    def __init__(self, value):
        self.roman_str = value.upper()
'''
Описание: 
Класс Roman представляет римское число. 
Внутри класса определен словарь roman_numerals, сопоставляющий римские цифры и их арабские эквиваленты. 
В конструкторе (__init__) инициализируется объект римского числа, приводя входное значение к верхнему регистру.
'''
'''
Шаг №2. Статический метод parse_input
'''
@staticmethod
def parse_input(input_str):
    try:
        return Roman.roman_numerals[input_str]
    except KeyError:
        raise ValueError(f"Invalid Roman numeral: {input_str}")
'''
Описание: 
Статический метод parse_input преобразует римское число в арабское число. 
Если входное значение не является допустимым римским числом, генерируется исключение ValueError.
'''
'''
Шаг №3. Статический метод int_to_roman
'''
@staticmethod
def int_to_roman(num):
    if not 0 < num < 4000:
        raise ValueError("Number out of range (1 to 3999)")

    result = ''
    for roman, value in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
        while num >= value:
            result += roman
            num -= value
    return result
'''
Описание: 
Статический метод int_to_roman преобразует арабское число в римское число. 
В случае, если число находится вне диапазона (1 до 3999), генерируется исключение ValueError.
'''
'''
Шаг №4. Метод to_int
'''
def to_int(self):
    result = 0
    i = 0
    while i < len(self.roman_str):
        if i + 1 < len(self.roman_str) and self.roman_str[i:i + 2] in Roman.roman_numerals:
            result += Roman.roman_numerals[self.roman_str[i:i + 2]]
            i += 2
        else:
            result += Roman.roman_numerals[self.roman_str[i]]
            i += 1
    return result
'''
Описание: 
Метод to_int преобразует римское число объекта в арабское число, используя словарь roman_numerals. 
Обрабатываются случаи, когда в римском числе есть комбинации (например, 'IV' для 4).
'''
'''
Шаг №5. Перегрузка операторов (__add__, __sub__, __mul__, __truediv__)
'''
def __add__(self, other):
    if isinstance(other, Roman):
        result = self.to_int() + other.to_int()
        return Roman(self.int_to_roman(result))
    else:
        raise TypeError("Unsupported operand type for +")


def __sub__(self, other):
    if isinstance(other, Roman):
        result = self.to_int() - other.to_int()
        return Roman(self.int_to_roman(result))
    else:
        raise TypeError("Unsupported operand type for -")


def __mul__(self, other):
    if isinstance(other, Roman):
        result = self.to_int() * other.to_int()
        return Roman(self.int_to_roman(result))
    else:
        raise TypeError("Unsupported operand type for *")


def __truediv__(self, other):
    if isinstance(other, Roman):
        result = self.to_int() / other.to_int()
        return Roman(self.int_to_roman(int(result)))
    else:
        raise TypeError("Unsupported operand type for /")
'''
Описание: 
Операторы сложения, вычитания, умножения и деления перегружены для класса Roman. 
Каждый из них проверяет тип второго операнда (other). Если он также является объектом класса Roman, 
выполняется соответствующая операция, а результат преобразуется обратно в римское число.
'''
'''
Шаг №6. Метод __str__
'''
def __str__(self):
    return self.roman_str
'''
Описание: Метод __str__ возвращает строковое представление объекта, используя римское число.
'''
'''
Пример использования
'''
# Пример использования
try:
    a = Roman("XII")
    b = Roman("V")

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a} + {b} = {result_add}")
    print(f"{a} - {b} = {result_sub}")
    print(f"{a} * {b} = {result_mul}")
    print(f"{a} / {b} = {result_div}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
'''
Описание: 
Создаются два объекта римских чисел a и b. 
Затем выполняются операции сложения, вычитания, умножения и деления, и результаты выводятся на экран. 
Если возникают ошибки (например, некорректное римское число или несовместимый операнд), выводится соответствующее 
сообщение об ошибке.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №2.
'''
import roman

class Roman:
    word_to_number = {
        'one': 1, 'first': 1, 'two': 2, 'second': 2, 'three': 3, 'third': 3,
        # Добавьте остальные числа по мере необходимости
    }

    roman_numerals = {
        'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
        'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
        'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
    }

    def __init__(self, value):
        self.value = self.parse_input(value)

    @staticmethod
    def parse_input(input_str):
        if isinstance(input_str, int):
            return input_str
        input_str_lower = input_str.lower()
        if input_str_lower in Roman.word_to_number:
            return Roman.word_to_number[input_str_lower]
        else:
            try:
                # Попробуем преобразовать введенное значение как римскую цифру
                if input_str_upper := input_str.upper():
                    return roman.fromRoman(input_str_upper)
                else:
                    raise ValueError("Invalid input: empty string")
            except roman.InvalidRomanNumeralError:
                raise ValueError(f"Invalid input: {input_str}")

    def __add__(self, other):
        if isinstance(other, Roman):
            result = self.value + other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for +")

    def __sub__(self, other):
        if isinstance(other, Roman):
            result = self.value - other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for -")

    def __mul__(self, other):
        if isinstance(other, Roman):
            result = self.value * other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for *")

    def __truediv__(self, other):
        if isinstance(other, Roman):
            result = self.value / other.value
            return Roman(int(result))
        else:
            raise TypeError("Unsupported operand type for /")

    def __str__(self):
        return str(self.value)

# Пример использования
try:
    a = Roman(input("Введите первое число: "))
    b = Roman(input("Введите второе число: "))

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a} + {b} = {result_add}")
    print(f"{a} - {b} = {result_sub}")
    print(f"{a} * {b} = {result_mul}")
    print(f"{a} / {b} = {result_div}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
X и I - I и X
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотеки roman
'''
import roman
'''
Этот шаг просто импортирует библиотеку roman, 
которая предоставляет функциональность для преобразования римских чисел в арабские и обратно.
'''
'''
Шаг 2: Определение класса Roman
'''
class Roman:
''''
Здесь создается класс Roman.
''''
'''
Шаг 3: Статические переменные класса
'''
word_to_number = {
    'one': 1, 'first': 1, 'two': 2, 'second': 2, 'three': 3, 'third': 3,
    # Добавьте остальные числа по мере необходимости
}

roman_numerals = {
    'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
    'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
    'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
}
'''
Здесь определены статические переменные класса, 
которые используются для преобразования словесных чисел в арабские (word_to_number) и для представления римских 
цифр в виде словаря (roman_numerals).
'''
'''
Шаг 4: Конструктор класса
'''
def __init__(self, value):
    self.value = self.parse_input(value)
'''
Конструктор класса принимает значение и вызывает метод parse_input, чтобы преобразовать его в числовое значение.
'''
'''
Шаг 5: Метод parse_input
'''


@staticmethod
def parse_input(input_str):
    if isinstance(input_str, int):
        return input_str
    input_str_lower = input_str.lower()
    if input_str_lower in Roman.word_to_number:
        return Roman.word_to_number[input_str_lower]
    else:
        try:
            # Попробуем преобразовать введенное значение как римскую цифру
            if input_str_upper := input_str.upper():
                return roman.fromRoman(input_str_upper)
            else:
                raise ValueError("Invalid input: empty string")
        except roman.InvalidRomanNumeralError:
            raise ValueError(f"Invalid input: {input_str}")
'''
Этот метод преобразует входные данные в числовое значение. Если входное значение является целым числом, оно 
возвращается как есть. 
В противном случае проверяется, является ли входное слово числом на языке (например, 'one') или римским числом.
'''
'''
Шаг 6-9: Перегрузка операторов
'''


def __add__(self, other):
    if isinstance(other, Roman):
        result = self.value + other.value
        return Roman(result)
    else:
        raise TypeError("Unsupported operand type for +")


def __sub__(self, other):
    if isinstance(other, Roman):
        result = self.value - other.value
        return Roman(result)
    else:
        raise TypeError("Unsupported operand type for -")


def __mul__(self, other):
    if isinstance(other, Roman):
        result = self.value * other.value
        return Roman(result)
    else:
        raise TypeError("Unsupported operand type for *")


def __truediv__(self, other):
    if isinstance(other, Roman):
        result = self.value / other.value
        return Roman(int(result))
    else:
        raise TypeError("Unsupported operand type for /")
'''
Эти методы перегружают операторы сложения, вычитания, умножения и деления для объектов класса Roman.
'''
'''
Шаг 10: Перегрузка метода __str__
'''
def __str__(self):
    return str(self.value)
'''
Этот метод перегружает метод __str__ и возвращает строковое представление числового значения.
'''
'''
Шаг 11-14: Пример использования
'''
# Пример использования
try:
    a = Roman(input("Введите первое число: "))
    b = Roman(input("Введите второе число: "))

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a} + {b} = {result_add}")
    print(f"{a} - {b} = {result_sub}")
    print(f"{a} * {b} = {result_mul}")
    print(f"{a} / {b} = {result_div}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
'''
Здесь пользователю предлагается ввести два числа (либо римские, либо словесные),
а затем выполняются операции сложения, вычитания, умножения и деления с выводом результатов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №3.
'''
import roman

class Roman:
    word_to_number = {
        'one': 1, 'first': 1, 'один': 1,
        'two': 2, 'second': 2, 'два': 2,
        'three': 3, 'third': 3, 'три': 3,
        'four': 4, 'четыре': 4,
        'five': 5, 'пять': 5,
        'six': 6, 'шесть': 6,
        'seven': 7, 'семь': 7,
        'eight': 8, 'восемь': 8,
        'nine': 9, 'девять': 9,
        'ten': 10, 'десять': 10,
        'eleven': 11, 'одиннадцать': 11,
        'twelve': 12, 'двенадцать': 12,
        'thirteen': 13, 'тринадцать': 13,
        'fourteen': 14, 'четырнадцать': 14,
        'fifteen': 15, 'пятнадцать': 15,
        'sixteen': 16, 'шестнадцать': 16,
        'seventeen': 17, 'семнадцать': 17,
        'eighteen': 18, 'восемнадцать': 18,
        'nineteen': 19, 'девятнадцать': 19,
        'twenty': 20, 'двадцать': 20,
        'thirty': 30, 'тридцать': 30,
        'forty': 40, 'сорок': 40,
        'fifty': 50, 'пятьдесят': 50,
        'sixty': 60, 'шестьдесят': 60,
        'seventy': 70, 'семьдесят': 70,
        'eighty': 80, 'восемьдесят': 80,
        'ninety': 90, 'девяносто': 90,
        'hundred': 100, 'сто': 100,
    }

    roman_numerals = {
        'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
        'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
        'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
    }

    def __init__(self, value):
        self.value = self.parse_input(value)

    @staticmethod
    def parse_input(input_str):
        if isinstance(input_str, int):
            return input_str
        input_str_lower = input_str.lower()
        if input_str_lower in Roman.word_to_number:
            return Roman.word_to_number[input_str_lower]
        else:
            try:
                return roman.fromRoman(input_str)
            except roman.InvalidRomanNumeralError:
                try:
                    return int(input_str)
                except ValueError:
                    raise ValueError(f"Invalid input: {input_str}")

    @staticmethod
    def int_to_roman(num):
        if not 0 < num < 10000000000:
            raise ValueError("Number out of range (1 to 9999999999)")

        result = ''
        for numeral, integer in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
            while num >= integer:
                result += numeral
                num -= integer

        return result

    def to_roman_string(self):
        result = ''
        remaining = self.value

        for numeral, integer in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
            while remaining >= integer:
                result += numeral
                remaining -= integer

        return result

    def to_word_number(self):
        for word, number in Roman.word_to_number.items():
            if number == self.value:
                return word.capitalize()
        return str(self.value)

    def __add__(self, other):
        if isinstance(other, Roman):
            result = self.value + other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for +")

    def __sub__(self, other):
        if isinstance(other, Roman):
            result = self.value - other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for -")

    def __mul__(self, other):
        if isinstance(other, Roman):
            result = self.value * other.value
            return Roman(result)
        else:
            raise TypeError("Unsupported operand type for *")

    def __truediv__(self, other):
        if isinstance(other, Roman):
            result = self.value / other.value
            return Roman(int(result))
        else:
            raise TypeError("Unsupported operand type for /")

    def __str__(self):
        if self.value in Roman.roman_numerals.values():
            # Если значение уже в римском формате, просто вернем его
            return roman.toRoman(self.value)
        elif self.value in Roman.word_to_number.values():
            # Если значение в числительной форме, преобразуем в римское число
            return Roman.int_to_roman(self.value)
        else:
            return str(self.value)

# Пример использования
try:
    a = Roman(input("Введите первое число: "))
    b = Roman(input("Введите второе число: "))

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a.to_roman_string()} + {b.to_roman_string()} = {result_add.to_roman_string()}")
    print(f"{a.to_roman_string()} - {b.to_roman_string()} = {result_sub.to_roman_string()}")
    print(f"{a.to_roman_string()} * {b.to_roman_string()} = {result_mul.to_roman_string()}")
    print(f"{a.to_roman_string()} / {b.to_roman_string()} = {result_div.to_roman_string()}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Словари для словесных и римских чисел
'''
word_to_number = {
    # ... (числа от 'one' до 'hundred')
}

roman_numerals = {
    'I': 1, 'IV': 4, 'V': 5, 'IX': 9, 'X': 10,
    'XL': 40, 'L': 50, 'XC': 90, 'C': 100,
    'CD': 400, 'D': 500, 'CM': 900, 'M': 1000
}
'''
Добавлен словарь word_to_number для соответствия словесных чисел и их арабских значений.
'''
'''
Шаг 2: Метод parse_input для обработки словесных чисел
'''
@staticmethod
def parse_input(input_str):
    if isinstance(input_str, int):
        return input_str
    input_str_lower = input_str.lower()
    if input_str_lower in Roman.word_to_number:
        return Roman.word_to_number[input_str_lower]
    else:
        try:
            return roman.fromRoman(input_str)
        except roman.InvalidRomanNumeralError:
            try:
                return int(input_str)
            except ValueError:
                raise ValueError(f"Invalid input: {input_str}")
'''
Этот метод обновлен для обработки ввода словесных чисел, преобразуя их в арабские числа.
'''
'''
Шаг 3: Метод to_word_number для представления числа словами
'''
def to_word_number(self):
    for word, number in Roman.word_to_number.items():
        if number == self.value:
            return word.capitalize()
    return str(self.value)
'''
Этот метод возвращает представление числа словами.
'''
'''
Шаг 4: Метод to_roman_string
'''
def to_roman_string(self):
    result = ''
    remaining = self.value

    for numeral, integer in sorted(Roman.roman_numerals.items(), key=lambda x: x[1], reverse=True):
        while remaining >= integer:
            result += numeral
            remaining -= integer

    return result
'''
Метод для представления числа в римской форме как строку.
'''
'''
Шаг 5: Обновленные операции (сложение, вычитание, умножение, деление)
'''
def __add__(self, other):
    if isinstance(other, Roman):
        result = self.value + other.value
        return Roman(result)
    else:
        raise TypeError("Unsupported operand type for +")

def __sub__(self, other):
    if isinstance(other, Roman):
        result = self.value - other.value
        return Roman(result)
    else:
        raise TypeError("Unsupported operand type for -")


def __mul__(self, other):
    if isinstance(other, Roman):
        result = self.value * other.value
        return Roman(result)
    else:
        raise TypeError("Unsupported operand type for *")


def __truediv__(self, other):
    if isinstance(other, Roman):
        result = self.value / other.value
        return Roman(int(result))
    else:
        raise TypeError("Unsupported operand type for /")
'''
Обновленные операции теперь возвращают объект класса Roman для обеспечения сохранения типов.
'''
'''
Шаг 6: Пример использования
'''
# Пример использования
try:
    a = Roman(input("Введите первое число: "))
    b = Roman(input("Введите второе число: "))

    result_add = a + b
    result_sub = a - b
    result_mul = a * b
    result_div = a / b

    print(f"{a.to_roman_string()} + {b.to_roman_string()} = {result_add.to_roman_string()}")
    print(f"{a.to_roman_string()} - {b.to_roman_string()} = {result_sub.to_roman_string()}")
    print(f"{a.to_roman_string()} * {b.to_roman_string()} = {result_mul.to_roman_string()}")
    print(f"{a.to_roman_string()} / {b.to_roman_string()} = {result_div.to_roman_string()}")

except ValueError as e:
    print(f"Ошибка: {e}")
except TypeError as e:
    print(f"Ошибка: {e}")
'''
Пример использования теперь выводит результаты в римской форме и числительной форме.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #

2024
Продолжение будет (но чуть позже)

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #


# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней работы: 08-09 - ЯНВАРЯ 2024 года.
''''
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы HTML и CSS !!! НО НАПИСАЛ Я КОД В PYCHARM - на PYTHON
'''
'''
Урок от 08.01.2024
Домашнее задание № 1 : Введение в Веб разработку
'''
'''
Выполните следующие задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
1. Установить Visual Studio Code.
2. Создать html файл с помощью VSCode.
3. Написать в html файле свои имя и фамилию.
4. Открыть в браузере с инструментами разработчика.
5. Сделать скриншот и отправить его вместе с вашим html файлом.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Скриншоты можно делать с помощью встроенного приложения Ножницы или с помощью бесплатной программы LightShot.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант - более правильный наверное. (Написанный на Python)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import re
import sqlite3
import html
import webbrowser

# Создаем подключение к базе данных (или создаем базу данных, если ее нет)
conn = sqlite3.connect('names_database.db')
cursor = conn.cursor()

# Создаем таблицу, если ее нет
cursor.execute('''
    CREATE TABLE IF NOT EXISTS names (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT,
        last_name TEXT
    )
''')
conn.commit()

def validate_name(name):
    # Проверка, что имя состоит только из букв, как в русском, так и в английском алфавите
    return re.match("^[а-яА-Яa-zA-Z]+$", name) is not None

# Получаем имя от пользователя с проверкой ввода
name = input("Введите ваше имя: ")
while not validate_name(name):
    print("Ошибка: Имя должно состоять только из букв русского или английского алфавита.")
    name = input("Введите ваше имя: ")

# Получаем фамилию от пользователя с проверкой ввода
surname = input("Введите вашу фамилию: ")
while not validate_name(surname):
    print("Ошибка: Фамилия должна состоять только из букв русского или английского алфавита.")
    surname = input("Введите вашу фамилию: ")

# Записываем введенные данные в базу данных
cursor.execute('INSERT INTO names (first_name, last_name) VALUES (?, ?)', (name, surname))
conn.commit()

# Записываем введенные данные в HTML-файл
with open('index.html', 'w', encoding='utf-8') as file:
    # Записываем начальный HTML-код
    file.write('''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Имя и Фамилия</title>
</head>
<body>
''')

    # Записываем введенные данные в HTML-файл с разными стилями
    file.write(f'    <h3>{html.escape(name)}</h3>\n')  # Стандартное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h2><i>{html.escape(name)}</i></h2>\n')  # Курсивное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h1><b>{html.escape(name)}</b></h1>\n')  # Жирное отображение

    file.write(f'    <h3>{html.escape(surname)}</h3>\n')  # Стандартное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h2><i>{html.escape(surname)}</i></h2>\n')  # Курсивное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h1><b>{html.escape(surname)}</b></h1>\n')  # Жирное отображение

    # Добавляем рамку для JPG-изображения
    file.write('    <div style="border: 1px solid #ccc; padding: 10px; margin-top: 20px; width: 500px; height: 500px;">\n')
    file.write('        <img src="https://sun1-13.userapi.com/s/v1/if1/ygP2SHKIEpSAcXSnjPOnyo_29b2oYdOwC-tYTq_g_pxbMtjq6LFPpYHALe4I2vpHCNC4RmWX.jpg?size=200x200&quality=96&crop=290,94,382,382&ava=1" style="width: 100%; height: 100%;">\n')
    file.write('    </div>\n')

    # Добавляем рамку для GIF-изображения
    file.write('    <div style="border: 1px solid #ccc; padding: 10px; margin-top: 20px; width: 500px; height: 500px;">\n')
    file.write('        <img src="https://cdn.humoraf.ru/wp-content/uploads/2019/04/ponedelnik-gifki-humoraf-53.gif" alt="GIF Картинка" style="width: 100%; height: 100%;">\n')
    file.write('    </div>\n')

    # Записываем закрывающий HTML-код
    file.write('''
</body>
</html>
''')

print("HTML-файл создан успешно.")

# Открываем HTML-файл в браузере
webbrowser.open('index.html', new=2)

# Закрываем соединение с базой данных
conn.close()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Подключение к базе данных

Название функции/команды: sqlite3.connect('names_database.db')

Пример:
'''
conn = sqlite3.connect('names_database.db')
'''
Описание:
Эта строка кода создает подключение к базе данных SQLite с именем "names_database.db" или открывает уже существующую 
базу данных с таким именем. Если базы данных нет, она будет создана. Объект conn представляет собой соединение
с базой данных.
'''
'''
Шаг №2: Создание таблицы в базе данных

Название функции/команды: cursor.execute('''CREATE TABLE IF NOT EXISTS names (id INTEGER PRIMARY KEY AUTOINCREMENT,
first_name TEXT, last_name TEXT)''')

Пример:
'''
cursor.execute('''
    CREATE TABLE IF NOT EXISTS names (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT,
        last_name TEXT
    )
''')
'''
Описание:
Эта строка кода создает таблицу "names" в базе данных с тремя полями: "id" (целочисленный идентификатор,
автоматически увеличиваемый), "first_name" и "last_name" (текстовые поля для хранения имени и фамилии соответственно).
'''
'''
Шаг №3: Валидация имени

Название функции/команды: re.match("^[а-яА-Яa-zA-Z]+$", name)

Пример:
'''
return re.match("^[а-яА-Яa-zA-Z]+$", name) is not None
'''
Описание:
Эта функция validate_name использует регулярное выражение для проверки того, что переданное имя состоит только из
букв русского или английского алфавита. Возвращает True, если условие выполняется, и False в противном случае.
'''
'''
Шаг №4: Запись данных в базу данных

Название функции/команды: cursor.execute('INSERT INTO names (first_name, last_name) VALUES (?, ?)', (name, surname))

Пример:
'''
cursor.execute('INSERT INTO names (first_name, last_name) VALUES (?, ?)', (name, surname))
'''
Описание:
Эта строка кода вставляет введенные пользователем имя и фамилию в таблицу "names" базы данных.
'''
'''
Шаг №5: Запись данных в HTML-файл

Название функции/команды: with open('index.html', 'w', encoding='utf-8') as file:

Пример:
'''
with open('index.html', 'w', encoding='utf-8') as file:
'''
Описание:
Этот блок кода открывает файл 'index.html' для записи в кодировке UTF-8. Файл будет использоваться для 
создания HTML-страницы.
'''
'''
Шаг №6: Запись HTML-кода в файл

Название функции/команды: file.write(...)

Пример:
'''
file.write('    <h3>{html.escape(name)}</h3>\n')
'''
Описание:
Этот блок кода записывает различные версии имени и фамилии в HTML-файл с разными стилями 
(стандартное, курсивное, жирное) и добавляет рамки для JPG и GIF изображений с соответствующими ссылками.
'''
'''
Шаг №7: Закрытие HTML-файла

Название функции/команды: file.write('...</body></html>')

Пример:
'''
file.write('''
</body>
</html>
''')
'''
Описание:
Эта строка кода закрывает HTML-код в файле после записи данных и завершает создание HTML-страницы.
'''
'''
Шаг №8: Открытие HTML-файла в браузере

Название функции/команды: webbrowser.open('index.html', new=2)

Пример:
'''
webbrowser.open('index.html', new=2)
'''
Описание:
Эта строка кода открывает созданный HTML-файл в браузере. Аргумент new=2 гарантирует, 
что файл открывается в новой вкладке.
'''
'''
Шаг №9: Закрытие соединения с базой данных

Название функции/команды: conn.close()

Пример:
'''
conn.close()
'''
Описание:
Эта строка кода закрывает соединение с базой данных после выполнения всех операций.
'''

'''
Как итог, этот код выполняет следующие функции:

1. Устанавливает соединение с базой данных SQLite, создает таблицу "names" и проверяет наличие или создает базу данных.
2. Вводит имя и фамилию от пользователя, проверяет их на корректность (только буквы русского или английского алфавита).
3. Записывает введенные данные в базу данных и создает HTML-файл с разными стилями отображения имени и фамилии,
а также добавляет рамки для JPG и GIF изображений.
4. Открывает созданный HTML-файл в браузере.
5. Закрывает соединение с базой данных.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Практической-Работы: 10-11 - ЯНВАРЯ 2024 года.
''''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением
Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №22: Консольные приложения (bash, shell). Упаковка приложений (.exe) - модуль auto_py_to_exe

Выполните следующие задания:

Задание №1
а) Напишите скрипт, который будет создавать виртуальное окружение и устанавливать туда библиотеку openpyxl.
б) Напишите пример приложения-калькулятор на input.
в) Соберите это приложение в исполняемый файл («.exe»).
'''
'''
Урок от 10.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
Само решение не не удалось записать сюда (СМ ЗАДАНИЕ) - Сам калькулятор - есть в отдельном файле .py
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 10-11 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашнее задание № 22: Консольные приложения (bash, shell). Упаковка приложений (.exe) модуль auto_py_to_exe
Выполните следующее задание:

Задание №1

а) Напишите скрипт, который будет создавать виртуальное окружение и устанавливать туда библиотеку python-docx.
б) Напишите пример приложения на input, которое получает текст от пользователя и создаёт word-файл с этим текстом.
в) Соберите это приложение в исполняемый файл («.exe»).
В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 10.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
Само решение не не удалось записать сюда (СМ ЗАДАНИЕ) - Сам word-файл - есть в отдельном файле .py
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Практической-Работы: 12-13 - ЯНВАРЯ 2024 года.
''''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением
Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №23: Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет складывать два числа, и спустя 1 секунду задержки возвращать результат.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал многопоточного запуска, с замером времени.
'''
'''
Урок от 12.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import threading
import time

# Задание №1

# a) Функция для сложения двух чисел с задержкой в 1 секунду
def add_numbers(x, y):
    time.sleep(1)
    return x + y

# b) Запуск циклом 100 таких функций и замер времени
start_time = time.time()

results = []
for _ in range(100):
    result = add_numbers(3, 4)  # Пример: сложение чисел 3 и 4
    results.append(result)

end_time = time.time()
print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))

# c) Функционал многопоточного запуска с замером времени
start_time = time.time()

results_mt = []
threads = []

def worker():
    result = add_numbers(3, 4)
    results_mt.append(result)

for _ in range(100):
    thread = threading.Thread(target=worker)
    thread.start()
    threads.append(thread)

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()
print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Далее более подробно:
'''
'''
Для начала импортируем все то что нам нужно:
'''
import threading
import time
'''
Шаг 1: Функция add_numbers
'''
def add_numbers(x, y):
    time.sleep(1)
    return x + y
'''
Описание:

Название функции: add_numbers.

Пример кода:

def add_numbers(x, y):: Определяет функцию с именем add_numbers, принимающую два аргумента x и y.
time.sleep(1): Вызывает функцию sleep из модуля time, приостанавливая выполнение программы на 1 секунду.
Это эмулирует задержку в функции.
return x + y: Возвращает сумму x и y.
Более подробное описание:

Функция add_numbers принимает два аргумента x и y и добавляет их. 
Однако, перед возвратом результата, функция "засыпает" на 1 секунду с использованием time.sleep(1).
Это добавляет задержку к выполнению функции, чтобы сделать ее более реалистичной для примера многопоточности.
'''
'''
Шаг 2: Блок кода для выполнения без использования потоков
'''
# b) Запуск циклом 100 таких функций и замер времени
start_time = time.time()

results = []
for _ in range(100):
    result = add_numbers(3, 4)  # Пример: сложение чисел 3 и 4
    results.append(result)

end_time = time.time()
print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Пример кода:

start_time = time.time(): Записывает текущее время в переменную start_time.
results = []: Создает пустой список для хранения результатов функций.
for _ in range(100):: Начинает цикл, который выполняется 100 раз.
result = add_numbers(3, 4): Вызывает функцию add_numbers с аргументами 3 и 4 (пример сложения чисел) и 
сохраняет результат в переменную result.
results.append(result): Добавляет результат в список results.
end_time = time.time(): Записывает текущее время в переменную end_time.
print("Time taken without threading: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время,
затраченное на выполнение цикла без использования потоков.

Более подробное описание:

В этом блоке кода мы запускаем цикл, который 100 раз вызывает функцию add_numbers с аргументами 3 и 4.
Результат каждого вызова добавляется в список results. Затем мы замеряем время, затраченное на выполнение цикла,
и выводим его в консоль.
'''
'''
Шаг 3: Блок кода для выполнения с использованием потоков
'''
# c) Функционал многопоточного запуска с замером времени
start_time = time.time()

results_mt = []
threads = []

def worker():
    result = add_numbers(3, 4)
    results_mt.append(result)

for _ in range(100):
    thread = threading.Thread(target=worker)
    thread.start()
    threads.append(thread)

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()
print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Пример кода:

start_time = time.time(): Записывает текущее время в переменную start_time.
results_mt = []: Создает пустой список для хранения результатов функций с использованием потоков.
threads = []: Создает пустой список для хранения объектов потоков.
def worker():: Определяет функцию worker, которая вызывает add_numbers и добавляет результат в список results_mt.
for _ in range(100):: Начинает цикл, который выполняется 100 раз.
thread = threading.Thread(target=worker): Создает объект потока, используя функцию worker в качестве цели.
thread.start(): Запускает поток.
threads.append(thread): Добавляет объект потока в список.
for thread in threads: и thread.join(): Ожидает завершения всех потоков.
end_time = time.time(): Записывает текущее время в переменную end_time.
print("Time taken with threading: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время,
затраченное на выполнение цикла с использованием потоков.

Более подробное описание:

В этом блоке кода мы используем многопоточность для выполнения функции add_numbers параллельно в 100 потоках.
Каждый поток вызывает функцию worker, которая в свою очередь вызывает add_numbers и добавляет результат 
в список results_mt. Мы замеряем время, затраченное на выполнение этого блока кода, и выводим его в консоль.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Возможные альтернативные варианты
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант 1: Использование concurrent.futures для многопоточности
'''
import concurrent.futures
import time

def add_numbers(x, y):
    time.sleep(1)
    return x + y

def worker(_):
    result = add_numbers(3, 4)
    return result

# Используем ThreadPoolExecutor для многопоточности
start_time = time.time()
with concurrent.futures.ThreadPoolExecutor() as executor:
    results_mt = list(executor.map(worker, range(100)))

end_time = time.time()
print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
В этом варианте используется модуль concurrent.futures, который предоставляет высокоуровневый интерфейс для работы
с потоками и процессами. Мы используем ThreadPoolExecutor для создания пула потоков и функцию executor.map для
многопоточного выполнения задач.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Функция add_numbers
'''
def add_numbers(x, y):
    time.sleep(1)
    return x + y
'''
Описание:

Название функции: add_numbers.

Пример кода:
def add_numbers(x, y):: Определяет функцию с именем add_numbers, принимающую два аргумента x и y.
time.sleep(1): Вызывает функцию sleep из модуля time, приостанавливая выполнение программы на 1 секунду. 
Это эмулирует задержку в функции.
return x + y: Возвращает сумму x и y.

Более подробное описание:

Функция add_numbers принимает два аргумента x и y и добавляет их. 
Однако, перед возвратом результата, функция "засыпает" на 1 секунду с использованием time.sleep(1). 
Это добавляет задержку к выполнению функции, чтобы сделать ее более реалистичной для примера многопоточности.
'''
'''
Шаг 2: Блок кода с использованием concurrent.futures
'''
# Используем ThreadPoolExecutor для многопоточности
start_time = time.time()
with concurrent.futures.ThreadPoolExecutor() as executor:
    results_mt = list(executor.map(worker, range(100)))

end_time = time.time()
print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Пример кода:
start_time = time.time(): Записывает текущее время в переменную start_time.
with concurrent.futures.ThreadPoolExecutor() as executor:: Создает ThreadPoolExecutor в контексте.
results_mt = list(executor.map(worker, range(100))): Запускает функцию worker в 100 потоках и получает результаты 
с использованием executor.map.
end_time = time.time(): Записывает текущее время в переменную end_time.
print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время,
затраченное на выполнение кода с использованием ThreadPoolExecutor.

Более подробное описание:

Мы используем ThreadPoolExecutor из concurrent.futures для создания пула потоков. 
В контексте блока with мы вызываем функцию worker в 100 потоках с использованием executor.map. 
Этот метод ожидает завершения всех потоков и возвращает результаты в виде списка. 
Затем мы замеряем время выполнения этого блока кода и выводим результат в консоль.
'''
'''
В итоге, код в Варианте 1 демонстрирует использование ThreadPoolExecutor для параллельного выполнения функции в
нескольких потоках, что может быть полезным при выполнении множества схожих задач.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант 2: Использование asyncio для асинхронного выполнения
'''
import asyncio
import time

async def add_numbers(x, y):
    await asyncio.sleep(1)
    return x + y

async def worker():
    result = await add_numbers(3, 4)
    return result

# Используем asyncio для асинхронного выполнения
start_time = time.time()

# Создаем новый цикл событий, если его нет
loop = asyncio.new_event_loop()
asyncio.set_event_loop(loop)

# Запускаем асинхронные задачи
results_mt = loop.run_until_complete(asyncio.gather(*(worker() for _ in range(100))))

end_time = time.time()

print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Определение асинхронных функций
'''
async def add_numbers(x, y):
    await asyncio.sleep(1)
    return x + y

async def worker():
    result = await add_numbers(3, 4)
    return result
'''
Описание:
add_numbers: Асинхронная функция, принимающая два аргумента x и y. 
Она приостанавливает выполнение на 1 секунду с помощью await asyncio.sleep(1) и затем возвращает сумму x и y.
worker: Еще одна асинхронная функция, вызывающая add_numbers с аргументами 3 и 4. 
Она также использует await, чтобы ждать завершения выполнения add_numbers и затем возвращает результат.
'''
'''
Шаг 2: Использование asyncio для асинхронного выполнения
'''
start_time = time.time()
loop = asyncio.get_event_loop()
tasks = [worker() for _ in range(100)]
results_mt = loop.run_until_complete(asyncio.gather(*tasks))
end_time = time.time()
'''
Описание:
start_time = time.time(): Записывает текущее время перед началом выполнения асинхронного кода.
loop = asyncio.get_event_loop(): Получает экземпляр цикла событий asyncio.
tasks = [worker() for _ in range(100)]: Создает список асинхронных задач, 
каждая из которых представляет собой вызов функции worker().
results_mt = loop.run_until_complete(asyncio.gather(*tasks)): Запускает выполнение всех асинхронных 
задач с помощью asyncio.gather() и блокирует выполнение до их завершения. 
Результаты сохраняются в results_mt.
end_time = time.time(): Записывает текущее время после завершения выполнения асинхронного кода.
'''
'''
Шаг 3: Вывод результатов
'''
print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time))
'''
Описание:
print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время, 
затраченное на выполнение всех асинхронных задач. В данном случае, это время выполнения 100 вызовов функции worker().
'''
'''
Более подробное описание:

Код использует асинхронные функции для эмуляции параллельного выполнения асинхронных задач. 
В цикле for создается список из 100 вызовов worker(). Затем с помощью asyncio.gather() 
запускаются все эти задачи параллельно, и код блокируется до их завершения. Время выполнения замеряется 
с использованием time.time().
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Можно прям и по этапно ("пошагово") - т.е. - а), б), в).
'''

'''
Шаг а): Напишите функцию, складывающую два числа с задержкой в 1 секунду
'''
import asyncio

async def add_numbers(x, y):
    await asyncio.sleep(1)  # Асинхронная задержка на 1 секунду
    return x + y
'''
Описание:

add_numbers: Асинхронная функция, принимающая два аргумента x и y.
Она приостанавливает выполнение на 1 секунду с помощью await asyncio.sleep(1) и затем возвращает сумму x и y.
'''

'''
Шаг 1б: Запустите циклом 100 таких функций, а также замерьте время
'''
import asyncio
import time

# Определение асинхронных функций
async def add_numbers(x, y):
    await asyncio.sleep(1)
    return x + y

async def worker():
    result = await add_numbers(3, 4)
    return result

# Использование asyncio для асинхронного выполнения
start_time = time.time()
loop = asyncio.get_event_loop()
tasks = [worker() for _ in range(100)]
results_mt = loop.run_until_complete(asyncio.gather(*tasks))
end_time = time.time()

print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

start_time = time.time(): Записывает текущее время перед началом выполнения асинхронного кода.
loop = asyncio.get_event_loop(): Получает экземпляр цикла событий asyncio.
tasks = [worker() for _ in range(100)]: Создает список асинхронных задач, каждая из которых представляет 
собой вызов функции worker().
results_mt = loop.run_until_complete(asyncio.gather(*tasks)): Запускает выполнение всех асинхронных задач 
с помощью asyncio.gather() и блокирует выполнение до их завершения. Результаты сохраняются в results_mt.
end_time = time.time(): Записывает текущее время после завершения выполнения асинхронного кода.
'''

'''
Шаг 1в: Добавьте функционал многопоточного запуска, с замером времени
'''
import concurrent.futures
import time

# Определение функции
def add_numbers(x, y):
    time.sleep(1)
    return x + y

# Использование ThreadPoolExecutor для многопоточности
start_time = time.time()
with concurrent.futures.ThreadPoolExecutor() as executor:
    results_mt = list(executor.map(add_numbers, [3] * 100, [4] * 100))
end_time = time.time()

print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

start_time = time.time(): Записывает текущее время перед началом выполнения кода с использованием ThreadPoolExecutor.
with concurrent.futures.ThreadPoolExecutor() as executor: Создает ThreadPoolExecutor для многопоточности.
results_mt = list(executor.map(add_numbers, [3] * 100, [4] * 100)): Запускает выполнение функции add_numbers в 
нескольких потоках с помощью executor.map() и сохраняет результаты в results_mt.
end_time = time.time(): Записывает текущее время после завершения выполнения кода с использованием ThreadPoolExecutor.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 12-13 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django

Дисциплина: Основы программирования на Python

Домашнее задание № 23 : Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1

а) Напишите функцию, которая будет создавать файл, с задержкой 1 секунду.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал многопоточного запуска, с замером времени.

** В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 12.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import threading
import time


# Шаг 1: Написание функции для создания файла с задержкой
def create_file_with_delay(file_number):
    time.sleep(1)  # Задержка в 1 секунду
    file_name = f"file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"File '{file_name}' created.")


# Шаг 2: Запуск циклом 100 функций и замер времени
start_time = time.time()

for i in range(1, 101):
    create_file_with_delay(i)

end_time = time.time()

print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))

# Шаг 3: Добавление функционала многопоточного запуска и замер времени
start_time = time.time()

threads = []
for i in range(1, 101):
    thread = threading.Thread(target=create_file_with_delay, args=(i,))
    threads.append(thread)
    thread.start()

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()

print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import threading
import time
'''
Описание: 
В этом шаге импортируются две библиотеки - threading для работы с многопоточностью и time для работы с временем.
'''
'''
Шаг 2: Определение функции для создания файла с задержкой
'''
def create_file_with_delay(file_number):
    time.sleep(1)  # Задержка в 1 секунду
    file_name = f"file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"File '{file_name}' created.")
'''
Описание: 
Эта функция create_file_with_delay принимает file_number в качестве аргумента, добавляет задержку в 1 секунду 
с использованием time.sleep(1), затем создает файл с именем, содержащим номер, и записывает в него текст.
В конце функция выводит сообщение о создании файла.
'''
'''
Шаг 3: Запуск циклом 100 функций и замер времени без использования многопоточности
'''
start_time = time.time()

for i in range(1, 101):
    create_file_with_delay(i)

end_time = time.time()

print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:
В этом шаге создается цикл, который запускает функцию create_file_with_delay 100 раз для создания файлов с задержкой.
Время начала измеряется с использованием time.time(), а затем выводится время выполнения после завершения цикла.
'''
'''
Шаг 4: Добавление функционала многопоточного запуска и замер времени
'''
start_time = time.time()

threads = []
for i in range(1, 101):
    thread = threading.Thread(target=create_file_with_delay, args=(i,))
    threads.append(thread)
    thread.start()

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()

print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:
В этом шаге создается новый цикл, который использует многопоточность. 
Создаются 100 потоков (threads), каждый из которых запускает функцию create_file_with_delay с уникальным file_number. 
Затем выполняется ожидание завершения всех потоков с использованием thread.join(). 
Время выполнения снова измеряется и выводится.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ЕЩЕ ВАРИАНТ
'''
import concurrent.futures
import asyncio
import time

def create_file_with_delay_thread(file_number):
    time.sleep(1)
    file_name = f"thread_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"Thread File '{file_name}' created.")

async def create_file_with_delay_asyncio(file_number):
    await asyncio.sleep(1)
    file_name = f"asyncio_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"AsyncIO File '{file_name}' created.")

async def main():
    # Используем ThreadPoolExecutor для многопоточности
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(create_file_with_delay_thread, range(1, 51))

    # Используем asyncio для асинхронного выполнения
    tasks = [create_file_with_delay_asyncio(i) for i in range(51, 101)]
    await asyncio.gather(*tasks)

start_time = time.time()
asyncio.run(main())
end_time = time.time()

print("Time taken with ThreadPoolExecutor and asyncio: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек само собой
'''
'''
Шаг 1:
'''
def create_file_with_delay_thread(file_number):
    time.sleep(1)
    file_name = f"thread_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"Thread File '{file_name}' created.")
'''
Описание:

Эта функция create_file_with_delay_thread представляет собой синхронную функцию,
которая создает файл с использованием задержки в 1 секунду. Она принимает аргумент file_number,
который используется для создания уникального имени файла. Функция использует time.sleep(1),
чтобы создать задержку в 1 секунду перед созданием файла. Затем она открывает файл с именем, 
сформированным на основе file_number, и записывает в него строку. Наконец, выводится сообщение в 
консоль о создании файла.
'''
'''
Шаг 2:
'''
async def create_file_with_delay_asyncio(file_number):
    await asyncio.sleep(1)
    file_name = f"asyncio_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"AsyncIO File '{file_name}' created.")
'''
Описание:

Эта функция create_file_with_delay_asyncio - это асинхронная функция, использующая async/await, 
чтобы создать файл с задержкой в 1 секунду. Она аналогична предыдущей функции, но с использованием асинхронных 
возможностей. Она также принимает аргумент file_number, формирует уникальное имя файла, создает задержку с 
использованием await asyncio.sleep(1), и затем создает и записывает файл, выводя сообщение о создании в консоль.
'''
'''
Шаг 3:
'''
async def main():
    # Используем ThreadPoolExecutor для многопоточности
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(create_file_with_delay_thread, range(1, 51))

    # Используем asyncio для асинхронного выполнения
    tasks = [create_file_with_delay_asyncio(i) for i in range(51, 101)]
    await asyncio.gather(*tasks)
'''
Описание:

Эта функция main является точкой входа для выполнения обоих подходов (многопоточного и асинхронного).
Внутри main сначала используется ThreadPoolExecutor для запуска функции create_file_with_delay_thread 50 раз в
многопоточной среде. Затем используется asyncio для выполнения функции create_file_with_delay_asyncio еще 50 раз в
асинхронной среде. Весь код выполняется асинхронно, и функция main ожидает завершения всех задач, 
используя await asyncio.gather(*tasks).
'''
start_time = time.time()
asyncio.run(main())
end_time = time.time()

print("Time taken with ThreadPoolExecutor and asyncio: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Здесь вычисляется и выводится общее время выполнения обоих подходов с использованием многопоточности и асинхронности.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #


< !--  ########################################################################################## -->
< !-- HTML  # 1 Recap + Вводный ПЕРВЫЙ УРОК - 08.01.2024 -->

< !-- < html >
< head >

< meta
charset = "UTF-8" / >
< title > Hello
HTML < / title >
< / head >
< body >

hello
world
Привет
Мир

< img
src = "https://pngicon.ru/file/uploads/serdce.png"
height = "150px"
width = "150px" / >
< h1 > Вау - я
программист < / h1 >
< / body >
< / html > -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !-- 2024
_HTML_LESSON_№2
_Структура_HTML.Теги_Атрибуты - 15.01
.2024 -->

< !--

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 2: Структура
HTML.Теги.Атрибуты.Правила
языка
разметки
-->

< !DOCTYPE
html >
< html
lang = "ru" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Страница
с
текстом
и
заголовками < / title >
< / head >
< body >

< !-- 1.
Создать
страницу
с
текстом
по
примеру
на
картинке. -->

< h2 >
Стихотворение
< / h2 >
< p >

Мириады
маленьких
дел < br >
Пьют
по
капле
гаснущий
день, < br >
А < i > дела
большие < / i > сушит
жажда. < br >
Оставляя
все
на «потом», < br >
Прозреваем
задним
числом, < br >
Только
год < b > не
повторится < / b > дважды. < br >
< br >
И.Тальков

< / p >

< !-- 2.
Повторите
по
данному
образцу: -->
< h1 > Это
заголовок < / h1 >
< h2 > Это
заголовок < / h2 >
< h3 > Это
заголовок < / h3 >
< h4 > Это
заголовок < / h4 >

< p > Это < b > абзац < / b >.< / p >
< p > Это
еще < i > абзац < / i >.< / p >

< h1 > < i > Это
заголовок
h1 < / i > < / h1 >

< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->







# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Дата выполнения Практической-Работы: 17-18 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Практическая работа №24: Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет загружать изображение.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал мультипроцессорного запуска, с замером времени.

*г) Необходимо провести вычисление факториала в параллельных процессах.
'''
'''
Урок от 17.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
import time
import multiprocessing
from PIL import Image

def load_image(image_path):
    """Функция для загрузки изображения."""
    with Image.open(image_path) as img:
        # Просто загрузка изображения без выполнения каких-либо операций.
        pass

def process_images_sequential(image_paths):
    """Функция для последовательной обработки изображений."""
    start_time = time.time()

    for path in image_paths:
        load_image(path)

    end_time = time.time()
    print(f"Последовательная обработка: {end_time - start_time} секунд")

def process_images_parallel(image_paths):
    """Функция для параллельной обработки изображений."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    pool.map(load_image, image_paths)
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельная обработка: {end_time - start_time} секунд")

def calculate_factorial(n):
    """Функция для вычисления факториала."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    return result

def calculate_factorial_parallel(n):
    """Функция для параллельного вычисления факториала."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    results = pool.map(calculate_factorial, range(1, n + 1))
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельное вычисление факториала: {end_time - start_time} секунд")
    return results

def show_images(image_paths):
    """Функция для отображения изображений."""
    for path in image_paths:
        img = Image.open(path)
        img.show()

if __name__ == "__main__":
    # Задайте список путей к изображениям
    image_paths = ["E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/1.jpg",
                   "E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/2.jpg"]  # Замените на реальные пути

    # Задание 1б: Запуск циклом 100 функций
    for _ in range(100):
        load_image(image_paths[0])  # Загружаем одно изображение

    # Задание 1в: Параллельная обработка с замером времени
    process_images_parallel(image_paths)

    # Задание 1г: Параллельное вычисление факториала
    n = 5
    result_parallel = calculate_factorial_parallel(n)
    print(f"Результаты параллельного вычисления факториала: {result_parallel}")

    # Отображение изображений
    show_images(image_paths)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1. Загрузка изображения
'''
def load_image(image_path):
    """Функция для загрузки изображения."""
    with Image.open(image_path) as img:
        # Просто загрузка изображения без выполнения каких-либо операций.
        pass
'''
Описание: 
Эта функция использует библиотеку PIL (Pillow) для загрузки изображения по указанному пути. 
В данном случае, операции с изображением не выполняются, и функция служит примером базовой операции загрузки.
'''
'''
Шаг №2. Последовательная обработка изображений
'''
def process_images_sequential(image_paths):
    """Функция для последовательной обработки изображений."""
    start_time = time.time()

    for path in image_paths:
        load_image(path)

    end_time = time.time()
    print(f"Последовательная обработка: {end_time - start_time} секунд")
'''
Описание: 
Эта функция выполняет последовательную обработку изображений. 
Она вызывает функцию load_image для каждого пути к изображению в списке image_paths. 
Замеряется время выполнения операции, и результат выводится на экран.
'''
'''
Шаг №3. Параллельная обработка изображений
'''
def process_images_parallel(image_paths):
    """Функция для параллельной обработки изображений."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    pool.map(load_image, image_paths)
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельная обработка: {end_time - start_time} секунд")
'''
Описание: 
Эта функция выполняет параллельную обработку изображений, используя модуль multiprocessing. 
Пул процессов создается с количеством процессов, равным количеству ядер процессора. 
Функция load_image применяется к каждому пути к изображению в списке image_paths. 
Замеряется время выполнения операции, и результат выводится на экран.
'''
'''
Шаг №4. Вычисление факториала
'''
def calculate_factorial(n):
    """Функция для вычисления факториала."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    return result
'''
Описание: 
Эта функция принимает целое число n и вычисляет его факториал.
'''
'''
Шаг №5. Параллельное вычисление факториала
'''
def calculate_factorial_parallel(n):
    """Функция для параллельного вычисления факториала."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    results = pool.map(calculate_factorial, range(1, n + 1))
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельное вычисление факториала: {end_time - start_time} секунд")
    return results
'''
Описание: 
Эта функция выполняет параллельное вычисление факториала для чисел от 1 до n с использованием пула процессов. 
Замеряется время выполнения операции, и результат выводится на экран.
'''
'''
Шаг №6. Отображение изображений
'''
def show_images(image_paths):
    """Функция для отображения изображений."""
    for path in image_paths:
        img = Image.open(path)
        img.show()
'''
Описание: 
Эта функция использует библиотеку PIL для открытия и отображения изображений, указанных в списке image_paths.
'''
'''
Шаг №7. Основной код (выполняется при запуске файла)
'''
if __name__ == "__main__":
    # Задайте список путей к изображениям
    image_paths = ["E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/1.jpg",
                   "E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/2.jpg"]  # Замените на реальные пути

    # Задание 1б: Запуск циклом 100 функций
    for _ in range(100):
        load_image(image_paths[0])  # Загружаем одно изображение

    # Задание 1в: Параллельная обработка с замером времени
    process_images_parallel(image_paths)

    # Задание 1г: Параллельное вычисление факториала
    n = 5
    result_parallel = calculate_factorial_parallel(n)
    print(f"Результаты параллельного вычисления факториала: {result_parallel}")

    # Отображение изображений
    show_images(image_paths)
'''
Описание: 
В основной части кода задаются пути к изображениям. 
Затем выполняются задания б, в и г. 
Задание б представляет собой цикл, в котором 100 раз загружается одно изображение. 
Задание в - параллельная обработка изображений с замером времени. 
Задание г - параллельное вычисление факториала для числа 5. 
Наконец, вызывается функция для отображения изображений.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #



# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 17-18 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Домашнее задание № 24 : Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1

а) Напишите функцию, которая будет создавать файл и записывать в него рандомное число, с задержкой 1 секунду.
б) Запустите циклом 1000 таких функций, а также замерьте время.
в) Добавьте функционал мультипоточного запуска, с замером времени. 
Обязательно посмотрите нагрузку на ЦП в этот момент (через диспетчер задач).

*г) Необходимо провести вычисление факториала в параллельных процессах.
** В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 12.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import time
import random
import threading
import multiprocessing

# Шаг а: Функция для создания файла и записи в него рандомного числа с задержкой
def write_random_number(file_path):
    """Создает файл и записывает в него рандомное число с задержкой 1 секунда."""
    with open(file_path, 'w') as file:
        random_number = random.randint(1, 100)
        file.write(str(random_number))
        time.sleep(1)

# Шаг б: Функция для запуска циклом 1000 функций с замером времени
def run_sequential():
    """Запускает циклом 1000 функций и замеряет время."""
    start_time = time.time()

    for _ in range(1000):
        write_random_number(f"file{_}.txt")

    end_time = time.time()
    print(f"Время, затраченное на последовательное выполнение: {end_time - start_time:.4f} секунд")

# Шаг в: Функция для многопоточного запуска с замером времени
def run_multithreaded():
    """Многопоточный запуск с замером времени."""
    start_time = time.time()

    threads = []
    for _ in range(1000):
        thread = threading.Thread(target=write_random_number, args=(f"file{_}.txt",))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    end_time = time.time()
    print(f"Время, затраченное на многопоточное выполнение: {end_time - start_time:.4f} секунд")

# Шаг г: Функция для параллельного вычисления факториала
def calculate_factorial(n, result_queue):
    """Вычисляет факториал числа n и помещает результат в очередь."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    result_queue.put(result)

# Шаг г (продолжение): Функция для параллельного запуска вычисления факториала
def run_multiprocess_factorial():
    """Параллельное вычисление факториала."""
    start_time = time.time()

    result_queue = multiprocessing.Queue()
    processes = []

    for _ in range(1000):
        process = multiprocessing.Process(target=calculate_factorial, args=(5, result_queue))
        processes.append(process)
        process.start()

    for process in processes:
        process.join()

    # Получаем результаты из очереди
    factorial_results = [result_queue.get() for _ in range(1000)]

    end_time = time.time()
    print(f"Время, затраченное на параллельное вычисление факториала: {end_time - start_time:.4f} секунд")
    print(f"Результаты факториала: {factorial_results}")

if __name__ == '__main__':
    # Шаг б: Запуск циклом 1000 функций
    run_sequential()

    # Шаг в: Многопоточный запуск
    run_multithreaded()

    # Шаг г: Параллельное вычисление факториала
    run_multiprocess_factorial()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг а: Функция для создания файла и записи в него рандомного числа с задержкой
'''
def write_random_number(file_path):
    """Создает файл и записывает в него рандомное число с задержкой 1 секунда."""
    with open(file_path, 'w') as file:
        random_number = random.randint(1, 100)
        file.write(str(random_number))
        time.sleep(1)
'''
Описание: 
Эта функция создает файл по указанному пути (file_path) и записывает в него случайное число от 1 до 100. 
После этого функция "засыпает" на 1 секунду с помощью time.sleep(1).
'''
'''
Шаг б: Функция для запуска циклом 1000 функций с замером времени
'''
def run_sequential():
    """Запускает циклом 1000 функций и замеряет время."""
    start_time = time.time()

    for _ in range(1000):
        write_random_number(f"file{_}.txt")

    end_time = time.time()
    print(f"Время, затраченное на последовательное выполнение: {end_time - start_time:.4f} секунд")
'''
Описание: 
Эта функция запускает цикл из 1000 итераций. 
На каждой итерации она вызывает функцию write_random_number, создавая 1000 файлов. 
Затем она замеряет время выполнения и выводит результат.
'''
'''
Шаг в: Функция для многопоточного запуска с замером времени
'''
def run_multithreaded():
    """Многопоточный запуск с замером времени."""
    start_time = time.time()

    threads = []
    for _ in range(1000):
        thread = threading.Thread(target=write_random_number, args=(f"file{_}.txt",))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    end_time = time.time()
    print(f"Время, затраченное на многопоточное выполнение: {end_time - start_time:.4f} секунд")
'''
Описание: 
Эта функция также создает 1000 файлов, но использует многопоточность. 
Для каждой итерации цикла создается новый поток (threading.Thread), который выполняет функцию write_random_number. 
Затем происходит ожидание завершения всех потоков (thread.join()), и замеряется время выполнения.
'''
'''
Шаг г: Функция для параллельного вычисления факториала
'''
def calculate_factorial(n, result_queue):
    """Вычисляет факториал числа n и помещает результат в очередь."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    result_queue.put(result)
'''
Описание: 
Эта функция вычисляет факториал числа n и помещает результат в указанную очередь (result_queue).
'''
'''
Шаг г (продолжение): Функция для параллельного запуска вычисления факториала
'''
def run_multiprocess_factorial():
    """Параллельное вычисление факториала."""
    start_time = time.time()

    result_queue = multiprocessing.Queue()
    processes = []

    for _ in range(1000):
        process = multiprocessing.Process(target=calculate_factorial, args=(5, result_queue))
        processes.append(process)
        process.start()

    for process in processes:
        process.join()

    # Получаем результаты из очереди
    factorial_results = [result_queue.get() for _ in range(1000)]

    end_time = time.time()
    print(f"Время, затраченное на параллельное вычисление факториала: {end_time - start_time:.4f} секунд")
    print(f"Результаты факториала: {factorial_results}")
'''
Описание: 
Эта функция создает 1000 процессов (multiprocessing.Process), каждый из которых вычисляет факториал числа 5. 
Затем происходит ожидание завершения всех процессов, получение результатов из очереди и вывод времени выполнения
и результатов.
'''
'''
На сладкое
'''
if __name__ == '__main__':
    # Шаг 1б: Запуск циклом 1000 функций
    run_sequential()

    # Шаг 1в: Многопоточный запуск
    run_multithreaded()

    # Шаг 1г: Параллельное вычисление факториала
    run_multiprocess_factorial()
'''
Описание: 
Скажем так, основная часть кода запускает три разные функции для сравнения времени выполнения: последовательное
выполнение, многопоточное выполнение и параллельное вычисление факториала.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Дата выполнения Практической-Работы: 19-20 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Практическая работа №25: Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет загружать данные с jsonplaceholder.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал асинхронной работы, с замером времени.

 - - - КАК ЭТО ПОНЯЛ Я - - - 
*) *) Запустите сначала 100, затем 200, и на последок 500 таких функций и поработайте с файлами. 
Сделайте запись чисел в файл 1 : 1 в файл 2 : 1, 2 в файл 3 : 1, 2, 3 и так далее,
чтобы задача была цикличной и вариативной.
Затем посмотрите сколько времени уходит на исполнение и насколько преимущественнее тот или иной подход.
Сравните и выведите результаты на экран. На сколько эффективна мультипоточность и, или  мультипроцессорность.
'''
'''
Урок от 19.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''' Вариант №1 - (учитываются пункты только а), б), в)) '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp  # Добавлен импорт aiohttp
import asyncio  # Добавлен импорт asyncio
import time  # Добавлен импорт time
import requests  # Добавлен импорт requests

# Асинхронная функция для загрузки данных по URL с использованием aiohttp
async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()

# Асинхронная функция для асинхронного выполнения запросов к нескольким URL
async def asynchronous_fetch(urls):
    # Создание сессии для aiohttp
    async with aiohttp.ClientSession() as session:
        # Создание списка задач (coroutines) для каждого URL
        tasks = [fetch_data(session, url) for url in urls]
        # Ожидание выполнения всех задач и возврат результатов
        return await asyncio.gather(*tasks)

# Синхронная функция для синхронного выполнения запросов к нескольким URL
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        # Синхронный запрос с использованием requests
        response = requests.get(url)
        data = response.json()
        # Обработка данных, если необходимо
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")

# Асинхронная точка входа программы
async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавьте остальные URL
    ]

    # Запуск задания №1 - б) (синхронный)
    synchronous_fetch(urls)

    # Запуск задания №1 - в) (асинхронный)
    start_time = time.time()
    await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения: {elapsed_time:.2f} секунд")

# Проверка, что скрипт выполняется как программа, а не импортируется как модуль
if __name__ == "__main__":
    # Запуск асинхронной программы с использованием asyncio
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек
'''
import aiohttp
import asyncio
import time
import requests
'''
aiohttp - библиотека для асинхронных HTTP-запросов.
asyncio - библиотека для асинхронного программирования.
time - для замера времени выполнения.
requests - библиотека для синхронных HTTP-запросов.
'''
'''
Асинхронная функция для загрузки данных по URL с использованием aiohttp
'''
async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()
'''
fetch_data - асинхронная функция, которая использует aiohttp для выполнения асинхронного HTTP-запроса.
Возвращает JSON-данные из ответа.
'''
'''
Асинхронная функция для асинхронного выполнения запросов к нескольким URL
'''
async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)
'''
asynchronous_fetch - асинхронная функция, которая создает сессию aiohttp и выполняет несколько асинхронных 
HTTP-запросов параллельно.
'''
'''
Синхронная функция для синхронного выполнения запросов к нескольким URL
'''
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")
'''
synchronous_fetch - синхронная функция, которая выполняет HTTP-запросы с использованием библиотеки requests.
'''
'''
Асинхронная точка входа программы
'''
async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавьте остальные URL
    ]

    synchronous_fetch(urls)
    start_time = time.time()
    await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения: {elapsed_time:.2f} секунд")
'''
main - асинхронная точка входа программы. 
Запускает сначала синхронное выполнение, затем асинхронное, и выводит время выполнения обеих операций.
'''
'''
Проверка, что скрипт выполняется как программа
'''
if __name__ == "__main__":
    asyncio.run(main())
'''
Проверяет, что скрипт выполняется как программа, а не импортируется как модуль,
и запускает асинхронную программу с использованием asyncio.run(main()).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''' Вариант №2 - (учитываются Все пункты)'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import requests
import json
import time
import concurrent.futures

def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)

def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(', '.join(map(str, data)) + '\n')

def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)
    write_to_file(f'file{task_id % 3 + 1}.txt', data)

def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")

def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")

def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")

if __name__ == "__main__":
    main()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек
'''
import requests
import json
import time
import concurrent.futures
'''
requests: библиотека для выполнения HTTP-запросов.
json: модуль для работы с данными в формате JSON.
time: модуль для работы со временем.
concurrent.futures: стандартная библиотека для работы с параллельным выполнением кода.
'''
'''
Функция load_data
'''
def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)
'''
load_data: 
функция выполняет HTTP GET-запрос по заданному URL, затем использует json.loads для загрузки данных из 
текстового ответа в формате JSON.
'''
'''
Функция write_to_file
'''
def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(', '.join(map(str, data)) + '\n')
'''
write_to_file: функция записывает данные в файл. Каждый вызов функции добавляет новую строку в файл.
'''
'''
Функция task
'''
def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)
    write_to_file(f'file{task_id % 3 + 1}.txt', data)
'''
task: функция, представляющая задачу. 
Она формирует URL с использованием task_id, загружает данные с помощью load_data и записывает их в файл.
'''
'''
Функция synchronous_execution
'''
def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")
'''
synchronous_execution: функция для синхронного выполнения 100 задач. Измеряется время начала и конца выполнения.
'''
'''
Функция asynchronous_execution
'''
def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")
'''
asynchronous_execution: функция для асинхронного выполнения 100 задач с использованием пула потоков.
Измеряется время начала и конца выполнения.
'''
'''
Функция main
'''
def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")
'''
main: функция, запускающая синхронное и асинхронное выполнение, а также асинхронное выполнение с разным числом задач.
Измеряется время каждого выполнения.
'''
'''
Проверка, что скрипт выполняется как программа
'''
if __name__ == "__main__":
    main()
'''
Проверка, что скрипт выполняется напрямую (а не импортирован как модуль), и запуск функции main().
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''' Еще '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import requests
import json
import time
import concurrent.futures


def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)


def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(f'{data}\n')


def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)

    # Модификация: запись данных в файлы
    write_to_file(f'file{task_id % 3 + 1}.txt', f'{task_id}: {data}')


def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")


def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")


def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")


if __name__ == "__main__":
    main()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Эти строки отвечают за импорт необходимых библиотек.
'''
import requests
import json
import time
import concurrent.futures
'''
Эта функция load_data принимает URL в качестве аргумента, отправляет GET-запрос по указанному URL с помощью библиотеки
requests, получает ответ и затем использует json.loads для преобразования JSON-строки в Python-объект. 
Возвращает этот объект.
'''
def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)
'''
Функция write_to_file принимает имя файла (filename) и данные (data) и открывает файл в режиме добавления ('a').
Затем она записывает данные в файл, добавляя символ новой строки после каждой записи.
'''
def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(f'{data}\n')
'''
Функция task принимает task_id, формирует URL с использованием этого идентификатора, 
загружает данные с помощью функции load_data, а затем записывает эти данные в соответствующий файл с использованием 
функции write_to_file. Файл выбирается с учетом остатка от деления task_id на 3, 
чтобы обеспечить цикличную запись в 3 файла.
'''
def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)

    # Модификация: запись данных в файлы
    write_to_file(f'file{task_id % 3 + 1}.txt', f'{task_id}: {data}')
'''
Функция synchronous_execution выполняет 100 задач синхронно (последовательно) и замеряет время выполнения.
'''
def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")
'''
Функция asynchronous_execution выполняет 100 
задач асинхронно с использованием многозадачности в пуле потоков и замеряет время выполнения.
'''
def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")
'''
В функции main вызываются три варианта выполнения: синхронное, асинхронное с 100 задачами и 
асинхронное с разным числом задач (100, 200, 500). Замеряется время выполнения каждого варианта.
'''
def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")

if __name__ == "__main__":
    main()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Можно модифицировать функцию main следующим образом, чтобы выполнять асинхронное выполнение с разным числом задач:
'''
def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение с разным числом задач
    for num_tasks in [100, 200, 500, 1000]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")

if __name__ == "__main__":
    main()
'''
Тут добавлена задача с 1000 задачами.
Можно добавить или изменить значения в списке [100, 200, 500, 1000], чтобы отражать нужные варианты выполнения.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''Рубрика "ЭКСПЕРЕМЕНТЫ" '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет загружать данные с jsonplaceholder.
б) Запустите циклом сначала 100 таких функций, затем 200 таких функций, 500 таких функций, 1000 таких функций,
2000 таких функций, 5000 таких функций, а также замерьте время.
в) Добавьте функционал асинхронной работы, с замером времени.


Дополнительно:

Запустите сначала 100, затем 200, и напоследок 500, 1000 таких функций и поработайте с файлами.
Затем сделайте запись чисел в файл 1 : 1 в файл 2 : 1, 2 в файл 3 : 1, 2, 3 и так далее,
чтобы задача была цикличной и вариативной.
Затем посмотрите сколько времени уходит на исполнение и насколько преимущественнее тот или иной подход.
Сравните и выведите результаты на экран. На сколько эффективна мультипоточность и, или  мультипроцессорность.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import time
import requests
import concurrent.futures

async def fetch_data(session, url):
    async with session.get(url) as response:
        if response.status == 200:
            return await response.json()
        else:
            response.raise_for_status()

async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)

def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
        # Обработка данных, если необходимо
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")

async def write_to_file(file, data):
    async with aiofiles.open(file, 'a') as f:
        await f.write(data + '\n')

async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавьте остальные URL
    ]

    # Задание 1б: Запуск циклом 100 функций
    start_time = time.time()
    for _ in range(100):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (100 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 200 функций
    start_time = time.time()
    for _ in range(200):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (200 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 500 функций
    start_time = time.time()
    for _ in range(500):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (500 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 1000 функций
    start_time = time.time()
    for _ in range(1000):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (1000 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 2000 функций
    start_time = time.time()
    for _ in range(2000):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (2000 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 5000 функций
    start_time = time.time()
    for _ in range(5000):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (5000 функций): {elapsed_time:.2f} секунд")

    # Дополнительно) Запуск сначала 100, затем 200, и на последок 500, 1000 функций
    for num_tasks in [100, 200, 500, 1000]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(asynchronous_fetch, [urls] * num_tasks)
        elapsed_time = time.time() - start_time
        print(f"Асинхронное выполнение с {num_tasks} задачами: {elapsed_time:.2f} секунд")

        # Запись чисел в файлы
        for i in range(num_tasks):
            data = ', '.join(map(str, range(1, i + 2)))
            await write_to_file(f"file{i + 1}.txt", data)

if __name__ == "__main__":
    asyncio.run(main())

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Функция fetch_data

Пример кода:
'''
async def fetch_data(session, url):
    async with session.get(url) as response:
        if response.status == 200:
            return await response.json()
        else:
            response.raise_for_status()
'''
Полное и подробное описание:
Эта функция представляет собой асинхронный метод, который использует библиотеку aiohttp для выполнения 
HTTP-запроса к указанному URL. Она принимает сессию (session) и URL, а затем асинхронно отправляет GET-запрос. 
Если статус ответа равен 200, функция возвращает данные в формате JSON. В противном случае вызывается исключение 
для обработки ошибки.
'''
'''
Шаг №2: Функция asynchronous_fetch

Пример кода:
'''
async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiohttp для асинхронного выполнения нескольких HTTP-запросов.
Она создает сессию (ClientSession), а затем формирует список задач (tasks), каждая из которых представляет собой
вызов функции fetch_data для каждого URL из переданного списка. Затем функция использует asyncio.gather для 
параллельного выполнения всех задач и возвращает результат.
'''
'''
Шаг №3: Функция synchronous_fetch

Пример кода:
'''
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
        # Обработка данных, если необходимо
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")
'''
Полное и подробное описание:
Эта функция представляет собой синхронный метод для выполнения HTTP-запросов с использованием библиотеки requests. 
Она проходит по каждому URL в переданном списке, отправляет GET-запрос и обрабатывает данные (если это необходимо). 
Время выполнения замеряется, и результат выводится на экран.
'''
'''
Шаг №4: Функция write_to_file

Пример кода:
'''
async def write_to_file(file, data):
    async with aiofiles.open(file, 'a') as f:
        await f.write(data + '\n')
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiofiles для асинхронной записи данных в файл. 
Она принимает имя файла (file) и данные (data). 
С помощью aiofiles.open открывается файл в режиме добавления ('a'), и затем происходит асинхронная запись 
переданных данных с добавлением новой строки.
'''
'''
Шаг №5: Функция main

Пример кода:
'''
async def main():
    # ... (Пропущен код для определения списка URLs)

    # Задание 1б: Запуск циклом 100 функций
    start_time = time.time()
    for _ in range(100):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (100 функций): {elapsed_time:.2f} секунд")

    # ... (Аналогичные блоки для 200, 500, 1000, 2000, 5000 функций)

    # Дополнительно) Запуск сначала 100, затем 200, и на последок 500, 1000 функций
    for num_tasks in [100, 200, 500, 1000]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(asynchronous_fetch, [urls] * num_tasks)
        elapsed_time = time.time() - start_time
        print(f"Асинхронное выполнение с {num_tasks} задачами: {elapsed_time:.2f} секунд")

        # Запись чисел в файлы
        for i in range(num_tasks):
            data = ', '.join(map(str, range(1, i + 2)))
            await write_to_file(f"file{i + 1}.txt", data)

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Эта асинхронная функция main выполняет несколько задач. 
В первой части она запускает циклы с разным числом асинхронных функций asynchronous_fetch для измерения времени 
выполнения. Затем, в дополнительном блоке, используется многопоточный подход (ThreadPoolExecutor), чтобы 
запустить асинхронные функции с разным числом задач. После этого данные записываются в файлы согласно заданным условиям.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 19-20 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Домашнее задание № 25 : Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1

а) Напишите функцию, которая будет загружать информацию сразу с 100 ссылок.
б) Запустите эту функцию, а также замерьте время.
в) Добавьте функционал асинхронного запуска, с замером времени. Обязательно посмотрите нагрузку на ЦП в этот момент (через диспетчер задач).

*) Запустите сначала 100, затем 200, и напоследок 500 таких функций и поработайте с файлами. 
Сделайте запись чисел в файл 1 : 1 в файл 2 : 1, 2 в файл 3 : 1, 2, 3 и так далее,
чтобы задача была цикличной и вариативной.
Затем посмотрите сколько времени уходит на исполнение и насколько преимущественнее тот или иной подход.
Сравните и выведите результаты на экран. На сколько эффективна мультипоточность и, или  мультипроцессорность.

** В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 19.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import time
import concurrent.futures

async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()

async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)

def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")

async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавить другие URL-адреса
    ]

    # Запустите асинхронную функцию с разным количеством задач и измерьте время выполнения
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        for _ in range(num_tasks):
            await asynchronous_fetch(urls)
        elapsed_time = time.time() - start_time
        print(f"Время асинхронного выполнения ({num_tasks} задач): {elapsed_time:.2f} секунд")

        # Задача *: Запись чисел в файлы
        for i in range(num_tasks):
            with open(f"file{i + 1}.txt", "a") as file:
                file.write(", ".join(map(str, range(1, i + 2))) + "\n")

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Функция fetch_data

Пример кода:
'''
async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiohttp для выполнения HTTP-запроса по указанному URL. 
Сначала создается асинхронный контекст сессии (async with session.get(url) as response), 
затем функция ожидает ответ (await response.json()), который интерпретируется как JSON. 
Функция возвращает полученные данные.
'''
'''
Шаг №2: Функция asynchronous_fetch

Пример кода:
'''
async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiohttp для асинхронного выполнения нескольких HTTP-запросов. 
Она создает сессию (ClientSession) и формирует список задач (tasks), каждая из которых представляет собой вызов 
функции fetch_data для каждого URL из переданного списка. 
Затем функция использует asyncio.gather для параллельного выполнения всех задач и возвращает результат.
'''
'''
Шаг №3: Функция synchronous_fetch

Пример кода:
'''
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")
'''
Полное и подробное описание:
Эта функция представляет собой синхронный метод для выполнения HTTP-запросов с использованием библиотеки requests. 
Она проходит по каждому URL в переданном списке, отправляет GET-запрос и обрабатывает данные (если это необходимо). 
Время выполнения замеряется, и результат выводится на экран.
'''
'''
Шаг №4: Функция main

Пример кода:
'''
async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавить другие URL-адреса
    ]

    # Запустите асинхронную функцию с разным количеством задач и измерьте время выполнения
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        for _ in range(num_tasks):
            await asynchronous_fetch(urls)
        elapsed_time = time.time() - start_time
        print(f"Время асинхронного выполнения ({num_tasks} задач): {elapsed_time:.2f} секунд")

        # Задача X: Запись чисел в файлы
        for i in range(num_tasks):
            with open(f"file{i + 1}.txt", "a") as file:
                file.write(", ".join(map(str, range(1, i + 2))) + "\n")

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Эта асинхронная функция main выполняет несколько задач. 
В первой части она запускает циклы с разным числом асинхронных функций asynchronous_fetch для измерения 
времени выполнения. Затем, в дополнительном блоке, используется многопоточный подход (ThreadPoolExecutor), 
чтобы запустить асинхронные функции с разным числом задач. После этого данные записываются в файлы согласно 
заданным условиям.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
22.01.2024 - Домашняя и практическая - работы.
<!-- ########################################################################################## -->
Практическая работа

Курс: Разработка интерфейса на JavaScript
Дисциплина: Основы HTML и CSS
Тема занятия №3 Введение в CSS. Форматирование текста при помощи CSS
1. Создать HTML-страницу “Vehicle”
Текст можете найти в архиве Classwork в папке Материалы к занятию.

2. Создать HTML-страницу “Lorem Ipsum”
Текст можете найти в архиве Classwork в папке Материалы к занятию.

3. Создать HTML-страницу “Математические формулы”

Текст можете найти в архиве Classwork в папке Материалы к занятию.
Используйте здесь теги h1-h6, span, p, sup, sub и спецсимволы.
<!-- ########################################################################################## -->

Шрифты находятся в папке Материалы к занятию в архиве Fonts.

1. Пример подключение шрифтов через ссылку
2. Пример подключения через локальные файлы шрифтов
<!-- ########################################################################################## -->

№1

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Двигатель</title>
    <style>
        body {
            font-family: 'Tahoma', serif;
            line-height: 1.6;
            margin: 18px;
        }

        h1 {
            color: #00c264;
        }
        h2 {
            color: #9b05ff;
        }
        h3 {
            color: #ff0505;
        }
        h4 {
            color: #00a0eb;
        }
        h5 {
            color: #00a0eb;
        }
        h6 {
            color: #ffff05;
        }
        p {
            text-align: justify;
        }
        .no-underline {
    text-decoration: none;
  }
        .larger-font {
    font-size: 20px;
  }
        .larger-font2 {
    font-size: 40px;
  }
  .smaller-text {
    font-size: 58%; /* Измените на желаемый размер шрифта */
  }
        .slightly-bold {
    font-weight: bold; /* или font-weight: 600; */
  }
         .subtle-line {
    border: 2px solid #9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; /* Добавьте отступы по желанию */
  }
         .colored-letter {
    color: #00a0eb; /* Измените на желаемый цвет */
  }
    </style>

</head>
<body>
    <h1
            class="larger-font2">
        Двигатель
    </h1>
    <!-- Текст о двигателе с сайта https://ru.wikipedia.org/wiki/Двигатель -->

    <p
        class="larger-font">
    <b>Дви́гатель</b> — устройство, преобразующее
        какой-либо вид <a href="https://ru.wikipedia.org/wiki/Энергия" class="no-underline">энергии</a>
         в <a href="https://ru.wikipedia.org/wiki/Механическая_энергия" class="no-underline">механическую</a> работу.
    Термин <b>мотор</b> <a href="https://ru.wikipedia.org/wiki/Заимствование"class="no-underline"> заимствован</a> в
        первой половине XIX века из немецкого языка<b
            class="smaller-text colored-letter">[1]</b>
        (нем. Motor — «двигатель», от лат. mōtor — «приводящий в движение») и преимущественно им называют электрические
        двигатели и двигатели внутреннего сгорания<b
            class="smaller-text colored-letter">
        [2]</b>.
    </p>
    <p
            class="larger-font">
        Двигатели подразделяют на первичные и вторичные.
    К первичным относят непосредственно преобразующие природные энергетические ресурсы в механическую работу,
    а ко вторичным — преобразующие энергию, выработанную или накопленную другими источниками<b
            class="smaller-text colored-letter">
        [3]</b>.
    </p>
<hr class="subtle-line">
    <p>
    <h4>Примечания</h4>
    <h5>[1]</h5>Шанский Н. М., Боброва Т. А. Кот // Школьный этимологический словарь русского языка.
    Происхождение слов. — 7-е изд., стереотип.. — М.: Дрофа, 2004. — 398, [2] с.
    <h5>[2]</h5>Крысин Л. П. Мотор // Толковый словарь иноязычных слов. — М.: Эксмо, 2008. — 944 с. —
    (Библиотека словарей).
    <h5>[3]</h5>Definition of motor | Dictionary.com (англ.). www.dictionary.com.
    Дата обращения: 27 января 2022. Архивировано 27 января 2022 года.
</p>
</body>
</html>
<!-- ########################################################################################## -->

№2

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Lorem Ipsum</title>
    <style>

        h1 {
            color: #5d0080;
        }
        h2 {
            color: #9b05ff;
        }
        h3 {
            color: #ff0505;
        }
        h4 {
            color: #00a0eb;
        }
        h5 {
            color: #00a0eb;
        }
        h6 {
            color: #ffff05;
        }
        p {
            text-align0: justify;
        }
          .left-aligned1 {
    text-align: left;
  }

  .right-aligned2 {
    text-align: right;
  }

  .custom-font {
    font-family: "Courier New", monospace;
  }
  .custom-font2 {
    font-family: "mv boli", monospace;
  }
  .custom-font3 {
    font-family: "segoe print", monospace;
  }
  .custom-font4 {
    font-family: "LUCIDA SANS CONSOLE", monospace;
  }
  .custom-font5 {
    font-family: "GABRIOLA", monospace;
  }
  .custom-font6 {
    font-family: "tempus sans itc", monospace;
  }
  .bold-text {
    font-weight: bold;
  }

  .italic-text {
    font-style: italic;
  }
        .no-underline {
    text-decoration: none;
  }
        .larger-font {
    font-size: 20px;
  }
        .larger-font2 {
    font-size: 40px;
  }
        .larger-font3 {
    font-size: 50px;
  }
  .centered-text {
    text-align: center;
  }

  .centered-container {
    margin: 0 auto;
    width: 50%; /* Измените на желаемую ширину */
  }

  .smaller-text {
    font-size: 95%; /* Измените на желаемый размер шрифта */
  }
   .smaller-text2 {
    font-size: 80%; /* Измените на желаемый размер шрифта */
  }
   .smaller-text3 {
    font-size: 115%; /* Измените на желаемый размер шрифта */
  }
   .smaller-text4 {
    font-size: 15%; /* Измените на желаемый размер шрифта */
  }
        .slightly-bold {
    font-weight: bold; /* или font-weight: 600; */
  }
         .subtle-line {
    border: 2px solid #9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; /* Добавьте отступы по желанию */
  }
  .subtle-line2 {
    border: 1px solid #f2f2f2; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; /* Добавьте отступы по желанию */
  }
         .colored-letter {
    color: #00a0eb; /* Измените на желаемый цвет */
  }
         .custom-color1 {
    color: #001fb8; /* Замените на желаемый цвет */
  }
         .custom-color2 {
    color: #6e6e6e; /* Замените на желаемый цвет */
  }
         .page-container {
    margin: 20px; /* Замените на желаемое значение отступа */
  }
    </style>
</head>
<body>
    <h1 class="centered-text larger-font3 custom-font4">Lorem Ipsum</h1>
<div class="right-aligned2 custom-color1 custom-font3 italic-text smaller-text">
  <p>"Neque porro quisquam est qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit...".</p>
</div>
    <div class="right-aligned2 custom-font2 custom-color2 italic-text smaller-text2 larger-font">
  <p>"Нет никого, кто любил бы боль саму по себе, кто искал бы её и кто хотел бы иметь её
      просто потому, что это боль..".</p>
</div>
    <hr class="subtle-line2">
    <!-- Пример текста Lorem Ipsum -->
    <div class="page-container">
    <p class="larger-font custom-font"><b>
        Классический текст Lorem Ipsum, используемый с XVI века
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>L</b>orem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod tempor incididunt
        ut labore et dolore magna aliqua. Ut enim ad minim veniam, quis nostrud exercitation ullamco laboris
        nisi ut aliquip ex ea commodo consequat. Duis aute irure dolor in reprehenderit in voluptate velit esse
        cillum dolore eu fugiat nulla pariatur. Excepteur sint occaecat cupidatat non proident, sunt in culpa qui
        officia deserunt mollit anim id est laborum."
</p>
    <p class="larger-font custom-font"><b>
        Абзац 1.10.32 "de Finibus Bonorum et Malorum", написанный Цицероном в 45 году н.э.
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>S</b>ed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem
    aperiam, eaque ipsa quae ab illo inventore veritatis et quasi architecto beatae vitae dicta sunt explicabo.
    Nemo enim ipsam voluptatem quia voluptas sit aspernatur aut odit aut fugit, sed quia consequuntur magni dolores
    eos qui ratione voluptatem sequi nesciunt. Neque porro quisquam est, qui dolorem ipsum quia dolor sit amet,
    consectetur, adipisci velit, sed quia non numquam eius modi tempora incidunt ut labore et dolore magnam aliquam
    quaerat voluptatem. Ut enim ad minima veniam, quis nostrum exercitationem ullam corporis suscipit laboriosam, nisi
    ut aliquid ex ea commodi consequatur? Quis autem vel eum iure reprehenderit qui in ea voluptate velit esse quam
    nihil molestiae consequatur, vel illum qui dolorem eum fugiat quo voluptas nulla pariatur?"
</p>
    <p class="larger-font custom-font"><b>
        Английский перевод 1914 года, H. Rackham
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>B</b>ut I must explain to you how all this mistaken idea of denouncing pleasure and praising pain was born and
    I will give you a complete account of the system, and expound the actual teachings of the great explorer of the
    truth, the master-builder of human happiness. No one rejects, dislikes, or avoids pleasure itself, because it is
    pleasure, but because those who do not know how to pursue pleasure rationally encounter consequences that are
    extremely painful. Nor again is there anyone who loves or pursues or desires to obtain pain of itself, because it
    is pain, but because occasionally circumstances occur in which toil and pain can procure him some great pleasure.
    To take a trivial example, which of us ever undertakes laborious physical exercise, except to obtain some advantage
    from it? But who has any right to find fault with a man who chooses to enjoy a pleasure that has no annoying
    consequences, or one who avoids a pain that produces no resultant pleasure?"
</p>
    <p class="larger-font custom-font"><b>
        Абзац 1.10.33 "de Finibus Bonorum et Malorum", написанный Цицероном в 45 году н.э.
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>A</b>t vero eos et accusamus et iusto odio dignissimos ducimus qui blanditiis praesentium voluptatum deleniti
    atque corrupti quos dolores et quas molestias excepturi sint occaecati cupiditate non provident, similique sunt
    in culpa qui officia deserunt mollitia animi, id est laborum et dolorum fuga. Et harum quidem rerum facilis est
    et expedita distinctio. Nam libero tempore, cum soluta nobis est eligendi optio cumque nihil impedit quo minus id
    quod maxime placeat facere possimus, omnis voluptas assumenda est, omnis dolor repellendus. Temporibus autem
    quibusdam et aut officiis debitis aut rerum necessitatibus saepe eveniet ut et voluptates repudiandae sint et
    molestiae non recusandae. Itaque earum rerum hic tenetur a sapiente delectus, ut aut reiciendis voluptatibus
    maiores alias consequatur aut perferendis doloribus asperiores repellat."
</p>
    <p class="larger-font custom-font"><b>
        Английский перевод 1914 года, H. Rackham
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>O</b>n the other hand, we denounce with righteous indignation and dislike men who are so beguiled and demoralized
    by the charms of pleasure of the moment, so blinded by desire, that they cannot foresee the pain and trouble that
    are bound to ensue; and equal blame belongs to those who fail in their duty through weakness of will, which is the
    same as saying through shrinking from toil and pain. These cases are perfectly simple and easy to distinguish.
    In a free hour, when our power of choice is untrammelled and when nothing prevents our being able to do what we
    like best, every pleasure is to be welcomed and every pain avoided. But in certain circumstances and owing to the
    claims of duty or the obligations of business it will frequently occur that pleasures have to be repudiated and
    annoyances accepted. The wise man therefore always holds in these matters to this principle of selection: he
    rejects pleasures to secure other greater pleasures, or else he endures pains to avoid worse pains."
</p>
        </div>
</body>
</html>

<!-- ########################################################################################## -->

№3

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Математические формулы</title>
    <style>
        /* Общие стили для всего документа */
        body {
            font-family: 'VERDANA', monospace; /* Устанавливаем шрифт для всего документа */
            line-height: 1.0; /* Задаем высоту строки текста */
            margin: 70px; /* Задаем отступы для всего документа */
        }

        /* Стили для заголовка первого уровня */
        h1 {
            color: #000000; /* Цвет текста */
            text-align: center; /* Выравнивание текста по центру */
            font-size: 50px; /* Размер шрифта */
            font-weight: bold; /* Жирный шрифт */
        }

        /* Стили для заголовка второго уровня */
        h2 {
            color: #22009e; /* Цвет текста */
            font-size: 35px; /* Размер шрифта */
            margin-top: 20px; /* Верхний отступ */
        }

        /* Стили для подзаголовка */
        h3 {
            color: #00ff49; /* Цвет текста */
            font-size: 25px; /* Размер шрифта */
            font-weight: bold; /* Жирный шрифт */
            font-style: italic; /* Курсив */
            text-align: left; /* Выравнивание текста влево */
            background-color: #a300a3b8; /* Цвет подложки */
            padding: 10px; /* Внутренний отступ */
            border-radius: 45px; /* Закругленные углы подложки */
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.2); /* Тень подложки */
            max-width: 320px; /* Максимальная ширина подложки */
            width: 25%; /* Ширина подложки равна 20% родительского контейнера */
            box-sizing: border-box; /* Учитываем padding и border в общей ширине */
            margin: left; /* Выравнивание по левой стороне */
        }

        /* Стили для обычного текста и вложенных элементов */
        p, span {
            text-align: justify; /* Выравнивание текста по ширине с обеих сторон */
        }

        /* Стили для верхних и нижних индексов */
        sup, sub {
            font-size: 80%; /* Размер шрифта для верхних и нижних индексов */
        }

        /* Отступ для блока с описанием формулы */
        .formula-description {
            margin-left: 20px; /* Отступ слева */
        }

        /* Стили для контейнера формулы и самой формулы */
        .formula-container, .formula {
            text-align: center; /* Выравнивание текста по центру */
            background-color: #ff00008f; /* Цвет подложки */
            padding: 2px; /* Внутренний отступ */
            border-radius: 55px; /* Закругленные углы подложки */
            box-shadow: 0 0 15px rgba(0, 0, 0, 0.2); /* Тень подложки */
            max-width: 670px; /* Максимальная ширина подложки */
            width: 50%; /* Ширина подложки равна 50% родительского контейнера */
            box-sizing: border-box; /* Учитываем padding и border в общей ширине */
            margin: auto; /* Выравнивание по центру */
        }

        /* Стили для самой формулы */
        .formula {
            text-align: center; /* Выравнивание текста по центру */
            color: #0700ff; /* Цвет формулы */
            font-size: 25px; /* Размер шрифта формулы */
            font-style: italic; /* Курсив */
            font-weight: bold; /* Жирный шрифт формулы */
        }

        /* Стили для раздела с примерами */
        .example-section {
            margin-top: 20px; /* Верхний отступ раздела с примерами */
        }

        /* Стили для примеров */
        .example {
            font-size: 20px; /* Размер шрифта примеров */
            color: #555; /* Цвет текста примеров */
        }
         /* Стили для цветной буквы */
        .colored-letter {
            color: #ff0000; /* Цвет буквы */
        }
        /* Стили для увеличенной буквы */
        .enlarged-letter {
            font-size: 24px; /* Размер шрифта для увеличенной буквы */
        }
          /* Стили для выделенных букв */
        .special-letter {
            font-family: 'Arial', sans-serif; /* Используемый шрифт */
            font-style: italic; /* Курсив */
            color: #00adff; /* Цвет текста */
        }
    </style>
</head>
<body>
    <!-- Заголовок первого уровня -->
    <h1>Математические формулы</h1>

    <!-- Формула окружности -->
    <h2>Формула окружности:</h2>
    <!-- Контейнер формулы -->
    <div class="formula-container">
        <!-- Сама формула -->
        <p class="formula"><b> x<sup>2</sup> + y<sup>2</sup> = r<sup>2</sup></b></p>
    </div>
    <!-- Описание формулы -->
    <div class="formula-description">
        <p><b>Эта формула описывает уравнение окружности в декартовой системе координат.</b></p>
        <ul><b>
            <li><span class="colored-letter enlarged-letter special-letter">x</span> и <span class="colored-letter enlarged-letter special-letter">y</span> - координаты точки на плоскости.</li>
            <li><span class="colored-letter enlarged-letter special-letter">r</span> - радиус окружности.</li>
        </ul></b>
    </div>
    <!-- Подробное описание формулы -->
    <h3>Описание формулы:</h3>
    <p class="formula-description"><b>Формула окружности описывает геометрическую фигуру,
        представляющую собой множество точек, равноудаленных от центра.</b></p>
    <p class="formula-description"><b>Здесь x и y - координаты точек,
        а r - радиус окружности.</b></p>

    <!-- Формула площади треугольника -->
    <h2>Формула площади треугольника:</h2>
    <!-- Контейнер формулы -->
    <div class="formula-container">
        <!-- Сама формула -->
        <p class="formula"><b> S = 0.5 * a * b * sin(C)</b></p>
    </div>
    <!-- Описание формулы -->
    <div class="formula-description">
        <p><b>Где:</b></p>
        <ul><b>
            <li><span class="colored-letter enlarged-letter special-letter">S</span> - площадь треугольника.</li>
            <li><span class="colored-letter enlarged-letter special-letter">a</span> и <span class="colored-letter enlarged-letter special-letter">b</span> - длины сторон треугольника.</li>
            <li><span class="colored-letter enlarged-letter special-letter">C</span> - угол между сторонами a и b (в радианах).</li>
        </ul></b>
    </div>
    <!-- Подробное описание формулы -->
    <h3>Описание формулы:</h3>
    <p class="formula-description"><b>Формула площади треугольника выражает площадь как половину произведения длин двух сторон на синус угла между ними.</b></p>

    <!-- Формула объема цилиндра -->
    <h2>Формула объема цилиндра:</h2>
    <!-- Контейнер формулы -->
    <div class="formula-container">
        <!-- Сама формула -->
        <p class="formula"><b> V = π * r<sup>2</sup> * h</b></p>
    </div>
    <!-- Описание формулы -->
    <div class="formula-description">
        <p><b>Где:</b></p>
        <ul><b>
            <li><span class="colored-letter enlarged-letter special-letter">V</span> - объем цилиндра.</li>
            <li><span class="colored-letter enlarged-letter special-letter">π</span> (пи) - математическая константа, приблизительно равная 3.14159.</li>
            <li><span class="colored-letter enlarged-letter special-letter">r</span> - радиус основания цилиндра.</li>
            <li><span class="colored-letter enlarged-letter special-letter">h</span> - высота цилиндра.</li>
        </ul></b>
    </div>
    <!-- Подробное описание формулы -->
    <h3>Описание формулы:</h3>
    <p class="formula-description"><b>Формула объема цилиндра выражает объем как произведение площади основания на высоту.</b></p>

    <!-- Раздел с примерами применения формул -->
    <div class="example-section">
        <h2>Сферы применения формул:</h2>
        <!-- Примеры -->
        <p class="example">1. <b>Геометрия:</b> расчет геометрических параметров фигур.</p>
        <p class="example">2. <b>Инженерия:</b> проектирование и расчеты конструкций.</p>
        <p class="example">3. <b>Физика:</b> моделирование и анализ физических явлений.</p>
    </div>

</body>
</html>

<!-- ########################################################################################## -->

Бонус

1. Пример подключение шрифтов через ссылку

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Your Title</title>

    <!-- Подключение внешних шрифтов -->
    <link rel="stylesheet" href="https://fonts.googleapis.com/css?family=YourFontFamily">

    <!-- Ваши стили -->
    <style>
        body {
            font-family: 'YourFontFamily', sans-serif;
            /* Другие стили... */
        }
    </style>
</head>
<body>
    <!-- Ваш контент -->
</body>
</html>

<!--
Замените "https://fonts.googleapis.com/css?family=YourFontFamily" на фактическую ссылку на шрифт, который вы хотите
 использовать. Например, если вы хотите использовать шрифт "Roboto", ссылка будет
  выглядеть так: "https://fonts.googleapis.com/css?family=Roboto".

Обратите внимание, что это только пример, и вам нужно заменить "YourFontFamily" и "Your Title" на фактические
 значения, используемые в вашем проекте.
 -->

2. Пример подключения через локальные файлы шрифтов

<!--
Поместите шрифтовые файлы в ваш проект:
Сначала загрузите файлы шрифтов (например, .woff, .woff2, .ttf) в ваш проект.
Создайте папку, например, "fonts", и поместите файлы шрифтов в неё.

Пример структуры проекта:
 -->
/your-project
    /fonts
        your-font.woff
        your-font.woff2
    index.html
    style.css
<!--
Добавьте стили в ваш CSS:
В вашем файле стилей (например, style.css), определите новый @font-face и используйте его в стилях элементов.

Пример:
 -->
@font-face {
    font-family: 'YourFontFamily';
    src: url('fonts/your-font.woff2') format('woff2'),
         url('fonts/your-font.woff') format('woff');
    /* Дополнительные параметры, если нужны */
}

body {
    font-family: 'YourFontFamily', sans-serif;
    /* Другие стили... */
}
<!--
Убедитесь, что путь к файлам шрифтов в url соответствует структуре вашего проекта.
 -->
<!--
Подключите стили в HTML:
В вашем файле HTML (например, index.html), убедитесь, что вы подключили ваш CSS файл.

Пример:
 -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Your Title</title>

    <!-- Подключение локальных стилей -->
    <link rel="stylesheet" href="style.css">
</head>
<body>
    <!-- Ваш контент -->
</body>
</html>
<!--
Замените "Your Title" на фактический заголовок вашей страницы.

Это позволит вам использовать локальные файлы шрифтов в вашем проекте. Убедитесь, что пути указаны правильно и что
 браузер может найти и загрузить файлы шрифтов.
 -->
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #

<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->

Домашнее задание

Курс: Разработка интерфейса на JavaScript
Дисциплина: Основы HTML и CSS
Домашнее задание № 3 : Введение в CSS. Форматирование текста при помощи CSS
1. Попробуйте каждый тег

На каждой новой строчке напишите по предложению используя все
пройденные теги.

К каждому тегу добавьте атрибуты class или id

Затем измените с помощью CSS цвет этим предложениям (все должны
быть разного цвета)

Чем разнообразнее будут использоваться CSS свойства, тем выше вы получите
балл. CSS свойства можно также подключать или использовать по разному. Чем
разнообразнее, тем лучше. Удачи!

2. Создайте страницу “Romeo and Juliet”

3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы.

4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии.

<!-- ########################################################################################## -->

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML and CSS Exercise</title>
    <style>
        /* Стили для разнообразных свойств */
        body {
            font-family: 'Arial', sans-serif;
            background-color: #f8f8f8;
            color: #333;
            margin: 20px;
        }

        #paragraph1 {
            font-size: 18px;
            font-weight: bold;
            color: #0066cc;
            text-decoration: underline;
        }

        .paragraph2 {
            font-style: italic;
            color: #cc0000;
            text-align: justify;
            margin-bottom: 20px;
        }

        .highlight {
            background-color: #ffcc00;
            padding: 5px;
        }

        #container {
            border: 2px solid #009900;
            padding: 10px;
            margin: 10px 0;
            background-color: #e6ffe6;
            border-radius: 10px;
        }

        .centered-text {
            text-align: center;
            color: #990099;
        }
    </style>
</head>
<body>
    <!-- Использование различных тегов с атрибутами и стилизацией -->
    <p id="paragraph1" class="highlight">Это первое предложение с различными тегами и атрибутами.</p>
    <p class="paragraph2" id="container">Второе предложение выделено разными стилями с использованием CSS.</p>
    <p class="centered-text">Третье предложение выровнено по центру и имеет свой цвет.</p>

    <!-- Примеры других тегов -->
    <div id="container">
        <h2>Пример заголовка второго уровня</h2>
        <ul>
            <li><strong>Список:</strong> элемент 1</li>
            <li><strong>Список:</strong> элемент 2</li>
        </ul>
        <a href="https://www.example.com" target="_blank" style="color: #cc00cc; text-decoration: none;">Ссылка на примерный сайт</a>
    </div>
</body>
</html>

<!-- ########################################################################################## -->


Домашнее задание

Курс: Разработка интерфейса на JavaScript
Дисциплина: Основы HTML и CSS

Домашнее задание №3: Введение в CSS. Форматирование текста при помощи CSS
<!-- #### -->
1. Попробуйте каждый тег
<!-- #### -->
На каждой новой строчке напишите по предложению используя все
пройденные теги.
К каждому тегу добавьте атрибуты class или id
Затем измените с помощью CSS цвет этим предложениям (все должны
быть разного цвета)
<!-- #### -->
Чем разнообразнее будут использоваться
CSS свойства, тем выше вы получите
балл. CSS свойства можно также подключать или использовать по разному. Чем
разнообразнее, тем лучше. Удачи!
<!-- #### -->
2. Создайте страницу “Romeo and Juliet”
<!-- #### -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы.
<!-- #### -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии.
<!-- #### -->

<!-- ########################################################################################## -->
1. Попробуйте каждый тег
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML and CSS Exercise</title>
    <style>
        /* Стили для разнообразных свойств */
        body {
            font-family: 'Arial', sans-serif;
            background-color: #f8f8f8;
            color: #333;
            margin: 20px;
        }

        #paragraph1 {
            font-size: 18px;
            font-weight: bold;
            color: #0066cc;
            text-decoration: underline;
        }

        .paragraph2 {
            font-style: italic;
            color: #cc0000;
            text-align: justify;
            margin-bottom: 20px;
        }

        .highlight {
            background-color: #ffcc00;
            padding: 5px;
        }

        #container {
            border: 2px solid #009900;
            padding: 10px;
            margin: 10px 0;
            background-color: #e6ffe6;
            border-radius: 10px;
        }

        .centered-text {
            text-align: center;
            color: #990099;
        }
    </style>
</head>
<body>
    <!-- Использование различных тегов с атрибутами и стилизацией -->
    <p id="paragraph1" class="highlight">Это первое предложение с различными тегами и атрибутами.</p>
    <p class="paragraph2" id="container">Второе предложение выделено разными стилями с использованием CSS.</p>
    <p class="centered-text">Третье предложение выровнено по центру и имеет свой цвет.</p>

    <!-- Примеры других тегов -->
    <div id="container">
        <h2>Пример заголовка второго уровня</h2>
        <ul>
            <li><strong>Список:</strong> элемент 1</li>
            <li><strong>Список:</strong> элемент 2</li>
        </ul>
        <a href="https://www.example.com" target="_blank" style="color: #cc00cc; text-decoration: none;">Ссылка на примерный сайт</a>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
2. Создайте страницу “Romario and Juluya”
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Romeo and Juliet</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            text-align: center;
        }

        #title {
            font-family: 'Georgia', serif;
            color: #FF4500; /* Цвет - оранжевый */
            font-style: italic;
            font-size: 36px;
        }

        #author {
            font-size: 18px;
            font-style: italic;
        }

        hr {
            width: 50%;
            margin: auto;
            margin-top: 20px;
            border: 1px solid #000;
        }

        #prologue {
            font-size: 24px;
            margin-top: 20px;
        }

        #left-text, #right-text {
            display: inline-block;
            width: 45%;
            vertical-align: top;
        }

        .act-title {
            font-size: 28px;
            margin-top: 20px;
        }

        .scene-title {
            font-size: 20px;
            font-weight: bold;
            margin-top: 10px;
        }

        .character {
            font-weight: bold;
        }

        .dialogue {
            text-align: left;
            margin-top: 10px;
            margin-bottom: 10px;
        }

        .romeo {
            color: #0000FF; /* Цвет - синий */
        }

        .juliet {
            color: #FF1493; /* Цвет - розовый */
        }

        .sampson {
            color: #008000; /* Цвет - зеленый */
        }

        .gregory {
            color: #FFD700; /* Цвет - золотой */
        }
    </style>
</head>
<body>
    <div id="title">ROMEO AND JULIET</div>
    <div id="author">By William Shakespeare</div>
    <hr>

    <div id="prologue">
        <div id="left-text">
            Two households, both alike in dignity,
            In fair Verona, where we lay our scene,
            From ancient grudge break to new mutiny,
            Where civil blood makes civil hands unclean.
            From forth the fatal loins of these two foes
            A pair of star-cross’d lovers take their life;
            Whole misadventured piteous overthrows
        </div>

        <div id="right-text">
            Do with their death bury their parents’ strife.
            The fearful passage of their death-mark’d love,
            And the continuance of their parents’ rage,
            Which, but their children’s end, nought could remove,
            Is now the two hours’ traffic of our stage;
            The which if you with patient ears attend,
            What here shall miss, our toil shall strive to mend.
        </div>
    </div>
    <hr>

    <div class="act-title">ACT I</div>
    <div class="scene-title">SCENE I. Verona. A public place.</div>
    <div class="scene-title"><i>Enter SAMPSON and GREGORY, of the house
    of CAPULET, armed with swords and bucklers</i></div>

    <div id="left-text" class="dialogue">
        <span class="character sampson">SAMPSON</span> Gregory, o’ my word, we’ll not carry coals.<br>
        <span class="character gregory">GREGORY</span> No, for then we should be colliers.<br>
        <span class="character sampson">SAMPSON</span> I mean, an we be in choler, we’ll draw.<br>
        <span class="character gregory">GREGORY</span> Ay, while you live, draw your neck out
        o’ the collar.<br>
        <span class="character sampson">SAMPSON</span> I strike quickly, being moved.<br>
        <span class="character gregory">GREGORY</span> But thou art not quickly moved to strike.<br>
        <span class="character sampson">SAMPSON</span> A dog of the house of Montague moves me.<br>
        <span class="character gregory">GREGORY</span> To move is to stir; and to be valiant is to
        stand: therefore, if thou art moved, thou runn’st away.<br>
    </div>

    <div id="right-text" class="dialogue">
        <span class="character sampson">SAMPSON</span> A dog of that house shall move me to stand:
        I will take the wall of any man or maid of Montague’s.<br>
        <span class="character gregory">GREGORY</span> That shows thee a weak slave; for the weakest
        goes to the wall.<br>
        <span class="character sampson">SAMPSON</span> True; and therefore women, being the
        weaker vessels, are ever thrust to the wall: therefore I
        will push Montague’s men from the wall, and thrust his
        maids to the wall.<br>
        <span class="character gregory">GREGORY</span> The quarrel is between our masters and us
        their men.<br>
        <span class="character sampson">SAMPSON</span> ’Tis all one, I will show myself a tyrant: when
        I have fought with the men, I will be cruel with the
        maids, and cut off their heads.<br>
        <span class="character gregory">GREGORY</span> The heads of the maids?<br>
        <span class="character sampson">SAMPSON</span> Ay, the heads of the maids, or their
        maidenheads; take it in what sense thou wilt.<br>
        <span class="character gregory">GREGORY</span> They must take it in sense that feel it.<br>
    </div>

</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №0.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 48px;
            margin-right: 15px;
            display: inline-block;
            border: 3px solid #000;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 15px;
            background-color: #fff;
            text-align: center;
        }

        .hearts { color: red; }
        .diamonds { color: #e74c3c; }
        .clubs { color: #2ecc71; }
        .spades { color: #3498db; }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card hearts">♥</div>
        <div class="card diamonds">♦</div>
        <div class="card diamonds">♦</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
        <div class="card hearts">♥</div>
        <!-- Дополнительные карты -->
        <div class="card diamonds">♦</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card hearts">♥</div>
        <div class="card diamonds">♦</div>
        <div class="card hearts">♥</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
        <div class="card hearts">♥</div>
        <!-- Дополнительные карты -->
        <div class="card diamonds">♦</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №1.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 24px;
            margin-right: 10px;
        }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card">&#x1F0A1; Двойка - крести</div>
        <div class="card">&#x1F0C9; Девятка - черви</div>
        <div class="card">&#x1F0C2; Шестерка - черви</div>
        <div class="card">&#x1F0B4; Туз - буби</div>
        <div class="card">&#x1F0A9; Дама - пики</div>
        <div class="card">&#x1F0AC; Валет - крести</div>
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card">&#x1F0C8; Тройка - черви</div>
        <div class="card">&#x1F0A1; Туз - крести</div>
        <div class="card">&#x1F0AA; Десятка - крести</div>
        <div class="card">&#x1F0B7; Король - буби</div>
        <div class="card">&#x1F0A6; Восмёрка - пики</div>
        <div class="card">&#x1F0C7; Тройка - черви</div>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №2.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 36px;
            margin-right: 10px;
            display: inline-block;
            border: 2px solid #000;
            padding: 10px;
            border-radius: 8px;
        }

        /* Цвета мастей */
        .hearts { color: red; }
        .diamonds { color: #e74c3c; }
        .clubs { color: #2ecc71; }
        .spades { color: #3498db; }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card hearts">&#x1F0A1;</div>
        <div class="card diamonds">&#x1F0C9;</div>
        <div class="card diamonds">&#x1F0C2;</div>
        <div class="card clubs">&#x1F0B4;</div>
        <div class="card spades">&#x1F0A9;</div>
        <div class="card hearts">&#x1F0AC;</div>
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card hearts">&#x1F0C8;</div>
        <div class="card diamonds">&#x1F0A1;</div>
        <div class="card hearts">&#x1F0AA;</div>
        <div class="card clubs">&#x1F0B7;</div>
        <div class="card spades">&#x1F0A6;</div>
        <div class="card hearts">&#x1F0C7;</div>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №3.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 190px;
            margin-right: 15px;
            display: inline-block;
            border: 3px solid #000;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 15px;
        }

        .hearts { color: red; }
        .diamonds { color: #e74c3c; }
        .clubs { color: #2ecc71; }
        .spades { color: #3498db; }

        /* Дополнительные стили для улучшения внешнего вида */
        p {
            font-size: 18px;
            margin-bottom: 10px;
        }

        strong {
            font-size: 24px;
        }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card hearts">&#x1F0A1;</div>
        <div class="card diamonds">&#x1F0C9;</div>
        <div class="card diamonds">&#x1F0C2;</div>
        <div class="card clubs">&#x1F0B4;</div>
        <div class="card spades">&#x1F0A9;</div>
        <div class="card hearts">&#x1F0AC;</div>
        <!-- Дополнительные карты -->
<!--        <div class="card diamonds">&#x1F0C4;</div>-->
<!--        <div class="card clubs">&#x1F0B3;</div>-->
<!--        <div class="card spades">&#x1F0AA;</div>-->
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card hearts">&#x1F0C8;</div>
        <div class="card spades">&#x1F0A8;</div>
        <div class="card hearts">&#x1F0AA;</div>
        <div class="card clubs">&#x1F0B7;</div>
        <div class="card spades">&#x1F0A6;</div>
        <div class="card hearts">&#x1F0C7;</div>
        <!-- Дополнительные карты -->
<!--        <div class="card diamonds">&#x1F0C3;</div>-->
<!--        <div class="card clubs">&#x1F0B2;</div>-->
<!--        <div class="card diamonds">&#x1F0A1;</div>-->
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №4.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        /* Общие стили для всего документа */
        body {
            font-family: 'Arial', sans-serif; /* Шрифт */
            line-height: 1.6; /* Межстрочный интервал */
            margin: 20px; /* Отступы от краев */
            background-color: #f8f9fa; /* Цвет фона */
        }

        /* Стили для заголовка страницы */
        h1 {
            color: #336699; /* Цвет текста */
        }

        /* Стили для контейнеров карт игроков */
        .player-cards {
            margin-top: 20px; /* Верхний отступ */
            display: flex; /* Отображение карт в строку */
            flex-wrap: wrap; /* Перенос карт на следующую строку, если не помещаются */
        }

        /* Общие стили для карт */
        .card {
            margin-right: 15px; /* Правый отступ между картами */
            margin-bottom: 15px; /* Нижний отступ между картами */
            padding: 15px; /* Внутренний отступ */
            border: 3px solid #000; /* Обводка карты */
            border-radius: 10px; /* Закругление углов */
            background-color: #fff; /* Цвет фона карты */
            text-align: center; /* Выравнивание текста по центру */
            transition: transform 0.3s ease-in-out; /* Плавное изменение при наведении */
        }

        /* Стили для масти "черви" */
        .hearts { color: red; }

        /* Стили для масти "бубны" */
        .diamonds { color: #e74c3c; }

        /* Стили для масти "трефы" */
        .clubs { color: #2ecc71; }

        /* Стили для масти "пики" */
        .spades { color: #3498db; }
    </style>
</head>
<body>
    <!-- Заголовок страницы -->
    <h1>Игра в карты</h1>

    <!-- Ползунок для изменения размера карт -->
    <label for="cardSize">Размер карт:</label>
    <input type="range" id="cardSize" min="20" max="100" value="36">

    <!-- Контейнер для карт первого игрока -->
    <div class="player-cards" id="player1Cards"></div>

    <!-- Контейнер для карт второго игрока -->
    <div class="player-cards" id="player2Cards"></div>

    <!-- Скрипт на JavaScript для генерации, перемешивания, изменения размера и отображения карт -->
    <script>
        /* Масти и их символы */
        const suits = ['clubs', 'spades', 'diamonds', 'hearts'];
        const suitsSymbols = {
            'hearts': '♥',
            'diamonds': '♦',
            'clubs': '♣',
            'spades': '♠'
        };

        /* Ранги карт */
        const ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

        /* Функция для перемешивания колоды */
        function shuffleDeck() {
            let deck = [];
            for (let suit of suits) {
                for (let rank of ranks) {
                    deck.push({ suit, rank });
                }
            }
            return deck.sort(() => Math.random() - 0.5);
        }

        /* Функция для отображения карт в указанном контейнере с учетом размера */
        function displayCards(playerId, cards, cardSize) {
            const playerCardsElement = document.getElementById(playerId);
            playerCardsElement.innerHTML = '';
            for (let card of cards) {
                const cardElement = document.createElement('div');
                cardElement.style.fontSize = `${cardSize}px`; /* Изменение размера карты */
                /* Добавляем класс масти для стилизации цвета */
                cardElement.className = `card ${card.suit}`;
                /* Отображаем ранг карты и символ масти */
                cardElement.innerText = `${card.rank} ${suitsSymbols[card.suit]}`;
                playerCardsElement.appendChild(cardElement);
            }
        }

        /* Обработчик изменения размера карт при перемещении ползунка */
        document.getElementById('cardSize').addEventListener('input', function() {
            const newSize = this.value;
            displayCards('player1Cards', player1Cards, newSize);
            displayCards('player2Cards', player2Cards, newSize);
        });

        /* Генерируем и перемешиваем колоду */
        const deck = shuffleDeck();

        /* Раздача карт игрокам (первому и второму) */
        const player1Cards = deck.slice(0, 9);
        const player2Cards = deck.slice(9, 18);

        /* Отображение карт для каждого игрока с начальным размером 36px */
        displayCards('player1Cards', player1Cards, 36);
        displayCards('player2Cards', player2Cards, 36);
    </script>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №5.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        /* Общие стили для всего документа */
        body {
            font-family: 'Arial', sans-serif; /* Шрифт */
            line-height: 1.6; /* Межстрочный интервал */
            margin: 20px; /* Отступы от краев */
            background-color: #f8f9fa; /* Цвет фона */
        }

        /* Стили для заголовка страницы */
        h1 {
            color: #336699; /* Цвет текста */
        }

        /* Стили для контейнеров карт игроков */
        .player-cards {
            margin-top: 20px; /* Верхний отступ */
            display: flex; /* Отображение карт в строку */
            flex-wrap: wrap; /* Перенос карт на следующую строку, если не помещаются */
        }

        /* Общие стили для карт */
        .card {
            width: 100px; /* Ширина карты */
            height: 150px; /* Высота карты */
            margin-right: 15px; /* Правый отступ между картами */
            margin-bottom: 15px; /* Нижний отступ между картами */
            padding: 15px; /* Внутренний отступ */
            border: 3px solid #000; /* Обводка карты */
            border-radius: 10px; /* Закругление углов */
            background-color: #fff; /* Цвет фона карты */
            text-align: center; /* Выравнивание текста по центру */
            position: relative; /* Позиционирование элементов внутри карты */
        }

        /* Стили для масти "черви" */
        .hearts { color: red; }

        /* Стили для масти "бубны" */
        .diamonds { color: #e74c3c; }

        /* Стили для масти "трефы" */
        .clubs { color: #2ecc71; }

        /* Стили для масти "пики" */
        .spades { color: #3498db; }

        /* Стили для значения карты */
        .card-value {
            font-size: 14px; /* Размер шрифта значения карты */
            color: #555; /* Цвет текста значения карты */
            position: absolute; /* Абсолютное позиционирование значения */
        }

        /* Стили для левого верхнего угла карты */
        .card-value.top-left {
            top: 10px; /* Отступ от верхнего края карты */
            left: 10px; /* Отступ от левого края карты */
        }

        /* Стили для правого нижнего угла карты */
        .card-value.bottom-right {
            bottom: 10px; /* Отступ от нижнего края карты */
            right: 10px; /* Отступ от правого края карты */
        }
    </style>
</head>
<body>
    <!-- Заголовок страницы -->
    <h1>Игра в карты</h1>

    <!-- Ползунок для изменения размера карт -->
    <label for="cardSize">Размер карт:</label>
    <input type="range" id="cardSize" min="20" max="100" value="50">

    <!-- Выпадающий список для выбора масти -->
    <label for="suitSelect">Масть карты:</label>
    <select id="suitSelect">
        <option value="all">Все масти</option>
        <option value="hearts">Черви</option>
        <option value="diamonds">Бубны</option>
        <option value="clubs">Трефы</option>
        <option value="spades">Пики</option>
    </select>

    <!-- Выпадающий список для выбора ранга карты -->
    <label for="rankSelect">Ранг карты:</label>
    <select id="rankSelect">
        <option value="all">Все ранги</option>
        <option value="A">Туз</option>
        <option value="2">2</option>
        <option value="3">3</option>
        <option value="4">4</option>
        <option value="5">5</option>
        <option value="6">6</option>
        <option value="7">7</option>
        <option value="8">8</option>
        <option value="9">9</option>
        <option value="10">10</option>
        <option value="J">Валет</option>
        <option value="Q">Дама</option>
        <option value="K">Король</option>
    </select>

    <!-- Контейнер для карт первого игрока -->
    <div class="player-cards" id="player1Cards"></div>

    <!-- Контейнер для карт второго игрока -->
    <div class="player-cards" id="player2Cards"></div>

    <!-- Скрипт на JavaScript для генерации, перемешивания, изменения размера и отображения карт -->
    <script>
        /* Масти и их символы */
        const suits = ['clubs', 'spades', 'diamonds', 'hearts'];
        const suitsSymbols = {
            'hearts': '♥',
            'diamonds': '♦',
            'clubs': '♣',
            'spades': '♠'
        };

        /* Ранги карт */
        const ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

        /* Функция для перемешивания колоды */
        function shuffleDeck() {
            let deck = [];
            for (let suit of suits) {
                for (let rank of ranks) {
                    deck.push({ suit, rank });
                }
            }
            return deck.sort(() => Math.random() - 0.5);
        }

        /* Функция для отображения карт в указанном контейнере с учетом размера, масти и ранга */
        function displayCards(playerId, cards, cardSize, selectedSuit, selectedRank) {
            const playerCardsElement = document.getElementById(playerId);
            playerCardsElement.innerHTML = '';
            for (let card of cards) {
                if ((selectedSuit === 'all' || card.suit === selectedSuit) && (selectedRank === 'all' || card.rank === selectedRank)) {
                    const cardElement = document.createElement('div');
                    cardElement.style.width = `${cardSize}px`; /* Изменение ширины карты */
                    cardElement.style.height = `${cardSize * 1.5}px`; /* Изменение высоты карты */
                    /* Добавляем класс масти для стилизации цвета */
                    cardElement.className = `card ${card.suit}`;
                    /* Добавляем значение карты в верхний и нижний угол карты */
                    cardElement.innerHTML = `
                        <div class="card-value top-left">${card.rank}</div>
                        ${suitsSymbols[card.suit]}
                        <div class="card-value bottom-right">${card.rank}</div>`;
                    playerCardsElement.appendChild(cardElement);
                }
            }
        }

        /* Обработчик изменения размера карт при перемещении ползунка */
        document.getElementById('cardSize').addEventListener('input', function() {
            const newSize = this.value;
            const selectedSuit = document.getElementById('suitSelect').value;
            const selectedRank = document.getElementById('rankSelect').value;
            displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
            displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
        });

        /* Обработчик изменения масти карты */
        document.getElementById('suitSelect').addEventListener('change', function() {
            const newSize = document.getElementById('cardSize').value;
            const selectedSuit = this.value;
            const selectedRank = document.getElementById('rankSelect').value;
            displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
            displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
        });

        /* Обработчик изменения ранга карты */
        document.getElementById('rankSelect').addEventListener('change', function() {
            const newSize = document.getElementById('cardSize').value;
            const selectedSuit = document.getElementById('suitSelect').value;
            const selectedRank = this.value;
            displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
            displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
        });

        /* Генерируем и перемешиваем колоду */
        const deck = shuffleDeck();

        /* Раздача карт игрокам (первому и второму) */
        const player1Cards = deck.slice(0, 9);
        const player2Cards = deck.slice(9, 18);

        /* Отображение карт для каждого игрока с начальным размером 50px, мастью "все" и рангом "все" */
        displayCards('player1Cards', player1Cards, 50, 'all', 'all');
        displayCards('player2Cards', player2Cards, 50, 'all', 'all');
    </script>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №0.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Пример стилей для страницы (можно дополнительно настроить) */
        body {
            font-family: Arial, sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
        }

        img {
            max-width: 100%;
            height: auto;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }
    </style>
</head>

<body>

    <header>
        <h1>Welcome to the HTML Tutorial</h1>
    </header>

    <section>
        <article>
            <p>This HTML tutorial covers various HTML elements and their usage. Learn how to create a basic HTML
                structure and enhance your web development skills.</p>
        </article>

        <article>
            <h2>Basic HTML Structure</h2>
            <p>HTML documents consist of the following elements:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code> declaration</li>
                <li><code>&lt;html&gt;</code> element</li>
                <li><code>&lt;head&gt;</code> element</li>
                <li><code>&lt;title&gt;</code> element</li>
                <li><code>&lt;body&gt;</code> element</li>
            </ul>
        </article>

        <article>
            <h2>Text Formatting</h2>
            <p>Use tags like <code>&lt;p&gt;</code>, <code>&lt;h1&gt;</code>, <code>&lt;strong&gt;</code>, <code>&lt;em&gt;</code>,
                etc., for text formatting.</p>
        </article>

        <article>
            <h2>Images</h2>
            <p>Include images using the <code>&lt;img&gt;</code> tag.</p>
            <img src="https://placekitten.com/400/200" alt="Cute Kitten">
        </article>

        <article>
            <h2>Links</h2>
            <p>Create hyperlinks with the <code>&lt;a&gt;</code> tag.</p>
            <p>Visit our <a href="https://example.com" target="_blank">example website</a>.</p>
        </article>
    </section>

    <footer>
        <p>Explore more HTML elements and features to enhance your web development skills.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №1.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="ru">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Стили для страницы */
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
            color: #555;
        }

        img {
            max-width: 100%;
            height: auto;
            margin-bottom: 20px;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
            color: #333;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }

        code {
            background-color: #f8f9fa;
            padding: 2px 5px;
            border: 1px solid #ccc;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            color: #333;
        }
    </style>
</head>

<body>

    <header>
        <h1>Добро пожаловать в HTML-туториал</h1>
    </header>

    <section>
        <article>
            <p>HTML (HyperText Markup Language) — это стандартный язык разметки для создания веб-страниц. Он
                используется для структурирования контента в интернете, предоставляя основу для отображения текста,
                изображений, ссылок, форм и мультимедийных элементов.</p>
        </article>

        <article>
            <h2>Основная структура HTML</h2>
            <p>HTML-документ состоит из следующих элементов:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code>: Объявление версии HTML.</li>
                <li><code>&lt;html&gt;</code>: Корневой элемент HTML-страницы.</li>
                <li><code>&lt;head&gt;</code>: Содержит мета-информацию о документе.</li>
                <li><code>&lt;title&gt;</code>: Устанавливает заголовок HTML-документа (отображается во вкладке
                    браузера).</li>
                <li><code>&lt;body&gt;</code>: Содержит содержимое HTML-документа.</li>
            </ul>
            <p>Вот пример простой структуры HTML:</p>
            <code>
                &lt;!DOCTYPE html&gt;<br>
                &lt;html&gt;<br>
                &nbsp;&nbsp;&lt;head&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;title&gt;Моя первая HTML-страница&lt;/title&gt;<br>
                &nbsp;&nbsp;&lt;/head&gt;<br>
                &nbsp;&nbsp;&lt;body&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;h1&gt;Привет, мир!&lt;/h1&gt;<br>
                &nbsp;&nbsp;&lt;/body&gt;<br>
                &lt;/html&gt;
            </code>
        </article>

        <article>
            <h2>Форматирование текста</h2>
            <p>HTML предоставляет различные теги для форматирования текста:</p>
            <ul>
                <li><code>&lt;p&gt;</code>: Параграф</li>
                <li><code>&lt;h1&gt;</code> до <code>&lt;h6&gt;</code>: Заголовки</li>
                <li><code>&lt;strong&gt;</code>: Жирный текст</li>
                <li><code>&lt;em&gt;</code>: Курсив</li>
            </ul>
            <p>Пример:</p>
            <code>
                &lt;p&gt;Это &lt;strong&gt;параграф&lt;/strong&gt; с жирным текстом.&lt;/p&gt;<br>
                &lt;h2&gt;Подзаголовок&lt;/h2&gt;<br>
                &lt;p&gt;Это &lt;em&gt;курсивизированное&lt;/em&gt; предложение.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Изображения</h2>
            <p>Вставка изображений с использованием тега <code>&lt;img&gt;</code>:</p>
            <code>
                &lt;img src="image.jpg" alt="Описание изображения"&gt;
            </code>
        </article>

        <article>
            <h2>Ссылки</h2>
            <p>Создание гиперссылок с тегом <code>&lt;a&gt;</code>:</p>
            <code>
                &lt;a href="https://example.com" target="_blank"&gt;Посетите веб-сайт примера&lt;/a&gt;
            </code>
        </article>

        <article>
            <h2>Цветной текст</h2>
            <p>Применение цветового стиля к тексту:</p>
            <code>
                &lt;p style="color: #e44d26;"&gt;Этот текст оранжевого цвета.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Списки</h2>
            <p>Использование тегов <code>&lt;ul&gt;</code> и <code>&lt;li&gt;</code> для создания списков:</p>
            <code>
                &lt;ul&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Первый элемент списка&lt;/li&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Второй элемент списка&lt;/li&gt;<br>
                &lt;/ul&gt;
            </code>
        </article>

        <article>
            <h2>Видео</h2>
            <p>Вставка видео с помощью тега <code>&lt;iframe&gt;</code>:</p>
            <code>
                &lt;iframe width="560" height="315" src="https://www.youtube.com/embed/ВАШ_KОД" frameborder="0"
                allowfullscreen&gt;&lt;/iframe&gt;
            </code>
        </article>
    </section>

    <footer>
        <p>Исследуйте больше элементов и возможностей HTML для улучшения своих навыков веб-разработки.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №1_1.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="ru">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Стили для страницы */
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
            color: #555;
        }

        img {
            max-width: 100%;
            height: auto;
            margin-bottom: 20px;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
            color: #333;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }

        code {
            background-color: #f8f9fa;
            padding: 2px 5px;
            border: 1px solid #ccc;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            color: #333;
        }

        .styled-text {
            font-size: 18px;
            font-weight: bold;
            color: #e44d26;
            font-style: italic;
            text-decoration: underline;
        }
    </style>
</head>

<body>

    <header>
        <h1>Добро пожаловать в HTML-туториал</h1>
    </header>

    <section>
        <article>
            <p>HTML (HyperText Markup Language) — это стандартный язык разметки для создания веб-страниц.</p> <p>Он
                используется для структурирования контента в интернете, предоставляя основу для отображения текста,
                изображений, ссылок, форм и мультимедийных элементов.</p>
        </article>

        <article>
            <h2>Основная структура HTML</h2>
            <p>HTML-документ состоит из следующих элементов:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code>: Объявление версии HTML.</li>
                <li><code>&lt;html&gt;</code>: Корневой элемент HTML-страницы.</li>
                <li><code>&lt;head&gt;</code>: Содержит мета-информацию о документе.</li>
                <li><code>&lt;title&gt;</code>: Устанавливает заголовок HTML-документа (отображается во вкладке
                    браузера).</li>
                <li><code>&lt;body&gt;</code>: Содержит содержимое HTML-документа.</li>
            </ul>
            <p>Вот пример простой структуры HTML:</p>
            <code>
                &lt;!DOCTYPE html&gt;<br>
                &lt;html&gt;<br>
                &nbsp;&nbsp;&lt;head&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;title&gt;Моя первая HTML-страница&lt;/title&gt;<br>
                &nbsp;&nbsp;&lt;/head&gt;<br>
                &nbsp;&nbsp;&lt;body&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;h1&gt;Привет, мир!&lt;/h1&gt;<br>
                &nbsp;&nbsp;&lt;/body&gt;<br>
                &lt;/html&gt;
            </code>
        </article>

        <article>
            <h2>Форматирование текста</h2>
            <p>HTML предоставляет различные теги для форматирования текста:</p>
            <ul>
                <li><code>&lt;p&gt;</code>: Параграф</li>
                <li><code>&lt;h1&gt;</code> до <code>&lt;h6&gt;</code>: Заголовки</li>
                <li><code>&lt;strong&gt;</code>: Жирный текст</li>
                <li><code>&lt;em&gt;</code>: Курсив</li>
            </ul>
            <p>Пример:</p>
            <code>
                &lt;p&gt;Это &lt;strong&gt;параграф&lt;/strong&gt; с жирным текстом.&lt;/p&gt;<br>
                &lt;h2&gt;Подзаголовок&lt;/h2&gt;<br>
                &lt;p&gt;Это &lt;em&gt;курсивизированное&lt;/em&gt; предложение.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Изображения</h2>
            <p>Вставка изображений с использованием тега <code>&lt;img&gt;</code>:</p>
            <img src="https://placekitten.com/800/400" alt="Котенок">
        </article>

        <article>
            <h2>Ссылки</h2>
            <p>Создание гиперссылок с тегом <code>&lt;a&gt;</code>:</p>
            <code>
                &lt;a href="https://example.com" target="_blank"&gt;Посетите веб-сайт примера&lt;/a&gt;
            </code>
        </article>

        <article>
            <h2>Цветной текст</h2>
            <p>Применение цветового стиля к тексту:</p>
            <code>
                &lt;p style="color: #e44d26;"&gt;Этот текст оранжевого цвета.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Стилизованный текст</h2>
            <p>Пример изменения шрифта, стиля, размера и цвета:</p>
            <span class="styled-text">Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого цвета и
                размера 18px.</span>
        </article>

        <article>
            <h2>Списки</h2>
            <p>Использование тегов <code>&lt;ul&gt;</code> и <code>&lt;li&gt;</code> для создания списков:</p>
            <code>
                &lt;ul&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Первый элемент списка&lt;/li&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Второй элемент списка&lt;/li&gt;<br>
                &lt;/ul&gt;
            </code>
        </article>

        <article>
            <h2>Видео</h2>
            <p>Вставка видео с помощью тега <code>&lt;iframe&gt;</code>:</p>
            <iframe width="560" height="315" src="https://www.youtube.com/embed/dQw4w9WgXcQ" frameborder="0"
                allowfullscreen></iframe>
        </article>
    </section>

    <footer>
        <p>Исследуйте больше элементов и возможностей HTML для улучшения своих навыков веб-разработки.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №2.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="ru">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Стили для страницы */
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
            color: #555;
        }

        img {
            max-width: 100%;
            height: auto;
            margin-bottom: 20px;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
            color: #333;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }

        code {
            background-color: #f8f9fa;
            padding: 2px 5px;
            border: 1px solid #ccc;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            color: #333;
        }

        .styled-text {
            font-size: 18px;
            font-weight: bold;
            color: #e44d26;
            font-style: italic;
            text-decoration: underline;
        }

        hr {
            margin: 20px 0;
            border: none;
            border-top: 2px solid #007bff;
        }
    </style>
</head>

<body>

    <header>
        <h1>Добро пожаловать в HTML-туториал</h1>
    </header>

    <section>
        <article>
            <p><b>HTML</b> (Hypertext Markup Language) - это код, который используется для структурирования и отображения веб-страницы и её контента.</p>
            <p>Например, контент может быть структурирован внутри множества параграфов, маркированных списков или с использованием изображений и таблиц данных.</p>
            <p>Как видно из названия, этот туториал постарается дать вам базовое понимание HTML и его функций.</p>
        </article>
        <hr>
        <p><b>HTML</b> не является языком программирования; это язык разметки, и используется, чтобы сообщать вашему браузеру, как отображать веб-страницы, которые вы посещаете.</p>
        <p>Он может быть сложным или простым, в зависимости от того, как хочет веб-дизайнер.</p>
        <p>HTML состоит из ряда элементов, которые вы используете, чтобы вкладывать или оборачивать различные части контента, чтобы заставить контент отображаться или действовать определённым образом.</p>
        <p>Ограждающие теги могут сделать слово или изображение ссылкой на что-то ещё, могут сделать слова курсивом, сделать шрифт больше или меньше и так далее.</p>
        <hr>

        <article>
            <h2>Основная структура HTML</h2>
            <p>HTML-документ состоит из следующих элементов:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code>: Объявление версии HTML.</li>
                <li><code>&lt;html&gt;</code>: Корневой элемент HTML-страницы.</li>
                <li><code>&lt;head&gt;</code>: Содержит мета-информацию о документе.</li>
                <li><code>&lt;title&gt;</code>: Устанавливает заголовок HTML-документа (отображается во вкладке
                    браузера).</li>
                <li><code>&lt;body&gt;</code>: Содержит содержимое HTML-документа.</li>
            </ul>
            <p>Вот пример простой структуры HTML:</p>
            <code>
                &lt;!DOCTYPE html&gt;<br>
                &lt;html&gt;<br>
                &nbsp;&nbsp;&lt;head&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;title&gt;Моя первая HTML-страница&lt;/title&gt;<br>
                &nbsp;&nbsp;&lt;/head&gt;<br>
                &nbsp;&nbsp;&lt;body&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;h1&gt;Привет, мир!&lt;/h1&gt;<br>
                &nbsp;&nbsp;&lt;/body&gt;<br>
                &lt;/html&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Форматирование текста</h2>
            <p>HTML предоставляет различные теги для форматирования текста:</p>
            <ul>
                <li><code>&lt;p&gt;</code>: Параграф</li>
                <li><code>&lt;h1&gt;</code> до <code>&lt;h6&gt;</code>: Заголовки</li>
                <li><code>&lt;strong&gt;</code>: Жирный текст</li>
                <li><code>&lt;em&gt;</code>: Курсив</li>
            </ul>
            <p>Пример использования тегов форматирования текста:</p>
            <code>
                &lt;p&gt;Это &lt;strong&gt;параграф&lt;/strong&gt; с жирным текстом.&lt;/p&gt;<br>
                &lt;h2&gt;Подзаголовок&lt;/h2&gt;<br>
                &lt;p&gt;Это &lt;em&gt;курсивизированное&lt;/em&gt; предложение.&lt;/p&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Изображения</h2>
            <p>Вставка изображений с использованием тега <code>&lt;img&gt;</code>:</p>
            <img src="https://placekitten.com/800/400" alt="Котенок">
        </article>

        <hr>

        <article>
            <h2>Ссылки</h2>
            <p>Создание гиперссылок с тегом <code>&lt;a&gt;</code>:</p>
            <code>
                &lt;a href="https://example.com" target="_blank"&gt;Посетите веб-сайт примера&lt;/a&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Цветной текст</h2>
            <p>Применение цветового стиля к тексту:</p>
            <code>
                &lt;p style="color: #e44d26;"&gt;Этот текст оранжевого цвета.&lt;/p&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Стилизованный текст</h2>
            <p>Пример изменения шрифта, стиля, размера и цвета:</p>
            <span class="styled-text">Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого цвета и
                размера 18px.</span>
        </article>

        <hr>

        <article>
            <h2>Списки</h2>
            <p>Использование тегов <code>&lt;ul&gt;</code> и <code>&lt;li&gt;</code> для создания списков:</p>
            <code>
                &lt;ul&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Первый элемент списка&lt;/li&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Второй элемент списка&lt;/li&gt;<br>
                &lt;/ul&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Видео</h2>
            <p>Вставка видео с помощью тега <code>&lt;iframe&gt;</code>:</p>
            <iframe width="560" height="315" src="https://www.youtube.com/embed/dQw4w9WgXcQ" frameborder="0"
                allowfullscreen></iframe>
        </article>
    </section>

    <footer>
        <p>Исследуйте больше элементов и возможностей HTML для улучшения своих навыков веб-разработки.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->

'''
Дата выполнения Практической-Работы: 19-20 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №26: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите одиночный json – объект с сайта jsonplaceholder, используя библиотеку requests.
б) Сохраните его в файл.

'''
'''
Урок от 24.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: (а и б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №1.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# Импорт библиотек
import requests
import json


# Определение функции для загрузки и сохранения JSON
def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        # Вывод сообщения об успешной загрузке и сохранении
        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        # Вывод сообщения об ошибке при загрузке или сохранении JSON
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


# Указание URL и имени файла, куда сохранить JSON
json_url = "https://jsonplaceholder.typicode.com/posts/1"
output_filename = "output.json"

# Вызов функции для загрузки и сохранения JSON
download_and_save_json(json_url, output_filename)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import requests
import json

'''
Описание:

requests: Эта библиотека используется для выполнения HTTP-запросов.
json: Эта библиотека предоставляет методы для работы с данными в формате JSON.
'''
'''
Шаг 2: Определение функции download_and_save_json
'''


def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


'''
Описание:

download_and_save_json: Это пользовательская функция, которая принимает два аргумента - url и filename.
requests.get(url): Этот метод выполняет GET-запрос к указанному url и возвращает объект Response.
response.raise_for_status(): Проверяет, были ли ошибки в результате запроса. Если есть, вызывает исключение HTTPError.
open(filename, 'w'): Открывает файл с указанным именем для записи.
json.dump(response.json(), file, indent=2): Записывает JSON-объект из ответа в файл с отступом в 2 пробела.
print(f"JSON-объект успешно загружен и сохранен в файл {filename}"): 
Выводит сообщение об успешной загрузке и сохранении.
'''
'''
Шаг 3: Указание URL и имени файла и вызов функции
'''
# Укажите URL и имя файла, куда сохранить JSON
json_url = "https://jsonplaceholder.typicode.com/posts/1"
output_filename = "output.json"

# Вызов функции для загрузки и сохранения JSON
download_and_save_json(json_url, output_filename)
'''
Описание:

Здесь вы устанавливаете URL и имя файла для JSON-данных.
Затем вызывается функция download_and_save_json с указанными аргументами.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №2.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# Импорт библиотек
import requests
import json


# Определение функции для загрузки и сохранения JSON
def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        # Вывод сообщения об успешной загрузке и сохранении
        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        # Вывод сообщения об ошибке при загрузке или сохранении JSON
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


# Указание URL и имени файла, куда сохранить JSON для поста
post_url = "https://jsonplaceholder.typicode.com/posts/1"
post_filename = "post_output.json"

# Вызов функции для загрузки и сохранения JSON для поста
download_and_save_json(post_url, post_filename)

# Указание URL и имени файла, куда сохранить JSON для комментариев
comments_url = "https://jsonplaceholder.typicode.com/comments?postId=1"
comments_filename = "comments_output.json"

# Вызов функции для загрузки и сохранения JSON для комментариев
download_and_save_json(comments_url, comments_filename)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import requests
import json

'''
Описание:

requests: Эта библиотека используется для выполнения HTTP-запросов.
json: Эта библиотека предоставляет методы для работы с данными в формате JSON.
'''
'''
Шаг 2: Определение функции download_and_save_json
'''


def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


'''
Описание:

download_and_save_json: Это пользовательская функция, которая принимает два аргумента - url и filename.
requests.get(url): Этот метод выполняет GET-запрос к указанному url и возвращает объект Response.
response.raise_for_status(): Проверяет, были ли ошибки в результате запроса. Если есть, вызывает исключение HTTPError.
open(filename, 'w'): Открывает файл с указанным именем для записи.
json.dump(response.json(), file, indent=2): Записывает JSON-объект из ответа в файл с отступом в 2 пробела.
print(f"JSON-объект успешно загружен и сохранен в файл {filename}"): 
Выводит сообщение об успешной загрузке и сохранении.
'''
'''
Шаг 3: Указание URL и имени файла, вызов функции для поста
'''
# Укажите URL и имя файла, куда сохранить JSON для поста
post_url = "https://jsonplaceholder.typicode.com/posts/1"
post_filename = "post_output.json"

# Вызов функции для загрузки и сохранения JSON для поста
download_and_save_json(post_url, post_filename)
'''
Описание:

Здесь вы устанавливаете URL и имя файла для JSON-данных поста.
Затем вызывается функция download_and_save_json с указанными аргументами.
'''
'''
Шаг 4: Указание URL и имени файла, вызов функции для комментариев
'''
# Укажите URL и имя файла, куда сохранить JSON для комментариев
comments_url = "https://jsonplaceholder.typicode.com/comments?postId=1"
comments_filename = "comments_output.json"

# Вызов функции для загрузки и сохранения JSON для комментариев
download_and_save_json(comments_url, comments_filename)
'''
Описание:

Здесь вы устанавливаете URL и имя файла для JSON-данных комментариев.
Затем снова вызывается функция download_and_save_json с указанными аргументами.
'''
'''
В итоге, код выполняет два HTTP-запроса, загружает JSON-объекты и сохраняет их в соответствующие файлы. 
Вся логика организована внутри функции download_and_save_json, 
что делает код более структурированным и переиспользуемым.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вы можете загружать JSON-объекты с различных открытых API, 
которые предоставляют данные в формате JSON. Например, некоторые 
популярные открытые API, которые предоставляют JSON-данные:

1. JSONPlaceholder: 
Это фейковый онлайн-сервис для тестирования и прототипирования, который предоставляет тестовые данные в формате JSON.
Вы уже использовали его в предыдущем примере.

    URL: https://jsonplaceholder.typicode.com

2. OpenWeatherMap API: Предоставляет информацию о погоде в формате JSON.

    URL: https://openweathermap.org/api

3. GitHub API: Предоставляет данные о репозиториях, пользователях и других аспектах GitHub.

    URL: https://developer.github.com/v3

4. REST Countries API: Предоставляет информацию о странах в формате JSON.

    URL: https://restcountries.com

Просто выберите API, которое соответствует вашим интересам, и используйте его URL для загрузки JSON-данных. 
Не забывайте, что при использовании сторонних API следует ознакомиться с их документацией и условиями использования.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


'''
Дата выполнения Домашней-Работы: 24-25 - ЯНВАРЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №26: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите массив json – объектов с сайта jsonplaceholder, используя библиотеку requests.
б) Сохраните циклом каждый в отдельный файл, в одну новую папку.

'''
'''
Урок от 24.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import requests
import json


def download_and_save_json_objects(url, output_folder):
    try:
        # Загрузка массива JSON-объектов с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Разбор JSON-объектов
        json_objects = response.json()

        # Создание новой папки, если она не существует
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Цикл по каждому JSON-объекту и сохранение в отдельный файл
        for i, json_object in enumerate(json_objects):
            filename = os.path.join(output_folder, f"object_{i + 1}.json")
            with open(filename, 'w') as file:
                json.dump(json_object, file, indent=2)

            print(f"JSON-объект {i + 1} успешно сохранен в файл {filename}")

    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


# Укажите URL и имя новой папки для сохранения JSON-объектов
json_url = "https://jsonplaceholder.typicode.com/posts"
output_folder = "json_objects_folder"

# Вызов функции для загрузки и сохранения JSON-объектов
download_and_save_json_objects(json_url, output_folder)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import os
import requests
import json

'''
Описание:

os: Эта библиотека предоставляет функционал для работы с операционной системой, в данном случае, для создания папок.
requests: Эта библиотека используется для выполнения HTTP-запросов.
json: Эта библиотека предоставляет методы для работы с данными в формате JSON.
'''
'''
Шаг 2: Определение функции download_and_save_json_objects
'''


def download_and_save_json_objects(url, output_folder):
    try:
        # Загрузка массива JSON-объектов с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Разбор JSON-объектов
        json_objects = response.json()

        # Создание новой папки, если она не существует
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Цикл по каждому JSON-объекту и сохранение в отдельный файл
        for i, json_object in enumerate(json_objects):
            filename = os.path.join(output_folder, f"object_{i + 1}.json")
            with open(filename, 'w') as file:
                json.dump(json_object, file, indent=2)

            print(f"JSON-объект {i + 1} успешно сохранен в файл {filename}")

    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


'''
Описание:

download_and_save_json_objects: Это пользовательская функция, которая принимает два аргумента - url и output_folder.
requests.get(url): Этот метод выполняет GET-запрос к указанному url и возвращает объект Response.
response.raise_for_status(): Проверяет, были ли ошибки в результате запроса. Если есть, вызывает исключение HTTPError.
json_objects = response.json(): Разбирает JSON-объекты из ответа.
os.makedirs(output_folder): Создает новую папку, если она не существует.
enumerate(json_objects): Возвращает индекс и текущий элемент из списка JSON-объектов.
os.path.join(output_folder, f"object_{i + 1}.json"): Формирует полный путь к файлу с учетом созданной папки и 
индекса объекта.
json.dump(json_object, file, indent=2): Записывает JSON-объект в файл с отступом в 2 пробела.
print(f"JSON-объект {i + 1} успешно сохранен в файл {filename}"): Выводит сообщение об успешном сохранении.
'''
'''
Шаг 3: Указание URL и имени новой папки, вызов функции
'''
# Укажите URL и имя новой папки для сохранения JSON-объектов
json_url = "https://jsonplaceholder.typicode.com/posts"
output_folder = "json_objects_folder"

# Вызов функции для загрузки и сохранения JSON-объектов
download_and_save_json_objects(json_url, output_folder)
'''
Описание:

Здесь мы устанавливаем URL и имя новой папки для JSON-данных.
Затем вызывается функция download_and_save_json_objects с указанными аргументами.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Да, так-же мы можем использовать любой из предложенных сервисов для загрузки массива JSON-объектов.
Все в точности как и в практической работе №26.

Примеры URL для загрузки данных:

1. JSONPlaceholder:

    URL для постов: https://jsonplaceholder.typicode.com/posts
    URL для комментариев: https://jsonplaceholder.typicode.com/comments
    Используем любой из этих URL в зависимости от ваших потребностей.

2. OpenWeatherMap API:

    Пример URL для получения данных о погоде 
    в Лондоне: https://api.openweathermap.org/data/2.5/weather?q=London&appid=YOUR_API_KEY
    Необходимо заменить YOUR_API_KEY на необходимый ключ API, полученный после регистрации на сайте OpenWeatherMap.

3. GitHub API:

    Документация GitHub API: https://developer.github.com/v3
    Мы можем использовать различные эндпойнты API в зависимости от того, какие данные нужны.

4. REST Countries API:

    Пример URL для получения информации о странах: https://restcountries.com/v3.1/all
    Этот URL предоставит информацию о всех странах. 
    Мы также можем использовать другие эндпойнты для более конкретных запросов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Дата выполнения Практической-Работы: 26-27 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №27: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите одиночный json – объект с сайта jsonplaceholder, используя библиотеку aiohttp.
б) Сохраните его в файл.

'''
'''
Урок от 26.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: (а и б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import json

async def download_and_save_json(url, filename):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                json_data = await response.json()

                with open(filename, 'w') as file:
                    json.dump(json_data, file, indent=2)

                print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")

async def main():
    # Укажите URL и имя файла, куда сохранить JSON
    json_url = "https://jsonplaceholder.typicode.com/users/1"
    output_filename = "output_async.json"

    # Вызов функции для загрузки и сохранения JSON
    await download_and_save_json(json_url, output_filename)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек:
'''
import aiohttp
import asyncio
import json
''''
Описание:

Импортируются необходимые библиотеки: aiohttp для асинхронных HTTP-запросов, asyncio для асинхронного 
программирования, json для работы с JSON-данными.
''''
'''
Шаг 2: Определение асинхронной функции download_and_save_json:
'''
async def download_and_save_json(url, filename):
''''
Описание:

mЭта функция асинхронная и предназначена для загрузки JSON-данных по заданному URL и сохранения их в файл.
''''
'''
Шаг 3: Блок try-except внутри функции:
'''
try:
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            response.raise_for_status()
            json_data = await response.json()

            with open(filename, 'w') as file:
                json.dump(json_data, file, indent=2)

            print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
except aiohttp.ClientError as e:
    print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")
'''
Описание:

Создается асинхронная сессия с помощью aiohttp.ClientSession() для выполнения HTTP-запросов.
Выполняется асинхронный GET-запрос с использованием session.get(url).
Проверяется наличие ошибок при запросе с response.raise_for_status().
Получается JSON-данные из ответа с использованием await response.json().
Сохраняются JSON-данные в файл с использованием json.dump.
Обрабатываются и выводятся ошибки с использованием aiohttp.ClientError.
'''
'''
Шаг 4: Определение асинхронной функции main:
'''
async def main():
''''
Описание:

Эта функция предназначена для запуска основной асинхронной программы.
'''''
'''
Шаг 5: Вызов функции download_and_save_json из main:
'''
await download_and_save_json(json_url, output_filename)
'''
Вызывается ранее определенная функция download_and_save_json.
'''
'''
Шаг 6: Запуск асинхронной программы:
'''
if __name__ == "__main__":
    asyncio.run(main())
'''
Описание:

Запуск основной асинхронной программы с использованием asyncio.run(main()).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'output_async.json'
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
{
  "id": 1,
  "name": "Leanne Graham",
  "username": "Bret",
  "email": "Sincere@april.biz",
  "address": {
    "street": "Kulas Light",
    "suite": "Apt. 556",
    "city": "Gwenborough",
    "zipcode": "92998-3874",
    "geo": {
      "lat": "-37.3159",
      "lng": "81.1496"
    }
  },
  "phone": "1-770-736-8031 x56442",
  "website": "hildegard.org",
  "company": {
    "name": "Romaguera-Crona",
    "catchPhrase": "Multi-layered client-server neural-net",
    "bs": "harness real-time e-markets"
  }
}
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Мы также можем загружать JSON-объекты с различных открытых API, которые предоставляют данные в формате JSON. 
Например, некоторые популярные открытые API, которые предоставляют JSON-данные:

1. JSONPlaceholder: 
Это фейковый онлайн-сервис для тестирования и прототипирования, который предоставляет тестовые данные в формате JSON.

    URL: https://jsonplaceholder.typicode.com

2. OpenWeatherMap API: Предоставляет информацию о погоде в формате JSON.

    URL: https://openweathermap.org/api

3. GitHub API: Предоставляет данные о репозиториях, пользователях и других аспектах GitHub.

    URL: https://developer.github.com/v3

4. REST Countries API: Предоставляет информацию о странах в формате JSON.

    URL: https://restcountries.com

Просто выберите API, которое соответствует вашим интересам, и используйте его URL для загрузки JSON-данных.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import json

async def download_and_save_json(url, filename):
    try:
        async with aiohttp.ClientSession() as session, session.get(url) as response:
            response.raise_for_status()
            json_data = await response.json()

            with open(filename, 'w') as file:
                json.dump(json_data, file, indent=2)

            print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")

async def main():
    # Укажите URL и имя файла, куда сохранить JSON
    json_url = "https://jsonplaceholder.typicode.com/users/1"
    output_filename = "output_async.json"

    # Запуск асинхронной программы
    await download_and_save_json(json_url, output_filename)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек:
'''
import aiohttp
import asyncio
import json
'''
Импортируются библиотеки для асинхронных HTTP-запросов (aiohttp), 
работы с асинхронными задачами (asyncio), и работы с JSON-данными (json).
'''
'''
Определение асинхронной функции download_and_save_json:
'''
async def download_and_save_json(url, filename):
''''
Эта функция асинхронная и предназначена для загрузки JSON-данных по заданному URL и сохранения их в файл.
''''
'''
Блок try-except внутри функции:
'''
try:
    async with aiohttp.ClientSession() as session, session.get(url) as response:
        response.raise_for_status()
        json_data = await response.json()

        with open(filename, 'w') as file:
            json.dump(json_data, file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

except aiohttp.ClientError as e:
    print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")
'''
Создается асинхронная сессия и выполняется асинхронный GET-запрос с использованием контекстного менеджера async with.
Проверяется наличие ошибок при запросе с response.raise_for_status().
Получаются JSON-данные из ответа с использованием await response.json().
Сохраняются JSON-данные в файл с использованием json.dump.
Обрабатываются и выводятся ошибки с использованием aiohttp.ClientError.
'''
'''
Определение асинхронной функции main:
'''
async def main():
''''
Эта функция предназначена для запуска основной асинхронной программы.
''''
'''
Вызов функции download_and_save_json:
'''
await download_and_save_json(json_url, output_filename)
'''
Вызывается ранее определенная функция download_and_save_json с указанным URL и именем файла.
'''
'''
Запуск асинхронной программы:
'''
if __name__ == "__main__":
    asyncio.run(main())
'''
Запуск основной асинхронной программы с использованием asyncio.run(main()).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Оба варианта кода выполняют асинхронное скачивание JSON-данных и сохранение их в файл.
Отличие в том, как они используют асинхронные контекстные менеджеры для удобства и читаемости кода.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Дата выполнения Домашней-Работы: 26-27 - ЯНВАРЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №27: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите массив json – объектов с сайта jsonplaceholder, используя библиотеку aiohttp.
б) Сохраните циклом каждый в отдельный файл, в одну новую папку.

'''
'''
Урок от 26.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: (а и б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import json
import os

async def download_and_save_json(session, url, filename):
    try:
        # Отправляем GET-запрос по указанному URL
        async with session.get(url) as response:
            response.raise_for_status()  # Проверка наличия ошибок при запросе
            json_data = await response.json()  # Получаем JSON-данные из ответа

            # Сохраняем JSON-объект в файл
            with open(filename, 'w') as file:
                json.dump(json_data, file, indent=2)

            print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")

async def download_and_save_all_jsons():
    # Указываем базовый URL для загрузки JSON-объектов
    json_url_base = "https://jsonplaceholder.typicode.com/posts/"
    output_folder = "json_files"  # Название папки для сохранения файлов

    # Создаем папку, если ее нет
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Создаем сессию aiohttp
    async with aiohttp.ClientSession() as session:
        # Запускаем цикл для загрузки и сохранения каждого JSON-объекта
        for post_id in range(1, 51):  # Пример: загрузим 50 JSON-объектов
            json_url = f"{json_url_base}{post_id}"
            output_filename = os.path.join(output_folder, f"post_{post_id}.json")

            # Вызываем функцию загрузки и сохранения JSON-объекта
            await download_and_save_json(session, json_url, output_filename)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(download_and_save_all_jsons())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Импорт библиотек:
'''
import aiohttp
import asyncio
import json
import os
'''
Импорт необходимых библиотек: aiohttp для асинхронных HTTP-запросов, asyncio для работы с
асинхронными задачами, json для работы с JSON-данными и os для работы с операционной системой.
'''
'''
Шаг №2: Определение асинхронной функции download_and_save_json:
'''
async def download_and_save_json(session, url, filename):
''''
Эта функция принимает асинхронную сессию (session), URL для запроса (url) и имя файла (filename).
''''
'''
Шаг №3: Блок try-except внутри функции:
'''
try:
    async with session.get(url) as response:
        response.raise_for_status()
        json_data = await response.json()

        with open(filename, 'w') as file:
            json.dump(json_data, file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

except aiohttp.ClientError as e:
    print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")
'''
Внутри try:
Отправляется асинхронный GET-запрос по указанному URL.
Проверяется наличие ошибок при запросе с использованием response.raise_for_status().
Получаются JSON-данные из ответа с использованием await response.json().
Сохраняются JSON-данные в файл.
Выводится сообщение об успешном сохранении.

В блоке except:
Обрабатываются ошибки с помощью aiohttp.ClientError.
'''
'''
Шаг №4: Определение асинхронной функции download_and_save_all_jsons:
'''
async def download_and_save_all_jsons():
''''
Эта функция предназначена для загрузки и сохранения JSON-объектов.
''''
'''
Шаг №5: Указание базового URL и создание папки:
'''
json_url_base = "https://jsonplaceholder.typicode.com/posts/"
output_folder = "json_files"

if not os.path.exists(output_folder):
    os.makedirs(output_folder)
'''
Указывается базовый URL для загрузки JSON-объектов и создается папка для сохранения файлов, если ее нет.
'''
'''
Шаг №6: Создание сессии aiohttp:
'''
async with aiohttp.ClientSession() as session:
'''
Создается асинхронная сессия aiohttp.ClientSession().
'''
'''
Шаг №7: Цикл для загрузки и сохранения JSON-объектов:
'''
for post_id in range(1, 51):
    json_url = f"{json_url_base}{post_id}"
    output_filename = os.path.join(output_folder, f"post_{post_id}.json")

    await download_and_save_json(session, json_url, output_filename)
'''
Для каждого post_id в диапазоне от 1 до 50:
Формируется URL для запроса.
Формируется имя файла для сохранения.
Вызывается асинхронная функция download_and_save_json.
'''
'''
Шаг №8: Запуск асинхронной программы:
'''
if __name__ == "__main__":
    asyncio.run(download_and_save_all_jsons())
'''
Проверяется, что скрипт был запущен как основной,
 и запускается асинхронная функция download_and_save_all_jsons с помощью asyncio.run().
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Мы также можем загружать JSON-объекты с различных открытых API, которые предоставляют данные в формате JSON. 
Например, некоторые популярные открытые API, которые предоставляют JSON-данные:

1. JSONPlaceholder: 
Это фейковый онлайн-сервис для тестирования и прототипирования, который предоставляет тестовые данные в формате JSON.

    URL: https://jsonplaceholder.typicode.com

2. OpenWeatherMap API: Предоставляет информацию о погоде в формате JSON.

    URL: https://openweathermap.org/api

3. GitHub API: Предоставляет данные о репозиториях, пользователях и других аспектах GitHub.

    URL: https://developer.github.com/v3

4. REST Countries API: Предоставляет информацию о странах в формате JSON.

    URL: https://restcountries.com

Просто выберите API, которое соответствует вашим интересам, и используйте его URL для загрузки JSON-данных.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Дата выполнения Практической-Работы: 31 - ЯНВАРЯ - 01 ФЕВРАЛЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №28: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1

а) Загрузите одиночную html страницу с сайта википедии, используя библиотеку requests, затем aiohttp.
б) Сохраните оба раза её в файл.
'''
'''
Урок от 31.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а) & б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import requests
import aiohttp
import asyncio

# Функция для загрузки HTML-страницы с использованием библиотеки requests
def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

# Асинхронная функция для загрузки HTML-страницы с использованием библиотеки aiohttp
async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Создание сессии aiohttp
        async with aiohttp.ClientSession() as session:
            # Запрос HTML-страницы по URL
            async with session.get(url) as response:
                response.raise_for_status()
                # Получение текста HTML-страницы
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                with open(filename, 'w', encoding='utf-8') as file:
                    file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

# Асинхронная функция для выполнения основной логики
async def main():
    # Укажите URL и имена файлов для сохранения HTML
    html_url = "https://ru.wikipedia.org/wiki/Отряд_731"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Вызов функций для загрузки и сохранения HTML
    download_and_save_html_with_requests(html_url, html_filename_requests)
    await download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ДАЛЕЕ ПОДРОБНО И ШАГ ЗА ШАГОМ
'''
'''
Шаг №1: Функция download_and_save_html_with_requests
Пример кода:
'''
# Функция для загрузки HTML-страницы с использованием библиотеки requests
def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")
'''
Полное и подробное описание:

Эта функция предназначена для загрузки HTML-страницы с использованием библиотеки requests. Вот ее основные шаги:

Загрузка HTML-страницы: Функция отправляет GET-запрос к указанному URL с помощью requests.get(url).

Проверка ошибок: С помощью response.raise_for_status() происходит проверка наличия ошибок при запросе.
Если запрос был неудачным (например, код ответа не в диапазоне 200-299), вызывается
исключение requests.exceptions.RequestException.

Сохранение HTML: Если запрос был успешным, HTML-страница сохраняется в указанный
файл с помощью with open(filename, 'w', encoding='utf-8') as file.
'''
'''
Шаг №2: Функция download_and_save_html_with_aiohttp
Пример кода:
'''
# Асинхронная функция для загрузки HTML-страницы с использованием библиотеки aiohttp
async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Создание сессии aiohttp
        async with aiohttp.ClientSession() as session:
            # Запрос HTML-страницы по URL
            async with session.get(url) as response:
                response.raise_for_status()
                # Получение текста HTML-страницы
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                with open(filename, 'w', encoding='utf-8') as file:
                    file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")
'''
Полное и подробное описание:

Эта асинхронная функция использует библиотеку aiohttp для загрузки HTML-страницы. Вот шаги, которые она выполняет:

Создание сессии aiohttp: Асинхронный контекст async with aiohttp.ClientSession() as session создает сессию, 
которую мы будем использовать для отправки запросов.

Запрос HTML-страницы: С помощью async with session.get(url) as response отправляется асинхронный GET-запрос. 
Как и в предыдущей функции, здесь также используется response.raise_for_status() для проверки ошибок при запросе.

Сохранение HTML: Если запрос был успешным, текст HTML-страницы сохраняется в указанный файл, 
аналогично функции с использованием requests.
'''
'''
Шаг №3: Функция main
Пример кода:
'''
# Асинхронная функция для выполнения основной логики
async def main():
    # Укажите URL и имена файлов для сохранения HTML
    html_url = "https://ru.wikipedia.org/wiki/Отряд_731"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Вызов функций для загрузки и сохранения HTML
    download_and_save_html_with_requests(html_url, html_filename_requests)
    await download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:

Эта асинхронная функция main представляет основную логику программы. Она выполняет следующие шаги:

Задание URL и имен файлов: Определяется URL HTML-страницы и имена файлов, в которые будут сохранены страницы,
загруженные с использованием requests и aiohttp.

Вызов функций для загрузки и сохранения HTML: 
Вызываются обе функции download_and_save_html_with_requests и download_and_save_html_with_aiohttp для загрузки
и сохранения HTML-страниц.

Запуск асинхронной программы: asyncio.run(main()) запускает асинхронную программу, выполняя основную логику.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import aiofiles
import asyncio
import requests

async def download_and_save_html_with_requests(url, filename):
    try:
        response = requests.get(url)
        response.raise_for_status()
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)
        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)
                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

async def main():
    html_url = "https://ru.wikipedia.org/wiki/Mortal_Kombat"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Используйте asyncio.gather для параллельного выполнения асинхронных функций
    await asyncio.gather(
        download_and_save_html_with_requests(html_url, html_filename_requests),
        download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)
    )

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ДАЛЕЕ ПОДРОБНО И ШАГ ЗА ШАГОМ
'''
'''
Шаг №1: Асинхронные функции download_and_save_html_with_requests и download_and_save_html_with_aiohttp
Пример кода:
'''
async def download_and_save_html_with_requests(url, filename):
    try:
        response = requests.get(url)
        response.raise_for_status()
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)
        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)
                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")
'''
Полное и подробное описание:
download_and_save_html_with_requests:

Эта асинхронная функция использует библиотеку requests для синхронной загрузки HTML-страницы.
Она отправляет GET-запрос по указанному URL с помощью requests.get(url).
Если запрос успешен, текст HTML-страницы записывается в файл с использованием стандартного модуля open.
download_and_save_html_with_aiohttp:

Эта асинхронная функция использует библиотеки aiohttp и aiofiles для асинхронной загрузки и сохранения HTML-страницы.
Создается асинхронная сессия с aiohttp.ClientSession().
Отправляется асинхронный GET-запрос, и при успешном выполнении 
текст HTML-страницы записывается в файл с помощью aiofiles.open.
'''
'''
Шаг №2: Асинхронная функция main
Пример кода:
'''
async def main():
    html_url = "https://ru.wikipedia.org/wiki/Mortal_Kombat"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Используйте asyncio.gather для параллельного выполнения асинхронных функций
    await asyncio.gather(
        download_and_save_html_with_requests(html_url, html_filename_requests),
        download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)
    )

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Эта асинхронная функция main является точкой входа программы.
Она задает URL HTML-страницы и имена файлов для сохранения результатов с использованием requests и aiohttp.
Для параллельного выполнения обеих асинхронных функций используется asyncio.gather.
'''
'''
Шаг №3: Запуск программы
Последний блок кода if __name__ == "__main__": использует asyncio.run(main()) для запуска асинхронной программы.
'''
'''
Когда программа выполняется, она асинхронно загружает HTML-страницу дважды: первый раз с использованием requests, 
второй раз с использованием aiohttp. Обе версии страницы сохраняются в соответствующие файлы.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №3
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import requests
import aiohttp
import aiofiles
import asyncio
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import os
import re

def clean_filename(filename):
    # Очистка имени файла от недопустимых символов
    return re.sub(r'[\/:*?"<>|]', '', filename)

async def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с использованием requests
        response = requests.get(url)
        response.raise_for_status()

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (requests): {e}")
        return None

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Загрузка HTML-страницы с использованием aiohttp
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
                return html_content
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (aiohttp): {e}")
        return None

async def download_and_save_images(html_content, base_url, image_folder):
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        image_tags = soup.find_all('img')

        # Проверка и создание директории, если она не существует
        if not os.path.exists(image_folder):
            os.makedirs(image_folder)

        for img_tag in image_tags:
            src = img_tag.get('src')
            if src:
                abs_url = urljoin(base_url, src)
                image_response = requests.get(abs_url)
                image_response.raise_for_status()

                filename = clean_filename(os.path.basename(urlparse(abs_url).path))
                image_filename = os.path.join(image_folder, filename)

                with open(image_filename, 'wb') as image_file:
                    image_file.write(image_response.content)

                print(f"Изображение успешно загружено и сохранено в файл {image_filename}")
    except Exception as e:
        print(f"Произошла ошибка при загрузке или сохранении изображений: {e}")

async def main():
    wikipedia_url = "https://ru.wikipedia.org/wiki/Призрак_в_доспехах"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"
    image_folder = "downloaded_images"

    html_content_requests = await download_and_save_html_with_requests(wikipedia_url, html_filename_requests)
    html_content_aiohttp = await download_and_save_html_with_aiohttp(wikipedia_url, html_filename_aiohttp)

    if html_content_requests:
        await download_and_save_images(html_content_requests, wikipedia_url, image_folder)

    if html_content_aiohttp:
        await download_and_save_images(html_content_aiohttp, wikipedia_url, image_folder)

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ДАЛЕЕ ПОДРОБНО И ШАГ ЗА ШАГОМ
'''
'''
Шаг №1: Функция clean_filename
Пример кода:
'''
def clean_filename(filename):
    # Очистка имени файла от недопустимых символов
    return re.sub(r'[\/:*?"<>|]', '', filename)
'''
Полное и подробное описание:
Эта функция принимает имя файла и удаляет из него недопустимые символы для использования в системе файлов.
'''
'''
Шаг №2: Асинхронные функции download_and_save_html_with_requests и download_and_save_html_with_aiohttp
Пример кода:
'''
async def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с использованием requests
        response = requests.get(url)
        response.raise_for_status()

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (requests): {e}")
        return None

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Загрузка HTML-страницы с использованием aiohttp
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
                return html_content
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (aiohttp): {e}")
        return None
'''
Полное и подробное описание:
download_and_save_html_with_requests:

Эта асинхронная функция использует библиотеку requests для синхронной загрузки HTML-страницы.
Она отправляет GET-запрос по указанному URL с помощью requests.get(url).
Если запрос успешен, текст HTML-страницы записывается в файл с использованием стандартного модуля open.
download_and_save_html_with_aiohttp:

Эта асинхронная функция использует библиотеки aiohttp и aiofiles для асинхронной загрузки и сохранения HTML-страницы.
Создается асинхронная сессия с aiohttp.ClientSession().
Отправляется асинхронный GET-запрос, и при успешном выполнении текст HTML-страницы записывается 
в файл с помощью aiofiles.open.
'''
'''
Шаг №3: Асинхронная функция download_and_save_images
Пример кода:
'''
async def download_and_save_images(html_content, base_url, image_folder):
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        image_tags = soup.find_all('img')

        # Проверка и создание директории, если она не существует
        if not os.path.exists(image_folder):
            os.makedirs(image_folder)

        for img_tag in image_tags:
            src = img_tag.get('src')
            if src:
                abs_url = urljoin(base_url, src)
                image_response = requests.get(abs_url)
                image_response.raise_for_status()

                filename = clean_filename(os.path.basename(urlparse(abs_url).path))
                image_filename = os.path.join(image_folder, filename)

                with open(image_filename, 'wb') as image_file:
                    image_file.write(image_response.content)

                print(f"Изображение успешно загружено и сохранено в файл {image_filename}")
    except Exception as e:
        print(f"Произошла ошибка при загрузке или сохранении изображений: {e}")
'''
Полное и подробное описание:
Эта асинхронная функция принимает HTML-контент, базовый URL и директорию для сохранения изображений.
С использованием BeautifulSoup находятся все теги <img> в HTML-странице.
Проверяется существование директории для сохранения изображений, и если ее нет, она создается.
Для каждого тега <img> извлекается атрибут src и формируется абсолютный URL.
Изображение загружается с использованием requests.get, сохраняется в файл, и выводится сообщение об успешной загрузке.
'''
'''
Шаг №4: Асинхронная функция main
Пример кода:
'''
async def main():
    wikipedia_url = "https://ru.wikipedia.org/wiki/Призрак_в_доспехах"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"
    image_folder = "downloaded_images"

    html_content_requests = await download_and_save_html_with_requests(wikipedia_url, html_filename_requests)
    html_content_aiohttp = await download_and_save_html_with_aiohttp(wikipedia_url, html_filename_aiohttp)

    if html_content_requests:
        await download_and_save_images(html_content_requests, wikipedia_url, image_folder)

    if html_content_aiohttp:
        await download_and_save_images(html_content_aiohttp, wikipedia_url, image_folder)

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Асинхронная функция main является точкой входа в программу.
Задается URL страницы Википедии, имена файлов для сохранения HTML, и директория для сохранения изображений.
Сначала вызываются асинхронные функции для загрузки HTML-страницы с использованием requests и aiohttp.
Затем, если HTML-контент был успешно получен, вызываются функции для загрузки и сохранения изображений.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~

'''
Дата выполнения Домашней-Работы: 31 - ЯНВАРЯ - 01 ФЕВРАЛЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №28: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1

а) Загрузите циклом 10 рандомных картинок с сайта используя библиотеку requests, затем aiohttp.
б) Сохраните их в разные папки циклом.
'''
'''
Урок от 31.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а) & б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import os
import requests
import asyncio
import aiohttp
import random
import string

# Создаем папки для сохранения изображений
os.makedirs('images_sync_lorem', exist_ok=True)
os.makedirs('images_async_lorem', exist_ok=True)

def generate_random_filename(length=10, extension='.jpg'):
    letters_and_digits = string.ascii_letters + string.digits
    random_str = ''.join(random.choice(letters_and_digits) for _ in range(length))
    return f"{random_str}{extension}"

async def download_image_async(session, url, folder, extension='.jpg'):
    async with session.get(url) as response:
        if response.status == 200:
            # Генерируем рандомное имя файла
            filename = os.path.join(folder, generate_random_filename(extension=extension))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Lorem Picsum: {filename}')

def download_image_sync(url, folder, extension='.jpg'):
    response = requests.get(url)
    if response.status_code == 200:
        # Генерируем рандомное имя файла
        filename = os.path.join(folder, generate_random_filename(extension=extension))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Lorem Picsum: {filename}')

async def main():
    # URL-адреса для загрузки с Lorem Picsum
    lorem_image_urls = [
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
    ]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_lorem') for url in lorem_image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in lorem_image_urls:
        download_image_sync(url, 'images_sync_lorem')

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Создание папок и функции генерации имени файла
'''
# Создаем папки для сохранения изображений
os.makedirs('images_sync_lorem', exist_ok=True)
os.makedirs('images_async_lorem', exist_ok=True)

def generate_random_filename(length=10, extension='.jpg'):
    letters_and_digits = string.ascii_letters + string.digits
    random_str = ''.join(random.choice(letters_and_digits) for _ in range(length))
    return f"{random_str}{extension}"
'''
Здесь создаются папки для сохранения изображений.
Функция generate_random_filename генерирует случайное имя файла заданной длины с указанным расширением.
'''
'''
Шаг №2: Асинхронные функции для загрузки изображений с использованием aiohttp
'''
async def download_image_async(session, url, folder, extension='.jpg'):
    async with session.get(url) as response:
        if response.status == 200:
            # Генерируем рандомное имя файла
            filename = os.path.join(folder, generate_random_filename(extension=extension))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Lorem Picsum: {filename}')
'''
Эта асинхронная функция использует библиотеку aiohttp для асинхронной загрузки изображения.
Она принимает сессию, URL изображения, папку для сохранения и расширение файла.
Загруженное изображение сохраняется в файл с рандомным именем в указанной папке.
'''
'''
Шаг №3: Синхронная функция для загрузки изображений с использованием requests
'''
def download_image_sync(url, folder, extension='.jpg'):
    response = requests.get(url)
    if response.status_code == 200:
        # Генерируем рандомное имя файла
        filename = os.path.join(folder, generate_random_filename(extension=extension))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Lorem Picsum: {filename}')
'''
Эта синхронная функция использует библиотеку requests для синхронной загрузки изображения.
Она принимает URL изображения, папку для сохранения и расширение файла.
Загруженное изображение также сохраняется в файл с рандомным именем в указанной папке.
'''
'''
Шаг №4: Асинхронная функция main
'''
async def main():
    # URL-адреса для загрузки с Lorem Picsum
    lorem_image_urls = [
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
    ]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_lorem') for url in lorem_image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in lorem_image_urls:
        download_image_sync(url, 'images_sync_lorem')

if __name__ == "__main__":
    asyncio.run(main())
'''
В асинхронной функции main создается список URL-адресов изображений с Lorem Picsum.
Затем используется aiohttp для асинхронной загрузки изображений в одну папку (images_async_lorem).
После этого используется requests для синхронной загрузки тех же изображений в другую папку (images_sync_lorem).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import os
import requests
import asyncio
import aiohttp

# Замените 'свой ключ API Pixabay' на ваш ключ API Pixabay
PIXABAY_API_KEY = 'свой ключ API Pixabay'

# Создаем папки для сохранения изображений
os.makedirs('images_sync_pixabay', exist_ok=True)
os.makedirs('images_async_pixabay', exist_ok=True)

async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем имя файла из URL
            filename = os.path.join(folder, os.path.basename(url))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Pixabay: {filename}')

def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем имя файла из URL
        filename = os.path.join(folder, os.path.basename(url))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Pixabay: {filename}')

async def main():
    # Запрос на получение 10 случайных изображений с Pixabay
    pixabay_url = f'https://pixabay.com/api/?key={PIXABAY_API_KEY}&per_page=10'
    response = requests.get(pixabay_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()['hits']
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['largeImageURL'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_pixabay') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_pixabay')

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Установка ключа API Pixabay
'''
PIXABAY_API_KEY = 'свой ключ API Pixabay'
'''
Здесь вы должны заменить 'свой ключ API Pixabay' на свой собственный ключ API Pixabay.
'''
'''
Шаг №2: Создание папок для сохранения изображений
'''
# Создаем папки для сохранения изображений
os.makedirs('images_sync_pixabay', exist_ok=True)
os.makedirs('images_async_pixabay', exist_ok=True)
'''
Этот код создает две папки: images_sync_pixabay и images_async_pixabay для сохранения изображений.
Опция exist_ok=True позволяет не вызывать ошибку, если папка уже существует.
'''
'''
Шаг №3: Асинхронные функции для загрузки изображений с использованием aiohttp
'''
async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем имя файла из URL
            filename = os.path.join(folder, os.path.basename(url))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Pixabay: {filename}')
'''
Эта асинхронная функция использует библиотеку aiohttp для асинхронной загрузки изображения.
Она принимает сессию, URL изображения и папку для сохранения.
Извлекается имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг №4: Синхронная функция для загрузки изображений с использованием requests
'''
def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем имя файла из URL
        filename = os.path.join(folder, os.path.basename(url))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Pixabay: {filename}')
'''
Эта синхронная функция использует библиотеку requests для синхронной загрузки изображения.
Она принимает URL изображения и папку для сохранения.
Извлекается имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг №5: Асинхронная функция main
'''
async def main():
    # Запрос на получение 10 случайных изображений с Pixabay
    pixabay_url = f'https://pixabay.com/api/?key={PIXABAY_API_KEY}&per_page=10'
    response = requests.get(pixabay_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()['hits']
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['largeImageURL'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_pixabay') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_pixabay')

if __name__ == "__main__":
    asyncio.run(main())
'''
В функции main отправляется запрос к Pixabay для получения 10 случайных изображений.
Если запрос успешен, извлекаются URL-адреса изображений.
Затем используется aiohttp для асинхронной загрузки изображений в папку images_async_pixabay.
Также используется requests для синхронной загрузки изображений в папку images_sync_pixabay.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №3
 '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import os
import requests
import asyncio
import aiohttp
from pathlib import Path
from urllib.parse import urlparse

# Замените 'ваш ключ API Unsplash' на ваш ключ API Unsplash
UNSPLASH_ACCESS_KEY = 'ваш ключ API Unsplash'

# Создаем папки для сохранения изображений
os.makedirs('images_sync_unsplash', exist_ok=True)
os.makedirs('images_async_unsplash', exist_ok=True)

async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем безопасное имя файла из URL с использованием urllib.parse
            filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Unsplash: {filename}')

def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем безопасное имя файла из URL с использованием urllib.parse
        filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')

        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Unsplash: {filename}')

async def main():
    # Запрос на получение 10 случайных изображений с Unsplash
    unsplash_url = f'https://api.unsplash.com/photos/random?count=10&client_id={UNSPLASH_ACCESS_KEY}'
    response = requests.get(unsplash_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['urls']['raw'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_unsplash') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_unsplash')

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Установка ключа API Unsplash
'''
# Замените 'ваш ключ API Unsplash' на ваш ключ API Unsplash
UNSPLASH_ACCESS_KEY = 'ваш ключ API Unsplash'
'''
В этой части кода вы должны заменить 'ваш ключ API Unsplash' на свой собственный ключ API Unsplash.
'''
'''
Шаг 2: Создание папок для сохранения изображений
'''
# Создаем папки для сохранения изображений
os.makedirs('images_sync_unsplash', exist_ok=True)
os.makedirs('images_async_unsplash', exist_ok=True)
'''
Этот код создает две папки: images_sync_unsplash и images_async_unsplash для сохранения изображений.
Опция exist_ok=True позволяет не вызывать ошибку, если папка уже существует.
'''
'''
Шаг 3: Асинхронные функции для загрузки изображений с использованием aiohttp
'''
async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем безопасное имя файла из URL с использованием urllib.parse
            filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Unsplash: {filename}')
'''
Эта асинхронная функция использует библиотеку aiohttp для асинхронной загрузки изображения.
Она принимает сессию, URL изображения и папку для сохранения.
Извлекается безопасное имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг 4: Синхронная функция для загрузки изображений с использованием requests
'''
def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем безопасное имя файла из URL с использованием urllib.parse
        filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')

        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Unsplash: {filename}')
'''
Эта синхронная функция использует библиотеку requests для синхронной загрузки изображения.
Она принимает URL изображения и папку для сохранения.
Извлекается безопасное имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг 5: Асинхронная функция main
'''
async def main():
    # Запрос на получение 10 случайных изображений с Unsplash
    unsplash_url = f'https://api.unsplash.com/photos/random?count=10&client_id={UNSPLASH_ACCESS_KEY}'
    response = requests.get(unsplash_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['urls']['raw'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_unsplash') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_unsplash')

if __name__ == "__main__":
    asyncio.run(main())
'''
В функции main отправляется запрос к Unsplash для получения 10 случайных изображений.
Если запрос успешен, извлекаются URL-адреса изображений.
Затем используется aiohttp для асинхронной загрузки изображений в папку images_async_unsplash.
Также используется requests для синхронной загрузки изображений в папку images_sync_unsplash.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~

'''
Дата выполнения Практической-Работы: 02 - ФЕВРАЛЯ - 03 ФЕВРАЛЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №29: Программы с интерфейсом (GUI) - tkinter и pyqt6

Выполните следующие задания:

Задание №1

а) Сделать набросок дизайна программы в figma / paint для программы, которая читает все excel-файлы в 
папке и выводит на экран общее количество строк.

б) Разработать эту программу на библиотеке tkinter.
'''
'''
Урок от 02.02.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а) & б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_lines += sheet.max_row
            except Exception as e:
                print(f"Считывание ошибки {filename}: {e}")

    return total_lines

def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        total_lines = count_lines_in_excel_files(folder_path)
        result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")

# GUI setup
window = tk.Tk()
window.title("Счетчик строк Excel")

# Установим размер окна 500x300
window.geometry('500x300')

browse_button = tk.Button(window, text="Просмотр папки", command=browse_folder)
browse_button.pack(pady=20)

result_label = tk.Label(window, text="")
result_label.pack()

window.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Импорт необходимых библиотек.
'''
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
'''
Полное и подробное описание:

os: Библиотека для взаимодействия с операционной системой, используется для работы с файловой системой.
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
filedialog: Модуль в Tkinter для диалогового взаимодействия с файлами и папками.
openpyxl: Библиотека для работы с файлами формата Excel (.xlsx).
'''
'''
Шаг №2: Определение функции count_lines_in_excel_files.
'''
def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_lines += sheet.max_row
            except Exception as e:
                print(f"Считывание ошибки {filename}: {e}")

    return total_lines
'''
Полное и подробное описание:

count_lines_in_excel_files: Функция, которая принимает путь к папке и возвращает общее количество строк во всех 
файлах Excel в данной папке.
total_lines: Переменная для хранения общего количества строк.
for filename in os.listdir(folder_path): Цикл проходит по всем файлам в указанной папке.
if filename.endswith(".xlsx"):: Проверка, что файл имеет расширение .xlsx.
file_path = os.path.join(folder_path, filename): Формирование полного пути к файлу.
try ... except Exception as e:: Обработка исключений при загрузке файла Excel. Если возникает ошибка, 
выводится сообщение об ошибке.
'''
'''
Шаг №3: Определение функции browse_folder.
'''
def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        total_lines = count_lines_in_excel_files(folder_path)
        result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")
'''
Полное и подробное описание:

browse_folder: Функция, вызываемая при нажатии кнопки "Просмотр папки".
Открывает диалог выбора папки и вызывает count_lines_in_excel_files для подсчета строк в файлах Excel.
filedialog.askdirectory(): Пользователь выбирает папку через диалоговое окно.
if folder_path:: Проверка, что пользователь выбрал папку.
total_lines = count_lines_in_excel_files(folder_path): Вызов функции count_lines_in_excel_files для подсчета строк.
result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}"): Обновление метки
на GUI с общим количеством строк.
'''
'''
Шаг №4: Настройка графического интерфейса (GUI).
'''
window = tk.Tk()
window.title("Счетчик строк Excel")

window.geometry('500x300')

browse_button = tk.Button(window, text="Просмотр папки", command=browse_folder)
browse_button.pack(pady=20)

result_label = tk.Label(window, text="")
result_label.pack()

window.mainloop()
'''
Полное и подробное описание:

tk.Tk(): Создание главного окна приложения.
window.title("Счетчик строк Excel"): Установка заголовка окна.
window.geometry('500x300'): Установка размеров окна.
tk.Button(window, text="Просмотр папки", command=browse_folder): Создание
кнопки "Просмотр папки", которая при нажатии вызывает функцию browse_folder.
tk.Label(window, text=""): Создание метки для вывода результата.
window.mainloop(): Запуск главного цикла обработки событий Tkinter.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook

result_label = None  # Объявляем переменную здесь

def count_lines_in_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        total_lines = sum(sheet.max_row for sheet in workbook)
        return total_lines
    except Exception as e:
        return f"Считывание ошибки {file_path}: {e}"

def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            total_lines += count_lines_in_excel_file(file_path)

    return total_lines

def browse_folder():
    global result_label  # Используем глобальную переменную
    try:
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Файлы Excel", "*.xlsx")])
        if file_path:
            total_lines = count_lines_in_excel_file(file_path)
            result_label.config(text=f"Общее количество строк в файле Excel: {total_lines}")
    except Exception as e:
        result_label.config(text=f"Ошибка: {e}")

def clear_result():
    global result_label  # Используем глобальную переменную
    result_label.config(text="")

# GUI setup
def create_gui():
    global result_label  # Используем глобальную переменную
    window = tk.Tk()
    window.title("Счетчик строк Excel")
    window.geometry('500x300')

    browse_button = tk.Button(window, text="Просмотр файла Excel", command=browse_folder)
    browse_button.pack(pady=10)

    clear_button = tk.Button(window, text="Очистить результат", command=clear_result)
    clear_button.pack()

    result_label = tk.Label(window, text="")
    result_label.pack()

    window.mainloop()

if __name__ == "__main__":
    create_gui()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
1. Импорт библиотек и объявление глобальных переменных:
'''
import os
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook

result_label = None  # Объявляем переменную здесь
'''
Описание:
os: Предоставляет функции для работы с операционной системой, такие как чтение содержимого папки.
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
filedialog: Подмодуль tkinter для работы с диалоговыми окнами выбора файлов и папок.
ttk: Подмодуль tkinter, предоставляющий расширенные виджеты.
openpyxl: Библиотека для работы с файлами Excel формата xlsx.
result_label: Глобальная переменная, которая будет использоваться для отображения результата в GUI.
'''
'''
2. Функция count_lines_in_excel_file:
'''
def count_lines_in_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        total_lines = sum(sheet.max_row for sheet in workbook)
        return total_lines
    except Exception as e:
        return f"Считывание ошибки {file_path}: {e}"
'''
Описание:
count_lines_in_excel_file: Функция, которая считает общее количество строк в файле Excel.
Открывает файл, считывает количество строк в каждом листе и возвращает сумму.
Обрабатывает возможные исключения, выводя сообщение об ошибке.
'''
'''
3. Функция count_lines_in_excel_files:
'''
def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            total_lines += count_lines_in_excel_file(file_path)

    return total_lines
'''
Описание:
count_lines_in_excel_files: Функция, которая проходит через все файлы с 
расширением ".xlsx" в указанной папке и считает общее количество строк.
Использует предыдущую функцию count_lines_in_excel_file.
'''
'''
4. Функция browse_folder:
'''
def browse_folder():
    global result_label  # Используем глобальную переменную
    try:
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Файлы Excel", "*.xlsx")])
        if file_path:
            total_lines = count_lines_in_excel_file(file_path)
            result_label.config(text=f"Общее количество строк в файле Excel: {total_lines}")
    except Exception as e:
        result_label.config(text=f"Ошибка: {e}")
'''
Описание:
browse_folder: Функция, вызываемая при нажатии кнопки "Просмотр файла Excel".
Открывает диалоговое окно выбора файла Excel, считает количество строк и обновляет result_label в GUI.
'''
'''
Функция clear_result:
'''
def clear_result():
    global result_label  # Используем глобальную переменную
    result_label.config(text="")
'''
Описание:
clear_result: Функция, вызываемая при нажатии кнопки "Очистить результат".
Очищает result_label в GUI.
'''
'''
6. Настройка GUI в функции create_gui:
'''
def create_gui():
    global result_label  # Используем глобальную переменную
    window = tk.Tk()
    window.title("Счетчик строк Excel")
    window.geometry('500x300')

    browse_button = tk.Button(window, text="Просмотр файла Excel", command=browse_folder)
    browse_button.pack(pady=10)

    clear_button = tk.Button(window, text="Очистить результат", command=clear_result)
    clear_button.pack()

    result_label = tk.Label(window, text="")
    result_label.pack()

    window.mainloop()
'''
Описание:
create_gui: Функция для создания основного GUI.
Создает окно, добавляет кнопки "Просмотр файла Excel" и "Очистить результат", а также result_label.
'''
'''
7. Запуск GUI, если код выполняется как самостоятельный скрипт:
'''
if __name__ == "__main__":
    create_gui()
'''
Описание:
Проверяет, запущен ли скрипт непосредственно (а не импортирован как модуль).
Если да, вызывает функцию create_gui для создания и запуска GUI.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~




'''
Дата выполнения Домашней-Работы: 02 - ФЕВРАЛЯ - 03 ФЕВРАЛЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №29: Программы с интерфейсом (GUI) tkinter и pyqt6

Выполните следующие задания:

Задание №1

а) Сделать набросок дизайна программы в figma / paint для программы, которая делает запрос на сайт jsonplaceholder с определённым id.
б) Разработать эту программу на библиотеке tkinter.
в) Реализовать сохранение полученного объекта в папку.
'''
'''
Урок от 02.02.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а), б) & в)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ только один (Возможно ОДИН)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
!!! 
(На первых скриншотах есть наименования в "" (ковычках) на ENG языке - так как писал код в "черновике", оставил как
есть - НО ДЛЯ "ЧИСТОВИКА" - перевел все на - РУССКИЙ ЯЗЫК - для удобства и красоты в интерфейсе.

* черновик - 02.02.2024_ONLY_TEST_CODE+.py
* ЧИСТОВИК - 2024_PYTHON_LESSON_№29_HW_02.02.2024_03.02.2024.py
'''
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import tkinter as tk
from tkinter import messagebox
import requests
import json

def fetch_data(user_id):
    try:
        # Замените URL на свой, если требуется
        url = f"https://jsonplaceholder.typicode.com/users/{user_id}"
        response = requests.get(url)
        response.raise_for_status()  # Проверяем наличие ошибок при запросе

        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Ошибка", f"Не удалось получить данные: {e}")
        return None

def save_data(data, file_path):
    try:
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=2)
        messagebox.showinfo("Успех", f"Данные, сохраненные в {file_path}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {e}")

def on_fetch_button_click():
    user_id = entry_user_id.get()
    if not user_id.isdigit():
        messagebox.showwarning("Предупреждение", "Идентификатор пользователя должен быть числом")
        return

    user_id = int(user_id)
    data = fetch_data(user_id)
    if data:
        save_data(data, "user_data.json")

# Настройка графического интерфейса пользователя
window = tk.Tk()
window.title("Средство извлечения данных JSONPlaceholder")
window.geometry('650x350')

label_user_id = tk.Label(window, text="Введите идентификатор пользователя:")
label_user_id.pack(pady=10)

entry_user_id = tk.Entry(window)
entry_user_id.pack(pady=10)

fetch_button = tk.Button(window, text="Извлечь данные", command=on_fetch_button_click)
fetch_button.pack(pady=10)

window.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Данный код представляет собой простое графическое приложение,
написанное с использованием библиотеки tkinter,
которое позволяет пользователю ввести идентификатор пользователя,
после чего отправляет запрос к JSONPlaceholder API для получения данных о пользователе с указанным идентификатором.
Полученные данные сохраняются в файл JSON.
Шаги по коду:
'''
'''
1. Импорт библиотек:
'''
import tkinter as tk
from tkinter import messagebox
import requests
import json
'''
Описание:
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
messagebox: Подмодуль tkinter для создания диалоговых окон сообщений.
requests: Библиотека для выполнения HTTP-запросов.
json: Библиотека для работы с форматом JSON.
'''
'''
2. Функция fetch_data:
'''
def fetch_data(user_id):
    try:
        url = f"https://jsonplaceholder.typicode.com/users/{user_id}"
        response = requests.get(url)
        response.raise_for_status()

        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Ошибка", f"Не удалось получить данные: {e}")
        return None
'''
Описание:
fetch_data: Функция, отправляющая запрос к JSONPlaceholder API для получения данных о пользователе.
Возвращает данные в формате JSON, если запрос успешен, или None, если возникла ошибка.
Если возникла ошибка, выводится диалоговое окно с сообщением об ошибке.
'''
'''
3. Функция save_data:
'''
def save_data(data, file_path):
    try:
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=2)
        messagebox.showinfo("Успех", f"Данные, сохраненные в {file_path}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {e}")
'''
Описание:
save_data: Функция, сохраняющая данные в файл JSON.
В случае успеха выводится сообщение об успешном сохранении, иначе - сообщение об ошибке.
'''
'''
4. Функция on_fetch_button_click:
'''
def on_fetch_button_click():
    user_id = entry_user_id.get()
    if not user_id.isdigit():
        messagebox.showwarning("Предупреждение", "Идентификатор пользователя должен быть числом")
        return

    user_id = int(user_id)
    data = fetch_data(user_id)
    if data:
        save_data(data, "user_data.json")
'''
Описание:
on_fetch_button_click: Функция, вызываемая при клике на кнопку "Извлечь данные".
Получает идентификатор пользователя из виджета ввода (entry_user_id).
Проверяет, что идентификатор является числом, и выводит предупреждение, если нет.
После получения данных о пользователе вызывает save_data для сохранения в файл "user_data.json".
'''
'''
5. Настройка графического интерфейса пользователя:
'''
window = tk.Tk()
window.title("Средство извлечения данных JSONPlaceholder")
window.geometry('650x350')

label_user_id = tk.Label(window, text="Введите идентификатор пользователя:")
label_user_id.pack(pady=10)

entry_user_id = tk.Entry(window)
entry_user_id.pack(pady=10)

fetch_button = tk.Button(window, text="Извлечь данные", command=on_fetch_button_click)
fetch_button.pack(pady=10)

window.mainloop()
'''
Описание:
Создает основное окно (Tk).
Добавляет виджеты: метку (Label), поле ввода (Entry), кнопку (Button).
Назначает функцию on_fetch_button_click обработчиком события для кнопки.
Запускает главный цикл событий (mainloop), который ожидает действий пользователя.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


< !--  ########################################################################################## -->
< !-- HTML  # 1 Recap + Вводный ПЕРВЫЙ УРОК - 08.01.2024 -->

< !-- < html >
< head >

< meta
charset = "UTF-8" / >
< title > Hello
HTML < / title >
< / head >
< body >

hello
world
Привет
Мир

< img
src = "https://pngicon.ru/file/uploads/serdce.png"
height = "150px"
width = "150px" / >
< h1 > Вау - я
программист < / h1 >
< / body >
< / html > -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !-- 2024
_HTML_LESSON_№2
_Структура_HTML.Теги_Атрибуты - 15.01
.2024 -->

< !--

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 2: Структура
HTML.Теги.Атрибуты.Правила
языка
разметки
-->

< !DOCTYPE
html >
< html
lang = "ru" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Страница
с
текстом
и
заголовками < / title >
< / head >
< body >

< !-- 1.
Создать
страницу
с
текстом
по
примеру
на
картинке. -->

< h2 >
Стихотворение
< / h2 >
< p >

Мириады
маленьких
дел < br >
Пьют
по
капле
гаснущий
день, < br >
А < i > дела
большие < / i > сушит
жажда. < br >
Оставляя
все
на «потом», < br >
Прозреваем
задним
числом, < br >
Только
год < b > не
повторится < / b > дважды. < br >
< br >
И.Тальков

< / p >

< !-- 2.
Повторите
по
данному
образцу: -->
< h1 > Это
заголовок < / h1 >
< h2 > Это
заголовок < / h2 >
< h3 > Это
заголовок < / h3 >
< h4 > Это
заголовок < / h4 >

< p > Это < b > абзац < / b >.< / p >
< p > Это
еще < i > абзац < / i >.< / p >

< h1 > < i > Это
заголовок
h1 < / i > < / h1 >

< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

03
_Введение
в
CSS.Форматирование
текста
при
помощи
CSS - 22.01
.2024
Классная
работа

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< link
rel = "preconnect"
href = "https://fonts.googleapis.com" >
< link
rel = "preconnect"
href = "https://fonts.gstatic.com"
crossorigin >
< link
href = "https://fonts.googleapis.com/css2?family=Rubik+Burned&display=swap"
rel = "stylesheet" >

< meta
charset = "UTF-8" >
< title > Lesson
2 < / title >
< script >
/ *легкий
init
js * /
< / script >
< meta
description = "Краткое описание содержимого/контента страницы" / >
< meta
keywords = "купить, по акции, новый" / >
< style >
html, body
{font - size: 10px;
font - family: "Rubik Burned", system - ui;}
h1
{font - size: 4rem}
h2
{font - size: 3rem}
h3
{font - size: 2rem}
h4
{font - size: 1rem}
< / style >
< link
rel = "stylesheet"
href = "./style.css" / >
< script
src = "https://cdn.tailwindcss.com" > < / script >
< / head >
< body >
< div
style = "display: flex" >
< img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
< img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
< img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
< / div >
< div >
< h1


class ="heading-of-section text-red text-krasnii" > Красный < / h1 >

< p


class ="text-orange" > Оранжевый < / p >

< div >
< p


class ="button default" > Нажми меня < / p >

< div


class ="button alert" > Нажми меня < / div >

< button


class ="button alert" alt="Для оформления заказа нажмите эту кнопку" > Нажми меня < / button >

< / div >
< div >
< div


class ="display-inline rounded-md shadow-md padding-md bg-gray" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-red" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-blue" > Нажми меня < / div >

< / div >
< / div >
< b > Жирный
текст < / b >
< h1
style = "text-decoration: line-through" > Панды, наконец - то
полетят
на
Марс < / h1 >
< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3


class ="heading" > Hello world < / h3 >

< h4


class ="text-red" > Hello world < / h4 >

< h5 > Hello
world < / h5 >
< h6


class ="bg-blue" > Hello world < / h6 >

< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3 > Hello
world < / h3 >
< h4
id = "heading-h4-1" > Hello
world < / h4 >
< h4
id = "heading-h4-2" > Hello
world < / h4 >
< h4 > Hello
world < / h4 >
< h3


class ="text-lg text-rose-500 bg-gray-100 rounded-md" > Hello world < / h3 >

< h3 > Hello
world < / h3 >

< p >
paragraph < b > lorem < / b >
< br / >
< br / >
< br / >
< br / >
< br / >
ipsum
set < i > < b > ВАЖНО! < / b > < / i > dolor
< br / >
ipsum
set < strong > < em > Кликни & rarr; < / em > < / strong > dolor
< br / >
2 * 2
< br / >
2
x2
< br / >
2 & cross;
2
< / p >
< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
22.01
.2024 - Домашняя
и
практическая - работы.
< !--  ########################################################################################## -->
Практическая
работа

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №3
Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Создать
HTML - страницу “Vehicle”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

2.
Создать
HTML - страницу “Lorem
Ipsum”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

3.
Создать
HTML - страницу “Математические
формулы”

Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.
Используйте
здесь
теги
h1 - h6, span, p, sup, sub
и
спецсимволы.
< !--  ########################################################################################## -->

Шрифты
находятся
в
папке
Материалы
к
занятию
в
архиве
Fonts.

1.
Пример
подключение
шрифтов
через
ссылку
2.
Пример
подключения
через
локальные
файлы
шрифтов
< !--  ########################################################################################## -->

№1

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Двигатель < / title >
< style >
body
{
    font - family: 'Tahoma', serif;
line - height: 1.6;
margin: 18
px;
}

h1
{
    color:  # 00c264;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align: justify;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.smaller - text
{
    font - size: 58 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
< / style >

< / head >
< body >
< h1


class ="larger-font2" >


Двигатель
< / h1 >
< !-- Текст
о
двигателе
с
сайта
https: // ru.wikipedia.org / wiki / Двигатель -->

< p


class ="larger-font" >

< b > Дви́гатель < / b > — устройство, преобразующее
какой - либо
вид < a
href = "https://ru.wikipedia.org/wiki/Энергия"


class ="no-underline" > энергии < / a >


в < a
href = "https://ru.wikipedia.org/wiki/Механическая_энергия"


class ="no-underline" > механическую < / a > работу.


Термин < b > мотор < / b > < a
href = "https://ru.wikipedia.org/wiki/Заимствование"


class ="no-underline" > заимствован < / a > в


первой
половине
XIX
века
из
немецкого
языка < b


class ="smaller-text colored-letter" >[1] < / b >


(нем.Motor — «двигатель», от лат.mōtor — «приводящий в движение»)
и
преимущественно
им
называют
электрические
двигатели
и
двигатели
внутреннего
сгорания < b


class ="smaller-text colored-letter" >


[2] < / b >.
< / p >
< p


class ="larger-font" >


Двигатели
подразделяют
на
первичные
и
вторичные.
К
первичным
относят
непосредственно
преобразующие
природные
энергетические
ресурсы
в
механическую
работу,
а
ко
вторичным — преобразующие
энергию, выработанную
или
накопленную
другими
источниками < b


class ="smaller-text colored-letter" >


[3] < / b >.
< / p >
< hr


class ="subtle-line" >

< p >
< h4 > Примечания < / h4 >
< h5 > [1] < / h5 > Шанский
Н.М., Боброва
Т.А.Кот // Школьный
этимологический
словарь
русского
языка.
Происхождение
слов. — 7 - е
изд., стереотип.. — М.: Дрофа, 2004. — 398, [2]
с.
< h5 > [2] < / h5 > Крысин
Л.П.Мотор // Толковый
словарь
иноязычных
слов. — М.: Эксмо, 2008. — 944
с. —
(Библиотека словарей).
< h5 > [3] < / h5 > Definition
of
motor | Dictionary.com(англ.).www.dictionary.com.
Дата
обращения: 27
января
2022.
Архивировано
27
января
2022
года.
< / p >
< / body >
< / html >
< !--  ########################################################################################## -->

№2

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Lorem
Ipsum < / title >
< style >

h1
{
    color:  # 5d0080;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align0: justify;
}
.left - aligned1
{
    text - align: left;
}

.right - aligned2
{
    text - align: right;
}

.custom - font
{
    font - family: "Courier New", monospace;
}
.custom - font2
{
    font - family: "mv boli", monospace;
}
.custom - font3
{
    font - family: "segoe print", monospace;
}
.custom - font4
{
    font - family: "LUCIDA SANS CONSOLE", monospace;
}
.custom - font5
{
    font - family: "GABRIOLA", monospace;
}
.custom - font6
{
    font - family: "tempus sans itc", monospace;
}
.bold - text
{
    font - weight: bold;
}

.italic - text
{
    font - style: italic;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.larger - font3
{
    font - size: 50px;
}
.centered - text
{
    text - align: center;
}

.centered - container
{
    margin: 0 auto;
width: 50 %; / *Измените
на
желаемую
ширину * /
}

.smaller - text
{
    font - size: 95 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text2
{
    font - size: 80 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text3
{
    font - size: 115 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text4
{
    font - size: 15 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.subtle - line2
{
    border: 1px solid  # f2f2f2; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
.custom - color1
{
    color:  # 001fb8; /* Замените на желаемый цвет */
}
.custom - color2
{
    color:  # 6e6e6e; /* Замените на желаемый цвет */
}
.page - container
{
    margin: 20px; / *Замените
на
желаемое
значение
отступа * /
}
< / style >
< / head >
< body >
< h1


class ="centered-text larger-font3 custom-font4" > Lorem Ipsum < / h1 >

< div


class ="right-aligned2 custom-color1 custom-font3 italic-text smaller-text" >

< p > "Neque porro quisquam est qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit...". < / p >
< / div >
< div


class ="right-aligned2 custom-font2 custom-color2 italic-text smaller-text2 larger-font" >

< p > "Нет никого, кто любил бы боль саму по себе, кто искал бы её и кто хотел бы иметь её
просто
потому, что
это
боль..
".</p>
< / div >
< hr


class ="subtle-line2" >

< !-- Пример
текста
Lorem
Ipsum -->
< div


class ="page-container" >

< p


class ="larger-font custom-font" > < b >


Классический
текст
Lorem
Ipsum, используемый
с
XVI
века
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>L</b>orem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod tempor incididunt
ut
labore
et
dolore
magna
aliqua.Ut
enim
ad
minim
veniam, quis
nostrud
exercitation
ullamco
laboris
nisi
ut
aliquip
ex
ea
commodo
consequat.Duis
aute
irure
dolor in reprehenderit in voluptate
velit
esse
cillum
dolore
eu
fugiat
nulla
pariatur.Excepteur
sint
occaecat
cupidatat
non
proident, sunt in culpa
qui
officia
deserunt
mollit
anim
id
est
laborum.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.32
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>S</b>ed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem
aperiam, eaque
ipsa
quae
ab
illo
inventore
veritatis
et
quasi
architecto
beatae
vitae
dicta
sunt
explicabo.
Nemo
enim
ipsam
voluptatem
quia
voluptas
sit
aspernatur
aut
odit
aut
fugit, sed
quia
consequuntur
magni
dolores
eos
qui
ratione
voluptatem
sequi
nesciunt.Neque
porro
quisquam
est, qui
dolorem
ipsum
quia
dolor
sit
amet,
consectetur, adipisci
velit, sed
quia
non
numquam
eius
modi
tempora
incidunt
ut
labore
et
dolore
magnam
aliquam
quaerat
voluptatem.Ut
enim
ad
minima
veniam, quis
nostrum
exercitationem
ullam
corporis
suscipit
laboriosam, nisi
ut
aliquid
ex
ea
commodi
consequatur? Quis
autem
vel
eum
iure
reprehenderit
qui in ea
voluptate
velit
esse
quam
nihil
molestiae
consequatur, vel
illum
qui
dolorem
eum
fugiat
quo
voluptas
nulla
pariatur?"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>B</b>ut I must explain to you how all this mistaken idea of denouncing pleasure and praising pain was born and
I
will
give
you
a
complete
account
of
the
system, and expound
the
actual
teachings
of
the
great
explorer
of
the
truth, the
master - builder
of
human
happiness.No
one
rejects, dislikes, or avoids
pleasure
itself, because
it is
pleasure, but
because
those
who
do
not know
how
to
pursue
pleasure
rationally
encounter
consequences
that
are
extremely
painful.Nor
again is there
anyone
who
loves or pursues or desires
to
obtain
pain
of
itself, because
it
is pain, but
because
occasionally
circumstances
occur in which
toil and pain
can
procure
him
some
great
pleasure.
To
take
a
trivial
example, which
of
us
ever
undertakes
laborious
physical
exercise, except to
obtain
some
advantage
from it? But
who
has
any
right
to
find
fault
with a man who chooses to enjoy a pleasure that has no annoying
consequences, or one
who
avoids
a
pain
that
produces
no
resultant
pleasure?"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.33
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>A</b>t vero eos et accusamus et iusto odio dignissimos ducimus qui blanditiis praesentium voluptatum deleniti
atque
corrupti
quos
dolores
et
quas
molestias
excepturi
sint
occaecati
cupiditate
non
provident, similique
sunt
in culpa
qui
officia
deserunt
mollitia
animi, id
est
laborum
et
dolorum
fuga.Et
harum
quidem
rerum
facilis
est
et
expedita
distinctio.Nam
libero
tempore, cum
soluta
nobis
est
eligendi
optio
cumque
nihil
impedit
quo
minus
id
quod
maxime
placeat
facere
possimus, omnis
voluptas
assumenda
est, omnis
dolor
repellendus.Temporibus
autem
quibusdam
et
aut
officiis
debitis
aut
rerum
necessitatibus
saepe
eveniet
ut
et
voluptates
repudiandae
sint
et
molestiae
non
recusandae.Itaque
earum
rerum
hic
tenetur
a
sapiente
delectus, ut
aut
reiciendis
voluptatibus
maiores
alias
consequatur
aut
perferendis
doloribus
asperiores
repellat.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>O</b>n the other hand, we denounce with righteous indignation and dislike men who are so beguiled and demoralized
by
the
charms
of
pleasure
of
the
moment, so
blinded
by
desire, that
they
cannot
foresee
the
pain and trouble
that
are
bound
to
ensue; and equal
blame
belongs
to
those
who
fail in their
duty
through
weakness
of
will, which is the
same as saying
through
shrinking
from toil and pain.These
cases
are
perfectly
simple and easy
to
distinguish.
In
a
free
hour, when
our
power
of
choice is untrammelled and when
nothing
prevents
our
being
able
to
do
what
we
like
best, every
pleasure is to
be
welcomed and every
pain
avoided.But in certain
circumstances and owing
to
the
claims
of
duty or the
obligations
of
business
it
will
frequently
occur
that
pleasures
have
to
be
repudiated and
annoyances
accepted.The
wise
man
therefore
always
holds in these
matters
to
this
principle
of
selection: he
rejects
pleasures
to
secure
other
greater
pleasures, or else he
endures
pains
to
avoid
worse
pains.
"
< / p >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

№3

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Математические
формулы < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'VERDANA', monospace; / *Устанавливаем
шрифт
для
всего
документа * /
line - height: 1.0; / *Задаем
высоту
строки
текста * /
margin: 70
px; / *Задаем
отступы
для
всего
документа * /
}

/ *Стили
для
заголовка
первого
уровня * /
h1
{
    color:  # 000000; /* Цвет текста */
        text - align: center; / *Выравнивание
текста
по
центру * /
font - size: 50
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
}

/ *Стили
для
заголовка
второго
уровня * /
h2
{
    color:  # 22009e; /* Цвет текста */
        font - size: 35
px; / *Размер
шрифта * /
margin - top: 20
px; / *Верхний
отступ * /
}

/ *Стили
для
подзаголовка * /
h3
{
    color:  # 00ff49; /* Цвет текста */
        font - size: 25
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
font - style: italic; / *Курсив * /
                         text - align: left; / *Выравнивание
текста
влево * /
background - color:  # a300a3b8; /* Цвет подложки */
padding: 10
px; / *Внутренний
отступ * /
border - radius: 45
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 320
px; / *Максимальная
ширина
подложки * /
width: 25 %; / *Ширина
подложки
равна
20 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: left; / *Выравнивание
по
левой
стороне * /
}

/ *Стили
для
обычного
текста
и
вложенных
элементов * /
p, span
{
    text - align: justify; / *Выравнивание
текста
по
ширине
с
обеих
сторон * /
}

/ *Стили
для
верхних
и
нижних
индексов * /
sup, sub
{
    font - size: 80 %; / *Размер
шрифта
для
верхних
и
нижних
индексов * /
}

/ *Отступ
для
блока
с
описанием
формулы * /
.formula - description
{
    margin - left: 20px; / *Отступ
слева * /
}

/ *Стили
для
контейнера
формулы
и
самой
формулы * /
.formula - container,.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
background - color:  # ff00008f; /* Цвет подложки */
padding: 2
px; / *Внутренний
отступ * /
border - radius: 55
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
15
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 670
px; / *Максимальная
ширина
подложки * /
width: 50 %; / *Ширина
подложки
равна
50 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: auto; / *Выравнивание
по
центру * /
}

/ *Стили
для
самой
формулы * /
.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
color:  # 0700ff; /* Цвет формулы */
font - size: 25
px; / *Размер
шрифта
формулы * /
font - style: italic; / *Курсив * /
                         font - weight: bold; / *Жирный
шрифт
формулы * /
}

/ *Стили
для
раздела
с
примерами * /
.example - section
{
    margin - top: 20px; / *Верхний
отступ
раздела
с
примерами * /
}

/ *Стили
для
примеров * /
.example
{
    font - size: 20px; / *Размер
шрифта
примеров * /
color:  # 555; /* Цвет текста примеров */
}
/ *Стили
для
цветной
буквы * /
.colored - letter
{
    color:  # ff0000; /* Цвет буквы */
}
/ *Стили
для
увеличенной
буквы * /
.enlarged - letter
{
    font - size: 24px; / *Размер
шрифта
для
увеличенной
буквы * /
}
/ *Стили
для
выделенных
букв * /
.special - letter
{
    font - family: 'Arial', sans - serif; / *Используемый
шрифт * /
font - style: italic; / *Курсив * /
                         color:  # 00adff; /* Цвет текста */
}
< / style >
< / head >
< body >
< !-- Заголовок
первого
уровня -->
< h1 > Математические
формулы < / h1 >

< !-- Формула
окружности -->
< h2 > Формула
окружности: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > x < sup > 2 < / sup > + y < sup > 2 < / sup > = r < sup > 2 < / sup > < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Эта
формула
описывает
уравнение
окружности
в
декартовой
системе
координат. < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > x < / span > и < span class ="colored-letter enlarged-letter special-letter" > y < / span > - координаты точки на плоскости.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус окружности.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула окружности описывает геометрическую фигуру,


представляющую
собой
множество
точек, равноудаленных
от
центра. < / b > < / p >
< p


class ="formula-description" > < b > Здесь x и y - координаты точек,


а
r - радиус
окружности. < / b > < / p >

< !-- Формула
площади
треугольника -->
< h2 > Формула
площади
треугольника: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > S = 0.5 * a * b * sin(C) < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > S < / span > - площадь треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > a < / span > и < span class ="colored-letter enlarged-letter special-letter" > b < / span > - длины сторон треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > C < / span > - угол между сторонами a и b (в радианах).< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула площади треугольника выражает площадь как половину произведения длин двух сторон на синус угла между ними.< / b > < / p >

< !-- Формула
объема
цилиндра -->
< h2 > Формула
объема
цилиндра: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > V = π * r < sup > 2 < / sup > * h < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > V < / span > - объем цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > π < / span > (пи) - математическая константа, приблизительно равная 3.14159.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус основания цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > h < / span > - высота цилиндра.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула объема цилиндра выражает объем как произведение площади основания на высоту.< / b > < / p >

< !-- Раздел
с
примерами
применения
формул -->
< div


class ="example-section" >

< h2 > Сферы
применения
формул: < / h2 >
< !-- Примеры -->
< p


class ="example" > 1. < b > Геометрия:<

    / b > расчет
геометрических
параметров
фигур. < / p >
< p


class ="example" > 2. < b > Инженерия:<

    / b > проектирование
и
расчеты
конструкций. < / p >
< p


class ="example" > 3. < b > Физика:<

    / b > моделирование
и
анализ
физических
явлений. < / p >
< / div >

< / body >
< / html >

< !--  ########################################################################################## -->

Бонус

1.
Пример
подключение
шрифтов
через
ссылку

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
внешних
шрифтов -->
< link
rel = "stylesheet"
href = "https://fonts.googleapis.com/css?family=YourFontFamily" >

< !-- Ваши
стили -->
< style >
body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< / style >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >

< !--
Замените
"https://fonts.googleapis.com/css?family=YourFontFamily"
на
фактическую
ссылку
на
шрифт, который
вы
хотите
использовать.Например, если
вы
хотите
использовать
шрифт
"Roboto", ссылка
будет
выглядеть
так: "https://fonts.googleapis.com/css?family=Roboto".

Обратите
внимание, что
это
только
пример, и
вам
нужно
заменить
"YourFontFamily"
и
"Your Title"
на
фактические
значения, используемые
в
вашем
проекте.
-->

2.
Пример
подключения
через
локальные
файлы
шрифтов

< !--
Поместите
шрифтовые
файлы
в
ваш
проект:
Сначала
загрузите
файлы
шрифтов(например,.woff,.woff2,.ttf) в
ваш
проект.
Создайте
папку, например, "fonts", и
поместите
файлы
шрифтов
в
неё.

Пример
структуры
проекта:
-->
/ your - project
/ fonts
your - font.woff
your - font.woff2
index.html
style.css
< !--
Добавьте
стили
в
ваш
CSS:
В
вашем
файле
стилей(например, style.css), определите
новый @ font - face
и
используйте
его
в
стилях
элементов.

Пример:
-->

@font - face


{
    font - family: 'YourFontFamily';
src: url('fonts/your-font.woff2')
format('woff2'),
url('fonts/your-font.woff')
format('woff');
/ *Дополнительные
параметры, если
нужны * /
}

body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< !--
Убедитесь, что
путь
к
файлам
шрифтов
в
url
соответствует
структуре
вашего
проекта.
-->
< !--
Подключите
стили
в
HTML:
В
вашем
файле
HTML(например, index.html), убедитесь, что
вы
подключили
ваш
CSS
файл.

Пример:
-->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
локальных
стилей -->
< link
rel = "stylesheet"
href = "style.css" >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >
< !--
Замените
"Your Title"
на
фактический
заголовок
вашей
страницы.

Это
позволит
вам
использовать
локальные
файлы
шрифтов
в
вашем
проекте.Убедитесь, что
пути
указаны
правильно
и
что
браузер
может
найти
и
загрузить
файлы
шрифтов.
-->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Попробуйте
каждый
тег

На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.

К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)

Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!

2.
Создайте
страницу “Romeo and Juliet”

3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.

4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.

< !--  ########################################################################################## -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS

Домашнее
задание №3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
< !--  #### -->
1.
Попробуйте
каждый
тег
< !--  #### -->
На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.
К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)
< !--  #### -->
Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!
< !--  #### -->
2.
Создайте
страницу “Romeo and Juliet”
< !--  #### -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.
< !--  #### -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.
< !--  #### -->

< !--  ########################################################################################## -->
1.
Попробуйте
каждый
тег
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
2.
Создайте
страницу “Romario and Juluya”
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Romeo and Juliet < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
text - align: center;
}

# title {
font - family: 'Georgia', serif;
color:  # FF4500; /* Цвет - оранжевый */
font - style: italic;
font - size: 36
px;
}

# author {
font - size: 18
px;
font - style: italic;
}

hr
{
    width: 50 %;
margin: auto;
margin - top: 20
px;
border: 1
px
solid  # 000;
}

# prologue {
font - size: 24
px;
margin - top: 20
px;
}

# left-text, #right-text {
display: inline - block;
width: 45 %;
vertical - align: top;
}

.act - title
{
    font - size: 28px;
margin - top: 20
px;
}

.scene - title
{
    font - size: 20px;
font - weight: bold;
margin - top: 10
px;
}

.character
{
    font - weight: bold;
}

.dialogue
{
    text - align: left;
margin - top: 10
px;
margin - bottom: 10
px;
}

.romeo
{
    color:  # 0000FF; /* Цвет - синий */
}

.juliet
{
    color:  # FF1493; /* Цвет - розовый */
}

.sampson
{
    color:  # 008000; /* Цвет - зеленый */
}

.gregory
{
    color:  # FFD700; /* Цвет - золотой */
}
< / style >
    < / head >
        < body >
        < div
id = "title" > ROMEO
AND
JULIET < / div >
           < div
id = "author" > By
William
Shakespeare < / div >
                < hr >

                < div
id = "prologue" >
     < div
id = "left-text" >
     Two
households, both
alike in dignity,
In
fair
Verona, where
we
lay
our
scene,
From
ancient
grudge
break
to
new
mutiny,
Where
civil
blood
makes
civil
hands
unclean.
From
forth
the
fatal
loins
of
these
two
foes
A
pair
of
star - cross’d
lovers
take
their
life;
Whole
misadventured
piteous
overthrows
< / div >

< div
id = "right-text" >
Do
with their death bury their parents’ strife.
The
fearful
passage
of
their
death - mark’d
love,
And
the
continuance
of
their
parents’ rage,
Which, but
their
children’s
end, nought
could
remove,
Is
now
the
two
hours’ traffic
of
our
stage;
The
which if you
with patient ears attend,
What
here
shall
miss, our
toil
shall
strive
to
mend.
< / div >
< / div >
< hr >

< div


class ="act-title" > ACT I < / div >

< div


class ="scene-title" > SCENE I.Verona.A public place.< / div >

< div


class ="scene-title" > < i > Enter SAMPSON and GREGORY, of the house


of
CAPULET, armed
with swords and bucklers </ i > < / div >

< div
id = "left-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > Gregory, o’ my word, we’ll not carry coals.< br >

< span


class ="character gregory" > GREGORY < / span > No, for then we should be colliers.< br >

< span


class ="character sampson" > SAMPSON < / span > I mean, an we be in choler, we’ll draw.< br >

< span


class ="character gregory" > GREGORY < / span > Ay, while you live, draw your neck out


o’ the
collar. < br >
< span


class ="character sampson" > SAMPSON < / span > I strike quickly, being moved.< br >

< span


class ="character gregory" > GREGORY < / span > But thou art not quickly moved to strike.< br >

< span


class ="character sampson" > SAMPSON < / span > A dog of the house of Montague moves me.< br >

< span


class ="character gregory" > GREGORY < / span > To move is to stir; and to be valiant is to


stand: therefore, if thou art moved, thou runn’st away.< br >
< / div >

< div
id = "right-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > A dog of that house shall move me to stand:
    I
    will
    take
    the
    wall
    of
    any
    man or maid
    of
    Montague’s. < br >

< span


class ="character gregory" > GREGORY < / span > That shows thee a weak slave; for the weakest


goes
to
the
wall. < br >
< span


class ="character sampson" > SAMPSON < / span > True; and therefore women, being the


weaker
vessels, are
ever
thrust
to
the
wall: therefore
I
will
push
Montague’s
men
from the wall, and thrust
his
maids
to
the
wall. < br >
< span


class ="character gregory" > GREGORY < / span > The quarrel is between our masters and us


their
men. < br >
< span


class ="character sampson" > SAMPSON < / span > ’Tis all one, I will show myself a tyrant: when


I
have
fought
with the men, I will be cruel with the
maids, and cut
off
their
heads. < br >
< span


class ="character gregory" > GREGORY < / span > The heads of the maids? < br >

< span


class ="character sampson" > SAMPSON < / span > Ay, the heads of the maids, or their


maidenheads;
take
it in what
sense
thou
wilt. < br >
< span


class ="character gregory" > GREGORY < / span > They must take it in sense that feel it.< br >

< / div >

< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №0.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 48px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
background - color:  # fff;
text - align: center;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card hearts" > ♥ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 24px;
margin - right: 10
px;
}
< / style >
< / head >
< body >
< h1 > Игра
в
карты < / h1 >

< div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card" > &  # x1F0A1; Двойка - крести</div>

< div


class ="card" > &  # x1F0C9; Девятка - черви</div>

< div


class ="card" > &  # x1F0C2; Шестерка - черви</div>

< div


class ="card" > &  # x1F0B4; Туз - буби</div>

< div


class ="card" > &  # x1F0A9; Дама - пики</div>

< div


class ="card" > &  # x1F0AC; Валет - крести</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card" > &  # x1F0C8; Тройка - черви</div>

< div


class ="card" > &  # x1F0A1; Туз - крести</div>

< div


class ="card" > &  # x1F0AA; Десятка - крести</div>

< div


class ="card" > &  # x1F0B7; Король - буби</div>

< div


class ="card" > &  # x1F0A6; Восмёрка - пики</div>

< div


class ="card" > &  # x1F0C7; Тройка - черви</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 36px;
margin - right: 10
px;
display: inline - block;
border: 2
px
solid  # 000;
padding: 10
px;
border - radius: 8
px;
}

/ *Цвета
мастей * /
.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card diamonds" > &  # x1F0A1;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №3.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 190px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }

/ * Дополнительные
стили
для
улучшения
внешнего
вида * /
p
{
    font - size: 18px;
margin - bottom: 10
px;
}

strong
{
    font - size: 24px;
}
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C4;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B3;</div>-->

< !-- < div


class ="card spades" > &  # x1F0AA;</div>-->

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card spades" > &  # x1F0A8;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C3;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B2;</div>-->

< !-- < div


class ="card diamonds" > &  # x1F0A1;</div>-->

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №4.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'Arial', sans - serif; / *Шрифт * /
                                             line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
    color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
    margin - top: 20px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
    margin - right: 15px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
transition: transform
0.3
s
ease - in -out; / *Плавное
изменение
при
наведении * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ * Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "36" >

        <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера * /
function
displayCards(playerId, cards, cardSize)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
    const cardElement = document.createElement('div');
cardElement.style.fontSize = `${cardSize}px`; / * Изменение размера карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Отображаем ранг карты и символ масти * /
cardElement.innerText = `${card.rank} ${suitsSymbols[card.suit]}`;
playerCardsElement.appendChild(cardElement);
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
displayCards('player1Cards', player1Cards, newSize);
displayCards('player2Cards', player2Cards, newSize);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
36
px * /
displayCards('player1Cards', player1Cards, 36);
displayCards('player2Cards', player2Cards, 36);
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №5.
                        <!--  ########################################################################################## -->
                          <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Игра
в
карты < / title >
          < style >
/ *Общие
стили
для
всего
документа * /
body
{
font - family: 'Arial', sans - serif; / *Шрифт * /
line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
margin - top: 20
px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
width: 100
px; / *Ширина
карты * /
height: 150
px; / *Высота
карты * /
margin - right: 15
px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
position: relative; / *Позиционирование
элементов
внутри
карты * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ *Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }

/ * Стили
для
значения
карты * /
.card - value
{
    font - size: 14px; / *Размер
шрифта
значения
карты * /
color:  # 555; /* Цвет текста значения карты */
position: absolute; / *Абсолютное
позиционирование
значения * /
}

/ *Стили
для
левого
верхнего
угла
карты * /
.card - value.top - left
{
    top: 10px; / *Отступ
от
верхнего
края
карты * /
left: 10
px; / *Отступ
от
левого
края
карты * /
}

/ *Стили
для
правого
нижнего
угла
карты * /
.card - value.bottom - right
{
    bottom: 10px; / *Отступ
от
нижнего
края
карты * /
right: 10
px; / *Отступ
от
правого
края
карты * /
}
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "50" >

        <!-- Выпадающий
список
для
выбора
масти -->
< label
for ="suitSelect" > Масть карты:< / label >
                                    < select
id = "suitSelect" >
     < option
value = "all" > Все
масти < / option >
          < option
value = "hearts" > Черви < / option >
                             < option
value = "diamonds" > Бубны < / option >
                               < option
value = "clubs" > Трефы < / option >
                            < option
value = "spades" > Пики < / option >
                            < / select >

                                <!-- Выпадающий
список
для
выбора
ранга
карты -->
< label
for ="rankSelect" > Ранг карты:< / label >
                                   < select
id = "rankSelect" >
     < option
value = "all" > Все
ранги < / option >
          < option
value = "A" > Туз < / option >
                      < option
value = "2" > 2 < / option >
                    < option
value = "3" > 3 < / option >
                    < option
value = "4" > 4 < / option >
                    < option
value = "5" > 5 < / option >
                    < option
value = "6" > 6 < / option >
                    < option
value = "7" > 7 < / option >
                    < option
value = "8" > 8 < / option >
                    < option
value = "9" > 9 < / option >
                    < option
value = "10" > 10 < / option >
                      < option
value = "J" > Валет < / option >
                        < option
value = "Q" > Дама < / option >
                       < option
value = "K" > Король < / option >
                         < / select >

                             <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера, масти
и
ранга * /
function
displayCards(playerId, cards, cardSize, selectedSuit, selectedRank)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
if ((selectedSuit === 'all' | | card.suit == = selectedSuit) & & (selectedRank ==
= 'all' | | card.rank == = selectedRank)) {
const cardElement = document.createElement('div');
cardElement.style.width = `${cardSize}px`; / * Изменение ширины карты * /
cardElement.style.height = `${cardSize * 1.5}px`; / * Изменение высоты карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Добавляем значение карты в верхний и нижний угол карты * /
cardElement.innerHTML = `
< div


class ="card-value top-left" > ${card.rank} < / div >

${suitsSymbols[card.suit]}
< div


class ="card-value bottom-right" > ${card.rank} < / div > `;


playerCardsElement.appendChild(cardElement);
}
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
масти
карты * /
document.getElementById('suitSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = this.value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
ранга
карты * /
document.getElementById('rankSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = this.value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
50
px, мастью
"все"
и
рангом
"все" * /
displayCards('player1Cards', player1Cards, 50, 'all', 'all');
displayCards('player2Cards', player2Cards, 50, 'all', 'all');
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №0.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "en" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Пример
стилей
для
страницы(можно
дополнительно
настроить) * /
body
{
    font - family: Arial, sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
}

img
{
    max - width: 100 %;
height: auto;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Welcome
to
the
HTML
Tutorial < / h1 >
             < / header >

                 < section >
                 < article >
                 < p > This
HTML
tutorial
covers
various
HTML
elements and their
usage.Learn
how
to
create
a
basic
HTML
structure and enhance
your
web
development
skills. < / p >
            < / article >

                < article >
                < h2 > Basic
HTML
Structure < / h2 >
              < p > HTML
documents
consist
of
the
following
elements: < / p >
              < ul >
              < li > < code > & lt;!DOCTYPE
html & gt; < / code > declaration < / li >
                                      < li > < code > & lt;
html & gt; < / code > element < / li >
                                  < li > < code > & lt;
head & gt; < / code > element < / li >
                                  < li > < code > & lt;
title & gt; < / code > element < / li >
                                   < li > < code > & lt;
body & gt; < / code > element < / li >
                                  < / ul >
                                      < / article >

                                          < article >
                                          < h2 > Text
Formatting < / h2 >
               < p > Use
tags
like < code > & lt;
p & gt; < / code >, < code > & lt;
h1 & gt; < / code >, < code > & lt;
strong & gt; < / code >, < code > & lt;
em & gt; < / code >,
etc.,
for text formatting. </ p >
< / article >

< article >
< h2 > Images < / h2 >
< p > Include images using the < code > & lt;img & gt; < / code > tag.< / p >
< img src="https://placekitten.com/400/200" alt="Cute Kitten" >
< / article >

< article >
< h2 > Links < / h2 >
< p > Create hyperlinks with the < code > & lt;a & gt; < / code > tag.< / p >
< p > Visit our < a href="https://example.com" target="_blank" > example website < / a >.< / p >
< / article >
< / section >

< footer >
< p > Explore more HTML elements and features to enhance your web development skills.< / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”.Для выполнения используйте все теги,
которые прошли на занятии.- вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE html >
< html lang="ru" >

< head >
< meta charset="UTF-8" >
< meta name="viewport" content="width=device-width, initial-scale=1.0" >
< title > HTML Tutorial < / title >
< style >
/ * Стили для страницы * /
body {
font-family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц.Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < code >
& lt;
img
src = "image.jpg"
alt = "Описание изображения" & gt;
< / code >
    < / article >

        < article >
        < h2 > Ссылки < / h2 >
                          < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Списки < / h2 >
                          < p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
             < code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
    < / article >

        < article >
        < h2 > Видео < / h2 >
                         < p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
                             < code >
& lt;
iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/ВАШ_KОД"
frameborder = "0"
allowfullscreen & gt; & lt; / iframe & gt;
< / code >
    < / article >
        < / section >

            < footer >
            < p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
                      < / footer >

                          < / body >

                              < / html >
                                  <!--  ########################################################################################## -->
                                    <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №1_1.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "ru" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц. < / p > < p > Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
      < / article >

          < article >
          < h2 > Ссылки < / h2 >
                            < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Стилизованный
текст < / h2 >
          < p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
           < span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "ru" >

< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML
Tutorial < / title >
< style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}

hr
{
    margin: 20px 0;
border: none;
border - top: 2
px
solid  # 007bff;
}
< / style >
< / head >

< body >

< header >
< h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
< / header >

< section >
< article >
< p > < b > HTML < / b > (Hypertext Markup Language) - это
код, который
используется
для
структурирования
и
отображения
веб - страницы
и
её
контента. < / p >
< p > Например, контент
может
быть
структурирован
внутри
множества
параграфов, маркированных
списков
или
с
использованием
изображений
и
таблиц
данных. < / p >
< p > Как
видно
из
названия, этот
туториал
постарается
дать
вам
базовое
понимание
HTML
и
его
функций. < / p >
< / article >
< hr >
< p > < b > HTML < / b > не
является
языком
программирования;
это
язык
разметки, и
используется, чтобы
сообщать
вашему
браузеру, как
отображать
веб - страницы, которые
вы
посещаете. < / p >
< p > Он
может
быть
сложным
или
простым, в
зависимости
от
того, как
хочет
веб - дизайнер. < / p >
< p > HTML
состоит
из
ряда
элементов, которые
вы
используете, чтобы
вкладывать
или
оборачивать
различные
части
контента, чтобы
заставить
контент
отображаться
или
действовать
определённым
образом. < / p >
< p > Ограждающие
теги
могут
сделать
слово
или
изображение
ссылкой
на
что - то
ещё, могут
сделать
слова
курсивом, сделать
шрифт
больше
или
меньше
и
так
далее. < / p >
< hr >

< article >
< h2 > Основная
структура
HTML < / h2 >
< p > HTML - документ
состоит
из
следующих
элементов: < / p >
< ul >
< li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
< li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
< li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
< li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
< li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
< / ul >
< p > Вот
пример
простой
структуры
HTML: < / p >
< code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Форматирование
текста < / h2 >
< p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
< ul >
< li > < code > & lt;
p & gt; < / code >: Параграф < / li >
< li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
< li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
< li > < code > & lt;
em & gt; < / code >: Курсив < / li >
< / ul >
< p > Пример
использования
тегов
форматирования
текста: < / p >
< code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Изображения < / h2 >
< p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
< img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
< / article >

< hr >

< article >
< h2 > Ссылки < / h2 >
< p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
< code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Цветной
текст < / h2 >
< p > Применение
цветового
стиля
к
тексту: < / p >
< code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Стилизованный
текст < / h2 >
< p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
< span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< hr >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !--  ########################################################################################## -->
< !-- HTML  # 4 - Блочная структура элементов. Свойство Display. Размеры - 29.01.2024 -->
< !-- ПРАКТИЧЕСКАЯ
РАБОТА
во
ВРЕМЯ
КЛАССНОГО
УРОКА(ПАРЫ) - 29.01
.2024 -->
< !-- ВАРИАНТ №1 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-rocket float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-refresh float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:
< !DOCTYPE
html >
< html
lang = "en" >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Заголовок
документа:

< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< title >: Заголовок
документа, отображается
во
вкладке
браузера.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Устанавливает
настройки
просмотра
для
устройств
с
различной
шириной
экрана.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
стилей:

< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

Подключаются
стили
Bootstrap
и
шрифтов
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
стилей
внутри
тега < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Здесь
определены
пользовательские
стили
для
элементов.Например, фон
body, стили
карточек.dashboard - card
и
размер
иконок
внутри
этих
карточек.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Основное
содержимое
документа:

< body >
< !-- ... -->
< / body >

Весь
контент
страницы
помещается
внутрь
тега < body >.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Контент
с
использованием
Bootstrap:

Внутри < body > размещаются
карточки, созданные
с
использованием
Bootstrap.
Каждая
карточка
находится
внутри < div


class ="card bg-... dashboard-card" > и имеет свой уникальный


цвет(bg - info, bg - success, bg - warning).
Иконки
и
текст
внутри
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- <!-- ВАРИАНТ №2 --> -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Продукция в пути < / h6 >

< h2


class ="text-right" > < i class ="fa fa-truck float-left" > < / i > < span > 123 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные доставки < span class ="float-right" > 87 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Товары на складе < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cubes float-left" > < / i > < span > 789 < / span > < / h2 >

< p


class ="m-b-0" > Товары в резерве < span class ="float-right" > 456 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-danger dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Возвраты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-undo float-left" > < / i > < span > 10 < / span > < / h2 >

< p


class ="m-b-0" > Обработанные возвраты < span class ="float-right" > 5 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-primary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Поддержка клиентов < / h6 >

< h2


class ="text-right" > < i class ="fa fa-headphones float-left" > < / i > < span > 24 / 7 < / span > < / h2 >

< p


class ="m-b-0" > Запросы < span class ="float-right" > 120 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-secondary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Новые клиенты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-users float-left" > < / i > < span > 50 < / span > < / h2 >

< p


class ="m-b-0" > Клиенты на обслуживании < span class ="float-right" > 20 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< / head >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей
в
теге < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Устанавливаются
пользовательские
стили
для
элементов
страницы, таких
как
body
и.dashboard - card.
Эти
стили
определяют
отступы, фон, тени
и
другие
характеристики.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Содержимое
страницы:

< body >
< div


class ="container" >

< div


class ="row" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >
Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4" > (каждая из них занимает 1 / 3 ширины экрана).


Класс
card
определяет, что
это
карточка
Bootstrap.Классы
bg - info, bg - success, bg - warning, и
т.д.,
устанавливают
цвет
фона
для
каждой
карточки.
Иконки
и
содержимое
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- ВАРИАНТ №3 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
/ *Дополнительные
стили
или
анимации
могут
быть
добавлены
здесь * /
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< div


class ="col-md-4 mb-3" >

< div


class ="card bg-info dashboard-card" id="card1" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Еда и Напитки < / h6 >

< i


class ="fa fa-cutlery float-left" > < / i >

< h2


class ="text-right" > < span id="value1" > 100 < / span > < / h2 >

< p


class ="m-b-0" > Выполненные заказы < span class ="float-right" > 50 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-success dashboard-card" id="card2" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Ноутбуки и ПК < / h6 >

< i


class ="fa fa-laptop float-left" > < / i >

< h2


class ="text-right" > < span id="value2" > 75 < / span > < / h2 >

< p


class ="m-b-0" > Товары в наличии < span class ="float-right" > 25 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-warning dashboard-card" id="card3" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Транспорт < / h6 >

< i


class ="fa fa-car float-left" > < / i >

< h2


class ="text-right" > < span id="value3" > 50 < / span > < / h2 >

< p


class ="m-b-0" > Автомобили в наличии < span class ="float-right" > 10 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-danger dashboard-card" id="card4" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Смартфоны и Гаджеты < / h6 >

< i


class ="fa fa-mobile float-left" > < / i >

< h2


class ="text-right" > < span id="value4" > 120 < / span > < / h2 >

< p


class ="m-b-0" > Гаджеты в наличии < span class ="float-right" > 80 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-primary dashboard-card" id="card5" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Доставка < / h6 >

< i


class ="fa fa-truck float-left" > < / i >

< h2


class ="text-right" > < span id="value5" > 60 < / span > < / h2 >

< p


class ="m-b-0" > Заказы в пути < span class ="float-right" > 400 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-secondary dashboard-card" id="card6" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Одежда и Аксессуары < / h6 >

< i


class ="fa fa-shopping-bag float-left" > < / i >

< h2


class ="text-right" > < span id="value6" > 30 < / span > < / h2 >

< p


class ="m-b-0" > Модные новинки 2024 < span class ="float-right" > 20 000 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >
< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >
    < / body >
        < / html >

            <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "utf-8" >
          < title > Градиентные
карточки
для
панели
управления < / title >
               < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          < / head >

              <!DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          <!-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
       < link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
       <!-- Стили -->
< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    < meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей:

< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    Содержимое
страницы:

< body >
  < div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >

Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4 mb-3" > (каждая из них занимает 1 / 3 ширины экрана).


Индивидуальные
идентификаторы(id)
присваиваются
каждой
карточке(card1, card2, и
т.д.).
Иконки
и
содержимое
карточек
заполняются
данными.Обратите
внимание
на
использование
спанов
с
id
для
отображения
и
обновления
случайных
значений.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
JavaScript - скрипт:

< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
// Обновление
случайных
значений
внутри
карточек
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >

    Создаются
две
функции: getRandomInt
для
получения
случайного
целого
числа
в
заданном
диапазоне
и
updateValues
для
обновления
значений
в
карточках.
setInterval(updateValues, 5000): Запуск
функции
обновления
значений
каждые
5
секунд.
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
ПРИСУЦТВУЮТ
НЕ
ВСЕ
ПРАКТИЧЕСКИЕ
И
НЕ
ВСЕ
ДОМАШНИЕ
ЗАДАНИЯ!
ОНИ
В
ОТДЕЛЬНЫХ
ФАЙЛАХ.html
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !--  ########################################################################################## -->
   <!-- HTML  # 1 Recap + Вводный ПЕРВЫЙ УРОК - 08.01.2024 -->

     <!-- < html >
       < head >

       < meta
charset = "UTF-8" / >
          < title > Hello
HTML < / title >
         < / head >
             < body >

             hello
world
Привет
Мир

< img
src = "https://pngicon.ru/file/uploads/serdce.png"
height = "150px"
width = "150px" / >
        < h1 > Вау - я
программист < / h1 >
                < / body >
                    < / html > -->
< !--  ########################################################################################## -->
   <!--  ########################################################################################## -->
     <!--  ########################################################################################## -->

       <!-- 2024
_HTML_LESSON_№2
_Структура_HTML.Теги_Атрибуты - 15.01
.2024 -->

< !--

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 2: Структура
HTML.Теги.Атрибуты.Правила
языка
разметки
-->

< !DOCTYPE
html >
< html
lang = "ru" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Страница
с
текстом
и
заголовками < / title >
                < / head >
                    < body >

                    <!-- 1.
Создать
страницу
с
текстом
по
примеру
на
картинке. -->

< h2 >
  Стихотворение
  < / h2 >
      < p >

      Мириады
маленьких
дел < br >
Пьют
по
капле
гаснущий
день, < br >
        А < i > дела
большие < / i > сушит
жажда. < br >
Оставляя
все
на «потом», < br >
              Прозреваем
задним
числом, < br >
          Только
год < b > не
повторится < / b > дважды. < br >
               < br >
               И.Тальков

               < / p >

                   <!-- 2.
Повторите
по
данному
образцу: -->
< h1 > Это
заголовок < / h1 >
              < h2 > Это
заголовок < / h2 >
              < h3 > Это
заголовок < / h3 >
              < h4 > Это
заголовок < / h4 >

              < p > Это < b > абзац < / b >.< / p >
                                                < p > Это
еще < i > абзац < / i >.< / p >

                            < h1 > < i > Это
заголовок
h1 < / i > < / h1 >

               < / body >
                   < / html >

                       <!--  ########################################################################################## -->
                         <!--  ########################################################################################## -->
                           <!--  ########################################################################################## -->

03
_Введение
в
CSS.Форматирование
текста
при
помощи
CSS - 22.01
.2024
Классная
работа

<!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < link
rel = "preconnect"
href = "https://fonts.googleapis.com" >
       < link
rel = "preconnect"
href = "https://fonts.gstatic.com"
crossorigin >
< link
href = "https://fonts.googleapis.com/css2?family=Rubik+Burned&display=swap"
rel = "stylesheet" >

      < meta
charset = "UTF-8" >
          < title > Lesson
2 < / title >
      < script >
/ *легкий
init
js * /
< / script >
    < meta
description = "Краткое описание содержимого/контента страницы" / >
              < meta
keywords = "купить, по акции, новый" / >
           < style >
           html, body
{font - size: 10px;
font - family: "Rubik Burned", system - ui;}
h1
{font - size: 4rem}
h2
{font - size: 3rem}
h3
{font - size: 2rem}
h4
{font - size: 1rem}
< / style >
    < link
rel = "stylesheet"
href = "./style.css" / >
       < script
src = "https://cdn.tailwindcss.com" > < / script >
                                          < / head >
                                              < body >
                                              < div
style = "display: flex" >
        < img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
      < img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
      < img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
      < / div >
          < div >
          < h1


class ="heading-of-section text-red text-krasnii" > Красный < / h1 >

< p


class ="text-orange" > Оранжевый < / p >

< div >
< p


class ="button default" > Нажми меня < / p >

< div


class ="button alert" > Нажми меня < / div >

< button


class ="button alert" alt="Для оформления заказа нажмите эту кнопку" > Нажми меня < / button >

< / div >
< div >
< div


class ="display-inline rounded-md shadow-md padding-md bg-gray" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-red" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-blue" > Нажми меня < / div >

< / div >
< / div >
< b > Жирный
текст < / b >
< h1
style = "text-decoration: line-through" > Панды, наконец - то
полетят
на
Марс < / h1 >
< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3


class ="heading" > Hello world < / h3 >

< h4


class ="text-red" > Hello world < / h4 >

< h5 > Hello
world < / h5 >
< h6


class ="bg-blue" > Hello world < / h6 >

< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3 > Hello
world < / h3 >
< h4
id = "heading-h4-1" > Hello
world < / h4 >
< h4
id = "heading-h4-2" > Hello
world < / h4 >
< h4 > Hello
world < / h4 >
< h3


class ="text-lg text-rose-500 bg-gray-100 rounded-md" > Hello world < / h3 >

< h3 > Hello
world < / h3 >

< p >
paragraph < b > lorem < / b >
< br / >
< br / >
< br / >
< br / >
< br / >
ipsum
set < i > < b > ВАЖНО! < / b > < / i > dolor
< br / >
ipsum
set < strong > < em > Кликни & rarr; < / em > < / strong > dolor
< br / >
2 * 2
< br / >
2
x2
< br / >
2 & cross;
2
< / p >
< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
22.01
.2024 - Домашняя
и
практическая - работы.
< !--  ########################################################################################## -->
Практическая
работа

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №3
Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Создать
HTML - страницу “Vehicle”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

2.
Создать
HTML - страницу “Lorem
Ipsum”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

3.
Создать
HTML - страницу “Математические
формулы”

Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.
Используйте
здесь
теги
h1 - h6, span, p, sup, sub
и
спецсимволы.
< !--  ########################################################################################## -->

Шрифты
находятся
в
папке
Материалы
к
занятию
в
архиве
Fonts.

1.
Пример
подключение
шрифтов
через
ссылку
2.
Пример
подключения
через
локальные
файлы
шрифтов
< !--  ########################################################################################## -->

№1

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Двигатель < / title >
< style >
body
{
    font - family: 'Tahoma', serif;
line - height: 1.6;
margin: 18
px;
}

h1
{
    color:  # 00c264;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align: justify;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.smaller - text
{
    font - size: 58 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
< / style >

< / head >
< body >
< h1


class ="larger-font2" >


Двигатель
< / h1 >
< !-- Текст
о
двигателе
с
сайта
https: // ru.wikipedia.org / wiki / Двигатель -->

< p


class ="larger-font" >

< b > Дви́гатель < / b > — устройство, преобразующее
какой - либо
вид < a
href = "https://ru.wikipedia.org/wiki/Энергия"


class ="no-underline" > энергии < / a >


в < a
href = "https://ru.wikipedia.org/wiki/Механическая_энергия"


class ="no-underline" > механическую < / a > работу.


Термин < b > мотор < / b > < a
href = "https://ru.wikipedia.org/wiki/Заимствование"


class ="no-underline" > заимствован < / a > в


первой
половине
XIX
века
из
немецкого
языка < b


class ="smaller-text colored-letter" >[1] < / b >


(нем.Motor — «двигатель», от лат.mōtor — «приводящий в движение»)
и
преимущественно
им
называют
электрические
двигатели
и
двигатели
внутреннего
сгорания < b


class ="smaller-text colored-letter" >


[2] < / b >.
< / p >
< p


class ="larger-font" >


Двигатели
подразделяют
на
первичные
и
вторичные.
К
первичным
относят
непосредственно
преобразующие
природные
энергетические
ресурсы
в
механическую
работу,
а
ко
вторичным — преобразующие
энергию, выработанную
или
накопленную
другими
источниками < b


class ="smaller-text colored-letter" >


[3] < / b >.
< / p >
< hr


class ="subtle-line" >

< p >
< h4 > Примечания < / h4 >
< h5 > [1] < / h5 > Шанский
Н.М., Боброва
Т.А.Кот // Школьный
этимологический
словарь
русского
языка.
Происхождение
слов. — 7 - е
изд., стереотип.. — М.: Дрофа, 2004. — 398, [2]
с.
< h5 > [2] < / h5 > Крысин
Л.П.Мотор // Толковый
словарь
иноязычных
слов. — М.: Эксмо, 2008. — 944
с. —
(Библиотека словарей).
< h5 > [3] < / h5 > Definition
of
motor | Dictionary.com(англ.).www.dictionary.com.
Дата
обращения: 27
января
2022.
Архивировано
27
января
2022
года.
< / p >
< / body >
< / html >
< !--  ########################################################################################## -->

№2

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Lorem
Ipsum < / title >
< style >

h1
{
    color:  # 5d0080;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align0: justify;
}
.left - aligned1
{
    text - align: left;
}

.right - aligned2
{
    text - align: right;
}

.custom - font
{
    font - family: "Courier New", monospace;
}
.custom - font2
{
    font - family: "mv boli", monospace;
}
.custom - font3
{
    font - family: "segoe print", monospace;
}
.custom - font4
{
    font - family: "LUCIDA SANS CONSOLE", monospace;
}
.custom - font5
{
    font - family: "GABRIOLA", monospace;
}
.custom - font6
{
    font - family: "tempus sans itc", monospace;
}
.bold - text
{
    font - weight: bold;
}

.italic - text
{
    font - style: italic;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.larger - font3
{
    font - size: 50px;
}
.centered - text
{
    text - align: center;
}

.centered - container
{
    margin: 0 auto;
width: 50 %; / *Измените
на
желаемую
ширину * /
}

.smaller - text
{
    font - size: 95 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text2
{
    font - size: 80 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text3
{
    font - size: 115 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text4
{
    font - size: 15 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.subtle - line2
{
    border: 1px solid  # f2f2f2; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
.custom - color1
{
    color:  # 001fb8; /* Замените на желаемый цвет */
}
.custom - color2
{
    color:  # 6e6e6e; /* Замените на желаемый цвет */
}
.page - container
{
    margin: 20px; / *Замените
на
желаемое
значение
отступа * /
}
< / style >
< / head >
< body >
< h1


class ="centered-text larger-font3 custom-font4" > Lorem Ipsum < / h1 >

< div


class ="right-aligned2 custom-color1 custom-font3 italic-text smaller-text" >

< p > "Neque porro quisquam est qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit...". < / p >
< / div >
< div


class ="right-aligned2 custom-font2 custom-color2 italic-text smaller-text2 larger-font" >

< p > "Нет никого, кто любил бы боль саму по себе, кто искал бы её и кто хотел бы иметь её
просто
потому, что
это
боль..
".</p>
< / div >
< hr


class ="subtle-line2" >

< !-- Пример
текста
Lorem
Ipsum -->
< div


class ="page-container" >

< p


class ="larger-font custom-font" > < b >


Классический
текст
Lorem
Ipsum, используемый
с
XVI
века
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>L</b>orem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod tempor incididunt
ut
labore
et
dolore
magna
aliqua.Ut
enim
ad
minim
veniam, quis
nostrud
exercitation
ullamco
laboris
nisi
ut
aliquip
ex
ea
commodo
consequat.Duis
aute
irure
dolor in reprehenderit in voluptate
velit
esse
cillum
dolore
eu
fugiat
nulla
pariatur.Excepteur
sint
occaecat
cupidatat
non
proident, sunt in culpa
qui
officia
deserunt
mollit
anim
id
est
laborum.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.32
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>S</b>ed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem
aperiam, eaque
ipsa
quae
ab
illo
inventore
veritatis
et
quasi
architecto
beatae
vitae
dicta
sunt
explicabo.
Nemo
enim
ipsam
voluptatem
quia
voluptas
sit
aspernatur
aut
odit
aut
fugit, sed
quia
consequuntur
magni
dolores
eos
qui
ratione
voluptatem
sequi
nesciunt.Neque
porro
quisquam
est, qui
dolorem
ipsum
quia
dolor
sit
amet,
consectetur, adipisci
velit, sed
quia
non
numquam
eius
modi
tempora
incidunt
ut
labore
et
dolore
magnam
aliquam
quaerat
voluptatem.Ut
enim
ad
minima
veniam, quis
nostrum
exercitationem
ullam
corporis
suscipit
laboriosam, nisi
ut
aliquid
ex
ea
commodi
consequatur? Quis
autem
vel
eum
iure
reprehenderit
qui in ea
voluptate
velit
esse
quam
nihil
molestiae
consequatur, vel
illum
qui
dolorem
eum
fugiat
quo
voluptas
nulla
pariatur?"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>B</b>ut I must explain to you how all this mistaken idea of denouncing pleasure and praising pain was born and
I
will
give
you
a
complete
account
of
the
system, and expound
the
actual
teachings
of
the
great
explorer
of
the
truth, the
master - builder
of
human
happiness.No
one
rejects, dislikes, or avoids
pleasure
itself, because
it is
pleasure, but
because
those
who
do
not know
how
to
pursue
pleasure
rationally
encounter
consequences
that
are
extremely
painful.Nor
again is there
anyone
who
loves or pursues or desires
to
obtain
pain
of
itself, because
it
is pain, but
because
occasionally
circumstances
occur in which
toil and pain
can
procure
him
some
great
pleasure.
To
take
a
trivial
example, which
of
us
ever
undertakes
laborious
physical
exercise, except to
obtain
some
advantage
from it? But
who
has
any
right
to
find
fault
with a man who chooses to enjoy a pleasure that has no annoying
consequences, or one
who
avoids
a
pain
that
produces
no
resultant
pleasure?"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.33
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>A</b>t vero eos et accusamus et iusto odio dignissimos ducimus qui blanditiis praesentium voluptatum deleniti
atque
corrupti
quos
dolores
et
quas
molestias
excepturi
sint
occaecati
cupiditate
non
provident, similique
sunt
in culpa
qui
officia
deserunt
mollitia
animi, id
est
laborum
et
dolorum
fuga.Et
harum
quidem
rerum
facilis
est
et
expedita
distinctio.Nam
libero
tempore, cum
soluta
nobis
est
eligendi
optio
cumque
nihil
impedit
quo
minus
id
quod
maxime
placeat
facere
possimus, omnis
voluptas
assumenda
est, omnis
dolor
repellendus.Temporibus
autem
quibusdam
et
aut
officiis
debitis
aut
rerum
necessitatibus
saepe
eveniet
ut
et
voluptates
repudiandae
sint
et
molestiae
non
recusandae.Itaque
earum
rerum
hic
tenetur
a
sapiente
delectus, ut
aut
reiciendis
voluptatibus
maiores
alias
consequatur
aut
perferendis
doloribus
asperiores
repellat.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>O</b>n the other hand, we denounce with righteous indignation and dislike men who are so beguiled and demoralized
by
the
charms
of
pleasure
of
the
moment, so
blinded
by
desire, that
they
cannot
foresee
the
pain and trouble
that
are
bound
to
ensue; and equal
blame
belongs
to
those
who
fail in their
duty
through
weakness
of
will, which is the
same as saying
through
shrinking
from toil and pain.These
cases
are
perfectly
simple and easy
to
distinguish.
In
a
free
hour, when
our
power
of
choice is untrammelled and when
nothing
prevents
our
being
able
to
do
what
we
like
best, every
pleasure is to
be
welcomed and every
pain
avoided.But in certain
circumstances and owing
to
the
claims
of
duty or the
obligations
of
business
it
will
frequently
occur
that
pleasures
have
to
be
repudiated and
annoyances
accepted.The
wise
man
therefore
always
holds in these
matters
to
this
principle
of
selection: he
rejects
pleasures
to
secure
other
greater
pleasures, or else he
endures
pains
to
avoid
worse
pains.
"
< / p >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

№3

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Математические
формулы < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'VERDANA', monospace; / *Устанавливаем
шрифт
для
всего
документа * /
line - height: 1.0; / *Задаем
высоту
строки
текста * /
margin: 70
px; / *Задаем
отступы
для
всего
документа * /
}

/ *Стили
для
заголовка
первого
уровня * /
h1
{
    color:  # 000000; /* Цвет текста */
        text - align: center; / *Выравнивание
текста
по
центру * /
font - size: 50
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
}

/ *Стили
для
заголовка
второго
уровня * /
h2
{
    color:  # 22009e; /* Цвет текста */
        font - size: 35
px; / *Размер
шрифта * /
margin - top: 20
px; / *Верхний
отступ * /
}

/ *Стили
для
подзаголовка * /
h3
{
    color:  # 00ff49; /* Цвет текста */
        font - size: 25
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
font - style: italic; / *Курсив * /
                         text - align: left; / *Выравнивание
текста
влево * /
background - color:  # a300a3b8; /* Цвет подложки */
padding: 10
px; / *Внутренний
отступ * /
border - radius: 45
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 320
px; / *Максимальная
ширина
подложки * /
width: 25 %; / *Ширина
подложки
равна
20 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: left; / *Выравнивание
по
левой
стороне * /
}

/ *Стили
для
обычного
текста
и
вложенных
элементов * /
p, span
{
    text - align: justify; / *Выравнивание
текста
по
ширине
с
обеих
сторон * /
}

/ *Стили
для
верхних
и
нижних
индексов * /
sup, sub
{
    font - size: 80 %; / *Размер
шрифта
для
верхних
и
нижних
индексов * /
}

/ *Отступ
для
блока
с
описанием
формулы * /
.formula - description
{
    margin - left: 20px; / *Отступ
слева * /
}

/ *Стили
для
контейнера
формулы
и
самой
формулы * /
.formula - container,.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
background - color:  # ff00008f; /* Цвет подложки */
padding: 2
px; / *Внутренний
отступ * /
border - radius: 55
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
15
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 670
px; / *Максимальная
ширина
подложки * /
width: 50 %; / *Ширина
подложки
равна
50 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: auto; / *Выравнивание
по
центру * /
}

/ *Стили
для
самой
формулы * /
.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
color:  # 0700ff; /* Цвет формулы */
font - size: 25
px; / *Размер
шрифта
формулы * /
font - style: italic; / *Курсив * /
                         font - weight: bold; / *Жирный
шрифт
формулы * /
}

/ *Стили
для
раздела
с
примерами * /
.example - section
{
    margin - top: 20px; / *Верхний
отступ
раздела
с
примерами * /
}

/ *Стили
для
примеров * /
.example
{
    font - size: 20px; / *Размер
шрифта
примеров * /
color:  # 555; /* Цвет текста примеров */
}
/ *Стили
для
цветной
буквы * /
.colored - letter
{
    color:  # ff0000; /* Цвет буквы */
}
/ *Стили
для
увеличенной
буквы * /
.enlarged - letter
{
    font - size: 24px; / *Размер
шрифта
для
увеличенной
буквы * /
}
/ *Стили
для
выделенных
букв * /
.special - letter
{
    font - family: 'Arial', sans - serif; / *Используемый
шрифт * /
font - style: italic; / *Курсив * /
                         color:  # 00adff; /* Цвет текста */
}
< / style >
< / head >
< body >
< !-- Заголовок
первого
уровня -->
< h1 > Математические
формулы < / h1 >

< !-- Формула
окружности -->
< h2 > Формула
окружности: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > x < sup > 2 < / sup > + y < sup > 2 < / sup > = r < sup > 2 < / sup > < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Эта
формула
описывает
уравнение
окружности
в
декартовой
системе
координат. < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > x < / span > и < span class ="colored-letter enlarged-letter special-letter" > y < / span > - координаты точки на плоскости.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус окружности.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула окружности описывает геометрическую фигуру,


представляющую
собой
множество
точек, равноудаленных
от
центра. < / b > < / p >
< p


class ="formula-description" > < b > Здесь x и y - координаты точек,


а
r - радиус
окружности. < / b > < / p >

< !-- Формула
площади
треугольника -->
< h2 > Формула
площади
треугольника: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > S = 0.5 * a * b * sin(C) < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > S < / span > - площадь треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > a < / span > и < span class ="colored-letter enlarged-letter special-letter" > b < / span > - длины сторон треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > C < / span > - угол между сторонами a и b (в радианах).< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула площади треугольника выражает площадь как половину произведения длин двух сторон на синус угла между ними.< / b > < / p >

< !-- Формула
объема
цилиндра -->
< h2 > Формула
объема
цилиндра: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > V = π * r < sup > 2 < / sup > * h < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > V < / span > - объем цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > π < / span > (пи) - математическая константа, приблизительно равная 3.14159.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус основания цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > h < / span > - высота цилиндра.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула объема цилиндра выражает объем как произведение площади основания на высоту.< / b > < / p >

< !-- Раздел
с
примерами
применения
формул -->
< div


class ="example-section" >

< h2 > Сферы
применения
формул: < / h2 >
< !-- Примеры -->
< p


class ="example" > 1. < b > Геометрия:<

    / b > расчет
геометрических
параметров
фигур. < / p >
< p


class ="example" > 2. < b > Инженерия:<

    / b > проектирование
и
расчеты
конструкций. < / p >
< p


class ="example" > 3. < b > Физика:<

    / b > моделирование
и
анализ
физических
явлений. < / p >
< / div >

< / body >
< / html >

< !--  ########################################################################################## -->

Бонус

1.
Пример
подключение
шрифтов
через
ссылку

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
внешних
шрифтов -->
< link
rel = "stylesheet"
href = "https://fonts.googleapis.com/css?family=YourFontFamily" >

< !-- Ваши
стили -->
< style >
body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< / style >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >

< !--
Замените
"https://fonts.googleapis.com/css?family=YourFontFamily"
на
фактическую
ссылку
на
шрифт, который
вы
хотите
использовать.Например, если
вы
хотите
использовать
шрифт
"Roboto", ссылка
будет
выглядеть
так: "https://fonts.googleapis.com/css?family=Roboto".

Обратите
внимание, что
это
только
пример, и
вам
нужно
заменить
"YourFontFamily"
и
"Your Title"
на
фактические
значения, используемые
в
вашем
проекте.
-->

2.
Пример
подключения
через
локальные
файлы
шрифтов

< !--
Поместите
шрифтовые
файлы
в
ваш
проект:
Сначала
загрузите
файлы
шрифтов(например,.woff,.woff2,.ttf) в
ваш
проект.
Создайте
папку, например, "fonts", и
поместите
файлы
шрифтов
в
неё.

Пример
структуры
проекта:
-->
/ your - project
/ fonts
your - font.woff
your - font.woff2
index.html
style.css
< !--
Добавьте
стили
в
ваш
CSS:
В
вашем
файле
стилей(например, style.css), определите
новый @ font - face
и
используйте
его
в
стилях
элементов.

Пример:
-->

@font - face


{
    font - family: 'YourFontFamily';
src: url('fonts/your-font.woff2')
format('woff2'),
url('fonts/your-font.woff')
format('woff');
/ *Дополнительные
параметры, если
нужны * /
}

body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< !--
Убедитесь, что
путь
к
файлам
шрифтов
в
url
соответствует
структуре
вашего
проекта.
-->
< !--
Подключите
стили
в
HTML:
В
вашем
файле
HTML(например, index.html), убедитесь, что
вы
подключили
ваш
CSS
файл.

Пример:
-->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
локальных
стилей -->
< link
rel = "stylesheet"
href = "style.css" >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >
< !--
Замените
"Your Title"
на
фактический
заголовок
вашей
страницы.

Это
позволит
вам
использовать
локальные
файлы
шрифтов
в
вашем
проекте.Убедитесь, что
пути
указаны
правильно
и
что
браузер
может
найти
и
загрузить
файлы
шрифтов.
-->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Попробуйте
каждый
тег

На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.

К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)

Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!

2.
Создайте
страницу “Romeo and Juliet”

3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.

4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.

< !--  ########################################################################################## -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS

Домашнее
задание №3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
< !--  #### -->
1.
Попробуйте
каждый
тег
< !--  #### -->
На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.
К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)
< !--  #### -->
Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!
< !--  #### -->
2.
Создайте
страницу “Romeo and Juliet”
< !--  #### -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.
< !--  #### -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.
< !--  #### -->

< !--  ########################################################################################## -->
1.
Попробуйте
каждый
тег
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
2.
Создайте
страницу “Romario and Juluya”
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Romeo and Juliet < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
text - align: center;
}

# title {
font - family: 'Georgia', serif;
color:  # FF4500; /* Цвет - оранжевый */
font - style: italic;
font - size: 36
px;
}

# author {
font - size: 18
px;
font - style: italic;
}

hr
{
    width: 50 %;
margin: auto;
margin - top: 20
px;
border: 1
px
solid  # 000;
}

# prologue {
font - size: 24
px;
margin - top: 20
px;
}

# left-text, #right-text {
display: inline - block;
width: 45 %;
vertical - align: top;
}

.act - title
{
    font - size: 28px;
margin - top: 20
px;
}

.scene - title
{
    font - size: 20px;
font - weight: bold;
margin - top: 10
px;
}

.character
{
    font - weight: bold;
}

.dialogue
{
    text - align: left;
margin - top: 10
px;
margin - bottom: 10
px;
}

.romeo
{
    color:  # 0000FF; /* Цвет - синий */
}

.juliet
{
    color:  # FF1493; /* Цвет - розовый */
}

.sampson
{
    color:  # 008000; /* Цвет - зеленый */
}

.gregory
{
    color:  # FFD700; /* Цвет - золотой */
}
< / style >
    < / head >
        < body >
        < div
id = "title" > ROMEO
AND
JULIET < / div >
           < div
id = "author" > By
William
Shakespeare < / div >
                < hr >

                < div
id = "prologue" >
     < div
id = "left-text" >
     Two
households, both
alike in dignity,
In
fair
Verona, where
we
lay
our
scene,
From
ancient
grudge
break
to
new
mutiny,
Where
civil
blood
makes
civil
hands
unclean.
From
forth
the
fatal
loins
of
these
two
foes
A
pair
of
star - cross’d
lovers
take
their
life;
Whole
misadventured
piteous
overthrows
< / div >

< div
id = "right-text" >
Do
with their death bury their parents’ strife.
The
fearful
passage
of
their
death - mark’d
love,
And
the
continuance
of
their
parents’ rage,
Which, but
their
children’s
end, nought
could
remove,
Is
now
the
two
hours’ traffic
of
our
stage;
The
which if you
with patient ears attend,
What
here
shall
miss, our
toil
shall
strive
to
mend.
< / div >
< / div >
< hr >

< div


class ="act-title" > ACT I < / div >

< div


class ="scene-title" > SCENE I.Verona.A public place.< / div >

< div


class ="scene-title" > < i > Enter SAMPSON and GREGORY, of the house


of
CAPULET, armed
with swords and bucklers </ i > < / div >

< div
id = "left-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > Gregory, o’ my word, we’ll not carry coals.< br >

< span


class ="character gregory" > GREGORY < / span > No, for then we should be colliers.< br >

< span


class ="character sampson" > SAMPSON < / span > I mean, an we be in choler, we’ll draw.< br >

< span


class ="character gregory" > GREGORY < / span > Ay, while you live, draw your neck out


o’ the
collar. < br >
< span


class ="character sampson" > SAMPSON < / span > I strike quickly, being moved.< br >

< span


class ="character gregory" > GREGORY < / span > But thou art not quickly moved to strike.< br >

< span


class ="character sampson" > SAMPSON < / span > A dog of the house of Montague moves me.< br >

< span


class ="character gregory" > GREGORY < / span > To move is to stir; and to be valiant is to


stand: therefore, if thou art moved, thou runn’st away.< br >
< / div >

< div
id = "right-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > A dog of that house shall move me to stand:
    I
    will
    take
    the
    wall
    of
    any
    man or maid
    of
    Montague’s. < br >

< span


class ="character gregory" > GREGORY < / span > That shows thee a weak slave; for the weakest


goes
to
the
wall. < br >
< span


class ="character sampson" > SAMPSON < / span > True; and therefore women, being the


weaker
vessels, are
ever
thrust
to
the
wall: therefore
I
will
push
Montague’s
men
from the wall, and thrust
his
maids
to
the
wall. < br >
< span


class ="character gregory" > GREGORY < / span > The quarrel is between our masters and us


their
men. < br >
< span


class ="character sampson" > SAMPSON < / span > ’Tis all one, I will show myself a tyrant: when


I
have
fought
with the men, I will be cruel with the
maids, and cut
off
their
heads. < br >
< span


class ="character gregory" > GREGORY < / span > The heads of the maids? < br >

< span


class ="character sampson" > SAMPSON < / span > Ay, the heads of the maids, or their


maidenheads;
take
it in what
sense
thou
wilt. < br >
< span


class ="character gregory" > GREGORY < / span > They must take it in sense that feel it.< br >

< / div >

< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №0.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 48px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
background - color:  # fff;
text - align: center;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card hearts" > ♥ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 24px;
margin - right: 10
px;
}
< / style >
< / head >
< body >
< h1 > Игра
в
карты < / h1 >

< div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card" > &  # x1F0A1; Двойка - крести</div>

< div


class ="card" > &  # x1F0C9; Девятка - черви</div>

< div


class ="card" > &  # x1F0C2; Шестерка - черви</div>

< div


class ="card" > &  # x1F0B4; Туз - буби</div>

< div


class ="card" > &  # x1F0A9; Дама - пики</div>

< div


class ="card" > &  # x1F0AC; Валет - крести</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card" > &  # x1F0C8; Тройка - черви</div>

< div


class ="card" > &  # x1F0A1; Туз - крести</div>

< div


class ="card" > &  # x1F0AA; Десятка - крести</div>

< div


class ="card" > &  # x1F0B7; Король - буби</div>

< div


class ="card" > &  # x1F0A6; Восмёрка - пики</div>

< div


class ="card" > &  # x1F0C7; Тройка - черви</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 36px;
margin - right: 10
px;
display: inline - block;
border: 2
px
solid  # 000;
padding: 10
px;
border - radius: 8
px;
}

/ *Цвета
мастей * /
.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card diamonds" > &  # x1F0A1;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №3.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 190px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }

/ * Дополнительные
стили
для
улучшения
внешнего
вида * /
p
{
    font - size: 18px;
margin - bottom: 10
px;
}

strong
{
    font - size: 24px;
}
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C4;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B3;</div>-->

< !-- < div


class ="card spades" > &  # x1F0AA;</div>-->

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card spades" > &  # x1F0A8;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C3;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B2;</div>-->

< !-- < div


class ="card diamonds" > &  # x1F0A1;</div>-->

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №4.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'Arial', sans - serif; / *Шрифт * /
                                             line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
    color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
    margin - top: 20px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
    margin - right: 15px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
transition: transform
0.3
s
ease - in -out; / *Плавное
изменение
при
наведении * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ * Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "36" >

        <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера * /
function
displayCards(playerId, cards, cardSize)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
    const cardElement = document.createElement('div');
cardElement.style.fontSize = `${cardSize}px`; / * Изменение размера карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Отображаем ранг карты и символ масти * /
cardElement.innerText = `${card.rank} ${suitsSymbols[card.suit]}`;
playerCardsElement.appendChild(cardElement);
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
displayCards('player1Cards', player1Cards, newSize);
displayCards('player2Cards', player2Cards, newSize);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
36
px * /
displayCards('player1Cards', player1Cards, 36);
displayCards('player2Cards', player2Cards, 36);
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №5.
                        <!--  ########################################################################################## -->
                          <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Игра
в
карты < / title >
          < style >
/ *Общие
стили
для
всего
документа * /
body
{
font - family: 'Arial', sans - serif; / *Шрифт * /
line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
margin - top: 20
px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
width: 100
px; / *Ширина
карты * /
height: 150
px; / *Высота
карты * /
margin - right: 15
px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
position: relative; / *Позиционирование
элементов
внутри
карты * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ *Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }

/ * Стили
для
значения
карты * /
.card - value
{
    font - size: 14px; / *Размер
шрифта
значения
карты * /
color:  # 555; /* Цвет текста значения карты */
position: absolute; / *Абсолютное
позиционирование
значения * /
}

/ *Стили
для
левого
верхнего
угла
карты * /
.card - value.top - left
{
    top: 10px; / *Отступ
от
верхнего
края
карты * /
left: 10
px; / *Отступ
от
левого
края
карты * /
}

/ *Стили
для
правого
нижнего
угла
карты * /
.card - value.bottom - right
{
    bottom: 10px; / *Отступ
от
нижнего
края
карты * /
right: 10
px; / *Отступ
от
правого
края
карты * /
}
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "50" >

        <!-- Выпадающий
список
для
выбора
масти -->
< label
for ="suitSelect" > Масть карты:< / label >
                                    < select
id = "suitSelect" >
     < option
value = "all" > Все
масти < / option >
          < option
value = "hearts" > Черви < / option >
                             < option
value = "diamonds" > Бубны < / option >
                               < option
value = "clubs" > Трефы < / option >
                            < option
value = "spades" > Пики < / option >
                            < / select >

                                <!-- Выпадающий
список
для
выбора
ранга
карты -->
< label
for ="rankSelect" > Ранг карты:< / label >
                                   < select
id = "rankSelect" >
     < option
value = "all" > Все
ранги < / option >
          < option
value = "A" > Туз < / option >
                      < option
value = "2" > 2 < / option >
                    < option
value = "3" > 3 < / option >
                    < option
value = "4" > 4 < / option >
                    < option
value = "5" > 5 < / option >
                    < option
value = "6" > 6 < / option >
                    < option
value = "7" > 7 < / option >
                    < option
value = "8" > 8 < / option >
                    < option
value = "9" > 9 < / option >
                    < option
value = "10" > 10 < / option >
                      < option
value = "J" > Валет < / option >
                        < option
value = "Q" > Дама < / option >
                       < option
value = "K" > Король < / option >
                         < / select >

                             <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера, масти
и
ранга * /
function
displayCards(playerId, cards, cardSize, selectedSuit, selectedRank)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
if ((selectedSuit === 'all' | | card.suit == = selectedSuit) & & (selectedRank ==
= 'all' | | card.rank == = selectedRank)) {
const cardElement = document.createElement('div');
cardElement.style.width = `${cardSize}px`; / * Изменение ширины карты * /
cardElement.style.height = `${cardSize * 1.5}px`; / * Изменение высоты карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Добавляем значение карты в верхний и нижний угол карты * /
cardElement.innerHTML = `
< div


class ="card-value top-left" > ${card.rank} < / div >

${suitsSymbols[card.suit]}
< div


class ="card-value bottom-right" > ${card.rank} < / div > `;


playerCardsElement.appendChild(cardElement);
}
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
масти
карты * /
document.getElementById('suitSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = this.value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
ранга
карты * /
document.getElementById('rankSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = this.value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
50
px, мастью
"все"
и
рангом
"все" * /
displayCards('player1Cards', player1Cards, 50, 'all', 'all');
displayCards('player2Cards', player2Cards, 50, 'all', 'all');
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №0.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "en" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Пример
стилей
для
страницы(можно
дополнительно
настроить) * /
body
{
    font - family: Arial, sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
}

img
{
    max - width: 100 %;
height: auto;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Welcome
to
the
HTML
Tutorial < / h1 >
             < / header >

                 < section >
                 < article >
                 < p > This
HTML
tutorial
covers
various
HTML
elements and their
usage.Learn
how
to
create
a
basic
HTML
structure and enhance
your
web
development
skills. < / p >
            < / article >

                < article >
                < h2 > Basic
HTML
Structure < / h2 >
              < p > HTML
documents
consist
of
the
following
elements: < / p >
              < ul >
              < li > < code > & lt;!DOCTYPE
html & gt; < / code > declaration < / li >
                                      < li > < code > & lt;
html & gt; < / code > element < / li >
                                  < li > < code > & lt;
head & gt; < / code > element < / li >
                                  < li > < code > & lt;
title & gt; < / code > element < / li >
                                   < li > < code > & lt;
body & gt; < / code > element < / li >
                                  < / ul >
                                      < / article >

                                          < article >
                                          < h2 > Text
Formatting < / h2 >
               < p > Use
tags
like < code > & lt;
p & gt; < / code >, < code > & lt;
h1 & gt; < / code >, < code > & lt;
strong & gt; < / code >, < code > & lt;
em & gt; < / code >,
etc.,
for text formatting. </ p >
< / article >

< article >
< h2 > Images < / h2 >
< p > Include images using the < code > & lt;img & gt; < / code > tag.< / p >
< img src="https://placekitten.com/400/200" alt="Cute Kitten" >
< / article >

< article >
< h2 > Links < / h2 >
< p > Create hyperlinks with the < code > & lt;a & gt; < / code > tag.< / p >
< p > Visit our < a href="https://example.com" target="_blank" > example website < / a >.< / p >
< / article >
< / section >

< footer >
< p > Explore more HTML elements and features to enhance your web development skills.< / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”.Для выполнения используйте все теги,
которые прошли на занятии.- вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE html >
< html lang="ru" >

< head >
< meta charset="UTF-8" >
< meta name="viewport" content="width=device-width, initial-scale=1.0" >
< title > HTML Tutorial < / title >
< style >
/ * Стили для страницы * /
body {
font-family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц.Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < code >
& lt;
img
src = "image.jpg"
alt = "Описание изображения" & gt;
< / code >
    < / article >

        < article >
        < h2 > Ссылки < / h2 >
                          < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Списки < / h2 >
                          < p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
             < code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
    < / article >

        < article >
        < h2 > Видео < / h2 >
                         < p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
                             < code >
& lt;
iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/ВАШ_KОД"
frameborder = "0"
allowfullscreen & gt; & lt; / iframe & gt;
< / code >
    < / article >
        < / section >

            < footer >
            < p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
                      < / footer >

                          < / body >

                              < / html >
                                  <!--  ########################################################################################## -->
                                    <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №1_1.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "ru" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц. < / p > < p > Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
      < / article >

          < article >
          < h2 > Ссылки < / h2 >
                            < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Стилизованный
текст < / h2 >
          < p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
           < span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "ru" >

< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML
Tutorial < / title >
< style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}

hr
{
    margin: 20px 0;
border: none;
border - top: 2
px
solid  # 007bff;
}
< / style >
< / head >

< body >

< header >
< h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
< / header >

< section >
< article >
< p > < b > HTML < / b > (Hypertext Markup Language) - это
код, который
используется
для
структурирования
и
отображения
веб - страницы
и
её
контента. < / p >
< p > Например, контент
может
быть
структурирован
внутри
множества
параграфов, маркированных
списков
или
с
использованием
изображений
и
таблиц
данных. < / p >
< p > Как
видно
из
названия, этот
туториал
постарается
дать
вам
базовое
понимание
HTML
и
его
функций. < / p >
< / article >
< hr >
< p > < b > HTML < / b > не
является
языком
программирования;
это
язык
разметки, и
используется, чтобы
сообщать
вашему
браузеру, как
отображать
веб - страницы, которые
вы
посещаете. < / p >
< p > Он
может
быть
сложным
или
простым, в
зависимости
от
того, как
хочет
веб - дизайнер. < / p >
< p > HTML
состоит
из
ряда
элементов, которые
вы
используете, чтобы
вкладывать
или
оборачивать
различные
части
контента, чтобы
заставить
контент
отображаться
или
действовать
определённым
образом. < / p >
< p > Ограждающие
теги
могут
сделать
слово
или
изображение
ссылкой
на
что - то
ещё, могут
сделать
слова
курсивом, сделать
шрифт
больше
или
меньше
и
так
далее. < / p >
< hr >

< article >
< h2 > Основная
структура
HTML < / h2 >
< p > HTML - документ
состоит
из
следующих
элементов: < / p >
< ul >
< li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
< li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
< li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
< li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
< li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
< / ul >
< p > Вот
пример
простой
структуры
HTML: < / p >
< code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Форматирование
текста < / h2 >
< p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
< ul >
< li > < code > & lt;
p & gt; < / code >: Параграф < / li >
< li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
< li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
< li > < code > & lt;
em & gt; < / code >: Курсив < / li >
< / ul >
< p > Пример
использования
тегов
форматирования
текста: < / p >
< code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Изображения < / h2 >
< p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
< img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
< / article >

< hr >

< article >
< h2 > Ссылки < / h2 >
< p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
< code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Цветной
текст < / h2 >
< p > Применение
цветового
стиля
к
тексту: < / p >
< code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Стилизованный
текст < / h2 >
< p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
< span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< hr >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !--  ########################################################################################## -->
< !-- HTML  # 4 - Блочная структура элементов. Свойство Display. Размеры - 29.01.2024 -->
< !-- ПРАКТИЧЕСКАЯ
РАБОТА
во
ВРЕМЯ
КЛАССНОГО
УРОКА(ПАРЫ) - 29.01
.2024 -->
< !-- ВАРИАНТ №1 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-rocket float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-refresh float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:
< !DOCTYPE
html >
< html
lang = "en" >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Заголовок
документа:

< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< title >: Заголовок
документа, отображается
во
вкладке
браузера.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Устанавливает
настройки
просмотра
для
устройств
с
различной
шириной
экрана.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
стилей:

< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

Подключаются
стили
Bootstrap
и
шрифтов
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
стилей
внутри
тега < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Здесь
определены
пользовательские
стили
для
элементов.Например, фон
body, стили
карточек.dashboard - card
и
размер
иконок
внутри
этих
карточек.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Основное
содержимое
документа:

< body >
< !-- ... -->
< / body >

Весь
контент
страницы
помещается
внутрь
тега < body >.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Контент
с
использованием
Bootstrap:

Внутри < body > размещаются
карточки, созданные
с
использованием
Bootstrap.
Каждая
карточка
находится
внутри < div


class ="card bg-... dashboard-card" > и имеет свой уникальный


цвет(bg - info, bg - success, bg - warning).
Иконки
и
текст
внутри
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- <!-- ВАРИАНТ №2 --> -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Продукция в пути < / h6 >

< h2


class ="text-right" > < i class ="fa fa-truck float-left" > < / i > < span > 123 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные доставки < span class ="float-right" > 87 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Товары на складе < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cubes float-left" > < / i > < span > 789 < / span > < / h2 >

< p


class ="m-b-0" > Товары в резерве < span class ="float-right" > 456 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-danger dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Возвраты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-undo float-left" > < / i > < span > 10 < / span > < / h2 >

< p


class ="m-b-0" > Обработанные возвраты < span class ="float-right" > 5 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-primary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Поддержка клиентов < / h6 >

< h2


class ="text-right" > < i class ="fa fa-headphones float-left" > < / i > < span > 24 / 7 < / span > < / h2 >

< p


class ="m-b-0" > Запросы < span class ="float-right" > 120 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-secondary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Новые клиенты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-users float-left" > < / i > < span > 50 < / span > < / h2 >

< p


class ="m-b-0" > Клиенты на обслуживании < span class ="float-right" > 20 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< / head >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей
в
теге < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Устанавливаются
пользовательские
стили
для
элементов
страницы, таких
как
body
и.dashboard - card.
Эти
стили
определяют
отступы, фон, тени
и
другие
характеристики.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Содержимое
страницы:

< body >
< div


class ="container" >

< div


class ="row" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >
Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4" > (каждая из них занимает 1 / 3 ширины экрана).


Класс
card
определяет, что
это
карточка
Bootstrap.Классы
bg - info, bg - success, bg - warning, и
т.д.,
устанавливают
цвет
фона
для
каждой
карточки.
Иконки
и
содержимое
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- ВАРИАНТ №3 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
/ *Дополнительные
стили
или
анимации
могут
быть
добавлены
здесь * /
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< div


class ="col-md-4 mb-3" >

< div


class ="card bg-info dashboard-card" id="card1" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Еда и Напитки < / h6 >

< i


class ="fa fa-cutlery float-left" > < / i >

< h2


class ="text-right" > < span id="value1" > 100 < / span > < / h2 >

< p


class ="m-b-0" > Выполненные заказы < span class ="float-right" > 50 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-success dashboard-card" id="card2" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Ноутбуки и ПК < / h6 >

< i


class ="fa fa-laptop float-left" > < / i >

< h2


class ="text-right" > < span id="value2" > 75 < / span > < / h2 >

< p


class ="m-b-0" > Товары в наличии < span class ="float-right" > 25 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-warning dashboard-card" id="card3" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Транспорт < / h6 >

< i


class ="fa fa-car float-left" > < / i >

< h2


class ="text-right" > < span id="value3" > 50 < / span > < / h2 >

< p


class ="m-b-0" > Автомобили в наличии < span class ="float-right" > 10 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-danger dashboard-card" id="card4" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Смартфоны и Гаджеты < / h6 >

< i


class ="fa fa-mobile float-left" > < / i >

< h2


class ="text-right" > < span id="value4" > 120 < / span > < / h2 >

< p


class ="m-b-0" > Гаджеты в наличии < span class ="float-right" > 80 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-primary dashboard-card" id="card5" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Доставка < / h6 >

< i


class ="fa fa-truck float-left" > < / i >

< h2


class ="text-right" > < span id="value5" > 60 < / span > < / h2 >

< p


class ="m-b-0" > Заказы в пути < span class ="float-right" > 400 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-secondary dashboard-card" id="card6" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Одежда и Аксессуары < / h6 >

< i


class ="fa fa-shopping-bag float-left" > < / i >

< h2


class ="text-right" > < span id="value6" > 30 < / span > < / h2 >

< p


class ="m-b-0" > Модные новинки 2024 < span class ="float-right" > 20 000 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >
< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >
    < / body >
        < / html >

            <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "utf-8" >
          < title > Градиентные
карточки
для
панели
управления < / title >
               < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          < / head >

              <!DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          <!-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
       < link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
       <!-- Стили -->
< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    < meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей:

< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    Содержимое
страницы:

< body >
  < div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >

Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4 mb-3" > (каждая из них занимает 1 / 3 ширины экрана).


Индивидуальные
идентификаторы(id)
присваиваются
каждой
карточке(card1, card2, и
т.д.).
Иконки
и
содержимое
карточек
заполняются
данными.Обратите
внимание
на
использование
спанов
с
id
для
отображения
и
обновления
случайных
значений.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
JavaScript - скрипт:

< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
// Обновление
случайных
значений
внутри
карточек
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >

    Создаются
две
функции: getRandomInt
для
получения
случайного
целого
числа
в
заданном
диапазоне
и
updateValues
для
обновления
значений
в
карточках.
setInterval(updateValues, 5000): Запуск
функции
обновления
значений
каждые
5
секунд.
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS

Домашнее
задание №4: Блочная
структура
элементов.Свойство
Display.Размеры
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
1.
Создать
страницу, как
на
примере: "Флексагон"
         <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !DOCTYPE
html >
< html >
< head >
< title > Флексагон < / title >
                        < meta
charset = "utf-8" >
          < link
rel = "stylesheet"
href = "style.css" >
       < style >
       body
{
font - family: Arial, Verdana, sans - serif; / *Семейство
шрифтов * /
font - size: 20
pt; / *Размер
основного
шрифта
в
пунктах * /
background - color:  # f0f0f0; /* Цвет фона веб-страницы */
color:  # 333; /* Цвет основного текста */
}
h1
{
color:  # a52a2a; /* Цвет заголовка */
font - size: 75
pt; / *Размер
шрифта
в
пунктах * /
font - family: Georgia, Times, serif; / *Семейство
шрифтов * /
font - weight: normal; / *Нормальное
начертание
текста * /
}
p
{
text - align: justify; / *Выравнивание
по
ширине * /
margin - left: 75
px; / *Отступ
слева
в
пикселах * /
margin - right: 10
px; / *Отступ
справа
в
пикселах * /
border - left: 1
px
solid  # 999; /* Параметры линии слева */
border - bottom: 1
px
solid  # 999; /* Параметры линии снизу */
padding - left: 10
px; / *Отступ
от
линии
слева
до
текста * /
padding - bottom: 10
px; / *Отступ
от
линии
снизу
до
текста * /
}
< / style >
    < / head >
        < body >
        < h1 > Флексагон < / h1 >
                             < p > < b > Флексагон
представляет
собой
бумажную
фигуру, которая
имеет
три
и
более
стороны.Поначалу
кажется, что
это
невозможно, но
вспомните
ленту
Мёбиуса, она
ведь
имеет
всего
одну
сторону, в
отличие
от
листа
бумаги,
и, тем
не
менее, реальна.Так
же
реален
и
флексагон, который
легко
сделать
и
склеить
в
домашних
условиях.Он
выглядит
как
двухсторонний
шестиугольник, но
стоит
согнуть
его
особым
образом, и
мы
увидим
третью
сторону.Легко
убедиться,
что
мы
имеем
дело
именно
с
тремя
сторонами, если
раскрасить
их
в
разные
цвета.
Перегибая
флексагон, по
очереди
будем
наблюдать
все
его
поверхности. < / b > < / p >
                         < / body >
                             < / html >
                                 <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
2.
Создать
страницу, как
на
примере: "Вот такой чай"
         <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !DOCTYPE
html >
< html >
< head >
< meta
charset = "UTF-8" / >
          < title > Блоки < / title >
                              < link
href = "style.css"
rel = "stylesheet"
type = "text/css" / >
       < / head >
           < style >
.text
{
margin: 20
px;
padding: 10
px;
}
.www
{
background - color:  # 569099;
color: white;
font - size: 80
px;
}
.sss
{
color: black;
background - color:  # CAD8D0;
font - size: 30
px;
}
< / style >
    < body >
    < div


class ="text www" > < b > Вот такой чай < / b > < / div >

< div


class ="text sss" >


История
о
том, как
один
человек
хотел
попить
чайку, но
по
ошибке
вместо
воды
попытался
налить
в
чайник
бензин, и
что
из
этого
получилось.
< / div >
< / body >
< / html >
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
ПРИСУЦТВУЮТ
НЕ
ВСЕ
ПРАКТИЧЕСКИЕ
И
НЕ
ВСЕ
ДОМАШНИЕ
ЗАДАНИЯ!
ОНИ
В
ОТДЕЛЬНЫХ
ФАЙЛАХ.html
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
Самый
простой
Вариант:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Список
фруктов < / title >
            < style >
            body
{
background - color:  # f0f0f0; /* Светлый фон у всего списка */
}

.fruit - list
{
list - style - type: none;
padding: 10
px; / *Внешний
отступ
для
списка * /
}

.fruit - item
{
background - color:  # e0e0e0; /* Более темный фон у одного элемента списка */
padding: 10
px; / *Внутренний
отступ
для
элемента
списка * /
margin - bottom: 5
px; / *Внешний
отступ
между
элементами
списка * /
}

.fruit - item: last - child
{
background - color:  # 333; /* Самый тёмный фон только под текстом последнего элемента */
color:  # fff; /* Белый текст на тёмном фоне */
}
< / style >
    < / head >
        < body >
        < ul


class ="fruit-list" >

< li


class ="fruit-item" > Яблоко < / li >

< li


class ="fruit-item" > Груша < / li >

< li


class ="fruit-item" > Банан < / li >

< li


class ="fruit-item" > Апельсин < / li >

< !-- Можно
добавить
и
другие
фрукты
по
аналогии -->
< / ul >
< / body >
< / html >
< !--  ########################################################################################## -->

< !--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для -->
< !--    мобильных
устройств. -->
< !-- < title > Список
фруктов < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    background - color:  # f0f0f0;-->
< !--}-->

< !--.fruit - list
{-->
< !--    list - style - type: none;
-->
< !--    padding: 10
px;
-->
< !--}-->

< !--.fruit - item
{-->
< !--    background - color:  # e0e0e0;-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--}-->

< !--.fruit - item: last - child
{-->
< !--    background - color:  # 333;-->
< !--    color:  # fff;-->
< !--}-->

< !--Описание: -->
< !--background - color: Устанавливает
цвет
фона. -->
< !--.fruit - list: Класс
для
стилизации
всего
списка. -->
< !--.fruit - item: Класс
для
стилизации
каждого
элемента
списка. -->
< !--padding: Устанавливает
внутренний
отступ. -->
< !--margin - bottom: Устанавливает
внешний
отступ
между
элементами
списка. -->
< !--.fruit - item: last - child: Стилизация
последнего
элемента
списка. -->
< !--color: Устанавливает
цвет
текста. -->

< !--3.
Структура
HTML - документа: -->

< !-- < ul


class ="fruit-list" > -->

< !-- < li


class ="fruit-item" > Яблоко < / li > -->

< !-- < li


class ="fruit-item" > Груша < / li > -->

< !-- < li


class ="fruit-item" > Банан < / li > -->

< !-- < li


class ="fruit-item" > Апельсин < / li > -->

< !-- & lt;! & ndash;
Можно
добавить
и
другие
фрукты
по
аналогии & ndash; & gt;
-->
< !-- < / ul > -->

< !--Описание: -->
< !--Используется
маркированный
список( < ul >) с
классом.fruit - list. -->
< !--Каждый
элемент
списка( < li >) имеет
класс.fruit - item. -->
< !--Элементы
списка
содержат
названия
фруктов. -->
< !--Последний
элемент
списка
имеет
дополнительные
стили
из
класса.fruit - item: last - child. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
Вариант №2(Немного
"иной"
вариант)
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Список
фруктов < / title >
            < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script >
                                                    < style >
                                                    body
{
background - color:  # f0f0f0;
font - family: Arial, sans - serif;
}

.container
{
display: flex;
justify - content: space - between;
padding: 20
px;
}

.fruit - list
{
list - style - type: none;
padding: 10
px;
width: 200
px;
background - color:  # e0e0e0;
border - radius: 8
px;
margin - right: 20
px;
}

.chart - container
{
width: 60 %;
}
< / style >
    < / head >
        < body >
        < div


class ="container" >

< ul


class ="fruit-list" id="fruits" >

< !-- Список
фруктов
будет
создан
здесь -->
< / ul >

< div


class ="chart-container" >

< canvas
id = "fruitChart" > < / canvas >
< / div >
< / div >

< script >
var
fruitsData = [];

function
generateFruits()
{
    var
fruitCountElements = document.querySelectorAll(".fruit-count");
var
fruitsList = document.getElementById("fruits");
var
chartData = [];

fruitsList.innerHTML = '';
fruitsData = [];

fruitCountElements.forEach(function(element)
{
    var
fruitType = element.getAttribute("data-type");
var
count = parseInt(element.value);

var
listItem = document.createElement("li");
listItem.textContent = `${fruitType} - ${count}
`;
fruitsList.appendChild(listItem);

fruitsData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});

chartData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});
});

updateChart(chartData);
}

function
getBackgroundColor(fruitType)
{
    switch(fruitType)
{
    case
"Яблоко":
return "rgba(255, 99, 132, 0.7)";
case
"Груша":
return "rgba(255, 206, 86, 0.7)";
case
"Банан":
return "rgba(75, 192, 192, 0.7)";
case
"Апельсин":
return "rgba(54, 162, 235, 0.7)";
default:
return "rgba(0, 0, 0, 0.7)";
}
}

function
updateChart(data)
{
    var
ctx = document.getElementById("fruitChart").getContext("2d");
var
myChart = new
Chart(ctx, {
    type: "bar",
    data: {
        labels: ["Количество"],
        datasets: data,
    },
    options: {
        scales: {
            x: {
                stacked: true,
            },
            y: {
                stacked: true,
                beginAtZero: true,
            },
        },
    },
});
}
< / script >

    < div >
    < label
for ="appleCount" > Яблоко:< / label >
                               < input
type = "number"
id = "appleCount"
min = "0"
value = "25"


class ="fruit-count" data-type="Яблоко" >

< / div >

< div >
< label
for ="pearCount" > Груша:< / label >
< input
type = "number"
id = "pearCount"
min = "0"
value = "45"


class ="fruit-count" data-type="Груша" >

< / div >

< div >
< label
for ="bananaCount" > Банан:< / label >
< input
type = "number"
id = "bananaCount"
min = "0"
value = "15"


class ="fruit-count" data-type="Банан" >

< / div >

< div >
< label
for ="orangeCount" > Апельсин:< / label >
< input
type = "number"
id = "orangeCount"
min = "0"
value = "35"


class ="fruit-count" data-type="Апельсин" >

< / div >

< button
onclick = "generateFruits()" > Создать
список
и
диаграмму < / button >
< / body >
< / html >
< !--  ########################################################################################## -->

< !--Шаг
1: Общее
описание -->

< !--Этот
код
представляет
собой
веб - страницу, на
которой
пользователь
может
вводить
количество
четырех
различных -->
< !--видов
фруктов(яблоки, груши, бананы, апельсины)
в
соответствующие
поля
ввода.После
ввода
данных -->
< !--и
нажатия
кнопки, на
странице
отображается
список
введенных
фруктов
и
строится
столбчатая
диаграмма, -->
< !--иллюстрирующая
количество
каждого
фрукта. -->

< !--Шаг
2: Объявление
документа -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script > -->
< !-- < style > -->
< !-- / * ...
стили
для
body,.container,.fruit - list,.chart - container... * / -->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
...
контент
страницы... & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробнее: -->
< !--Здесь
создается
структура
HTML - документа
с
объявлением
типа
документа, -->
< !--установкой
кодировки, определением
вида
и
масштабирования, а
также
заданием
заголовка
страницы. -->
< !--Также
включается
скрипт
библиотеки
Chart.js
и
описываются
стили
для
элементов
страницы. -->

< !--Шаг
3: Стилизация
страницы -->

< !--body
{-->
< !--    background - color:  # f0f0f0;-->
< !--    font - family: Arial, sans - serif;
-->
< !--}-->

< !--.container
{-->
< !--    display: flex;
-->
< !--    justify - content: space - between;
-->
< !--    padding: 20
px;
-->
< !--}-->

< !--.fruit - list
{-->
< !-- / * ...
стили
для
списка
фруктов... * / -->
< !--}-->

< !--.chart - container
{-->
< !--    width: 60 %;
-->
< !--}-->

< !--Стилизация
страницы: -->
< !--Устанавливаются
общие
стили
для
body, определяется
структура
контейнера
с
фруктами
и
графикой, -->
< !--а
также
стили
для
списка
фруктов
и
контейнера
с
графикой. -->

< !--Шаг
4: Контент
страницы -->
< !-- < body > -->
< !-- < div


class ="container" > -->

< !-- < ul


class ="fruit-list" id="fruits" > -->

< !-- & lt;! & ndash;
Список
фруктов
будет
создан
здесь & ndash; & gt;
-->
< !-- < / ul > -->

< !-- < div


class ="chart-container" > -->

< !-- < canvas
id = "fruitChart" > < / canvas > -->
< !-- < / div > -->
< !-- < / div > -->
< !-- & lt;! & ndash;
...
ввод
данных
и
кнопка... & ndash; & gt;
-->
< !-- < / body > -->
< !--Контент
страницы: -->
< !--Определен
контейнер
для
расположения
списка
фруктов
и
графика. -->
< !--Эти
элементы
будут
созданы
и
отображены
динамически. -->

< !--Шаг
5: JavaScript - логика -->

< !-- < script > -->
< !--    var
fruitsData = [];
-->

< !--    function
generateFruits()
{-->
< !-- // ...
логика
создания
списка
фруктов
и
данных
для
графика... -->
< !--}-->

< !--    function
getBackgroundColor(fruitType)
{-->
< !-- // ...
логика
определения
цвета
для
фрукта... -->
< !--}-->

< !--    function
updateChart(data)
{-->
< !-- // ...
логика
обновления
графика... -->
< !--}-->
< !-- < / script > -->

< !--JavaScript - логика: -->
< !--В
этом
скрипте
объявляются
функции
для
генерации
списка
фруктов, определения
цвета
для
фрукта
и
обновления
графика. -->
< !--Также
создается
массив
fruitsData
для
хранения
данных
о
фруктах. -->

< !--Шаг
6: Создание
и
ввод
данных -->

< !-- < div > -->
< !-- < label
for ="appleCount" > Яблоко:< / label > -->
< !-- < input
type = "number"
id = "appleCount"
min = "0"
value = "25"


class ="fruit-count" data-type="Яблоко" > -->

< !-- < / div > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < button
onclick = "generateFruits()" > Создать
список
и
диаграмму < / button > -->

< !--Создание
и
ввод
данных: -->
< !--Здесь
предоставлены
поля
ввода
для
каждого
типа
фрукта
и
кнопка
"Создать список и диаграмму". -->
< !--По
нажатии
кнопки
вызывается
функция
generateFruits(), которая
обновляет
список
фруктов
и
строит
график. -->

< !--Шаг
7: Динамическое
обновление
списка
фруктов
и
графика -->

< !--function
generateFruits()
{-->
< !--    var
fruitCountElements = document.querySelectorAll(".fruit-count");
-->
< !--    var
fruitsList = document.getElementById("fruits");
-->
< !--    var
chartData = [];
-->

< !--    fruitsList.innerHTML = '';
-->
< !--    fruitsData = [];
-->

< !--    fruitCountElements.forEach(function(element)
{-->
< !--        var
fruitType = element.getAttribute("data-type");
-->
< !--        var
count = parseInt(element.value);
-->

< !--        var
listItem = document.createElement("li");
-->
< !--        listItem.textContent = `${fruitType} - ${count}
`;
-->
< !--        fruitsList.appendChild(listItem);
-->

< !--        fruitsData.push({-->
< !--            label: fruitType, -->
< !--            data: [count], -->
< !--            backgroundColor: getBackgroundColor(fruitType), -->
< !--});-->

< !--        chartData.push({-->
< !--            label: fruitType, -->
< !--            data: [count], -->
< !--            backgroundColor: getBackgroundColor(fruitType), -->
< !--});-->
< !--});-->

< !--    updateChart(chartData);
-->
< !--}-->

< !--function
getBackgroundColor(fruitType)
{-->
< !--    switch(fruitType)
{-->
< !--        case
"Яблоко": -->
< !--
return "rgba(255, 99, 132, 0.7)";
-->
< !--        case
"Груша": -->
< !--
return "rgba(255, 206, 86, 0.7)";
-->
< !--        case
"Банан": -->
< !--
return "rgba(75, 192, 192, 0.7)";
-->
< !--        case
"Апельсин": -->
< !--
return "rgba(54, 162, 235, 0.7)";
-->
< !--        default: -->
< !--
return "rgba(0, 0, 0, 0.7)";
-->
< !--}-->
< !--}-->

< !--function
updateChart(data)
{-->
< !--    var
ctx = document.getElementById("fruitChart").getContext("2d");
-->
< !--    var
myChart = new
Chart(ctx, {-->
< !--        type: "bar", -->
< !--        data: {-->
< !--            labels: ["Количество"], -->
< !--            datasets: data, -->
< !--}, -->
< !--        options: {-->
< !--            scales: {-->
< !--                x: {-->
< !--                    stacked: true, -->
< !--}, -->
< !--                y: {-->
< !--                    stacked: true, -->
< !--                    beginAtZero: true, -->
< !--}, -->
< !--}, -->
< !--}, -->
< !--});-->
< !--}-->

< !--Динамическое
обновление
списка
фруктов
и
графика: -->
< !--Функция
generateFruits()
собирает
данные
из
полей
ввода, обновляет
список
фруктов
и -->
< !--создает
массив
данных
для
графика.Функции
getBackgroundColor()
и
updateChart()
используются
для -->
< !--определения
цвета
фрукта
и
обновления
столбчатой
диаграммы
соответственно. -->

< !--Общий
вывод: -->
< !--Этот
код
создает
интерактивную
страницу
для
ввода
данных
о
фруктах, -->
< !--отображения
списка
фруктов
и
построения
столбчатой
диаграммы
на
основе
введенных
значений. -->
< !--Каждая
строчка
кода
выполняет
конкретную
задачу, начиная
от
объявления
документа, стилей
и
заканчивая -->
< !--динамическим
обновлением
контента
и
графика. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №3: Чуть - Чуть
красивее(Но
не
все
еще
не
так
как
в
примере
в
ДЗ)
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script >
                                                    < style >
                                                    body
{
    background - color:  # f5f5f5;
        font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
display: flex;
align - items: center;
justify - content: center;
height: 100
vh;
margin: 0;
}

.container
{
    background - color:  # fff;
        border - radius: 8
px;
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.1);
padding: 20
px;
width: 80 %;
max - width: 600
px;
animation: fadeInUp
0.5
s
ease - out;
}

@keyframes


fadeInUp
{
from

{
    opacity: 0;
transform: translateY(20
px);
}
to
{
    opacity: 1;
transform: translateY(0);
}
}

.chart - container
{
    margin - top: 20px;
}

.fruit - input
{
    display: flex;
justify - content: space - between;
align - items: center;
margin - bottom: 10
px;
}

.fruit - label
{
    width: 70px;
font - weight: bold;
}

.fruit - count
{
    width: 50px;
padding: 5
px;
border: 1
px
solid  # ddd;
border - radius: 4
px;
}

.generate - button
{
    background - color:  # 4caf50;
        color:  # fff;
padding: 10
px;
border: none;
border - radius: 4
px;
cursor: pointer;
transition: background - color
0.3
s
ease;
}

.generate - button: hover
{
    background - color:  # 45a049;
}

.fruits - list
{
    list - style - type: none;
padding: 10
px;
margin: 0;
}

.fruits - list
li
{
    margin - bottom: 5px;
display: flex;
align - items: center;
}

.fruit - emoji
{
    margin - right: 10px;
font - size: 20
px;
}
< / style >
    < / head >
        < body >
        < div


class ="container" >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="appleCount" > < span class ="fruit-emoji" > 🍎 < / span > Яблоко:<

    / label >
< input
type = "number"
id = "appleCount"
min = "0"
value = "25"


class ="fruit-count" data-type="Яблоко" >

< / div >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="pearCount" > < span class ="fruit-emoji" > 🍐 < / span > Груша:<

    / label >
< input
type = "number"
id = "pearCount"
min = "0"
value = "45"


class ="fruit-count" data-type="Груша" >

< / div >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="bananaCount" > < span class ="fruit-emoji" > 🍌 < / span > Банан:<

    / label >
< input
type = "number"
id = "bananaCount"
min = "0"
value = "15"


class ="fruit-count" data-type="Банан" >

< / div >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="orangeCount" > < span class ="fruit-emoji" > 🍊 < / span > Апельсин:<

    / label >
< input
type = "number"
id = "orangeCount"
min = "0"
value = "35"


class ="fruit-count" data-type="Апельсин" >

< / div >

< button


class ="generate-button" onclick="generateFruits()" > Создать список и диаграмму < / button >

< ul
id = "fruits-list"


class ="fruits-list" > < / ul >

< div


class ="chart-container" >

< canvas
id = "fruitChart" > < / canvas >
< / div >
< / div >

< script >
var
fruitsData = [];

function
generateFruits()
{
    var
fruitCountElements = document.querySelectorAll(".fruit-count");
var
fruitsList = document.getElementById("fruits-list");
var
chartData = [];

fruitsList.innerHTML = '';
fruitsData = [];

fruitCountElements.forEach(function(element)
{
    var
fruitType = element.getAttribute("data-type");
var
count = parseInt(element.value);

var
listItem = document.createElement("li");
listItem.innerHTML = ` < span


class ="fruit-emoji" > ${getFruitEmoji(fruitType)} < / span > ${fruitType} - ${count}`;


fruitsList.appendChild(listItem);

fruitsData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});

chartData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});
});

updateChart(chartData);
}

function
getFruitEmoji(fruitType)
{
    switch(fruitType)
{
    case
"Яблоко":
return "🍎";
case
"Груша":
return "🍐";
case
"Банан":
return "🍌";
case
"Апельсин":
return "🍊";
default:
return "❓";
}
}

function
getBackgroundColor(fruitType)
{
    switch(fruitType)
{
case
"Яблоко":
return "rgba(255, 99, 132, 0.7)";
case
"Груша":
return "rgba(255, 204, 0, 0.7)";
case
"Банан":
return "rgba(255, 215, 0, 0.7)";
case
"Апельсин":
return "rgba(255, 165, 0, 0.7)";
default:
return "rgba(0, 0, 0, 0.7)";
}
}

function
updateChart(data)
{
    var
ctx = document.getElementById("fruitChart").getContext("2d");
var
myChart = new
Chart(ctx, {
    type: "bar",
    data: {
        labels: ["Количество"],
        datasets: data,
    },
    options: {
        scales: {
            x: {
                stacked: true,
            },
            y: {
                stacked: true,
                beginAtZero: true,
            },
        },
    },
});
}
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--1.
Шаг №1: Объявление
документа -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script > -->
< !-- < style > -->
< !-- / * ...
стили
для
body,.container,.fruit - input, и
другие... * / -->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
...
контент
страницы... & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание
этого
кода: -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешней
библиотеки
Chart.js. -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для -->
< !--    устройств
с
различной
шириной
экрана. -->
< !-- < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script >: Подключение
библиотеки
Chart.js
для
создания
графиков. -->
< !-- < style >: Секция
стилей
для
определения
внешнего
вида
элементов
на
странице. -->

< !--Шаг №2: Стилизация
страницы -->

< !--Пример
выполнения
этого
шага: -->

< !--body
{-->
< !--    background - color:  # f5f5f5;-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    display: flex;
-->
< !--    align - items: center;
-->
< !--    justify - content: center;
-->
< !--    height: 100
vh;
-->
< !--    margin: 0;
-->
< !--}-->

< !--.container
{-->
< !-- / * ...
стили
для
контейнера... * / -->
< !--}-->

< !-- / * ...
стили
для
других
элементов... * / -->

< !--Подробное
описание
этого
кода: -->

< !--body: Определение
общих
стилей
для
всего
документа, таких
как
цвет
фона, шрифт
и
центрирование
содержимого. -->
< !--.container: Стили
для
основного
контейнера, в
котором
размещены
остальные
элементы
страницы. -->
< !-- @ keyframes
fadeInUp: Анимация, делающая
элементы
появляющимися
с
небольшим
подъемом
сверху
при
загрузке
страницы. -->
< !--Стили
для
различных
элементов
формы, ввода
данных
и
графика. -->

< !--Шаг №3: Контент
страницы -->

< !--Пример
выполнения
этого
шага: -->

< !-- < body > -->
< !-- < div


class ="container" > -->

< !-- < div


class ="fruit-input" > -->

< !-- & lt;! & ndash;
...
ввод
данных
для
яблок... & ndash; & gt;
-->
< !-- < / div > -->

< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->

< !-- < button


class ="generate-button" onclick="generateFruits()" > Создать список и диаграмму < / button > -->

< !-- < ul
id = "fruits-list"


class ="fruits-list" > < / ul > -->

< !-- < div


class ="chart-container" > -->

< !-- < canvas
id = "fruitChart" > < / canvas > -->
< !-- < / div > -->
< !-- < / div > -->

< !-- < script > -->
< !-- // ...
скрипт
JavaScript
для
обработки
данных... -->
< !-- < / script > -->
< !-- < / body > -->

< !--Подробное
описание
этого
кода: -->

< !--div
с
классом
container: Основной
контейнер
страницы. -->
< !--div
с
классом
fruit - input: Блок
для
ввода
данных
о
фруктах. -->
< !--button
с
классом
generate - button: Кнопка
для
генерации
списка
фруктов
и
построения
диаграммы. -->
< !--ul
с
id
fruits - list: Список, в
который
будут
добавляться
элементы
с
данными
о
фруктах. -->
< !--div
с
классом
chart - container: Контейнер
для
размещения
графика. -->

< !--Шаг №4: JavaScript - логика -->

< !--Пример
выполнения
этого
шага: -->

< !-- < script > -->
< !--    var
fruitsData = [];
-->

< !--    function
generateFruits()
{-->
< !-- // ...
логика
создания
списка
фруктов
и
данных
для
графика... -->
< !--}-->

< !--    function
getFruitEmoji(fruitType)
{-->
< !-- // ...
логика
определения
эмодзи
для
фрукта... -->
< !--}-->

< !--    function
getBackgroundColor(fruitType)
{-->
< !-- // ...
логика
определения
цвета
для
фрукта... -->
< !--}-->

< !--    function
updateChart(data)
{-->
< !-- // ...
логика
обновления
графика... -->
< !--}-->
< !-- < / script > -->

< !--Подробное
описание
этого
кода: -->

< !--var
fruitsData = [];: Объявление
массива
для
хранения
данных
о
фруктах. -->
< !--Функция
generateFruits(): Обработка
данных
ввода, создание
списка
фруктов
и
массива
данных
для
графика. -->
< !--Функции
getFruitEmoji(), getBackgroundColor(): Логика
определения
эмодзи
и
цвета
для
фруктов. -->
< !--Функция
updateChart(data): Логика
обновления
графика
на
основе
предоставленных
данных. -->

< !--Общий
вывод: -->

< !--Этот
код
создает
интерактивную
страницу
для
ввода
данных
о
фруктах, -->
< !--отображения
списка
фруктов
и
построения
столбчатой
диаграммы
на
основе
введенных
значений. -->
< !--Он
включает
в
себя
HTML - разметку, стили
CSS
для
оформления, а
также
JavaScript - логику -->
< !--для
обработки
данных
и
обновления
графика. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №4: "Новый взгляд"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < style >
          body
{
background - color:  # f5f5f5;
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
display: flex;
align - items: center;
justify - content: center;
height: 100
vh;
margin: 0;
}

ul
{
list - style - type: none;
padding: 10
px;
margin: 0;
background - color:  # fff; /* Светлый фон у всего списка */
border - radius: 8
px;
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.1);
}

li
{
margin - bottom: 10
px;
padding: 10
px;
border - radius: 4
px;
background - color:  # f0f0f0; /* Более темный фон у одного элемента списка */
transition: background - color
0.3
s
ease - in -out;
display: flex;
justify - content: space - between;
align - items: center;
}

p
{
margin: 0;
padding: 5
px;
border - radius: 4
px;
background - color:  # d9d9d9; /* Самый тёмный фон только под текстом */
}

/ *Добавим
стили
для
цветов
фруктов * /
.apple
{
background - color:  # ff6666; /* Цвет фона для яблока */
}

.banana
{
background - color:  # ffe066; /* Цвет фона для банана */
}

.orange
{
background - color:  # ffa366; /* Цвет фона для апельсина */
}

.pear
{
background - color:  # b3ff99; /* Цвет фона для груши */
}

/ *и
так
далее
для
других
фруктов * /
.kiwi
{
background - color:  # 99cc66;
}

.watermelon
{
background - color:  # ff5050;
}

.grape
{
background - color:  # 990099;
}

.strawberry
{
background - color:  # ff66b2;
}

.pineapple
{
background - color:  # ffcc66;
}

/ *Добавим
стили
для
эмодзи
фруктов * /
.emoji
{
font - size: 20
px;
margin - right: 5
px;
}

/ *Анимация * /
   @ keyframes
countUp
{
from

{
    opacity: 0;
transform: translateY(10
px);
}
to
{
    opacity: 1;
transform: translateY(0);
}
}

span
{
animation: countUp
0.5
s
ease - in -out;
display: inline - block;
}
< / style >
    < / head >
        < body >
        < ul >
        < li


class ="apple" >

< p


class ="emoji" > 🍎 < / p > Яблоко - < span id="appleCount" > < / span >

< / li >
< li


class ="banana" >

< p


class ="emoji" > 🍌 < / p > Банан - < span id="bananaCount" > < / span >

< / li >
< li


class ="orange" >

< p


class ="emoji" > 🍊 < / p > Апельсин - < span id="orangeCount" > < / span >

< / li >
< li


class ="pear" >

< p


class ="emoji" > 🍐 < / p > Груша - < span id="pearCount" > < / span >

< / li >
< li


class ="kiwi" >

< p


class ="emoji" > 🥝 < / p > Киви - < span id="kiwiCount" > < / span >

< / li >
< li


class ="watermelon" >

< p


class ="emoji" > 🍉 < / p > Арбуз - < span id="watermelonCount" > < / span >

< / li >
< li


class ="grape" >

< p


class ="emoji" > 🍇 < / p > Виноград - < span id="grapeCount" > < / span >

< / li >
< li


class ="strawberry" >

< p


class ="emoji" > 🍓 < / p > Клубника - < span id="strawberryCount" > < / span >

< / li >
< li


class ="pineapple" >

< p


class ="emoji" > 🍍 < / p > Ананас - < span id="pineappleCount" > < / span >

< / li >
< / ul >

< script >
function
getRandomCount()
{
return Math.floor(Math.random() * 100) + 1;
}

document.getElementById("appleCount").textContent = getRandomCount();
document.getElementById("bananaCount").textContent = getRandomCount();
document.getElementById("orangeCount").textContent = getRandomCount();
document.getElementById("pearCount").textContent = getRandomCount();
document.getElementById("kiwiCount").textContent = getRandomCount();
document.getElementById("watermelonCount").textContent = getRandomCount();
document.getElementById("grapeCount").textContent = getRandomCount();
document.getElementById("strawberryCount").textContent = getRandomCount();
document.getElementById("pineappleCount").textContent = getRandomCount();
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--1.
Шаг №1: Объявление
документа
и
настройка
вида
страницы -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < style > -->
< !-- / * ...
стили
для
body, ul, li, p
и
другие... * / -->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
...
контент
страницы... & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание
этого
кода: -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
стили
для
определения
внешнего
вида
элементов
на
странице. -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для -->
< !--    устройств
с
различной
шириной
экрана. -->
< !-- < style >: Секция
стилей
для
определения
внешнего
вида
элементов
страницы. -->

< !--Шаг №2: Стилизация
элементов
списка -->

< !--Пример
выполнения
этого
шага: -->

< !--ul
{-->
< !-- / * ...
стили
для
списка... * / -->
< !--}-->

< !--li
{-->
< !-- / * ...
стили
для
элементов
списка... * / -->
< !--}-->

< !--p
{-->
< !-- / * ...
стили
для
абзаца... * / -->
< !--}-->

< !-- / * ...
стили
для
цветов
фруктов... * / -->
< !-- / * ...
стили
для
эмодзи
фруктов... * / -->

< !-- @ keyframes
countUp
{-->
< !-- / * ...
стили
для
анимации... * / -->
< !--}-->

< !--span
{-->
< !-- / * ...
стили
для
элемента
span... * / -->
< !--}-->

< !--Подробное
описание
этого
кода: -->

< !--Стили
для
общего
вида
ul(списка)
и
li(элементов
списка).-->
< !--p: Стили
для
абзаца
внутри
элемента
списка. -->
< !--Стили
для
цветов
фруктов, каждый
фрукт
имеет
свой
уникальный
цвет
фона. -->
< !--Стили
для
эмодзи
фруктов, в
данном
случае, устанавливается
размер
и
отступ
справа. -->

< !--Шаг №3: Контент
страницы
и
JavaScript
для
случайных
значений -->

< !--Пример
выполнения
этого
шага: -->

< !-- < body > -->
< !-- < ul > -->
< !-- < li


class ="apple" > -->

< !-- < p


class ="emoji" > 🍎 < / p > Яблоко - < span id="appleCount" > < / span > -->

< !-- < / li > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ul > -->

< !-- < script > -->
< !--        function
getRandomCount()
{-->
< !--
return Math.floor(Math.random() * 100) + 1;
-->
< !--}-->

< !--        document.getElementById("appleCount").textContent = getRandomCount();
-->
< !-- // Аналогичные
строки
для
других
фруктов -->
< !-- < / script > -->
< !-- < / body > -->

< !--Подробное
описание
этого
кода: -->

< !--ul: Определение
списка
для
отображения
фруктов. -->
< !--li: Элемент
списка
для
каждого
фрукта
с
указанием
эмодзи, названия
и
значения. -->
< !--Функция
getRandomCount(): Генерирует
случайное
число
для
отображения
количества
фруктов. -->
< !--JavaScript - код: Устанавливает
случайные
значения
для
каждого
фрукта
в
соответствующие
элементы. -->

< !--Общий
вывод: -->
< !--Этот
код
создает
страницу, на
которой
отображаются
различные
фрукты
с
их
количеством. -->
< !--Каждый
фрукт
имеет
свой
уникальный
цвет
фона, эмодзи
и
анимацию
при
первоначальной
загрузке. -->
< !--JavaScript
используется
для
генерации
случайных
значений
количества
фруктов
при
каждой
загрузке
страницы. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №5: "По заданию с изменениями"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "2_fruit_list.css" >
       < title > Fruit
list < / title >
         < / head >
             < style >

*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # ede8e8;
}

ol
{
    width: 440px;
padding: 20
px;
list - style - position: inside;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
opacity: 0;
}
to
{
opacity: 1;
}
}

li
{
    margin: 10px 0 10px 0;
padding: 5
px
5
px
5
px
10
px;
background - color:  # c66376;
transition: background - color
0.3
s
ease - in -out;
}

li: hover
{
    background - color:  # a54257;
}

span
{
    margin - left: 8px;
padding: 5
px
15
px
5
px
15
px;
background - color:  # 9e354a;
}

.size_list
{
    margin: 0px;
}

.color_list
{
    background - color:  # ed9eae;
}

li: first - child
{
    margin - top: 0px;
}

li: last - child
{
    margin - bottom: 0px;
}
< / style >
    < body >
    < ol


class ="color_list size_list" >

< li > < span > Apple < / span > < / li >
< li > < span > Orange < / span > < / li >
< li > < span > Pineapple < / span > < / li >
< li > < span > Pear < / span > < / li >
< li > < span > Cherry < / span > < / li >
< li > < span > Banana < / span > < / li >
< li > < span > Grapes < / span > < / li >
< li > < span > Strawberry < / span > < / li >
< li > < span > Mango < / span > < / li >
< li > < span > Watermelon < / span > < / li >
< li > < span > Blueberry < / span > < / li >
< li > < span > Kiwi < / span > < / li >
< li > < span > Peach < / span > < / li >
< li > < span > Raspberry < / span > < / li >
< / ol >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--1.
Шаг №1: Настройка
и
подключение
внешних
ресурсов -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Fruit
list < / title > -->
< !-- < / head > -->

< !--Подробное
описание
этого
кода: -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"2_fruit_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования -->
< !--    для
устройств
с
различной
шириной
экрана. -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" >: Подключение
внешнего
файла
стилей. -->

< !--Шаг №2: Стилизация
страницы
и
анимация -->

< !--Пример
выполнения
этого
шага: -->

< !-- < style > -->

< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->
< !-- < / style > -->

< !--Подробное
описание
этого
кода: -->

< !--Общие
стили
для
всех
элементов
на
странице: установка
шрифта, цвета
фона
и
текста. -->
< !--Стили
для
упорядоченного
списка(ol): фиксированная
ширина, внутренний
отступ, позиция
маркера -->
< !--    списка
внутри
элемента, и
анимация
появления. -->
< !--Анимация @ keyframes
fadeIn: Плавное
появление
элементов
списка. -->
< !--Стили
для
элементов
списка(li): внешний
и
внутренний
отступы, цвет
фона
и
переход
при
наведении. -->
< !--Стили
для
вложенного
элемента
span: отступ
слева, внутренний
отступ
и
цвет
фона. -->
< !--Дополнительные
стили
для
списков
с
классами
size_list
и
color_list, а
также
для -->
< !--    первого
и
последнего
элементов
списка. -->

< !--Шаг №3: Разметка
и
отображение
данных -->

< !--Пример
выполнения
этого
шага: -->

< !-- < body > -->
< !-- < ol


class ="color_list size_list" > -->

< !-- < li > < span > Apple < / span > < / li > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание
этого
кода: -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ol


class ="color_list size_list" >: Упорядоченный


список
с
указанием
классов
для
стилизации. -->
< !-- < li > < span > Apple < / span > < / li >: Элемент
списка
с
текстом
"Apple"
и
вложенным
элементом
span. -->

< !--Как
итог: Этот
код
создает
страницу
с
стилизованным
упорядоченным
списком
фруктов. -->
< !--Каждый
фрукт
имеет
свой
цвет
фона, отличающийся
при
наведении. -->
< !--Анимация
появления
списка
при
загрузке
страницы
создает
плавный
эффект. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №6: "Попытка не пытка"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < / head > -->
< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->
< !-- < / style > -->
< !-- < body > -->
< !-- < ol


class ="color_list size_list" id="fruitList" > -->

< !-- < li > < span > Яблоко < / span > < / li > -->
< !-- < li > < span > Апельсин < / span > < / li > -->
< !-- < li > < span > Ананас < / span > < / li > -->
< !-- < li > < span > Груша < / span > < / li > -->
< !-- < li > < span > Вишня < / span > < / li > -->
< !-- < li > < span > Банан < / span > < / li > -->
< !-- < li > < span > Виноград < / span > < / li > -->
< !-- < li > < span > Клубника < / span > < / li > -->
< !-- < li > < span > Манго < / span > < / li > -->
< !-- < li > < span > Арбуз < / span > < / li > -->
< !-- < li > < span > Голубика < / span > < / li > -->
< !-- < li > < span > Киви < / span > < / li > -->
< !-- < li > < span > Персик < / span > < / li > -->
< !-- < li > < span > Малина < / span > < / li > -->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->
< !--  ########################################################################################## -->
< !-- ДЛЯ
КРАСОТЫ -->
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "ru" >
< head >
< meta
charset = "UTF-8" >
< meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< link
rel = "stylesheet"
href = "2_fruit_list.css" >
< title > Список
фруктов < / title >
< / head >
< style >
*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # ede8e8;
}

ol
{
    width: 440px;
padding: 20
px;
list - style - position: inside;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
    opacity: 0;
}
to
{
    opacity: 1;
}
}

li
{
    margin: 10px 0 10px 0;
padding: 5
px
5
px
5
px
10
px;
background - color:  # c66376;
transition: background - color
0.3
s
ease - in -out;
}

li: hover
{
    background - color:  # a54257;
}

span
{
    margin - left: 8px;
padding: 5
px
15
px
5
px
15
px;
background - color:  # 9e354a;
transition: background - color
0.3
s
ease - in -out;
}

.size_list
{
    margin: 0px;
}

.color_list
{
    background - color:  # ed9eae;
}

li: first - child
{
    margin - top: 0px;
}

li: last - child
{
    margin - bottom: 0px;
}
< / style >
< body >
< ol


class ="color_list size_list" id="fruitList" >

< li > < span >🍎 Яблоко < / span > < / li >
< li > < span >🍊 Апельсин < / span > < / li >
< li > < span >🍍 Ананас < / span > < / li >
< li > < span >🍐 Груша < / span > < / li >
< li > < span >🍒 Вишня < / span > < / li >
< li > < span >🍌 Банан < / span > < / li >
< li > < span >🍇 Виноград < / span > < / li >
< li > < span >🍓 Клубника < / span > < / li >
< li > < span >🥭 Манго < / span > < / li >
< li > < span >🍉 Арбуз < / span > < / li >
< li > < span >🫐 Голубика < / span > < / li >
< li > < span >🥝 Киви < / span > < / li >
< li > < span >🍑 Персик < / span > < / li >
< li > < span >🍇 Малина < / span > < / li >
< / ol >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--1.
Шаг №1: Настройка
и
подключение
стилей -->
< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "ru" >: Начало
HTML - разметки
с
указанием
языка(русский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"2_fruit_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для -->
< !--    устройств
с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
анимация -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: -->
< !--    Задание
общих
стилей
для
всех
элементов
на
странице, таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ol(упорядоченного
списка): Фиксированная
ширина, внутренний
отступ, позиция
маркера -->
< !--    списка
внутри
элемента, анимация
появления. -->
< !--Анимация @ keyframes
fadeIn: -->
< !--    Плавное
появление
элементов
списка
при
загрузке
страницы. -->
< !--Стили
для
li(элементов
списка):-->
< !--    Внешний
и
внутренний
отступы, цвет
фона
и
эффект
перехода
при
наведении. -->
< !--Стили
для
span(вложенного
элемента):-->
< !--    Отступ
слева, внутренний
отступ
и
цвет
фона, а
также
эффект
перехода
при
наведении. -->
< !--Дополнительные
стили
для
списков
с
классами
size_list
и
color_list, а
также
для
первого -->
< !--    и
последнего
элементов
списка. -->

< !--3.
Шаг №3: Разметка
и
отображение
данных -->

< !-- < body > -->
< !-- < ol


class ="color_list size_list" id="fruitList" > -->

< !-- < li > < span >🍎 Яблоко < / span > < / li > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ol


class ="color_list size_list" id="fruitList" >: Упорядоченный


список
с
указанием
классов
для -->
< !--    стилизации
и
идентификатора
"fruitList". -->
< !-- < li > < span >🍎 Яблоко < / span > < / li >: Элемент
списка
с
текстом, включая
эмодзи
для
каждого
фрукта. -->
< !--    -->
< !--Ну
и
снова
опять
двадцать
пять: -->
< !--    -->
< !--Этот
код
создает
стильную
страницу
со
списком
фруктов. -->
< !--    Упорядоченный
список
имеет
анимированный
эффект
появления, -->
< !--    каждый
элемент
списка
стилизован
с
использованием
цветов
и
переходов
при
наведении, -->
< !--    а
каждый
фрукт
дополнен
соответствующим
эмодзи. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №6: "Попытка не пытка"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "ru" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "2_fruit_list.css" >
       < title > Список
фруктов < / title >
            < / head >
                < style >
*{
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # ede8e8;
}

ol
{
width: 440
px;
padding: 20
px;
list - style - position: inside;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
    opacity: 0;
}
to
{
    opacity: 1;
}
}

li
{
margin: 10
px
0
10
px
0;
padding: 5
px
5
px
5
px
10
px;
background - color:  # c66376;
transition: background - color
0.3
s
ease - in -out;
display: flex;
align - items: center;
}

li: hover
{
background - color:  # a54257;
}

span
{
margin - left: 8
px;
padding: 5
px
15
px
5
px
15
px;
background - color:  # 9e354a;
transition: background - color
0.3
s
ease - in -out;
}

.size_list
{
margin: 0
px;
}

.color_list
{
background - color:  # ed9eae;
}

li: first - child
{
margin - top: 0
px;
}

li: last - child
{
margin - bottom: 0
px;
}

.emoji
{
font - size: 20
px;
margin - right: 8
px;
display: inline - block;
animation: bounce
0.8
s
ease
infinite
alternate;
}

@keyframes


bounce
{
0 % {
    transform: translateY(0);
}
100 % {
    transform: translateY(-5px);
}
}
< / style >
    < body >
    < ol


class ="color_list size_list" id="fruitList" >

< li > < span


class ="emoji" > 🍏🍏🍏 < / span > < span > Яблоко < / span > < / li >

< li > < span


class ="emoji" > 🍊🍊 < / span > < span > Апельсин < / span > < / li >

< li > < span


class ="emoji" > 🍍🍍🍍🍍 < / span > < span > Ананас < / span > < / li >

< li > < span


class ="emoji" > 🍐🍐 < / span > < span > Груша < / span > < / li >

< li > < span


class ="emoji" > 🍒🍒🍒🍒🍒 < / span > < span > Вишня < / span > < / li >

< li > < span


class ="emoji" > 🍌 < / span > < span > Банан < / span > < / li >

< li > < span


class ="emoji" > 🍇🍇🍇 < / span > < span > Виноград < / span > < / li >

< li > < span


class ="emoji" > 🍓🍓🍓 < / span > < span > Клубника < / span > < / li >

< li > < span


class ="emoji" > 🥭🥭 < / span > < span > Манго < / span > < / li >

< li > < span


class ="emoji" > 🍉🍉🍉 < / span > < span > Арбуз < / span > < / li >

< li > < span


class ="emoji" > 🫐🫐🫐 < / span > < span > Голубика < / span > < / li >

< li > < span


class ="emoji" > 🥝🥝 < / span > < span > Киви < / span > < / li >

< li > < span


class ="emoji" > 🍑🍑 < / span > < span > Персик < / span > < / li >

< li > < span


class ="emoji" > 🍇🍇 < / span > < span > Малина < / span > < / li >

< / ol >
< / body >
< / html >
< !--  ###################################################################### -->
< !--1.
Шаг №1: Настройка
и
подключение
стилей -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "ru" >: Начало
HTML - разметки
с
указанием
языка(русский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"2_fruit_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования -->
< !--    для
устройств
с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
анимация -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--    display: flex;
-->
< !--    align - items: center;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->

< !--.emoji
{-->
< !--    font - size: 20
px;
-->
< !--    margin - right: 8
px;
-->
< !--    display: inline - block;
-->
< !--    animation: bounce
0.8
s
ease
infinite
alternate;
-->
< !--}-->

< !-- @ keyframes
bounce
{-->
< !--    0 % {-->
< !--        transform: translateY(0);
-->
< !--}-->
< !--    100 % {-->
< !--        transform: translateY(-5
px);-->
< !--}-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: Задание
общих
стилей
для
всех
элементов
на
странице, таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ol(упорядоченного
списка): Фиксированная
ширина, внутренний
отступ, позиция
маркера -->
< !--    списка
внутри
элемента, анимация
появления. -->
< !--Анимация @ keyframes
fadeIn: Плавное
появление
элементов
списка
при
загрузке
страницы. -->
< !--Стили
для
li(элементов
списка): Внешний
и
внутренний
отступы, цвет
фона
и
эффект
перехода -->
< !--    при
наведении, использование
flex
для
выравнивания
элементов. -->
< !--Стили
для
span(вложенного
элемента): Отступ
слева, внутренний
отступ
и
цвет -->
< !--    фона, а
также
эффект
перехода
при
наведении. -->
< !--Дополнительные
стили
для
списков
с
классами
size_list
и
color_list, а
также
для
первого -->
< !--    и
последнего
элементов
списка. -->
< !--Анимация
для.emoji
с
использованием @ keyframes: Прыжковая
анимация
для
эмодзи. -->

< !--3.
Шаг №3: Разметка
и
отображение
данных -->

< !-- < body > -->
< !-- < ol


class ="color_list size_list" id="fruitList" > -->

< !-- < li > < span


class ="emoji" > 🍏🍏🍏 < / span > < span > Яблоко < / span > < / li > -->

< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ol


class ="color_list size_list" id="fruitList" >: Упорядоченный


список
с
указанием -->
< !--    классов
для
стилизации
и
идентификатора
"fruitList". -->
< !-- < li > < span


class ="emoji" > 🍏🍏🍏 < / span > < span > Яблоко < / span > < / li >: Элемент


списка
с -->
< !--    эмодзи
и
текстом
для
каждого
фрукта. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Самый
простой
Вариант:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "task_list.css" >
       < title > Список
задач < / title >
          < / head >
              < style >

*{
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 333;
}

ul
{
list - style: none;
padding: 0;
width: 300
px;
}

li
{
padding: 10
px;
margin - bottom: 5
px;
border - radius: 5
px;
}

/ *Четные
элементы
списка * /
li: nth - child(even)
{
background - color:  # d3e0ea;
}

/ *Нечетные
элементы
списка * /
li: nth - child(odd)
{
background - color:  # a8bdd4;
}

/ *Отмеченные
элементы * /
li.checked
{
background - color:  # ccc;
text - decoration: line - through;
}

/ *Неотмеченные
элементы * /
li: not (.checked)::before
{
content: "❌";
color: red;
margin - right: 5
px;
}

/ *Отмеченные
элементы * /
li.checked::before
{
content: "✔";
color: green;
margin - right: 5
px;
}

< / style >
    < body >
    < ul >
    < li


class ="checked" > Выполнить задачу 1 < / li >

< li > Выполнить
задачу
2 < / li >
< li


class ="checked" > Выполнить задачу 3 < / li >

< li > Выполнить
задачу
4 < / li >
< li > Выполнить
задачу
5 < / li >
< / ul >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--1.
Шаг №1: Настройка
и
подключение
стилей -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"task_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для
устройств -->
< !--    с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
определение
задач -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 333;-->
< !--}-->

< !--ul
{-->
< !--    list - style: none;
-->
< !--    padding: 0;
-->
< !--    width: 300
px;
-->
< !--}-->

< !--li
{-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--}-->

< !-- / * Четные
элементы
списка * / -->
< !--li: nth - child(even)
{-->
< !--    background - color:  # d3e0ea;-->
< !--}-->

< !-- / * Нечетные
элементы
списка * / -->
< !--li: nth - child(odd)
{-->
< !--    background - color:  # a8bdd4;-->
< !--}-->

< !-- / * Отмеченные
элементы * / -->
< !--li.checked
{-->
< !--    background - color:  # ccc;-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !-- / * Неотмеченные
элементы * / -->
< !--li: not (.checked)::before
{-->
< !--    content: "❌";
-->
< !--    color: red;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !-- / * Отмеченные
элементы * / -->
< !--li.checked::before
{-->
< !--    content: "✔";
-->
< !--    color: green;
-->
< !--    margin - right: 5
px;
-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: Задание
общих
стилей
для
всех
элементов
на
странице, -->
< !--    таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ul(списка): Сброс
стилей
маркера
списка
и
установка
ширины. -->
< !--Стили
для
li(элементов
списка): Определение
внешнего
и
внутреннего
отступов, -->
< !--    а
также
скругления
углов. -->
< !--Стили
для
li: nth - child(even)
и
li: nth - child(odd): Определение
цветов
фона
для
четных
и
нечетных -->
< !--    элементов
списка
соответственно. -->
< !--Стили
для
li.checked: Определение
стилей
для
отмеченных
элементов, включая
цвет
фона -->
< !--    и
зачеркивание
текста. -->
< !--Стили
для
li: not (.checked)::before
и
li.checked::before: Добавление
графической
метки(✔ или ❌) перед -->
< !--    неотмеченными
и
отмеченными
задачами. -->

< !--3.
Шаг №3: Разметка
и
отображение
задач -->

< !-- < body > -->
< !-- < ul > -->
< !-- < li


class ="checked" > Выполнить задачу 1 < / li > -->

< !-- < li > Выполнить
задачу
2 < / li > -->
< !-- < li


class ="checked" > Выполнить задачу 3 < / li > -->

< !-- < li > Выполнить
задачу
4 < / li > -->
< !-- < li > Выполнить
задачу
5 < / li > -->
< !-- < / ul > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ul >: Начало
упорядоченного
списка
задач. -->
< !-- < li


class ="checked" > Выполнить задачу 1 < / li > и аналогичные элементы: -

->
< !--    Элементы
списка
с
текстом
задач
и
классами
для
стилизации(отмеченные
и
неотмеченные).-->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Самый
простой
Вариант - Вариант
2:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "ru" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "styles.css" >
       < title > Список
задач < / title >
          < / head >
              < style >
*{
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 1a1a1a;
}

ul
{
list - style - type: none;
padding: 0;
}

li
{
padding: 10
px;
margin: 5
px;
border - radius: 5
px;
}

li: nth - child(odd)
{
background - color:  # d9edf7; /* оттенок для четных элементов */
}

li: nth - child(even)
{
background - color:  # dff0d8; /* оттенок для нечетных элементов */
}

.checked
{
background - color:  # ccc !important; /* серый фон для отмеченных элементов */
text - decoration: line - through;
}

.delete
{
color:  # d9534f; /* красный цвет крестика */
cursor: pointer;
}

.delete: hover
{
text - decoration: underline;
}
< / style >
    < body >
    < ul
id = "taskList" >
     < li


class ="checked" > ✔️ Выполненная задача < span class ="delete" > ❌ < / span > < / li >

< li > Невыполненная
задача < span


class ="delete" > ❌ < / span > < / li >

< li


class ="checked" > ✔️ Еще одна выполненная задача < span class ="delete" > ❌ < / span > < / li >

< li > Еще
одна
невыполненная
задача < span


class ="delete" > ❌ < / span > < / li >

< / ul >

< script >
document.addEventListener('DOMContentLoaded', function()
{
    const
taskList = document.getElementById('taskList');
const
deleteButtons = document.querySelectorAll('.delete');

deleteButtons.forEach(button= > {
    button.addEventListener('click', function()
{
    const
listItem = this.parentNode;
listItem.parentNode.removeChild(listItem);
});
});
});
< / script >
< / body >
< / html >

< !--  ########################################################################################## -->

< !--1.
Шаг №1: Настройка
и
подключение
стилей -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "styles.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "ru" >: Начало
HTML - разметки
с
указанием
языка(русский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего -->
< !--    файла
стилей
"styles.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости -->
< !--    с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования -->
< !--    для
устройств
с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
определение
задач -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 1a1a1a;-->
< !--}-->

< !--  ul
{-->
< !--    list - style - type: none;
-->
< !--    padding: 0;
-->
< !--}-->

< !--  li
{-->
< !--    padding: 10
px;
-->
< !--    margin: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--}-->

< !--  li: nth - child(odd)
{-->
< !--    background - color:  # d9edf7; /* оттенок для четных элементов */-->
< !--}-->

< !--  li: nth - child(even)
{-->
< !--    background - color:  # dff0d8; /* оттенок для нечетных элементов */-->
< !--}-->

< !--.checked
{-->
< !--    background - color:  # ccc !important; /* серый фон для отмеченных элементов */-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !--.delete
{-->
< !--    color:  # d9534f; /* красный цвет крестика */-->
< !--    cursor: pointer;
-->
< !--}-->

< !--.delete: hover
{-->
< !--    text - decoration: underline;
-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: Задание
общих
стилей
для
всех
элементов -->
< !--    на
странице, таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ul(списка): Сброс
стилей
маркера
списка
и -->
< !--    установка
отступов. -->
< !--Стили
для
li(элементов
списка): Определение
внешнего
и -->
< !--    внутреннего
отступов, а
также
скругления
углов. -->
< !--Стили
для
li: nth - child(odd)
и
li: nth - child(even): Определение
цветов
фона -->
< !--    для
четных
и
нечетных
элементов
списка. -->
< !--Стили
для.checked: Определение
стилей
для
отмеченных
элементов, включая -->
< !--    цвет
фона
и
зачеркивание
текста.Важность(!important) используется, -->
< !--    чтобы
переопределить
стандартные
стили. -->
< !--Стили
для.delete
и.delete: hover: Определение
стилей
для
кнопки
удаления -->
< !--    задачи, включая
цвет
и
эффект
при
наведении. -->

< !--3.
Шаг №3: Разметка
и
удаление
задач -->

< !-- < body > -->
< !-- < ul
id = "taskList" > -->
< !-- < li


class ="checked" > ✔️ Выполненная задача < span class ="delete" > ❌ < / span > < / li > -->

< !-- < li > Невыполненная
задача < span


class ="delete" > ❌ < / span > < / li > -->

< !-- < li


class ="checked" > ✔️ Еще одна выполненная задача < span class ="delete" > ❌ < / span > < / li > -->

< !-- < li > Еще
одна
невыполненная
задача < span


class ="delete" > ❌ < / span > < / li > -->

< !-- < / ul > -->

< !-- < script > -->
< !--        document.addEventListener('DOMContentLoaded', function()
{-->
< !--            const
taskList = document.getElementById('taskList');
-->
< !--            const
deleteButtons = document.querySelectorAll('.delete');
-->

< !--            deleteButtons.forEach(button= > {-->
< !--                button.addEventListener('click', function()
{-->
< !--                    const
listItem = this.parentNode;
-->
< !--                    listItem.parentNode.removeChild(listItem);
-->
< !--});-->
< !--});-->
< !--});-->
< !-- < / script > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ul
id = "taskList" >: Начало
упорядоченного
списка
задач
с
идентификатором
"taskList". -->
< !-- < li > элементы: Элементы
списка
с
текстом
задач
и
кнопкой
удаления( < span


class ="delete" > ❌ < / span > ).-->
< !--Скрипт для удаления задач: JavaScript - скрипт, использующий


событие
DOMContentLoaded
для -->
< !--    обработки
события
клика
на
кнопке
удаления.Когда
кнопка
нажимается, -->
< !--    соответствующий
элемент
списка
удаляется
из
DOM - структуры. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Вариант
3:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "task_list.css" >
       < title > Список
задач < / title >
          < / head >
              < style >

*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 333;
}

ul
{
    list - style: none;
padding: 0;
width: 300
px;
}

li
{
    padding: 10px;
margin - bottom: 5
px;
border - radius: 5
px;
}

/ *Четные
элементы
списка * /
li: nth - child(even)
{
    background - color:  # d3e0ea;
}

/ * Нечетные
элементы
списка * /
li: nth - child(odd)
{
    background - color:  # a8bdd4;
}

/ * Отмеченные
элементы * /
li.checked
{
    background - color:  # ccc;
        text - decoration: line - through;
}

/ *Неотмеченные
элементы * /
li: not (.checked)::before
{
    content: "❌";
color: red;
margin - right: 5
px;
}

/ *Отмеченные
элементы * /
li.checked::before
{
    content: "✔";
color: green;
margin - right: 5
px;
}

< / style >
    < body >
    < ul
id = "taskList" > < / ul >

                      < script >
// Случайные
названия
задач
const
taskNames = [
    "Подготовить отчет",
    "Заказать канцтовары",
    "Провести совещание",
    "Написать письмо клиенту",
    "Подготовить презентацию",
    "Проверить почту",
    "Организовать встречу",
    "Сделать звонок",
    "Закончить проект",
    "Подготовить план маркетинга"
];

// Генерация
случайных
задач
const
taskList = document.getElementById("taskList");

for (let i = 0; i < 10; i++)
{
    const
randomIndex = Math.floor(Math.random() * taskNames.length);
const
taskName = taskNames[randomIndex];

const
li = document.createElement("li");
li.textContent = taskName;

// Случайным
образом
отмечаем
некоторые
задачи
if (Math.random() > 0.5)
{
li.classList.add("checked");
}

taskList.appendChild(li);
}
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--Шапка
документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Это
объявление
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включая
мета - теги, -->
< !--    подключение
внешнего
файла
стилей
"task_list.css"
и
настройки
визуализации
на
различных
устройствах. -->

< !--    -->
< !--    Стили: -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 333;-->
< !--}-->

< !--ul
{-->
< !--    list - style: none;
-->
< !--    padding: 0;
-->
< !--    width: 300
px;
-->
< !--}-->

< !--li
{-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--}-->

< !--li: nth - child(even)
{-->
< !--    background - color:  # d3e0ea;-->
< !--}-->

< !--li: nth - child(odd)
{-->
< !--    background - color:  # a8bdd4;-->
< !--}-->

< !--li.checked
{-->
< !--    background - color:  # ccc;-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !--li: not (.checked)::before
{-->
< !--    content: "❌";
-->
< !--    color: red;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !--li.checked::before
{-->
< !--    content: "✔";
-->
< !--    color: green;
-->
< !--    margin - right: 5
px;
-->
< !--}-->
< !-- < / style > -->
< !--    -->
< !--Общие
стили: -->
< !--    Задают
общие
стили
для
всех
элементов
на
странице, такие
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ul(списка): -->
< !--    Определяют
стили
для
упорядоченного
списка, убирая
стандартные
маркеры
и
устанавливая
отступы. -->
< !--Стили
для
li(элементов
списка): -->
< !--    Задают
внешний
и
внутренний
отступы, а
также
скругление
углов. -->
< !--Стили
для
четных
и
нечетных
элементов: -->
< !--    Задают
цвет
фона
для
четных
и
нечетных
элементов
списка. -->
< !--Стили
для
отмеченных
и
неотмеченных
элементов: -->
< !--    Определяют
цвет
фона, стиль
зачеркивания
и
добавляют
графические -->
< !--    символы(✔ и ❌) для
отмеченных
и
неотмеченных
элементов
соответственно. -->

< !--    -->
< !--Тело
документа: -->

< !-- < body > -->
< !-- < ul
id = "taskList" > < / ul > -->

< !-- < script > -->
< !-- // Случайные
названия
задач -->
< !--        const
taskNames = [-->
< !--            "Подготовить отчет", -->
< !--            "Заказать канцтовары", -->
< !--            "Провести совещание", -->
< !--            "Написать письмо клиенту", -->
< !--            "Подготовить презентацию", -->
< !--            "Проверить почту", -->
< !--            "Организовать встречу", -->
< !--            "Сделать звонок", -->
< !--            "Закончить проект", -->
< !--            "Подготовить план маркетинга" -->
< !--];-->

< !-- // Генерация
случайных
задач -->
< !--        const
taskList = document.getElementById("taskList");
-->

< !--
for (let i = 0; i < 10; i++) {-->
< !--            const randomIndex = Math.floor(Math.random() * taskNames.length);-->
< !--            const taskName = taskNames[randomIndex];-->

< !--            const li = document.createElement("li");-->
< !--            li.textContent = taskName;-->

< !-- // Случайным образом отмечаем некоторые задачи-->
< !-- if (Math.random() > 0.5) {-->
< !--                li.classList.add("checked");
-->
< !--}-->

< !--            taskList.appendChild(li);
-->
< !--}-->
< !-- < / script > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Открывает
тело
документа. -->
< !-- < ul
id = "taskList" > < / ul >: Создает
пустой
упорядоченный
список
с
идентификатором
"taskList". -->

< !--JavaScript - скрипт: -->

< !-- < script > -->
< !-- // Случайные
названия
задач -->
< !--    const
taskNames = [-->
< !--        "Подготовить отчет", -->
< !--        "Заказать канцтовары", -->
< !--        "Провести совещание", -->
< !--        "Написать письмо клиенту", -->
< !--        "Подготовить презентацию", -->
< !--        "Проверить почту", -->
< !--        "Организовать встречу", -->
< !--        "Сделать звонок", -->
< !--        "Закончить проект", -->
< !--        "Подготовить план маркетинга" -->
< !--];-->

< !-- // Генерация
случайных
задач -->
< !--    const
taskList = document.getElementById("taskList");
-->

< !--
for (let i = 0; i < 10; i++) {-->
< !--        const randomIndex = Math.floor(Math.random() * taskNames.length);-->
< !--        const taskName = taskNames[randomIndex];-->

< !--        const li = document.createElement("li");-->
< !--        li.textContent = taskName;-->

< !-- // Случайным образом отмечаем некоторые задачи-->
< !-- if (Math.random() > 0.5) {-->
< !--            li.classList.add("checked");
-->
< !--}-->

< !--        taskList.appendChild(li);
-->
< !--}-->
< !-- < / script > -->

< !--Генерация
задач: -->
< !--Создает
массив
taskNames
с
возможными
именами
задач. -->
< !--Создание
элементов
списка: -->
< !--В
цикле
создается
элемент
li
для
каждой
задачи, случайным
образом
выбранной
из
taskNames. -->
< !--Отметка
выполненных
задач: -->
< !--Если
случайное
число
больше
0.5, задача
отмечается
как
выполненная, добавляя
класс
"checked". -->
< !--Добавление
элемента
в
список: -->
< !--Каждый
элемент
списка
добавляется
в
taskList. -->

< !--Итог: -->
< !--Этот
код
демонстрирует
использование
HTML, CSS
и
JavaScript
для -->
< !--создания
динамического
списка
задач
с
различными
стилями, -->
< !--включая
случайное
выделение
выполненных
задач. -->

< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Вариант
3:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "task_list.css" >
       < title > Список
задач < / title >
          < / head >
              < style >

*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 333;
}

ul
{
    list - style: none;
padding: 0;
width: 300
px;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
    opacity: 0;
}
to
{
    opacity: 1;
}
}

li
{
    padding: 10px;
margin - bottom: 5
px;
border - radius: 5
px;
opacity: 0;
animation: slideIn
0.5
s
ease - in -out
forwards;
}

@keyframes


slideIn
{
from

{
    opacity: 0;
transform: translateY(-20
px);
}
to
{
    opacity: 1;
transform: translateY(0);
}
}

/ *Четные
элементы
списка * /
li: nth - child(even)
{
    background - color:  # d3e0ea;
}

/ * Нечетные
элементы
списка * /
li: nth - child(odd)
{
    background - color:  # a8bdd4;
}

/ * Отмеченные
элементы * /
li.checked
{
    background - color:  # ccc;
        text - decoration: line - through;
}

/ *Неотмеченные
элементы * /
li: not (.checked)::before
{
    content: "❌";
color: red;
margin - right: 5
px;
}

/ *Отмеченные
элементы * /
li.checked::before
{
    content: "✔";
color: green;
margin - right: 5
px;
}

< / style >
    < body >
    < ul
id = "taskList" > < / ul >

                      < script >
// Случайные
названия
задач
const
taskNames = [
    "Подготовить отчет",
    "Заказать канцтовары",
    "Провести совещание",
    "Написать письмо клиенту",
    "Подготовить презентацию",
    "Проверить почту",
    "Организовать встречу",
    "Сделать звонок",
    "Закончить проект",
    "Подготовить план маркетинга"
];

// Генерация
случайных
задач
const
taskList = document.getElementById("taskList");

for (let i = 0; i < 10; i++)
{
    const
randomIndex = Math.floor(Math.random() * taskNames.length);
const
taskName = taskNames[randomIndex];

const
li = document.createElement("li");
li.textContent = taskName;

// Случайным
образом
отмечаем
некоторые
задачи
if (Math.random() > 0.5)
{
    li.classList.add("checked");
}

taskList.appendChild(li);
}
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->

              <!--Шаг №1: HTML - структура
документа -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->
< !-- < body > -->
< !-- < ul
id = "taskList" > < / ul > -->

< !-- < script > -->
< !-- // JavaScript - код
будет
вставлен
здесь -->
< !-- < / script > -->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание: -->

< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Начало
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа.Здесь
подключены
мета - теги, стили
и
указан
заголовок
страницы. -->
< !-- < meta
charset = "UTF-8" >: Указание
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Настройка
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования -->
< !--    для
мобильных
устройств. -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" >: Подключение
внешних
стилей
из
файла
task_list.css. -->
< !-- < title > Список
задач < / title >: Заголовок
страницы. -->
< !-- < body >: Открывающий
тег
тела
документа, где
будет
размещен
контент
страницы. -->
< !-- < ul
id = "taskList" > < / ul >: Пустой
список
задач
с
идентификатором
taskList. -->
< !-- < script >: Блок
JavaScript
для
динамического
создания
задач.Здесь
он
пока
пуст, -->
< !-- так
как
будет
заполнен
на
следующих
шагах. -->
< !-- < / body > и < / html >: Закрывающие
теги
для
тела
и
HTML - документа. -->

< !--Шаг №2: CSS - стили
для
общих
элементов -->

< !--Пример
выполнения
этого
шага: -->

< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 333;-->
< !--}-->

< !--Подробное
описание: -->

< !-- *: Устанавливает
стили
для
всех
элементов
на
странице. -->
< !--font - family: Задает
шрифт
текста. -->
< !--background - color: Устанавливает
цвет
фона. -->
< !--color: Задает
цвет
текста. -->

< !--Шаг №3: Стили
для
списка
задач -->

< !--Пример
выполнения
этого
шага: -->

< !--ul
{-->
< !--    list - style: none;
-->
< !--    padding: 0;
-->
< !--    width: 300
px;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !--Подробное
описание: -->

< !--ul: Стили
для
списка. -->
< !--list - style: Убирает
маркеры
списка. -->
< !--padding: Устанавливает
отступы
внутри
списка. -->
< !--width: Задает
ширину
списка. -->
< !--opacity: Устанавливает
начальную
прозрачность
списка. -->
< !--animation: Применяет
анимацию
fadeIn
для
плавного
появления
списка. -->

< !--Шаг №4: Анимации -->

< !--Пример
выполнения
этого
шага: -->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--Подробное
описание: -->

< !-- @ keyframes
fadeIn: Объявление
анимации
с
названием
fadeIn. -->
< !--
from: Начальное
состояние
анимации(прозрачность
0).-->
< !--to: Конечное
состояние
анимации(прозрачность
1).-->

< !--Шаг №5: Стили
для
задач -->

< !--Пример
выполнения
этого
шага: -->

< !--li
{-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--    opacity: 0;
-->
< !--    animation: slideIn
0.5
s
ease - in -out
forwards;
-->
< !--}-->

< !--Подробное
описание: -->

< !--li: Стили
для
элементов
списка
задач. -->
< !--padding: Задает
внутренний
отступ
для
задач. -->
< !--margin - bottom: Задает
отступ
снизу
между
задачами. -->
< !--border - radius: Округляет
углы
задач. -->
< !--opacity: Устанавливает
начальную
прозрачность
задач. -->
< !--animation: Применяет
анимацию
slideIn
для
плавного
появления
задач. -->

< !--Шаг №6: Анимация
появления
задач -->

< !--Пример
выполнения
этого
шага: -->

< !-- @ keyframes
slideIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--        transform: translateY(-20
px);-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--        transform: translateY(0);
-->
< !--}-->
< !--}-->

< !--Подробное
описание: -->

< !-- @ keyframes
slideIn: Объявление
анимации
с
названием
slideIn. -->
< !--
from: Начальное
состояние
анимации(прозрачность
0, смещение
вверх
на
20
пикселей).-->
< !--to: Конечное
состояние
анимации(прозрачность
1, без
смещения).-->

< !--Шаг №7: Стили
для
четных
и
нечетных
элементов -->

< !--Пример
выполнения
этого
шага: -->

< !--li: nth - child(even)
{-->
< !--    background - color:  # d3e0ea;-->
< !--}-->

< !--li: nth - child(odd)
{-->
< !--    background - color:  # a8bdd4;-->
< !--}-->

< !--Подробное
описание: -->

< !--li: nth - child(even): Стили
для
четных
элементов
списка(задается
цвет
фона).-->
< !--li: nth - child(odd): Стили
для
нечетных
элементов
списка(задается
другой
цвет
фона).-->

< !--Шаг №8: Стили
для
отмеченных
и
неотмеченных
элементов -->

< !--Пример
выполнения
этого
шага: -->

< !--li.checked
{-->
< !--    background - color:  # ccc;-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !--li: not (.checked)::before
{-->
< !--    content: "❌";
-->
< !--    color: red;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !--li.checked::before
{-->
< !--    content: "✔";
-->
< !--    color: green;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !--Подробное
описание: -->

< !--li.checked: Стили
для
отмеченных
задач(задается
цвет
фона
и
зачеркивание
текста).-->
< !--li: not (.checked)::before: Стили
для
неотмеченных
задач(вставляется
символ
"❌"
перед
текстом).-->
< !--li.checked::before: Стили
для
отмеченных
задач(вставляется
символ
"✔"
перед
текстом).-->

< !--Шаг №9: JavaScript
для
генерации
задач -->

< !--Пример
выполнения
этого
шага: -->

< !--const
taskNames = [-->
< !--    "Подготовить отчет", -->
< !--    "Заказать канцтовары", -->
< !--    "Провести совещание", -->
< !--    "Написать письмо клиенту", -->
< !--    "Подготовить презентацию", -->
< !--    "Проверить почту", -->
< !--    "Организовать встречу", -->
< !--    "Сделать звонок", -->
< !--    "Закончить проект", -->
< !--    "Подготовить план маркетинга" -->
< !--];-->

< !--const
taskList = document.getElementById("taskList");
-->

< !--
for (let i = 0; i < 10; i++) {-->
< !--    const randomIndex = Math.floor(Math.random() * taskNames.length);-->
< !--    const taskName = taskNames[randomIndex];-->

< !--    const li = document.createElement("li");-->
< !--    li.textContent = taskName;-->

< !-- if (Math.random() > 0.5) {-->
< !--        li.classList.add("checked");-->
< !--}-->

< !--    taskList.appendChild(li);-->
< !--}-->

< !--Подробное описание:-->

< !--const
taskNames: Массив
строк
с
названиями
задач. -->
< !--const
taskList = document.getElementById("taskList"): Получение
ссылки
на
элемент
списка
задач
по
его
идентификатору. -->
< !--
for (let i = 0; i < 10; i++): Цикл
для
создания
10
задач. -->
< !--const
randomIndex = Math.floor(Math.random() * taskNames.length): Генерация
случайного
индекса
из
массива
taskNames. -->
< !--const
taskName = taskNames[randomIndex]: Получение
случайного
названия
задачи. -->
< !--const
li = document.createElement("li"): Создание
элемента
списка( < li >).-->
< !--li.textContent = taskName: Установка
текста
задачи. -->
< !-- if (Math.random() > 0.5)
{li.classList.add("checked");}: Случайное
добавление
класса
"checked"
для
отмеченных
задач. -->
< !--taskList.appendChild(li): Добавление
созданной
задачи
в
список. -->

< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
Думаю
что
одного - ВАРИАНТА
будет
достаточно
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "styles.css" >
       < style >
:root
{
    --season - font - size: 35px;
--season - font - color:  # 333;
--month - font - size: 16
px;
--month - font - color:  # 555;
--month - icon - size: 50
px;
--month - icon - margin: 2, 5
px;
}

ul
{
    list - style - type: none;
margin: 0;
padding: 0;
}

li
{
    margin - bottom: 10px;
font - size: var(--season - font - size);
color: var(--season - font - color);
}

.month - icon
{
    font - size: var(--month - icon - size);
margin - right: var(--month - icon - margin);
}

ul
ul
{
    margin - left: 25px;
}

ul
ul
ul
{
    margin - left: 50px;
}
< / style >
    < / head >
        < body >
        < ul >
        < li >
        Времена
года
< ul >
< li >
Зима
< ul >
< li > < span


class ="month-icon" > 🎄 < / span > Декабрь < / li >

< li > < span


class ="month-icon" > ☃️ < / span > Январь < / li >

< li > < span


class ="month-icon" > ❄️ < / span > Февраль < / li >

< / ul >
< / li >
< li >
Весна
< ul >
< li > < span


class ="month-icon" > 💐 < / span > Март < / li >

< li > < span


class ="month-icon" > 🌷 < / span > Апрель < / li >

< li > < span


class ="month-icon" > 🌺 < / span > Май < / li >

< / ul >
< / li >
< li >
Лето
< ul >
< li > < span


class ="month-icon" > 🍉 < / span > Июнь < / li >

< li > < span


class ="month-icon" > 🌴 < / span > Июль < / li >

< li > < span


class ="month-icon" > 🎣 < / span > Август < / li >

< / ul >
< / li >
< li >
Осень
< ul >
< li > < span


class ="month-icon" > 🍃 < / span > Сентябрь < / li >

< li > < span


class ="month-icon" > 🍁 < / span > Октябрь < / li >

< li > < span


class ="month-icon" > 🍂 < / span > Ноябрь < / li >

< / ul >
< / li >
< / ul >
< / li >
< / ul >
< / body >
< / html >
< !--  ########################################################################################## -->

< !--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "styles.css" > -->
< !-- < style > -->
< !-- & lt;! & ndash;
CSS - стили
встроены
непосредственно
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < link
rel = "stylesheet"
href = "styles.css" >: Подключение
внешнего
файла
стилей(styles.css). -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Объявление
переменных
CSS(в
теге < style >):-->

< !--: root
{-->
< !-- &  # 45;&#45;season-font-size: 35px;-->
   <!-- &  # 45;&#45;season-font-color: #333;-->
     <!-- &  # 45;&#45;month-font-size: 16px;-->
       <!-- &  # 45;&#45;month-font-color: #555;-->
         <!-- &  # 45;&#45;month-icon-size: 50px;-->
           <!-- &  # 45;&#45;month-icon-margin: 2.5px;-->
             <!--}-->

< !--Описание: -->
< !--: root: Псевдокласс, представляющий
корневой
элемент
документа, используется
для
объявления
пользовательских
CSS - переменных. -->
< !--Переменные, устанавливающие
размеры
шрифта, цвета
и
размеры
иконок
для
времен
года
и
месяцев. -->

< !--3.
Стили
CSS: -->

< !--ul
{-->
< !--    list - style - type: none;
-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--li
{-->
< !--    margin - bottom: 10
px;
-->
< !--    font - size: var( &  # 45;&#45;season-font-size);-->
                      <!--    color: var( &  # 45;&#45;season-font-color);-->
                                     <!--}-->

< !--.month - icon
{-->
< !--    font - size: var( &  # 45;&#45;month-icon-size);-->
                      <!--    margin - right: var( &  # 45;&#45;month-icon-margin);-->
                                              <!--}-->

< !--ul
ul
{-->
< !--    margin - left: 25
px;
-->
< !--}-->

< !--ul
ul
ul
{-->
< !--    margin - left: 50
px;
-->
< !--}-->

< !--Описание: -->
< !--Удаление
маркеров
списка
и
отступов. -->
< !--Настройка
отступов
и
размеров
шрифта
для
элементов
списка. -->
< !--Определение
стилей
для
класса.month - icon, который
отвечает
за
размер
иконок
месяцев. -->
< !--Настройка
вложенных
списков
с
различными
отступами. -->

< !--4.
Структура
HTML - документа: -->
< !-- < ul > -->
< !-- < li > -->
< !--        Времена
года -->
< !-- < ul > -->
< !-- & lt;! & ndash;
Вложенные
списки
для
каждого
времени
года & ndash; & gt;
-->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->

< !--Описание: -->
< !-- < ul >: Начало
списка
без
маркеров. -->
< !-- < li >: Элемент
списка
для
времен
года. -->
< !--Вложенные < ul > и < li > для
каждого
времени
года
и
их
месяцев. -->

< !--5.
Иконки
месяцев
в
эмодзи: -->
< !-- < li > < span


class ="month-icon" > 🎄 < / span > Декабрь < / li > -->

< !-- < li > < span


class ="month-icon" > ☃️ < / span > Январь < / li > -->

< !-- & lt;! & ndash;
...
остальные
месяцы... & ndash; & gt;
-->

< !--Описание: -->
< !-- < span


class ="month-icon" >: Элемент - контейнер


для
применения
стилей
к
иконке
месяца. -->
< !--Иконки
месяцев
представлены
символами
эмодзи. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №1 - Чем
проще, тем
лучше.
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
font - family: Arial, sans - serif;
}

h1
{
font - weight: bold;
}

h2
{
margin - left: 20
px;
}

h3
{
margin - left: 40
px;
}

h4
{
margin - left: 60
px;
}

h5
{
margin - left: 80
px;
}
< / style >
    < / head >
        < body >
        < h1 > Contents < / h1 >

                            < h2 > I.History < / h2 >
                                                 < h3 > 1.
Development < / h3 >
                < h3 > 2.
HTML
versions
timeline < / h3 >
             < h4 > a.HTML
draft
version
timeline < / h4 >
             < h4 > b.XHTML
versions < / h4 >

             < h2 > II.Markup < / h2 >
                                  < h3 > 1.
Elements < / h3 >
             < h4 > a.Element
examples < / h4 >
             < h4 > b.Attributes < / h4 >
                                     < h3 > 2.
Character and entity
references < / h3 >
               < h3 > 3.
Data
types < / h3 >
          < h3 > 4.
Document
type declaration < / h3 >

                     < h2 > III.Semantic
HTML < / h2 >

         < h2 > IV.Delivery < / h2 >
                                < h3 > 1.
HTTP < / h3 >
         < / body >
             < / html >
                 <!--  ############################################################################################################## -->

                   <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: Arial, sans - serif;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--}-->

< !--h2
{-->
< !--    margin - left: 20
px;
-->
< !--}-->

< !--h3
{-->
< !--    margin - left: 40
px;
-->
< !--}-->

< !--h4
{-->
< !--    margin - left: 60
px;
-->
< !--}-->

< !--h5
{-->
< !--    margin - left: 80
px;
-->
< !--}-->

< !--Описание: -->
< !--font - family: Arial, sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--Настройка
отступов
для
заголовков
разного
уровня
с
использованием
margin - left. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < h2 > I.History < / h2 > -->
< !-- < h3 > 1.
Development < / h3 > -->
< !-- < h3 > 2.
HTML
versions
timeline < / h3 > -->
< !-- < h4 > a.HTML
draft
version
timeline < / h4 > -->
< !-- < h4 > b.XHTML
versions < / h4 > -->

< !-- < h2 > II.Markup < / h2 > -->
< !-- < h3 > 1.
Elements < / h3 > -->
< !-- < h4 > a.Element
examples < / h4 > -->
< !-- < h4 > b.Attributes < / h4 > -->
< !-- < h3 > 2.
Character and entity
references < / h3 > -->
< !-- < h3 > 3.
Data
types < / h3 > -->
< !-- < h3 > 4.
Document
type declaration < / h3 > -->

< !-- < h2 > III.Semantic
HTML < / h2 > -->

< !-- < h2 > IV.Delivery < / h2 > -->
< !-- < h3 > 1.
HTTP < / h3 > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >, < h2 >, < h3 >, < h4 >, < h5 >), чтобы
структурировать
содержание
страницы. -->
< !--Каждый
заголовок
имеет
отступ, который
указан
в
стилях
CSS. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--
ВАРИАНТ №2 - Уже
лучше.
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
    font - family: 'Roboto', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 0;
padding: 0;
}

h1
{
    font - weight: bold;
color:  # 0066cc;
background - color:  # f2f2f2;
padding: 10
px;
margin: 0;
}

h2, h3, h4, h5
{
    margin - left: 20px;
}

h2
{
    color:  # 009900;
        font - weight: bold;
}

h3
{
    color:  # cc3300;
        font - weight: normal;
}

h4
{
    color:  # 990099;
        font - weight: normal;
}

h5
{
    color:  # ff6600;
        font - weight: normal;
}

h4, h5
{
    margin - left: 40px;
}
< / style >
    < / head >
        < body >
        < h1 > Contents < / h1 >

                            < h2 > I.History < / h2 >
                                                 < h3 > 1.
Development < / h3 >
                < h3 > 2.
HTML
versions
timeline < / h3 >
             < h4 > a.HTML
draft
version
timeline < / h4 >
             < h4 > b.XHTML
versions < / h4 >

             < h2 > II.Markup < / h2 >
                                  < h3 > 1.
Elements < / h3 >
             < h4 > a.Element
examples < / h4 >
             < h4 > b.Attributes < / h4 >
                                     < h3 > 2.
Character and entity
references < / h3 >
               < h3 > 3.
Data
types < / h3 >
          < h3 > 4.
Document
type declaration < / h3 >

                     < h2 > III.Semantic
HTML < / h2 >

         < h2 > IV.Delivery < / h2 >
                                < h3 > 1.
HTTP < / h3 >
         < / body >
             < / html >
                 <!--  ############################################################################################################## -->

                   <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: 'Roboto', sans - serif;
-->
< !--    background - color:  # f8f8f8;-->
< !--    color:  # 333;-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--    color:  # 0066cc;-->
< !--    background - color:  # f2f2f2;-->
< !--    padding: 10
px;
-->
< !--    margin: 0;
-->
< !--}-->

< !--h2, h3, h4, h5
{-->
< !--    margin - left: 20
px;
-->
< !--}-->

< !--h2
{-->
< !--    color:  # 009900;-->
< !--    font - weight: bold;
-->
< !--}-->

< !--h3
{-->
< !--    color:  # cc3300;-->
< !--    font - weight: normal;
-->
< !--}-->

< !--h4
{-->
< !--    color:  # 990099;-->
< !--    font - weight: normal;
-->
< !--}-->

< !--h5
{-->
< !--    color:  # ff6600;-->
< !--    font - weight: normal;
-->
< !--}-->

< !--h4, h5
{-->
< !--    margin - left: 40
px;
-->
< !--}-->

< !--Описание: -->
< !--font - family: 'Roboto', sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--background - color:  # f8f8f8;: Задание цвета фона.-->
< !--color:  # 333;: Установка цвета текста.-->
< !--margin: 0;
padding: 0;: Убирают
отступы
и
внутренний
отступ
для
тела
документа. -->
< !--Стили
для
заголовков
разных
уровней
с
различными
цветами
и
отступами. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < h2 > I.History < / h2 > -->
< !-- < h3 > 1.
Development < / h3 > -->
< !-- < h3 > 2.
HTML
versions
timeline < / h3 > -->
< !-- < h4 > a.HTML
draft
version
timeline < / h4 > -->
< !-- < h4 > b.XHTML
versions < / h4 > -->

< !-- < h2 > II.Markup < / h2 > -->
< !-- < h3 > 1.
Elements < / h3 > -->
< !-- < h4 > a.Element
examples < / h4 > -->
< !-- < h4 > b.Attributes < / h4 > -->
< !-- < h3 > 2.
Character and entity
references < / h3 > -->
< !-- < h3 > 3.
Data
types < / h3 > -->
< !-- < h3 > 4.
Document
type declaration < / h3 > -->

< !-- < h2 > III.Semantic
HTML < / h2 > -->

< !-- < h2 > IV.Delivery < / h2 > -->
< !-- < h3 > 1.
HTTP < / h3 > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >, < h2 >, < h3 >, < h4 >, < h5 >), чтобы
структурировать
содержание
страницы. -->
< !--Каждый
заголовок
имеет
определенные
стили, включая
цвет
текста
и
фона, а
также
отступы, указанные
в
стилях
CSS. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №3 - Красота.
         -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
    font - family: 'Arial', sans - serif;
background - color:  # 282c35;
color:  # fff;
margin: 0;
padding: 0;
display: flex;
justify - content: center;
align - items: center;
height: 100
vh;
}

h1
{
    font - weight: bold;
color:  # 61dafb;
text - shadow: 2
px
2
px
4
px  # 333;
animation: rainbowAnimation
5
s
infinite;
}

ul
{
    list - style - type: none;
margin: 0;
padding: 0;
}

li
{
    margin - left: 20px;
margin - bottom: 10
px;
padding: 8
px;
background - color:  # 61dafb;
border - radius: 5
px;
transition: background - color
0.3
s
ease;
}

li: hover
{
    background - color:  # 4fa3d1;
}

@ keyframes
rainbowAnimation
{
    0 % {color:  # ff0000; }
             25 % {color:  # ff9900; }
50 % {color:  # ffff00; }
          75 % {color:  # 33cc33; }
                    100 % {color:  # 3366ff; }
                           }
                    < / style >
          < / head >
< body >
< h1 > Contents < / h1 >

                    < ul >
                    < li > I.History
                    < ul >
                    < li > 1.
Development < / li >
                < li > 2.
HTML
versions
timeline
< ul >
< li > a.HTML
draft
version
timeline < / li >
             < li > b.XHTML
versions < / li >
             < / ul >
                 < / li >
                     < / ul >
                         < / li >

                             < li > II.Markup
                             < ul >
                             < li > 1.
Elements
< ul >
< li > a.Element
examples < / li >
             < li > b.Attributes < / li >
                                     < / ul >
                                         < / li >
                                             < li > 2.
Character and entity
references < / li >
               < li > 3.
Data
types < / li >
          < li > 4.
Document
type declaration < / li >
                     < / ul >
                         < / li >

                             < li > III.Semantic
HTML < / li >

         < li > IV.Delivery
         < ul >
         < li > 1.
HTTP < / li >
         < / ul >
             < / li >
                 < / ul >
                     < / body >
                         < / html >
                             <!--  ############################################################################################################## -->

                               <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для -->
< !--    мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: 'Arial', sans - serif;
-->
< !--    background - color:  # 282c35;-->
< !--    color:  # fff;-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--    display: flex;
-->
< !--    justify - content: center;
-->
< !--    align - items: center;
-->
< !--    height: 100
vh;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--    color:  # 61dafb;-->
< !--    text - shadow: 2
px
2
px
4
px  # 333;-->
<!--    animation: rainbowAnimation
5
s
infinite;
-->
< !--}-->

< !--ul
{-->
< !--    list - style - type: none;
-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--li
{-->
< !--    margin - left: 20
px;
-->
< !--    margin - bottom: 10
px;
-->
< !--    padding: 8
px;
-->
< !--    background - color:  # 61dafb;-->
< !--    border - radius: 5
px;
-->
< !--    transition: background - color
0.3
s
ease;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # 4fa3d1;-->
< !--}-->

< !-- @ keyframes
rainbowAnimation
{-->
< !--    0 % {color:  # ff0000; }-->
   <!--    25 % {color:  # ff9900; }-->
     <!--    50 % {color:  # ffff00; }-->
       <!--    75 % {color:  # 33cc33; }-->
         <!--    100 % {color:  # 3366ff; }-->
           <!--}-->

< !--Описание: -->
< !--font - family: 'Arial', sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--background - color:  # 282c35;: Задание цвета фона.-->
< !--color:  # fff;: Установка цвета текста.-->
< !--margin: 0;
padding: 0;: Убирают
отступы
и
внутренний
отступ
для
тела
документа. -->
< !--display: flex;
justify - content: center;
align - items: center;: Выравнивание
содержимого
по
центру. -->
< !--height: 100
vh;: Задание
высоты
в
100 % от
высоты
видимой
области
окна. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < ul > -->
< !-- < li > I.History -->
< !-- < ul > -->
< !-- < li > 1.
Development < / li > -->
< !-- < li > 2.
HTML
versions
timeline -->
< !-- < ul > -->
< !-- < li > a.HTML
draft
version
timeline < / li > -->
< !-- < li > b.XHTML
versions < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > II.Markup -->
< !-- < ul > -->
< !-- < li > 1.
Elements -->
< !-- < ul > -->
< !-- < li > a.Element
examples < / li > -->
< !-- < li > b.Attributes < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < li > 2.
Character and entity
references < / li > -->
< !-- < li > 3.
Data
types < / li > -->
< !-- < li > 4.
Document
type declaration < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > III.Semantic
HTML < / li > -->

< !-- < li > IV.Delivery -->
< !-- < ul > -->
< !-- < li > 1.
HTTP < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >) и
маркированный
список( < ul >, < li >), чтобы
структурировать -->
< !--содержание
страницы. -->
< !--Каждый
заголовок
и
элемент
списка
имеют
определенные
стили, включая
цвет
текста
и
фона. -->
< !--Например, при
наведении
на
элемент
списка(li: hover), цвет
фона
изменяется
с
анимацией. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №4 - Красота(Светлый
вариант).
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
    font - family: 'Arial', sans - serif;
background - color:  # f0f0f0;
color:  # 333;
margin: 0;
padding: 0;
display: flex;
justify - content: center;
align - items: center;
height: 100
vh;
}

h1
{
    font - weight: bold;
color:  # 333;
}

ul
{
    list - style - type: none;
margin: 0;
padding: 0;
}

li
{
    margin - left: 20px;
margin - bottom: 10
px;
padding: 8
px;
background - color:  # fff;
border - radius: 5
px;
box - shadow: 0
2
px
4
px
rgba(0, 0, 0, 0.1);
transition: background - color
0.3
s
ease;
}

li: hover
{
    background - color:  # f0f0f0;
}
< / style >
    < / head >
        < body >
        < h1 > Contents < / h1 >

                            < ul >
                            < li > I.History
                            < ul >
                            < li > 1.
Development < / li >
                < li > 2.
HTML
versions
timeline
< ul >
< li > a.HTML
draft
version
timeline < / li >
             < li > b.XHTML
versions < / li >
             < / ul >
                 < / li >
                     < / ul >
                         < / li >

                             < li > II.Markup
                             < ul >
                             < li > 1.
Elements
< ul >
< li > a.Element
examples < / li >
             < li > b.Attributes < / li >
                                     < / ul >
                                         < / li >
                                             < li > 2.
Character and entity
references < / li >
               < li > 3.
Data
types < / li >
          < li > 4.
Document
type declaration < / li >
                     < / ul >
                         < / li >

                             < li > III.Semantic
HTML < / li >

         < li > IV.Delivery
         < ul >
         < li > 1.
HTTP < / li >
         < / ul >
             < / li >
                 < / ul >
                     < / body >
                         < / html >
                             <!--  ############################################################################################################## -->

                               <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: 'Arial', sans - serif;
-->
< !--    background - color:  # f0f0f0;-->
< !--    color:  # 333;-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--    display: flex;
-->
< !--    justify - content: center;
-->
< !--    align - items: center;
-->
< !--    height: 100
vh;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--    color:  # 333;-->
< !--}-->

< !--ul
{-->
< !--    list - style - type: none;
-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--li
{-->
< !--    margin - left: 20
px;
-->
< !--    margin - bottom: 10
px;
-->
< !--    padding: 8
px;
-->
< !--    background - color:  # fff;-->
< !--    border - radius: 5
px;
-->
< !--    box - shadow: 0
2
px
4
px
rgba(0, 0, 0, 0.1);
-->
< !--    transition: background - color
0.3
s
ease;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # f0f0f0;-->
< !--}-->

< !--Описание: -->
< !--font - family: 'Arial', sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--background - color:  # f0f0f0;: Задание цвета фона.-->
< !--color:  # 333;: Установка цвета текста.-->
< !--margin: 0;
padding: 0;: Убирают
отступы
и
внутренний
отступ
для
тела
документа. -->
< !--display: flex;
justify - content: center;
align - items: center;: Выравнивание
содержимого
по
центру. -->
< !--height: 100
vh;: Задание
высоты
в
100 % от
высоты
видимой
области
окна. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < ul > -->
< !-- < li > I.History -->
< !-- < ul > -->
< !-- < li > 1.
Development < / li > -->
< !-- < li > 2.
HTML
versions
timeline -->
< !-- < ul > -->
< !-- < li > a.HTML
draft
version
timeline < / li > -->
< !-- < li > b.XHTML
versions < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > II.Markup -->
< !-- < ul > -->
< !-- < li > 1.
Elements -->
< !-- < ul > -->
< !-- < li > a.Element
examples < / li > -->
< !-- < li > b.Attributes < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < li > 2.
Character and entity
references < / li > -->
< !-- < li > 3.
Data
types < / li > -->
< !-- < li > 4.
Document
type declaration < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > III.Semantic
HTML < / li > -->

< !-- < li > IV.Delivery -->
< !-- < ul > -->
< !-- < li > 1.
HTTP < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >) и
маркированный
список( < ul >, < li >), -->
< !--чтобы
структурировать
содержание
страницы. -->
< !--Каждый
заголовок
и
элемент
списка
имеют
определенные
стили, включая
цвет
текста, -->
< !--фона, тени
и
эффект
при
наведении. -->
< !--  ############################################################################################################## -->
   <!--  ############################################################################################################## -->
     <!--  ############################################################################################################## -->


# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
