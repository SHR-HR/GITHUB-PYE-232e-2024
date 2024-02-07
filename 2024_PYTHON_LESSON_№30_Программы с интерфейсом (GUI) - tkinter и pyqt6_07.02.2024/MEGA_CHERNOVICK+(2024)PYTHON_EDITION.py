# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней работы: 08-09 - ЯНВАРЯ 2024 года.
''''
'''
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы HTML и CSS !!! НО НАПИСАЛ Я КОД В PYCHARM - на PYTHON
'''
'''
Урок от 08.01.2024
Домашнее задание № 1 : Введение в Веб разработку
'''
'''
Выполните следующие задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
1. Установить Visual Studio Code.
2. Создать html файл с помощью VSCode.
3. Написать в html файле свои имя и фамилию.
4. Открыть в браузере с инструментами разработчика.
5. Сделать скриншот и отправить его вместе с вашим html файлом.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Скриншоты можно делать с помощью встроенного приложения Ножницы или с помощью бесплатной программы LightShot.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант - более правильный наверное. (Написанный на Python)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import re
import sqlite3
import html
import webbrowser

# Создаем подключение к базе данных (или создаем базу данных, если ее нет)
conn = sqlite3.connect('names_database.db')
cursor = conn.cursor()

# Создаем таблицу, если ее нет
cursor.execute('''
    CREATE TABLE IF NOT EXISTS names (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT,
        last_name TEXT
    )
''')
conn.commit()

def validate_name(name):
    # Проверка, что имя состоит только из букв, как в русском, так и в английском алфавите
    return re.match("^[а-яА-Яa-zA-Z]+$", name) is not None

# Получаем имя от пользователя с проверкой ввода
name = input("Введите ваше имя: ")
while not validate_name(name):
    print("Ошибка: Имя должно состоять только из букв русского или английского алфавита.")
    name = input("Введите ваше имя: ")

# Получаем фамилию от пользователя с проверкой ввода
surname = input("Введите вашу фамилию: ")
while not validate_name(surname):
    print("Ошибка: Фамилия должна состоять только из букв русского или английского алфавита.")
    surname = input("Введите вашу фамилию: ")

# Записываем введенные данные в базу данных
cursor.execute('INSERT INTO names (first_name, last_name) VALUES (?, ?)', (name, surname))
conn.commit()

# Записываем введенные данные в HTML-файл
with open('index.html', 'w', encoding='utf-8') as file:
    # Записываем начальный HTML-код
    file.write('''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Имя и Фамилия</title>
</head>
<body>
''')

    # Записываем введенные данные в HTML-файл с разными стилями
    file.write(f'    <h3>{html.escape(name)}</h3>\n')  # Стандартное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h2><i>{html.escape(name)}</i></h2>\n')  # Курсивное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h1><b>{html.escape(name)}</b></h1>\n')  # Жирное отображение

    file.write(f'    <h3>{html.escape(surname)}</h3>\n')  # Стандартное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h2><i>{html.escape(surname)}</i></h2>\n')  # Курсивное отображение (для наглядности чуть уменьшил шрифт)
    file.write(f'    <h1><b>{html.escape(surname)}</b></h1>\n')  # Жирное отображение

    # Добавляем рамку для JPG-изображения
    file.write('    <div style="border: 1px solid #ccc; padding: 10px; margin-top: 20px; width: 500px; height: 500px;">\n')
    file.write('        <img src="https://sun1-13.userapi.com/s/v1/if1/ygP2SHKIEpSAcXSnjPOnyo_29b2oYdOwC-tYTq_g_pxbMtjq6LFPpYHALe4I2vpHCNC4RmWX.jpg?size=200x200&quality=96&crop=290,94,382,382&ava=1" style="width: 100%; height: 100%;">\n')
    file.write('    </div>\n')

    # Добавляем рамку для GIF-изображения
    file.write('    <div style="border: 1px solid #ccc; padding: 10px; margin-top: 20px; width: 500px; height: 500px;">\n')
    file.write('        <img src="https://cdn.humoraf.ru/wp-content/uploads/2019/04/ponedelnik-gifki-humoraf-53.gif" alt="GIF Картинка" style="width: 100%; height: 100%;">\n')
    file.write('    </div>\n')

    # Записываем закрывающий HTML-код
    file.write('''
</body>
</html>
''')

print("HTML-файл создан успешно.")

# Открываем HTML-файл в браузере
webbrowser.open('index.html', new=2)

# Закрываем соединение с базой данных
conn.close()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Подключение к базе данных

Название функции/команды: sqlite3.connect('names_database.db')

Пример:
'''
conn = sqlite3.connect('names_database.db')
'''
Описание:
Эта строка кода создает подключение к базе данных SQLite с именем "names_database.db" или открывает уже существующую 
базу данных с таким именем. Если базы данных нет, она будет создана. Объект conn представляет собой соединение
с базой данных.
'''
'''
Шаг №2: Создание таблицы в базе данных

Название функции/команды: cursor.execute('''CREATE TABLE IF NOT EXISTS names (id INTEGER PRIMARY KEY AUTOINCREMENT,
first_name TEXT, last_name TEXT)''')

Пример:
'''
cursor.execute('''
    CREATE TABLE IF NOT EXISTS names (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT,
        last_name TEXT
    )
''')
'''
Описание:
Эта строка кода создает таблицу "names" в базе данных с тремя полями: "id" (целочисленный идентификатор,
автоматически увеличиваемый), "first_name" и "last_name" (текстовые поля для хранения имени и фамилии соответственно).
'''
'''
Шаг №3: Валидация имени

Название функции/команды: re.match("^[а-яА-Яa-zA-Z]+$", name)

Пример:
'''
return re.match("^[а-яА-Яa-zA-Z]+$", name) is not None
'''
Описание:
Эта функция validate_name использует регулярное выражение для проверки того, что переданное имя состоит только из
букв русского или английского алфавита. Возвращает True, если условие выполняется, и False в противном случае.
'''
'''
Шаг №4: Запись данных в базу данных

Название функции/команды: cursor.execute('INSERT INTO names (first_name, last_name) VALUES (?, ?)', (name, surname))

Пример:
'''
cursor.execute('INSERT INTO names (first_name, last_name) VALUES (?, ?)', (name, surname))
'''
Описание:
Эта строка кода вставляет введенные пользователем имя и фамилию в таблицу "names" базы данных.
'''
'''
Шаг №5: Запись данных в HTML-файл

Название функции/команды: with open('index.html', 'w', encoding='utf-8') as file:

Пример:
'''
with open('index.html', 'w', encoding='utf-8') as file:
'''
Описание:
Этот блок кода открывает файл 'index.html' для записи в кодировке UTF-8. Файл будет использоваться для 
создания HTML-страницы.
'''
'''
Шаг №6: Запись HTML-кода в файл

Название функции/команды: file.write(...)

Пример:
'''
file.write('    <h3>{html.escape(name)}</h3>\n')
'''
Описание:
Этот блок кода записывает различные версии имени и фамилии в HTML-файл с разными стилями 
(стандартное, курсивное, жирное) и добавляет рамки для JPG и GIF изображений с соответствующими ссылками.
'''
'''
Шаг №7: Закрытие HTML-файла

Название функции/команды: file.write('...</body></html>')

Пример:
'''
file.write('''
</body>
</html>
''')
'''
Описание:
Эта строка кода закрывает HTML-код в файле после записи данных и завершает создание HTML-страницы.
'''
'''
Шаг №8: Открытие HTML-файла в браузере

Название функции/команды: webbrowser.open('index.html', new=2)

Пример:
'''
webbrowser.open('index.html', new=2)
'''
Описание:
Эта строка кода открывает созданный HTML-файл в браузере. Аргумент new=2 гарантирует, 
что файл открывается в новой вкладке.
'''
'''
Шаг №9: Закрытие соединения с базой данных

Название функции/команды: conn.close()

Пример:
'''
conn.close()
'''
Описание:
Эта строка кода закрывает соединение с базой данных после выполнения всех операций.
'''

'''
Как итог, этот код выполняет следующие функции:

1. Устанавливает соединение с базой данных SQLite, создает таблицу "names" и проверяет наличие или создает базу данных.
2. Вводит имя и фамилию от пользователя, проверяет их на корректность (только буквы русского или английского алфавита).
3. Записывает введенные данные в базу данных и создает HTML-файл с разными стилями отображения имени и фамилии,
а также добавляет рамки для JPG и GIF изображений.
4. Открывает созданный HTML-файл в браузере.
5. Закрывает соединение с базой данных.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Практической-Работы: 10-11 - ЯНВАРЯ 2024 года.
''''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением
Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №22: Консольные приложения (bash, shell). Упаковка приложений (.exe) - модуль auto_py_to_exe

Выполните следующие задания:

Задание №1
а) Напишите скрипт, который будет создавать виртуальное окружение и устанавливать туда библиотеку openpyxl.
б) Напишите пример приложения-калькулятор на input.
в) Соберите это приложение в исполняемый файл («.exe»).
'''
'''
Урок от 10.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
Само решение не не удалось записать сюда (СМ ЗАДАНИЕ) - Сам калькулятор - есть в отдельном файле .py
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 10-11 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашнее задание № 22: Консольные приложения (bash, shell). Упаковка приложений (.exe) модуль auto_py_to_exe
Выполните следующее задание:

Задание №1

а) Напишите скрипт, который будет создавать виртуальное окружение и устанавливать туда библиотеку python-docx.
б) Напишите пример приложения на input, которое получает текст от пользователя и создаёт word-файл с этим текстом.
в) Соберите это приложение в исполняемый файл («.exe»).
В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 10.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
Само решение не не удалось записать сюда (СМ ЗАДАНИЕ) - Сам word-файл - есть в отдельном файле .py
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Практической-Работы: 12-13 - ЯНВАРЯ 2024 года.
''''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением
Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №23: Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет складывать два числа, и спустя 1 секунду задержки возвращать результат.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал многопоточного запуска, с замером времени.
'''
'''
Урок от 12.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import threading
import time

# Задание №1

# a) Функция для сложения двух чисел с задержкой в 1 секунду
def add_numbers(x, y):
    time.sleep(1)
    return x + y

# b) Запуск циклом 100 таких функций и замер времени
start_time = time.time()

results = []
for _ in range(100):
    result = add_numbers(3, 4)  # Пример: сложение чисел 3 и 4
    results.append(result)

end_time = time.time()
print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))

# c) Функционал многопоточного запуска с замером времени
start_time = time.time()

results_mt = []
threads = []

def worker():
    result = add_numbers(3, 4)
    results_mt.append(result)

for _ in range(100):
    thread = threading.Thread(target=worker)
    thread.start()
    threads.append(thread)

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()
print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Далее более подробно:
'''
'''
Для начала импортируем все то что нам нужно:
'''
import threading
import time
'''
Шаг 1: Функция add_numbers
'''
def add_numbers(x, y):
    time.sleep(1)
    return x + y
'''
Описание:

Название функции: add_numbers.

Пример кода:

def add_numbers(x, y):: Определяет функцию с именем add_numbers, принимающую два аргумента x и y.
time.sleep(1): Вызывает функцию sleep из модуля time, приостанавливая выполнение программы на 1 секунду.
Это эмулирует задержку в функции.
return x + y: Возвращает сумму x и y.
Более подробное описание:

Функция add_numbers принимает два аргумента x и y и добавляет их. 
Однако, перед возвратом результата, функция "засыпает" на 1 секунду с использованием time.sleep(1).
Это добавляет задержку к выполнению функции, чтобы сделать ее более реалистичной для примера многопоточности.
'''
'''
Шаг 2: Блок кода для выполнения без использования потоков
'''
# b) Запуск циклом 100 таких функций и замер времени
start_time = time.time()

results = []
for _ in range(100):
    result = add_numbers(3, 4)  # Пример: сложение чисел 3 и 4
    results.append(result)

end_time = time.time()
print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Пример кода:

start_time = time.time(): Записывает текущее время в переменную start_time.
results = []: Создает пустой список для хранения результатов функций.
for _ in range(100):: Начинает цикл, который выполняется 100 раз.
result = add_numbers(3, 4): Вызывает функцию add_numbers с аргументами 3 и 4 (пример сложения чисел) и 
сохраняет результат в переменную result.
results.append(result): Добавляет результат в список results.
end_time = time.time(): Записывает текущее время в переменную end_time.
print("Time taken without threading: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время,
затраченное на выполнение цикла без использования потоков.

Более подробное описание:

В этом блоке кода мы запускаем цикл, который 100 раз вызывает функцию add_numbers с аргументами 3 и 4.
Результат каждого вызова добавляется в список results. Затем мы замеряем время, затраченное на выполнение цикла,
и выводим его в консоль.
'''
'''
Шаг 3: Блок кода для выполнения с использованием потоков
'''
# c) Функционал многопоточного запуска с замером времени
start_time = time.time()

results_mt = []
threads = []

def worker():
    result = add_numbers(3, 4)
    results_mt.append(result)

for _ in range(100):
    thread = threading.Thread(target=worker)
    thread.start()
    threads.append(thread)

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()
print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Пример кода:

start_time = time.time(): Записывает текущее время в переменную start_time.
results_mt = []: Создает пустой список для хранения результатов функций с использованием потоков.
threads = []: Создает пустой список для хранения объектов потоков.
def worker():: Определяет функцию worker, которая вызывает add_numbers и добавляет результат в список results_mt.
for _ in range(100):: Начинает цикл, который выполняется 100 раз.
thread = threading.Thread(target=worker): Создает объект потока, используя функцию worker в качестве цели.
thread.start(): Запускает поток.
threads.append(thread): Добавляет объект потока в список.
for thread in threads: и thread.join(): Ожидает завершения всех потоков.
end_time = time.time(): Записывает текущее время в переменную end_time.
print("Time taken with threading: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время,
затраченное на выполнение цикла с использованием потоков.

Более подробное описание:

В этом блоке кода мы используем многопоточность для выполнения функции add_numbers параллельно в 100 потоках.
Каждый поток вызывает функцию worker, которая в свою очередь вызывает add_numbers и добавляет результат 
в список results_mt. Мы замеряем время, затраченное на выполнение этого блока кода, и выводим его в консоль.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Возможные альтернативные варианты
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант 1: Использование concurrent.futures для многопоточности
'''
import concurrent.futures
import time

def add_numbers(x, y):
    time.sleep(1)
    return x + y

def worker(_):
    result = add_numbers(3, 4)
    return result

# Используем ThreadPoolExecutor для многопоточности
start_time = time.time()
with concurrent.futures.ThreadPoolExecutor() as executor:
    results_mt = list(executor.map(worker, range(100)))

end_time = time.time()
print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
В этом варианте используется модуль concurrent.futures, который предоставляет высокоуровневый интерфейс для работы
с потоками и процессами. Мы используем ThreadPoolExecutor для создания пула потоков и функцию executor.map для
многопоточного выполнения задач.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Функция add_numbers
'''
def add_numbers(x, y):
    time.sleep(1)
    return x + y
'''
Описание:

Название функции: add_numbers.

Пример кода:
def add_numbers(x, y):: Определяет функцию с именем add_numbers, принимающую два аргумента x и y.
time.sleep(1): Вызывает функцию sleep из модуля time, приостанавливая выполнение программы на 1 секунду. 
Это эмулирует задержку в функции.
return x + y: Возвращает сумму x и y.

Более подробное описание:

Функция add_numbers принимает два аргумента x и y и добавляет их. 
Однако, перед возвратом результата, функция "засыпает" на 1 секунду с использованием time.sleep(1). 
Это добавляет задержку к выполнению функции, чтобы сделать ее более реалистичной для примера многопоточности.
'''
'''
Шаг 2: Блок кода с использованием concurrent.futures
'''
# Используем ThreadPoolExecutor для многопоточности
start_time = time.time()
with concurrent.futures.ThreadPoolExecutor() as executor:
    results_mt = list(executor.map(worker, range(100)))

end_time = time.time()
print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Пример кода:
start_time = time.time(): Записывает текущее время в переменную start_time.
with concurrent.futures.ThreadPoolExecutor() as executor:: Создает ThreadPoolExecutor в контексте.
results_mt = list(executor.map(worker, range(100))): Запускает функцию worker в 100 потоках и получает результаты 
с использованием executor.map.
end_time = time.time(): Записывает текущее время в переменную end_time.
print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время,
затраченное на выполнение кода с использованием ThreadPoolExecutor.

Более подробное описание:

Мы используем ThreadPoolExecutor из concurrent.futures для создания пула потоков. 
В контексте блока with мы вызываем функцию worker в 100 потоках с использованием executor.map. 
Этот метод ожидает завершения всех потоков и возвращает результаты в виде списка. 
Затем мы замеряем время выполнения этого блока кода и выводим результат в консоль.
'''
'''
В итоге, код в Варианте 1 демонстрирует использование ThreadPoolExecutor для параллельного выполнения функции в
нескольких потоках, что может быть полезным при выполнении множества схожих задач.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант 2: Использование asyncio для асинхронного выполнения
'''
import asyncio
import time

async def add_numbers(x, y):
    await asyncio.sleep(1)
    return x + y

async def worker():
    result = await add_numbers(3, 4)
    return result

# Используем asyncio для асинхронного выполнения
start_time = time.time()

# Создаем новый цикл событий, если его нет
loop = asyncio.new_event_loop()
asyncio.set_event_loop(loop)

# Запускаем асинхронные задачи
results_mt = loop.run_until_complete(asyncio.gather(*(worker() for _ in range(100))))

end_time = time.time()

print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Определение асинхронных функций
'''
async def add_numbers(x, y):
    await asyncio.sleep(1)
    return x + y

async def worker():
    result = await add_numbers(3, 4)
    return result
'''
Описание:
add_numbers: Асинхронная функция, принимающая два аргумента x и y. 
Она приостанавливает выполнение на 1 секунду с помощью await asyncio.sleep(1) и затем возвращает сумму x и y.
worker: Еще одна асинхронная функция, вызывающая add_numbers с аргументами 3 и 4. 
Она также использует await, чтобы ждать завершения выполнения add_numbers и затем возвращает результат.
'''
'''
Шаг 2: Использование asyncio для асинхронного выполнения
'''
start_time = time.time()
loop = asyncio.get_event_loop()
tasks = [worker() for _ in range(100)]
results_mt = loop.run_until_complete(asyncio.gather(*tasks))
end_time = time.time()
'''
Описание:
start_time = time.time(): Записывает текущее время перед началом выполнения асинхронного кода.
loop = asyncio.get_event_loop(): Получает экземпляр цикла событий asyncio.
tasks = [worker() for _ in range(100)]: Создает список асинхронных задач, 
каждая из которых представляет собой вызов функции worker().
results_mt = loop.run_until_complete(asyncio.gather(*tasks)): Запускает выполнение всех асинхронных 
задач с помощью asyncio.gather() и блокирует выполнение до их завершения. 
Результаты сохраняются в results_mt.
end_time = time.time(): Записывает текущее время после завершения выполнения асинхронного кода.
'''
'''
Шаг 3: Вывод результатов
'''
print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time))
'''
Описание:
print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time)): Выводит в консоль время, 
затраченное на выполнение всех асинхронных задач. В данном случае, это время выполнения 100 вызовов функции worker().
'''
'''
Более подробное описание:

Код использует асинхронные функции для эмуляции параллельного выполнения асинхронных задач. 
В цикле for создается список из 100 вызовов worker(). Затем с помощью asyncio.gather() 
запускаются все эти задачи параллельно, и код блокируется до их завершения. Время выполнения замеряется 
с использованием time.time().
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Можно прям и по этапно ("пошагово") - т.е. - а), б), в).
'''

'''
Шаг а): Напишите функцию, складывающую два числа с задержкой в 1 секунду
'''
import asyncio

async def add_numbers(x, y):
    await asyncio.sleep(1)  # Асинхронная задержка на 1 секунду
    return x + y
'''
Описание:

add_numbers: Асинхронная функция, принимающая два аргумента x и y.
Она приостанавливает выполнение на 1 секунду с помощью await asyncio.sleep(1) и затем возвращает сумму x и y.
'''

'''
Шаг 1б: Запустите циклом 100 таких функций, а также замерьте время
'''
import asyncio
import time

# Определение асинхронных функций
async def add_numbers(x, y):
    await asyncio.sleep(1)
    return x + y

async def worker():
    result = await add_numbers(3, 4)
    return result

# Использование asyncio для асинхронного выполнения
start_time = time.time()
loop = asyncio.get_event_loop()
tasks = [worker() for _ in range(100)]
results_mt = loop.run_until_complete(asyncio.gather(*tasks))
end_time = time.time()

print("Time taken with asyncio: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

start_time = time.time(): Записывает текущее время перед началом выполнения асинхронного кода.
loop = asyncio.get_event_loop(): Получает экземпляр цикла событий asyncio.
tasks = [worker() for _ in range(100)]: Создает список асинхронных задач, каждая из которых представляет 
собой вызов функции worker().
results_mt = loop.run_until_complete(asyncio.gather(*tasks)): Запускает выполнение всех асинхронных задач 
с помощью asyncio.gather() и блокирует выполнение до их завершения. Результаты сохраняются в results_mt.
end_time = time.time(): Записывает текущее время после завершения выполнения асинхронного кода.
'''

'''
Шаг 1в: Добавьте функционал многопоточного запуска, с замером времени
'''
import concurrent.futures
import time

# Определение функции
def add_numbers(x, y):
    time.sleep(1)
    return x + y

# Использование ThreadPoolExecutor для многопоточности
start_time = time.time()
with concurrent.futures.ThreadPoolExecutor() as executor:
    results_mt = list(executor.map(add_numbers, [3] * 100, [4] * 100))
end_time = time.time()

print("Time taken with concurrent.futures: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

start_time = time.time(): Записывает текущее время перед началом выполнения кода с использованием ThreadPoolExecutor.
with concurrent.futures.ThreadPoolExecutor() as executor: Создает ThreadPoolExecutor для многопоточности.
results_mt = list(executor.map(add_numbers, [3] * 100, [4] * 100)): Запускает выполнение функции add_numbers в 
нескольких потоках с помощью executor.map() и сохраняет результаты в results_mt.
end_time = time.time(): Записывает текущее время после завершения выполнения кода с использованием ThreadPoolExecutor.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 12-13 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django

Дисциплина: Основы программирования на Python

Домашнее задание № 23 : Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1

а) Напишите функцию, которая будет создавать файл, с задержкой 1 секунду.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал многопоточного запуска, с замером времени.

** В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 12.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import threading
import time


# Шаг 1: Написание функции для создания файла с задержкой
def create_file_with_delay(file_number):
    time.sleep(1)  # Задержка в 1 секунду
    file_name = f"file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"File '{file_name}' created.")


# Шаг 2: Запуск циклом 100 функций и замер времени
start_time = time.time()

for i in range(1, 101):
    create_file_with_delay(i)

end_time = time.time()

print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))

# Шаг 3: Добавление функционала многопоточного запуска и замер времени
start_time = time.time()

threads = []
for i in range(1, 101):
    thread = threading.Thread(target=create_file_with_delay, args=(i,))
    threads.append(thread)
    thread.start()

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()

print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import threading
import time
'''
Описание: 
В этом шаге импортируются две библиотеки - threading для работы с многопоточностью и time для работы с временем.
'''
'''
Шаг 2: Определение функции для создания файла с задержкой
'''
def create_file_with_delay(file_number):
    time.sleep(1)  # Задержка в 1 секунду
    file_name = f"file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"File '{file_name}' created.")
'''
Описание: 
Эта функция create_file_with_delay принимает file_number в качестве аргумента, добавляет задержку в 1 секунду 
с использованием time.sleep(1), затем создает файл с именем, содержащим номер, и записывает в него текст.
В конце функция выводит сообщение о создании файла.
'''
'''
Шаг 3: Запуск циклом 100 функций и замер времени без использования многопоточности
'''
start_time = time.time()

for i in range(1, 101):
    create_file_with_delay(i)

end_time = time.time()

print("Time taken without threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:
В этом шаге создается цикл, который запускает функцию create_file_with_delay 100 раз для создания файлов с задержкой.
Время начала измеряется с использованием time.time(), а затем выводится время выполнения после завершения цикла.
'''
'''
Шаг 4: Добавление функционала многопоточного запуска и замер времени
'''
start_time = time.time()

threads = []
for i in range(1, 101):
    thread = threading.Thread(target=create_file_with_delay, args=(i,))
    threads.append(thread)
    thread.start()

# Ожидание завершения всех потоков
for thread in threads:
    thread.join()

end_time = time.time()

print("Time taken with threading: {:.2f} seconds".format(end_time - start_time))
'''
Описание:
В этом шаге создается новый цикл, который использует многопоточность. 
Создаются 100 потоков (threads), каждый из которых запускает функцию create_file_with_delay с уникальным file_number. 
Затем выполняется ожидание завершения всех потоков с использованием thread.join(). 
Время выполнения снова измеряется и выводится.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ЕЩЕ ВАРИАНТ
'''
import concurrent.futures
import asyncio
import time

def create_file_with_delay_thread(file_number):
    time.sleep(1)
    file_name = f"thread_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"Thread File '{file_name}' created.")

async def create_file_with_delay_asyncio(file_number):
    await asyncio.sleep(1)
    file_name = f"asyncio_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"AsyncIO File '{file_name}' created.")

async def main():
    # Используем ThreadPoolExecutor для многопоточности
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(create_file_with_delay_thread, range(1, 51))

    # Используем asyncio для асинхронного выполнения
    tasks = [create_file_with_delay_asyncio(i) for i in range(51, 101)]
    await asyncio.gather(*tasks)

start_time = time.time()
asyncio.run(main())
end_time = time.time()

print("Time taken with ThreadPoolExecutor and asyncio: {:.2f} seconds".format(end_time - start_time))
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек само собой
'''
'''
Шаг 1:
'''
def create_file_with_delay_thread(file_number):
    time.sleep(1)
    file_name = f"thread_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"Thread File '{file_name}' created.")
'''
Описание:

Эта функция create_file_with_delay_thread представляет собой синхронную функцию,
которая создает файл с использованием задержки в 1 секунду. Она принимает аргумент file_number,
который используется для создания уникального имени файла. Функция использует time.sleep(1),
чтобы создать задержку в 1 секунду перед созданием файла. Затем она открывает файл с именем, 
сформированным на основе file_number, и записывает в него строку. Наконец, выводится сообщение в 
консоль о создании файла.
'''
'''
Шаг 2:
'''
async def create_file_with_delay_asyncio(file_number):
    await asyncio.sleep(1)
    file_name = f"asyncio_file_{file_number}.txt"

    with open(file_name, "w") as file:
        file.write("This is a sample file.")

    print(f"AsyncIO File '{file_name}' created.")
'''
Описание:

Эта функция create_file_with_delay_asyncio - это асинхронная функция, использующая async/await, 
чтобы создать файл с задержкой в 1 секунду. Она аналогична предыдущей функции, но с использованием асинхронных 
возможностей. Она также принимает аргумент file_number, формирует уникальное имя файла, создает задержку с 
использованием await asyncio.sleep(1), и затем создает и записывает файл, выводя сообщение о создании в консоль.
'''
'''
Шаг 3:
'''
async def main():
    # Используем ThreadPoolExecutor для многопоточности
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(create_file_with_delay_thread, range(1, 51))

    # Используем asyncio для асинхронного выполнения
    tasks = [create_file_with_delay_asyncio(i) for i in range(51, 101)]
    await asyncio.gather(*tasks)
'''
Описание:

Эта функция main является точкой входа для выполнения обоих подходов (многопоточного и асинхронного).
Внутри main сначала используется ThreadPoolExecutor для запуска функции create_file_with_delay_thread 50 раз в
многопоточной среде. Затем используется asyncio для выполнения функции create_file_with_delay_asyncio еще 50 раз в
асинхронной среде. Весь код выполняется асинхронно, и функция main ожидает завершения всех задач, 
используя await asyncio.gather(*tasks).
'''
start_time = time.time()
asyncio.run(main())
end_time = time.time()

print("Time taken with ThreadPoolExecutor and asyncio: {:.2f} seconds".format(end_time - start_time))
'''
Описание:

Здесь вычисляется и выводится общее время выполнения обоих подходов с использованием многопоточности и асинхронности.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #


< !--  ########################################################################################## -->
< !-- HTML  # 1 Recap + Вводный ПЕРВЫЙ УРОК - 08.01.2024 -->

< !-- < html >
< head >

< meta
charset = "UTF-8" / >
< title > Hello
HTML < / title >
< / head >
< body >

hello
world
Привет
Мир

< img
src = "https://pngicon.ru/file/uploads/serdce.png"
height = "150px"
width = "150px" / >
< h1 > Вау - я
программист < / h1 >
< / body >
< / html > -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !-- 2024
_HTML_LESSON_№2
_Структура_HTML.Теги_Атрибуты - 15.01
.2024 -->

< !--

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 2: Структура
HTML.Теги.Атрибуты.Правила
языка
разметки
-->

< !DOCTYPE
html >
< html
lang = "ru" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Страница
с
текстом
и
заголовками < / title >
< / head >
< body >

< !-- 1.
Создать
страницу
с
текстом
по
примеру
на
картинке. -->

< h2 >
Стихотворение
< / h2 >
< p >

Мириады
маленьких
дел < br >
Пьют
по
капле
гаснущий
день, < br >
А < i > дела
большие < / i > сушит
жажда. < br >
Оставляя
все
на «потом», < br >
Прозреваем
задним
числом, < br >
Только
год < b > не
повторится < / b > дважды. < br >
< br >
И.Тальков

< / p >

< !-- 2.
Повторите
по
данному
образцу: -->
< h1 > Это
заголовок < / h1 >
< h2 > Это
заголовок < / h2 >
< h3 > Это
заголовок < / h3 >
< h4 > Это
заголовок < / h4 >

< p > Это < b > абзац < / b >.< / p >
< p > Это
еще < i > абзац < / i >.< / p >

< h1 > < i > Это
заголовок
h1 < / i > < / h1 >

< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Дата выполнения Практической-Работы: 17-18 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Практическая работа №24: Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет загружать изображение.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал мультипроцессорного запуска, с замером времени.

*г) Необходимо провести вычисление факториала в параллельных процессах.
'''
'''
Урок от 17.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
import time
import multiprocessing
from PIL import Image

def load_image(image_path):
    """Функция для загрузки изображения."""
    with Image.open(image_path) as img:
        # Просто загрузка изображения без выполнения каких-либо операций.
        pass

def process_images_sequential(image_paths):
    """Функция для последовательной обработки изображений."""
    start_time = time.time()

    for path in image_paths:
        load_image(path)

    end_time = time.time()
    print(f"Последовательная обработка: {end_time - start_time} секунд")

def process_images_parallel(image_paths):
    """Функция для параллельной обработки изображений."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    pool.map(load_image, image_paths)
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельная обработка: {end_time - start_time} секунд")

def calculate_factorial(n):
    """Функция для вычисления факториала."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    return result

def calculate_factorial_parallel(n):
    """Функция для параллельного вычисления факториала."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    results = pool.map(calculate_factorial, range(1, n + 1))
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельное вычисление факториала: {end_time - start_time} секунд")
    return results

def show_images(image_paths):
    """Функция для отображения изображений."""
    for path in image_paths:
        img = Image.open(path)
        img.show()

if __name__ == "__main__":
    # Задайте список путей к изображениям
    image_paths = ["E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/1.jpg",
                   "E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/2.jpg"]  # Замените на реальные пути

    # Задание 1б: Запуск циклом 100 функций
    for _ in range(100):
        load_image(image_paths[0])  # Загружаем одно изображение

    # Задание 1в: Параллельная обработка с замером времени
    process_images_parallel(image_paths)

    # Задание 1г: Параллельное вычисление факториала
    n = 5
    result_parallel = calculate_factorial_parallel(n)
    print(f"Результаты параллельного вычисления факториала: {result_parallel}")

    # Отображение изображений
    show_images(image_paths)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1. Загрузка изображения
'''
def load_image(image_path):
    """Функция для загрузки изображения."""
    with Image.open(image_path) as img:
        # Просто загрузка изображения без выполнения каких-либо операций.
        pass
'''
Описание: 
Эта функция использует библиотеку PIL (Pillow) для загрузки изображения по указанному пути. 
В данном случае, операции с изображением не выполняются, и функция служит примером базовой операции загрузки.
'''
'''
Шаг №2. Последовательная обработка изображений
'''
def process_images_sequential(image_paths):
    """Функция для последовательной обработки изображений."""
    start_time = time.time()

    for path in image_paths:
        load_image(path)

    end_time = time.time()
    print(f"Последовательная обработка: {end_time - start_time} секунд")
'''
Описание: 
Эта функция выполняет последовательную обработку изображений. 
Она вызывает функцию load_image для каждого пути к изображению в списке image_paths. 
Замеряется время выполнения операции, и результат выводится на экран.
'''
'''
Шаг №3. Параллельная обработка изображений
'''
def process_images_parallel(image_paths):
    """Функция для параллельной обработки изображений."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    pool.map(load_image, image_paths)
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельная обработка: {end_time - start_time} секунд")
'''
Описание: 
Эта функция выполняет параллельную обработку изображений, используя модуль multiprocessing. 
Пул процессов создается с количеством процессов, равным количеству ядер процессора. 
Функция load_image применяется к каждому пути к изображению в списке image_paths. 
Замеряется время выполнения операции, и результат выводится на экран.
'''
'''
Шаг №4. Вычисление факториала
'''
def calculate_factorial(n):
    """Функция для вычисления факториала."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    return result
'''
Описание: 
Эта функция принимает целое число n и вычисляет его факториал.
'''
'''
Шаг №5. Параллельное вычисление факториала
'''
def calculate_factorial_parallel(n):
    """Функция для параллельного вычисления факториала."""
    start_time = time.time()

    # Создаем пул процессов, количество которых равно количеству ядер процессора
    pool = multiprocessing.Pool()
    results = pool.map(calculate_factorial, range(1, n + 1))
    pool.close()
    pool.join()

    end_time = time.time()
    print(f"Параллельное вычисление факториала: {end_time - start_time} секунд")
    return results
'''
Описание: 
Эта функция выполняет параллельное вычисление факториала для чисел от 1 до n с использованием пула процессов. 
Замеряется время выполнения операции, и результат выводится на экран.
'''
'''
Шаг №6. Отображение изображений
'''
def show_images(image_paths):
    """Функция для отображения изображений."""
    for path in image_paths:
        img = Image.open(path)
        img.show()
'''
Описание: 
Эта функция использует библиотеку PIL для открытия и отображения изображений, указанных в списке image_paths.
'''
'''
Шаг №7. Основной код (выполняется при запуске файла)
'''
if __name__ == "__main__":
    # Задайте список путей к изображениям
    image_paths = ["E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/1.jpg",
                   "E:/PyCharm - IT-Step/CHERNOWICK_+_WORK_IN_CLASS_ROOM/+2024+/2.jpg"]  # Замените на реальные пути

    # Задание 1б: Запуск циклом 100 функций
    for _ in range(100):
        load_image(image_paths[0])  # Загружаем одно изображение

    # Задание 1в: Параллельная обработка с замером времени
    process_images_parallel(image_paths)

    # Задание 1г: Параллельное вычисление факториала
    n = 5
    result_parallel = calculate_factorial_parallel(n)
    print(f"Результаты параллельного вычисления факториала: {result_parallel}")

    # Отображение изображений
    show_images(image_paths)
'''
Описание: 
В основной части кода задаются пути к изображениям. 
Затем выполняются задания б, в и г. 
Задание б представляет собой цикл, в котором 100 раз загружается одно изображение. 
Задание в - параллельная обработка изображений с замером времени. 
Задание г - параллельное вычисление факториала для числа 5. 
Наконец, вызывается функция для отображения изображений.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #




# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 17-18 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Домашнее задание № 24 : Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1

а) Напишите функцию, которая будет создавать файл и записывать в него рандомное число, с задержкой 1 секунду.
б) Запустите циклом 1000 таких функций, а также замерьте время.
в) Добавьте функционал мультипоточного запуска, с замером времени. 
Обязательно посмотрите нагрузку на ЦП в этот момент (через диспетчер задач).

*г) Необходимо провести вычисление факториала в параллельных процессах.
** В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 12.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import time
import random
import threading
import multiprocessing

# Шаг а: Функция для создания файла и записи в него рандомного числа с задержкой
def write_random_number(file_path):
    """Создает файл и записывает в него рандомное число с задержкой 1 секунда."""
    with open(file_path, 'w') as file:
        random_number = random.randint(1, 100)
        file.write(str(random_number))
        time.sleep(1)

# Шаг б: Функция для запуска циклом 1000 функций с замером времени
def run_sequential():
    """Запускает циклом 1000 функций и замеряет время."""
    start_time = time.time()

    for _ in range(1000):
        write_random_number(f"file{_}.txt")

    end_time = time.time()
    print(f"Время, затраченное на последовательное выполнение: {end_time - start_time:.4f} секунд")

# Шаг в: Функция для многопоточного запуска с замером времени
def run_multithreaded():
    """Многопоточный запуск с замером времени."""
    start_time = time.time()

    threads = []
    for _ in range(1000):
        thread = threading.Thread(target=write_random_number, args=(f"file{_}.txt",))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    end_time = time.time()
    print(f"Время, затраченное на многопоточное выполнение: {end_time - start_time:.4f} секунд")

# Шаг г: Функция для параллельного вычисления факториала
def calculate_factorial(n, result_queue):
    """Вычисляет факториал числа n и помещает результат в очередь."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    result_queue.put(result)

# Шаг г (продолжение): Функция для параллельного запуска вычисления факториала
def run_multiprocess_factorial():
    """Параллельное вычисление факториала."""
    start_time = time.time()

    result_queue = multiprocessing.Queue()
    processes = []

    for _ in range(1000):
        process = multiprocessing.Process(target=calculate_factorial, args=(5, result_queue))
        processes.append(process)
        process.start()

    for process in processes:
        process.join()

    # Получаем результаты из очереди
    factorial_results = [result_queue.get() for _ in range(1000)]

    end_time = time.time()
    print(f"Время, затраченное на параллельное вычисление факториала: {end_time - start_time:.4f} секунд")
    print(f"Результаты факториала: {factorial_results}")

if __name__ == '__main__':
    # Шаг б: Запуск циклом 1000 функций
    run_sequential()

    # Шаг в: Многопоточный запуск
    run_multithreaded()

    # Шаг г: Параллельное вычисление факториала
    run_multiprocess_factorial()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг а: Функция для создания файла и записи в него рандомного числа с задержкой
'''
def write_random_number(file_path):
    """Создает файл и записывает в него рандомное число с задержкой 1 секунда."""
    with open(file_path, 'w') as file:
        random_number = random.randint(1, 100)
        file.write(str(random_number))
        time.sleep(1)
'''
Описание: 
Эта функция создает файл по указанному пути (file_path) и записывает в него случайное число от 1 до 100. 
После этого функция "засыпает" на 1 секунду с помощью time.sleep(1).
'''
'''
Шаг б: Функция для запуска циклом 1000 функций с замером времени
'''
def run_sequential():
    """Запускает циклом 1000 функций и замеряет время."""
    start_time = time.time()

    for _ in range(1000):
        write_random_number(f"file{_}.txt")

    end_time = time.time()
    print(f"Время, затраченное на последовательное выполнение: {end_time - start_time:.4f} секунд")
'''
Описание: 
Эта функция запускает цикл из 1000 итераций. 
На каждой итерации она вызывает функцию write_random_number, создавая 1000 файлов. 
Затем она замеряет время выполнения и выводит результат.
'''
'''
Шаг в: Функция для многопоточного запуска с замером времени
'''
def run_multithreaded():
    """Многопоточный запуск с замером времени."""
    start_time = time.time()

    threads = []
    for _ in range(1000):
        thread = threading.Thread(target=write_random_number, args=(f"file{_}.txt",))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    end_time = time.time()
    print(f"Время, затраченное на многопоточное выполнение: {end_time - start_time:.4f} секунд")
'''
Описание: 
Эта функция также создает 1000 файлов, но использует многопоточность. 
Для каждой итерации цикла создается новый поток (threading.Thread), который выполняет функцию write_random_number. 
Затем происходит ожидание завершения всех потоков (thread.join()), и замеряется время выполнения.
'''
'''
Шаг г: Функция для параллельного вычисления факториала
'''
def calculate_factorial(n, result_queue):
    """Вычисляет факториал числа n и помещает результат в очередь."""
    result = 1
    for i in range(1, n + 1):
        result *= i
    result_queue.put(result)
'''
Описание: 
Эта функция вычисляет факториал числа n и помещает результат в указанную очередь (result_queue).
'''
'''
Шаг г (продолжение): Функция для параллельного запуска вычисления факториала
'''
def run_multiprocess_factorial():
    """Параллельное вычисление факториала."""
    start_time = time.time()

    result_queue = multiprocessing.Queue()
    processes = []

    for _ in range(1000):
        process = multiprocessing.Process(target=calculate_factorial, args=(5, result_queue))
        processes.append(process)
        process.start()

    for process in processes:
        process.join()

    # Получаем результаты из очереди
    factorial_results = [result_queue.get() for _ in range(1000)]

    end_time = time.time()
    print(f"Время, затраченное на параллельное вычисление факториала: {end_time - start_time:.4f} секунд")
    print(f"Результаты факториала: {factorial_results}")
'''
Описание: 
Эта функция создает 1000 процессов (multiprocessing.Process), каждый из которых вычисляет факториал числа 5. 
Затем происходит ожидание завершения всех процессов, получение результатов из очереди и вывод времени выполнения
и результатов.
'''
'''
На сладкое
'''
if __name__ == '__main__':
    # Шаг 1б: Запуск циклом 1000 функций
    run_sequential()

    # Шаг 1в: Многопоточный запуск
    run_multithreaded()

    # Шаг 1г: Параллельное вычисление факториала
    run_multiprocess_factorial()
'''
Описание: 
Скажем так, основная часть кода запускает три разные функции для сравнения времени выполнения: последовательное
выполнение, многопоточное выполнение и параллельное вычисление факториала.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Дата выполнения Практической-Работы: 19-20 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Практическая работа №25: Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет загружать данные с jsonplaceholder.
б) Запустите циклом 100 таких функций, а также замерьте время.
в) Добавьте функционал асинхронной работы, с замером времени.

 - - - КАК ЭТО ПОНЯЛ Я - - - 
*) *) Запустите сначала 100, затем 200, и на последок 500 таких функций и поработайте с файлами. 
Сделайте запись чисел в файл 1 : 1 в файл 2 : 1, 2 в файл 3 : 1, 2, 3 и так далее,
чтобы задача была цикличной и вариативной.
Затем посмотрите сколько времени уходит на исполнение и насколько преимущественнее тот или иной подход.
Сравните и выведите результаты на экран. На сколько эффективна мультипоточность и, или  мультипроцессорность.
'''
'''
Урок от 19.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''' Вариант №1 - (учитываются пункты только а), б), в)) '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp  # Добавлен импорт aiohttp
import asyncio  # Добавлен импорт asyncio
import time  # Добавлен импорт time
import requests  # Добавлен импорт requests

# Асинхронная функция для загрузки данных по URL с использованием aiohttp
async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()

# Асинхронная функция для асинхронного выполнения запросов к нескольким URL
async def asynchronous_fetch(urls):
    # Создание сессии для aiohttp
    async with aiohttp.ClientSession() as session:
        # Создание списка задач (coroutines) для каждого URL
        tasks = [fetch_data(session, url) for url in urls]
        # Ожидание выполнения всех задач и возврат результатов
        return await asyncio.gather(*tasks)

# Синхронная функция для синхронного выполнения запросов к нескольким URL
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        # Синхронный запрос с использованием requests
        response = requests.get(url)
        data = response.json()
        # Обработка данных, если необходимо
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")

# Асинхронная точка входа программы
async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавьте остальные URL
    ]

    # Запуск задания №1 - б) (синхронный)
    synchronous_fetch(urls)

    # Запуск задания №1 - в) (асинхронный)
    start_time = time.time()
    await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения: {elapsed_time:.2f} секунд")

# Проверка, что скрипт выполняется как программа, а не импортируется как модуль
if __name__ == "__main__":
    # Запуск асинхронной программы с использованием asyncio
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек
'''
import aiohttp
import asyncio
import time
import requests
'''
aiohttp - библиотека для асинхронных HTTP-запросов.
asyncio - библиотека для асинхронного программирования.
time - для замера времени выполнения.
requests - библиотека для синхронных HTTP-запросов.
'''
'''
Асинхронная функция для загрузки данных по URL с использованием aiohttp
'''
async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()
'''
fetch_data - асинхронная функция, которая использует aiohttp для выполнения асинхронного HTTP-запроса.
Возвращает JSON-данные из ответа.
'''
'''
Асинхронная функция для асинхронного выполнения запросов к нескольким URL
'''
async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)
'''
asynchronous_fetch - асинхронная функция, которая создает сессию aiohttp и выполняет несколько асинхронных 
HTTP-запросов параллельно.
'''
'''
Синхронная функция для синхронного выполнения запросов к нескольким URL
'''
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")
'''
synchronous_fetch - синхронная функция, которая выполняет HTTP-запросы с использованием библиотеки requests.
'''
'''
Асинхронная точка входа программы
'''
async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавьте остальные URL
    ]

    synchronous_fetch(urls)
    start_time = time.time()
    await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения: {elapsed_time:.2f} секунд")
'''
main - асинхронная точка входа программы. 
Запускает сначала синхронное выполнение, затем асинхронное, и выводит время выполнения обеих операций.
'''
'''
Проверка, что скрипт выполняется как программа
'''
if __name__ == "__main__":
    asyncio.run(main())
'''
Проверяет, что скрипт выполняется как программа, а не импортируется как модуль,
и запускает асинхронную программу с использованием asyncio.run(main()).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''' Вариант №2 - (учитываются Все пункты)'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import requests
import json
import time
import concurrent.futures

def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)

def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(', '.join(map(str, data)) + '\n')

def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)
    write_to_file(f'file{task_id % 3 + 1}.txt', data)

def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")

def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")

def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")

if __name__ == "__main__":
    main()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек
'''
import requests
import json
import time
import concurrent.futures
'''
requests: библиотека для выполнения HTTP-запросов.
json: модуль для работы с данными в формате JSON.
time: модуль для работы со временем.
concurrent.futures: стандартная библиотека для работы с параллельным выполнением кода.
'''
'''
Функция load_data
'''
def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)
'''
load_data: 
функция выполняет HTTP GET-запрос по заданному URL, затем использует json.loads для загрузки данных из 
текстового ответа в формате JSON.
'''
'''
Функция write_to_file
'''
def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(', '.join(map(str, data)) + '\n')
'''
write_to_file: функция записывает данные в файл. Каждый вызов функции добавляет новую строку в файл.
'''
'''
Функция task
'''
def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)
    write_to_file(f'file{task_id % 3 + 1}.txt', data)
'''
task: функция, представляющая задачу. 
Она формирует URL с использованием task_id, загружает данные с помощью load_data и записывает их в файл.
'''
'''
Функция synchronous_execution
'''
def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")
'''
synchronous_execution: функция для синхронного выполнения 100 задач. Измеряется время начала и конца выполнения.
'''
'''
Функция asynchronous_execution
'''
def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")
'''
asynchronous_execution: функция для асинхронного выполнения 100 задач с использованием пула потоков.
Измеряется время начала и конца выполнения.
'''
'''
Функция main
'''
def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")
'''
main: функция, запускающая синхронное и асинхронное выполнение, а также асинхронное выполнение с разным числом задач.
Измеряется время каждого выполнения.
'''
'''
Проверка, что скрипт выполняется как программа
'''
if __name__ == "__main__":
    main()
'''
Проверка, что скрипт выполняется напрямую (а не импортирован как модуль), и запуск функции main().
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
''' Еще '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import requests
import json
import time
import concurrent.futures


def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)


def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(f'{data}\n')


def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)

    # Модификация: запись данных в файлы
    write_to_file(f'file{task_id % 3 + 1}.txt', f'{task_id}: {data}')


def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")


def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")


def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")


if __name__ == "__main__":
    main()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Эти строки отвечают за импорт необходимых библиотек.
'''
import requests
import json
import time
import concurrent.futures
'''
Эта функция load_data принимает URL в качестве аргумента, отправляет GET-запрос по указанному URL с помощью библиотеки
requests, получает ответ и затем использует json.loads для преобразования JSON-строки в Python-объект. 
Возвращает этот объект.
'''
def load_data(url):
    response = requests.get(url)
    return json.loads(response.text)
'''
Функция write_to_file принимает имя файла (filename) и данные (data) и открывает файл в режиме добавления ('a').
Затем она записывает данные в файл, добавляя символ новой строки после каждой записи.
'''
def write_to_file(filename, data):
    with open(filename, 'a') as file:
        file.write(f'{data}\n')
'''
Функция task принимает task_id, формирует URL с использованием этого идентификатора, 
загружает данные с помощью функции load_data, а затем записывает эти данные в соответствующий файл с использованием 
функции write_to_file. Файл выбирается с учетом остатка от деления task_id на 3, 
чтобы обеспечить цикличную запись в 3 файла.
'''
def task(task_id):
    url = f'https://jsonplaceholder.typicode.com/todos/{task_id}'
    data = load_data(url)

    # Модификация: запись данных в файлы
    write_to_file(f'file{task_id % 3 + 1}.txt', f'{task_id}: {data}')
'''
Функция synchronous_execution выполняет 100 задач синхронно (последовательно) и замеряет время выполнения.
'''
def synchronous_execution():
    start_time = time.time()
    for i in range(1, 101):
        task(i)
    end_time = time.time()
    print(f"Время синхронного выполнения: {end_time - start_time} секунд")
'''
Функция asynchronous_execution выполняет 100 
задач асинхронно с использованием многозадачности в пуле потоков и замеряет время выполнения.
'''
def asynchronous_execution():
    start_time = time.time()
    with concurrent.futures.ThreadPoolExecutor() as executor:
        executor.map(task, range(1, 101))
    end_time = time.time()
    print(f"Время асинхронного выполнения: {end_time - start_time} секунд")
'''
В функции main вызываются три варианта выполнения: синхронное, асинхронное с 100 задачами и 
асинхронное с разным числом задач (100, 200, 500). Замеряется время выполнения каждого варианта.
'''
def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение
    asynchronous_execution()

    # Исполнение с разным числом задач
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")

if __name__ == "__main__":
    main()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Можно модифицировать функцию main следующим образом, чтобы выполнять асинхронное выполнение с разным числом задач:
'''
def main():
    # Синхронное выполнение
    synchronous_execution()

    # Асинхронное выполнение с разным числом задач
    for num_tasks in [100, 200, 500, 1000]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(task, range(1, num_tasks + 1))
        end_time = time.time()
        print(f"Асинхронное выполнение с {num_tasks} задачами: {end_time - start_time} секунд")

if __name__ == "__main__":
    main()
'''
Тут добавлена задача с 1000 задачами.
Можно добавить или изменить значения в списке [100, 200, 500, 1000], чтобы отражать нужные варианты выполнения.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''Рубрика "ЭКСПЕРЕМЕНТЫ" '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''

Выполните следующие задания:

Задание №1
а) Напишите функцию, которая будет загружать данные с jsonplaceholder.
б) Запустите циклом сначала 100 таких функций, затем 200 таких функций, 500 таких функций, 1000 таких функций,
2000 таких функций, 5000 таких функций, а также замерьте время.
в) Добавьте функционал асинхронной работы, с замером времени.


Дополнительно:

Запустите сначала 100, затем 200, и напоследок 500, 1000 таких функций и поработайте с файлами.
Затем сделайте запись чисел в файл 1 : 1 в файл 2 : 1, 2 в файл 3 : 1, 2, 3 и так далее,
чтобы задача была цикличной и вариативной.
Затем посмотрите сколько времени уходит на исполнение и насколько преимущественнее тот или иной подход.
Сравните и выведите результаты на экран. На сколько эффективна мультипоточность и, или  мультипроцессорность.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import time
import requests
import concurrent.futures

async def fetch_data(session, url):
    async with session.get(url) as response:
        if response.status == 200:
            return await response.json()
        else:
            response.raise_for_status()

async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)

def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
        # Обработка данных, если необходимо
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")

async def write_to_file(file, data):
    async with aiofiles.open(file, 'a') as f:
        await f.write(data + '\n')

async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавьте остальные URL
    ]

    # Задание 1б: Запуск циклом 100 функций
    start_time = time.time()
    for _ in range(100):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (100 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 200 функций
    start_time = time.time()
    for _ in range(200):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (200 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 500 функций
    start_time = time.time()
    for _ in range(500):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (500 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 1000 функций
    start_time = time.time()
    for _ in range(1000):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (1000 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 2000 функций
    start_time = time.time()
    for _ in range(2000):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (2000 функций): {elapsed_time:.2f} секунд")

    # Задание 1б: Запуск циклом 5000 функций
    start_time = time.time()
    for _ in range(5000):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (5000 функций): {elapsed_time:.2f} секунд")

    # Дополнительно) Запуск сначала 100, затем 200, и на последок 500, 1000 функций
    for num_tasks in [100, 200, 500, 1000]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(asynchronous_fetch, [urls] * num_tasks)
        elapsed_time = time.time() - start_time
        print(f"Асинхронное выполнение с {num_tasks} задачами: {elapsed_time:.2f} секунд")

        # Запись чисел в файлы
        for i in range(num_tasks):
            data = ', '.join(map(str, range(1, i + 2)))
            await write_to_file(f"file{i + 1}.txt", data)

if __name__ == "__main__":
    asyncio.run(main())

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Функция fetch_data

Пример кода:
'''
async def fetch_data(session, url):
    async with session.get(url) as response:
        if response.status == 200:
            return await response.json()
        else:
            response.raise_for_status()
'''
Полное и подробное описание:
Эта функция представляет собой асинхронный метод, который использует библиотеку aiohttp для выполнения 
HTTP-запроса к указанному URL. Она принимает сессию (session) и URL, а затем асинхронно отправляет GET-запрос. 
Если статус ответа равен 200, функция возвращает данные в формате JSON. В противном случае вызывается исключение 
для обработки ошибки.
'''
'''
Шаг №2: Функция asynchronous_fetch

Пример кода:
'''
async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiohttp для асинхронного выполнения нескольких HTTP-запросов.
Она создает сессию (ClientSession), а затем формирует список задач (tasks), каждая из которых представляет собой
вызов функции fetch_data для каждого URL из переданного списка. Затем функция использует asyncio.gather для 
параллельного выполнения всех задач и возвращает результат.
'''
'''
Шаг №3: Функция synchronous_fetch

Пример кода:
'''
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
        # Обработка данных, если необходимо
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")
'''
Полное и подробное описание:
Эта функция представляет собой синхронный метод для выполнения HTTP-запросов с использованием библиотеки requests. 
Она проходит по каждому URL в переданном списке, отправляет GET-запрос и обрабатывает данные (если это необходимо). 
Время выполнения замеряется, и результат выводится на экран.
'''
'''
Шаг №4: Функция write_to_file

Пример кода:
'''
async def write_to_file(file, data):
    async with aiofiles.open(file, 'a') as f:
        await f.write(data + '\n')
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiofiles для асинхронной записи данных в файл. 
Она принимает имя файла (file) и данные (data). 
С помощью aiofiles.open открывается файл в режиме добавления ('a'), и затем происходит асинхронная запись 
переданных данных с добавлением новой строки.
'''
'''
Шаг №5: Функция main

Пример кода:
'''
async def main():
    # ... (Пропущен код для определения списка URLs)

    # Задание 1б: Запуск циклом 100 функций
    start_time = time.time()
    for _ in range(100):
        await asynchronous_fetch(urls)
    elapsed_time = time.time() - start_time
    print(f"Время асинхронного выполнения (100 функций): {elapsed_time:.2f} секунд")

    # ... (Аналогичные блоки для 200, 500, 1000, 2000, 5000 функций)

    # Дополнительно) Запуск сначала 100, затем 200, и на последок 500, 1000 функций
    for num_tasks in [100, 200, 500, 1000]:
        start_time = time.time()
        with concurrent.futures.ThreadPoolExecutor() as executor:
            executor.map(asynchronous_fetch, [urls] * num_tasks)
        elapsed_time = time.time() - start_time
        print(f"Асинхронное выполнение с {num_tasks} задачами: {elapsed_time:.2f} секунд")

        # Запись чисел в файлы
        for i in range(num_tasks):
            data = ', '.join(map(str, range(1, i + 2)))
            await write_to_file(f"file{i + 1}.txt", data)

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Эта асинхронная функция main выполняет несколько задач. 
В первой части она запускает циклы с разным числом асинхронных функций asynchronous_fetch для измерения времени 
выполнения. Затем, в дополнительном блоке, используется многопоточный подход (ThreadPoolExecutor), чтобы 
запустить асинхронные функции с разным числом задач. После этого данные записываются в файлы согласно заданным условиям.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
''''
Дата выполнения Домашней-Работы: 19-20 - ЯНВАРЯ 2024 года.
''''
'''
Домашнее задание
Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python
Домашнее задание № 25 : Многопоточное, асинхронное и мультипроцессорное программирование. GIL

Выполните следующие задания:

Задание №1

а) Напишите функцию, которая будет загружать информацию сразу с 100 ссылок.
б) Запустите эту функцию, а также замерьте время.
в) Добавьте функционал асинхронного запуска, с замером времени. Обязательно посмотрите нагрузку на ЦП в этот момент (через диспетчер задач).

*) Запустите сначала 100, затем 200, и напоследок 500 таких функций и поработайте с файлами. 
Сделайте запись чисел в файл 1 : 1 в файл 2 : 1, 2 в файл 3 : 1, 2, 3 и так далее,
чтобы задача была цикличной и вариативной.
Затем посмотрите сколько времени уходит на исполнение и насколько преимущественнее тот или иной подход.
Сравните и выведите результаты на экран. На сколько эффективна мультипоточность и, или  мультипроцессорность.

** В качестве ответа прикрепить скрин и ссылку на github.
'''
'''
Урок от 19.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import time
import concurrent.futures

async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()

async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)

def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")

async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавить другие URL-адреса
    ]

    # Запустите асинхронную функцию с разным количеством задач и измерьте время выполнения
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        for _ in range(num_tasks):
            await asynchronous_fetch(urls)
        elapsed_time = time.time() - start_time
        print(f"Время асинхронного выполнения ({num_tasks} задач): {elapsed_time:.2f} секунд")

        # Задача *: Запись чисел в файлы
        for i in range(num_tasks):
            with open(f"file{i + 1}.txt", "a") as file:
                file.write(", ".join(map(str, range(1, i + 2))) + "\n")

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Функция fetch_data

Пример кода:
'''
async def fetch_data(session, url):
    async with session.get(url) as response:
        return await response.json()
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiohttp для выполнения HTTP-запроса по указанному URL. 
Сначала создается асинхронный контекст сессии (async with session.get(url) as response), 
затем функция ожидает ответ (await response.json()), который интерпретируется как JSON. 
Функция возвращает полученные данные.
'''
'''
Шаг №2: Функция asynchronous_fetch

Пример кода:
'''
async def asynchronous_fetch(urls):
    async with aiohttp.ClientSession() as session:
        tasks = [fetch_data(session, url) for url in urls]
        return await asyncio.gather(*tasks)
'''
Полное и подробное описание:
Эта асинхронная функция использует библиотеку aiohttp для асинхронного выполнения нескольких HTTP-запросов. 
Она создает сессию (ClientSession) и формирует список задач (tasks), каждая из которых представляет собой вызов 
функции fetch_data для каждого URL из переданного списка. 
Затем функция использует asyncio.gather для параллельного выполнения всех задач и возвращает результат.
'''
'''
Шаг №3: Функция synchronous_fetch

Пример кода:
'''
def synchronous_fetch(urls):
    start_time = time.time()
    for url in urls:
        response = requests.get(url)
        data = response.json()
    elapsed_time = time.time() - start_time
    print(f"Время синхронного выполнения: {elapsed_time:.2f} секунд")
'''
Полное и подробное описание:
Эта функция представляет собой синхронный метод для выполнения HTTP-запросов с использованием библиотеки requests. 
Она проходит по каждому URL в переданном списке, отправляет GET-запрос и обрабатывает данные (если это необходимо). 
Время выполнения замеряется, и результат выводится на экран.
'''
'''
Шаг №4: Функция main

Пример кода:
'''
async def main():
    urls = [
        "https://jsonplaceholder.typicode.com/posts/1",
        "https://jsonplaceholder.typicode.com/posts/2",
        # Добавить другие URL-адреса
    ]

    # Запустите асинхронную функцию с разным количеством задач и измерьте время выполнения
    for num_tasks in [100, 200, 500]:
        start_time = time.time()
        for _ in range(num_tasks):
            await asynchronous_fetch(urls)
        elapsed_time = time.time() - start_time
        print(f"Время асинхронного выполнения ({num_tasks} задач): {elapsed_time:.2f} секунд")

        # Задача X: Запись чисел в файлы
        for i in range(num_tasks):
            with open(f"file{i + 1}.txt", "a") as file:
                file.write(", ".join(map(str, range(1, i + 2))) + "\n")

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Эта асинхронная функция main выполняет несколько задач. 
В первой части она запускает циклы с разным числом асинхронных функций asynchronous_fetch для измерения 
времени выполнения. Затем, в дополнительном блоке, используется многопоточный подход (ThreadPoolExecutor), 
чтобы запустить асинхронные функции с разным числом задач. После этого данные записываются в файлы согласно 
заданным условиям.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
22.01.2024 - Домашняя и практическая - работы.
<!-- ########################################################################################## -->
Практическая работа

Курс: Разработка интерфейса на JavaScript
Дисциплина: Основы HTML и CSS
Тема занятия №3 Введение в CSS. Форматирование текста при помощи CSS
1. Создать HTML-страницу “Vehicle”
Текст можете найти в архиве Classwork в папке Материалы к занятию.

2. Создать HTML-страницу “Lorem Ipsum”
Текст можете найти в архиве Classwork в папке Материалы к занятию.

3. Создать HTML-страницу “Математические формулы”

Текст можете найти в архиве Classwork в папке Материалы к занятию.
Используйте здесь теги h1-h6, span, p, sup, sub и спецсимволы.
<!-- ########################################################################################## -->

Шрифты находятся в папке Материалы к занятию в архиве Fonts.

1. Пример подключение шрифтов через ссылку
2. Пример подключения через локальные файлы шрифтов
<!-- ########################################################################################## -->

№1

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Двигатель</title>
    <style>
        body {
            font-family: 'Tahoma', serif;
            line-height: 1.6;
            margin: 18px;
        }

        h1 {
            color: #00c264;
        }
        h2 {
            color: #9b05ff;
        }
        h3 {
            color: #ff0505;
        }
        h4 {
            color: #00a0eb;
        }
        h5 {
            color: #00a0eb;
        }
        h6 {
            color: #ffff05;
        }
        p {
            text-align: justify;
        }
        .no-underline {
    text-decoration: none;
  }
        .larger-font {
    font-size: 20px;
  }
        .larger-font2 {
    font-size: 40px;
  }
  .smaller-text {
    font-size: 58%; /* Измените на желаемый размер шрифта */
  }
        .slightly-bold {
    font-weight: bold; /* или font-weight: 600; */
  }
         .subtle-line {
    border: 2px solid #9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; /* Добавьте отступы по желанию */
  }
         .colored-letter {
    color: #00a0eb; /* Измените на желаемый цвет */
  }
    </style>

</head>
<body>
    <h1
            class="larger-font2">
        Двигатель
    </h1>
    <!-- Текст о двигателе с сайта https://ru.wikipedia.org/wiki/Двигатель -->

    <p
        class="larger-font">
    <b>Дви́гатель</b> — устройство, преобразующее
        какой-либо вид <a href="https://ru.wikipedia.org/wiki/Энергия" class="no-underline">энергии</a>
         в <a href="https://ru.wikipedia.org/wiki/Механическая_энергия" class="no-underline">механическую</a> работу.
    Термин <b>мотор</b> <a href="https://ru.wikipedia.org/wiki/Заимствование"class="no-underline"> заимствован</a> в
        первой половине XIX века из немецкого языка<b
            class="smaller-text colored-letter">[1]</b>
        (нем. Motor — «двигатель», от лат. mōtor — «приводящий в движение») и преимущественно им называют электрические
        двигатели и двигатели внутреннего сгорания<b
            class="smaller-text colored-letter">
        [2]</b>.
    </p>
    <p
            class="larger-font">
        Двигатели подразделяют на первичные и вторичные.
    К первичным относят непосредственно преобразующие природные энергетические ресурсы в механическую работу,
    а ко вторичным — преобразующие энергию, выработанную или накопленную другими источниками<b
            class="smaller-text colored-letter">
        [3]</b>.
    </p>
<hr class="subtle-line">
    <p>
    <h4>Примечания</h4>
    <h5>[1]</h5>Шанский Н. М., Боброва Т. А. Кот // Школьный этимологический словарь русского языка.
    Происхождение слов. — 7-е изд., стереотип.. — М.: Дрофа, 2004. — 398, [2] с.
    <h5>[2]</h5>Крысин Л. П. Мотор // Толковый словарь иноязычных слов. — М.: Эксмо, 2008. — 944 с. —
    (Библиотека словарей).
    <h5>[3]</h5>Definition of motor | Dictionary.com (англ.). www.dictionary.com.
    Дата обращения: 27 января 2022. Архивировано 27 января 2022 года.
</p>
</body>
</html>
<!-- ########################################################################################## -->

№2

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Lorem Ipsum</title>
    <style>

        h1 {
            color: #5d0080;
        }
        h2 {
            color: #9b05ff;
        }
        h3 {
            color: #ff0505;
        }
        h4 {
            color: #00a0eb;
        }
        h5 {
            color: #00a0eb;
        }
        h6 {
            color: #ffff05;
        }
        p {
            text-align0: justify;
        }
          .left-aligned1 {
    text-align: left;
  }

  .right-aligned2 {
    text-align: right;
  }

  .custom-font {
    font-family: "Courier New", monospace;
  }
  .custom-font2 {
    font-family: "mv boli", monospace;
  }
  .custom-font3 {
    font-family: "segoe print", monospace;
  }
  .custom-font4 {
    font-family: "LUCIDA SANS CONSOLE", monospace;
  }
  .custom-font5 {
    font-family: "GABRIOLA", monospace;
  }
  .custom-font6 {
    font-family: "tempus sans itc", monospace;
  }
  .bold-text {
    font-weight: bold;
  }

  .italic-text {
    font-style: italic;
  }
        .no-underline {
    text-decoration: none;
  }
        .larger-font {
    font-size: 20px;
  }
        .larger-font2 {
    font-size: 40px;
  }
        .larger-font3 {
    font-size: 50px;
  }
  .centered-text {
    text-align: center;
  }

  .centered-container {
    margin: 0 auto;
    width: 50%; /* Измените на желаемую ширину */
  }

  .smaller-text {
    font-size: 95%; /* Измените на желаемый размер шрифта */
  }
   .smaller-text2 {
    font-size: 80%; /* Измените на желаемый размер шрифта */
  }
   .smaller-text3 {
    font-size: 115%; /* Измените на желаемый размер шрифта */
  }
   .smaller-text4 {
    font-size: 15%; /* Измените на желаемый размер шрифта */
  }
        .slightly-bold {
    font-weight: bold; /* или font-weight: 600; */
  }
         .subtle-line {
    border: 2px solid #9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; /* Добавьте отступы по желанию */
  }
  .subtle-line2 {
    border: 1px solid #f2f2f2; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; /* Добавьте отступы по желанию */
  }
         .colored-letter {
    color: #00a0eb; /* Измените на желаемый цвет */
  }
         .custom-color1 {
    color: #001fb8; /* Замените на желаемый цвет */
  }
         .custom-color2 {
    color: #6e6e6e; /* Замените на желаемый цвет */
  }
         .page-container {
    margin: 20px; /* Замените на желаемое значение отступа */
  }
    </style>
</head>
<body>
    <h1 class="centered-text larger-font3 custom-font4">Lorem Ipsum</h1>
<div class="right-aligned2 custom-color1 custom-font3 italic-text smaller-text">
  <p>"Neque porro quisquam est qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit...".</p>
</div>
    <div class="right-aligned2 custom-font2 custom-color2 italic-text smaller-text2 larger-font">
  <p>"Нет никого, кто любил бы боль саму по себе, кто искал бы её и кто хотел бы иметь её
      просто потому, что это боль..".</p>
</div>
    <hr class="subtle-line2">
    <!-- Пример текста Lorem Ipsum -->
    <div class="page-container">
    <p class="larger-font custom-font"><b>
        Классический текст Lorem Ipsum, используемый с XVI века
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>L</b>orem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod tempor incididunt
        ut labore et dolore magna aliqua. Ut enim ad minim veniam, quis nostrud exercitation ullamco laboris
        nisi ut aliquip ex ea commodo consequat. Duis aute irure dolor in reprehenderit in voluptate velit esse
        cillum dolore eu fugiat nulla pariatur. Excepteur sint occaecat cupidatat non proident, sunt in culpa qui
        officia deserunt mollit anim id est laborum."
</p>
    <p class="larger-font custom-font"><b>
        Абзац 1.10.32 "de Finibus Bonorum et Malorum", написанный Цицероном в 45 году н.э.
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>S</b>ed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem
    aperiam, eaque ipsa quae ab illo inventore veritatis et quasi architecto beatae vitae dicta sunt explicabo.
    Nemo enim ipsam voluptatem quia voluptas sit aspernatur aut odit aut fugit, sed quia consequuntur magni dolores
    eos qui ratione voluptatem sequi nesciunt. Neque porro quisquam est, qui dolorem ipsum quia dolor sit amet,
    consectetur, adipisci velit, sed quia non numquam eius modi tempora incidunt ut labore et dolore magnam aliquam
    quaerat voluptatem. Ut enim ad minima veniam, quis nostrum exercitationem ullam corporis suscipit laboriosam, nisi
    ut aliquid ex ea commodi consequatur? Quis autem vel eum iure reprehenderit qui in ea voluptate velit esse quam
    nihil molestiae consequatur, vel illum qui dolorem eum fugiat quo voluptas nulla pariatur?"
</p>
    <p class="larger-font custom-font"><b>
        Английский перевод 1914 года, H. Rackham
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>B</b>ut I must explain to you how all this mistaken idea of denouncing pleasure and praising pain was born and
    I will give you a complete account of the system, and expound the actual teachings of the great explorer of the
    truth, the master-builder of human happiness. No one rejects, dislikes, or avoids pleasure itself, because it is
    pleasure, but because those who do not know how to pursue pleasure rationally encounter consequences that are
    extremely painful. Nor again is there anyone who loves or pursues or desires to obtain pain of itself, because it
    is pain, but because occasionally circumstances occur in which toil and pain can procure him some great pleasure.
    To take a trivial example, which of us ever undertakes laborious physical exercise, except to obtain some advantage
    from it? But who has any right to find fault with a man who chooses to enjoy a pleasure that has no annoying
    consequences, or one who avoids a pain that produces no resultant pleasure?"
</p>
    <p class="larger-font custom-font"><b>
        Абзац 1.10.33 "de Finibus Bonorum et Malorum", написанный Цицероном в 45 году н.э.
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>A</b>t vero eos et accusamus et iusto odio dignissimos ducimus qui blanditiis praesentium voluptatum deleniti
    atque corrupti quos dolores et quas molestias excepturi sint occaecati cupiditate non provident, similique sunt
    in culpa qui officia deserunt mollitia animi, id est laborum et dolorum fuga. Et harum quidem rerum facilis est
    et expedita distinctio. Nam libero tempore, cum soluta nobis est eligendi optio cumque nihil impedit quo minus id
    quod maxime placeat facere possimus, omnis voluptas assumenda est, omnis dolor repellendus. Temporibus autem
    quibusdam et aut officiis debitis aut rerum necessitatibus saepe eveniet ut et voluptates repudiandae sint et
    molestiae non recusandae. Itaque earum rerum hic tenetur a sapiente delectus, ut aut reiciendis voluptatibus
    maiores alias consequatur aut perferendis doloribus asperiores repellat."
</p>
    <p class="larger-font custom-font"><b>
        Английский перевод 1914 года, H. Rackham
    </p></b>
<p class="custom-font6 smaller-text3">
        "<b>O</b>n the other hand, we denounce with righteous indignation and dislike men who are so beguiled and demoralized
    by the charms of pleasure of the moment, so blinded by desire, that they cannot foresee the pain and trouble that
    are bound to ensue; and equal blame belongs to those who fail in their duty through weakness of will, which is the
    same as saying through shrinking from toil and pain. These cases are perfectly simple and easy to distinguish.
    In a free hour, when our power of choice is untrammelled and when nothing prevents our being able to do what we
    like best, every pleasure is to be welcomed and every pain avoided. But in certain circumstances and owing to the
    claims of duty or the obligations of business it will frequently occur that pleasures have to be repudiated and
    annoyances accepted. The wise man therefore always holds in these matters to this principle of selection: he
    rejects pleasures to secure other greater pleasures, or else he endures pains to avoid worse pains."
</p>
        </div>
</body>
</html>

<!-- ########################################################################################## -->

№3

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Математические формулы</title>
    <style>
        /* Общие стили для всего документа */
        body {
            font-family: 'VERDANA', monospace; /* Устанавливаем шрифт для всего документа */
            line-height: 1.0; /* Задаем высоту строки текста */
            margin: 70px; /* Задаем отступы для всего документа */
        }

        /* Стили для заголовка первого уровня */
        h1 {
            color: #000000; /* Цвет текста */
            text-align: center; /* Выравнивание текста по центру */
            font-size: 50px; /* Размер шрифта */
            font-weight: bold; /* Жирный шрифт */
        }

        /* Стили для заголовка второго уровня */
        h2 {
            color: #22009e; /* Цвет текста */
            font-size: 35px; /* Размер шрифта */
            margin-top: 20px; /* Верхний отступ */
        }

        /* Стили для подзаголовка */
        h3 {
            color: #00ff49; /* Цвет текста */
            font-size: 25px; /* Размер шрифта */
            font-weight: bold; /* Жирный шрифт */
            font-style: italic; /* Курсив */
            text-align: left; /* Выравнивание текста влево */
            background-color: #a300a3b8; /* Цвет подложки */
            padding: 10px; /* Внутренний отступ */
            border-radius: 45px; /* Закругленные углы подложки */
            box-shadow: 0 0 10px rgba(0, 0, 0, 0.2); /* Тень подложки */
            max-width: 320px; /* Максимальная ширина подложки */
            width: 25%; /* Ширина подложки равна 20% родительского контейнера */
            box-sizing: border-box; /* Учитываем padding и border в общей ширине */
            margin: left; /* Выравнивание по левой стороне */
        }

        /* Стили для обычного текста и вложенных элементов */
        p, span {
            text-align: justify; /* Выравнивание текста по ширине с обеих сторон */
        }

        /* Стили для верхних и нижних индексов */
        sup, sub {
            font-size: 80%; /* Размер шрифта для верхних и нижних индексов */
        }

        /* Отступ для блока с описанием формулы */
        .formula-description {
            margin-left: 20px; /* Отступ слева */
        }

        /* Стили для контейнера формулы и самой формулы */
        .formula-container, .formula {
            text-align: center; /* Выравнивание текста по центру */
            background-color: #ff00008f; /* Цвет подложки */
            padding: 2px; /* Внутренний отступ */
            border-radius: 55px; /* Закругленные углы подложки */
            box-shadow: 0 0 15px rgba(0, 0, 0, 0.2); /* Тень подложки */
            max-width: 670px; /* Максимальная ширина подложки */
            width: 50%; /* Ширина подложки равна 50% родительского контейнера */
            box-sizing: border-box; /* Учитываем padding и border в общей ширине */
            margin: auto; /* Выравнивание по центру */
        }

        /* Стили для самой формулы */
        .formula {
            text-align: center; /* Выравнивание текста по центру */
            color: #0700ff; /* Цвет формулы */
            font-size: 25px; /* Размер шрифта формулы */
            font-style: italic; /* Курсив */
            font-weight: bold; /* Жирный шрифт формулы */
        }

        /* Стили для раздела с примерами */
        .example-section {
            margin-top: 20px; /* Верхний отступ раздела с примерами */
        }

        /* Стили для примеров */
        .example {
            font-size: 20px; /* Размер шрифта примеров */
            color: #555; /* Цвет текста примеров */
        }
         /* Стили для цветной буквы */
        .colored-letter {
            color: #ff0000; /* Цвет буквы */
        }
        /* Стили для увеличенной буквы */
        .enlarged-letter {
            font-size: 24px; /* Размер шрифта для увеличенной буквы */
        }
          /* Стили для выделенных букв */
        .special-letter {
            font-family: 'Arial', sans-serif; /* Используемый шрифт */
            font-style: italic; /* Курсив */
            color: #00adff; /* Цвет текста */
        }
    </style>
</head>
<body>
    <!-- Заголовок первого уровня -->
    <h1>Математические формулы</h1>

    <!-- Формула окружности -->
    <h2>Формула окружности:</h2>
    <!-- Контейнер формулы -->
    <div class="formula-container">
        <!-- Сама формула -->
        <p class="formula"><b> x<sup>2</sup> + y<sup>2</sup> = r<sup>2</sup></b></p>
    </div>
    <!-- Описание формулы -->
    <div class="formula-description">
        <p><b>Эта формула описывает уравнение окружности в декартовой системе координат.</b></p>
        <ul><b>
            <li><span class="colored-letter enlarged-letter special-letter">x</span> и <span class="colored-letter enlarged-letter special-letter">y</span> - координаты точки на плоскости.</li>
            <li><span class="colored-letter enlarged-letter special-letter">r</span> - радиус окружности.</li>
        </ul></b>
    </div>
    <!-- Подробное описание формулы -->
    <h3>Описание формулы:</h3>
    <p class="formula-description"><b>Формула окружности описывает геометрическую фигуру,
        представляющую собой множество точек, равноудаленных от центра.</b></p>
    <p class="formula-description"><b>Здесь x и y - координаты точек,
        а r - радиус окружности.</b></p>

    <!-- Формула площади треугольника -->
    <h2>Формула площади треугольника:</h2>
    <!-- Контейнер формулы -->
    <div class="formula-container">
        <!-- Сама формула -->
        <p class="formula"><b> S = 0.5 * a * b * sin(C)</b></p>
    </div>
    <!-- Описание формулы -->
    <div class="formula-description">
        <p><b>Где:</b></p>
        <ul><b>
            <li><span class="colored-letter enlarged-letter special-letter">S</span> - площадь треугольника.</li>
            <li><span class="colored-letter enlarged-letter special-letter">a</span> и <span class="colored-letter enlarged-letter special-letter">b</span> - длины сторон треугольника.</li>
            <li><span class="colored-letter enlarged-letter special-letter">C</span> - угол между сторонами a и b (в радианах).</li>
        </ul></b>
    </div>
    <!-- Подробное описание формулы -->
    <h3>Описание формулы:</h3>
    <p class="formula-description"><b>Формула площади треугольника выражает площадь как половину произведения длин двух сторон на синус угла между ними.</b></p>

    <!-- Формула объема цилиндра -->
    <h2>Формула объема цилиндра:</h2>
    <!-- Контейнер формулы -->
    <div class="formula-container">
        <!-- Сама формула -->
        <p class="formula"><b> V = π * r<sup>2</sup> * h</b></p>
    </div>
    <!-- Описание формулы -->
    <div class="formula-description">
        <p><b>Где:</b></p>
        <ul><b>
            <li><span class="colored-letter enlarged-letter special-letter">V</span> - объем цилиндра.</li>
            <li><span class="colored-letter enlarged-letter special-letter">π</span> (пи) - математическая константа, приблизительно равная 3.14159.</li>
            <li><span class="colored-letter enlarged-letter special-letter">r</span> - радиус основания цилиндра.</li>
            <li><span class="colored-letter enlarged-letter special-letter">h</span> - высота цилиндра.</li>
        </ul></b>
    </div>
    <!-- Подробное описание формулы -->
    <h3>Описание формулы:</h3>
    <p class="formula-description"><b>Формула объема цилиндра выражает объем как произведение площади основания на высоту.</b></p>

    <!-- Раздел с примерами применения формул -->
    <div class="example-section">
        <h2>Сферы применения формул:</h2>
        <!-- Примеры -->
        <p class="example">1. <b>Геометрия:</b> расчет геометрических параметров фигур.</p>
        <p class="example">2. <b>Инженерия:</b> проектирование и расчеты конструкций.</p>
        <p class="example">3. <b>Физика:</b> моделирование и анализ физических явлений.</p>
    </div>

</body>
</html>

<!-- ########################################################################################## -->

Бонус

1. Пример подключение шрифтов через ссылку

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Your Title</title>

    <!-- Подключение внешних шрифтов -->
    <link rel="stylesheet" href="https://fonts.googleapis.com/css?family=YourFontFamily">

    <!-- Ваши стили -->
    <style>
        body {
            font-family: 'YourFontFamily', sans-serif;
            /* Другие стили... */
        }
    </style>
</head>
<body>
    <!-- Ваш контент -->
</body>
</html>

<!--
Замените "https://fonts.googleapis.com/css?family=YourFontFamily" на фактическую ссылку на шрифт, который вы хотите
 использовать. Например, если вы хотите использовать шрифт "Roboto", ссылка будет
  выглядеть так: "https://fonts.googleapis.com/css?family=Roboto".

Обратите внимание, что это только пример, и вам нужно заменить "YourFontFamily" и "Your Title" на фактические
 значения, используемые в вашем проекте.
 -->

2. Пример подключения через локальные файлы шрифтов

<!--
Поместите шрифтовые файлы в ваш проект:
Сначала загрузите файлы шрифтов (например, .woff, .woff2, .ttf) в ваш проект.
Создайте папку, например, "fonts", и поместите файлы шрифтов в неё.

Пример структуры проекта:
 -->
/your-project
    /fonts
        your-font.woff
        your-font.woff2
    index.html
    style.css
<!--
Добавьте стили в ваш CSS:
В вашем файле стилей (например, style.css), определите новый @font-face и используйте его в стилях элементов.

Пример:
 -->
@font-face {
    font-family: 'YourFontFamily';
    src: url('fonts/your-font.woff2') format('woff2'),
         url('fonts/your-font.woff') format('woff');
    /* Дополнительные параметры, если нужны */
}

body {
    font-family: 'YourFontFamily', sans-serif;
    /* Другие стили... */
}
<!--
Убедитесь, что путь к файлам шрифтов в url соответствует структуре вашего проекта.
 -->
<!--
Подключите стили в HTML:
В вашем файле HTML (например, index.html), убедитесь, что вы подключили ваш CSS файл.

Пример:
 -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Your Title</title>

    <!-- Подключение локальных стилей -->
    <link rel="stylesheet" href="style.css">
</head>
<body>
    <!-- Ваш контент -->
</body>
</html>
<!--
Замените "Your Title" на фактический заголовок вашей страницы.

Это позволит вам использовать локальные файлы шрифтов в вашем проекте. Убедитесь, что пути указаны правильно и что
 браузер может найти и загрузить файлы шрифтов.
 -->
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->

Домашнее задание

Курс: Разработка интерфейса на JavaScript
Дисциплина: Основы HTML и CSS
Домашнее задание № 3 : Введение в CSS. Форматирование текста при помощи CSS
1. Попробуйте каждый тег

На каждой новой строчке напишите по предложению используя все
пройденные теги.

К каждому тегу добавьте атрибуты class или id

Затем измените с помощью CSS цвет этим предложениям (все должны
быть разного цвета)

Чем разнообразнее будут использоваться CSS свойства, тем выше вы получите
балл. CSS свойства можно также подключать или использовать по разному. Чем
разнообразнее, тем лучше. Удачи!

2. Создайте страницу “Romeo and Juliet”

3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы.

4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии.

<!-- ########################################################################################## -->

<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML and CSS Exercise</title>
    <style>
        /* Стили для разнообразных свойств */
        body {
            font-family: 'Arial', sans-serif;
            background-color: #f8f8f8;
            color: #333;
            margin: 20px;
        }

        #paragraph1 {
            font-size: 18px;
            font-weight: bold;
            color: #0066cc;
            text-decoration: underline;
        }

        .paragraph2 {
            font-style: italic;
            color: #cc0000;
            text-align: justify;
            margin-bottom: 20px;
        }

        .highlight {
            background-color: #ffcc00;
            padding: 5px;
        }

        #container {
            border: 2px solid #009900;
            padding: 10px;
            margin: 10px 0;
            background-color: #e6ffe6;
            border-radius: 10px;
        }

        .centered-text {
            text-align: center;
            color: #990099;
        }
    </style>
</head>
<body>
    <!-- Использование различных тегов с атрибутами и стилизацией -->
    <p id="paragraph1" class="highlight">Это первое предложение с различными тегами и атрибутами.</p>
    <p class="paragraph2" id="container">Второе предложение выделено разными стилями с использованием CSS.</p>
    <p class="centered-text">Третье предложение выровнено по центру и имеет свой цвет.</p>

    <!-- Примеры других тегов -->
    <div id="container">
        <h2>Пример заголовка второго уровня</h2>
        <ul>
            <li><strong>Список:</strong> элемент 1</li>
            <li><strong>Список:</strong> элемент 2</li>
        </ul>
        <a href="https://www.example.com" target="_blank" style="color: #cc00cc; text-decoration: none;">Ссылка на примерный сайт</a>
    </div>
</body>
</html>

<!-- ########################################################################################## -->


Домашнее задание

Курс: Разработка интерфейса на JavaScript
Дисциплина: Основы HTML и CSS

Домашнее задание №3: Введение в CSS. Форматирование текста при помощи CSS
<!-- #### -->
1. Попробуйте каждый тег
<!-- #### -->
На каждой новой строчке напишите по предложению используя все
пройденные теги.
К каждому тегу добавьте атрибуты class или id
Затем измените с помощью CSS цвет этим предложениям (все должны
быть разного цвета)
<!-- #### -->
Чем разнообразнее будут использоваться
CSS свойства, тем выше вы получите
балл. CSS свойства можно также подключать или использовать по разному. Чем
разнообразнее, тем лучше. Удачи!
<!-- #### -->
2. Создайте страницу “Romeo and Juliet”
<!-- #### -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы.
<!-- #### -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии.
<!-- #### -->

<!-- ########################################################################################## -->
1. Попробуйте каждый тег
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML and CSS Exercise</title>
    <style>
        /* Стили для разнообразных свойств */
        body {
            font-family: 'Arial', sans-serif;
            background-color: #f8f8f8;
            color: #333;
            margin: 20px;
        }

        #paragraph1 {
            font-size: 18px;
            font-weight: bold;
            color: #0066cc;
            text-decoration: underline;
        }

        .paragraph2 {
            font-style: italic;
            color: #cc0000;
            text-align: justify;
            margin-bottom: 20px;
        }

        .highlight {
            background-color: #ffcc00;
            padding: 5px;
        }

        #container {
            border: 2px solid #009900;
            padding: 10px;
            margin: 10px 0;
            background-color: #e6ffe6;
            border-radius: 10px;
        }

        .centered-text {
            text-align: center;
            color: #990099;
        }
    </style>
</head>
<body>
    <!-- Использование различных тегов с атрибутами и стилизацией -->
    <p id="paragraph1" class="highlight">Это первое предложение с различными тегами и атрибутами.</p>
    <p class="paragraph2" id="container">Второе предложение выделено разными стилями с использованием CSS.</p>
    <p class="centered-text">Третье предложение выровнено по центру и имеет свой цвет.</p>

    <!-- Примеры других тегов -->
    <div id="container">
        <h2>Пример заголовка второго уровня</h2>
        <ul>
            <li><strong>Список:</strong> элемент 1</li>
            <li><strong>Список:</strong> элемент 2</li>
        </ul>
        <a href="https://www.example.com" target="_blank" style="color: #cc00cc; text-decoration: none;">Ссылка на примерный сайт</a>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
2. Создайте страницу “Romario and Juluya”
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Romeo and Juliet</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            text-align: center;
        }

        #title {
            font-family: 'Georgia', serif;
            color: #FF4500; /* Цвет - оранжевый */
            font-style: italic;
            font-size: 36px;
        }

        #author {
            font-size: 18px;
            font-style: italic;
        }

        hr {
            width: 50%;
            margin: auto;
            margin-top: 20px;
            border: 1px solid #000;
        }

        #prologue {
            font-size: 24px;
            margin-top: 20px;
        }

        #left-text, #right-text {
            display: inline-block;
            width: 45%;
            vertical-align: top;
        }

        .act-title {
            font-size: 28px;
            margin-top: 20px;
        }

        .scene-title {
            font-size: 20px;
            font-weight: bold;
            margin-top: 10px;
        }

        .character {
            font-weight: bold;
        }

        .dialogue {
            text-align: left;
            margin-top: 10px;
            margin-bottom: 10px;
        }

        .romeo {
            color: #0000FF; /* Цвет - синий */
        }

        .juliet {
            color: #FF1493; /* Цвет - розовый */
        }

        .sampson {
            color: #008000; /* Цвет - зеленый */
        }

        .gregory {
            color: #FFD700; /* Цвет - золотой */
        }
    </style>
</head>
<body>
    <div id="title">ROMEO AND JULIET</div>
    <div id="author">By William Shakespeare</div>
    <hr>

    <div id="prologue">
        <div id="left-text">
            Two households, both alike in dignity,
            In fair Verona, where we lay our scene,
            From ancient grudge break to new mutiny,
            Where civil blood makes civil hands unclean.
            From forth the fatal loins of these two foes
            A pair of star-cross’d lovers take their life;
            Whole misadventured piteous overthrows
        </div>

        <div id="right-text">
            Do with their death bury their parents’ strife.
            The fearful passage of their death-mark’d love,
            And the continuance of their parents’ rage,
            Which, but their children’s end, nought could remove,
            Is now the two hours’ traffic of our stage;
            The which if you with patient ears attend,
            What here shall miss, our toil shall strive to mend.
        </div>
    </div>
    <hr>

    <div class="act-title">ACT I</div>
    <div class="scene-title">SCENE I. Verona. A public place.</div>
    <div class="scene-title"><i>Enter SAMPSON and GREGORY, of the house
    of CAPULET, armed with swords and bucklers</i></div>

    <div id="left-text" class="dialogue">
        <span class="character sampson">SAMPSON</span> Gregory, o’ my word, we’ll not carry coals.<br>
        <span class="character gregory">GREGORY</span> No, for then we should be colliers.<br>
        <span class="character sampson">SAMPSON</span> I mean, an we be in choler, we’ll draw.<br>
        <span class="character gregory">GREGORY</span> Ay, while you live, draw your neck out
        o’ the collar.<br>
        <span class="character sampson">SAMPSON</span> I strike quickly, being moved.<br>
        <span class="character gregory">GREGORY</span> But thou art not quickly moved to strike.<br>
        <span class="character sampson">SAMPSON</span> A dog of the house of Montague moves me.<br>
        <span class="character gregory">GREGORY</span> To move is to stir; and to be valiant is to
        stand: therefore, if thou art moved, thou runn’st away.<br>
    </div>

    <div id="right-text" class="dialogue">
        <span class="character sampson">SAMPSON</span> A dog of that house shall move me to stand:
        I will take the wall of any man or maid of Montague’s.<br>
        <span class="character gregory">GREGORY</span> That shows thee a weak slave; for the weakest
        goes to the wall.<br>
        <span class="character sampson">SAMPSON</span> True; and therefore women, being the
        weaker vessels, are ever thrust to the wall: therefore I
        will push Montague’s men from the wall, and thrust his
        maids to the wall.<br>
        <span class="character gregory">GREGORY</span> The quarrel is between our masters and us
        their men.<br>
        <span class="character sampson">SAMPSON</span> ’Tis all one, I will show myself a tyrant: when
        I have fought with the men, I will be cruel with the
        maids, and cut off their heads.<br>
        <span class="character gregory">GREGORY</span> The heads of the maids?<br>
        <span class="character sampson">SAMPSON</span> Ay, the heads of the maids, or their
        maidenheads; take it in what sense thou wilt.<br>
        <span class="character gregory">GREGORY</span> They must take it in sense that feel it.<br>
    </div>

</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №0.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 48px;
            margin-right: 15px;
            display: inline-block;
            border: 3px solid #000;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 15px;
            background-color: #fff;
            text-align: center;
        }

        .hearts { color: red; }
        .diamonds { color: #e74c3c; }
        .clubs { color: #2ecc71; }
        .spades { color: #3498db; }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card hearts">♥</div>
        <div class="card diamonds">♦</div>
        <div class="card diamonds">♦</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
        <div class="card hearts">♥</div>
        <!-- Дополнительные карты -->
        <div class="card diamonds">♦</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card hearts">♥</div>
        <div class="card diamonds">♦</div>
        <div class="card hearts">♥</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
        <div class="card hearts">♥</div>
        <!-- Дополнительные карты -->
        <div class="card diamonds">♦</div>
        <div class="card clubs">♣</div>
        <div class="card spades">♠</div>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №1.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 24px;
            margin-right: 10px;
        }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card">&#x1F0A1; Двойка - крести</div>
        <div class="card">&#x1F0C9; Девятка - черви</div>
        <div class="card">&#x1F0C2; Шестерка - черви</div>
        <div class="card">&#x1F0B4; Туз - буби</div>
        <div class="card">&#x1F0A9; Дама - пики</div>
        <div class="card">&#x1F0AC; Валет - крести</div>
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card">&#x1F0C8; Тройка - черви</div>
        <div class="card">&#x1F0A1; Туз - крести</div>
        <div class="card">&#x1F0AA; Десятка - крести</div>
        <div class="card">&#x1F0B7; Король - буби</div>
        <div class="card">&#x1F0A6; Восмёрка - пики</div>
        <div class="card">&#x1F0C7; Тройка - черви</div>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №2.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 36px;
            margin-right: 10px;
            display: inline-block;
            border: 2px solid #000;
            padding: 10px;
            border-radius: 8px;
        }

        /* Цвета мастей */
        .hearts { color: red; }
        .diamonds { color: #e74c3c; }
        .clubs { color: #2ecc71; }
        .spades { color: #3498db; }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card hearts">&#x1F0A1;</div>
        <div class="card diamonds">&#x1F0C9;</div>
        <div class="card diamonds">&#x1F0C2;</div>
        <div class="card clubs">&#x1F0B4;</div>
        <div class="card spades">&#x1F0A9;</div>
        <div class="card hearts">&#x1F0AC;</div>
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card hearts">&#x1F0C8;</div>
        <div class="card diamonds">&#x1F0A1;</div>
        <div class="card hearts">&#x1F0AA;</div>
        <div class="card clubs">&#x1F0B7;</div>
        <div class="card spades">&#x1F0A6;</div>
        <div class="card hearts">&#x1F0C7;</div>
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №3.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
        }

        h1 {
            color: #336699;
        }

        .player-cards {
            margin-top: 20px;
        }

        .card {
            font-size: 190px;
            margin-right: 15px;
            display: inline-block;
            border: 3px solid #000;
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 15px;
        }

        .hearts { color: red; }
        .diamonds { color: #e74c3c; }
        .clubs { color: #2ecc71; }
        .spades { color: #3498db; }

        /* Дополнительные стили для улучшения внешнего вида */
        p {
            font-size: 18px;
            margin-bottom: 10px;
        }

        strong {
            font-size: 24px;
        }
    </style>
</head>
<body>
    <h1>Игра в карты</h1>

    <div class="player-cards">
        <p><strong>У игрока 1:</strong></p>
        <div class="card hearts">&#x1F0A1;</div>
        <div class="card diamonds">&#x1F0C9;</div>
        <div class="card diamonds">&#x1F0C2;</div>
        <div class="card clubs">&#x1F0B4;</div>
        <div class="card spades">&#x1F0A9;</div>
        <div class="card hearts">&#x1F0AC;</div>
        <!-- Дополнительные карты -->
<!--        <div class="card diamonds">&#x1F0C4;</div>-->
<!--        <div class="card clubs">&#x1F0B3;</div>-->
<!--        <div class="card spades">&#x1F0AA;</div>-->
    </div>

    <div class="player-cards">
        <p><strong>У игрока 2:</strong></p>
        <div class="card hearts">&#x1F0C8;</div>
        <div class="card spades">&#x1F0A8;</div>
        <div class="card hearts">&#x1F0AA;</div>
        <div class="card clubs">&#x1F0B7;</div>
        <div class="card spades">&#x1F0A6;</div>
        <div class="card hearts">&#x1F0C7;</div>
        <!-- Дополнительные карты -->
<!--        <div class="card diamonds">&#x1F0C3;</div>-->
<!--        <div class="card clubs">&#x1F0B2;</div>-->
<!--        <div class="card diamonds">&#x1F0A1;</div>-->
    </div>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №4.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        /* Общие стили для всего документа */
        body {
            font-family: 'Arial', sans-serif; /* Шрифт */
            line-height: 1.6; /* Межстрочный интервал */
            margin: 20px; /* Отступы от краев */
            background-color: #f8f9fa; /* Цвет фона */
        }

        /* Стили для заголовка страницы */
        h1 {
            color: #336699; /* Цвет текста */
        }

        /* Стили для контейнеров карт игроков */
        .player-cards {
            margin-top: 20px; /* Верхний отступ */
            display: flex; /* Отображение карт в строку */
            flex-wrap: wrap; /* Перенос карт на следующую строку, если не помещаются */
        }

        /* Общие стили для карт */
        .card {
            margin-right: 15px; /* Правый отступ между картами */
            margin-bottom: 15px; /* Нижний отступ между картами */
            padding: 15px; /* Внутренний отступ */
            border: 3px solid #000; /* Обводка карты */
            border-radius: 10px; /* Закругление углов */
            background-color: #fff; /* Цвет фона карты */
            text-align: center; /* Выравнивание текста по центру */
            transition: transform 0.3s ease-in-out; /* Плавное изменение при наведении */
        }

        /* Стили для масти "черви" */
        .hearts { color: red; }

        /* Стили для масти "бубны" */
        .diamonds { color: #e74c3c; }

        /* Стили для масти "трефы" */
        .clubs { color: #2ecc71; }

        /* Стили для масти "пики" */
        .spades { color: #3498db; }
    </style>
</head>
<body>
    <!-- Заголовок страницы -->
    <h1>Игра в карты</h1>

    <!-- Ползунок для изменения размера карт -->
    <label for="cardSize">Размер карт:</label>
    <input type="range" id="cardSize" min="20" max="100" value="36">

    <!-- Контейнер для карт первого игрока -->
    <div class="player-cards" id="player1Cards"></div>

    <!-- Контейнер для карт второго игрока -->
    <div class="player-cards" id="player2Cards"></div>

    <!-- Скрипт на JavaScript для генерации, перемешивания, изменения размера и отображения карт -->
    <script>
        /* Масти и их символы */
        const suits = ['clubs', 'spades', 'diamonds', 'hearts'];
        const suitsSymbols = {
            'hearts': '♥',
            'diamonds': '♦',
            'clubs': '♣',
            'spades': '♠'
        };

        /* Ранги карт */
        const ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

        /* Функция для перемешивания колоды */
        function shuffleDeck() {
            let deck = [];
            for (let suit of suits) {
                for (let rank of ranks) {
                    deck.push({ suit, rank });
                }
            }
            return deck.sort(() => Math.random() - 0.5);
        }

        /* Функция для отображения карт в указанном контейнере с учетом размера */
        function displayCards(playerId, cards, cardSize) {
            const playerCardsElement = document.getElementById(playerId);
            playerCardsElement.innerHTML = '';
            for (let card of cards) {
                const cardElement = document.createElement('div');
                cardElement.style.fontSize = `${cardSize}px`; /* Изменение размера карты */
                /* Добавляем класс масти для стилизации цвета */
                cardElement.className = `card ${card.suit}`;
                /* Отображаем ранг карты и символ масти */
                cardElement.innerText = `${card.rank} ${suitsSymbols[card.suit]}`;
                playerCardsElement.appendChild(cardElement);
            }
        }

        /* Обработчик изменения размера карт при перемещении ползунка */
        document.getElementById('cardSize').addEventListener('input', function() {
            const newSize = this.value;
            displayCards('player1Cards', player1Cards, newSize);
            displayCards('player2Cards', player2Cards, newSize);
        });

        /* Генерируем и перемешиваем колоду */
        const deck = shuffleDeck();

        /* Раздача карт игрокам (первому и второму) */
        const player1Cards = deck.slice(0, 9);
        const player2Cards = deck.slice(9, 18);

        /* Отображение карт для каждого игрока с начальным размером 36px */
        displayCards('player1Cards', player1Cards, 36);
        displayCards('player2Cards', player2Cards, 36);
    </script>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
3. Создайте страницу “Card Game”. Для отображения карт используйте
спецсимволы. - вариант №5.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Игра в карты</title>
    <style>
        /* Общие стили для всего документа */
        body {
            font-family: 'Arial', sans-serif; /* Шрифт */
            line-height: 1.6; /* Межстрочный интервал */
            margin: 20px; /* Отступы от краев */
            background-color: #f8f9fa; /* Цвет фона */
        }

        /* Стили для заголовка страницы */
        h1 {
            color: #336699; /* Цвет текста */
        }

        /* Стили для контейнеров карт игроков */
        .player-cards {
            margin-top: 20px; /* Верхний отступ */
            display: flex; /* Отображение карт в строку */
            flex-wrap: wrap; /* Перенос карт на следующую строку, если не помещаются */
        }

        /* Общие стили для карт */
        .card {
            width: 100px; /* Ширина карты */
            height: 150px; /* Высота карты */
            margin-right: 15px; /* Правый отступ между картами */
            margin-bottom: 15px; /* Нижний отступ между картами */
            padding: 15px; /* Внутренний отступ */
            border: 3px solid #000; /* Обводка карты */
            border-radius: 10px; /* Закругление углов */
            background-color: #fff; /* Цвет фона карты */
            text-align: center; /* Выравнивание текста по центру */
            position: relative; /* Позиционирование элементов внутри карты */
        }

        /* Стили для масти "черви" */
        .hearts { color: red; }

        /* Стили для масти "бубны" */
        .diamonds { color: #e74c3c; }

        /* Стили для масти "трефы" */
        .clubs { color: #2ecc71; }

        /* Стили для масти "пики" */
        .spades { color: #3498db; }

        /* Стили для значения карты */
        .card-value {
            font-size: 14px; /* Размер шрифта значения карты */
            color: #555; /* Цвет текста значения карты */
            position: absolute; /* Абсолютное позиционирование значения */
        }

        /* Стили для левого верхнего угла карты */
        .card-value.top-left {
            top: 10px; /* Отступ от верхнего края карты */
            left: 10px; /* Отступ от левого края карты */
        }

        /* Стили для правого нижнего угла карты */
        .card-value.bottom-right {
            bottom: 10px; /* Отступ от нижнего края карты */
            right: 10px; /* Отступ от правого края карты */
        }
    </style>
</head>
<body>
    <!-- Заголовок страницы -->
    <h1>Игра в карты</h1>

    <!-- Ползунок для изменения размера карт -->
    <label for="cardSize">Размер карт:</label>
    <input type="range" id="cardSize" min="20" max="100" value="50">

    <!-- Выпадающий список для выбора масти -->
    <label for="suitSelect">Масть карты:</label>
    <select id="suitSelect">
        <option value="all">Все масти</option>
        <option value="hearts">Черви</option>
        <option value="diamonds">Бубны</option>
        <option value="clubs">Трефы</option>
        <option value="spades">Пики</option>
    </select>

    <!-- Выпадающий список для выбора ранга карты -->
    <label for="rankSelect">Ранг карты:</label>
    <select id="rankSelect">
        <option value="all">Все ранги</option>
        <option value="A">Туз</option>
        <option value="2">2</option>
        <option value="3">3</option>
        <option value="4">4</option>
        <option value="5">5</option>
        <option value="6">6</option>
        <option value="7">7</option>
        <option value="8">8</option>
        <option value="9">9</option>
        <option value="10">10</option>
        <option value="J">Валет</option>
        <option value="Q">Дама</option>
        <option value="K">Король</option>
    </select>

    <!-- Контейнер для карт первого игрока -->
    <div class="player-cards" id="player1Cards"></div>

    <!-- Контейнер для карт второго игрока -->
    <div class="player-cards" id="player2Cards"></div>

    <!-- Скрипт на JavaScript для генерации, перемешивания, изменения размера и отображения карт -->
    <script>
        /* Масти и их символы */
        const suits = ['clubs', 'spades', 'diamonds', 'hearts'];
        const suitsSymbols = {
            'hearts': '♥',
            'diamonds': '♦',
            'clubs': '♣',
            'spades': '♠'
        };

        /* Ранги карт */
        const ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

        /* Функция для перемешивания колоды */
        function shuffleDeck() {
            let deck = [];
            for (let suit of suits) {
                for (let rank of ranks) {
                    deck.push({ suit, rank });
                }
            }
            return deck.sort(() => Math.random() - 0.5);
        }

        /* Функция для отображения карт в указанном контейнере с учетом размера, масти и ранга */
        function displayCards(playerId, cards, cardSize, selectedSuit, selectedRank) {
            const playerCardsElement = document.getElementById(playerId);
            playerCardsElement.innerHTML = '';
            for (let card of cards) {
                if ((selectedSuit === 'all' || card.suit === selectedSuit) && (selectedRank === 'all' || card.rank === selectedRank)) {
                    const cardElement = document.createElement('div');
                    cardElement.style.width = `${cardSize}px`; /* Изменение ширины карты */
                    cardElement.style.height = `${cardSize * 1.5}px`; /* Изменение высоты карты */
                    /* Добавляем класс масти для стилизации цвета */
                    cardElement.className = `card ${card.suit}`;
                    /* Добавляем значение карты в верхний и нижний угол карты */
                    cardElement.innerHTML = `
                        <div class="card-value top-left">${card.rank}</div>
                        ${suitsSymbols[card.suit]}
                        <div class="card-value bottom-right">${card.rank}</div>`;
                    playerCardsElement.appendChild(cardElement);
                }
            }
        }

        /* Обработчик изменения размера карт при перемещении ползунка */
        document.getElementById('cardSize').addEventListener('input', function() {
            const newSize = this.value;
            const selectedSuit = document.getElementById('suitSelect').value;
            const selectedRank = document.getElementById('rankSelect').value;
            displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
            displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
        });

        /* Обработчик изменения масти карты */
        document.getElementById('suitSelect').addEventListener('change', function() {
            const newSize = document.getElementById('cardSize').value;
            const selectedSuit = this.value;
            const selectedRank = document.getElementById('rankSelect').value;
            displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
            displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
        });

        /* Обработчик изменения ранга карты */
        document.getElementById('rankSelect').addEventListener('change', function() {
            const newSize = document.getElementById('cardSize').value;
            const selectedSuit = document.getElementById('suitSelect').value;
            const selectedRank = this.value;
            displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
            displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
        });

        /* Генерируем и перемешиваем колоду */
        const deck = shuffleDeck();

        /* Раздача карт игрокам (первому и второму) */
        const player1Cards = deck.slice(0, 9);
        const player2Cards = deck.slice(9, 18);

        /* Отображение карт для каждого игрока с начальным размером 50px, мастью "все" и рангом "все" */
        displayCards('player1Cards', player1Cards, 50, 'all', 'all');
        displayCards('player2Cards', player2Cards, 50, 'all', 'all');
    </script>
</body>
</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №0.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="en">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Пример стилей для страницы (можно дополнительно настроить) */
        body {
            font-family: Arial, sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
        }

        img {
            max-width: 100%;
            height: auto;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }
    </style>
</head>

<body>

    <header>
        <h1>Welcome to the HTML Tutorial</h1>
    </header>

    <section>
        <article>
            <p>This HTML tutorial covers various HTML elements and their usage. Learn how to create a basic HTML
                structure and enhance your web development skills.</p>
        </article>

        <article>
            <h2>Basic HTML Structure</h2>
            <p>HTML documents consist of the following elements:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code> declaration</li>
                <li><code>&lt;html&gt;</code> element</li>
                <li><code>&lt;head&gt;</code> element</li>
                <li><code>&lt;title&gt;</code> element</li>
                <li><code>&lt;body&gt;</code> element</li>
            </ul>
        </article>

        <article>
            <h2>Text Formatting</h2>
            <p>Use tags like <code>&lt;p&gt;</code>, <code>&lt;h1&gt;</code>, <code>&lt;strong&gt;</code>, <code>&lt;em&gt;</code>,
                etc., for text formatting.</p>
        </article>

        <article>
            <h2>Images</h2>
            <p>Include images using the <code>&lt;img&gt;</code> tag.</p>
            <img src="https://placekitten.com/400/200" alt="Cute Kitten">
        </article>

        <article>
            <h2>Links</h2>
            <p>Create hyperlinks with the <code>&lt;a&gt;</code> tag.</p>
            <p>Visit our <a href="https://example.com" target="_blank">example website</a>.</p>
        </article>
    </section>

    <footer>
        <p>Explore more HTML elements and features to enhance your web development skills.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №1.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="ru">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Стили для страницы */
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
            color: #555;
        }

        img {
            max-width: 100%;
            height: auto;
            margin-bottom: 20px;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
            color: #333;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }

        code {
            background-color: #f8f9fa;
            padding: 2px 5px;
            border: 1px solid #ccc;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            color: #333;
        }
    </style>
</head>

<body>

    <header>
        <h1>Добро пожаловать в HTML-туториал</h1>
    </header>

    <section>
        <article>
            <p>HTML (HyperText Markup Language) — это стандартный язык разметки для создания веб-страниц. Он
                используется для структурирования контента в интернете, предоставляя основу для отображения текста,
                изображений, ссылок, форм и мультимедийных элементов.</p>
        </article>

        <article>
            <h2>Основная структура HTML</h2>
            <p>HTML-документ состоит из следующих элементов:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code>: Объявление версии HTML.</li>
                <li><code>&lt;html&gt;</code>: Корневой элемент HTML-страницы.</li>
                <li><code>&lt;head&gt;</code>: Содержит мета-информацию о документе.</li>
                <li><code>&lt;title&gt;</code>: Устанавливает заголовок HTML-документа (отображается во вкладке
                    браузера).</li>
                <li><code>&lt;body&gt;</code>: Содержит содержимое HTML-документа.</li>
            </ul>
            <p>Вот пример простой структуры HTML:</p>
            <code>
                &lt;!DOCTYPE html&gt;<br>
                &lt;html&gt;<br>
                &nbsp;&nbsp;&lt;head&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;title&gt;Моя первая HTML-страница&lt;/title&gt;<br>
                &nbsp;&nbsp;&lt;/head&gt;<br>
                &nbsp;&nbsp;&lt;body&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;h1&gt;Привет, мир!&lt;/h1&gt;<br>
                &nbsp;&nbsp;&lt;/body&gt;<br>
                &lt;/html&gt;
            </code>
        </article>

        <article>
            <h2>Форматирование текста</h2>
            <p>HTML предоставляет различные теги для форматирования текста:</p>
            <ul>
                <li><code>&lt;p&gt;</code>: Параграф</li>
                <li><code>&lt;h1&gt;</code> до <code>&lt;h6&gt;</code>: Заголовки</li>
                <li><code>&lt;strong&gt;</code>: Жирный текст</li>
                <li><code>&lt;em&gt;</code>: Курсив</li>
            </ul>
            <p>Пример:</p>
            <code>
                &lt;p&gt;Это &lt;strong&gt;параграф&lt;/strong&gt; с жирным текстом.&lt;/p&gt;<br>
                &lt;h2&gt;Подзаголовок&lt;/h2&gt;<br>
                &lt;p&gt;Это &lt;em&gt;курсивизированное&lt;/em&gt; предложение.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Изображения</h2>
            <p>Вставка изображений с использованием тега <code>&lt;img&gt;</code>:</p>
            <code>
                &lt;img src="image.jpg" alt="Описание изображения"&gt;
            </code>
        </article>

        <article>
            <h2>Ссылки</h2>
            <p>Создание гиперссылок с тегом <code>&lt;a&gt;</code>:</p>
            <code>
                &lt;a href="https://example.com" target="_blank"&gt;Посетите веб-сайт примера&lt;/a&gt;
            </code>
        </article>

        <article>
            <h2>Цветной текст</h2>
            <p>Применение цветового стиля к тексту:</p>
            <code>
                &lt;p style="color: #e44d26;"&gt;Этот текст оранжевого цвета.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Списки</h2>
            <p>Использование тегов <code>&lt;ul&gt;</code> и <code>&lt;li&gt;</code> для создания списков:</p>
            <code>
                &lt;ul&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Первый элемент списка&lt;/li&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Второй элемент списка&lt;/li&gt;<br>
                &lt;/ul&gt;
            </code>
        </article>

        <article>
            <h2>Видео</h2>
            <p>Вставка видео с помощью тега <code>&lt;iframe&gt;</code>:</p>
            <code>
                &lt;iframe width="560" height="315" src="https://www.youtube.com/embed/ВАШ_KОД" frameborder="0"
                allowfullscreen&gt;&lt;/iframe&gt;
            </code>
        </article>
    </section>

    <footer>
        <p>Исследуйте больше элементов и возможностей HTML для улучшения своих навыков веб-разработки.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №1_1.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="ru">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Стили для страницы */
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
            color: #555;
        }

        img {
            max-width: 100%;
            height: auto;
            margin-bottom: 20px;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
            color: #333;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }

        code {
            background-color: #f8f9fa;
            padding: 2px 5px;
            border: 1px solid #ccc;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            color: #333;
        }

        .styled-text {
            font-size: 18px;
            font-weight: bold;
            color: #e44d26;
            font-style: italic;
            text-decoration: underline;
        }
    </style>
</head>

<body>

    <header>
        <h1>Добро пожаловать в HTML-туториал</h1>
    </header>

    <section>
        <article>
            <p>HTML (HyperText Markup Language) — это стандартный язык разметки для создания веб-страниц.</p> <p>Он
                используется для структурирования контента в интернете, предоставляя основу для отображения текста,
                изображений, ссылок, форм и мультимедийных элементов.</p>
        </article>

        <article>
            <h2>Основная структура HTML</h2>
            <p>HTML-документ состоит из следующих элементов:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code>: Объявление версии HTML.</li>
                <li><code>&lt;html&gt;</code>: Корневой элемент HTML-страницы.</li>
                <li><code>&lt;head&gt;</code>: Содержит мета-информацию о документе.</li>
                <li><code>&lt;title&gt;</code>: Устанавливает заголовок HTML-документа (отображается во вкладке
                    браузера).</li>
                <li><code>&lt;body&gt;</code>: Содержит содержимое HTML-документа.</li>
            </ul>
            <p>Вот пример простой структуры HTML:</p>
            <code>
                &lt;!DOCTYPE html&gt;<br>
                &lt;html&gt;<br>
                &nbsp;&nbsp;&lt;head&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;title&gt;Моя первая HTML-страница&lt;/title&gt;<br>
                &nbsp;&nbsp;&lt;/head&gt;<br>
                &nbsp;&nbsp;&lt;body&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;h1&gt;Привет, мир!&lt;/h1&gt;<br>
                &nbsp;&nbsp;&lt;/body&gt;<br>
                &lt;/html&gt;
            </code>
        </article>

        <article>
            <h2>Форматирование текста</h2>
            <p>HTML предоставляет различные теги для форматирования текста:</p>
            <ul>
                <li><code>&lt;p&gt;</code>: Параграф</li>
                <li><code>&lt;h1&gt;</code> до <code>&lt;h6&gt;</code>: Заголовки</li>
                <li><code>&lt;strong&gt;</code>: Жирный текст</li>
                <li><code>&lt;em&gt;</code>: Курсив</li>
            </ul>
            <p>Пример:</p>
            <code>
                &lt;p&gt;Это &lt;strong&gt;параграф&lt;/strong&gt; с жирным текстом.&lt;/p&gt;<br>
                &lt;h2&gt;Подзаголовок&lt;/h2&gt;<br>
                &lt;p&gt;Это &lt;em&gt;курсивизированное&lt;/em&gt; предложение.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Изображения</h2>
            <p>Вставка изображений с использованием тега <code>&lt;img&gt;</code>:</p>
            <img src="https://placekitten.com/800/400" alt="Котенок">
        </article>

        <article>
            <h2>Ссылки</h2>
            <p>Создание гиперссылок с тегом <code>&lt;a&gt;</code>:</p>
            <code>
                &lt;a href="https://example.com" target="_blank"&gt;Посетите веб-сайт примера&lt;/a&gt;
            </code>
        </article>

        <article>
            <h2>Цветной текст</h2>
            <p>Применение цветового стиля к тексту:</p>
            <code>
                &lt;p style="color: #e44d26;"&gt;Этот текст оранжевого цвета.&lt;/p&gt;
            </code>
        </article>

        <article>
            <h2>Стилизованный текст</h2>
            <p>Пример изменения шрифта, стиля, размера и цвета:</p>
            <span class="styled-text">Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого цвета и
                размера 18px.</span>
        </article>

        <article>
            <h2>Списки</h2>
            <p>Использование тегов <code>&lt;ul&gt;</code> и <code>&lt;li&gt;</code> для создания списков:</p>
            <code>
                &lt;ul&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Первый элемент списка&lt;/li&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Второй элемент списка&lt;/li&gt;<br>
                &lt;/ul&gt;
            </code>
        </article>

        <article>
            <h2>Видео</h2>
            <p>Вставка видео с помощью тега <code>&lt;iframe&gt;</code>:</p>
            <iframe width="560" height="315" src="https://www.youtube.com/embed/dQw4w9WgXcQ" frameborder="0"
                allowfullscreen></iframe>
        </article>
    </section>

    <footer>
        <p>Исследуйте больше элементов и возможностей HTML для улучшения своих навыков веб-разработки.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”. Для выполнения используйте все теги,
которые прошли на занятии. - вариант №2.
<!-- ########################################################################################## -->
<!DOCTYPE html>
<html lang="ru">

<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>HTML Tutorial</title>
    <style>
        /* Стили для страницы */
        body {
            font-family: 'Arial', sans-serif;
            line-height: 1.6;
            margin: 20px;
            background-color: #f8f9fa;
            color: #333;
        }

        h1 {
            color: #007bff;
        }

        p {
            margin-bottom: 20px;
            color: #555;
        }

        img {
            max-width: 100%;
            height: auto;
            margin-bottom: 20px;
        }

        ul {
            list-style-type: disc;
            margin-left: 20px;
        }

        li {
            margin-bottom: 10px;
            color: #333;
        }

        a {
            color: #28a745;
            text-decoration: none;
            font-weight: bold;
        }

        a:hover {
            text-decoration: underline;
        }

        code {
            background-color: #f8f9fa;
            padding: 2px 5px;
            border: 1px solid #ccc;
            border-radius: 3px;
            font-family: 'Courier New', monospace;
            color: #333;
        }

        .styled-text {
            font-size: 18px;
            font-weight: bold;
            color: #e44d26;
            font-style: italic;
            text-decoration: underline;
        }

        hr {
            margin: 20px 0;
            border: none;
            border-top: 2px solid #007bff;
        }
    </style>
</head>

<body>

    <header>
        <h1>Добро пожаловать в HTML-туториал</h1>
    </header>

    <section>
        <article>
            <p><b>HTML</b> (Hypertext Markup Language) - это код, который используется для структурирования и отображения веб-страницы и её контента.</p>
            <p>Например, контент может быть структурирован внутри множества параграфов, маркированных списков или с использованием изображений и таблиц данных.</p>
            <p>Как видно из названия, этот туториал постарается дать вам базовое понимание HTML и его функций.</p>
        </article>
        <hr>
        <p><b>HTML</b> не является языком программирования; это язык разметки, и используется, чтобы сообщать вашему браузеру, как отображать веб-страницы, которые вы посещаете.</p>
        <p>Он может быть сложным или простым, в зависимости от того, как хочет веб-дизайнер.</p>
        <p>HTML состоит из ряда элементов, которые вы используете, чтобы вкладывать или оборачивать различные части контента, чтобы заставить контент отображаться или действовать определённым образом.</p>
        <p>Ограждающие теги могут сделать слово или изображение ссылкой на что-то ещё, могут сделать слова курсивом, сделать шрифт больше или меньше и так далее.</p>
        <hr>

        <article>
            <h2>Основная структура HTML</h2>
            <p>HTML-документ состоит из следующих элементов:</p>
            <ul>
                <li><code>&lt;!DOCTYPE html&gt;</code>: Объявление версии HTML.</li>
                <li><code>&lt;html&gt;</code>: Корневой элемент HTML-страницы.</li>
                <li><code>&lt;head&gt;</code>: Содержит мета-информацию о документе.</li>
                <li><code>&lt;title&gt;</code>: Устанавливает заголовок HTML-документа (отображается во вкладке
                    браузера).</li>
                <li><code>&lt;body&gt;</code>: Содержит содержимое HTML-документа.</li>
            </ul>
            <p>Вот пример простой структуры HTML:</p>
            <code>
                &lt;!DOCTYPE html&gt;<br>
                &lt;html&gt;<br>
                &nbsp;&nbsp;&lt;head&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;title&gt;Моя первая HTML-страница&lt;/title&gt;<br>
                &nbsp;&nbsp;&lt;/head&gt;<br>
                &nbsp;&nbsp;&lt;body&gt;<br>
                &nbsp;&nbsp;&nbsp;&nbsp;&lt;h1&gt;Привет, мир!&lt;/h1&gt;<br>
                &nbsp;&nbsp;&lt;/body&gt;<br>
                &lt;/html&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Форматирование текста</h2>
            <p>HTML предоставляет различные теги для форматирования текста:</p>
            <ul>
                <li><code>&lt;p&gt;</code>: Параграф</li>
                <li><code>&lt;h1&gt;</code> до <code>&lt;h6&gt;</code>: Заголовки</li>
                <li><code>&lt;strong&gt;</code>: Жирный текст</li>
                <li><code>&lt;em&gt;</code>: Курсив</li>
            </ul>
            <p>Пример использования тегов форматирования текста:</p>
            <code>
                &lt;p&gt;Это &lt;strong&gt;параграф&lt;/strong&gt; с жирным текстом.&lt;/p&gt;<br>
                &lt;h2&gt;Подзаголовок&lt;/h2&gt;<br>
                &lt;p&gt;Это &lt;em&gt;курсивизированное&lt;/em&gt; предложение.&lt;/p&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Изображения</h2>
            <p>Вставка изображений с использованием тега <code>&lt;img&gt;</code>:</p>
            <img src="https://placekitten.com/800/400" alt="Котенок">
        </article>

        <hr>

        <article>
            <h2>Ссылки</h2>
            <p>Создание гиперссылок с тегом <code>&lt;a&gt;</code>:</p>
            <code>
                &lt;a href="https://example.com" target="_blank"&gt;Посетите веб-сайт примера&lt;/a&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Цветной текст</h2>
            <p>Применение цветового стиля к тексту:</p>
            <code>
                &lt;p style="color: #e44d26;"&gt;Этот текст оранжевого цвета.&lt;/p&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Стилизованный текст</h2>
            <p>Пример изменения шрифта, стиля, размера и цвета:</p>
            <span class="styled-text">Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого цвета и
                размера 18px.</span>
        </article>

        <hr>

        <article>
            <h2>Списки</h2>
            <p>Использование тегов <code>&lt;ul&gt;</code> и <code>&lt;li&gt;</code> для создания списков:</p>
            <code>
                &lt;ul&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Первый элемент списка&lt;/li&gt;<br>
                &nbsp;&nbsp;&lt;li&gt;Второй элемент списка&lt;/li&gt;<br>
                &lt;/ul&gt;
            </code>
        </article>

        <hr>

        <article>
            <h2>Видео</h2>
            <p>Вставка видео с помощью тега <code>&lt;iframe&gt;</code>:</p>
            <iframe width="560" height="315" src="https://www.youtube.com/embed/dQw4w9WgXcQ" frameborder="0"
                allowfullscreen></iframe>
        </article>
    </section>

    <footer>
        <p>Исследуйте больше элементов и возможностей HTML для улучшения своих навыков веб-разработки.</p>
    </footer>

</body>

</html>
<!-- ########################################################################################## -->
<!-- ########################################################################################## -->

'''
Дата выполнения Практической-Работы: 19-20 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №26: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите одиночный json – объект с сайта jsonplaceholder, используя библиотеку requests.
б) Сохраните его в файл.

'''
'''
Урок от 24.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: (а и б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №1.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# Импорт библиотек
import requests
import json


# Определение функции для загрузки и сохранения JSON
def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        # Вывод сообщения об успешной загрузке и сохранении
        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        # Вывод сообщения об ошибке при загрузке или сохранении JSON
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


# Указание URL и имени файла, куда сохранить JSON
json_url = "https://jsonplaceholder.typicode.com/posts/1"
output_filename = "output.json"

# Вызов функции для загрузки и сохранения JSON
download_and_save_json(json_url, output_filename)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import requests
import json

'''
Описание:

requests: Эта библиотека используется для выполнения HTTP-запросов.
json: Эта библиотека предоставляет методы для работы с данными в формате JSON.
'''
'''
Шаг 2: Определение функции download_and_save_json
'''


def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


'''
Описание:

download_and_save_json: Это пользовательская функция, которая принимает два аргумента - url и filename.
requests.get(url): Этот метод выполняет GET-запрос к указанному url и возвращает объект Response.
response.raise_for_status(): Проверяет, были ли ошибки в результате запроса. Если есть, вызывает исключение HTTPError.
open(filename, 'w'): Открывает файл с указанным именем для записи.
json.dump(response.json(), file, indent=2): Записывает JSON-объект из ответа в файл с отступом в 2 пробела.
print(f"JSON-объект успешно загружен и сохранен в файл {filename}"): 
Выводит сообщение об успешной загрузке и сохранении.
'''
'''
Шаг 3: Указание URL и имени файла и вызов функции
'''
# Укажите URL и имя файла, куда сохранить JSON
json_url = "https://jsonplaceholder.typicode.com/posts/1"
output_filename = "output.json"

# Вызов функции для загрузки и сохранения JSON
download_and_save_json(json_url, output_filename)
'''
Описание:

Здесь вы устанавливаете URL и имя файла для JSON-данных.
Затем вызывается функция download_and_save_json с указанными аргументами.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вариант №2.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# Импорт библиотек
import requests
import json


# Определение функции для загрузки и сохранения JSON
def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        # Вывод сообщения об успешной загрузке и сохранении
        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        # Вывод сообщения об ошибке при загрузке или сохранении JSON
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


# Указание URL и имени файла, куда сохранить JSON для поста
post_url = "https://jsonplaceholder.typicode.com/posts/1"
post_filename = "post_output.json"

# Вызов функции для загрузки и сохранения JSON для поста
download_and_save_json(post_url, post_filename)

# Указание URL и имени файла, куда сохранить JSON для комментариев
comments_url = "https://jsonplaceholder.typicode.com/comments?postId=1"
comments_filename = "comments_output.json"

# Вызов функции для загрузки и сохранения JSON для комментариев
download_and_save_json(comments_url, comments_filename)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import requests
import json

'''
Описание:

requests: Эта библиотека используется для выполнения HTTP-запросов.
json: Эта библиотека предоставляет методы для работы с данными в формате JSON.
'''
'''
Шаг 2: Определение функции download_and_save_json
'''


def download_and_save_json(url, filename):
    try:
        # Загрузка JSON-объекта с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение JSON-объекта в файл
        with open(filename, 'w') as file:
            json.dump(response.json(), file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


'''
Описание:

download_and_save_json: Это пользовательская функция, которая принимает два аргумента - url и filename.
requests.get(url): Этот метод выполняет GET-запрос к указанному url и возвращает объект Response.
response.raise_for_status(): Проверяет, были ли ошибки в результате запроса. Если есть, вызывает исключение HTTPError.
open(filename, 'w'): Открывает файл с указанным именем для записи.
json.dump(response.json(), file, indent=2): Записывает JSON-объект из ответа в файл с отступом в 2 пробела.
print(f"JSON-объект успешно загружен и сохранен в файл {filename}"): 
Выводит сообщение об успешной загрузке и сохранении.
'''
'''
Шаг 3: Указание URL и имени файла, вызов функции для поста
'''
# Укажите URL и имя файла, куда сохранить JSON для поста
post_url = "https://jsonplaceholder.typicode.com/posts/1"
post_filename = "post_output.json"

# Вызов функции для загрузки и сохранения JSON для поста
download_and_save_json(post_url, post_filename)
'''
Описание:

Здесь вы устанавливаете URL и имя файла для JSON-данных поста.
Затем вызывается функция download_and_save_json с указанными аргументами.
'''
'''
Шаг 4: Указание URL и имени файла, вызов функции для комментариев
'''
# Укажите URL и имя файла, куда сохранить JSON для комментариев
comments_url = "https://jsonplaceholder.typicode.com/comments?postId=1"
comments_filename = "comments_output.json"

# Вызов функции для загрузки и сохранения JSON для комментариев
download_and_save_json(comments_url, comments_filename)
'''
Описание:

Здесь вы устанавливаете URL и имя файла для JSON-данных комментариев.
Затем снова вызывается функция download_and_save_json с указанными аргументами.
'''
'''
В итоге, код выполняет два HTTP-запроса, загружает JSON-объекты и сохраняет их в соответствующие файлы. 
Вся логика организована внутри функции download_and_save_json, 
что делает код более структурированным и переиспользуемым.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Вы можете загружать JSON-объекты с различных открытых API, 
которые предоставляют данные в формате JSON. Например, некоторые 
популярные открытые API, которые предоставляют JSON-данные:

1. JSONPlaceholder: 
Это фейковый онлайн-сервис для тестирования и прототипирования, который предоставляет тестовые данные в формате JSON.
Вы уже использовали его в предыдущем примере.

    URL: https://jsonplaceholder.typicode.com

2. OpenWeatherMap API: Предоставляет информацию о погоде в формате JSON.

    URL: https://openweathermap.org/api

3. GitHub API: Предоставляет данные о репозиториях, пользователях и других аспектах GitHub.

    URL: https://developer.github.com/v3

4. REST Countries API: Предоставляет информацию о странах в формате JSON.

    URL: https://restcountries.com

Просто выберите API, которое соответствует вашим интересам, и используйте его URL для загрузки JSON-данных. 
Не забывайте, что при использовании сторонних API следует ознакомиться с их документацией и условиями использования.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


'''
Дата выполнения Домашней-Работы: 24-25 - ЯНВАРЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №26: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите массив json – объектов с сайта jsonplaceholder, используя библиотеку requests.
б) Сохраните циклом каждый в отдельный файл, в одну новую папку.

'''
'''
Урок от 24.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import requests
import json


def download_and_save_json_objects(url, output_folder):
    try:
        # Загрузка массива JSON-объектов с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Разбор JSON-объектов
        json_objects = response.json()

        # Создание новой папки, если она не существует
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Цикл по каждому JSON-объекту и сохранение в отдельный файл
        for i, json_object in enumerate(json_objects):
            filename = os.path.join(output_folder, f"object_{i + 1}.json")
            with open(filename, 'w') as file:
                json.dump(json_object, file, indent=2)

            print(f"JSON-объект {i + 1} успешно сохранен в файл {filename}")

    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


# Укажите URL и имя новой папки для сохранения JSON-объектов
json_url = "https://jsonplaceholder.typicode.com/posts"
output_folder = "json_objects_folder"

# Вызов функции для загрузки и сохранения JSON-объектов
download_and_save_json_objects(json_url, output_folder)
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек
'''
import os
import requests
import json

'''
Описание:

os: Эта библиотека предоставляет функционал для работы с операционной системой, в данном случае, для создания папок.
requests: Эта библиотека используется для выполнения HTTP-запросов.
json: Эта библиотека предоставляет методы для работы с данными в формате JSON.
'''
'''
Шаг 2: Определение функции download_and_save_json_objects
'''


def download_and_save_json_objects(url, output_folder):
    try:
        # Загрузка массива JSON-объектов с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Разбор JSON-объектов
        json_objects = response.json()

        # Создание новой папки, если она не существует
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # Цикл по каждому JSON-объекту и сохранение в отдельный файл
        for i, json_object in enumerate(json_objects):
            filename = os.path.join(output_folder, f"object_{i + 1}.json")
            with open(filename, 'w') as file:
                json.dump(json_object, file, indent=2)

            print(f"JSON-объект {i + 1} успешно сохранен в файл {filename}")

    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")


'''
Описание:

download_and_save_json_objects: Это пользовательская функция, которая принимает два аргумента - url и output_folder.
requests.get(url): Этот метод выполняет GET-запрос к указанному url и возвращает объект Response.
response.raise_for_status(): Проверяет, были ли ошибки в результате запроса. Если есть, вызывает исключение HTTPError.
json_objects = response.json(): Разбирает JSON-объекты из ответа.
os.makedirs(output_folder): Создает новую папку, если она не существует.
enumerate(json_objects): Возвращает индекс и текущий элемент из списка JSON-объектов.
os.path.join(output_folder, f"object_{i + 1}.json"): Формирует полный путь к файлу с учетом созданной папки и 
индекса объекта.
json.dump(json_object, file, indent=2): Записывает JSON-объект в файл с отступом в 2 пробела.
print(f"JSON-объект {i + 1} успешно сохранен в файл {filename}"): Выводит сообщение об успешном сохранении.
'''
'''
Шаг 3: Указание URL и имени новой папки, вызов функции
'''
# Укажите URL и имя новой папки для сохранения JSON-объектов
json_url = "https://jsonplaceholder.typicode.com/posts"
output_folder = "json_objects_folder"

# Вызов функции для загрузки и сохранения JSON-объектов
download_and_save_json_objects(json_url, output_folder)
'''
Описание:

Здесь мы устанавливаем URL и имя новой папки для JSON-данных.
Затем вызывается функция download_and_save_json_objects с указанными аргументами.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Да, так-же мы можем использовать любой из предложенных сервисов для загрузки массива JSON-объектов.
Все в точности как и в практической работе №26.

Примеры URL для загрузки данных:

1. JSONPlaceholder:

    URL для постов: https://jsonplaceholder.typicode.com/posts
    URL для комментариев: https://jsonplaceholder.typicode.com/comments
    Используем любой из этих URL в зависимости от ваших потребностей.

2. OpenWeatherMap API:

    Пример URL для получения данных о погоде 
    в Лондоне: https://api.openweathermap.org/data/2.5/weather?q=London&appid=YOUR_API_KEY
    Необходимо заменить YOUR_API_KEY на необходимый ключ API, полученный после регистрации на сайте OpenWeatherMap.

3. GitHub API:

    Документация GitHub API: https://developer.github.com/v3
    Мы можем использовать различные эндпойнты API в зависимости от того, какие данные нужны.

4. REST Countries API:

    Пример URL для получения информации о странах: https://restcountries.com/v3.1/all
    Этот URL предоставит информацию о всех странах. 
    Мы также можем использовать другие эндпойнты для более конкретных запросов.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Дата выполнения Практической-Работы: 26-27 - ЯНВАРЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №27: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите одиночный json – объект с сайта jsonplaceholder, используя библиотеку aiohttp.
б) Сохраните его в файл.

'''
'''
Урок от 26.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: (а и б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import json

async def download_and_save_json(url, filename):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                json_data = await response.json()

                with open(filename, 'w') as file:
                    json.dump(json_data, file, indent=2)

                print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")

async def main():
    # Укажите URL и имя файла, куда сохранить JSON
    json_url = "https://jsonplaceholder.typicode.com/users/1"
    output_filename = "output_async.json"

    # Вызов функции для загрузки и сохранения JSON
    await download_and_save_json(json_url, output_filename)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Импорт библиотек:
'''
import aiohttp
import asyncio
import json
''''
Описание:

Импортируются необходимые библиотеки: aiohttp для асинхронных HTTP-запросов, asyncio для асинхронного 
программирования, json для работы с JSON-данными.
''''
'''
Шаг 2: Определение асинхронной функции download_and_save_json:
'''
async def download_and_save_json(url, filename):
''''
Описание:

mЭта функция асинхронная и предназначена для загрузки JSON-данных по заданному URL и сохранения их в файл.
''''
'''
Шаг 3: Блок try-except внутри функции:
'''
try:
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            response.raise_for_status()
            json_data = await response.json()

            with open(filename, 'w') as file:
                json.dump(json_data, file, indent=2)

            print(f"JSON-объект успешно загружен и сохранен в файл {filename}")
except aiohttp.ClientError as e:
    print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")
'''
Описание:

Создается асинхронная сессия с помощью aiohttp.ClientSession() для выполнения HTTP-запросов.
Выполняется асинхронный GET-запрос с использованием session.get(url).
Проверяется наличие ошибок при запросе с response.raise_for_status().
Получается JSON-данные из ответа с использованием await response.json().
Сохраняются JSON-данные в файл с использованием json.dump.
Обрабатываются и выводятся ошибки с использованием aiohttp.ClientError.
'''
'''
Шаг 4: Определение асинхронной функции main:
'''
async def main():
''''
Описание:

Эта функция предназначена для запуска основной асинхронной программы.
'''''
'''
Шаг 5: Вызов функции download_and_save_json из main:
'''
await download_and_save_json(json_url, output_filename)
'''
Вызывается ранее определенная функция download_and_save_json.
'''
'''
Шаг 6: Запуск асинхронной программы:
'''
if __name__ == "__main__":
    asyncio.run(main())
'''
Описание:

Запуск основной асинхронной программы с использованием asyncio.run(main()).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'output_async.json'
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
{
  "id": 1,
  "name": "Leanne Graham",
  "username": "Bret",
  "email": "Sincere@april.biz",
  "address": {
    "street": "Kulas Light",
    "suite": "Apt. 556",
    "city": "Gwenborough",
    "zipcode": "92998-3874",
    "geo": {
      "lat": "-37.3159",
      "lng": "81.1496"
    }
  },
  "phone": "1-770-736-8031 x56442",
  "website": "hildegard.org",
  "company": {
    "name": "Romaguera-Crona",
    "catchPhrase": "Multi-layered client-server neural-net",
    "bs": "harness real-time e-markets"
  }
}
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Мы также можем загружать JSON-объекты с различных открытых API, которые предоставляют данные в формате JSON. 
Например, некоторые популярные открытые API, которые предоставляют JSON-данные:

1. JSONPlaceholder: 
Это фейковый онлайн-сервис для тестирования и прототипирования, который предоставляет тестовые данные в формате JSON.

    URL: https://jsonplaceholder.typicode.com

2. OpenWeatherMap API: Предоставляет информацию о погоде в формате JSON.

    URL: https://openweathermap.org/api

3. GitHub API: Предоставляет данные о репозиториях, пользователях и других аспектах GitHub.

    URL: https://developer.github.com/v3

4. REST Countries API: Предоставляет информацию о странах в формате JSON.

    URL: https://restcountries.com

Просто выберите API, которое соответствует вашим интересам, и используйте его URL для загрузки JSON-данных.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import json

async def download_and_save_json(url, filename):
    try:
        async with aiohttp.ClientSession() as session, session.get(url) as response:
            response.raise_for_status()
            json_data = await response.json()

            with open(filename, 'w') as file:
                json.dump(json_data, file, indent=2)

            print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")

async def main():
    # Укажите URL и имя файла, куда сохранить JSON
    json_url = "https://jsonplaceholder.typicode.com/users/1"
    output_filename = "output_async.json"

    # Запуск асинхронной программы
    await download_and_save_json(json_url, output_filename)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Импорт библиотек:
'''
import aiohttp
import asyncio
import json
'''
Импортируются библиотеки для асинхронных HTTP-запросов (aiohttp), 
работы с асинхронными задачами (asyncio), и работы с JSON-данными (json).
'''
'''
Определение асинхронной функции download_and_save_json:
'''
async def download_and_save_json(url, filename):
''''
Эта функция асинхронная и предназначена для загрузки JSON-данных по заданному URL и сохранения их в файл.
''''
'''
Блок try-except внутри функции:
'''
try:
    async with aiohttp.ClientSession() as session, session.get(url) as response:
        response.raise_for_status()
        json_data = await response.json()

        with open(filename, 'w') as file:
            json.dump(json_data, file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

except aiohttp.ClientError as e:
    print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")
'''
Создается асинхронная сессия и выполняется асинхронный GET-запрос с использованием контекстного менеджера async with.
Проверяется наличие ошибок при запросе с response.raise_for_status().
Получаются JSON-данные из ответа с использованием await response.json().
Сохраняются JSON-данные в файл с использованием json.dump.
Обрабатываются и выводятся ошибки с использованием aiohttp.ClientError.
'''
'''
Определение асинхронной функции main:
'''
async def main():
''''
Эта функция предназначена для запуска основной асинхронной программы.
''''
'''
Вызов функции download_and_save_json:
'''
await download_and_save_json(json_url, output_filename)
'''
Вызывается ранее определенная функция download_and_save_json с указанным URL и именем файла.
'''
'''
Запуск асинхронной программы:
'''
if __name__ == "__main__":
    asyncio.run(main())
'''
Запуск основной асинхронной программы с использованием asyncio.run(main()).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Оба варианта кода выполняют асинхронное скачивание JSON-данных и сохранение их в файл.
Отличие в том, как они используют асинхронные контекстные менеджеры для удобства и читаемости кода.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Дата выполнения Домашней-Работы: 26-27 - ЯНВАРЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №27: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1
а) Загрузите массив json – объектов с сайта jsonplaceholder, используя библиотеку aiohttp.
б) Сохраните циклом каждый в отдельный файл, в одну новую папку.

'''
'''
Урок от 26.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: (а и б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import asyncio
import json
import os

async def download_and_save_json(session, url, filename):
    try:
        # Отправляем GET-запрос по указанному URL
        async with session.get(url) as response:
            response.raise_for_status()  # Проверка наличия ошибок при запросе
            json_data = await response.json()  # Получаем JSON-данные из ответа

            # Сохраняем JSON-объект в файл
            with open(filename, 'w') as file:
                json.dump(json_data, file, indent=2)

            print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")

async def download_and_save_all_jsons():
    # Указываем базовый URL для загрузки JSON-объектов
    json_url_base = "https://jsonplaceholder.typicode.com/posts/"
    output_folder = "json_files"  # Название папки для сохранения файлов

    # Создаем папку, если ее нет
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Создаем сессию aiohttp
    async with aiohttp.ClientSession() as session:
        # Запускаем цикл для загрузки и сохранения каждого JSON-объекта
        for post_id in range(1, 51):  # Пример: загрузим 50 JSON-объектов
            json_url = f"{json_url_base}{post_id}"
            output_filename = os.path.join(output_folder, f"post_{post_id}.json")

            # Вызываем функцию загрузки и сохранения JSON-объекта
            await download_and_save_json(session, json_url, output_filename)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(download_and_save_all_jsons())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Импорт библиотек:
'''
import aiohttp
import asyncio
import json
import os
'''
Импорт необходимых библиотек: aiohttp для асинхронных HTTP-запросов, asyncio для работы с
асинхронными задачами, json для работы с JSON-данными и os для работы с операционной системой.
'''
'''
Шаг №2: Определение асинхронной функции download_and_save_json:
'''
async def download_and_save_json(session, url, filename):
''''
Эта функция принимает асинхронную сессию (session), URL для запроса (url) и имя файла (filename).
''''
'''
Шаг №3: Блок try-except внутри функции:
'''
try:
    async with session.get(url) as response:
        response.raise_for_status()
        json_data = await response.json()

        with open(filename, 'w') as file:
            json.dump(json_data, file, indent=2)

        print(f"JSON-объект успешно загружен и сохранен в файл {filename}")

except aiohttp.ClientError as e:
    print(f"Произошла ошибка при загрузке или сохранении JSON: {e}")
'''
Внутри try:
Отправляется асинхронный GET-запрос по указанному URL.
Проверяется наличие ошибок при запросе с использованием response.raise_for_status().
Получаются JSON-данные из ответа с использованием await response.json().
Сохраняются JSON-данные в файл.
Выводится сообщение об успешном сохранении.

В блоке except:
Обрабатываются ошибки с помощью aiohttp.ClientError.
'''
'''
Шаг №4: Определение асинхронной функции download_and_save_all_jsons:
'''
async def download_and_save_all_jsons():
''''
Эта функция предназначена для загрузки и сохранения JSON-объектов.
''''
'''
Шаг №5: Указание базового URL и создание папки:
'''
json_url_base = "https://jsonplaceholder.typicode.com/posts/"
output_folder = "json_files"

if not os.path.exists(output_folder):
    os.makedirs(output_folder)
'''
Указывается базовый URL для загрузки JSON-объектов и создается папка для сохранения файлов, если ее нет.
'''
'''
Шаг №6: Создание сессии aiohttp:
'''
async with aiohttp.ClientSession() as session:
'''
Создается асинхронная сессия aiohttp.ClientSession().
'''
'''
Шаг №7: Цикл для загрузки и сохранения JSON-объектов:
'''
for post_id in range(1, 51):
    json_url = f"{json_url_base}{post_id}"
    output_filename = os.path.join(output_folder, f"post_{post_id}.json")

    await download_and_save_json(session, json_url, output_filename)
'''
Для каждого post_id в диапазоне от 1 до 50:
Формируется URL для запроса.
Формируется имя файла для сохранения.
Вызывается асинхронная функция download_and_save_json.
'''
'''
Шаг №8: Запуск асинхронной программы:
'''
if __name__ == "__main__":
    asyncio.run(download_and_save_all_jsons())
'''
Проверяется, что скрипт был запущен как основной,
 и запускается асинхронная функция download_and_save_all_jsons с помощью asyncio.run().
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Мы также можем загружать JSON-объекты с различных открытых API, которые предоставляют данные в формате JSON. 
Например, некоторые популярные открытые API, которые предоставляют JSON-данные:

1. JSONPlaceholder: 
Это фейковый онлайн-сервис для тестирования и прототипирования, который предоставляет тестовые данные в формате JSON.

    URL: https://jsonplaceholder.typicode.com

2. OpenWeatherMap API: Предоставляет информацию о погоде в формате JSON.

    URL: https://openweathermap.org/api

3. GitHub API: Предоставляет данные о репозиториях, пользователях и других аспектах GitHub.

    URL: https://developer.github.com/v3

4. REST Countries API: Предоставляет информацию о странах в формате JSON.

    URL: https://restcountries.com

Просто выберите API, которое соответствует вашим интересам, и используйте его URL для загрузки JSON-данных.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~

# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


'''
Дата выполнения Практической-Работы: 31 - ЯНВАРЯ - 01 ФЕВРАЛЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №28: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1

а) Загрузите одиночную html страницу с сайта википедии, используя библиотеку requests, затем aiohttp.
б) Сохраните оба раза её в файл.
'''
'''
Урок от 31.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а) & б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import requests
import aiohttp
import asyncio

# Функция для загрузки HTML-страницы с использованием библиотеки requests
def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

# Асинхронная функция для загрузки HTML-страницы с использованием библиотеки aiohttp
async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Создание сессии aiohttp
        async with aiohttp.ClientSession() as session:
            # Запрос HTML-страницы по URL
            async with session.get(url) as response:
                response.raise_for_status()
                # Получение текста HTML-страницы
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                with open(filename, 'w', encoding='utf-8') as file:
                    file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

# Асинхронная функция для выполнения основной логики
async def main():
    # Укажите URL и имена файлов для сохранения HTML
    html_url = "https://ru.wikipedia.org/wiki/Отряд_731"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Вызов функций для загрузки и сохранения HTML
    download_and_save_html_with_requests(html_url, html_filename_requests)
    await download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ДАЛЕЕ ПОДРОБНО И ШАГ ЗА ШАГОМ
'''
'''
Шаг №1: Функция download_and_save_html_with_requests
Пример кода:
'''
# Функция для загрузки HTML-страницы с использованием библиотеки requests
def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с сайта
        response = requests.get(url)
        response.raise_for_status()  # Проверка наличия ошибок при запросе

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")
'''
Полное и подробное описание:

Эта функция предназначена для загрузки HTML-страницы с использованием библиотеки requests. Вот ее основные шаги:

Загрузка HTML-страницы: Функция отправляет GET-запрос к указанному URL с помощью requests.get(url).

Проверка ошибок: С помощью response.raise_for_status() происходит проверка наличия ошибок при запросе.
Если запрос был неудачным (например, код ответа не в диапазоне 200-299), вызывается
исключение requests.exceptions.RequestException.

Сохранение HTML: Если запрос был успешным, HTML-страница сохраняется в указанный
файл с помощью with open(filename, 'w', encoding='utf-8') as file.
'''
'''
Шаг №2: Функция download_and_save_html_with_aiohttp
Пример кода:
'''
# Асинхронная функция для загрузки HTML-страницы с использованием библиотеки aiohttp
async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Создание сессии aiohttp
        async with aiohttp.ClientSession() as session:
            # Запрос HTML-страницы по URL
            async with session.get(url) as response:
                response.raise_for_status()
                # Получение текста HTML-страницы
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                with open(filename, 'w', encoding='utf-8') as file:
                    file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")

    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")
'''
Полное и подробное описание:

Эта асинхронная функция использует библиотеку aiohttp для загрузки HTML-страницы. Вот шаги, которые она выполняет:

Создание сессии aiohttp: Асинхронный контекст async with aiohttp.ClientSession() as session создает сессию, 
которую мы будем использовать для отправки запросов.

Запрос HTML-страницы: С помощью async with session.get(url) as response отправляется асинхронный GET-запрос. 
Как и в предыдущей функции, здесь также используется response.raise_for_status() для проверки ошибок при запросе.

Сохранение HTML: Если запрос был успешным, текст HTML-страницы сохраняется в указанный файл, 
аналогично функции с использованием requests.
'''
'''
Шаг №3: Функция main
Пример кода:
'''
# Асинхронная функция для выполнения основной логики
async def main():
    # Укажите URL и имена файлов для сохранения HTML
    html_url = "https://ru.wikipedia.org/wiki/Отряд_731"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Вызов функций для загрузки и сохранения HTML
    download_and_save_html_with_requests(html_url, html_filename_requests)
    await download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)

# Запуск асинхронной программы
if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:

Эта асинхронная функция main представляет основную логику программы. Она выполняет следующие шаги:

Задание URL и имен файлов: Определяется URL HTML-страницы и имена файлов, в которые будут сохранены страницы,
загруженные с использованием requests и aiohttp.

Вызов функций для загрузки и сохранения HTML: 
Вызываются обе функции download_and_save_html_with_requests и download_and_save_html_with_aiohttp для загрузки
и сохранения HTML-страниц.

Запуск асинхронной программы: asyncio.run(main()) запускает асинхронную программу, выполняя основную логику.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import aiohttp
import aiofiles
import asyncio
import requests

async def download_and_save_html_with_requests(url, filename):
    try:
        response = requests.get(url)
        response.raise_for_status()
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)
        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)
                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

async def main():
    html_url = "https://ru.wikipedia.org/wiki/Mortal_Kombat"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Используйте asyncio.gather для параллельного выполнения асинхронных функций
    await asyncio.gather(
        download_and_save_html_with_requests(html_url, html_filename_requests),
        download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)
    )

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ДАЛЕЕ ПОДРОБНО И ШАГ ЗА ШАГОМ
'''
'''
Шаг №1: Асинхронные функции download_and_save_html_with_requests и download_and_save_html_with_aiohttp
Пример кода:
'''
async def download_and_save_html_with_requests(url, filename):
    try:
        response = requests.get(url)
        response.raise_for_status()
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)
        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)
                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML: {e}")
'''
Полное и подробное описание:
download_and_save_html_with_requests:

Эта асинхронная функция использует библиотеку requests для синхронной загрузки HTML-страницы.
Она отправляет GET-запрос по указанному URL с помощью requests.get(url).
Если запрос успешен, текст HTML-страницы записывается в файл с использованием стандартного модуля open.
download_and_save_html_with_aiohttp:

Эта асинхронная функция использует библиотеки aiohttp и aiofiles для асинхронной загрузки и сохранения HTML-страницы.
Создается асинхронная сессия с aiohttp.ClientSession().
Отправляется асинхронный GET-запрос, и при успешном выполнении 
текст HTML-страницы записывается в файл с помощью aiofiles.open.
'''
'''
Шаг №2: Асинхронная функция main
Пример кода:
'''
async def main():
    html_url = "https://ru.wikipedia.org/wiki/Mortal_Kombat"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"

    # Используйте asyncio.gather для параллельного выполнения асинхронных функций
    await asyncio.gather(
        download_and_save_html_with_requests(html_url, html_filename_requests),
        download_and_save_html_with_aiohttp(html_url, html_filename_aiohttp)
    )

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Эта асинхронная функция main является точкой входа программы.
Она задает URL HTML-страницы и имена файлов для сохранения результатов с использованием requests и aiohttp.
Для параллельного выполнения обеих асинхронных функций используется asyncio.gather.
'''
'''
Шаг №3: Запуск программы
Последний блок кода if __name__ == "__main__": использует asyncio.run(main()) для запуска асинхронной программы.
'''
'''
Когда программа выполняется, она асинхронно загружает HTML-страницу дважды: первый раз с использованием requests, 
второй раз с использованием aiohttp. Обе версии страницы сохраняются в соответствующие файлы.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №3
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import requests
import aiohttp
import aiofiles
import asyncio
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import os
import re

def clean_filename(filename):
    # Очистка имени файла от недопустимых символов
    return re.sub(r'[\/:*?"<>|]', '', filename)

async def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с использованием requests
        response = requests.get(url)
        response.raise_for_status()

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (requests): {e}")
        return None

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Загрузка HTML-страницы с использованием aiohttp
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
                return html_content
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (aiohttp): {e}")
        return None

async def download_and_save_images(html_content, base_url, image_folder):
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        image_tags = soup.find_all('img')

        # Проверка и создание директории, если она не существует
        if not os.path.exists(image_folder):
            os.makedirs(image_folder)

        for img_tag in image_tags:
            src = img_tag.get('src')
            if src:
                abs_url = urljoin(base_url, src)
                image_response = requests.get(abs_url)
                image_response.raise_for_status()

                filename = clean_filename(os.path.basename(urlparse(abs_url).path))
                image_filename = os.path.join(image_folder, filename)

                with open(image_filename, 'wb') as image_file:
                    image_file.write(image_response.content)

                print(f"Изображение успешно загружено и сохранено в файл {image_filename}")
    except Exception as e:
        print(f"Произошла ошибка при загрузке или сохранении изображений: {e}")

async def main():
    wikipedia_url = "https://ru.wikipedia.org/wiki/Призрак_в_доспехах"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"
    image_folder = "downloaded_images"

    html_content_requests = await download_and_save_html_with_requests(wikipedia_url, html_filename_requests)
    html_content_aiohttp = await download_and_save_html_with_aiohttp(wikipedia_url, html_filename_aiohttp)

    if html_content_requests:
        await download_and_save_images(html_content_requests, wikipedia_url, image_folder)

    if html_content_aiohttp:
        await download_and_save_images(html_content_aiohttp, wikipedia_url, image_folder)

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ДАЛЕЕ ПОДРОБНО И ШАГ ЗА ШАГОМ
'''
'''
Шаг №1: Функция clean_filename
Пример кода:
'''
def clean_filename(filename):
    # Очистка имени файла от недопустимых символов
    return re.sub(r'[\/:*?"<>|]', '', filename)
'''
Полное и подробное описание:
Эта функция принимает имя файла и удаляет из него недопустимые символы для использования в системе файлов.
'''
'''
Шаг №2: Асинхронные функции download_and_save_html_with_requests и download_and_save_html_with_aiohttp
Пример кода:
'''
async def download_and_save_html_with_requests(url, filename):
    try:
        # Загрузка HTML-страницы с использованием requests
        response = requests.get(url)
        response.raise_for_status()

        # Сохранение HTML-страницы в файл
        with open(filename, 'w', encoding='utf-8') as file:
            file.write(response.text)

        print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
        return response.text
    except requests.exceptions.RequestException as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (requests): {e}")
        return None

async def download_and_save_html_with_aiohttp(url, filename):
    try:
        # Загрузка HTML-страницы с использованием aiohttp
        async with aiohttp.ClientSession() as session:
            async with session.get(url) as response:
                response.raise_for_status()
                html_content = await response.text()

                # Сохранение HTML-страницы в файл
                async with aiofiles.open(filename, 'w', encoding='utf-8') as file:
                    await file.write(html_content)

                print(f"HTML-страница успешно загружена и сохранена в файл {filename}")
                return html_content
    except aiohttp.ClientError as e:
        print(f"Произошла ошибка при загрузке или сохранении HTML (aiohttp): {e}")
        return None
'''
Полное и подробное описание:
download_and_save_html_with_requests:

Эта асинхронная функция использует библиотеку requests для синхронной загрузки HTML-страницы.
Она отправляет GET-запрос по указанному URL с помощью requests.get(url).
Если запрос успешен, текст HTML-страницы записывается в файл с использованием стандартного модуля open.
download_and_save_html_with_aiohttp:

Эта асинхронная функция использует библиотеки aiohttp и aiofiles для асинхронной загрузки и сохранения HTML-страницы.
Создается асинхронная сессия с aiohttp.ClientSession().
Отправляется асинхронный GET-запрос, и при успешном выполнении текст HTML-страницы записывается 
в файл с помощью aiofiles.open.
'''
'''
Шаг №3: Асинхронная функция download_and_save_images
Пример кода:
'''
async def download_and_save_images(html_content, base_url, image_folder):
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        image_tags = soup.find_all('img')

        # Проверка и создание директории, если она не существует
        if not os.path.exists(image_folder):
            os.makedirs(image_folder)

        for img_tag in image_tags:
            src = img_tag.get('src')
            if src:
                abs_url = urljoin(base_url, src)
                image_response = requests.get(abs_url)
                image_response.raise_for_status()

                filename = clean_filename(os.path.basename(urlparse(abs_url).path))
                image_filename = os.path.join(image_folder, filename)

                with open(image_filename, 'wb') as image_file:
                    image_file.write(image_response.content)

                print(f"Изображение успешно загружено и сохранено в файл {image_filename}")
    except Exception as e:
        print(f"Произошла ошибка при загрузке или сохранении изображений: {e}")
'''
Полное и подробное описание:
Эта асинхронная функция принимает HTML-контент, базовый URL и директорию для сохранения изображений.
С использованием BeautifulSoup находятся все теги <img> в HTML-странице.
Проверяется существование директории для сохранения изображений, и если ее нет, она создается.
Для каждого тега <img> извлекается атрибут src и формируется абсолютный URL.
Изображение загружается с использованием requests.get, сохраняется в файл, и выводится сообщение об успешной загрузке.
'''
'''
Шаг №4: Асинхронная функция main
Пример кода:
'''
async def main():
    wikipedia_url = "https://ru.wikipedia.org/wiki/Призрак_в_доспехах"
    html_filename_requests = "html_with_requests.html"
    html_filename_aiohttp = "html_with_aiohttp.html"
    image_folder = "downloaded_images"

    html_content_requests = await download_and_save_html_with_requests(wikipedia_url, html_filename_requests)
    html_content_aiohttp = await download_and_save_html_with_aiohttp(wikipedia_url, html_filename_aiohttp)

    if html_content_requests:
        await download_and_save_images(html_content_requests, wikipedia_url, image_folder)

    if html_content_aiohttp:
        await download_and_save_images(html_content_aiohttp, wikipedia_url, image_folder)

if __name__ == "__main__":
    asyncio.run(main())
'''
Полное и подробное описание:
Асинхронная функция main является точкой входа в программу.
Задается URL страницы Википедии, имена файлов для сохранения HTML, и директория для сохранения изображений.
Сначала вызываются асинхронные функции для загрузки HTML-страницы с использованием requests и aiohttp.
Затем, если HTML-контент был успешно получен, вызываются функции для загрузки и сохранения изображений.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


'''
Дата выполнения Домашней-Работы: 31 - ЯНВАРЯ - 01 ФЕВРАЛЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №28: Сетевое программирование. Библиотеки requests и aiohttp

Выполните следующие задания:

Задание №1

а) Загрузите циклом 10 рандомных картинок с сайта используя библиотеку requests, затем aiohttp.
б) Сохраните их в разные папки циклом.
'''
'''
Урок от 31.01.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а) & б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import os
import requests
import asyncio
import aiohttp
import random
import string

# Создаем папки для сохранения изображений
os.makedirs('images_sync_lorem', exist_ok=True)
os.makedirs('images_async_lorem', exist_ok=True)

def generate_random_filename(length=10, extension='.jpg'):
    letters_and_digits = string.ascii_letters + string.digits
    random_str = ''.join(random.choice(letters_and_digits) for _ in range(length))
    return f"{random_str}{extension}"

async def download_image_async(session, url, folder, extension='.jpg'):
    async with session.get(url) as response:
        if response.status == 200:
            # Генерируем рандомное имя файла
            filename = os.path.join(folder, generate_random_filename(extension=extension))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Lorem Picsum: {filename}')

def download_image_sync(url, folder, extension='.jpg'):
    response = requests.get(url)
    if response.status_code == 200:
        # Генерируем рандомное имя файла
        filename = os.path.join(folder, generate_random_filename(extension=extension))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Lorem Picsum: {filename}')

async def main():
    # URL-адреса для загрузки с Lorem Picsum
    lorem_image_urls = [
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
    ]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_lorem') for url in lorem_image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in lorem_image_urls:
        download_image_sync(url, 'images_sync_lorem')

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Создание папок и функции генерации имени файла
'''
# Создаем папки для сохранения изображений
os.makedirs('images_sync_lorem', exist_ok=True)
os.makedirs('images_async_lorem', exist_ok=True)

def generate_random_filename(length=10, extension='.jpg'):
    letters_and_digits = string.ascii_letters + string.digits
    random_str = ''.join(random.choice(letters_and_digits) for _ in range(length))
    return f"{random_str}{extension}"
'''
Здесь создаются папки для сохранения изображений.
Функция generate_random_filename генерирует случайное имя файла заданной длины с указанным расширением.
'''
'''
Шаг №2: Асинхронные функции для загрузки изображений с использованием aiohttp
'''
async def download_image_async(session, url, folder, extension='.jpg'):
    async with session.get(url) as response:
        if response.status == 200:
            # Генерируем рандомное имя файла
            filename = os.path.join(folder, generate_random_filename(extension=extension))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Lorem Picsum: {filename}')
'''
Эта асинхронная функция использует библиотеку aiohttp для асинхронной загрузки изображения.
Она принимает сессию, URL изображения, папку для сохранения и расширение файла.
Загруженное изображение сохраняется в файл с рандомным именем в указанной папке.
'''
'''
Шаг №3: Синхронная функция для загрузки изображений с использованием requests
'''
def download_image_sync(url, folder, extension='.jpg'):
    response = requests.get(url)
    if response.status_code == 200:
        # Генерируем рандомное имя файла
        filename = os.path.join(folder, generate_random_filename(extension=extension))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Lorem Picsum: {filename}')
'''
Эта синхронная функция использует библиотеку requests для синхронной загрузки изображения.
Она принимает URL изображения, папку для сохранения и расширение файла.
Загруженное изображение также сохраняется в файл с рандомным именем в указанной папке.
'''
'''
Шаг №4: Асинхронная функция main
'''
async def main():
    # URL-адреса для загрузки с Lorem Picsum
    lorem_image_urls = [
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
        'https://picsum.photos/800/600',
    ]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_lorem') for url in lorem_image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in lorem_image_urls:
        download_image_sync(url, 'images_sync_lorem')

if __name__ == "__main__":
    asyncio.run(main())
'''
В асинхронной функции main создается список URL-адресов изображений с Lorem Picsum.
Затем используется aiohttp для асинхронной загрузки изображений в одну папку (images_async_lorem).
После этого используется requests для синхронной загрузки тех же изображений в другую папку (images_sync_lorem).
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import os
import requests
import asyncio
import aiohttp

# Замените 'свой ключ API Pixabay' на ваш ключ API Pixabay
PIXABAY_API_KEY = 'свой ключ API Pixabay'

# Создаем папки для сохранения изображений
os.makedirs('images_sync_pixabay', exist_ok=True)
os.makedirs('images_async_pixabay', exist_ok=True)

async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем имя файла из URL
            filename = os.path.join(folder, os.path.basename(url))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Pixabay: {filename}')

def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем имя файла из URL
        filename = os.path.join(folder, os.path.basename(url))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Pixabay: {filename}')

async def main():
    # Запрос на получение 10 случайных изображений с Pixabay
    pixabay_url = f'https://pixabay.com/api/?key={PIXABAY_API_KEY}&per_page=10'
    response = requests.get(pixabay_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()['hits']
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['largeImageURL'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_pixabay') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_pixabay')

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Установка ключа API Pixabay
'''
PIXABAY_API_KEY = 'свой ключ API Pixabay'
'''
Здесь вы должны заменить 'свой ключ API Pixabay' на свой собственный ключ API Pixabay.
'''
'''
Шаг №2: Создание папок для сохранения изображений
'''
# Создаем папки для сохранения изображений
os.makedirs('images_sync_pixabay', exist_ok=True)
os.makedirs('images_async_pixabay', exist_ok=True)
'''
Этот код создает две папки: images_sync_pixabay и images_async_pixabay для сохранения изображений.
Опция exist_ok=True позволяет не вызывать ошибку, если папка уже существует.
'''
'''
Шаг №3: Асинхронные функции для загрузки изображений с использованием aiohttp
'''
async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем имя файла из URL
            filename = os.path.join(folder, os.path.basename(url))
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Pixabay: {filename}')
'''
Эта асинхронная функция использует библиотеку aiohttp для асинхронной загрузки изображения.
Она принимает сессию, URL изображения и папку для сохранения.
Извлекается имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг №4: Синхронная функция для загрузки изображений с использованием requests
'''
def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем имя файла из URL
        filename = os.path.join(folder, os.path.basename(url))
        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Pixabay: {filename}')
'''
Эта синхронная функция использует библиотеку requests для синхронной загрузки изображения.
Она принимает URL изображения и папку для сохранения.
Извлекается имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг №5: Асинхронная функция main
'''
async def main():
    # Запрос на получение 10 случайных изображений с Pixabay
    pixabay_url = f'https://pixabay.com/api/?key={PIXABAY_API_KEY}&per_page=10'
    response = requests.get(pixabay_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()['hits']
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['largeImageURL'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_pixabay') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_pixabay')

if __name__ == "__main__":
    asyncio.run(main())
'''
В функции main отправляется запрос к Pixabay для получения 10 случайных изображений.
Если запрос успешен, извлекаются URL-адреса изображений.
Затем используется aiohttp для асинхронной загрузки изображений в папку images_async_pixabay.
Также используется requests для синхронной загрузки изображений в папку images_sync_pixabay.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №3
 '''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
import os
import requests
import asyncio
import aiohttp
from pathlib import Path
from urllib.parse import urlparse

# Замените 'ваш ключ API Unsplash' на ваш ключ API Unsplash
UNSPLASH_ACCESS_KEY = 'ваш ключ API Unsplash'

# Создаем папки для сохранения изображений
os.makedirs('images_sync_unsplash', exist_ok=True)
os.makedirs('images_async_unsplash', exist_ok=True)

async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем безопасное имя файла из URL с использованием urllib.parse
            filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Unsplash: {filename}')

def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем безопасное имя файла из URL с использованием urllib.parse
        filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')

        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Unsplash: {filename}')

async def main():
    # Запрос на получение 10 случайных изображений с Unsplash
    unsplash_url = f'https://api.unsplash.com/photos/random?count=10&client_id={UNSPLASH_ACCESS_KEY}'
    response = requests.get(unsplash_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['urls']['raw'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_unsplash') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_unsplash')

if __name__ == "__main__":
    asyncio.run(main())
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг 1: Установка ключа API Unsplash
'''
# Замените 'ваш ключ API Unsplash' на ваш ключ API Unsplash
UNSPLASH_ACCESS_KEY = 'ваш ключ API Unsplash'
'''
В этой части кода вы должны заменить 'ваш ключ API Unsplash' на свой собственный ключ API Unsplash.
'''
'''
Шаг 2: Создание папок для сохранения изображений
'''
# Создаем папки для сохранения изображений
os.makedirs('images_sync_unsplash', exist_ok=True)
os.makedirs('images_async_unsplash', exist_ok=True)
'''
Этот код создает две папки: images_sync_unsplash и images_async_unsplash для сохранения изображений.
Опция exist_ok=True позволяет не вызывать ошибку, если папка уже существует.
'''
'''
Шаг 3: Асинхронные функции для загрузки изображений с использованием aiohttp
'''
async def download_image_async(session, url, folder):
    async with session.get(url) as response:
        if response.status == 200:
            # Извлекаем безопасное имя файла из URL с использованием urllib.parse
            filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')
            # Сохраняем изображение
            with open(filename, 'wb') as f:
                f.write(await response.read())
            print(f'Image downloaded asynchronously from Unsplash: {filename}')
'''
Эта асинхронная функция использует библиотеку aiohttp для асинхронной загрузки изображения.
Она принимает сессию, URL изображения и папку для сохранения.
Извлекается безопасное имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг 4: Синхронная функция для загрузки изображений с использованием requests
'''
def download_image_sync(url, folder):
    response = requests.get(url)
    if response.status_code == 200:
        # Извлекаем безопасное имя файла из URL с использованием urllib.parse
        filename = os.path.join(folder, f'image_{Path(urlparse(url).path).name}.jpg')

        # Сохраняем изображение
        with open(filename, 'wb') as f:
            f.write(response.content)
        print(f'Image downloaded synchronously from Unsplash: {filename}')
'''
Эта синхронная функция использует библиотеку requests для синхронной загрузки изображения.
Она принимает URL изображения и папку для сохранения.
Извлекается безопасное имя файла из URL, и изображение сохраняется в файл.
'''
'''
Шаг 5: Асинхронная функция main
'''
async def main():
    # Запрос на получение 10 случайных изображений с Unsplash
    unsplash_url = f'https://api.unsplash.com/photos/random?count=10&client_id={UNSPLASH_ACCESS_KEY}'
    response = requests.get(unsplash_url)

    # Проверяем успешность запроса
    if response.status_code == 200:
        image_data = response.json()
    else:
        print(f"Failed to retrieve images. Status code: {response.status_code}")
        return

    # Список URL для загрузки
    image_urls = [photo['urls']['raw'] for photo in image_data]

    # Загрузка изображений с использованием aiohttp
    async with aiohttp.ClientSession() as session:
        tasks = [download_image_async(session, url, 'images_async_unsplash') for url in image_urls]
        await asyncio.gather(*tasks)

    # Загрузка изображений с использованием requests
    for url in image_urls:
        download_image_sync(url, 'images_sync_unsplash')

if __name__ == "__main__":
    asyncio.run(main())
'''
В функции main отправляется запрос к Unsplash для получения 10 случайных изображений.
Если запрос успешен, извлекаются URL-адреса изображений.
Затем используется aiohttp для асинхронной загрузки изображений в папку images_async_unsplash.
Также используется requests для синхронной загрузки изображений в папку images_sync_unsplash.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~

'''
Дата выполнения Практической-Работы: 02 - ФЕВРАЛЯ - 03 ФЕВРАЛЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №29: Программы с интерфейсом (GUI) - tkinter и pyqt6

Выполните следующие задания:

Задание №1

а) Сделать набросок дизайна программы в figma / paint для программы, которая читает все excel-файлы в 
папке и выводит на экран общее количество строк.

б) Разработать эту программу на библиотеке tkinter.
'''
'''
Урок от 02.02.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а) & б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_lines += sheet.max_row
            except Exception as e:
                print(f"Считывание ошибки {filename}: {e}")

    return total_lines

def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        total_lines = count_lines_in_excel_files(folder_path)
        result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")

# GUI setup
window = tk.Tk()
window.title("Счетчик строк Excel")

# Установим размер окна 500x300
window.geometry('500x300')

browse_button = tk.Button(window, text="Просмотр папки", command=browse_folder)
browse_button.pack(pady=20)

result_label = tk.Label(window, text="")
result_label.pack()

window.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Импорт необходимых библиотек.
'''
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
'''
Полное и подробное описание:

os: Библиотека для взаимодействия с операционной системой, используется для работы с файловой системой.
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
filedialog: Модуль в Tkinter для диалогового взаимодействия с файлами и папками.
openpyxl: Библиотека для работы с файлами формата Excel (.xlsx).
'''
'''
Шаг №2: Определение функции count_lines_in_excel_files.
'''
def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_lines += sheet.max_row
            except Exception as e:
                print(f"Считывание ошибки {filename}: {e}")

    return total_lines
'''
Полное и подробное описание:

count_lines_in_excel_files: Функция, которая принимает путь к папке и возвращает общее количество строк во всех 
файлах Excel в данной папке.
total_lines: Переменная для хранения общего количества строк.
for filename in os.listdir(folder_path): Цикл проходит по всем файлам в указанной папке.
if filename.endswith(".xlsx"):: Проверка, что файл имеет расширение .xlsx.
file_path = os.path.join(folder_path, filename): Формирование полного пути к файлу.
try ... except Exception as e:: Обработка исключений при загрузке файла Excel. Если возникает ошибка, 
выводится сообщение об ошибке.
'''
'''
Шаг №3: Определение функции browse_folder.
'''
def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        total_lines = count_lines_in_excel_files(folder_path)
        result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")
'''
Полное и подробное описание:

browse_folder: Функция, вызываемая при нажатии кнопки "Просмотр папки".
Открывает диалог выбора папки и вызывает count_lines_in_excel_files для подсчета строк в файлах Excel.
filedialog.askdirectory(): Пользователь выбирает папку через диалоговое окно.
if folder_path:: Проверка, что пользователь выбрал папку.
total_lines = count_lines_in_excel_files(folder_path): Вызов функции count_lines_in_excel_files для подсчета строк.
result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}"): Обновление метки
на GUI с общим количеством строк.
'''
'''
Шаг №4: Настройка графического интерфейса (GUI).
'''
window = tk.Tk()
window.title("Счетчик строк Excel")

window.geometry('500x300')

browse_button = tk.Button(window, text="Просмотр папки", command=browse_folder)
browse_button.pack(pady=20)

result_label = tk.Label(window, text="")
result_label.pack()

window.mainloop()
'''
Полное и подробное описание:

tk.Tk(): Создание главного окна приложения.
window.title("Счетчик строк Excel"): Установка заголовка окна.
window.geometry('500x300'): Установка размеров окна.
tk.Button(window, text="Просмотр папки", command=browse_folder): Создание
кнопки "Просмотр папки", которая при нажатии вызывает функцию browse_folder.
tk.Label(window, text=""): Создание метки для вывода результата.
window.mainloop(): Запуск главного цикла обработки событий Tkinter.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook

result_label = None  # Объявляем переменную здесь

def count_lines_in_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        total_lines = sum(sheet.max_row for sheet in workbook)
        return total_lines
    except Exception as e:
        return f"Считывание ошибки {file_path}: {e}"

def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            total_lines += count_lines_in_excel_file(file_path)

    return total_lines

def browse_folder():
    global result_label  # Используем глобальную переменную
    try:
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Файлы Excel", "*.xlsx")])
        if file_path:
            total_lines = count_lines_in_excel_file(file_path)
            result_label.config(text=f"Общее количество строк в файле Excel: {total_lines}")
    except Exception as e:
        result_label.config(text=f"Ошибка: {e}")

def clear_result():
    global result_label  # Используем глобальную переменную
    result_label.config(text="")

# GUI setup
def create_gui():
    global result_label  # Используем глобальную переменную
    window = tk.Tk()
    window.title("Счетчик строк Excel")
    window.geometry('500x300')

    browse_button = tk.Button(window, text="Просмотр файла Excel", command=browse_folder)
    browse_button.pack(pady=10)

    clear_button = tk.Button(window, text="Очистить результат", command=clear_result)
    clear_button.pack()

    result_label = tk.Label(window, text="")
    result_label.pack()

    window.mainloop()

if __name__ == "__main__":
    create_gui()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
1. Импорт библиотек и объявление глобальных переменных:
'''
import os
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook

result_label = None  # Объявляем переменную здесь
'''
Описание:
os: Предоставляет функции для работы с операционной системой, такие как чтение содержимого папки.
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
filedialog: Подмодуль tkinter для работы с диалоговыми окнами выбора файлов и папок.
ttk: Подмодуль tkinter, предоставляющий расширенные виджеты.
openpyxl: Библиотека для работы с файлами Excel формата xlsx.
result_label: Глобальная переменная, которая будет использоваться для отображения результата в GUI.
'''
'''
2. Функция count_lines_in_excel_file:
'''
def count_lines_in_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        total_lines = sum(sheet.max_row for sheet in workbook)
        return total_lines
    except Exception as e:
        return f"Считывание ошибки {file_path}: {e}"
'''
Описание:
count_lines_in_excel_file: Функция, которая считает общее количество строк в файле Excel.
Открывает файл, считывает количество строк в каждом листе и возвращает сумму.
Обрабатывает возможные исключения, выводя сообщение об ошибке.
'''
'''
3. Функция count_lines_in_excel_files:
'''
def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            total_lines += count_lines_in_excel_file(file_path)

    return total_lines
'''
Описание:
count_lines_in_excel_files: Функция, которая проходит через все файлы с 
расширением ".xlsx" в указанной папке и считает общее количество строк.
Использует предыдущую функцию count_lines_in_excel_file.
'''
'''
4. Функция browse_folder:
'''
def browse_folder():
    global result_label  # Используем глобальную переменную
    try:
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Файлы Excel", "*.xlsx")])
        if file_path:
            total_lines = count_lines_in_excel_file(file_path)
            result_label.config(text=f"Общее количество строк в файле Excel: {total_lines}")
    except Exception as e:
        result_label.config(text=f"Ошибка: {e}")
'''
Описание:
browse_folder: Функция, вызываемая при нажатии кнопки "Просмотр файла Excel".
Открывает диалоговое окно выбора файла Excel, считает количество строк и обновляет result_label в GUI.
'''
'''
Функция clear_result:
'''
def clear_result():
    global result_label  # Используем глобальную переменную
    result_label.config(text="")
'''
Описание:
clear_result: Функция, вызываемая при нажатии кнопки "Очистить результат".
Очищает result_label в GUI.
'''
'''
6. Настройка GUI в функции create_gui:
'''
def create_gui():
    global result_label  # Используем глобальную переменную
    window = tk.Tk()
    window.title("Счетчик строк Excel")
    window.geometry('500x300')

    browse_button = tk.Button(window, text="Просмотр файла Excel", command=browse_folder)
    browse_button.pack(pady=10)

    clear_button = tk.Button(window, text="Очистить результат", command=clear_result)
    clear_button.pack()

    result_label = tk.Label(window, text="")
    result_label.pack()

    window.mainloop()
'''
Описание:
create_gui: Функция для создания основного GUI.
Создает окно, добавляет кнопки "Просмотр файла Excel" и "Очистить результат", а также result_label.
'''
'''
7. Запуск GUI, если код выполняется как самостоятельный скрипт:
'''
if __name__ == "__main__":
    create_gui()
'''
Описание:
Проверяет, запущен ли скрипт непосредственно (а не импортирован как модуль).
Если да, вызывает функцию create_gui для создания и запуска GUI.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


'''
Дата выполнения Домашней-Работы: 02 - ФЕВРАЛЯ - 03 ФЕВРАЛЯ 2024 года.
'''
'''
Домашняя работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Домашняя работа №29: Программы с интерфейсом (GUI) tkinter и pyqt6

Выполните следующие задания:

Задание №1

а) Сделать набросок дизайна программы в figma / paint для программы, которая делает запрос на сайт jsonplaceholder с определённым id.
б) Разработать эту программу на библиотеке tkinter.
в) Реализовать сохранение полученного объекта в папку.
'''
'''
Урок от 02.02.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а), б) & в)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ только один (Возможно ОДИН)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
!!! 
(На первых скриншотах есть наименования в "" (ковычках) на ENG языке - так как писал код в "черновике", оставил как
есть - НО ДЛЯ "ЧИСТОВИКА" - перевел все на - РУССКИЙ ЯЗЫК - для удобства и красоты в интерфейсе.

* черновик - 02.02.2024_ONLY_TEST_CODE+.py
* ЧИСТОВИК - 2024_PYTHON_LESSON_№29_HW_02.02.2024_03.02.2024.py
'''
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import tkinter as tk
from tkinter import messagebox
import requests
import json

def fetch_data(user_id):
    try:
        # Замените URL на свой, если требуется
        url = f"https://jsonplaceholder.typicode.com/users/{user_id}"
        response = requests.get(url)
        response.raise_for_status()  # Проверяем наличие ошибок при запросе

        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Ошибка", f"Не удалось получить данные: {e}")
        return None

def save_data(data, file_path):
    try:
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=2)
        messagebox.showinfo("Успех", f"Данные, сохраненные в {file_path}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {e}")

def on_fetch_button_click():
    user_id = entry_user_id.get()
    if not user_id.isdigit():
        messagebox.showwarning("Предупреждение", "Идентификатор пользователя должен быть числом")
        return

    user_id = int(user_id)
    data = fetch_data(user_id)
    if data:
        save_data(data, "user_data.json")

# Настройка графического интерфейса пользователя
window = tk.Tk()
window.title("Средство извлечения данных JSONPlaceholder")
window.geometry('650x350')

label_user_id = tk.Label(window, text="Введите идентификатор пользователя:")
label_user_id.pack(pady=10)

entry_user_id = tk.Entry(window)
entry_user_id.pack(pady=10)

fetch_button = tk.Button(window, text="Извлечь данные", command=on_fetch_button_click)
fetch_button.pack(pady=10)

window.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Данный код представляет собой простое графическое приложение,
написанное с использованием библиотеки tkinter,
которое позволяет пользователю ввести идентификатор пользователя,
после чего отправляет запрос к JSONPlaceholder API для получения данных о пользователе с указанным идентификатором.
Полученные данные сохраняются в файл JSON.
Шаги по коду:
'''
'''
1. Импорт библиотек:
'''
import tkinter as tk
from tkinter import messagebox
import requests
import json
'''
Описание:
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
messagebox: Подмодуль tkinter для создания диалоговых окон сообщений.
requests: Библиотека для выполнения HTTP-запросов.
json: Библиотека для работы с форматом JSON.
'''
'''
2. Функция fetch_data:
'''
def fetch_data(user_id):
    try:
        url = f"https://jsonplaceholder.typicode.com/users/{user_id}"
        response = requests.get(url)
        response.raise_for_status()

        data = response.json()
        return data
    except requests.exceptions.RequestException as e:
        messagebox.showerror("Ошибка", f"Не удалось получить данные: {e}")
        return None
'''
Описание:
fetch_data: Функция, отправляющая запрос к JSONPlaceholder API для получения данных о пользователе.
Возвращает данные в формате JSON, если запрос успешен, или None, если возникла ошибка.
Если возникла ошибка, выводится диалоговое окно с сообщением об ошибке.
'''
'''
3. Функция save_data:
'''
def save_data(data, file_path):
    try:
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=2)
        messagebox.showinfo("Успех", f"Данные, сохраненные в {file_path}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {e}")
'''
Описание:
save_data: Функция, сохраняющая данные в файл JSON.
В случае успеха выводится сообщение об успешном сохранении, иначе - сообщение об ошибке.
'''
'''
4. Функция on_fetch_button_click:
'''
def on_fetch_button_click():
    user_id = entry_user_id.get()
    if not user_id.isdigit():
        messagebox.showwarning("Предупреждение", "Идентификатор пользователя должен быть числом")
        return

    user_id = int(user_id)
    data = fetch_data(user_id)
    if data:
        save_data(data, "user_data.json")
'''
Описание:
on_fetch_button_click: Функция, вызываемая при клике на кнопку "Извлечь данные".
Получает идентификатор пользователя из виджета ввода (entry_user_id).
Проверяет, что идентификатор является числом, и выводит предупреждение, если нет.
После получения данных о пользователе вызывает save_data для сохранения в файл "user_data.json".
'''
'''
5. Настройка графического интерфейса пользователя:
'''
window = tk.Tk()
window.title("Средство извлечения данных JSONPlaceholder")
window.geometry('650x350')

label_user_id = tk.Label(window, text="Введите идентификатор пользователя:")
label_user_id.pack(pady=10)

entry_user_id = tk.Entry(window)
entry_user_id.pack(pady=10)

fetch_button = tk.Button(window, text="Извлечь данные", command=on_fetch_button_click)
fetch_button.pack(pady=10)

window.mainloop()
'''
Описание:
Создает основное окно (Tk).
Добавляет виджеты: метку (Label), поле ввода (Entry), кнопку (Button).
Назначает функцию on_fetch_button_click обработчиком события для кнопки.
Запускает главный цикл событий (mainloop), который ожидает действий пользователя.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


< !--  ########################################################################################## -->
< !-- HTML  # 1 Recap + Вводный ПЕРВЫЙ УРОК - 08.01.2024 -->

< !-- < html >
< head >

< meta
charset = "UTF-8" / >
< title > Hello
HTML < / title >
< / head >
< body >

hello
world
Привет
Мир

< img
src = "https://pngicon.ru/file/uploads/serdce.png"
height = "150px"
width = "150px" / >
< h1 > Вау - я
программист < / h1 >
< / body >
< / html > -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !-- 2024
_HTML_LESSON_№2
_Структура_HTML.Теги_Атрибуты - 15.01
.2024 -->

< !--

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 2: Структура
HTML.Теги.Атрибуты.Правила
языка
разметки
-->

< !DOCTYPE
html >
< html
lang = "ru" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Страница
с
текстом
и
заголовками < / title >
< / head >
< body >

< !-- 1.
Создать
страницу
с
текстом
по
примеру
на
картинке. -->

< h2 >
Стихотворение
< / h2 >
< p >

Мириады
маленьких
дел < br >
Пьют
по
капле
гаснущий
день, < br >
А < i > дела
большие < / i > сушит
жажда. < br >
Оставляя
все
на «потом», < br >
Прозреваем
задним
числом, < br >
Только
год < b > не
повторится < / b > дважды. < br >
< br >
И.Тальков

< / p >

< !-- 2.
Повторите
по
данному
образцу: -->
< h1 > Это
заголовок < / h1 >
< h2 > Это
заголовок < / h2 >
< h3 > Это
заголовок < / h3 >
< h4 > Это
заголовок < / h4 >

< p > Это < b > абзац < / b >.< / p >
< p > Это
еще < i > абзац < / i >.< / p >

< h1 > < i > Это
заголовок
h1 < / i > < / h1 >

< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

03
_Введение
в
CSS.Форматирование
текста
при
помощи
CSS - 22.01
.2024
Классная
работа

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< link
rel = "preconnect"
href = "https://fonts.googleapis.com" >
< link
rel = "preconnect"
href = "https://fonts.gstatic.com"
crossorigin >
< link
href = "https://fonts.googleapis.com/css2?family=Rubik+Burned&display=swap"
rel = "stylesheet" >

< meta
charset = "UTF-8" >
< title > Lesson
2 < / title >
< script >
/ *легкий
init
js * /
< / script >
< meta
description = "Краткое описание содержимого/контента страницы" / >
< meta
keywords = "купить, по акции, новый" / >
< style >
html, body
{font - size: 10px;
font - family: "Rubik Burned", system - ui;}
h1
{font - size: 4rem}
h2
{font - size: 3rem}
h3
{font - size: 2rem}
h4
{font - size: 1rem}
< / style >
< link
rel = "stylesheet"
href = "./style.css" / >
< script
src = "https://cdn.tailwindcss.com" > < / script >
< / head >
< body >
< div
style = "display: flex" >
< img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
< img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
< img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
< / div >
< div >
< h1


class ="heading-of-section text-red text-krasnii" > Красный < / h1 >

< p


class ="text-orange" > Оранжевый < / p >

< div >
< p


class ="button default" > Нажми меня < / p >

< div


class ="button alert" > Нажми меня < / div >

< button


class ="button alert" alt="Для оформления заказа нажмите эту кнопку" > Нажми меня < / button >

< / div >
< div >
< div


class ="display-inline rounded-md shadow-md padding-md bg-gray" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-red" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-blue" > Нажми меня < / div >

< / div >
< / div >
< b > Жирный
текст < / b >
< h1
style = "text-decoration: line-through" > Панды, наконец - то
полетят
на
Марс < / h1 >
< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3


class ="heading" > Hello world < / h3 >

< h4


class ="text-red" > Hello world < / h4 >

< h5 > Hello
world < / h5 >
< h6


class ="bg-blue" > Hello world < / h6 >

< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3 > Hello
world < / h3 >
< h4
id = "heading-h4-1" > Hello
world < / h4 >
< h4
id = "heading-h4-2" > Hello
world < / h4 >
< h4 > Hello
world < / h4 >
< h3


class ="text-lg text-rose-500 bg-gray-100 rounded-md" > Hello world < / h3 >

< h3 > Hello
world < / h3 >

< p >
paragraph < b > lorem < / b >
< br / >
< br / >
< br / >
< br / >
< br / >
ipsum
set < i > < b > ВАЖНО! < / b > < / i > dolor
< br / >
ipsum
set < strong > < em > Кликни & rarr; < / em > < / strong > dolor
< br / >
2 * 2
< br / >
2
x2
< br / >
2 & cross;
2
< / p >
< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
22.01
.2024 - Домашняя
и
практическая - работы.
< !--  ########################################################################################## -->
Практическая
работа

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №3
Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Создать
HTML - страницу “Vehicle”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

2.
Создать
HTML - страницу “Lorem
Ipsum”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

3.
Создать
HTML - страницу “Математические
формулы”

Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.
Используйте
здесь
теги
h1 - h6, span, p, sup, sub
и
спецсимволы.
< !--  ########################################################################################## -->

Шрифты
находятся
в
папке
Материалы
к
занятию
в
архиве
Fonts.

1.
Пример
подключение
шрифтов
через
ссылку
2.
Пример
подключения
через
локальные
файлы
шрифтов
< !--  ########################################################################################## -->

№1

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Двигатель < / title >
< style >
body
{
    font - family: 'Tahoma', serif;
line - height: 1.6;
margin: 18
px;
}

h1
{
    color:  # 00c264;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align: justify;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.smaller - text
{
    font - size: 58 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
< / style >

< / head >
< body >
< h1


class ="larger-font2" >


Двигатель
< / h1 >
< !-- Текст
о
двигателе
с
сайта
https: // ru.wikipedia.org / wiki / Двигатель -->

< p


class ="larger-font" >

< b > Дви́гатель < / b > — устройство, преобразующее
какой - либо
вид < a
href = "https://ru.wikipedia.org/wiki/Энергия"


class ="no-underline" > энергии < / a >


в < a
href = "https://ru.wikipedia.org/wiki/Механическая_энергия"


class ="no-underline" > механическую < / a > работу.


Термин < b > мотор < / b > < a
href = "https://ru.wikipedia.org/wiki/Заимствование"


class ="no-underline" > заимствован < / a > в


первой
половине
XIX
века
из
немецкого
языка < b


class ="smaller-text colored-letter" >[1] < / b >


(нем.Motor — «двигатель», от лат.mōtor — «приводящий в движение»)
и
преимущественно
им
называют
электрические
двигатели
и
двигатели
внутреннего
сгорания < b


class ="smaller-text colored-letter" >


[2] < / b >.
< / p >
< p


class ="larger-font" >


Двигатели
подразделяют
на
первичные
и
вторичные.
К
первичным
относят
непосредственно
преобразующие
природные
энергетические
ресурсы
в
механическую
работу,
а
ко
вторичным — преобразующие
энергию, выработанную
или
накопленную
другими
источниками < b


class ="smaller-text colored-letter" >


[3] < / b >.
< / p >
< hr


class ="subtle-line" >

< p >
< h4 > Примечания < / h4 >
< h5 > [1] < / h5 > Шанский
Н.М., Боброва
Т.А.Кот // Школьный
этимологический
словарь
русского
языка.
Происхождение
слов. — 7 - е
изд., стереотип.. — М.: Дрофа, 2004. — 398, [2]
с.
< h5 > [2] < / h5 > Крысин
Л.П.Мотор // Толковый
словарь
иноязычных
слов. — М.: Эксмо, 2008. — 944
с. —
(Библиотека словарей).
< h5 > [3] < / h5 > Definition
of
motor | Dictionary.com(англ.).www.dictionary.com.
Дата
обращения: 27
января
2022.
Архивировано
27
января
2022
года.
< / p >
< / body >
< / html >
< !--  ########################################################################################## -->

№2

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Lorem
Ipsum < / title >
< style >

h1
{
    color:  # 5d0080;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align0: justify;
}
.left - aligned1
{
    text - align: left;
}

.right - aligned2
{
    text - align: right;
}

.custom - font
{
    font - family: "Courier New", monospace;
}
.custom - font2
{
    font - family: "mv boli", monospace;
}
.custom - font3
{
    font - family: "segoe print", monospace;
}
.custom - font4
{
    font - family: "LUCIDA SANS CONSOLE", monospace;
}
.custom - font5
{
    font - family: "GABRIOLA", monospace;
}
.custom - font6
{
    font - family: "tempus sans itc", monospace;
}
.bold - text
{
    font - weight: bold;
}

.italic - text
{
    font - style: italic;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.larger - font3
{
    font - size: 50px;
}
.centered - text
{
    text - align: center;
}

.centered - container
{
    margin: 0 auto;
width: 50 %; / *Измените
на
желаемую
ширину * /
}

.smaller - text
{
    font - size: 95 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text2
{
    font - size: 80 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text3
{
    font - size: 115 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text4
{
    font - size: 15 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.subtle - line2
{
    border: 1px solid  # f2f2f2; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
.custom - color1
{
    color:  # 001fb8; /* Замените на желаемый цвет */
}
.custom - color2
{
    color:  # 6e6e6e; /* Замените на желаемый цвет */
}
.page - container
{
    margin: 20px; / *Замените
на
желаемое
значение
отступа * /
}
< / style >
< / head >
< body >
< h1


class ="centered-text larger-font3 custom-font4" > Lorem Ipsum < / h1 >

< div


class ="right-aligned2 custom-color1 custom-font3 italic-text smaller-text" >

< p > "Neque porro quisquam est qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit...". < / p >
< / div >
< div


class ="right-aligned2 custom-font2 custom-color2 italic-text smaller-text2 larger-font" >

< p > "Нет никого, кто любил бы боль саму по себе, кто искал бы её и кто хотел бы иметь её
просто
потому, что
это
боль..
".</p>
< / div >
< hr


class ="subtle-line2" >

< !-- Пример
текста
Lorem
Ipsum -->
< div


class ="page-container" >

< p


class ="larger-font custom-font" > < b >


Классический
текст
Lorem
Ipsum, используемый
с
XVI
века
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>L</b>orem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod tempor incididunt
ut
labore
et
dolore
magna
aliqua.Ut
enim
ad
minim
veniam, quis
nostrud
exercitation
ullamco
laboris
nisi
ut
aliquip
ex
ea
commodo
consequat.Duis
aute
irure
dolor in reprehenderit in voluptate
velit
esse
cillum
dolore
eu
fugiat
nulla
pariatur.Excepteur
sint
occaecat
cupidatat
non
proident, sunt in culpa
qui
officia
deserunt
mollit
anim
id
est
laborum.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.32
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>S</b>ed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem
aperiam, eaque
ipsa
quae
ab
illo
inventore
veritatis
et
quasi
architecto
beatae
vitae
dicta
sunt
explicabo.
Nemo
enim
ipsam
voluptatem
quia
voluptas
sit
aspernatur
aut
odit
aut
fugit, sed
quia
consequuntur
magni
dolores
eos
qui
ratione
voluptatem
sequi
nesciunt.Neque
porro
quisquam
est, qui
dolorem
ipsum
quia
dolor
sit
amet,
consectetur, adipisci
velit, sed
quia
non
numquam
eius
modi
tempora
incidunt
ut
labore
et
dolore
magnam
aliquam
quaerat
voluptatem.Ut
enim
ad
minima
veniam, quis
nostrum
exercitationem
ullam
corporis
suscipit
laboriosam, nisi
ut
aliquid
ex
ea
commodi
consequatur? Quis
autem
vel
eum
iure
reprehenderit
qui in ea
voluptate
velit
esse
quam
nihil
molestiae
consequatur, vel
illum
qui
dolorem
eum
fugiat
quo
voluptas
nulla
pariatur?"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>B</b>ut I must explain to you how all this mistaken idea of denouncing pleasure and praising pain was born and
I
will
give
you
a
complete
account
of
the
system, and expound
the
actual
teachings
of
the
great
explorer
of
the
truth, the
master - builder
of
human
happiness.No
one
rejects, dislikes, or avoids
pleasure
itself, because
it is
pleasure, but
because
those
who
do
not know
how
to
pursue
pleasure
rationally
encounter
consequences
that
are
extremely
painful.Nor
again is there
anyone
who
loves or pursues or desires
to
obtain
pain
of
itself, because
it
is pain, but
because
occasionally
circumstances
occur in which
toil and pain
can
procure
him
some
great
pleasure.
To
take
a
trivial
example, which
of
us
ever
undertakes
laborious
physical
exercise, except to
obtain
some
advantage
from it? But
who
has
any
right
to
find
fault
with a man who chooses to enjoy a pleasure that has no annoying
consequences, or one
who
avoids
a
pain
that
produces
no
resultant
pleasure?"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.33
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>A</b>t vero eos et accusamus et iusto odio dignissimos ducimus qui blanditiis praesentium voluptatum deleniti
atque
corrupti
quos
dolores
et
quas
molestias
excepturi
sint
occaecati
cupiditate
non
provident, similique
sunt
in culpa
qui
officia
deserunt
mollitia
animi, id
est
laborum
et
dolorum
fuga.Et
harum
quidem
rerum
facilis
est
et
expedita
distinctio.Nam
libero
tempore, cum
soluta
nobis
est
eligendi
optio
cumque
nihil
impedit
quo
minus
id
quod
maxime
placeat
facere
possimus, omnis
voluptas
assumenda
est, omnis
dolor
repellendus.Temporibus
autem
quibusdam
et
aut
officiis
debitis
aut
rerum
necessitatibus
saepe
eveniet
ut
et
voluptates
repudiandae
sint
et
molestiae
non
recusandae.Itaque
earum
rerum
hic
tenetur
a
sapiente
delectus, ut
aut
reiciendis
voluptatibus
maiores
alias
consequatur
aut
perferendis
doloribus
asperiores
repellat.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>O</b>n the other hand, we denounce with righteous indignation and dislike men who are so beguiled and demoralized
by
the
charms
of
pleasure
of
the
moment, so
blinded
by
desire, that
they
cannot
foresee
the
pain and trouble
that
are
bound
to
ensue; and equal
blame
belongs
to
those
who
fail in their
duty
through
weakness
of
will, which is the
same as saying
through
shrinking
from toil and pain.These
cases
are
perfectly
simple and easy
to
distinguish.
In
a
free
hour, when
our
power
of
choice is untrammelled and when
nothing
prevents
our
being
able
to
do
what
we
like
best, every
pleasure is to
be
welcomed and every
pain
avoided.But in certain
circumstances and owing
to
the
claims
of
duty or the
obligations
of
business
it
will
frequently
occur
that
pleasures
have
to
be
repudiated and
annoyances
accepted.The
wise
man
therefore
always
holds in these
matters
to
this
principle
of
selection: he
rejects
pleasures
to
secure
other
greater
pleasures, or else he
endures
pains
to
avoid
worse
pains.
"
< / p >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

№3

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Математические
формулы < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'VERDANA', monospace; / *Устанавливаем
шрифт
для
всего
документа * /
line - height: 1.0; / *Задаем
высоту
строки
текста * /
margin: 70
px; / *Задаем
отступы
для
всего
документа * /
}

/ *Стили
для
заголовка
первого
уровня * /
h1
{
    color:  # 000000; /* Цвет текста */
        text - align: center; / *Выравнивание
текста
по
центру * /
font - size: 50
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
}

/ *Стили
для
заголовка
второго
уровня * /
h2
{
    color:  # 22009e; /* Цвет текста */
        font - size: 35
px; / *Размер
шрифта * /
margin - top: 20
px; / *Верхний
отступ * /
}

/ *Стили
для
подзаголовка * /
h3
{
    color:  # 00ff49; /* Цвет текста */
        font - size: 25
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
font - style: italic; / *Курсив * /
                         text - align: left; / *Выравнивание
текста
влево * /
background - color:  # a300a3b8; /* Цвет подложки */
padding: 10
px; / *Внутренний
отступ * /
border - radius: 45
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 320
px; / *Максимальная
ширина
подложки * /
width: 25 %; / *Ширина
подложки
равна
20 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: left; / *Выравнивание
по
левой
стороне * /
}

/ *Стили
для
обычного
текста
и
вложенных
элементов * /
p, span
{
    text - align: justify; / *Выравнивание
текста
по
ширине
с
обеих
сторон * /
}

/ *Стили
для
верхних
и
нижних
индексов * /
sup, sub
{
    font - size: 80 %; / *Размер
шрифта
для
верхних
и
нижних
индексов * /
}

/ *Отступ
для
блока
с
описанием
формулы * /
.formula - description
{
    margin - left: 20px; / *Отступ
слева * /
}

/ *Стили
для
контейнера
формулы
и
самой
формулы * /
.formula - container,.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
background - color:  # ff00008f; /* Цвет подложки */
padding: 2
px; / *Внутренний
отступ * /
border - radius: 55
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
15
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 670
px; / *Максимальная
ширина
подложки * /
width: 50 %; / *Ширина
подложки
равна
50 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: auto; / *Выравнивание
по
центру * /
}

/ *Стили
для
самой
формулы * /
.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
color:  # 0700ff; /* Цвет формулы */
font - size: 25
px; / *Размер
шрифта
формулы * /
font - style: italic; / *Курсив * /
                         font - weight: bold; / *Жирный
шрифт
формулы * /
}

/ *Стили
для
раздела
с
примерами * /
.example - section
{
    margin - top: 20px; / *Верхний
отступ
раздела
с
примерами * /
}

/ *Стили
для
примеров * /
.example
{
    font - size: 20px; / *Размер
шрифта
примеров * /
color:  # 555; /* Цвет текста примеров */
}
/ *Стили
для
цветной
буквы * /
.colored - letter
{
    color:  # ff0000; /* Цвет буквы */
}
/ *Стили
для
увеличенной
буквы * /
.enlarged - letter
{
    font - size: 24px; / *Размер
шрифта
для
увеличенной
буквы * /
}
/ *Стили
для
выделенных
букв * /
.special - letter
{
    font - family: 'Arial', sans - serif; / *Используемый
шрифт * /
font - style: italic; / *Курсив * /
                         color:  # 00adff; /* Цвет текста */
}
< / style >
< / head >
< body >
< !-- Заголовок
первого
уровня -->
< h1 > Математические
формулы < / h1 >

< !-- Формула
окружности -->
< h2 > Формула
окружности: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > x < sup > 2 < / sup > + y < sup > 2 < / sup > = r < sup > 2 < / sup > < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Эта
формула
описывает
уравнение
окружности
в
декартовой
системе
координат. < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > x < / span > и < span class ="colored-letter enlarged-letter special-letter" > y < / span > - координаты точки на плоскости.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус окружности.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула окружности описывает геометрическую фигуру,


представляющую
собой
множество
точек, равноудаленных
от
центра. < / b > < / p >
< p


class ="formula-description" > < b > Здесь x и y - координаты точек,


а
r - радиус
окружности. < / b > < / p >

< !-- Формула
площади
треугольника -->
< h2 > Формула
площади
треугольника: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > S = 0.5 * a * b * sin(C) < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > S < / span > - площадь треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > a < / span > и < span class ="colored-letter enlarged-letter special-letter" > b < / span > - длины сторон треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > C < / span > - угол между сторонами a и b (в радианах).< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула площади треугольника выражает площадь как половину произведения длин двух сторон на синус угла между ними.< / b > < / p >

< !-- Формула
объема
цилиндра -->
< h2 > Формула
объема
цилиндра: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > V = π * r < sup > 2 < / sup > * h < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > V < / span > - объем цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > π < / span > (пи) - математическая константа, приблизительно равная 3.14159.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус основания цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > h < / span > - высота цилиндра.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула объема цилиндра выражает объем как произведение площади основания на высоту.< / b > < / p >

< !-- Раздел
с
примерами
применения
формул -->
< div


class ="example-section" >

< h2 > Сферы
применения
формул: < / h2 >
< !-- Примеры -->
< p


class ="example" > 1. < b > Геометрия:<

    / b > расчет
геометрических
параметров
фигур. < / p >
< p


class ="example" > 2. < b > Инженерия:<

    / b > проектирование
и
расчеты
конструкций. < / p >
< p


class ="example" > 3. < b > Физика:<

    / b > моделирование
и
анализ
физических
явлений. < / p >
< / div >

< / body >
< / html >

< !--  ########################################################################################## -->

Бонус

1.
Пример
подключение
шрифтов
через
ссылку

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
внешних
шрифтов -->
< link
rel = "stylesheet"
href = "https://fonts.googleapis.com/css?family=YourFontFamily" >

< !-- Ваши
стили -->
< style >
body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< / style >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >

< !--
Замените
"https://fonts.googleapis.com/css?family=YourFontFamily"
на
фактическую
ссылку
на
шрифт, который
вы
хотите
использовать.Например, если
вы
хотите
использовать
шрифт
"Roboto", ссылка
будет
выглядеть
так: "https://fonts.googleapis.com/css?family=Roboto".

Обратите
внимание, что
это
только
пример, и
вам
нужно
заменить
"YourFontFamily"
и
"Your Title"
на
фактические
значения, используемые
в
вашем
проекте.
-->

2.
Пример
подключения
через
локальные
файлы
шрифтов

< !--
Поместите
шрифтовые
файлы
в
ваш
проект:
Сначала
загрузите
файлы
шрифтов(например,.woff,.woff2,.ttf) в
ваш
проект.
Создайте
папку, например, "fonts", и
поместите
файлы
шрифтов
в
неё.

Пример
структуры
проекта:
-->
/ your - project
/ fonts
your - font.woff
your - font.woff2
index.html
style.css
< !--
Добавьте
стили
в
ваш
CSS:
В
вашем
файле
стилей(например, style.css), определите
новый @ font - face
и
используйте
его
в
стилях
элементов.

Пример:
-->

@font - face


{
    font - family: 'YourFontFamily';
src: url('fonts/your-font.woff2')
format('woff2'),
url('fonts/your-font.woff')
format('woff');
/ *Дополнительные
параметры, если
нужны * /
}

body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< !--
Убедитесь, что
путь
к
файлам
шрифтов
в
url
соответствует
структуре
вашего
проекта.
-->
< !--
Подключите
стили
в
HTML:
В
вашем
файле
HTML(например, index.html), убедитесь, что
вы
подключили
ваш
CSS
файл.

Пример:
-->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
локальных
стилей -->
< link
rel = "stylesheet"
href = "style.css" >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >
< !--
Замените
"Your Title"
на
фактический
заголовок
вашей
страницы.

Это
позволит
вам
использовать
локальные
файлы
шрифтов
в
вашем
проекте.Убедитесь, что
пути
указаны
правильно
и
что
браузер
может
найти
и
загрузить
файлы
шрифтов.
-->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Попробуйте
каждый
тег

На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.

К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)

Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!

2.
Создайте
страницу “Romeo and Juliet”

3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.

4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.

< !--  ########################################################################################## -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS

Домашнее
задание №3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
< !--  #### -->
1.
Попробуйте
каждый
тег
< !--  #### -->
На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.
К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)
< !--  #### -->
Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!
< !--  #### -->
2.
Создайте
страницу “Romeo and Juliet”
< !--  #### -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.
< !--  #### -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.
< !--  #### -->

< !--  ########################################################################################## -->
1.
Попробуйте
каждый
тег
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
2.
Создайте
страницу “Romario and Juluya”
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Romeo and Juliet < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
text - align: center;
}

# title {
font - family: 'Georgia', serif;
color:  # FF4500; /* Цвет - оранжевый */
font - style: italic;
font - size: 36
px;
}

# author {
font - size: 18
px;
font - style: italic;
}

hr
{
    width: 50 %;
margin: auto;
margin - top: 20
px;
border: 1
px
solid  # 000;
}

# prologue {
font - size: 24
px;
margin - top: 20
px;
}

# left-text, #right-text {
display: inline - block;
width: 45 %;
vertical - align: top;
}

.act - title
{
    font - size: 28px;
margin - top: 20
px;
}

.scene - title
{
    font - size: 20px;
font - weight: bold;
margin - top: 10
px;
}

.character
{
    font - weight: bold;
}

.dialogue
{
    text - align: left;
margin - top: 10
px;
margin - bottom: 10
px;
}

.romeo
{
    color:  # 0000FF; /* Цвет - синий */
}

.juliet
{
    color:  # FF1493; /* Цвет - розовый */
}

.sampson
{
    color:  # 008000; /* Цвет - зеленый */
}

.gregory
{
    color:  # FFD700; /* Цвет - золотой */
}
< / style >
    < / head >
        < body >
        < div
id = "title" > ROMEO
AND
JULIET < / div >
           < div
id = "author" > By
William
Shakespeare < / div >
                < hr >

                < div
id = "prologue" >
     < div
id = "left-text" >
     Two
households, both
alike in dignity,
In
fair
Verona, where
we
lay
our
scene,
From
ancient
grudge
break
to
new
mutiny,
Where
civil
blood
makes
civil
hands
unclean.
From
forth
the
fatal
loins
of
these
two
foes
A
pair
of
star - cross’d
lovers
take
their
life;
Whole
misadventured
piteous
overthrows
< / div >

< div
id = "right-text" >
Do
with their death bury their parents’ strife.
The
fearful
passage
of
their
death - mark’d
love,
And
the
continuance
of
their
parents’ rage,
Which, but
their
children’s
end, nought
could
remove,
Is
now
the
two
hours’ traffic
of
our
stage;
The
which if you
with patient ears attend,
What
here
shall
miss, our
toil
shall
strive
to
mend.
< / div >
< / div >
< hr >

< div


class ="act-title" > ACT I < / div >

< div


class ="scene-title" > SCENE I.Verona.A public place.< / div >

< div


class ="scene-title" > < i > Enter SAMPSON and GREGORY, of the house


of
CAPULET, armed
with swords and bucklers </ i > < / div >

< div
id = "left-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > Gregory, o’ my word, we’ll not carry coals.< br >

< span


class ="character gregory" > GREGORY < / span > No, for then we should be colliers.< br >

< span


class ="character sampson" > SAMPSON < / span > I mean, an we be in choler, we’ll draw.< br >

< span


class ="character gregory" > GREGORY < / span > Ay, while you live, draw your neck out


o’ the
collar. < br >
< span


class ="character sampson" > SAMPSON < / span > I strike quickly, being moved.< br >

< span


class ="character gregory" > GREGORY < / span > But thou art not quickly moved to strike.< br >

< span


class ="character sampson" > SAMPSON < / span > A dog of the house of Montague moves me.< br >

< span


class ="character gregory" > GREGORY < / span > To move is to stir; and to be valiant is to


stand: therefore, if thou art moved, thou runn’st away.< br >
< / div >

< div
id = "right-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > A dog of that house shall move me to stand:
    I
    will
    take
    the
    wall
    of
    any
    man or maid
    of
    Montague’s. < br >

< span


class ="character gregory" > GREGORY < / span > That shows thee a weak slave; for the weakest


goes
to
the
wall. < br >
< span


class ="character sampson" > SAMPSON < / span > True; and therefore women, being the


weaker
vessels, are
ever
thrust
to
the
wall: therefore
I
will
push
Montague’s
men
from the wall, and thrust
his
maids
to
the
wall. < br >
< span


class ="character gregory" > GREGORY < / span > The quarrel is between our masters and us


their
men. < br >
< span


class ="character sampson" > SAMPSON < / span > ’Tis all one, I will show myself a tyrant: when


I
have
fought
with the men, I will be cruel with the
maids, and cut
off
their
heads. < br >
< span


class ="character gregory" > GREGORY < / span > The heads of the maids? < br >

< span


class ="character sampson" > SAMPSON < / span > Ay, the heads of the maids, or their


maidenheads;
take
it in what
sense
thou
wilt. < br >
< span


class ="character gregory" > GREGORY < / span > They must take it in sense that feel it.< br >

< / div >

< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №0.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 48px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
background - color:  # fff;
text - align: center;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card hearts" > ♥ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 24px;
margin - right: 10
px;
}
< / style >
< / head >
< body >
< h1 > Игра
в
карты < / h1 >

< div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card" > &  # x1F0A1; Двойка - крести</div>

< div


class ="card" > &  # x1F0C9; Девятка - черви</div>

< div


class ="card" > &  # x1F0C2; Шестерка - черви</div>

< div


class ="card" > &  # x1F0B4; Туз - буби</div>

< div


class ="card" > &  # x1F0A9; Дама - пики</div>

< div


class ="card" > &  # x1F0AC; Валет - крести</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card" > &  # x1F0C8; Тройка - черви</div>

< div


class ="card" > &  # x1F0A1; Туз - крести</div>

< div


class ="card" > &  # x1F0AA; Десятка - крести</div>

< div


class ="card" > &  # x1F0B7; Король - буби</div>

< div


class ="card" > &  # x1F0A6; Восмёрка - пики</div>

< div


class ="card" > &  # x1F0C7; Тройка - черви</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 36px;
margin - right: 10
px;
display: inline - block;
border: 2
px
solid  # 000;
padding: 10
px;
border - radius: 8
px;
}

/ *Цвета
мастей * /
.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card diamonds" > &  # x1F0A1;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №3.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 190px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }

/ * Дополнительные
стили
для
улучшения
внешнего
вида * /
p
{
    font - size: 18px;
margin - bottom: 10
px;
}

strong
{
    font - size: 24px;
}
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C4;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B3;</div>-->

< !-- < div


class ="card spades" > &  # x1F0AA;</div>-->

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card spades" > &  # x1F0A8;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C3;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B2;</div>-->

< !-- < div


class ="card diamonds" > &  # x1F0A1;</div>-->

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №4.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'Arial', sans - serif; / *Шрифт * /
                                             line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
    color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
    margin - top: 20px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
    margin - right: 15px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
transition: transform
0.3
s
ease - in -out; / *Плавное
изменение
при
наведении * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ * Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "36" >

        <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера * /
function
displayCards(playerId, cards, cardSize)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
    const cardElement = document.createElement('div');
cardElement.style.fontSize = `${cardSize}px`; / * Изменение размера карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Отображаем ранг карты и символ масти * /
cardElement.innerText = `${card.rank} ${suitsSymbols[card.suit]}`;
playerCardsElement.appendChild(cardElement);
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
displayCards('player1Cards', player1Cards, newSize);
displayCards('player2Cards', player2Cards, newSize);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
36
px * /
displayCards('player1Cards', player1Cards, 36);
displayCards('player2Cards', player2Cards, 36);
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №5.
                        <!--  ########################################################################################## -->
                          <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Игра
в
карты < / title >
          < style >
/ *Общие
стили
для
всего
документа * /
body
{
font - family: 'Arial', sans - serif; / *Шрифт * /
line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
margin - top: 20
px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
width: 100
px; / *Ширина
карты * /
height: 150
px; / *Высота
карты * /
margin - right: 15
px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
position: relative; / *Позиционирование
элементов
внутри
карты * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ *Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }

/ * Стили
для
значения
карты * /
.card - value
{
    font - size: 14px; / *Размер
шрифта
значения
карты * /
color:  # 555; /* Цвет текста значения карты */
position: absolute; / *Абсолютное
позиционирование
значения * /
}

/ *Стили
для
левого
верхнего
угла
карты * /
.card - value.top - left
{
    top: 10px; / *Отступ
от
верхнего
края
карты * /
left: 10
px; / *Отступ
от
левого
края
карты * /
}

/ *Стили
для
правого
нижнего
угла
карты * /
.card - value.bottom - right
{
    bottom: 10px; / *Отступ
от
нижнего
края
карты * /
right: 10
px; / *Отступ
от
правого
края
карты * /
}
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "50" >

        <!-- Выпадающий
список
для
выбора
масти -->
< label
for ="suitSelect" > Масть карты:< / label >
                                    < select
id = "suitSelect" >
     < option
value = "all" > Все
масти < / option >
          < option
value = "hearts" > Черви < / option >
                             < option
value = "diamonds" > Бубны < / option >
                               < option
value = "clubs" > Трефы < / option >
                            < option
value = "spades" > Пики < / option >
                            < / select >

                                <!-- Выпадающий
список
для
выбора
ранга
карты -->
< label
for ="rankSelect" > Ранг карты:< / label >
                                   < select
id = "rankSelect" >
     < option
value = "all" > Все
ранги < / option >
          < option
value = "A" > Туз < / option >
                      < option
value = "2" > 2 < / option >
                    < option
value = "3" > 3 < / option >
                    < option
value = "4" > 4 < / option >
                    < option
value = "5" > 5 < / option >
                    < option
value = "6" > 6 < / option >
                    < option
value = "7" > 7 < / option >
                    < option
value = "8" > 8 < / option >
                    < option
value = "9" > 9 < / option >
                    < option
value = "10" > 10 < / option >
                      < option
value = "J" > Валет < / option >
                        < option
value = "Q" > Дама < / option >
                       < option
value = "K" > Король < / option >
                         < / select >

                             <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера, масти
и
ранга * /
function
displayCards(playerId, cards, cardSize, selectedSuit, selectedRank)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
if ((selectedSuit === 'all' | | card.suit == = selectedSuit) & & (selectedRank ==
= 'all' | | card.rank == = selectedRank)) {
const cardElement = document.createElement('div');
cardElement.style.width = `${cardSize}px`; / * Изменение ширины карты * /
cardElement.style.height = `${cardSize * 1.5}px`; / * Изменение высоты карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Добавляем значение карты в верхний и нижний угол карты * /
cardElement.innerHTML = `
< div


class ="card-value top-left" > ${card.rank} < / div >

${suitsSymbols[card.suit]}
< div


class ="card-value bottom-right" > ${card.rank} < / div > `;


playerCardsElement.appendChild(cardElement);
}
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
масти
карты * /
document.getElementById('suitSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = this.value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
ранга
карты * /
document.getElementById('rankSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = this.value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
50
px, мастью
"все"
и
рангом
"все" * /
displayCards('player1Cards', player1Cards, 50, 'all', 'all');
displayCards('player2Cards', player2Cards, 50, 'all', 'all');
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №0.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "en" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Пример
стилей
для
страницы(можно
дополнительно
настроить) * /
body
{
    font - family: Arial, sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
}

img
{
    max - width: 100 %;
height: auto;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Welcome
to
the
HTML
Tutorial < / h1 >
             < / header >

                 < section >
                 < article >
                 < p > This
HTML
tutorial
covers
various
HTML
elements and their
usage.Learn
how
to
create
a
basic
HTML
structure and enhance
your
web
development
skills. < / p >
            < / article >

                < article >
                < h2 > Basic
HTML
Structure < / h2 >
              < p > HTML
documents
consist
of
the
following
elements: < / p >
              < ul >
              < li > < code > & lt;!DOCTYPE
html & gt; < / code > declaration < / li >
                                      < li > < code > & lt;
html & gt; < / code > element < / li >
                                  < li > < code > & lt;
head & gt; < / code > element < / li >
                                  < li > < code > & lt;
title & gt; < / code > element < / li >
                                   < li > < code > & lt;
body & gt; < / code > element < / li >
                                  < / ul >
                                      < / article >

                                          < article >
                                          < h2 > Text
Formatting < / h2 >
               < p > Use
tags
like < code > & lt;
p & gt; < / code >, < code > & lt;
h1 & gt; < / code >, < code > & lt;
strong & gt; < / code >, < code > & lt;
em & gt; < / code >,
etc.,
for text formatting. </ p >
< / article >

< article >
< h2 > Images < / h2 >
< p > Include images using the < code > & lt;img & gt; < / code > tag.< / p >
< img src="https://placekitten.com/400/200" alt="Cute Kitten" >
< / article >

< article >
< h2 > Links < / h2 >
< p > Create hyperlinks with the < code > & lt;a & gt; < / code > tag.< / p >
< p > Visit our < a href="https://example.com" target="_blank" > example website < / a >.< / p >
< / article >
< / section >

< footer >
< p > Explore more HTML elements and features to enhance your web development skills.< / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”.Для выполнения используйте все теги,
которые прошли на занятии.- вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE html >
< html lang="ru" >

< head >
< meta charset="UTF-8" >
< meta name="viewport" content="width=device-width, initial-scale=1.0" >
< title > HTML Tutorial < / title >
< style >
/ * Стили для страницы * /
body {
font-family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц.Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < code >
& lt;
img
src = "image.jpg"
alt = "Описание изображения" & gt;
< / code >
    < / article >

        < article >
        < h2 > Ссылки < / h2 >
                          < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Списки < / h2 >
                          < p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
             < code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
    < / article >

        < article >
        < h2 > Видео < / h2 >
                         < p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
                             < code >
& lt;
iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/ВАШ_KОД"
frameborder = "0"
allowfullscreen & gt; & lt; / iframe & gt;
< / code >
    < / article >
        < / section >

            < footer >
            < p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
                      < / footer >

                          < / body >

                              < / html >
                                  <!--  ########################################################################################## -->
                                    <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №1_1.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "ru" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц. < / p > < p > Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
      < / article >

          < article >
          < h2 > Ссылки < / h2 >
                            < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Стилизованный
текст < / h2 >
          < p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
           < span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "ru" >

< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML
Tutorial < / title >
< style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}

hr
{
    margin: 20px 0;
border: none;
border - top: 2
px
solid  # 007bff;
}
< / style >
< / head >

< body >

< header >
< h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
< / header >

< section >
< article >
< p > < b > HTML < / b > (Hypertext Markup Language) - это
код, который
используется
для
структурирования
и
отображения
веб - страницы
и
её
контента. < / p >
< p > Например, контент
может
быть
структурирован
внутри
множества
параграфов, маркированных
списков
или
с
использованием
изображений
и
таблиц
данных. < / p >
< p > Как
видно
из
названия, этот
туториал
постарается
дать
вам
базовое
понимание
HTML
и
его
функций. < / p >
< / article >
< hr >
< p > < b > HTML < / b > не
является
языком
программирования;
это
язык
разметки, и
используется, чтобы
сообщать
вашему
браузеру, как
отображать
веб - страницы, которые
вы
посещаете. < / p >
< p > Он
может
быть
сложным
или
простым, в
зависимости
от
того, как
хочет
веб - дизайнер. < / p >
< p > HTML
состоит
из
ряда
элементов, которые
вы
используете, чтобы
вкладывать
или
оборачивать
различные
части
контента, чтобы
заставить
контент
отображаться
или
действовать
определённым
образом. < / p >
< p > Ограждающие
теги
могут
сделать
слово
или
изображение
ссылкой
на
что - то
ещё, могут
сделать
слова
курсивом, сделать
шрифт
больше
или
меньше
и
так
далее. < / p >
< hr >

< article >
< h2 > Основная
структура
HTML < / h2 >
< p > HTML - документ
состоит
из
следующих
элементов: < / p >
< ul >
< li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
< li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
< li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
< li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
< li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
< / ul >
< p > Вот
пример
простой
структуры
HTML: < / p >
< code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Форматирование
текста < / h2 >
< p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
< ul >
< li > < code > & lt;
p & gt; < / code >: Параграф < / li >
< li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
< li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
< li > < code > & lt;
em & gt; < / code >: Курсив < / li >
< / ul >
< p > Пример
использования
тегов
форматирования
текста: < / p >
< code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Изображения < / h2 >
< p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
< img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
< / article >

< hr >

< article >
< h2 > Ссылки < / h2 >
< p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
< code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Цветной
текст < / h2 >
< p > Применение
цветового
стиля
к
тексту: < / p >
< code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Стилизованный
текст < / h2 >
< p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
< span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< hr >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !--  ########################################################################################## -->
< !-- HTML  # 4 - Блочная структура элементов. Свойство Display. Размеры - 29.01.2024 -->
< !-- ПРАКТИЧЕСКАЯ
РАБОТА
во
ВРЕМЯ
КЛАССНОГО
УРОКА(ПАРЫ) - 29.01
.2024 -->
< !-- ВАРИАНТ №1 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-rocket float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-refresh float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:
< !DOCTYPE
html >
< html
lang = "en" >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Заголовок
документа:

< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< title >: Заголовок
документа, отображается
во
вкладке
браузера.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Устанавливает
настройки
просмотра
для
устройств
с
различной
шириной
экрана.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
стилей:

< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

Подключаются
стили
Bootstrap
и
шрифтов
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
стилей
внутри
тега < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Здесь
определены
пользовательские
стили
для
элементов.Например, фон
body, стили
карточек.dashboard - card
и
размер
иконок
внутри
этих
карточек.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Основное
содержимое
документа:

< body >
< !-- ... -->
< / body >

Весь
контент
страницы
помещается
внутрь
тега < body >.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Контент
с
использованием
Bootstrap:

Внутри < body > размещаются
карточки, созданные
с
использованием
Bootstrap.
Каждая
карточка
находится
внутри < div


class ="card bg-... dashboard-card" > и имеет свой уникальный


цвет(bg - info, bg - success, bg - warning).
Иконки
и
текст
внутри
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- <!-- ВАРИАНТ №2 --> -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Продукция в пути < / h6 >

< h2


class ="text-right" > < i class ="fa fa-truck float-left" > < / i > < span > 123 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные доставки < span class ="float-right" > 87 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Товары на складе < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cubes float-left" > < / i > < span > 789 < / span > < / h2 >

< p


class ="m-b-0" > Товары в резерве < span class ="float-right" > 456 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-danger dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Возвраты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-undo float-left" > < / i > < span > 10 < / span > < / h2 >

< p


class ="m-b-0" > Обработанные возвраты < span class ="float-right" > 5 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-primary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Поддержка клиентов < / h6 >

< h2


class ="text-right" > < i class ="fa fa-headphones float-left" > < / i > < span > 24 / 7 < / span > < / h2 >

< p


class ="m-b-0" > Запросы < span class ="float-right" > 120 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-secondary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Новые клиенты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-users float-left" > < / i > < span > 50 < / span > < / h2 >

< p


class ="m-b-0" > Клиенты на обслуживании < span class ="float-right" > 20 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< / head >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей
в
теге < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Устанавливаются
пользовательские
стили
для
элементов
страницы, таких
как
body
и.dashboard - card.
Эти
стили
определяют
отступы, фон, тени
и
другие
характеристики.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Содержимое
страницы:

< body >
< div


class ="container" >

< div


class ="row" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >
Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4" > (каждая из них занимает 1 / 3 ширины экрана).


Класс
card
определяет, что
это
карточка
Bootstrap.Классы
bg - info, bg - success, bg - warning, и
т.д.,
устанавливают
цвет
фона
для
каждой
карточки.
Иконки
и
содержимое
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- ВАРИАНТ №3 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
/ *Дополнительные
стили
или
анимации
могут
быть
добавлены
здесь * /
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< div


class ="col-md-4 mb-3" >

< div


class ="card bg-info dashboard-card" id="card1" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Еда и Напитки < / h6 >

< i


class ="fa fa-cutlery float-left" > < / i >

< h2


class ="text-right" > < span id="value1" > 100 < / span > < / h2 >

< p


class ="m-b-0" > Выполненные заказы < span class ="float-right" > 50 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-success dashboard-card" id="card2" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Ноутбуки и ПК < / h6 >

< i


class ="fa fa-laptop float-left" > < / i >

< h2


class ="text-right" > < span id="value2" > 75 < / span > < / h2 >

< p


class ="m-b-0" > Товары в наличии < span class ="float-right" > 25 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-warning dashboard-card" id="card3" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Транспорт < / h6 >

< i


class ="fa fa-car float-left" > < / i >

< h2


class ="text-right" > < span id="value3" > 50 < / span > < / h2 >

< p


class ="m-b-0" > Автомобили в наличии < span class ="float-right" > 10 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-danger dashboard-card" id="card4" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Смартфоны и Гаджеты < / h6 >

< i


class ="fa fa-mobile float-left" > < / i >

< h2


class ="text-right" > < span id="value4" > 120 < / span > < / h2 >

< p


class ="m-b-0" > Гаджеты в наличии < span class ="float-right" > 80 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-primary dashboard-card" id="card5" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Доставка < / h6 >

< i


class ="fa fa-truck float-left" > < / i >

< h2


class ="text-right" > < span id="value5" > 60 < / span > < / h2 >

< p


class ="m-b-0" > Заказы в пути < span class ="float-right" > 400 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-secondary dashboard-card" id="card6" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Одежда и Аксессуары < / h6 >

< i


class ="fa fa-shopping-bag float-left" > < / i >

< h2


class ="text-right" > < span id="value6" > 30 < / span > < / h2 >

< p


class ="m-b-0" > Модные новинки 2024 < span class ="float-right" > 20 000 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >
< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >
    < / body >
        < / html >

            <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "utf-8" >
          < title > Градиентные
карточки
для
панели
управления < / title >
               < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          < / head >

              <!DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          <!-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
       < link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
       <!-- Стили -->
< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    < meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей:

< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    Содержимое
страницы:

< body >
  < div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >

Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4 mb-3" > (каждая из них занимает 1 / 3 ширины экрана).


Индивидуальные
идентификаторы(id)
присваиваются
каждой
карточке(card1, card2, и
т.д.).
Иконки
и
содержимое
карточек
заполняются
данными.Обратите
внимание
на
использование
спанов
с
id
для
отображения
и
обновления
случайных
значений.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
JavaScript - скрипт:

< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
// Обновление
случайных
значений
внутри
карточек
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >

    Создаются
две
функции: getRandomInt
для
получения
случайного
целого
числа
в
заданном
диапазоне
и
updateValues
для
обновления
значений
в
карточках.
setInterval(updateValues, 5000): Запуск
функции
обновления
значений
каждые
5
секунд.
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
ПРИСУЦТВУЮТ
НЕ
ВСЕ
ПРАКТИЧЕСКИЕ
И
НЕ
ВСЕ
ДОМАШНИЕ
ЗАДАНИЯ!
ОНИ
В
ОТДЕЛЬНЫХ
ФАЙЛАХ.html
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !--  ########################################################################################## -->
   <!-- HTML  # 1 Recap + Вводный ПЕРВЫЙ УРОК - 08.01.2024 -->

     <!-- < html >
       < head >

       < meta
charset = "UTF-8" / >
          < title > Hello
HTML < / title >
         < / head >
             < body >

             hello
world
Привет
Мир

< img
src = "https://pngicon.ru/file/uploads/serdce.png"
height = "150px"
width = "150px" / >
        < h1 > Вау - я
программист < / h1 >
                < / body >
                    < / html > -->
< !--  ########################################################################################## -->
   <!--  ########################################################################################## -->
     <!--  ########################################################################################## -->

       <!-- 2024
_HTML_LESSON_№2
_Структура_HTML.Теги_Атрибуты - 15.01
.2024 -->

< !--

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 2: Структура
HTML.Теги.Атрибуты.Правила
языка
разметки
-->

< !DOCTYPE
html >
< html
lang = "ru" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Страница
с
текстом
и
заголовками < / title >
                < / head >
                    < body >

                    <!-- 1.
Создать
страницу
с
текстом
по
примеру
на
картинке. -->

< h2 >
  Стихотворение
  < / h2 >
      < p >

      Мириады
маленьких
дел < br >
Пьют
по
капле
гаснущий
день, < br >
        А < i > дела
большие < / i > сушит
жажда. < br >
Оставляя
все
на «потом», < br >
              Прозреваем
задним
числом, < br >
          Только
год < b > не
повторится < / b > дважды. < br >
               < br >
               И.Тальков

               < / p >

                   <!-- 2.
Повторите
по
данному
образцу: -->
< h1 > Это
заголовок < / h1 >
              < h2 > Это
заголовок < / h2 >
              < h3 > Это
заголовок < / h3 >
              < h4 > Это
заголовок < / h4 >

              < p > Это < b > абзац < / b >.< / p >
                                                < p > Это
еще < i > абзац < / i >.< / p >

                            < h1 > < i > Это
заголовок
h1 < / i > < / h1 >

               < / body >
                   < / html >

                       <!--  ########################################################################################## -->
                         <!--  ########################################################################################## -->
                           <!--  ########################################################################################## -->

03
_Введение
в
CSS.Форматирование
текста
при
помощи
CSS - 22.01
.2024
Классная
работа

<!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < link
rel = "preconnect"
href = "https://fonts.googleapis.com" >
       < link
rel = "preconnect"
href = "https://fonts.gstatic.com"
crossorigin >
< link
href = "https://fonts.googleapis.com/css2?family=Rubik+Burned&display=swap"
rel = "stylesheet" >

      < meta
charset = "UTF-8" >
          < title > Lesson
2 < / title >
      < script >
/ *легкий
init
js * /
< / script >
    < meta
description = "Краткое описание содержимого/контента страницы" / >
              < meta
keywords = "купить, по акции, новый" / >
           < style >
           html, body
{font - size: 10px;
font - family: "Rubik Burned", system - ui;}
h1
{font - size: 4rem}
h2
{font - size: 3rem}
h3
{font - size: 2rem}
h4
{font - size: 1rem}
< / style >
    < link
rel = "stylesheet"
href = "./style.css" / >
       < script
src = "https://cdn.tailwindcss.com" > < / script >
                                          < / head >
                                              < body >
                                              < div
style = "display: flex" >
        < img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
      < img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
      < img
style = "width: 33.333333%"
src = "https://mir-s3-cdn-cf.behance.net/projects/202/08d50f175132235.Y3JvcCwxODIzLDE0MjYsMTQ4LDEzMA.jpeg" / >
      < / div >
          < div >
          < h1


class ="heading-of-section text-red text-krasnii" > Красный < / h1 >

< p


class ="text-orange" > Оранжевый < / p >

< div >
< p


class ="button default" > Нажми меня < / p >

< div


class ="button alert" > Нажми меня < / div >

< button


class ="button alert" alt="Для оформления заказа нажмите эту кнопку" > Нажми меня < / button >

< / div >
< div >
< div


class ="display-inline rounded-md shadow-md padding-md bg-gray" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-red" > Нажми меня < / div >

< div


class ="display-inline rounded-md shadow-md padding-md bg-blue" > Нажми меня < / div >

< / div >
< / div >
< b > Жирный
текст < / b >
< h1
style = "text-decoration: line-through" > Панды, наконец - то
полетят
на
Марс < / h1 >
< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3


class ="heading" > Hello world < / h3 >

< h4


class ="text-red" > Hello world < / h4 >

< h5 > Hello
world < / h5 >
< h6


class ="bg-blue" > Hello world < / h6 >

< h2


class ="heading" > Глава 1: далекий


путь
с
одного
шага < / h2 >
< h3 > Hello
world < / h3 >
< h4
id = "heading-h4-1" > Hello
world < / h4 >
< h4
id = "heading-h4-2" > Hello
world < / h4 >
< h4 > Hello
world < / h4 >
< h3


class ="text-lg text-rose-500 bg-gray-100 rounded-md" > Hello world < / h3 >

< h3 > Hello
world < / h3 >

< p >
paragraph < b > lorem < / b >
< br / >
< br / >
< br / >
< br / >
< br / >
ipsum
set < i > < b > ВАЖНО! < / b > < / i > dolor
< br / >
ipsum
set < strong > < em > Кликни & rarr; < / em > < / strong > dolor
< br / >
2 * 2
< br / >
2
x2
< br / >
2 & cross;
2
< / p >
< / body >
< / html >

< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
22.01
.2024 - Домашняя
и
практическая - работы.
< !--  ########################################################################################## -->
Практическая
работа

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №3
Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Создать
HTML - страницу “Vehicle”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

2.
Создать
HTML - страницу “Lorem
Ipsum”
Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.

3.
Создать
HTML - страницу “Математические
формулы”

Текст
можете
найти
в
архиве
Classwork
в
папке
Материалы
к
занятию.
Используйте
здесь
теги
h1 - h6, span, p, sup, sub
и
спецсимволы.
< !--  ########################################################################################## -->

Шрифты
находятся
в
папке
Материалы
к
занятию
в
архиве
Fonts.

1.
Пример
подключение
шрифтов
через
ссылку
2.
Пример
подключения
через
локальные
файлы
шрифтов
< !--  ########################################################################################## -->

№1

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Двигатель < / title >
< style >
body
{
    font - family: 'Tahoma', serif;
line - height: 1.6;
margin: 18
px;
}

h1
{
    color:  # 00c264;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align: justify;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.smaller - text
{
    font - size: 58 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
< / style >

< / head >
< body >
< h1


class ="larger-font2" >


Двигатель
< / h1 >
< !-- Текст
о
двигателе
с
сайта
https: // ru.wikipedia.org / wiki / Двигатель -->

< p


class ="larger-font" >

< b > Дви́гатель < / b > — устройство, преобразующее
какой - либо
вид < a
href = "https://ru.wikipedia.org/wiki/Энергия"


class ="no-underline" > энергии < / a >


в < a
href = "https://ru.wikipedia.org/wiki/Механическая_энергия"


class ="no-underline" > механическую < / a > работу.


Термин < b > мотор < / b > < a
href = "https://ru.wikipedia.org/wiki/Заимствование"


class ="no-underline" > заимствован < / a > в


первой
половине
XIX
века
из
немецкого
языка < b


class ="smaller-text colored-letter" >[1] < / b >


(нем.Motor — «двигатель», от лат.mōtor — «приводящий в движение»)
и
преимущественно
им
называют
электрические
двигатели
и
двигатели
внутреннего
сгорания < b


class ="smaller-text colored-letter" >


[2] < / b >.
< / p >
< p


class ="larger-font" >


Двигатели
подразделяют
на
первичные
и
вторичные.
К
первичным
относят
непосредственно
преобразующие
природные
энергетические
ресурсы
в
механическую
работу,
а
ко
вторичным — преобразующие
энергию, выработанную
или
накопленную
другими
источниками < b


class ="smaller-text colored-letter" >


[3] < / b >.
< / p >
< hr


class ="subtle-line" >

< p >
< h4 > Примечания < / h4 >
< h5 > [1] < / h5 > Шанский
Н.М., Боброва
Т.А.Кот // Школьный
этимологический
словарь
русского
языка.
Происхождение
слов. — 7 - е
изд., стереотип.. — М.: Дрофа, 2004. — 398, [2]
с.
< h5 > [2] < / h5 > Крысин
Л.П.Мотор // Толковый
словарь
иноязычных
слов. — М.: Эксмо, 2008. — 944
с. —
(Библиотека словарей).
< h5 > [3] < / h5 > Definition
of
motor | Dictionary.com(англ.).www.dictionary.com.
Дата
обращения: 27
января
2022.
Архивировано
27
января
2022
года.
< / p >
< / body >
< / html >
< !--  ########################################################################################## -->

№2

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Lorem
Ipsum < / title >
< style >

h1
{
    color:  # 5d0080;
}
h2
{
    color:  # 9b05ff;
}
h3
{
    color:  # ff0505;
}
h4
{
    color:  # 00a0eb;
}
h5
{
    color:  # 00a0eb;
}
h6
{
    color:  # ffff05;
}
p
{
    text - align0: justify;
}
.left - aligned1
{
    text - align: left;
}

.right - aligned2
{
    text - align: right;
}

.custom - font
{
    font - family: "Courier New", monospace;
}
.custom - font2
{
    font - family: "mv boli", monospace;
}
.custom - font3
{
    font - family: "segoe print", monospace;
}
.custom - font4
{
    font - family: "LUCIDA SANS CONSOLE", monospace;
}
.custom - font5
{
    font - family: "GABRIOLA", monospace;
}
.custom - font6
{
    font - family: "tempus sans itc", monospace;
}
.bold - text
{
    font - weight: bold;
}

.italic - text
{
    font - style: italic;
}
.no - underline
{
    text - decoration: none;
}
.larger - font
{
    font - size: 20px;
}
.larger - font2
{
    font - size: 40px;
}
.larger - font3
{
    font - size: 50px;
}
.centered - text
{
    text - align: center;
}

.centered - container
{
    margin: 0 auto;
width: 50 %; / *Измените
на
желаемую
ширину * /
}

.smaller - text
{
    font - size: 95 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text2
{
    font - size: 80 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text3
{
    font - size: 115 %; / *Измените
на
желаемый
размер
шрифта * /
}
.smaller - text4
{
    font - size: 15 %; / *Измените
на
желаемый
размер
шрифта * /
}
.slightly - bold
{
    font - weight: bold; / *или
font - weight: 600; * /
}
.subtle - line
{
    border: 2px solid  # 9e9e9e; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.subtle - line2
{
    border: 1px solid  # f2f2f2; /* Измените на нужный цвет и стиль линии */
    margin: 0px 0; / *Добавьте
отступы
по
желанию * /
}
.colored - letter
{
    color:  # 00a0eb; /* Измените на желаемый цвет */
}
.custom - color1
{
    color:  # 001fb8; /* Замените на желаемый цвет */
}
.custom - color2
{
    color:  # 6e6e6e; /* Замените на желаемый цвет */
}
.page - container
{
    margin: 20px; / *Замените
на
желаемое
значение
отступа * /
}
< / style >
< / head >
< body >
< h1


class ="centered-text larger-font3 custom-font4" > Lorem Ipsum < / h1 >

< div


class ="right-aligned2 custom-color1 custom-font3 italic-text smaller-text" >

< p > "Neque porro quisquam est qui dolorem ipsum quia dolor sit amet, consectetur, adipisci velit...". < / p >
< / div >
< div


class ="right-aligned2 custom-font2 custom-color2 italic-text smaller-text2 larger-font" >

< p > "Нет никого, кто любил бы боль саму по себе, кто искал бы её и кто хотел бы иметь её
просто
потому, что
это
боль..
".</p>
< / div >
< hr


class ="subtle-line2" >

< !-- Пример
текста
Lorem
Ipsum -->
< div


class ="page-container" >

< p


class ="larger-font custom-font" > < b >


Классический
текст
Lorem
Ipsum, используемый
с
XVI
века
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>L</b>orem ipsum dolor sit amet, consectetur adipiscing elit, sed do eiusmod tempor incididunt
ut
labore
et
dolore
magna
aliqua.Ut
enim
ad
minim
veniam, quis
nostrud
exercitation
ullamco
laboris
nisi
ut
aliquip
ex
ea
commodo
consequat.Duis
aute
irure
dolor in reprehenderit in voluptate
velit
esse
cillum
dolore
eu
fugiat
nulla
pariatur.Excepteur
sint
occaecat
cupidatat
non
proident, sunt in culpa
qui
officia
deserunt
mollit
anim
id
est
laborum.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.32
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>S</b>ed ut perspiciatis unde omnis iste natus error sit voluptatem accusantium doloremque laudantium, totam rem
aperiam, eaque
ipsa
quae
ab
illo
inventore
veritatis
et
quasi
architecto
beatae
vitae
dicta
sunt
explicabo.
Nemo
enim
ipsam
voluptatem
quia
voluptas
sit
aspernatur
aut
odit
aut
fugit, sed
quia
consequuntur
magni
dolores
eos
qui
ratione
voluptatem
sequi
nesciunt.Neque
porro
quisquam
est, qui
dolorem
ipsum
quia
dolor
sit
amet,
consectetur, adipisci
velit, sed
quia
non
numquam
eius
modi
tempora
incidunt
ut
labore
et
dolore
magnam
aliquam
quaerat
voluptatem.Ut
enim
ad
minima
veniam, quis
nostrum
exercitationem
ullam
corporis
suscipit
laboriosam, nisi
ut
aliquid
ex
ea
commodi
consequatur? Quis
autem
vel
eum
iure
reprehenderit
qui in ea
voluptate
velit
esse
quam
nihil
molestiae
consequatur, vel
illum
qui
dolorem
eum
fugiat
quo
voluptas
nulla
pariatur?"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>B</b>ut I must explain to you how all this mistaken idea of denouncing pleasure and praising pain was born and
I
will
give
you
a
complete
account
of
the
system, and expound
the
actual
teachings
of
the
great
explorer
of
the
truth, the
master - builder
of
human
happiness.No
one
rejects, dislikes, or avoids
pleasure
itself, because
it is
pleasure, but
because
those
who
do
not know
how
to
pursue
pleasure
rationally
encounter
consequences
that
are
extremely
painful.Nor
again is there
anyone
who
loves or pursues or desires
to
obtain
pain
of
itself, because
it
is pain, but
because
occasionally
circumstances
occur in which
toil and pain
can
procure
him
some
great
pleasure.
To
take
a
trivial
example, which
of
us
ever
undertakes
laborious
physical
exercise, except to
obtain
some
advantage
from it? But
who
has
any
right
to
find
fault
with a man who chooses to enjoy a pleasure that has no annoying
consequences, or one
who
avoids
a
pain
that
produces
no
resultant
pleasure?"
< / p >
< p


class ="larger-font custom-font" > < b >


Абзац
1.10
.33
"de Finibus Bonorum et Malorum", написанный
Цицероном
в
45
году
н.э.
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>A</b>t vero eos et accusamus et iusto odio dignissimos ducimus qui blanditiis praesentium voluptatum deleniti
atque
corrupti
quos
dolores
et
quas
molestias
excepturi
sint
occaecati
cupiditate
non
provident, similique
sunt
in culpa
qui
officia
deserunt
mollitia
animi, id
est
laborum
et
dolorum
fuga.Et
harum
quidem
rerum
facilis
est
et
expedita
distinctio.Nam
libero
tempore, cum
soluta
nobis
est
eligendi
optio
cumque
nihil
impedit
quo
minus
id
quod
maxime
placeat
facere
possimus, omnis
voluptas
assumenda
est, omnis
dolor
repellendus.Temporibus
autem
quibusdam
et
aut
officiis
debitis
aut
rerum
necessitatibus
saepe
eveniet
ut
et
voluptates
repudiandae
sint
et
molestiae
non
recusandae.Itaque
earum
rerum
hic
tenetur
a
sapiente
delectus, ut
aut
reiciendis
voluptatibus
maiores
alias
consequatur
aut
perferendis
doloribus
asperiores
repellat.
"
< / p >
< p


class ="larger-font custom-font" > < b >


Английский
перевод
1914
года, H.Rackham
< / p > < / b >
< p


class ="custom-font6 smaller-text3" >


"<b>O</b>n the other hand, we denounce with righteous indignation and dislike men who are so beguiled and demoralized
by
the
charms
of
pleasure
of
the
moment, so
blinded
by
desire, that
they
cannot
foresee
the
pain and trouble
that
are
bound
to
ensue; and equal
blame
belongs
to
those
who
fail in their
duty
through
weakness
of
will, which is the
same as saying
through
shrinking
from toil and pain.These
cases
are
perfectly
simple and easy
to
distinguish.
In
a
free
hour, when
our
power
of
choice is untrammelled and when
nothing
prevents
our
being
able
to
do
what
we
like
best, every
pleasure is to
be
welcomed and every
pain
avoided.But in certain
circumstances and owing
to
the
claims
of
duty or the
obligations
of
business
it
will
frequently
occur
that
pleasures
have
to
be
repudiated and
annoyances
accepted.The
wise
man
therefore
always
holds in these
matters
to
this
principle
of
selection: he
rejects
pleasures
to
secure
other
greater
pleasures, or else he
endures
pains
to
avoid
worse
pains.
"
< / p >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

№3

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Математические
формулы < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'VERDANA', monospace; / *Устанавливаем
шрифт
для
всего
документа * /
line - height: 1.0; / *Задаем
высоту
строки
текста * /
margin: 70
px; / *Задаем
отступы
для
всего
документа * /
}

/ *Стили
для
заголовка
первого
уровня * /
h1
{
    color:  # 000000; /* Цвет текста */
        text - align: center; / *Выравнивание
текста
по
центру * /
font - size: 50
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
}

/ *Стили
для
заголовка
второго
уровня * /
h2
{
    color:  # 22009e; /* Цвет текста */
        font - size: 35
px; / *Размер
шрифта * /
margin - top: 20
px; / *Верхний
отступ * /
}

/ *Стили
для
подзаголовка * /
h3
{
    color:  # 00ff49; /* Цвет текста */
        font - size: 25
px; / *Размер
шрифта * /
font - weight: bold; / *Жирный
шрифт * /
font - style: italic; / *Курсив * /
                         text - align: left; / *Выравнивание
текста
влево * /
background - color:  # a300a3b8; /* Цвет подложки */
padding: 10
px; / *Внутренний
отступ * /
border - radius: 45
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 320
px; / *Максимальная
ширина
подложки * /
width: 25 %; / *Ширина
подложки
равна
20 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: left; / *Выравнивание
по
левой
стороне * /
}

/ *Стили
для
обычного
текста
и
вложенных
элементов * /
p, span
{
    text - align: justify; / *Выравнивание
текста
по
ширине
с
обеих
сторон * /
}

/ *Стили
для
верхних
и
нижних
индексов * /
sup, sub
{
    font - size: 80 %; / *Размер
шрифта
для
верхних
и
нижних
индексов * /
}

/ *Отступ
для
блока
с
описанием
формулы * /
.formula - description
{
    margin - left: 20px; / *Отступ
слева * /
}

/ *Стили
для
контейнера
формулы
и
самой
формулы * /
.formula - container,.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
background - color:  # ff00008f; /* Цвет подложки */
padding: 2
px; / *Внутренний
отступ * /
border - radius: 55
px; / *Закругленные
углы
подложки * /
box - shadow: 0
0
15
px
rgba(0, 0, 0, 0.2); / *Тень
подложки * /
max - width: 670
px; / *Максимальная
ширина
подложки * /
width: 50 %; / *Ширина
подложки
равна
50 % родительского
контейнера * /
box - sizing: border - box; / *Учитываем
padding
и
border
в
общей
ширине * /
margin: auto; / *Выравнивание
по
центру * /
}

/ *Стили
для
самой
формулы * /
.formula
{
    text - align: center; / *Выравнивание
текста
по
центру * /
color:  # 0700ff; /* Цвет формулы */
font - size: 25
px; / *Размер
шрифта
формулы * /
font - style: italic; / *Курсив * /
                         font - weight: bold; / *Жирный
шрифт
формулы * /
}

/ *Стили
для
раздела
с
примерами * /
.example - section
{
    margin - top: 20px; / *Верхний
отступ
раздела
с
примерами * /
}

/ *Стили
для
примеров * /
.example
{
    font - size: 20px; / *Размер
шрифта
примеров * /
color:  # 555; /* Цвет текста примеров */
}
/ *Стили
для
цветной
буквы * /
.colored - letter
{
    color:  # ff0000; /* Цвет буквы */
}
/ *Стили
для
увеличенной
буквы * /
.enlarged - letter
{
    font - size: 24px; / *Размер
шрифта
для
увеличенной
буквы * /
}
/ *Стили
для
выделенных
букв * /
.special - letter
{
    font - family: 'Arial', sans - serif; / *Используемый
шрифт * /
font - style: italic; / *Курсив * /
                         color:  # 00adff; /* Цвет текста */
}
< / style >
< / head >
< body >
< !-- Заголовок
первого
уровня -->
< h1 > Математические
формулы < / h1 >

< !-- Формула
окружности -->
< h2 > Формула
окружности: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > x < sup > 2 < / sup > + y < sup > 2 < / sup > = r < sup > 2 < / sup > < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Эта
формула
описывает
уравнение
окружности
в
декартовой
системе
координат. < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > x < / span > и < span class ="colored-letter enlarged-letter special-letter" > y < / span > - координаты точки на плоскости.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус окружности.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула окружности описывает геометрическую фигуру,


представляющую
собой
множество
точек, равноудаленных
от
центра. < / b > < / p >
< p


class ="formula-description" > < b > Здесь x и y - координаты точек,


а
r - радиус
окружности. < / b > < / p >

< !-- Формула
площади
треугольника -->
< h2 > Формула
площади
треугольника: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > S = 0.5 * a * b * sin(C) < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > S < / span > - площадь треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > a < / span > и < span class ="colored-letter enlarged-letter special-letter" > b < / span > - длины сторон треугольника.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > C < / span > - угол между сторонами a и b (в радианах).< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула площади треугольника выражает площадь как половину произведения длин двух сторон на синус угла между ними.< / b > < / p >

< !-- Формула
объема
цилиндра -->
< h2 > Формула
объема
цилиндра: < / h2 >
< !-- Контейнер
формулы -->
< div


class ="formula-container" >

< !-- Сама
формула -->
< p


class ="formula" > < b > V = π * r < sup > 2 < / sup > * h < / b > < / p >

< / div >
< !-- Описание
формулы -->
< div


class ="formula-description" >

< p > < b > Где: < / b > < / p >
< ul > < b >
< li > < span


class ="colored-letter enlarged-letter special-letter" > V < / span > - объем цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > π < / span > (пи) - математическая константа, приблизительно равная 3.14159.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > r < / span > - радиус основания цилиндра.< / li >

< li > < span


class ="colored-letter enlarged-letter special-letter" > h < / span > - высота цилиндра.< / li >

< / ul > < / b >
< / div >
< !-- Подробное
описание
формулы -->
< h3 > Описание
формулы: < / h3 >
< p


class ="formula-description" > < b > Формула объема цилиндра выражает объем как произведение площади основания на высоту.< / b > < / p >

< !-- Раздел
с
примерами
применения
формул -->
< div


class ="example-section" >

< h2 > Сферы
применения
формул: < / h2 >
< !-- Примеры -->
< p


class ="example" > 1. < b > Геометрия:<

    / b > расчет
геометрических
параметров
фигур. < / p >
< p


class ="example" > 2. < b > Инженерия:<

    / b > проектирование
и
расчеты
конструкций. < / p >
< p


class ="example" > 3. < b > Физика:<

    / b > моделирование
и
анализ
физических
явлений. < / p >
< / div >

< / body >
< / html >

< !--  ########################################################################################## -->

Бонус

1.
Пример
подключение
шрифтов
через
ссылку

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
внешних
шрифтов -->
< link
rel = "stylesheet"
href = "https://fonts.googleapis.com/css?family=YourFontFamily" >

< !-- Ваши
стили -->
< style >
body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< / style >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >

< !--
Замените
"https://fonts.googleapis.com/css?family=YourFontFamily"
на
фактическую
ссылку
на
шрифт, который
вы
хотите
использовать.Например, если
вы
хотите
использовать
шрифт
"Roboto", ссылка
будет
выглядеть
так: "https://fonts.googleapis.com/css?family=Roboto".

Обратите
внимание, что
это
только
пример, и
вам
нужно
заменить
"YourFontFamily"
и
"Your Title"
на
фактические
значения, используемые
в
вашем
проекте.
-->

2.
Пример
подключения
через
локальные
файлы
шрифтов

< !--
Поместите
шрифтовые
файлы
в
ваш
проект:
Сначала
загрузите
файлы
шрифтов(например,.woff,.woff2,.ttf) в
ваш
проект.
Создайте
папку, например, "fonts", и
поместите
файлы
шрифтов
в
неё.

Пример
структуры
проекта:
-->
/ your - project
/ fonts
your - font.woff
your - font.woff2
index.html
style.css
< !--
Добавьте
стили
в
ваш
CSS:
В
вашем
файле
стилей(например, style.css), определите
новый @ font - face
и
используйте
его
в
стилях
элементов.

Пример:
-->

@font - face


{
    font - family: 'YourFontFamily';
src: url('fonts/your-font.woff2')
format('woff2'),
url('fonts/your-font.woff')
format('woff');
/ *Дополнительные
параметры, если
нужны * /
}

body
{
    font - family: 'YourFontFamily', sans - serif;
/ *Другие
стили... * /
}
< !--
Убедитесь, что
путь
к
файлам
шрифтов
в
url
соответствует
структуре
вашего
проекта.
-->
< !--
Подключите
стили
в
HTML:
В
вашем
файле
HTML(например, index.html), убедитесь, что
вы
подключили
ваш
CSS
файл.

Пример:
-->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Your
Title < / title >

< !-- Подключение
локальных
стилей -->
< link
rel = "stylesheet"
href = "style.css" >
< / head >
< body >
< !-- Ваш
контент -->
< / body >
< / html >
< !--
Замените
"Your Title"
на
фактический
заголовок
вашей
страницы.

Это
позволит
вам
использовать
локальные
файлы
шрифтов
в
вашем
проекте.Убедитесь, что
пути
указаны
правильно
и
что
браузер
может
найти
и
загрузить
файлы
шрифтов.
-->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Домашнее
задание № 3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
1.
Попробуйте
каждый
тег

На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.

К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)

Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!

2.
Создайте
страницу “Romeo and Juliet”

3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.

4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.

< !--  ########################################################################################## -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >

< !--  ########################################################################################## -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS

Домашнее
задание №3: Введение
в
CSS.Форматирование
текста
при
помощи
CSS
< !--  #### -->
1.
Попробуйте
каждый
тег
< !--  #### -->
На
каждой
новой
строчке
напишите
по
предложению
используя
все
пройденные
теги.
К
каждому
тегу
добавьте
атрибуты


class или id


Затем
измените
с
помощью
CSS
цвет
этим
предложениям(все
должны
быть
разного
цвета)
< !--  #### -->
Чем
разнообразнее
будут
использоваться
CSS
свойства, тем
выше
вы
получите
балл.CSS
свойства
можно
также
подключать
или
использовать
по
разному.Чем
разнообразнее, тем
лучше.Удачи!
< !--  #### -->
2.
Создайте
страницу “Romeo and Juliet”
< !--  #### -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы.
< !--  #### -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии.
< !--  #### -->

< !--  ########################################################################################## -->
1.
Попробуйте
каждый
тег
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML and CSS
Exercise < / title >
< style >
/ *Стили
для
разнообразных
свойств * /
body
{
    font - family: 'Arial', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 20
px;
}

# paragraph1 {
font - size: 18
px;
font - weight: bold;
color:  # 0066cc;
text - decoration: underline;
}

.paragraph2
{
font - style: italic;
color:  # cc0000;
text - align: justify;
margin - bottom: 20
px;
}

.highlight
{
background - color:  # ffcc00;
padding: 5
px;
}

# container {
border: 2
px
solid  # 009900;
padding: 10
px;
margin: 10
px
0;
background - color:  # e6ffe6;
border - radius: 10
px;
}

.centered - text
{
    text - align: center;
color:  # 990099;
}
< / style >
    < / head >
        < body >
        <!-- Использование
различных
тегов
с
атрибутами
и
стилизацией -->
< p
id = "paragraph1"


class ="highlight" > Это первое предложение с различными тегами и атрибутами.< / p >

< p


class ="paragraph2" id="container" > Второе предложение выделено разными стилями с использованием CSS.< / p >

< p


class ="centered-text" > Третье предложение выровнено по центру и имеет свой цвет.< / p >

< !-- Примеры
других
тегов -->
< div
id = "container" >
< h2 > Пример
заголовка
второго
уровня < / h2 >
< ul >
< li > < strong > Список: < / strong > элемент
1 < / li >
< li > < strong > Список: < / strong > элемент
2 < / li >
< / ul >
< a
href = "https://www.example.com"
target = "_blank"
style = "color: #cc00cc; text-decoration: none;" > Ссылка
на
примерный
сайт < / a >
< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
2.
Создайте
страницу “Romario and Juluya”
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Romeo and Juliet < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
text - align: center;
}

# title {
font - family: 'Georgia', serif;
color:  # FF4500; /* Цвет - оранжевый */
font - style: italic;
font - size: 36
px;
}

# author {
font - size: 18
px;
font - style: italic;
}

hr
{
    width: 50 %;
margin: auto;
margin - top: 20
px;
border: 1
px
solid  # 000;
}

# prologue {
font - size: 24
px;
margin - top: 20
px;
}

# left-text, #right-text {
display: inline - block;
width: 45 %;
vertical - align: top;
}

.act - title
{
    font - size: 28px;
margin - top: 20
px;
}

.scene - title
{
    font - size: 20px;
font - weight: bold;
margin - top: 10
px;
}

.character
{
    font - weight: bold;
}

.dialogue
{
    text - align: left;
margin - top: 10
px;
margin - bottom: 10
px;
}

.romeo
{
    color:  # 0000FF; /* Цвет - синий */
}

.juliet
{
    color:  # FF1493; /* Цвет - розовый */
}

.sampson
{
    color:  # 008000; /* Цвет - зеленый */
}

.gregory
{
    color:  # FFD700; /* Цвет - золотой */
}
< / style >
    < / head >
        < body >
        < div
id = "title" > ROMEO
AND
JULIET < / div >
           < div
id = "author" > By
William
Shakespeare < / div >
                < hr >

                < div
id = "prologue" >
     < div
id = "left-text" >
     Two
households, both
alike in dignity,
In
fair
Verona, where
we
lay
our
scene,
From
ancient
grudge
break
to
new
mutiny,
Where
civil
blood
makes
civil
hands
unclean.
From
forth
the
fatal
loins
of
these
two
foes
A
pair
of
star - cross’d
lovers
take
their
life;
Whole
misadventured
piteous
overthrows
< / div >

< div
id = "right-text" >
Do
with their death bury their parents’ strife.
The
fearful
passage
of
their
death - mark’d
love,
And
the
continuance
of
their
parents’ rage,
Which, but
their
children’s
end, nought
could
remove,
Is
now
the
two
hours’ traffic
of
our
stage;
The
which if you
with patient ears attend,
What
here
shall
miss, our
toil
shall
strive
to
mend.
< / div >
< / div >
< hr >

< div


class ="act-title" > ACT I < / div >

< div


class ="scene-title" > SCENE I.Verona.A public place.< / div >

< div


class ="scene-title" > < i > Enter SAMPSON and GREGORY, of the house


of
CAPULET, armed
with swords and bucklers </ i > < / div >

< div
id = "left-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > Gregory, o’ my word, we’ll not carry coals.< br >

< span


class ="character gregory" > GREGORY < / span > No, for then we should be colliers.< br >

< span


class ="character sampson" > SAMPSON < / span > I mean, an we be in choler, we’ll draw.< br >

< span


class ="character gregory" > GREGORY < / span > Ay, while you live, draw your neck out


o’ the
collar. < br >
< span


class ="character sampson" > SAMPSON < / span > I strike quickly, being moved.< br >

< span


class ="character gregory" > GREGORY < / span > But thou art not quickly moved to strike.< br >

< span


class ="character sampson" > SAMPSON < / span > A dog of the house of Montague moves me.< br >

< span


class ="character gregory" > GREGORY < / span > To move is to stir; and to be valiant is to


stand: therefore, if thou art moved, thou runn’st away.< br >
< / div >

< div
id = "right-text"


class ="dialogue" >

< span


class ="character sampson" > SAMPSON < / span > A dog of that house shall move me to stand:
    I
    will
    take
    the
    wall
    of
    any
    man or maid
    of
    Montague’s. < br >

< span


class ="character gregory" > GREGORY < / span > That shows thee a weak slave; for the weakest


goes
to
the
wall. < br >
< span


class ="character sampson" > SAMPSON < / span > True; and therefore women, being the


weaker
vessels, are
ever
thrust
to
the
wall: therefore
I
will
push
Montague’s
men
from the wall, and thrust
his
maids
to
the
wall. < br >
< span


class ="character gregory" > GREGORY < / span > The quarrel is between our masters and us


their
men. < br >
< span


class ="character sampson" > SAMPSON < / span > ’Tis all one, I will show myself a tyrant: when


I
have
fought
with the men, I will be cruel with the
maids, and cut
off
their
heads. < br >
< span


class ="character gregory" > GREGORY < / span > The heads of the maids? < br >

< span


class ="character sampson" > SAMPSON < / span > Ay, the heads of the maids, or their


maidenheads;
take
it in what
sense
thou
wilt. < br >
< span


class ="character gregory" > GREGORY < / span > They must take it in sense that feel it.< br >

< / div >

< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №0.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 48px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
background - color:  # fff;
text - align: center;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > ♥ < / div >

< div


class ="card diamonds" > ♦ < / div >

< div


class ="card hearts" > ♥ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< div


class ="card hearts" > ♥ < / div >

< !-- Дополнительные
карты -->
< div


class ="card diamonds" > ♦ < / div >

< div


class ="card clubs" > ♣ < / div >

< div


class ="card spades" > ♠ < / div >

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 24px;
margin - right: 10
px;
}
< / style >
< / head >
< body >
< h1 > Игра
в
карты < / h1 >

< div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card" > &  # x1F0A1; Двойка - крести</div>

< div


class ="card" > &  # x1F0C9; Девятка - черви</div>

< div


class ="card" > &  # x1F0C2; Шестерка - черви</div>

< div


class ="card" > &  # x1F0B4; Туз - буби</div>

< div


class ="card" > &  # x1F0A9; Дама - пики</div>

< div


class ="card" > &  # x1F0AC; Валет - крести</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card" > &  # x1F0C8; Тройка - черви</div>

< div


class ="card" > &  # x1F0A1; Туз - крести</div>

< div


class ="card" > &  # x1F0AA; Десятка - крести</div>

< div


class ="card" > &  # x1F0B7; Король - буби</div>

< div


class ="card" > &  # x1F0A6; Восмёрка - пики</div>

< div


class ="card" > &  # x1F0C7; Тройка - черви</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 36px;
margin - right: 10
px;
display: inline - block;
border: 2
px
solid  # 000;
padding: 10
px;
border - radius: 8
px;
}

/ *Цвета
мастей * /
.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card diamonds" > &  # x1F0A1;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №3.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
}

h1
{
    color:  # 336699;
}

.player - cards
{
    margin - top: 20px;
}

.card
{
    font - size: 190px;
margin - right: 15
px;
display: inline - block;
border: 3
px
solid  # 000;
padding: 15
px;
border - radius: 10
px;
margin - bottom: 15
px;
}

.hearts
{color: red;}
.diamonds
{color:  # e74c3c; }
.clubs
{color:  # 2ecc71; }
.spades
{color:  # 3498db; }

/ * Дополнительные
стили
для
улучшения
внешнего
вида * /
p
{
    font - size: 18px;
margin - bottom: 10
px;
}

strong
{
    font - size: 24px;
}
< / style >
    < / head >
        < body >
        < h1 > Игра
в
карты < / h1 >

          < div


class ="player-cards" >

< p > < strong > У
игрока
1: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0A1;</div>

< div


class ="card diamonds" > &  # x1F0C9;</div>

< div


class ="card diamonds" > &  # x1F0C2;</div>

< div


class ="card clubs" > &  # x1F0B4;</div>

< div


class ="card spades" > &  # x1F0A9;</div>

< div


class ="card hearts" > &  # x1F0AC;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C4;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B3;</div>-->

< !-- < div


class ="card spades" > &  # x1F0AA;</div>-->

< / div >

< div


class ="player-cards" >

< p > < strong > У
игрока
2: < / strong > < / p >
< div


class ="card hearts" > &  # x1F0C8;</div>

< div


class ="card spades" > &  # x1F0A8;</div>

< div


class ="card hearts" > &  # x1F0AA;</div>

< div


class ="card clubs" > &  # x1F0B7;</div>

< div


class ="card spades" > &  # x1F0A6;</div>

< div


class ="card hearts" > &  # x1F0C7;</div>

< !-- Дополнительные
карты -->
< !-- < div


class ="card diamonds" > &  # x1F0C3;</div>-->

< !-- < div


class ="card clubs" > &  # x1F0B2;</div>-->

< !-- < div


class ="card diamonds" > &  # x1F0A1;</div>-->

< / div >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №4.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > Игра
в
карты < / title >
< style >
/ *Общие
стили
для
всего
документа * /
body
{
    font - family: 'Arial', sans - serif; / *Шрифт * /
                                             line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
    color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
    margin - top: 20px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
    margin - right: 15px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
transition: transform
0.3
s
ease - in -out; / *Плавное
изменение
при
наведении * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ * Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "36" >

        <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера * /
function
displayCards(playerId, cards, cardSize)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
    const cardElement = document.createElement('div');
cardElement.style.fontSize = `${cardSize}px`; / * Изменение размера карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Отображаем ранг карты и символ масти * /
cardElement.innerText = `${card.rank} ${suitsSymbols[card.suit]}`;
playerCardsElement.appendChild(cardElement);
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
displayCards('player1Cards', player1Cards, newSize);
displayCards('player2Cards', player2Cards, newSize);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
36
px * /
displayCards('player1Cards', player1Cards, 36);
displayCards('player2Cards', player2Cards, 36);
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
3.
Создайте
страницу “Card
Game”.Для
отображения
карт
используйте
спецсимволы. - вариант №5.
                        <!--  ########################################################################################## -->
                          <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Игра
в
карты < / title >
          < style >
/ *Общие
стили
для
всего
документа * /
body
{
font - family: 'Arial', sans - serif; / *Шрифт * /
line - height: 1.6; / *Межстрочный
интервал * /
margin: 20
px; / *Отступы
от
краев * /
background - color:  # f8f9fa; /* Цвет фона */
}

/ *Стили
для
заголовка
страницы * /
h1
{
color:  # 336699; /* Цвет текста */
}

/ *Стили
для
контейнеров
карт
игроков * /
.player - cards
{
margin - top: 20
px; / *Верхний
отступ * /
display: flex; / *Отображение
карт
в
строку * /
flex - wrap: wrap; / *Перенос
карт
на
следующую
строку, если
не
помещаются * /
}

/ *Общие
стили
для
карт * /
.card
{
width: 100
px; / *Ширина
карты * /
height: 150
px; / *Высота
карты * /
margin - right: 15
px; / *Правый
отступ
между
картами * /
margin - bottom: 15
px; / *Нижний
отступ
между
картами * /
padding: 15
px; / *Внутренний
отступ * /
border: 3
px
solid  # 000; /* Обводка карты */
border - radius: 10
px; / *Закругление
углов * /
background - color:  # fff; /* Цвет фона карты */
text - align: center; / *Выравнивание
текста
по
центру * /
position: relative; / *Позиционирование
элементов
внутри
карты * /
}

/ *Стили
для
масти
"черви" * /
.hearts
{color: red;}

/ *Стили
для
масти
"бубны" * /
.diamonds
{color:  # e74c3c; }

/ *Стили
для
масти
"трефы" * /
.clubs
{color:  # 2ecc71; }

/ * Стили
для
масти
"пики" * /
.spades
{color:  # 3498db; }

/ * Стили
для
значения
карты * /
.card - value
{
    font - size: 14px; / *Размер
шрифта
значения
карты * /
color:  # 555; /* Цвет текста значения карты */
position: absolute; / *Абсолютное
позиционирование
значения * /
}

/ *Стили
для
левого
верхнего
угла
карты * /
.card - value.top - left
{
    top: 10px; / *Отступ
от
верхнего
края
карты * /
left: 10
px; / *Отступ
от
левого
края
карты * /
}

/ *Стили
для
правого
нижнего
угла
карты * /
.card - value.bottom - right
{
    bottom: 10px; / *Отступ
от
нижнего
края
карты * /
right: 10
px; / *Отступ
от
правого
края
карты * /
}
< / style >
    < / head >
        < body >
        <!-- Заголовок
страницы -->
< h1 > Игра
в
карты < / h1 >

          <!-- Ползунок
для
изменения
размера
карт -->
< label
for ="cardSize" > Размер карт:< / label >
                                  < input
type = "range"
id = "cardSize"
min = "20"
max = "100"
value = "50" >

        <!-- Выпадающий
список
для
выбора
масти -->
< label
for ="suitSelect" > Масть карты:< / label >
                                    < select
id = "suitSelect" >
     < option
value = "all" > Все
масти < / option >
          < option
value = "hearts" > Черви < / option >
                             < option
value = "diamonds" > Бубны < / option >
                               < option
value = "clubs" > Трефы < / option >
                            < option
value = "spades" > Пики < / option >
                            < / select >

                                <!-- Выпадающий
список
для
выбора
ранга
карты -->
< label
for ="rankSelect" > Ранг карты:< / label >
                                   < select
id = "rankSelect" >
     < option
value = "all" > Все
ранги < / option >
          < option
value = "A" > Туз < / option >
                      < option
value = "2" > 2 < / option >
                    < option
value = "3" > 3 < / option >
                    < option
value = "4" > 4 < / option >
                    < option
value = "5" > 5 < / option >
                    < option
value = "6" > 6 < / option >
                    < option
value = "7" > 7 < / option >
                    < option
value = "8" > 8 < / option >
                    < option
value = "9" > 9 < / option >
                    < option
value = "10" > 10 < / option >
                      < option
value = "J" > Валет < / option >
                        < option
value = "Q" > Дама < / option >
                       < option
value = "K" > Король < / option >
                         < / select >

                             <!-- Контейнер
для
карт
первого
игрока -->
< div


class ="player-cards" id="player1Cards" > < / div >

< !-- Контейнер
для
карт
второго
игрока -->
< div


class ="player-cards" id="player2Cards" > < / div >

< !-- Скрипт
на
JavaScript
для
генерации, перемешивания, изменения
размера
и
отображения
карт -->
< script >
/ *Масти
и
их
символы * /
const
suits = ['clubs', 'spades', 'diamonds', 'hearts'];
const
suitsSymbols = {
    'hearts': '♥',
    'diamonds': '♦',
    'clubs': '♣',
    'spades': '♠'
};

/ *Ранги
карт * /
const
ranks = ['A', '2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K'];

/ *Функция
для
перемешивания
колоды * /
function
shuffleDeck()
{
    let
deck = [];
for (let suit of suits)
{
for (let rank of ranks) {
    deck.push({suit, rank});
}
}
return deck.sort(() = > Math.random() - 0.5);
}

/ *Функция
для
отображения
карт
в
указанном
контейнере
с
учетом
размера, масти
и
ранга * /
function
displayCards(playerId, cards, cardSize, selectedSuit, selectedRank)
{
const
playerCardsElement = document.getElementById(playerId);
playerCardsElement.innerHTML = '';
for (let card of cards) {
if ((selectedSuit === 'all' | | card.suit == = selectedSuit) & & (selectedRank ==
= 'all' | | card.rank == = selectedRank)) {
const cardElement = document.createElement('div');
cardElement.style.width = `${cardSize}px`; / * Изменение ширины карты * /
cardElement.style.height = `${cardSize * 1.5}px`; / * Изменение высоты карты * /
/ * Добавляем класс масти для стилизации цвета * /
cardElement.className = `card ${card.suit}`;
/ * Добавляем значение карты в верхний и нижний угол карты * /
cardElement.innerHTML = `
< div


class ="card-value top-left" > ${card.rank} < / div >

${suitsSymbols[card.suit]}
< div


class ="card-value bottom-right" > ${card.rank} < / div > `;


playerCardsElement.appendChild(cardElement);
}
}
}

/ *Обработчик
изменения
размера
карт
при
перемещении
ползунка * /
document.getElementById('cardSize').addEventListener('input', function()
{
    const
newSize = this.value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
масти
карты * /
document.getElementById('suitSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = this.value;
const
selectedRank = document.getElementById('rankSelect').value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Обработчик
изменения
ранга
карты * /
document.getElementById('rankSelect').addEventListener('change', function()
{
    const
newSize = document.getElementById('cardSize').value;
const
selectedSuit = document.getElementById('suitSelect').value;
const
selectedRank = this.value;
displayCards('player1Cards', player1Cards, newSize, selectedSuit, selectedRank);
displayCards('player2Cards', player2Cards, newSize, selectedSuit, selectedRank);
});

/ *Генерируем
и
перемешиваем
колоду * /
const
deck = shuffleDeck();

/ *Раздача
карт
игрокам(первому
и
второму) * /
const
player1Cards = deck.slice(0, 9);
const
player2Cards = deck.slice(9, 18);

/ *Отображение
карт
для
каждого
игрока
с
начальным
размером
50
px, мастью
"все"
и
рангом
"все" * /
displayCards('player1Cards', player1Cards, 50, 'all', 'all');
displayCards('player2Cards', player2Cards, 50, 'all', 'all');
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №0.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "en" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Пример
стилей
для
страницы(можно
дополнительно
настроить) * /
body
{
    font - family: Arial, sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
}

img
{
    max - width: 100 %;
height: auto;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Welcome
to
the
HTML
Tutorial < / h1 >
             < / header >

                 < section >
                 < article >
                 < p > This
HTML
tutorial
covers
various
HTML
elements and their
usage.Learn
how
to
create
a
basic
HTML
structure and enhance
your
web
development
skills. < / p >
            < / article >

                < article >
                < h2 > Basic
HTML
Structure < / h2 >
              < p > HTML
documents
consist
of
the
following
elements: < / p >
              < ul >
              < li > < code > & lt;!DOCTYPE
html & gt; < / code > declaration < / li >
                                      < li > < code > & lt;
html & gt; < / code > element < / li >
                                  < li > < code > & lt;
head & gt; < / code > element < / li >
                                  < li > < code > & lt;
title & gt; < / code > element < / li >
                                   < li > < code > & lt;
body & gt; < / code > element < / li >
                                  < / ul >
                                      < / article >

                                          < article >
                                          < h2 > Text
Formatting < / h2 >
               < p > Use
tags
like < code > & lt;
p & gt; < / code >, < code > & lt;
h1 & gt; < / code >, < code > & lt;
strong & gt; < / code >, < code > & lt;
em & gt; < / code >,
etc.,
for text formatting. </ p >
< / article >

< article >
< h2 > Images < / h2 >
< p > Include images using the < code > & lt;img & gt; < / code > tag.< / p >
< img src="https://placekitten.com/400/200" alt="Cute Kitten" >
< / article >

< article >
< h2 > Links < / h2 >
< p > Create hyperlinks with the < code > & lt;a & gt; < / code > tag.< / p >
< p > Visit our < a href="https://example.com" target="_blank" > example website < / a >.< / p >
< / article >
< / section >

< footer >
< p > Explore more HTML elements and features to enhance your web development skills.< / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4. Создайте страницу “HTML Tutorial”.Для выполнения используйте все теги,
которые прошли на занятии.- вариант №1.
< !--  ########################################################################################## -->
< !DOCTYPE html >
< html lang="ru" >

< head >
< meta charset="UTF-8" >
< meta name="viewport" content="width=device-width, initial-scale=1.0" >
< title > HTML Tutorial < / title >
< style >
/ * Стили для страницы * /
body {
font-family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц.Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < code >
& lt;
img
src = "image.jpg"
alt = "Описание изображения" & gt;
< / code >
    < / article >

        < article >
        < h2 > Ссылки < / h2 >
                          < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Списки < / h2 >
                          < p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
             < code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
    < / article >

        < article >
        < h2 > Видео < / h2 >
                         < p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
                             < code >
& lt;
iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/ВАШ_KОД"
frameborder = "0"
allowfullscreen & gt; & lt; / iframe & gt;
< / code >
    < / article >
        < / section >

            < footer >
            < p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
                      < / footer >

                          < / body >

                              < / html >
                                  <!--  ########################################################################################## -->
                                    <!--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №1_1.
                    <!--  ########################################################################################## -->
                      <!DOCTYPE
html >
< html
lang = "ru" >

       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > HTML
Tutorial < / title >
             < style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}
< / style >
    < / head >

        < body >

        < header >
        < h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
                    < / header >

                        < section >
                        < article >
                        < p > HTML(HyperText
Markup
Language) — это
стандартный
язык
разметки
для
создания
веб - страниц. < / p > < p > Он
используется
для
структурирования
контента
в
интернете, предоставляя
основу
для
отображения
текста,
изображений, ссылок, форм
и
мультимедийных
элементов. < / p >
               < / article >

                   < article >
                   < h2 > Основная
структура
HTML < / h2 >
         < p > HTML - документ
состоит
из
следующих
элементов: < / p >
               < ul >
               < li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
          < li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
                     < li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
               < li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
              < li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
                      < / ul >
                          < p > Вот
пример
простой
структуры
HTML: < / p >
          < code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
    < / article >

        < article >
        < h2 > Форматирование
текста < / h2 >
           < p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
            < ul >
            < li > < code > & lt;
p & gt; < / code >: Параграф < / li >
                                 < li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
                                   < li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
          < li > < code > & lt;
em & gt; < / code >: Курсив < / li >
                                < / ul >
                                    < p > Пример: < / p >
                                                      < code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Изображения < / h2 >
                               < p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
                          < img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
      < / article >

          < article >
          < h2 > Ссылки < / h2 >
                            < p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
                        < code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
    < / article >

        < article >
        < h2 > Цветной
текст < / h2 >
          < p > Применение
цветового
стиля
к
тексту: < / p >
            < code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
    < / article >

        < article >
        < h2 > Стилизованный
текст < / h2 >
          < p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
           < span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
4.
Создайте
страницу “HTML
Tutorial”.Для
выполнения
используйте
все
теги,
которые
прошли
на
занятии. - вариант №2.
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "ru" >

< head >
< meta
charset = "UTF-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< title > HTML
Tutorial < / title >
< style >
/ *Стили
для
страницы * /
body
{
    font - family: 'Arial', sans - serif;
line - height: 1.6;
margin: 20
px;
background - color:  # f8f9fa;
color:  # 333;
}

h1
{
    color:  # 007bff;
}

p
{
    margin - bottom: 20px;
color:  # 555;
}

img
{
    max - width: 100 %;
height: auto;
margin - bottom: 20
px;
}

ul
{
    list - style - type: disc;
margin - left: 20
px;
}

li
{
    margin - bottom: 10px;
color:  # 333;
}

a
{
    color:  # 28a745;
        text - decoration: none;
font - weight: bold;
}

a: hover
{
    text - decoration: underline;
}

code
{
    background - color:  # f8f9fa;
        padding: 2
px
5
px;
border: 1
px
solid  # ccc;
border - radius: 3
px;
font - family: 'Courier New', monospace;
color:  # 333;
}

.styled - text
{
    font - size: 18px;
font - weight: bold;
color:  # e44d26;
font - style: italic;
text - decoration: underline;
}

hr
{
    margin: 20px 0;
border: none;
border - top: 2
px
solid  # 007bff;
}
< / style >
< / head >

< body >

< header >
< h1 > Добро
пожаловать
в
HTML - туториал < / h1 >
< / header >

< section >
< article >
< p > < b > HTML < / b > (Hypertext Markup Language) - это
код, который
используется
для
структурирования
и
отображения
веб - страницы
и
её
контента. < / p >
< p > Например, контент
может
быть
структурирован
внутри
множества
параграфов, маркированных
списков
или
с
использованием
изображений
и
таблиц
данных. < / p >
< p > Как
видно
из
названия, этот
туториал
постарается
дать
вам
базовое
понимание
HTML
и
его
функций. < / p >
< / article >
< hr >
< p > < b > HTML < / b > не
является
языком
программирования;
это
язык
разметки, и
используется, чтобы
сообщать
вашему
браузеру, как
отображать
веб - страницы, которые
вы
посещаете. < / p >
< p > Он
может
быть
сложным
или
простым, в
зависимости
от
того, как
хочет
веб - дизайнер. < / p >
< p > HTML
состоит
из
ряда
элементов, которые
вы
используете, чтобы
вкладывать
или
оборачивать
различные
части
контента, чтобы
заставить
контент
отображаться
или
действовать
определённым
образом. < / p >
< p > Ограждающие
теги
могут
сделать
слово
или
изображение
ссылкой
на
что - то
ещё, могут
сделать
слова
курсивом, сделать
шрифт
больше
или
меньше
и
так
далее. < / p >
< hr >

< article >
< h2 > Основная
структура
HTML < / h2 >
< p > HTML - документ
состоит
из
следующих
элементов: < / p >
< ul >
< li > < code > & lt;!DOCTYPE
html & gt; < / code >: Объявление
версии
HTML. < / li >
< li > < code > & lt;
html & gt; < / code >: Корневой
элемент
HTML - страницы. < / li >
< li > < code > & lt;
head & gt; < / code >: Содержит
мета - информацию
о
документе. < / li >
< li > < code > & lt;
title & gt; < / code >: Устанавливает
заголовок
HTML - документа(отображается
во
вкладке
браузера).< / li >
< li > < code > & lt;
body & gt; < / code >: Содержит
содержимое
HTML - документа. < / li >
< / ul >
< p > Вот
пример
простой
структуры
HTML: < / p >
< code >
& lt;!DOCTYPE
html & gt; < br >
& lt;
html & gt; < br >
& nbsp; & nbsp; & lt;
head & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
title & gt;
Моя
первая
HTML - страница & lt; / title & gt; < br >
& nbsp; & nbsp; & lt; / head & gt; < br >
& nbsp; & nbsp; & lt;
body & gt; < br >
& nbsp; & nbsp; & nbsp; & nbsp; & lt;
h1 & gt;
Привет, мир! & lt; / h1 & gt; < br >
& nbsp; & nbsp; & lt; / body & gt; < br >
& lt; / html & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Форматирование
текста < / h2 >
< p > HTML
предоставляет
различные
теги
для
форматирования
текста: < / p >
< ul >
< li > < code > & lt;
p & gt; < / code >: Параграф < / li >
< li > < code > & lt;
h1 & gt; < / code > до < code > & lt;
h6 & gt; < / code >: Заголовки < / li >
< li > < code > & lt;
strong & gt; < / code >: Жирный
текст < / li >
< li > < code > & lt;
em & gt; < / code >: Курсив < / li >
< / ul >
< p > Пример
использования
тегов
форматирования
текста: < / p >
< code >
& lt;
p & gt;
Это & lt;
strong & gt;
параграф & lt; / strong & gt;
с
жирным
текстом. & lt; / p & gt; < br >
& lt;
h2 & gt;
Подзаголовок & lt; / h2 & gt; < br >
& lt;
p & gt;
Это & lt;
em & gt;
курсивизированное & lt; / em & gt;
предложение. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Изображения < / h2 >
< p > Вставка
изображений
с
использованием
тега < code > & lt;
img & gt; < / code >: < / p >
< img
src = "https://placekitten.com/800/400"
alt = "Котенок" >
< / article >

< hr >

< article >
< h2 > Ссылки < / h2 >
< p > Создание
гиперссылок
с
тегом < code > & lt;
a & gt; < / code >: < / p >
< code >
& lt;
a
href = "https://example.com"
target = "_blank" & gt;
Посетите
веб - сайт
примера & lt; / a & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Цветной
текст < / h2 >
< p > Применение
цветового
стиля
к
тексту: < / p >
< code >
& lt;
p
style = "color: #e44d26;" & gt;
Этот
текст
оранжевого
цвета. & lt; / p & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Стилизованный
текст < / h2 >
< p > Пример
изменения
шрифта, стиля, размера
и
цвета: < / p >
< span


class ="styled-text" > Этот текст имеет стилизацию: жирный, курсив, подчеркнутый, оранжевого


цвета
и
размера
18
px. < / span >
< / article >

< hr >

< article >
< h2 > Списки < / h2 >
< p > Использование
тегов < code > & lt;
ul & gt; < / code > и < code > & lt;
li & gt; < / code > для
создания
списков: < / p >
< code >
& lt;
ul & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Первый
элемент
списка & lt; / li & gt; < br >
& nbsp; & nbsp; & lt;
li & gt;
Второй
элемент
списка & lt; / li & gt; < br >
& lt; / ul & gt;
< / code >
< / article >

< hr >

< article >
< h2 > Видео < / h2 >
< p > Вставка
видео
с
помощью
тега < code > & lt;
iframe & gt; < / code >: < / p >
< iframe
width = "560"
height = "315"
src = "https://www.youtube.com/embed/dQw4w9WgXcQ"
frameborder = "0"
allowfullscreen > < / iframe >
< / article >
< / section >

< footer >
< p > Исследуйте
больше
элементов
и
возможностей
HTML
для
улучшения
своих
навыков
веб - разработки. < / p >
< / footer >

< / body >

< / html >
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->
< !--  ########################################################################################## -->

< !--  ########################################################################################## -->
< !-- HTML  # 4 - Блочная структура элементов. Свойство Display. Размеры - 29.01.2024 -->
< !-- ПРАКТИЧЕСКАЯ
РАБОТА
во
ВРЕМЯ
КЛАССНОГО
УРОКА(ПАРЫ) - 29.01
.2024 -->
< !-- ВАРИАНТ №1 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-rocket float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-refresh float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:
< !DOCTYPE
html >
< html
lang = "en" >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Заголовок
документа:

< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< title >: Заголовок
документа, отображается
во
вкладке
браузера.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Устанавливает
настройки
просмотра
для
устройств
с
различной
шириной
экрана.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
стилей:

< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

Подключаются
стили
Bootstrap
и
шрифтов
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
стилей
внутри
тега < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Здесь
определены
пользовательские
стили
для
элементов.Например, фон
body, стили
карточек.dashboard - card
и
размер
иконок
внутри
этих
карточек.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Основное
содержимое
документа:

< body >
< !-- ... -->
< / body >

Весь
контент
страницы
помещается
внутрь
тега < body >.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Контент
с
использованием
Bootstrap:

Внутри < body > размещаются
карточки, созданные
с
использованием
Bootstrap.
Каждая
карточка
находится
внутри < div


class ="card bg-... dashboard-card" > и имеет свой уникальный


цвет(bg - info, bg - success, bg - warning).
Иконки
и
текст
внутри
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- <!-- ВАРИАНТ №2 --> -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" >

< div


class ="col-md-4" >

< div


class ="card bg-info dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Заказы получены < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cart-plus float-left" > < / i > < span > 486 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные заказы < span class ="float-right" > 351 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-success dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Продукция в пути < / h6 >

< h2


class ="text-right" > < i class ="fa fa-truck float-left" > < / i > < span > 123 < / span > < / h2 >

< p


class ="m-b-0" > Завершенные доставки < span class ="float-right" > 87 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-warning dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Товары на складе < / h6 >

< h2


class ="text-right" > < i class ="fa fa-cubes float-left" > < / i > < span > 789 < / span > < / h2 >

< p


class ="m-b-0" > Товары в резерве < span class ="float-right" > 456 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-danger dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Возвраты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-undo float-left" > < / i > < span > 10 < / span > < / h2 >

< p


class ="m-b-0" > Обработанные возвраты < span class ="float-right" > 5 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-primary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Поддержка клиентов < / h6 >

< h2


class ="text-right" > < i class ="fa fa-headphones float-left" > < / i > < span > 24 / 7 < / span > < / h2 >

< p


class ="m-b-0" > Запросы < span class ="float-right" > 120 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4" >

< div


class ="card bg-secondary dashboard-card" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Новые клиенты < / h6 >

< h2


class ="text-right" > < i class ="fa fa-users float-left" > < / i > < span > 50 < / span > < / h2 >

< p


class ="m-b-0" > Клиенты на обслуживании < span class ="float-right" > 20 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< !-- Внешние
скрипты(CDN) -->
< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >
< / body >
< / html >

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< / head >

< !DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >

< meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей
в
теге < style >:

< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
< / style >

Устанавливаются
пользовательские
стили
для
элементов
страницы, таких
как
body
и.dashboard - card.
Эти
стили
определяют
отступы, фон, тени
и
другие
характеристики.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Содержимое
страницы:

< body >
< div


class ="container" >

< div


class ="row" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >
Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4" > (каждая из них занимает 1 / 3 ширины экрана).


Класс
card
определяет, что
это
карточка
Bootstrap.Классы
bg - info, bg - success, bg - warning, и
т.д.,
устанавливают
цвет
фона
для
каждой
карточки.
Иконки
и
содержимое
карточек
заполняются
данными.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.2.1.slim.min.js" > < / script >
< script
src = "https://cdnjs.cloudflare.com/ajax/libs/popper.js/1.12.9/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.0.0/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- ВАРИАНТ №3 -->

< !DOCTYPE
html >
< html
lang = "en" >
< head >
< meta
charset = "utf-8" >
< title > Градиентные
карточки
для
панели
управления < / title >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
< !-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
< !-- Стили -->
< style >
body
{
    margin - top: 20px;
background:  # FAFAFA;
}
.dashboard - card
{
    color:  # fff;
        border - radius: 5
px;
box - shadow: 0
1
px
2.94
px
0.06
px
rgba(4, 26, 55, 0.16);
border: none;
margin - bottom: 30
px;
transition: all
0.3
s
ease - in -out;
}
.dashboard - card.card - block
{
    padding: 25px;
}
.dashboard - card
i
{
    font - size: 26px;
}
/ *Дополнительные
стили
или
анимации
могут
быть
добавлены
здесь * /
< / style >
< / head >
< body >
< div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< div


class ="col-md-4 mb-3" >

< div


class ="card bg-info dashboard-card" id="card1" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Еда и Напитки < / h6 >

< i


class ="fa fa-cutlery float-left" > < / i >

< h2


class ="text-right" > < span id="value1" > 100 < / span > < / h2 >

< p


class ="m-b-0" > Выполненные заказы < span class ="float-right" > 50 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-success dashboard-card" id="card2" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Ноутбуки и ПК < / h6 >

< i


class ="fa fa-laptop float-left" > < / i >

< h2


class ="text-right" > < span id="value2" > 75 < / span > < / h2 >

< p


class ="m-b-0" > Товары в наличии < span class ="float-right" > 25 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-warning dashboard-card" id="card3" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Транспорт < / h6 >

< i


class ="fa fa-car float-left" > < / i >

< h2


class ="text-right" > < span id="value3" > 50 < / span > < / h2 >

< p


class ="m-b-0" > Автомобили в наличии < span class ="float-right" > 10 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-danger dashboard-card" id="card4" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Смартфоны и Гаджеты < / h6 >

< i


class ="fa fa-mobile float-left" > < / i >

< h2


class ="text-right" > < span id="value4" > 120 < / span > < / h2 >

< p


class ="m-b-0" > Гаджеты в наличии < span class ="float-right" > 80 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-primary dashboard-card" id="card5" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Доставка < / h6 >

< i


class ="fa fa-truck float-left" > < / i >

< h2


class ="text-right" > < span id="value5" > 60 < / span > < / h2 >

< p


class ="m-b-0" > Заказы в пути < span class ="float-right" > 400 000 < / span > < / p >

< / div >
< / div >
< / div >
< div


class ="col-md-4 mb-3" >

< div


class ="card bg-secondary dashboard-card" id="card6" >

< div


class ="card-block" >

< h6


class ="m-b-20" > Одежда и Аксессуары < / h6 >

< i


class ="fa fa-shopping-bag float-left" > < / i >

< h2


class ="text-right" > < span id="value6" > 30 < / span > < / h2 >

< p


class ="m-b-0" > Модные новинки 2024 < span class ="float-right" > 20 000 < / span > < / p >

< / div >
< / div >
< / div >
< / div >
< / div >
< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >
< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >
    < / body >
        < / html >

            <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- По
подробнее
как
в
Python -->

Объявление
HTML - документа:

< !DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "utf-8" >
          < title > Градиентные
карточки
для
панели
управления < / title >
               < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          < / head >

              <!DOCTYPE
html >: Объявление
типа
документа
HTML5.
< html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский).
< head >: Секция
заголовка
документа.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Метаинформация
и
подключение
стилей:

< meta
charset = "utf-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1" >
          <!-- Внешние
библиотеки(CDN) -->
< link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css" >
       < link
rel = "stylesheet"
href = "https://maxcdn.bootstrapcdn.com/font-awesome/4.7.0/css/font-awesome.min.css" >
       <!-- Стили -->
< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    < meta
charset = "utf-8" >: Устанавливает
кодировку
документа.
< meta
name = "viewport"
content = "width=device-width, initial-scale=1" >: Настройка
просмотра
для
устройств
с
различной
шириной
экрана.
Подключение
внешних
стилей
Bootstrap
и
Font
Awesome
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Описание
пользовательских
стилей:

< style >
/ *Стили
остаются
теми
же, что
и
в
предыдущем
примере * /
< / style >

    - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
    Содержимое
страницы:

< body >
  < div


class ="container" >

< div


class ="row" id="dashboard-cards" >

< !-- Карточки
с
информацией -->
< / div >
< / div >
< / body >

Контент
страницы
находится
внутри < body >.Карточки
размещаются
внутри
контейнера
с
классом
container
и
строки
с
классом
row.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Карточки:

Каждая
карточка
находится
внутри < div


class ="col-md-4 mb-3" > (каждая из них занимает 1 / 3 ширины экрана).


Индивидуальные
идентификаторы(id)
присваиваются
каждой
карточке(card1, card2, и
т.д.).
Иконки
и
содержимое
карточек
заполняются
данными.Обратите
внимание
на
использование
спанов
с
id
для
отображения
и
обновления
случайных
значений.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
Подключение
внешних
скриптов:

< script
src = "https://code.jquery.com/jquery-3.5.1.slim.min.js" > < / script >
< script
src = "https://cdn.jsdelivr.net/npm/@popperjs/core@2.9.2/dist/umd/popper.min.js" > < / script >
< script
src = "https://maxcdn.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js" > < / script >

Подключаются
внешние
скрипты(jQuery, Popper.js, Bootstrap.js)
с
использованием
CDN.
- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
JavaScript - скрипт:

< script >
function
getRandomInt(min, max)
{
return Math.floor(Math.random() * (max - min + 1)) + min;
}

function
updateValues()
{
// Обновление
случайных
значений
внутри
карточек
document.getElementById('value1').innerText = getRandomInt(1, 200);
document.getElementById('value2').innerText = getRandomInt(1, 200);
document.getElementById('value3').innerText = getRandomInt(1, 200);
document.getElementById('value4').innerText = getRandomInt(1, 200);
document.getElementById('value5').innerText = getRandomInt(1, 200);
document.getElementById('value6').innerText = getRandomInt(1, 200);
}

// Обновление
значений
каждые
5
секунд
setInterval(updateValues, 5000);
< / script >

    Создаются
две
функции: getRandomInt
для
получения
случайного
целого
числа
в
заданном
диапазоне
и
updateValues
для
обновления
значений
в
карточках.
setInterval(updateValues, 5000): Запуск
функции
обновления
значений
каждые
5
секунд.
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

Домашнее
задание

Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS

Домашнее
задание №4: Блочная
структура
элементов.Свойство
Display.Размеры
<!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
1.
Создать
страницу, как
на
примере: "Флексагон"
         <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !DOCTYPE
html >
< html >
< head >
< title > Флексагон < / title >
                        < meta
charset = "utf-8" >
          < link
rel = "stylesheet"
href = "style.css" >
       < style >
       body
{
font - family: Arial, Verdana, sans - serif; / *Семейство
шрифтов * /
font - size: 20
pt; / *Размер
основного
шрифта
в
пунктах * /
background - color:  # f0f0f0; /* Цвет фона веб-страницы */
color:  # 333; /* Цвет основного текста */
}
h1
{
color:  # a52a2a; /* Цвет заголовка */
font - size: 75
pt; / *Размер
шрифта
в
пунктах * /
font - family: Georgia, Times, serif; / *Семейство
шрифтов * /
font - weight: normal; / *Нормальное
начертание
текста * /
}
p
{
text - align: justify; / *Выравнивание
по
ширине * /
margin - left: 75
px; / *Отступ
слева
в
пикселах * /
margin - right: 10
px; / *Отступ
справа
в
пикселах * /
border - left: 1
px
solid  # 999; /* Параметры линии слева */
border - bottom: 1
px
solid  # 999; /* Параметры линии снизу */
padding - left: 10
px; / *Отступ
от
линии
слева
до
текста * /
padding - bottom: 10
px; / *Отступ
от
линии
снизу
до
текста * /
}
< / style >
    < / head >
        < body >
        < h1 > Флексагон < / h1 >
                             < p > < b > Флексагон
представляет
собой
бумажную
фигуру, которая
имеет
три
и
более
стороны.Поначалу
кажется, что
это
невозможно, но
вспомните
ленту
Мёбиуса, она
ведь
имеет
всего
одну
сторону, в
отличие
от
листа
бумаги,
и, тем
не
менее, реальна.Так
же
реален
и
флексагон, который
легко
сделать
и
склеить
в
домашних
условиях.Он
выглядит
как
двухсторонний
шестиугольник, но
стоит
согнуть
его
особым
образом, и
мы
увидим
третью
сторону.Легко
убедиться,
что
мы
имеем
дело
именно
с
тремя
сторонами, если
раскрасить
их
в
разные
цвета.
Перегибая
флексагон, по
очереди
будем
наблюдать
все
его
поверхности. < / b > < / p >
                         < / body >
                             < / html >
                                 <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
2.
Создать
страницу, как
на
примере: "Вот такой чай"
         <!-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !DOCTYPE
html >
< html >
< head >
< meta
charset = "UTF-8" / >
          < title > Блоки < / title >
                              < link
href = "style.css"
rel = "stylesheet"
type = "text/css" / >
       < / head >
           < style >
.text
{
margin: 20
px;
padding: 10
px;
}
.www
{
background - color:  # 569099;
color: white;
font - size: 80
px;
}
.sss
{
color: black;
background - color:  # CAD8D0;
font - size: 30
px;
}
< / style >
    < body >
    < div


class ="text www" > < b > Вот такой чай < / b > < / div >

< div


class ="text sss" >


История
о
том, как
один
человек
хотел
попить
чайку, но
по
ошибке
вместо
воды
попытался
налить
в
чайник
бензин, и
что
из
этого
получилось.
< / div >
< / body >
< / html >
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
ПРИСУЦТВУЮТ
НЕ
ВСЕ
ПРАКТИЧЕСКИЕ
И
НЕ
ВСЕ
ДОМАШНИЕ
ЗАДАНИЯ!
ОНИ
В
ОТДЕЛЬНЫХ
ФАЙЛАХ.html
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->
< !-- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
Самый
простой
Вариант:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Список
фруктов < / title >
            < style >
            body
{
background - color:  # f0f0f0; /* Светлый фон у всего списка */
}

.fruit - list
{
list - style - type: none;
padding: 10
px; / *Внешний
отступ
для
списка * /
}

.fruit - item
{
background - color:  # e0e0e0; /* Более темный фон у одного элемента списка */
padding: 10
px; / *Внутренний
отступ
для
элемента
списка * /
margin - bottom: 5
px; / *Внешний
отступ
между
элементами
списка * /
}

.fruit - item: last - child
{
background - color:  # 333; /* Самый тёмный фон только под текстом последнего элемента */
color:  # fff; /* Белый текст на тёмном фоне */
}
< / style >
    < / head >
        < body >
        < ul


class ="fruit-list" >

< li


class ="fruit-item" > Яблоко < / li >

< li


class ="fruit-item" > Груша < / li >

< li


class ="fruit-item" > Банан < / li >

< li


class ="fruit-item" > Апельсин < / li >

< !-- Можно
добавить
и
другие
фрукты
по
аналогии -->
< / ul >
< / body >
< / html >
< !--  ########################################################################################## -->

< !--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для -->
< !--    мобильных
устройств. -->
< !-- < title > Список
фруктов < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    background - color:  # f0f0f0;-->
< !--}-->

< !--.fruit - list
{-->
< !--    list - style - type: none;
-->
< !--    padding: 10
px;
-->
< !--}-->

< !--.fruit - item
{-->
< !--    background - color:  # e0e0e0;-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--}-->

< !--.fruit - item: last - child
{-->
< !--    background - color:  # 333;-->
< !--    color:  # fff;-->
< !--}-->

< !--Описание: -->
< !--background - color: Устанавливает
цвет
фона. -->
< !--.fruit - list: Класс
для
стилизации
всего
списка. -->
< !--.fruit - item: Класс
для
стилизации
каждого
элемента
списка. -->
< !--padding: Устанавливает
внутренний
отступ. -->
< !--margin - bottom: Устанавливает
внешний
отступ
между
элементами
списка. -->
< !--.fruit - item: last - child: Стилизация
последнего
элемента
списка. -->
< !--color: Устанавливает
цвет
текста. -->

< !--3.
Структура
HTML - документа: -->

< !-- < ul


class ="fruit-list" > -->

< !-- < li


class ="fruit-item" > Яблоко < / li > -->

< !-- < li


class ="fruit-item" > Груша < / li > -->

< !-- < li


class ="fruit-item" > Банан < / li > -->

< !-- < li


class ="fruit-item" > Апельсин < / li > -->

< !-- & lt;! & ndash;
Можно
добавить
и
другие
фрукты
по
аналогии & ndash; & gt;
-->
< !-- < / ul > -->

< !--Описание: -->
< !--Используется
маркированный
список( < ul >) с
классом.fruit - list. -->
< !--Каждый
элемент
списка( < li >) имеет
класс.fruit - item. -->
< !--Элементы
списка
содержат
названия
фруктов. -->
< !--Последний
элемент
списка
имеет
дополнительные
стили
из
класса.fruit - item: last - child. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
Вариант №2(Немного
"иной"
вариант)
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Список
фруктов < / title >
            < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script >
                                                    < style >
                                                    body
{
background - color:  # f0f0f0;
font - family: Arial, sans - serif;
}

.container
{
display: flex;
justify - content: space - between;
padding: 20
px;
}

.fruit - list
{
list - style - type: none;
padding: 10
px;
width: 200
px;
background - color:  # e0e0e0;
border - radius: 8
px;
margin - right: 20
px;
}

.chart - container
{
width: 60 %;
}
< / style >
    < / head >
        < body >
        < div


class ="container" >

< ul


class ="fruit-list" id="fruits" >

< !-- Список
фруктов
будет
создан
здесь -->
< / ul >

< div


class ="chart-container" >

< canvas
id = "fruitChart" > < / canvas >
< / div >
< / div >

< script >
var
fruitsData = [];

function
generateFruits()
{
    var
fruitCountElements = document.querySelectorAll(".fruit-count");
var
fruitsList = document.getElementById("fruits");
var
chartData = [];

fruitsList.innerHTML = '';
fruitsData = [];

fruitCountElements.forEach(function(element)
{
    var
fruitType = element.getAttribute("data-type");
var
count = parseInt(element.value);

var
listItem = document.createElement("li");
listItem.textContent = `${fruitType} - ${count}
`;
fruitsList.appendChild(listItem);

fruitsData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});

chartData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});
});

updateChart(chartData);
}

function
getBackgroundColor(fruitType)
{
    switch(fruitType)
{
    case
"Яблоко":
return "rgba(255, 99, 132, 0.7)";
case
"Груша":
return "rgba(255, 206, 86, 0.7)";
case
"Банан":
return "rgba(75, 192, 192, 0.7)";
case
"Апельсин":
return "rgba(54, 162, 235, 0.7)";
default:
return "rgba(0, 0, 0, 0.7)";
}
}

function
updateChart(data)
{
    var
ctx = document.getElementById("fruitChart").getContext("2d");
var
myChart = new
Chart(ctx, {
    type: "bar",
    data: {
        labels: ["Количество"],
        datasets: data,
    },
    options: {
        scales: {
            x: {
                stacked: true,
            },
            y: {
                stacked: true,
                beginAtZero: true,
            },
        },
    },
});
}
< / script >

    < div >
    < label
for ="appleCount" > Яблоко:< / label >
                               < input
type = "number"
id = "appleCount"
min = "0"
value = "25"


class ="fruit-count" data-type="Яблоко" >

< / div >

< div >
< label
for ="pearCount" > Груша:< / label >
< input
type = "number"
id = "pearCount"
min = "0"
value = "45"


class ="fruit-count" data-type="Груша" >

< / div >

< div >
< label
for ="bananaCount" > Банан:< / label >
< input
type = "number"
id = "bananaCount"
min = "0"
value = "15"


class ="fruit-count" data-type="Банан" >

< / div >

< div >
< label
for ="orangeCount" > Апельсин:< / label >
< input
type = "number"
id = "orangeCount"
min = "0"
value = "35"


class ="fruit-count" data-type="Апельсин" >

< / div >

< button
onclick = "generateFruits()" > Создать
список
и
диаграмму < / button >
< / body >
< / html >
< !--  ########################################################################################## -->

< !--Шаг
1: Общее
описание -->

< !--Этот
код
представляет
собой
веб - страницу, на
которой
пользователь
может
вводить
количество
четырех
различных -->
< !--видов
фруктов(яблоки, груши, бананы, апельсины)
в
соответствующие
поля
ввода.После
ввода
данных -->
< !--и
нажатия
кнопки, на
странице
отображается
список
введенных
фруктов
и
строится
столбчатая
диаграмма, -->
< !--иллюстрирующая
количество
каждого
фрукта. -->

< !--Шаг
2: Объявление
документа -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script > -->
< !-- < style > -->
< !-- / * ...
стили
для
body,.container,.fruit - list,.chart - container... * / -->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
...
контент
страницы... & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробнее: -->
< !--Здесь
создается
структура
HTML - документа
с
объявлением
типа
документа, -->
< !--установкой
кодировки, определением
вида
и
масштабирования, а
также
заданием
заголовка
страницы. -->
< !--Также
включается
скрипт
библиотеки
Chart.js
и
описываются
стили
для
элементов
страницы. -->

< !--Шаг
3: Стилизация
страницы -->

< !--body
{-->
< !--    background - color:  # f0f0f0;-->
< !--    font - family: Arial, sans - serif;
-->
< !--}-->

< !--.container
{-->
< !--    display: flex;
-->
< !--    justify - content: space - between;
-->
< !--    padding: 20
px;
-->
< !--}-->

< !--.fruit - list
{-->
< !-- / * ...
стили
для
списка
фруктов... * / -->
< !--}-->

< !--.chart - container
{-->
< !--    width: 60 %;
-->
< !--}-->

< !--Стилизация
страницы: -->
< !--Устанавливаются
общие
стили
для
body, определяется
структура
контейнера
с
фруктами
и
графикой, -->
< !--а
также
стили
для
списка
фруктов
и
контейнера
с
графикой. -->

< !--Шаг
4: Контент
страницы -->
< !-- < body > -->
< !-- < div


class ="container" > -->

< !-- < ul


class ="fruit-list" id="fruits" > -->

< !-- & lt;! & ndash;
Список
фруктов
будет
создан
здесь & ndash; & gt;
-->
< !-- < / ul > -->

< !-- < div


class ="chart-container" > -->

< !-- < canvas
id = "fruitChart" > < / canvas > -->
< !-- < / div > -->
< !-- < / div > -->
< !-- & lt;! & ndash;
...
ввод
данных
и
кнопка... & ndash; & gt;
-->
< !-- < / body > -->
< !--Контент
страницы: -->
< !--Определен
контейнер
для
расположения
списка
фруктов
и
графика. -->
< !--Эти
элементы
будут
созданы
и
отображены
динамически. -->

< !--Шаг
5: JavaScript - логика -->

< !-- < script > -->
< !--    var
fruitsData = [];
-->

< !--    function
generateFruits()
{-->
< !-- // ...
логика
создания
списка
фруктов
и
данных
для
графика... -->
< !--}-->

< !--    function
getBackgroundColor(fruitType)
{-->
< !-- // ...
логика
определения
цвета
для
фрукта... -->
< !--}-->

< !--    function
updateChart(data)
{-->
< !-- // ...
логика
обновления
графика... -->
< !--}-->
< !-- < / script > -->

< !--JavaScript - логика: -->
< !--В
этом
скрипте
объявляются
функции
для
генерации
списка
фруктов, определения
цвета
для
фрукта
и
обновления
графика. -->
< !--Также
создается
массив
fruitsData
для
хранения
данных
о
фруктах. -->

< !--Шаг
6: Создание
и
ввод
данных -->

< !-- < div > -->
< !-- < label
for ="appleCount" > Яблоко:< / label > -->
< !-- < input
type = "number"
id = "appleCount"
min = "0"
value = "25"


class ="fruit-count" data-type="Яблоко" > -->

< !-- < / div > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < button
onclick = "generateFruits()" > Создать
список
и
диаграмму < / button > -->

< !--Создание
и
ввод
данных: -->
< !--Здесь
предоставлены
поля
ввода
для
каждого
типа
фрукта
и
кнопка
"Создать список и диаграмму". -->
< !--По
нажатии
кнопки
вызывается
функция
generateFruits(), которая
обновляет
список
фруктов
и
строит
график. -->

< !--Шаг
7: Динамическое
обновление
списка
фруктов
и
графика -->

< !--function
generateFruits()
{-->
< !--    var
fruitCountElements = document.querySelectorAll(".fruit-count");
-->
< !--    var
fruitsList = document.getElementById("fruits");
-->
< !--    var
chartData = [];
-->

< !--    fruitsList.innerHTML = '';
-->
< !--    fruitsData = [];
-->

< !--    fruitCountElements.forEach(function(element)
{-->
< !--        var
fruitType = element.getAttribute("data-type");
-->
< !--        var
count = parseInt(element.value);
-->

< !--        var
listItem = document.createElement("li");
-->
< !--        listItem.textContent = `${fruitType} - ${count}
`;
-->
< !--        fruitsList.appendChild(listItem);
-->

< !--        fruitsData.push({-->
< !--            label: fruitType, -->
< !--            data: [count], -->
< !--            backgroundColor: getBackgroundColor(fruitType), -->
< !--});-->

< !--        chartData.push({-->
< !--            label: fruitType, -->
< !--            data: [count], -->
< !--            backgroundColor: getBackgroundColor(fruitType), -->
< !--});-->
< !--});-->

< !--    updateChart(chartData);
-->
< !--}-->

< !--function
getBackgroundColor(fruitType)
{-->
< !--    switch(fruitType)
{-->
< !--        case
"Яблоко": -->
< !--
return "rgba(255, 99, 132, 0.7)";
-->
< !--        case
"Груша": -->
< !--
return "rgba(255, 206, 86, 0.7)";
-->
< !--        case
"Банан": -->
< !--
return "rgba(75, 192, 192, 0.7)";
-->
< !--        case
"Апельсин": -->
< !--
return "rgba(54, 162, 235, 0.7)";
-->
< !--        default: -->
< !--
return "rgba(0, 0, 0, 0.7)";
-->
< !--}-->
< !--}-->

< !--function
updateChart(data)
{-->
< !--    var
ctx = document.getElementById("fruitChart").getContext("2d");
-->
< !--    var
myChart = new
Chart(ctx, {-->
< !--        type: "bar", -->
< !--        data: {-->
< !--            labels: ["Количество"], -->
< !--            datasets: data, -->
< !--}, -->
< !--        options: {-->
< !--            scales: {-->
< !--                x: {-->
< !--                    stacked: true, -->
< !--}, -->
< !--                y: {-->
< !--                    stacked: true, -->
< !--                    beginAtZero: true, -->
< !--}, -->
< !--}, -->
< !--}, -->
< !--});-->
< !--}-->

< !--Динамическое
обновление
списка
фруктов
и
графика: -->
< !--Функция
generateFruits()
собирает
данные
из
полей
ввода, обновляет
список
фруктов
и -->
< !--создает
массив
данных
для
графика.Функции
getBackgroundColor()
и
updateChart()
используются
для -->
< !--определения
цвета
фрукта
и
обновления
столбчатой
диаграммы
соответственно. -->

< !--Общий
вывод: -->
< !--Этот
код
создает
интерактивную
страницу
для
ввода
данных
о
фруктах, -->
< !--отображения
списка
фруктов
и
построения
столбчатой
диаграммы
на
основе
введенных
значений. -->
< !--Каждая
строчка
кода
выполняет
конкретную
задачу, начиная
от
объявления
документа, стилей
и
заканчивая -->
< !--динамическим
обновлением
контента
и
графика. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №3: Чуть - Чуть
красивее(Но
не
все
еще
не
так
как
в
примере
в
ДЗ)
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script >
                                                    < style >
                                                    body
{
    background - color:  # f5f5f5;
        font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
display: flex;
align - items: center;
justify - content: center;
height: 100
vh;
margin: 0;
}

.container
{
    background - color:  # fff;
        border - radius: 8
px;
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.1);
padding: 20
px;
width: 80 %;
max - width: 600
px;
animation: fadeInUp
0.5
s
ease - out;
}

@keyframes


fadeInUp
{
from

{
    opacity: 0;
transform: translateY(20
px);
}
to
{
    opacity: 1;
transform: translateY(0);
}
}

.chart - container
{
    margin - top: 20px;
}

.fruit - input
{
    display: flex;
justify - content: space - between;
align - items: center;
margin - bottom: 10
px;
}

.fruit - label
{
    width: 70px;
font - weight: bold;
}

.fruit - count
{
    width: 50px;
padding: 5
px;
border: 1
px
solid  # ddd;
border - radius: 4
px;
}

.generate - button
{
    background - color:  # 4caf50;
        color:  # fff;
padding: 10
px;
border: none;
border - radius: 4
px;
cursor: pointer;
transition: background - color
0.3
s
ease;
}

.generate - button: hover
{
    background - color:  # 45a049;
}

.fruits - list
{
    list - style - type: none;
padding: 10
px;
margin: 0;
}

.fruits - list
li
{
    margin - bottom: 5px;
display: flex;
align - items: center;
}

.fruit - emoji
{
    margin - right: 10px;
font - size: 20
px;
}
< / style >
    < / head >
        < body >
        < div


class ="container" >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="appleCount" > < span class ="fruit-emoji" > 🍎 < / span > Яблоко:<

    / label >
< input
type = "number"
id = "appleCount"
min = "0"
value = "25"


class ="fruit-count" data-type="Яблоко" >

< / div >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="pearCount" > < span class ="fruit-emoji" > 🍐 < / span > Груша:<

    / label >
< input
type = "number"
id = "pearCount"
min = "0"
value = "45"


class ="fruit-count" data-type="Груша" >

< / div >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="bananaCount" > < span class ="fruit-emoji" > 🍌 < / span > Банан:<

    / label >
< input
type = "number"
id = "bananaCount"
min = "0"
value = "15"


class ="fruit-count" data-type="Банан" >

< / div >

< div


class ="fruit-input" >

< label


class ="fruit-label" for ="orangeCount" > < span class ="fruit-emoji" > 🍊 < / span > Апельсин:<

    / label >
< input
type = "number"
id = "orangeCount"
min = "0"
value = "35"


class ="fruit-count" data-type="Апельсин" >

< / div >

< button


class ="generate-button" onclick="generateFruits()" > Создать список и диаграмму < / button >

< ul
id = "fruits-list"


class ="fruits-list" > < / ul >

< div


class ="chart-container" >

< canvas
id = "fruitChart" > < / canvas >
< / div >
< / div >

< script >
var
fruitsData = [];

function
generateFruits()
{
    var
fruitCountElements = document.querySelectorAll(".fruit-count");
var
fruitsList = document.getElementById("fruits-list");
var
chartData = [];

fruitsList.innerHTML = '';
fruitsData = [];

fruitCountElements.forEach(function(element)
{
    var
fruitType = element.getAttribute("data-type");
var
count = parseInt(element.value);

var
listItem = document.createElement("li");
listItem.innerHTML = ` < span


class ="fruit-emoji" > ${getFruitEmoji(fruitType)} < / span > ${fruitType} - ${count}`;


fruitsList.appendChild(listItem);

fruitsData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});

chartData.push({
    label: fruitType,
    data: [count],
    backgroundColor: getBackgroundColor(fruitType),
});
});

updateChart(chartData);
}

function
getFruitEmoji(fruitType)
{
    switch(fruitType)
{
    case
"Яблоко":
return "🍎";
case
"Груша":
return "🍐";
case
"Банан":
return "🍌";
case
"Апельсин":
return "🍊";
default:
return "❓";
}
}

function
getBackgroundColor(fruitType)
{
    switch(fruitType)
{
case
"Яблоко":
return "rgba(255, 99, 132, 0.7)";
case
"Груша":
return "rgba(255, 204, 0, 0.7)";
case
"Банан":
return "rgba(255, 215, 0, 0.7)";
case
"Апельсин":
return "rgba(255, 165, 0, 0.7)";
default:
return "rgba(0, 0, 0, 0.7)";
}
}

function
updateChart(data)
{
    var
ctx = document.getElementById("fruitChart").getContext("2d");
var
myChart = new
Chart(ctx, {
    type: "bar",
    data: {
        labels: ["Количество"],
        datasets: data,
    },
    options: {
        scales: {
            x: {
                stacked: true,
            },
            y: {
                stacked: true,
                beginAtZero: true,
            },
        },
    },
});
}
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--1.
Шаг №1: Объявление
документа -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script > -->
< !-- < style > -->
< !-- / * ...
стили
для
body,.container,.fruit - input, и
другие... * / -->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
...
контент
страницы... & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание
этого
кода: -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешней
библиотеки
Chart.js. -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для -->
< !--    устройств
с
различной
шириной
экрана. -->
< !-- < script
src = "https://cdn.jsdelivr.net/npm/chart.js" > < / script >: Подключение
библиотеки
Chart.js
для
создания
графиков. -->
< !-- < style >: Секция
стилей
для
определения
внешнего
вида
элементов
на
странице. -->

< !--Шаг №2: Стилизация
страницы -->

< !--Пример
выполнения
этого
шага: -->

< !--body
{-->
< !--    background - color:  # f5f5f5;-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    display: flex;
-->
< !--    align - items: center;
-->
< !--    justify - content: center;
-->
< !--    height: 100
vh;
-->
< !--    margin: 0;
-->
< !--}-->

< !--.container
{-->
< !-- / * ...
стили
для
контейнера... * / -->
< !--}-->

< !-- / * ...
стили
для
других
элементов... * / -->

< !--Подробное
описание
этого
кода: -->

< !--body: Определение
общих
стилей
для
всего
документа, таких
как
цвет
фона, шрифт
и
центрирование
содержимого. -->
< !--.container: Стили
для
основного
контейнера, в
котором
размещены
остальные
элементы
страницы. -->
< !-- @ keyframes
fadeInUp: Анимация, делающая
элементы
появляющимися
с
небольшим
подъемом
сверху
при
загрузке
страницы. -->
< !--Стили
для
различных
элементов
формы, ввода
данных
и
графика. -->

< !--Шаг №3: Контент
страницы -->

< !--Пример
выполнения
этого
шага: -->

< !-- < body > -->
< !-- < div


class ="container" > -->

< !-- < div


class ="fruit-input" > -->

< !-- & lt;! & ndash;
...
ввод
данных
для
яблок... & ndash; & gt;
-->
< !-- < / div > -->

< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->

< !-- < button


class ="generate-button" onclick="generateFruits()" > Создать список и диаграмму < / button > -->

< !-- < ul
id = "fruits-list"


class ="fruits-list" > < / ul > -->

< !-- < div


class ="chart-container" > -->

< !-- < canvas
id = "fruitChart" > < / canvas > -->
< !-- < / div > -->
< !-- < / div > -->

< !-- < script > -->
< !-- // ...
скрипт
JavaScript
для
обработки
данных... -->
< !-- < / script > -->
< !-- < / body > -->

< !--Подробное
описание
этого
кода: -->

< !--div
с
классом
container: Основной
контейнер
страницы. -->
< !--div
с
классом
fruit - input: Блок
для
ввода
данных
о
фруктах. -->
< !--button
с
классом
generate - button: Кнопка
для
генерации
списка
фруктов
и
построения
диаграммы. -->
< !--ul
с
id
fruits - list: Список, в
который
будут
добавляться
элементы
с
данными
о
фруктах. -->
< !--div
с
классом
chart - container: Контейнер
для
размещения
графика. -->

< !--Шаг №4: JavaScript - логика -->

< !--Пример
выполнения
этого
шага: -->

< !-- < script > -->
< !--    var
fruitsData = [];
-->

< !--    function
generateFruits()
{-->
< !-- // ...
логика
создания
списка
фруктов
и
данных
для
графика... -->
< !--}-->

< !--    function
getFruitEmoji(fruitType)
{-->
< !-- // ...
логика
определения
эмодзи
для
фрукта... -->
< !--}-->

< !--    function
getBackgroundColor(fruitType)
{-->
< !-- // ...
логика
определения
цвета
для
фрукта... -->
< !--}-->

< !--    function
updateChart(data)
{-->
< !-- // ...
логика
обновления
графика... -->
< !--}-->
< !-- < / script > -->

< !--Подробное
описание
этого
кода: -->

< !--var
fruitsData = [];: Объявление
массива
для
хранения
данных
о
фруктах. -->
< !--Функция
generateFruits(): Обработка
данных
ввода, создание
списка
фруктов
и
массива
данных
для
графика. -->
< !--Функции
getFruitEmoji(), getBackgroundColor(): Логика
определения
эмодзи
и
цвета
для
фруктов. -->
< !--Функция
updateChart(data): Логика
обновления
графика
на
основе
предоставленных
данных. -->

< !--Общий
вывод: -->

< !--Этот
код
создает
интерактивную
страницу
для
ввода
данных
о
фруктах, -->
< !--отображения
списка
фруктов
и
построения
столбчатой
диаграммы
на
основе
введенных
значений. -->
< !--Он
включает
в
себя
HTML - разметку, стили
CSS
для
оформления, а
также
JavaScript - логику -->
< !--для
обработки
данных
и
обновления
графика. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №4: "Новый взгляд"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < style >
          body
{
background - color:  # f5f5f5;
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
display: flex;
align - items: center;
justify - content: center;
height: 100
vh;
margin: 0;
}

ul
{
list - style - type: none;
padding: 10
px;
margin: 0;
background - color:  # fff; /* Светлый фон у всего списка */
border - radius: 8
px;
box - shadow: 0
0
10
px
rgba(0, 0, 0, 0.1);
}

li
{
margin - bottom: 10
px;
padding: 10
px;
border - radius: 4
px;
background - color:  # f0f0f0; /* Более темный фон у одного элемента списка */
transition: background - color
0.3
s
ease - in -out;
display: flex;
justify - content: space - between;
align - items: center;
}

p
{
margin: 0;
padding: 5
px;
border - radius: 4
px;
background - color:  # d9d9d9; /* Самый тёмный фон только под текстом */
}

/ *Добавим
стили
для
цветов
фруктов * /
.apple
{
background - color:  # ff6666; /* Цвет фона для яблока */
}

.banana
{
background - color:  # ffe066; /* Цвет фона для банана */
}

.orange
{
background - color:  # ffa366; /* Цвет фона для апельсина */
}

.pear
{
background - color:  # b3ff99; /* Цвет фона для груши */
}

/ *и
так
далее
для
других
фруктов * /
.kiwi
{
background - color:  # 99cc66;
}

.watermelon
{
background - color:  # ff5050;
}

.grape
{
background - color:  # 990099;
}

.strawberry
{
background - color:  # ff66b2;
}

.pineapple
{
background - color:  # ffcc66;
}

/ *Добавим
стили
для
эмодзи
фруктов * /
.emoji
{
font - size: 20
px;
margin - right: 5
px;
}

/ *Анимация * /
   @ keyframes
countUp
{
from

{
    opacity: 0;
transform: translateY(10
px);
}
to
{
    opacity: 1;
transform: translateY(0);
}
}

span
{
animation: countUp
0.5
s
ease - in -out;
display: inline - block;
}
< / style >
    < / head >
        < body >
        < ul >
        < li


class ="apple" >

< p


class ="emoji" > 🍎 < / p > Яблоко - < span id="appleCount" > < / span >

< / li >
< li


class ="banana" >

< p


class ="emoji" > 🍌 < / p > Банан - < span id="bananaCount" > < / span >

< / li >
< li


class ="orange" >

< p


class ="emoji" > 🍊 < / p > Апельсин - < span id="orangeCount" > < / span >

< / li >
< li


class ="pear" >

< p


class ="emoji" > 🍐 < / p > Груша - < span id="pearCount" > < / span >

< / li >
< li


class ="kiwi" >

< p


class ="emoji" > 🥝 < / p > Киви - < span id="kiwiCount" > < / span >

< / li >
< li


class ="watermelon" >

< p


class ="emoji" > 🍉 < / p > Арбуз - < span id="watermelonCount" > < / span >

< / li >
< li


class ="grape" >

< p


class ="emoji" > 🍇 < / p > Виноград - < span id="grapeCount" > < / span >

< / li >
< li


class ="strawberry" >

< p


class ="emoji" > 🍓 < / p > Клубника - < span id="strawberryCount" > < / span >

< / li >
< li


class ="pineapple" >

< p


class ="emoji" > 🍍 < / p > Ананас - < span id="pineappleCount" > < / span >

< / li >
< / ul >

< script >
function
getRandomCount()
{
return Math.floor(Math.random() * 100) + 1;
}

document.getElementById("appleCount").textContent = getRandomCount();
document.getElementById("bananaCount").textContent = getRandomCount();
document.getElementById("orangeCount").textContent = getRandomCount();
document.getElementById("pearCount").textContent = getRandomCount();
document.getElementById("kiwiCount").textContent = getRandomCount();
document.getElementById("watermelonCount").textContent = getRandomCount();
document.getElementById("grapeCount").textContent = getRandomCount();
document.getElementById("strawberryCount").textContent = getRandomCount();
document.getElementById("pineappleCount").textContent = getRandomCount();
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--1.
Шаг №1: Объявление
документа
и
настройка
вида
страницы -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < style > -->
< !-- / * ...
стили
для
body, ul, li, p
и
другие... * / -->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
...
контент
страницы... & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание
этого
кода: -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
стили
для
определения
внешнего
вида
элементов
на
странице. -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для -->
< !--    устройств
с
различной
шириной
экрана. -->
< !-- < style >: Секция
стилей
для
определения
внешнего
вида
элементов
страницы. -->

< !--Шаг №2: Стилизация
элементов
списка -->

< !--Пример
выполнения
этого
шага: -->

< !--ul
{-->
< !-- / * ...
стили
для
списка... * / -->
< !--}-->

< !--li
{-->
< !-- / * ...
стили
для
элементов
списка... * / -->
< !--}-->

< !--p
{-->
< !-- / * ...
стили
для
абзаца... * / -->
< !--}-->

< !-- / * ...
стили
для
цветов
фруктов... * / -->
< !-- / * ...
стили
для
эмодзи
фруктов... * / -->

< !-- @ keyframes
countUp
{-->
< !-- / * ...
стили
для
анимации... * / -->
< !--}-->

< !--span
{-->
< !-- / * ...
стили
для
элемента
span... * / -->
< !--}-->

< !--Подробное
описание
этого
кода: -->

< !--Стили
для
общего
вида
ul(списка)
и
li(элементов
списка).-->
< !--p: Стили
для
абзаца
внутри
элемента
списка. -->
< !--Стили
для
цветов
фруктов, каждый
фрукт
имеет
свой
уникальный
цвет
фона. -->
< !--Стили
для
эмодзи
фруктов, в
данном
случае, устанавливается
размер
и
отступ
справа. -->

< !--Шаг №3: Контент
страницы
и
JavaScript
для
случайных
значений -->

< !--Пример
выполнения
этого
шага: -->

< !-- < body > -->
< !-- < ul > -->
< !-- < li


class ="apple" > -->

< !-- < p


class ="emoji" > 🍎 < / p > Яблоко - < span id="appleCount" > < / span > -->

< !-- < / li > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ul > -->

< !-- < script > -->
< !--        function
getRandomCount()
{-->
< !--
return Math.floor(Math.random() * 100) + 1;
-->
< !--}-->

< !--        document.getElementById("appleCount").textContent = getRandomCount();
-->
< !-- // Аналогичные
строки
для
других
фруктов -->
< !-- < / script > -->
< !-- < / body > -->

< !--Подробное
описание
этого
кода: -->

< !--ul: Определение
списка
для
отображения
фруктов. -->
< !--li: Элемент
списка
для
каждого
фрукта
с
указанием
эмодзи, названия
и
значения. -->
< !--Функция
getRandomCount(): Генерирует
случайное
число
для
отображения
количества
фруктов. -->
< !--JavaScript - код: Устанавливает
случайные
значения
для
каждого
фрукта
в
соответствующие
элементы. -->

< !--Общий
вывод: -->
< !--Этот
код
создает
страницу, на
которой
отображаются
различные
фрукты
с
их
количеством. -->
< !--Каждый
фрукт
имеет
свой
уникальный
цвет
фона, эмодзи
и
анимацию
при
первоначальной
загрузке. -->
< !--JavaScript
используется
для
генерации
случайных
значений
количества
фруктов
при
каждой
загрузке
страницы. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №5: "По заданию с изменениями"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "2_fruit_list.css" >
       < title > Fruit
list < / title >
         < / head >
             < style >

*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # ede8e8;
}

ol
{
    width: 440px;
padding: 20
px;
list - style - position: inside;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
opacity: 0;
}
to
{
opacity: 1;
}
}

li
{
    margin: 10px 0 10px 0;
padding: 5
px
5
px
5
px
10
px;
background - color:  # c66376;
transition: background - color
0.3
s
ease - in -out;
}

li: hover
{
    background - color:  # a54257;
}

span
{
    margin - left: 8px;
padding: 5
px
15
px
5
px
15
px;
background - color:  # 9e354a;
}

.size_list
{
    margin: 0px;
}

.color_list
{
    background - color:  # ed9eae;
}

li: first - child
{
    margin - top: 0px;
}

li: last - child
{
    margin - bottom: 0px;
}
< / style >
    < body >
    < ol


class ="color_list size_list" >

< li > < span > Apple < / span > < / li >
< li > < span > Orange < / span > < / li >
< li > < span > Pineapple < / span > < / li >
< li > < span > Pear < / span > < / li >
< li > < span > Cherry < / span > < / li >
< li > < span > Banana < / span > < / li >
< li > < span > Grapes < / span > < / li >
< li > < span > Strawberry < / span > < / li >
< li > < span > Mango < / span > < / li >
< li > < span > Watermelon < / span > < / li >
< li > < span > Blueberry < / span > < / li >
< li > < span > Kiwi < / span > < / li >
< li > < span > Peach < / span > < / li >
< li > < span > Raspberry < / span > < / li >
< / ol >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--1.
Шаг №1: Настройка
и
подключение
внешних
ресурсов -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Fruit
list < / title > -->
< !-- < / head > -->

< !--Подробное
описание
этого
кода: -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"2_fruit_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования -->
< !--    для
устройств
с
различной
шириной
экрана. -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" >: Подключение
внешнего
файла
стилей. -->

< !--Шаг №2: Стилизация
страницы
и
анимация -->

< !--Пример
выполнения
этого
шага: -->

< !-- < style > -->

< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->
< !-- < / style > -->

< !--Подробное
описание
этого
кода: -->

< !--Общие
стили
для
всех
элементов
на
странице: установка
шрифта, цвета
фона
и
текста. -->
< !--Стили
для
упорядоченного
списка(ol): фиксированная
ширина, внутренний
отступ, позиция
маркера -->
< !--    списка
внутри
элемента, и
анимация
появления. -->
< !--Анимация @ keyframes
fadeIn: Плавное
появление
элементов
списка. -->
< !--Стили
для
элементов
списка(li): внешний
и
внутренний
отступы, цвет
фона
и
переход
при
наведении. -->
< !--Стили
для
вложенного
элемента
span: отступ
слева, внутренний
отступ
и
цвет
фона. -->
< !--Дополнительные
стили
для
списков
с
классами
size_list
и
color_list, а
также
для -->
< !--    первого
и
последнего
элементов
списка. -->

< !--Шаг №3: Разметка
и
отображение
данных -->

< !--Пример
выполнения
этого
шага: -->

< !-- < body > -->
< !-- < ol


class ="color_list size_list" > -->

< !-- < li > < span > Apple < / span > < / li > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание
этого
кода: -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ol


class ="color_list size_list" >: Упорядоченный


список
с
указанием
классов
для
стилизации. -->
< !-- < li > < span > Apple < / span > < / li >: Элемент
списка
с
текстом
"Apple"
и
вложенным
элементом
span. -->

< !--Как
итог: Этот
код
создает
страницу
с
стилизованным
упорядоченным
списком
фруктов. -->
< !--Каждый
фрукт
имеет
свой
цвет
фона, отличающийся
при
наведении. -->
< !--Анимация
появления
списка
при
загрузке
страницы
создает
плавный
эффект. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №6: "Попытка не пытка"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < / head > -->
< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->
< !-- < / style > -->
< !-- < body > -->
< !-- < ol


class ="color_list size_list" id="fruitList" > -->

< !-- < li > < span > Яблоко < / span > < / li > -->
< !-- < li > < span > Апельсин < / span > < / li > -->
< !-- < li > < span > Ананас < / span > < / li > -->
< !-- < li > < span > Груша < / span > < / li > -->
< !-- < li > < span > Вишня < / span > < / li > -->
< !-- < li > < span > Банан < / span > < / li > -->
< !-- < li > < span > Виноград < / span > < / li > -->
< !-- < li > < span > Клубника < / span > < / li > -->
< !-- < li > < span > Манго < / span > < / li > -->
< !-- < li > < span > Арбуз < / span > < / li > -->
< !-- < li > < span > Голубика < / span > < / li > -->
< !-- < li > < span > Киви < / span > < / li > -->
< !-- < li > < span > Персик < / span > < / li > -->
< !-- < li > < span > Малина < / span > < / li > -->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->
< !--  ########################################################################################## -->
< !-- ДЛЯ
КРАСОТЫ -->
< !--  ########################################################################################## -->
< !DOCTYPE
html >
< html
lang = "ru" >
< head >
< meta
charset = "UTF-8" >
< meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
< meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
< link
rel = "stylesheet"
href = "2_fruit_list.css" >
< title > Список
фруктов < / title >
< / head >
< style >
*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # ede8e8;
}

ol
{
    width: 440px;
padding: 20
px;
list - style - position: inside;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
    opacity: 0;
}
to
{
    opacity: 1;
}
}

li
{
    margin: 10px 0 10px 0;
padding: 5
px
5
px
5
px
10
px;
background - color:  # c66376;
transition: background - color
0.3
s
ease - in -out;
}

li: hover
{
    background - color:  # a54257;
}

span
{
    margin - left: 8px;
padding: 5
px
15
px
5
px
15
px;
background - color:  # 9e354a;
transition: background - color
0.3
s
ease - in -out;
}

.size_list
{
    margin: 0px;
}

.color_list
{
    background - color:  # ed9eae;
}

li: first - child
{
    margin - top: 0px;
}

li: last - child
{
    margin - bottom: 0px;
}
< / style >
< body >
< ol


class ="color_list size_list" id="fruitList" >

< li > < span >🍎 Яблоко < / span > < / li >
< li > < span >🍊 Апельсин < / span > < / li >
< li > < span >🍍 Ананас < / span > < / li >
< li > < span >🍐 Груша < / span > < / li >
< li > < span >🍒 Вишня < / span > < / li >
< li > < span >🍌 Банан < / span > < / li >
< li > < span >🍇 Виноград < / span > < / li >
< li > < span >🍓 Клубника < / span > < / li >
< li > < span >🥭 Манго < / span > < / li >
< li > < span >🍉 Арбуз < / span > < / li >
< li > < span >🫐 Голубика < / span > < / li >
< li > < span >🥝 Киви < / span > < / li >
< li > < span >🍑 Персик < / span > < / li >
< li > < span >🍇 Малина < / span > < / li >
< / ol >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--1.
Шаг №1: Настройка
и
подключение
стилей -->
< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "ru" >: Начало
HTML - разметки
с
указанием
языка(русский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"2_fruit_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для -->
< !--    устройств
с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
анимация -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: -->
< !--    Задание
общих
стилей
для
всех
элементов
на
странице, таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ol(упорядоченного
списка): Фиксированная
ширина, внутренний
отступ, позиция
маркера -->
< !--    списка
внутри
элемента, анимация
появления. -->
< !--Анимация @ keyframes
fadeIn: -->
< !--    Плавное
появление
элементов
списка
при
загрузке
страницы. -->
< !--Стили
для
li(элементов
списка):-->
< !--    Внешний
и
внутренний
отступы, цвет
фона
и
эффект
перехода
при
наведении. -->
< !--Стили
для
span(вложенного
элемента):-->
< !--    Отступ
слева, внутренний
отступ
и
цвет
фона, а
также
эффект
перехода
при
наведении. -->
< !--Дополнительные
стили
для
списков
с
классами
size_list
и
color_list, а
также
для
первого -->
< !--    и
последнего
элементов
списка. -->

< !--3.
Шаг №3: Разметка
и
отображение
данных -->

< !-- < body > -->
< !-- < ol


class ="color_list size_list" id="fruitList" > -->

< !-- < li > < span >🍎 Яблоко < / span > < / li > -->
< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ol


class ="color_list size_list" id="fruitList" >: Упорядоченный


список
с
указанием
классов
для -->
< !--    стилизации
и
идентификатора
"fruitList". -->
< !-- < li > < span >🍎 Яблоко < / span > < / li >: Элемент
списка
с
текстом, включая
эмодзи
для
каждого
фрукта. -->
< !--    -->
< !--Ну
и
снова
опять
двадцать
пять: -->
< !--    -->
< !--Этот
код
создает
стильную
страницу
со
списком
фруктов. -->
< !--    Упорядоченный
список
имеет
анимированный
эффект
появления, -->
< !--    каждый
элемент
списка
стилизован
с
использованием
цветов
и
переходов
при
наведении, -->
< !--    а
каждый
фрукт
дополнен
соответствующим
эмодзи. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №6: "Попытка не пытка"
            -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "ru" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "2_fruit_list.css" >
       < title > Список
фруктов < / title >
            < / head >
                < style >
*{
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # ede8e8;
}

ol
{
width: 440
px;
padding: 20
px;
list - style - position: inside;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
    opacity: 0;
}
to
{
    opacity: 1;
}
}

li
{
margin: 10
px
0
10
px
0;
padding: 5
px
5
px
5
px
10
px;
background - color:  # c66376;
transition: background - color
0.3
s
ease - in -out;
display: flex;
align - items: center;
}

li: hover
{
background - color:  # a54257;
}

span
{
margin - left: 8
px;
padding: 5
px
15
px
5
px
15
px;
background - color:  # 9e354a;
transition: background - color
0.3
s
ease - in -out;
}

.size_list
{
margin: 0
px;
}

.color_list
{
background - color:  # ed9eae;
}

li: first - child
{
margin - top: 0
px;
}

li: last - child
{
margin - bottom: 0
px;
}

.emoji
{
font - size: 20
px;
margin - right: 8
px;
display: inline - block;
animation: bounce
0.8
s
ease
infinite
alternate;
}

@keyframes


bounce
{
0 % {
    transform: translateY(0);
}
100 % {
    transform: translateY(-5px);
}
}
< / style >
    < body >
    < ol


class ="color_list size_list" id="fruitList" >

< li > < span


class ="emoji" > 🍏🍏🍏 < / span > < span > Яблоко < / span > < / li >

< li > < span


class ="emoji" > 🍊🍊 < / span > < span > Апельсин < / span > < / li >

< li > < span


class ="emoji" > 🍍🍍🍍🍍 < / span > < span > Ананас < / span > < / li >

< li > < span


class ="emoji" > 🍐🍐 < / span > < span > Груша < / span > < / li >

< li > < span


class ="emoji" > 🍒🍒🍒🍒🍒 < / span > < span > Вишня < / span > < / li >

< li > < span


class ="emoji" > 🍌 < / span > < span > Банан < / span > < / li >

< li > < span


class ="emoji" > 🍇🍇🍇 < / span > < span > Виноград < / span > < / li >

< li > < span


class ="emoji" > 🍓🍓🍓 < / span > < span > Клубника < / span > < / li >

< li > < span


class ="emoji" > 🥭🥭 < / span > < span > Манго < / span > < / li >

< li > < span


class ="emoji" > 🍉🍉🍉 < / span > < span > Арбуз < / span > < / li >

< li > < span


class ="emoji" > 🫐🫐🫐 < / span > < span > Голубика < / span > < / li >

< li > < span


class ="emoji" > 🥝🥝 < / span > < span > Киви < / span > < / li >

< li > < span


class ="emoji" > 🍑🍑 < / span > < span > Персик < / span > < / li >

< li > < span


class ="emoji" > 🍇🍇 < / span > < span > Малина < / span > < / li >

< / ol >
< / body >
< / html >
< !--  ###################################################################### -->
< !--1.
Шаг №1: Настройка
и
подключение
стилей -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "2_fruit_list.css" > -->
< !-- < title > Список
фруктов < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "ru" >: Начало
HTML - разметки
с
указанием
языка(русский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"2_fruit_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования -->
< !--    для
устройств
с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
анимация -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # ede8e8;-->
< !--}-->

< !--ol
{-->
< !--    width: 440
px;
-->
< !--    padding: 20
px;
-->
< !--    list - style - position: inside;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--li
{-->
< !--    margin: 10
px
0
10
px
0;
-->
< !--    padding: 5
px
5
px
5
px
10
px;
-->
< !--    background - color:  # c66376;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--    display: flex;
-->
< !--    align - items: center;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # a54257;-->
< !--}-->

< !--span
{-->
< !--    margin - left: 8
px;
-->
< !--    padding: 5
px
15
px
5
px
15
px;
-->
< !--    background - color:  # 9e354a;-->
< !--    transition: background - color
0.3
s
ease - in -out;
-->
< !--}-->

< !--.size_list
{-->
< !--    margin: 0
px;
-->
< !--}-->

< !--.color_list
{-->
< !--    background - color:  # ed9eae;-->
< !--}-->

< !--li: first - child
{-->
< !--    margin - top: 0
px;
-->
< !--}-->

< !--li: last - child
{-->
< !--    margin - bottom: 0
px;
-->
< !--}-->

< !--.emoji
{-->
< !--    font - size: 20
px;
-->
< !--    margin - right: 8
px;
-->
< !--    display: inline - block;
-->
< !--    animation: bounce
0.8
s
ease
infinite
alternate;
-->
< !--}-->

< !-- @ keyframes
bounce
{-->
< !--    0 % {-->
< !--        transform: translateY(0);
-->
< !--}-->
< !--    100 % {-->
< !--        transform: translateY(-5
px);-->
< !--}-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: Задание
общих
стилей
для
всех
элементов
на
странице, таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ol(упорядоченного
списка): Фиксированная
ширина, внутренний
отступ, позиция
маркера -->
< !--    списка
внутри
элемента, анимация
появления. -->
< !--Анимация @ keyframes
fadeIn: Плавное
появление
элементов
списка
при
загрузке
страницы. -->
< !--Стили
для
li(элементов
списка): Внешний
и
внутренний
отступы, цвет
фона
и
эффект
перехода -->
< !--    при
наведении, использование
flex
для
выравнивания
элементов. -->
< !--Стили
для
span(вложенного
элемента): Отступ
слева, внутренний
отступ
и
цвет -->
< !--    фона, а
также
эффект
перехода
при
наведении. -->
< !--Дополнительные
стили
для
списков
с
классами
size_list
и
color_list, а
также
для
первого -->
< !--    и
последнего
элементов
списка. -->
< !--Анимация
для.emoji
с
использованием @ keyframes: Прыжковая
анимация
для
эмодзи. -->

< !--3.
Шаг №3: Разметка
и
отображение
данных -->

< !-- < body > -->
< !-- < ol


class ="color_list size_list" id="fruitList" > -->

< !-- < li > < span


class ="emoji" > 🍏🍏🍏 < / span > < span > Яблоко < / span > < / li > -->

< !-- & lt;! & ndash;
Аналогичные
блоки
для
других
фруктов & ndash; & gt;
-->
< !-- < / ol > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ol


class ="color_list size_list" id="fruitList" >: Упорядоченный


список
с
указанием -->
< !--    классов
для
стилизации
и
идентификатора
"fruitList". -->
< !-- < li > < span


class ="emoji" > 🍏🍏🍏 < / span > < span > Яблоко < / span > < / li >: Элемент


списка
с -->
< !--    эмодзи
и
текстом
для
каждого
фрукта. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Самый
простой
Вариант:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "task_list.css" >
       < title > Список
задач < / title >
          < / head >
              < style >

*{
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 333;
}

ul
{
list - style: none;
padding: 0;
width: 300
px;
}

li
{
padding: 10
px;
margin - bottom: 5
px;
border - radius: 5
px;
}

/ *Четные
элементы
списка * /
li: nth - child(even)
{
background - color:  # d3e0ea;
}

/ *Нечетные
элементы
списка * /
li: nth - child(odd)
{
background - color:  # a8bdd4;
}

/ *Отмеченные
элементы * /
li.checked
{
background - color:  # ccc;
text - decoration: line - through;
}

/ *Неотмеченные
элементы * /
li: not (.checked)::before
{
content: "❌";
color: red;
margin - right: 5
px;
}

/ *Отмеченные
элементы * /
li.checked::before
{
content: "✔";
color: green;
margin - right: 5
px;
}

< / style >
    < body >
    < ul >
    < li


class ="checked" > Выполнить задачу 1 < / li >

< li > Выполнить
задачу
2 < / li >
< li


class ="checked" > Выполнить задачу 3 < / li >

< li > Выполнить
задачу
4 < / li >
< li > Выполнить
задачу
5 < / li >
< / ul >
< / body >
< / html >
< !--  ########################################################################################## -->
< !--1.
Шаг №1: Настройка
и
подключение
стилей -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего
файла
стилей
"task_list.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования
для
устройств -->
< !--    с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
определение
задач -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 333;-->
< !--}-->

< !--ul
{-->
< !--    list - style: none;
-->
< !--    padding: 0;
-->
< !--    width: 300
px;
-->
< !--}-->

< !--li
{-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--}-->

< !-- / * Четные
элементы
списка * / -->
< !--li: nth - child(even)
{-->
< !--    background - color:  # d3e0ea;-->
< !--}-->

< !-- / * Нечетные
элементы
списка * / -->
< !--li: nth - child(odd)
{-->
< !--    background - color:  # a8bdd4;-->
< !--}-->

< !-- / * Отмеченные
элементы * / -->
< !--li.checked
{-->
< !--    background - color:  # ccc;-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !-- / * Неотмеченные
элементы * / -->
< !--li: not (.checked)::before
{-->
< !--    content: "❌";
-->
< !--    color: red;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !-- / * Отмеченные
элементы * / -->
< !--li.checked::before
{-->
< !--    content: "✔";
-->
< !--    color: green;
-->
< !--    margin - right: 5
px;
-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: Задание
общих
стилей
для
всех
элементов
на
странице, -->
< !--    таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ul(списка): Сброс
стилей
маркера
списка
и
установка
ширины. -->
< !--Стили
для
li(элементов
списка): Определение
внешнего
и
внутреннего
отступов, -->
< !--    а
также
скругления
углов. -->
< !--Стили
для
li: nth - child(even)
и
li: nth - child(odd): Определение
цветов
фона
для
четных
и
нечетных -->
< !--    элементов
списка
соответственно. -->
< !--Стили
для
li.checked: Определение
стилей
для
отмеченных
элементов, включая
цвет
фона -->
< !--    и
зачеркивание
текста. -->
< !--Стили
для
li: not (.checked)::before
и
li.checked::before: Добавление
графической
метки(✔ или ❌) перед -->
< !--    неотмеченными
и
отмеченными
задачами. -->

< !--3.
Шаг №3: Разметка
и
отображение
задач -->

< !-- < body > -->
< !-- < ul > -->
< !-- < li


class ="checked" > Выполнить задачу 1 < / li > -->

< !-- < li > Выполнить
задачу
2 < / li > -->
< !-- < li


class ="checked" > Выполнить задачу 3 < / li > -->

< !-- < li > Выполнить
задачу
4 < / li > -->
< !-- < li > Выполнить
задачу
5 < / li > -->
< !-- < / ul > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ul >: Начало
упорядоченного
списка
задач. -->
< !-- < li


class ="checked" > Выполнить задачу 1 < / li > и аналогичные элементы: -

->
< !--    Элементы
списка
с
текстом
задач
и
классами
для
стилизации(отмеченные
и
неотмеченные).-->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Самый
простой
Вариант - Вариант
2:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "ru" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "styles.css" >
       < title > Список
задач < / title >
          < / head >
              < style >
*{
font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 1a1a1a;
}

ul
{
list - style - type: none;
padding: 0;
}

li
{
padding: 10
px;
margin: 5
px;
border - radius: 5
px;
}

li: nth - child(odd)
{
background - color:  # d9edf7; /* оттенок для четных элементов */
}

li: nth - child(even)
{
background - color:  # dff0d8; /* оттенок для нечетных элементов */
}

.checked
{
background - color:  # ccc !important; /* серый фон для отмеченных элементов */
text - decoration: line - through;
}

.delete
{
color:  # d9534f; /* красный цвет крестика */
cursor: pointer;
}

.delete: hover
{
text - decoration: underline;
}
< / style >
    < body >
    < ul
id = "taskList" >
     < li


class ="checked" > ✔️ Выполненная задача < span class ="delete" > ❌ < / span > < / li >

< li > Невыполненная
задача < span


class ="delete" > ❌ < / span > < / li >

< li


class ="checked" > ✔️ Еще одна выполненная задача < span class ="delete" > ❌ < / span > < / li >

< li > Еще
одна
невыполненная
задача < span


class ="delete" > ❌ < / span > < / li >

< / ul >

< script >
document.addEventListener('DOMContentLoaded', function()
{
    const
taskList = document.getElementById('taskList');
const
deleteButtons = document.querySelectorAll('.delete');

deleteButtons.forEach(button= > {
    button.addEventListener('click', function()
{
    const
listItem = this.parentNode;
listItem.parentNode.removeChild(listItem);
});
});
});
< / script >
< / body >
< / html >

< !--  ########################################################################################## -->

< !--1.
Шаг №1: Настройка
и
подключение
стилей -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "ru" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "styles.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Определение
типа
документа - HTML5. -->
< !-- < html
lang = "ru" >: Начало
HTML - разметки
с
указанием
языка(русский). -->
< !-- < head >: Секция
заголовка
документа, включает
мета - теги
и
подключение
внешнего -->
< !--    файла
стилей
"styles.css". -->
< !-- < meta
charset = "UTF-8" >: Установка
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Задание
совместимости -->
< !--    с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Контроль
масштабирования -->
< !--    для
устройств
с
различной
шириной
экрана. -->

< !--2.
Шаг №2: Стилизация
страницы
и
определение
задач -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 1a1a1a;-->
< !--}-->

< !--  ul
{-->
< !--    list - style - type: none;
-->
< !--    padding: 0;
-->
< !--}-->

< !--  li
{-->
< !--    padding: 10
px;
-->
< !--    margin: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--}-->

< !--  li: nth - child(odd)
{-->
< !--    background - color:  # d9edf7; /* оттенок для четных элементов */-->
< !--}-->

< !--  li: nth - child(even)
{-->
< !--    background - color:  # dff0d8; /* оттенок для нечетных элементов */-->
< !--}-->

< !--.checked
{-->
< !--    background - color:  # ccc !important; /* серый фон для отмеченных элементов */-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !--.delete
{-->
< !--    color:  # d9534f; /* красный цвет крестика */-->
< !--    cursor: pointer;
-->
< !--}-->

< !--.delete: hover
{-->
< !--    text - decoration: underline;
-->
< !--}-->
< !-- < / style > -->

< !--Общие
стили: Задание
общих
стилей
для
всех
элементов -->
< !--    на
странице, таких
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ul(списка): Сброс
стилей
маркера
списка
и -->
< !--    установка
отступов. -->
< !--Стили
для
li(элементов
списка): Определение
внешнего
и -->
< !--    внутреннего
отступов, а
также
скругления
углов. -->
< !--Стили
для
li: nth - child(odd)
и
li: nth - child(even): Определение
цветов
фона -->
< !--    для
четных
и
нечетных
элементов
списка. -->
< !--Стили
для.checked: Определение
стилей
для
отмеченных
элементов, включая -->
< !--    цвет
фона
и
зачеркивание
текста.Важность(!important) используется, -->
< !--    чтобы
переопределить
стандартные
стили. -->
< !--Стили
для.delete
и.delete: hover: Определение
стилей
для
кнопки
удаления -->
< !--    задачи, включая
цвет
и
эффект
при
наведении. -->

< !--3.
Шаг №3: Разметка
и
удаление
задач -->

< !-- < body > -->
< !-- < ul
id = "taskList" > -->
< !-- < li


class ="checked" > ✔️ Выполненная задача < span class ="delete" > ❌ < / span > < / li > -->

< !-- < li > Невыполненная
задача < span


class ="delete" > ❌ < / span > < / li > -->

< !-- < li


class ="checked" > ✔️ Еще одна выполненная задача < span class ="delete" > ❌ < / span > < / li > -->

< !-- < li > Еще
одна
невыполненная
задача < span


class ="delete" > ❌ < / span > < / li > -->

< !-- < / ul > -->

< !-- < script > -->
< !--        document.addEventListener('DOMContentLoaded', function()
{-->
< !--            const
taskList = document.getElementById('taskList');
-->
< !--            const
deleteButtons = document.querySelectorAll('.delete');
-->

< !--            deleteButtons.forEach(button= > {-->
< !--                button.addEventListener('click', function()
{-->
< !--                    const
listItem = this.parentNode;
-->
< !--                    listItem.parentNode.removeChild(listItem);
-->
< !--});-->
< !--});-->
< !--});-->
< !-- < / script > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Начало
тела
документа. -->
< !-- < ul
id = "taskList" >: Начало
упорядоченного
списка
задач
с
идентификатором
"taskList". -->
< !-- < li > элементы: Элементы
списка
с
текстом
задач
и
кнопкой
удаления( < span


class ="delete" > ❌ < / span > ).-->
< !--Скрипт для удаления задач: JavaScript - скрипт, использующий


событие
DOMContentLoaded
для -->
< !--    обработки
события
клика
на
кнопке
удаления.Когда
кнопка
нажимается, -->
< !--    соответствующий
элемент
списка
удаляется
из
DOM - структуры. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Вариант
3:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "task_list.css" >
       < title > Список
задач < / title >
          < / head >
              < style >

*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 333;
}

ul
{
    list - style: none;
padding: 0;
width: 300
px;
}

li
{
    padding: 10px;
margin - bottom: 5
px;
border - radius: 5
px;
}

/ *Четные
элементы
списка * /
li: nth - child(even)
{
    background - color:  # d3e0ea;
}

/ * Нечетные
элементы
списка * /
li: nth - child(odd)
{
    background - color:  # a8bdd4;
}

/ * Отмеченные
элементы * /
li.checked
{
    background - color:  # ccc;
        text - decoration: line - through;
}

/ *Неотмеченные
элементы * /
li: not (.checked)::before
{
    content: "❌";
color: red;
margin - right: 5
px;
}

/ *Отмеченные
элементы * /
li.checked::before
{
    content: "✔";
color: green;
margin - right: 5
px;
}

< / style >
    < body >
    < ul
id = "taskList" > < / ul >

                      < script >
// Случайные
названия
задач
const
taskNames = [
    "Подготовить отчет",
    "Заказать канцтовары",
    "Провести совещание",
    "Написать письмо клиенту",
    "Подготовить презентацию",
    "Проверить почту",
    "Организовать встречу",
    "Сделать звонок",
    "Закончить проект",
    "Подготовить план маркетинга"
];

// Генерация
случайных
задач
const
taskList = document.getElementById("taskList");

for (let i = 0; i < 10; i++)
{
    const
randomIndex = Math.floor(Math.random() * taskNames.length);
const
taskName = taskNames[randomIndex];

const
li = document.createElement("li");
li.textContent = taskName;

// Случайным
образом
отмечаем
некоторые
задачи
if (Math.random() > 0.5)
{
li.classList.add("checked");
}

taskList.appendChild(li);
}
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->
              <!--Шапка
документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->

< !-- <!DOCTYPE
html >: Это
объявление
типа
документа - HTML5. -->
< !-- < html
lang = "en" >: Начало
HTML - разметки
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа, включая
мета - теги, -->
< !--    подключение
внешнего
файла
стилей
"task_list.css"
и
настройки
визуализации
на
различных
устройствах. -->

< !--    -->
< !--    Стили: -->

< !-- < style > -->
< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 333;-->
< !--}-->

< !--ul
{-->
< !--    list - style: none;
-->
< !--    padding: 0;
-->
< !--    width: 300
px;
-->
< !--}-->

< !--li
{-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--}-->

< !--li: nth - child(even)
{-->
< !--    background - color:  # d3e0ea;-->
< !--}-->

< !--li: nth - child(odd)
{-->
< !--    background - color:  # a8bdd4;-->
< !--}-->

< !--li.checked
{-->
< !--    background - color:  # ccc;-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !--li: not (.checked)::before
{-->
< !--    content: "❌";
-->
< !--    color: red;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !--li.checked::before
{-->
< !--    content: "✔";
-->
< !--    color: green;
-->
< !--    margin - right: 5
px;
-->
< !--}-->
< !-- < / style > -->
< !--    -->
< !--Общие
стили: -->
< !--    Задают
общие
стили
для
всех
элементов
на
странице, такие
как
шрифт, цвет
фона
и
текста. -->
< !--Стили
для
ul(списка): -->
< !--    Определяют
стили
для
упорядоченного
списка, убирая
стандартные
маркеры
и
устанавливая
отступы. -->
< !--Стили
для
li(элементов
списка): -->
< !--    Задают
внешний
и
внутренний
отступы, а
также
скругление
углов. -->
< !--Стили
для
четных
и
нечетных
элементов: -->
< !--    Задают
цвет
фона
для
четных
и
нечетных
элементов
списка. -->
< !--Стили
для
отмеченных
и
неотмеченных
элементов: -->
< !--    Определяют
цвет
фона, стиль
зачеркивания
и
добавляют
графические -->
< !--    символы(✔ и ❌) для
отмеченных
и
неотмеченных
элементов
соответственно. -->

< !--    -->
< !--Тело
документа: -->

< !-- < body > -->
< !-- < ul
id = "taskList" > < / ul > -->

< !-- < script > -->
< !-- // Случайные
названия
задач -->
< !--        const
taskNames = [-->
< !--            "Подготовить отчет", -->
< !--            "Заказать канцтовары", -->
< !--            "Провести совещание", -->
< !--            "Написать письмо клиенту", -->
< !--            "Подготовить презентацию", -->
< !--            "Проверить почту", -->
< !--            "Организовать встречу", -->
< !--            "Сделать звонок", -->
< !--            "Закончить проект", -->
< !--            "Подготовить план маркетинга" -->
< !--];-->

< !-- // Генерация
случайных
задач -->
< !--        const
taskList = document.getElementById("taskList");
-->

< !--
for (let i = 0; i < 10; i++) {-->
< !--            const randomIndex = Math.floor(Math.random() * taskNames.length);-->
< !--            const taskName = taskNames[randomIndex];-->

< !--            const li = document.createElement("li");-->
< !--            li.textContent = taskName;-->

< !-- // Случайным образом отмечаем некоторые задачи-->
< !-- if (Math.random() > 0.5) {-->
< !--                li.classList.add("checked");
-->
< !--}-->

< !--            taskList.appendChild(li);
-->
< !--}-->
< !-- < / script > -->
< !-- < / body > -->
< !-- < / html > -->

< !-- < body >: Открывает
тело
документа. -->
< !-- < ul
id = "taskList" > < / ul >: Создает
пустой
упорядоченный
список
с
идентификатором
"taskList". -->

< !--JavaScript - скрипт: -->

< !-- < script > -->
< !-- // Случайные
названия
задач -->
< !--    const
taskNames = [-->
< !--        "Подготовить отчет", -->
< !--        "Заказать канцтовары", -->
< !--        "Провести совещание", -->
< !--        "Написать письмо клиенту", -->
< !--        "Подготовить презентацию", -->
< !--        "Проверить почту", -->
< !--        "Организовать встречу", -->
< !--        "Сделать звонок", -->
< !--        "Закончить проект", -->
< !--        "Подготовить план маркетинга" -->
< !--];-->

< !-- // Генерация
случайных
задач -->
< !--    const
taskList = document.getElementById("taskList");
-->

< !--
for (let i = 0; i < 10; i++) {-->
< !--        const randomIndex = Math.floor(Math.random() * taskNames.length);-->
< !--        const taskName = taskNames[randomIndex];-->

< !--        const li = document.createElement("li");-->
< !--        li.textContent = taskName;-->

< !-- // Случайным образом отмечаем некоторые задачи-->
< !-- if (Math.random() > 0.5) {-->
< !--            li.classList.add("checked");
-->
< !--}-->

< !--        taskList.appendChild(li);
-->
< !--}-->
< !-- < / script > -->

< !--Генерация
задач: -->
< !--Создает
массив
taskNames
с
возможными
именами
задач. -->
< !--Создание
элементов
списка: -->
< !--В
цикле
создается
элемент
li
для
каждой
задачи, случайным
образом
выбранной
из
taskNames. -->
< !--Отметка
выполненных
задач: -->
< !--Если
случайное
число
больше
0.5, задача
отмечается
как
выполненная, добавляя
класс
"checked". -->
< !--Добавление
элемента
в
список: -->
< !--Каждый
элемент
списка
добавляется
в
taskList. -->

< !--Итог: -->
< !--Этот
код
демонстрирует
использование
HTML, CSS
и
JavaScript
для -->
< !--создания
динамического
списка
задач
с
различными
стилями, -->
< !--включая
случайное
выделение
выполненных
задач. -->

< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Домашней - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Домашнее
задание
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Реализовать
страницу
со
списком
фруктов.
Необходимо
стилизовать
список
следующим
образом:
• Светлый
фон
у
всего
списка
• Более
темный
фон
у
одного
элемента
списка
• Самый
тёмный
фон
только
под
текстом
• Добавить
внешние
и
внутренние
отступы

2.
Реализовать
HTML - страницу
со
списком
задач.
• Чётные
и
нечётные
элементы
списка
имеют
разные
оттенки
• Отмеченные
элементы
имеют
серый
фон.Иконку
с
галочкой
и
зачёркнутый
текст
• Неотмеченные
без
галочки
• Крестик
красный
отобразить
с
помощью
спецсимвола.

Иконки
ко
второму
заданию
находятся
в
архиве
Homework
в
папке
Материалы
к
заданию
Скинуть
итоговые
задания
в
архиве!
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
Вариант
3:
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "task_list.css" >
       < title > Список
задач < / title >
          < / head >
              < style >

*{
    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
background - color:  # f5f5f5;
color:  # 333;
}

ul
{
    list - style: none;
padding: 0;
width: 300
px;
opacity: 0;
animation: fadeIn
1
s
ease - in -out
forwards;
}

@keyframes


fadeIn
{
from

{
    opacity: 0;
}
to
{
    opacity: 1;
}
}

li
{
    padding: 10px;
margin - bottom: 5
px;
border - radius: 5
px;
opacity: 0;
animation: slideIn
0.5
s
ease - in -out
forwards;
}

@keyframes


slideIn
{
from

{
    opacity: 0;
transform: translateY(-20
px);
}
to
{
    opacity: 1;
transform: translateY(0);
}
}

/ *Четные
элементы
списка * /
li: nth - child(even)
{
    background - color:  # d3e0ea;
}

/ * Нечетные
элементы
списка * /
li: nth - child(odd)
{
    background - color:  # a8bdd4;
}

/ * Отмеченные
элементы * /
li.checked
{
    background - color:  # ccc;
        text - decoration: line - through;
}

/ *Неотмеченные
элементы * /
li: not (.checked)::before
{
    content: "❌";
color: red;
margin - right: 5
px;
}

/ *Отмеченные
элементы * /
li.checked::before
{
    content: "✔";
color: green;
margin - right: 5
px;
}

< / style >
    < body >
    < ul
id = "taskList" > < / ul >

                      < script >
// Случайные
названия
задач
const
taskNames = [
    "Подготовить отчет",
    "Заказать канцтовары",
    "Провести совещание",
    "Написать письмо клиенту",
    "Подготовить презентацию",
    "Проверить почту",
    "Организовать встречу",
    "Сделать звонок",
    "Закончить проект",
    "Подготовить план маркетинга"
];

// Генерация
случайных
задач
const
taskList = document.getElementById("taskList");

for (let i = 0; i < 10; i++)
{
    const
randomIndex = Math.floor(Math.random() * taskNames.length);
const
taskName = taskNames[randomIndex];

const
li = document.createElement("li");
li.textContent = taskName;

// Случайным
образом
отмечаем
некоторые
задачи
if (Math.random() > 0.5)
{
    li.classList.add("checked");
}

taskList.appendChild(li);
}
< / script >
    < / body >
        < / html >
            <!--  ########################################################################################## -->

              <!--Шаг №1: HTML - структура
документа -->

< !--Пример
выполнения
этого
шага: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" > -->
< !-- < title > Список
задач < / title > -->
< !-- < / head > -->
< !-- < body > -->
< !-- < ul
id = "taskList" > < / ul > -->

< !-- < script > -->
< !-- // JavaScript - код
будет
вставлен
здесь -->
< !-- < / script > -->
< !-- < / body > -->
< !-- < / html > -->

< !--Подробное
описание: -->

< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Начало
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
заголовка
документа.Здесь
подключены
мета - теги, стили
и
указан
заголовок
страницы. -->
< !-- < meta
charset = "UTF-8" >: Указание
кодировки
документа. -->
< !-- < meta
http - equiv = "X-UA-Compatible"
content = "IE=edge" >: Настройка
совместимости
с
Internet
Explorer. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования -->
< !--    для
мобильных
устройств. -->
< !-- < link
rel = "stylesheet"
href = "task_list.css" >: Подключение
внешних
стилей
из
файла
task_list.css. -->
< !-- < title > Список
задач < / title >: Заголовок
страницы. -->
< !-- < body >: Открывающий
тег
тела
документа, где
будет
размещен
контент
страницы. -->
< !-- < ul
id = "taskList" > < / ul >: Пустой
список
задач
с
идентификатором
taskList. -->
< !-- < script >: Блок
JavaScript
для
динамического
создания
задач.Здесь
он
пока
пуст, -->
< !-- так
как
будет
заполнен
на
следующих
шагах. -->
< !-- < / body > и < / html >: Закрывающие
теги
для
тела
и
HTML - документа. -->

< !--Шаг №2: CSS - стили
для
общих
элементов -->

< !--Пример
выполнения
этого
шага: -->

< !-- * {-->
< !--    font - family: 'Segoe UI', Tahoma, Geneva, Verdana, sans - serif;
-->
< !--    background - color:  # f5f5f5;-->
< !--    color:  # 333;-->
< !--}-->

< !--Подробное
описание: -->

< !-- *: Устанавливает
стили
для
всех
элементов
на
странице. -->
< !--font - family: Задает
шрифт
текста. -->
< !--background - color: Устанавливает
цвет
фона. -->
< !--color: Задает
цвет
текста. -->

< !--Шаг №3: Стили
для
списка
задач -->

< !--Пример
выполнения
этого
шага: -->

< !--ul
{-->
< !--    list - style: none;
-->
< !--    padding: 0;
-->
< !--    width: 300
px;
-->
< !--    opacity: 0;
-->
< !--    animation: fadeIn
1
s
ease - in -out
forwards;
-->
< !--}-->

< !--Подробное
описание: -->

< !--ul: Стили
для
списка. -->
< !--list - style: Убирает
маркеры
списка. -->
< !--padding: Устанавливает
отступы
внутри
списка. -->
< !--width: Задает
ширину
списка. -->
< !--opacity: Устанавливает
начальную
прозрачность
списка. -->
< !--animation: Применяет
анимацию
fadeIn
для
плавного
появления
списка. -->

< !--Шаг №4: Анимации -->

< !--Пример
выполнения
этого
шага: -->

< !-- @ keyframes
fadeIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--}-->
< !--}-->

< !--Подробное
описание: -->

< !-- @ keyframes
fadeIn: Объявление
анимации
с
названием
fadeIn. -->
< !--
from: Начальное
состояние
анимации(прозрачность
0).-->
< !--to: Конечное
состояние
анимации(прозрачность
1).-->

< !--Шаг №5: Стили
для
задач -->

< !--Пример
выполнения
этого
шага: -->

< !--li
{-->
< !--    padding: 10
px;
-->
< !--    margin - bottom: 5
px;
-->
< !--    border - radius: 5
px;
-->
< !--    opacity: 0;
-->
< !--    animation: slideIn
0.5
s
ease - in -out
forwards;
-->
< !--}-->

< !--Подробное
описание: -->

< !--li: Стили
для
элементов
списка
задач. -->
< !--padding: Задает
внутренний
отступ
для
задач. -->
< !--margin - bottom: Задает
отступ
снизу
между
задачами. -->
< !--border - radius: Округляет
углы
задач. -->
< !--opacity: Устанавливает
начальную
прозрачность
задач. -->
< !--animation: Применяет
анимацию
slideIn
для
плавного
появления
задач. -->

< !--Шаг №6: Анимация
появления
задач -->

< !--Пример
выполнения
этого
шага: -->

< !-- @ keyframes
slideIn
{-->
< !--
from

{-->
< !--        opacity: 0;
-->
< !--        transform: translateY(-20
px);-->
< !--}-->
< !--    to
{-->
< !--        opacity: 1;
-->
< !--        transform: translateY(0);
-->
< !--}-->
< !--}-->

< !--Подробное
описание: -->

< !-- @ keyframes
slideIn: Объявление
анимации
с
названием
slideIn. -->
< !--
from: Начальное
состояние
анимации(прозрачность
0, смещение
вверх
на
20
пикселей).-->
< !--to: Конечное
состояние
анимации(прозрачность
1, без
смещения).-->

< !--Шаг №7: Стили
для
четных
и
нечетных
элементов -->

< !--Пример
выполнения
этого
шага: -->

< !--li: nth - child(even)
{-->
< !--    background - color:  # d3e0ea;-->
< !--}-->

< !--li: nth - child(odd)
{-->
< !--    background - color:  # a8bdd4;-->
< !--}-->

< !--Подробное
описание: -->

< !--li: nth - child(even): Стили
для
четных
элементов
списка(задается
цвет
фона).-->
< !--li: nth - child(odd): Стили
для
нечетных
элементов
списка(задается
другой
цвет
фона).-->

< !--Шаг №8: Стили
для
отмеченных
и
неотмеченных
элементов -->

< !--Пример
выполнения
этого
шага: -->

< !--li.checked
{-->
< !--    background - color:  # ccc;-->
< !--    text - decoration: line - through;
-->
< !--}-->

< !--li: not (.checked)::before
{-->
< !--    content: "❌";
-->
< !--    color: red;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !--li.checked::before
{-->
< !--    content: "✔";
-->
< !--    color: green;
-->
< !--    margin - right: 5
px;
-->
< !--}-->

< !--Подробное
описание: -->

< !--li.checked: Стили
для
отмеченных
задач(задается
цвет
фона
и
зачеркивание
текста).-->
< !--li: not (.checked)::before: Стили
для
неотмеченных
задач(вставляется
символ
"❌"
перед
текстом).-->
< !--li.checked::before: Стили
для
отмеченных
задач(вставляется
символ
"✔"
перед
текстом).-->

< !--Шаг №9: JavaScript
для
генерации
задач -->

< !--Пример
выполнения
этого
шага: -->

< !--const
taskNames = [-->
< !--    "Подготовить отчет", -->
< !--    "Заказать канцтовары", -->
< !--    "Провести совещание", -->
< !--    "Написать письмо клиенту", -->
< !--    "Подготовить презентацию", -->
< !--    "Проверить почту", -->
< !--    "Организовать встречу", -->
< !--    "Сделать звонок", -->
< !--    "Закончить проект", -->
< !--    "Подготовить план маркетинга" -->
< !--];-->

< !--const
taskList = document.getElementById("taskList");
-->

< !--
for (let i = 0; i < 10; i++) {-->
< !--    const randomIndex = Math.floor(Math.random() * taskNames.length);-->
< !--    const taskName = taskNames[randomIndex];-->

< !--    const li = document.createElement("li");-->
< !--    li.textContent = taskName;-->

< !-- if (Math.random() > 0.5) {-->
< !--        li.classList.add("checked");-->
< !--}-->

< !--    taskList.appendChild(li);-->
< !--}-->

< !--Подробное описание:-->

< !--const
taskNames: Массив
строк
с
названиями
задач. -->
< !--const
taskList = document.getElementById("taskList"): Получение
ссылки
на
элемент
списка
задач
по
его
идентификатору. -->
< !--
for (let i = 0; i < 10; i++): Цикл
для
создания
10
задач. -->
< !--const
randomIndex = Math.floor(Math.random() * taskNames.length): Генерация
случайного
индекса
из
массива
taskNames. -->
< !--const
taskName = taskNames[randomIndex]: Получение
случайного
названия
задачи. -->
< !--const
li = document.createElement("li"): Создание
элемента
списка( < li >).-->
< !--li.textContent = taskName: Установка
текста
задачи. -->
< !-- if (Math.random() > 0.5)
{li.classList.add("checked");}: Случайное
добавление
класса
"checked"
для
отмеченных
задач. -->
< !--taskList.appendChild(li): Добавление
созданной
задачи
в
список. -->

< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 1)
-->
< !--  ###################################################################### -->
   <!--
Думаю
что
одного - ВАРИАНТА
будет
достаточно
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < link
rel = "stylesheet"
href = "styles.css" >
       < style >
:root
{
    --season - font - size: 35px;
--season - font - color:  # 333;
--month - font - size: 16
px;
--month - font - color:  # 555;
--month - icon - size: 50
px;
--month - icon - margin: 2, 5
px;
}

ul
{
    list - style - type: none;
margin: 0;
padding: 0;
}

li
{
    margin - bottom: 10px;
font - size: var(--season - font - size);
color: var(--season - font - color);
}

.month - icon
{
    font - size: var(--month - icon - size);
margin - right: var(--month - icon - margin);
}

ul
ul
{
    margin - left: 25px;
}

ul
ul
ul
{
    margin - left: 50px;
}
< / style >
    < / head >
        < body >
        < ul >
        < li >
        Времена
года
< ul >
< li >
Зима
< ul >
< li > < span


class ="month-icon" > 🎄 < / span > Декабрь < / li >

< li > < span


class ="month-icon" > ☃️ < / span > Январь < / li >

< li > < span


class ="month-icon" > ❄️ < / span > Февраль < / li >

< / ul >
< / li >
< li >
Весна
< ul >
< li > < span


class ="month-icon" > 💐 < / span > Март < / li >

< li > < span


class ="month-icon" > 🌷 < / span > Апрель < / li >

< li > < span


class ="month-icon" > 🌺 < / span > Май < / li >

< / ul >
< / li >
< li >
Лето
< ul >
< li > < span


class ="month-icon" > 🍉 < / span > Июнь < / li >

< li > < span


class ="month-icon" > 🌴 < / span > Июль < / li >

< li > < span


class ="month-icon" > 🎣 < / span > Август < / li >

< / ul >
< / li >
< li >
Осень
< ul >
< li > < span


class ="month-icon" > 🍃 < / span > Сентябрь < / li >

< li > < span


class ="month-icon" > 🍁 < / span > Октябрь < / li >

< li > < span


class ="month-icon" > 🍂 < / span > Ноябрь < / li >

< / ul >
< / li >
< / ul >
< / li >
< / ul >
< / body >
< / html >
< !--  ########################################################################################## -->

< !--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < link
rel = "stylesheet"
href = "styles.css" > -->
< !-- < style > -->
< !-- & lt;! & ndash;
CSS - стили
встроены
непосредственно
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < link
rel = "stylesheet"
href = "styles.css" >: Подключение
внешнего
файла
стилей(styles.css). -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Объявление
переменных
CSS(в
теге < style >):-->

< !--: root
{-->
< !-- &  # 45;&#45;season-font-size: 35px;-->
   <!-- &  # 45;&#45;season-font-color: #333;-->
     <!-- &  # 45;&#45;month-font-size: 16px;-->
       <!-- &  # 45;&#45;month-font-color: #555;-->
         <!-- &  # 45;&#45;month-icon-size: 50px;-->
           <!-- &  # 45;&#45;month-icon-margin: 2.5px;-->
             <!--}-->

< !--Описание: -->
< !--: root: Псевдокласс, представляющий
корневой
элемент
документа, используется
для
объявления
пользовательских
CSS - переменных. -->
< !--Переменные, устанавливающие
размеры
шрифта, цвета
и
размеры
иконок
для
времен
года
и
месяцев. -->

< !--3.
Стили
CSS: -->

< !--ul
{-->
< !--    list - style - type: none;
-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--li
{-->
< !--    margin - bottom: 10
px;
-->
< !--    font - size: var( &  # 45;&#45;season-font-size);-->
                      <!--    color: var( &  # 45;&#45;season-font-color);-->
                                     <!--}-->

< !--.month - icon
{-->
< !--    font - size: var( &  # 45;&#45;month-icon-size);-->
                      <!--    margin - right: var( &  # 45;&#45;month-icon-margin);-->
                                              <!--}-->

< !--ul
ul
{-->
< !--    margin - left: 25
px;
-->
< !--}-->

< !--ul
ul
ul
{-->
< !--    margin - left: 50
px;
-->
< !--}-->

< !--Описание: -->
< !--Удаление
маркеров
списка
и
отступов. -->
< !--Настройка
отступов
и
размеров
шрифта
для
элементов
списка. -->
< !--Определение
стилей
для
класса.month - icon, который
отвечает
за
размер
иконок
месяцев. -->
< !--Настройка
вложенных
списков
с
различными
отступами. -->

< !--4.
Структура
HTML - документа: -->
< !-- < ul > -->
< !-- < li > -->
< !--        Времена
года -->
< !-- < ul > -->
< !-- & lt;! & ndash;
Вложенные
списки
для
каждого
времени
года & ndash; & gt;
-->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->

< !--Описание: -->
< !-- < ul >: Начало
списка
без
маркеров. -->
< !-- < li >: Элемент
списка
для
времен
года. -->
< !--Вложенные < ul > и < li > для
каждого
времени
года
и
их
месяцев. -->

< !--5.
Иконки
месяцев
в
эмодзи: -->
< !-- < li > < span


class ="month-icon" > 🎄 < / span > Декабрь < / li > -->

< !-- < li > < span


class ="month-icon" > ☃️ < / span > Январь < / li > -->

< !-- & lt;! & ndash;
...
остальные
месяцы... & ndash; & gt;
-->

< !--Описание: -->
< !-- < span


class ="month-icon" >: Элемент - контейнер


для
применения
стилей
к
иконке
месяца. -->
< !--Иконки
месяцев
представлены
символами
эмодзи. -->
< !--  ############################################################################################################## -->

< !--  ############################################################################################################## -->

< !--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
< !--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №1 - Чем
проще, тем
лучше.
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
font - family: Arial, sans - serif;
}

h1
{
font - weight: bold;
}

h2
{
margin - left: 20
px;
}

h3
{
margin - left: 40
px;
}

h4
{
margin - left: 60
px;
}

h5
{
margin - left: 80
px;
}
< / style >
    < / head >
        < body >
        < h1 > Contents < / h1 >

                            < h2 > I.History < / h2 >
                                                 < h3 > 1.
Development < / h3 >
                < h3 > 2.
HTML
versions
timeline < / h3 >
             < h4 > a.HTML
draft
version
timeline < / h4 >
             < h4 > b.XHTML
versions < / h4 >

             < h2 > II.Markup < / h2 >
                                  < h3 > 1.
Elements < / h3 >
             < h4 > a.Element
examples < / h4 >
             < h4 > b.Attributes < / h4 >
                                     < h3 > 2.
Character and entity
references < / h3 >
               < h3 > 3.
Data
types < / h3 >
          < h3 > 4.
Document
type declaration < / h3 >

                     < h2 > III.Semantic
HTML < / h2 >

         < h2 > IV.Delivery < / h2 >
                                < h3 > 1.
HTTP < / h3 >
         < / body >
             < / html >
                 <!--  ############################################################################################################## -->

                   <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: Arial, sans - serif;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--}-->

< !--h2
{-->
< !--    margin - left: 20
px;
-->
< !--}-->

< !--h3
{-->
< !--    margin - left: 40
px;
-->
< !--}-->

< !--h4
{-->
< !--    margin - left: 60
px;
-->
< !--}-->

< !--h5
{-->
< !--    margin - left: 80
px;
-->
< !--}-->

< !--Описание: -->
< !--font - family: Arial, sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--Настройка
отступов
для
заголовков
разного
уровня
с
использованием
margin - left. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < h2 > I.History < / h2 > -->
< !-- < h3 > 1.
Development < / h3 > -->
< !-- < h3 > 2.
HTML
versions
timeline < / h3 > -->
< !-- < h4 > a.HTML
draft
version
timeline < / h4 > -->
< !-- < h4 > b.XHTML
versions < / h4 > -->

< !-- < h2 > II.Markup < / h2 > -->
< !-- < h3 > 1.
Elements < / h3 > -->
< !-- < h4 > a.Element
examples < / h4 > -->
< !-- < h4 > b.Attributes < / h4 > -->
< !-- < h3 > 2.
Character and entity
references < / h3 > -->
< !-- < h3 > 3.
Data
types < / h3 > -->
< !-- < h3 > 4.
Document
type declaration < / h3 > -->

< !-- < h2 > III.Semantic
HTML < / h2 > -->

< !-- < h2 > IV.Delivery < / h2 > -->
< !-- < h3 > 1.
HTTP < / h3 > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >, < h2 >, < h3 >, < h4 >, < h5 >), чтобы
структурировать
содержание
страницы. -->
< !--Каждый
заголовок
имеет
отступ, который
указан
в
стилях
CSS. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--
ВАРИАНТ №2 - Уже
лучше.
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
    font - family: 'Roboto', sans - serif;
background - color:  # f8f8f8;
color:  # 333;
margin: 0;
padding: 0;
}

h1
{
    font - weight: bold;
color:  # 0066cc;
background - color:  # f2f2f2;
padding: 10
px;
margin: 0;
}

h2, h3, h4, h5
{
    margin - left: 20px;
}

h2
{
    color:  # 009900;
        font - weight: bold;
}

h3
{
    color:  # cc3300;
        font - weight: normal;
}

h4
{
    color:  # 990099;
        font - weight: normal;
}

h5
{
    color:  # ff6600;
        font - weight: normal;
}

h4, h5
{
    margin - left: 40px;
}
< / style >
    < / head >
        < body >
        < h1 > Contents < / h1 >

                            < h2 > I.History < / h2 >
                                                 < h3 > 1.
Development < / h3 >
                < h3 > 2.
HTML
versions
timeline < / h3 >
             < h4 > a.HTML
draft
version
timeline < / h4 >
             < h4 > b.XHTML
versions < / h4 >

             < h2 > II.Markup < / h2 >
                                  < h3 > 1.
Elements < / h3 >
             < h4 > a.Element
examples < / h4 >
             < h4 > b.Attributes < / h4 >
                                     < h3 > 2.
Character and entity
references < / h3 >
               < h3 > 3.
Data
types < / h3 >
          < h3 > 4.
Document
type declaration < / h3 >

                     < h2 > III.Semantic
HTML < / h2 >

         < h2 > IV.Delivery < / h2 >
                                < h3 > 1.
HTTP < / h3 >
         < / body >
             < / html >
                 <!--  ############################################################################################################## -->

                   <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: 'Roboto', sans - serif;
-->
< !--    background - color:  # f8f8f8;-->
< !--    color:  # 333;-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--    color:  # 0066cc;-->
< !--    background - color:  # f2f2f2;-->
< !--    padding: 10
px;
-->
< !--    margin: 0;
-->
< !--}-->

< !--h2, h3, h4, h5
{-->
< !--    margin - left: 20
px;
-->
< !--}-->

< !--h2
{-->
< !--    color:  # 009900;-->
< !--    font - weight: bold;
-->
< !--}-->

< !--h3
{-->
< !--    color:  # cc3300;-->
< !--    font - weight: normal;
-->
< !--}-->

< !--h4
{-->
< !--    color:  # 990099;-->
< !--    font - weight: normal;
-->
< !--}-->

< !--h5
{-->
< !--    color:  # ff6600;-->
< !--    font - weight: normal;
-->
< !--}-->

< !--h4, h5
{-->
< !--    margin - left: 40
px;
-->
< !--}-->

< !--Описание: -->
< !--font - family: 'Roboto', sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--background - color:  # f8f8f8;: Задание цвета фона.-->
< !--color:  # 333;: Установка цвета текста.-->
< !--margin: 0;
padding: 0;: Убирают
отступы
и
внутренний
отступ
для
тела
документа. -->
< !--Стили
для
заголовков
разных
уровней
с
различными
цветами
и
отступами. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < h2 > I.History < / h2 > -->
< !-- < h3 > 1.
Development < / h3 > -->
< !-- < h3 > 2.
HTML
versions
timeline < / h3 > -->
< !-- < h4 > a.HTML
draft
version
timeline < / h4 > -->
< !-- < h4 > b.XHTML
versions < / h4 > -->

< !-- < h2 > II.Markup < / h2 > -->
< !-- < h3 > 1.
Elements < / h3 > -->
< !-- < h4 > a.Element
examples < / h4 > -->
< !-- < h4 > b.Attributes < / h4 > -->
< !-- < h3 > 2.
Character and entity
references < / h3 > -->
< !-- < h3 > 3.
Data
types < / h3 > -->
< !-- < h3 > 4.
Document
type declaration < / h3 > -->

< !-- < h2 > III.Semantic
HTML < / h2 > -->

< !-- < h2 > IV.Delivery < / h2 > -->
< !-- < h3 > 1.
HTTP < / h3 > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >, < h2 >, < h3 >, < h4 >, < h5 >), чтобы
структурировать
содержание
страницы. -->
< !--Каждый
заголовок
имеет
определенные
стили, включая
цвет
текста
и
фона, а
также
отступы, указанные
в
стилях
CSS. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №3 - Красота.
         -->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
    font - family: 'Arial', sans - serif;
background - color:  # 282c35;
color:  # fff;
margin: 0;
padding: 0;
display: flex;
justify - content: center;
align - items: center;
height: 100
vh;
}

h1
{
    font - weight: bold;
color:  # 61dafb;
text - shadow: 2
px
2
px
4
px  # 333;
animation: rainbowAnimation
5
s
infinite;
}

ul
{
    list - style - type: none;
margin: 0;
padding: 0;
}

li
{
    margin - left: 20px;
margin - bottom: 10
px;
padding: 8
px;
background - color:  # 61dafb;
border - radius: 5
px;
transition: background - color
0.3
s
ease;
}

li: hover
{
    background - color:  # 4fa3d1;
}

@ keyframes
rainbowAnimation
{
    0 % {color:  # ff0000; }
             25 % {color:  # ff9900; }
50 % {color:  # ffff00; }
          75 % {color:  # 33cc33; }
                    100 % {color:  # 3366ff; }
                           }
                    < / style >
          < / head >
< body >
< h1 > Contents < / h1 >

                    < ul >
                    < li > I.History
                    < ul >
                    < li > 1.
Development < / li >
                < li > 2.
HTML
versions
timeline
< ul >
< li > a.HTML
draft
version
timeline < / li >
             < li > b.XHTML
versions < / li >
             < / ul >
                 < / li >
                     < / ul >
                         < / li >

                             < li > II.Markup
                             < ul >
                             < li > 1.
Elements
< ul >
< li > a.Element
examples < / li >
             < li > b.Attributes < / li >
                                     < / ul >
                                         < / li >
                                             < li > 2.
Character and entity
references < / li >
               < li > 3.
Data
types < / li >
          < li > 4.
Document
type declaration < / li >
                     < / ul >
                         < / li >

                             < li > III.Semantic
HTML < / li >

         < li > IV.Delivery
         < ul >
         < li > 1.
HTTP < / li >
         < / ul >
             < / li >
                 < / ul >
                     < / body >
                         < / html >
                             <!--  ############################################################################################################## -->

                               <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для -->
< !--    мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: 'Arial', sans - serif;
-->
< !--    background - color:  # 282c35;-->
< !--    color:  # fff;-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--    display: flex;
-->
< !--    justify - content: center;
-->
< !--    align - items: center;
-->
< !--    height: 100
vh;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--    color:  # 61dafb;-->
< !--    text - shadow: 2
px
2
px
4
px  # 333;-->
<!--    animation: rainbowAnimation
5
s
infinite;
-->
< !--}-->

< !--ul
{-->
< !--    list - style - type: none;
-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--li
{-->
< !--    margin - left: 20
px;
-->
< !--    margin - bottom: 10
px;
-->
< !--    padding: 8
px;
-->
< !--    background - color:  # 61dafb;-->
< !--    border - radius: 5
px;
-->
< !--    transition: background - color
0.3
s
ease;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # 4fa3d1;-->
< !--}-->

< !-- @ keyframes
rainbowAnimation
{-->
< !--    0 % {color:  # ff0000; }-->
   <!--    25 % {color:  # ff9900; }-->
     <!--    50 % {color:  # ffff00; }-->
       <!--    75 % {color:  # 33cc33; }-->
         <!--    100 % {color:  # 3366ff; }-->
           <!--}-->

< !--Описание: -->
< !--font - family: 'Arial', sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--background - color:  # 282c35;: Задание цвета фона.-->
< !--color:  # fff;: Установка цвета текста.-->
< !--margin: 0;
padding: 0;: Убирают
отступы
и
внутренний
отступ
для
тела
документа. -->
< !--display: flex;
justify - content: center;
align - items: center;: Выравнивание
содержимого
по
центру. -->
< !--height: 100
vh;: Задание
высоты
в
100 % от
высоты
видимой
области
окна. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < ul > -->
< !-- < li > I.History -->
< !-- < ul > -->
< !-- < li > 1.
Development < / li > -->
< !-- < li > 2.
HTML
versions
timeline -->
< !-- < ul > -->
< !-- < li > a.HTML
draft
version
timeline < / li > -->
< !-- < li > b.XHTML
versions < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > II.Markup -->
< !-- < ul > -->
< !-- < li > 1.
Elements -->
< !-- < ul > -->
< !-- < li > a.Element
examples < / li > -->
< !-- < li > b.Attributes < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < li > 2.
Character and entity
references < / li > -->
< !-- < li > 3.
Data
types < / li > -->
< !-- < li > 4.
Document
type declaration < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > III.Semantic
HTML < / li > -->

< !-- < li > IV.Delivery -->
< !-- < ul > -->
< !-- < li > 1.
HTTP < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >) и
маркированный
список( < ul >, < li >), чтобы
структурировать -->
< !--содержание
страницы. -->
< !--Каждый
заголовок
и
элемент
списка
имеют
определенные
стили, включая
цвет
текста
и
фона. -->
< !--Например, при
наведении
на
элемент
списка(li: hover), цвет
фона
изменяется
с
анимацией. -->
< !--  ############################################################################################################## -->

   <!--  ############################################################################################################## -->

     <!--
Дата
выполнения
Практической - Работы №5: 05 - ФЕВРАЛЯ - 06
ФЕВРАЛЯ
2024
года.

Практическая
работа
Курс: Разработка
интерфейса
на
JavaScript
Дисциплина: Основы
HTML
и
CSS
Тема
занятия №5: Списки.Вложенные
списки
-->
< !--  ############################################################################################################## -->
   <!--
Выполните
следующие
задания:

1.
Сделать
список
с
иконками.
Иконки
к
заданиям
находятся
в
архиве
Classwork
в
папке
Материалы
к
занятию

2.
Сделать
следующий
список
с
разными
цифрами.Римские, арабские
и
буквы.
Текст
для
примера
находится
в
архиве
Classwork
в
папке
Материалы
к
занятию
-->
< !--
Урок
от
05.02
.2024
-->
< !--
Выполнение
задания: 2)
-->
< !--  ###################################################################### -->
   <!--
ВАРИАНТ №4 - Красота(Светлый
вариант).
-->
< !--  ###################################################################### -->
   <!--
КОД:
-->
< !--  ###################################################################### -->
   <!DOCTYPE
html >
< html
lang = "en" >
       < head >
       < meta
charset = "UTF-8" / >
          < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >
          < title > Contents < / title >
                                 < style >
                                 body
{
    font - family: 'Arial', sans - serif;
background - color:  # f0f0f0;
color:  # 333;
margin: 0;
padding: 0;
display: flex;
justify - content: center;
align - items: center;
height: 100
vh;
}

h1
{
    font - weight: bold;
color:  # 333;
}

ul
{
    list - style - type: none;
margin: 0;
padding: 0;
}

li
{
    margin - left: 20px;
margin - bottom: 10
px;
padding: 8
px;
background - color:  # fff;
border - radius: 5
px;
box - shadow: 0
2
px
4
px
rgba(0, 0, 0, 0.1);
transition: background - color
0.3
s
ease;
}

li: hover
{
    background - color:  # f0f0f0;
}
< / style >
    < / head >
        < body >
        < h1 > Contents < / h1 >

                            < ul >
                            < li > I.History
                            < ul >
                            < li > 1.
Development < / li >
                < li > 2.
HTML
versions
timeline
< ul >
< li > a.HTML
draft
version
timeline < / li >
             < li > b.XHTML
versions < / li >
             < / ul >
                 < / li >
                     < / ul >
                         < / li >

                             < li > II.Markup
                             < ul >
                             < li > 1.
Elements
< ul >
< li > a.Element
examples < / li >
             < li > b.Attributes < / li >
                                     < / ul >
                                         < / li >
                                             < li > 2.
Character and entity
references < / li >
               < li > 3.
Data
types < / li >
          < li > 4.
Document
type declaration < / li >
                     < / ul >
                         < / li >

                             < li > III.Semantic
HTML < / li >

         < li > IV.Delivery
         < ul >
         < li > 1.
HTTP < / li >
         < / ul >
             < / li >
                 < / ul >
                     < / body >
                         < / html >
                             <!--  ############################################################################################################## -->

                               <!--1.
Объявление
HTML - документа: -->

< !-- <!DOCTYPE
html > -->
< !-- < html
lang = "en" > -->
< !-- < head > -->
< !-- < meta
charset = "UTF-8" / > -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" > -->
< !-- < title > Contents < / title > -->
< !-- < style > -->
< !-- & lt;! & ndash;
Стили
встроены
в
тег < style > & ndash; & gt;
-->
< !-- < / style > -->
< !-- < / head > -->
< !-- < body > -->
< !-- & lt;! & ndash;
Тело
HTML - документа & ndash; & gt;
-->
< !-- < / body > -->
< !-- < / html > -->

< !--Описание: -->
< !-- <!DOCTYPE
html >: Объявление
типа
документа. -->
< !-- < html
lang = "en" >: Открывающий
тег
для
HTML - документа
с
указанием
языка(английский). -->
< !-- < head >: Секция
метаданных
документа. -->
< !-- < meta
charset = "UTF-8" / >: Указание
кодировки
документа. -->
< !-- < meta
name = "viewport"
content = "width=device-width, initial-scale=1.0" >: Настройка
масштабирования
для
мобильных
устройств. -->
< !-- < title > Contents < / title >: Заголовок
страницы. -->
< !--Встроенные
стили
в
теге < style >.-->

< !--2.
Стили
CSS: -->

< !--body
{-->
< !--    font - family: 'Arial', sans - serif;
-->
< !--    background - color:  # f0f0f0;-->
< !--    color:  # 333;-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--    display: flex;
-->
< !--    justify - content: center;
-->
< !--    align - items: center;
-->
< !--    height: 100
vh;
-->
< !--}-->

< !--h1
{-->
< !--    font - weight: bold;
-->
< !--    color:  # 333;-->
< !--}-->

< !--ul
{-->
< !--    list - style - type: none;
-->
< !--    margin: 0;
-->
< !--    padding: 0;
-->
< !--}-->

< !--li
{-->
< !--    margin - left: 20
px;
-->
< !--    margin - bottom: 10
px;
-->
< !--    padding: 8
px;
-->
< !--    background - color:  # fff;-->
< !--    border - radius: 5
px;
-->
< !--    box - shadow: 0
2
px
4
px
rgba(0, 0, 0, 0.1);
-->
< !--    transition: background - color
0.3
s
ease;
-->
< !--}-->

< !--li: hover
{-->
< !--    background - color:  # f0f0f0;-->
< !--}-->

< !--Описание: -->
< !--font - family: 'Arial', sans - serif;: Установка
шрифта
для
всего
тела
документа. -->
< !--background - color:  # f0f0f0;: Задание цвета фона.-->
< !--color:  # 333;: Установка цвета текста.-->
< !--margin: 0;
padding: 0;: Убирают
отступы
и
внутренний
отступ
для
тела
документа. -->
< !--display: flex;
justify - content: center;
align - items: center;: Выравнивание
содержимого
по
центру. -->
< !--height: 100
vh;: Задание
высоты
в
100 % от
высоты
видимой
области
окна. -->

< !--3.
Структура
HTML - документа: -->

< !-- < h1 > Contents < / h1 > -->

< !-- < ul > -->
< !-- < li > I.History -->
< !-- < ul > -->
< !-- < li > 1.
Development < / li > -->
< !-- < li > 2.
HTML
versions
timeline -->
< !-- < ul > -->
< !-- < li > a.HTML
draft
version
timeline < / li > -->
< !-- < li > b.XHTML
versions < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > II.Markup -->
< !-- < ul > -->
< !-- < li > 1.
Elements -->
< !-- < ul > -->
< !-- < li > a.Element
examples < / li > -->
< !-- < li > b.Attributes < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < li > 2.
Character and entity
references < / li > -->
< !-- < li > 3.
Data
types < / li > -->
< !-- < li > 4.
Document
type declaration < / li > -->
< !-- < / ul > -->
< !-- < / li > -->

< !-- < li > III.Semantic
HTML < / li > -->

< !-- < li > IV.Delivery -->
< !-- < ul > -->
< !-- < li > 1.
HTTP < / li > -->
< !-- < / ul > -->
< !-- < / li > -->
< !-- < / ul > -->

< !--Описание: -->
< !--Используются
заголовки
разных
уровней( < h1 >) и
маркированный
список( < ul >, < li >), -->
< !--чтобы
структурировать
содержание
страницы. -->
< !--Каждый
заголовок
и
элемент
списка
имеют
определенные
стили, включая
цвет
текста, -->
< !--фона, тени
и
эффект
при
наведении. -->
< !--  ############################################################################################################## -->
   <!--  ############################################################################################################## -->
     <!--  ############################################################################################################## -->






