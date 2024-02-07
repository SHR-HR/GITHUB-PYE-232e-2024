'''
Дата выполнения Практической-Работы: 02 - ФЕВРАЛЯ - 03 ФЕВРАЛЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №29: Программы с интерфейсом (GUI) - tkinter и pyqt6

Выполните следующие задания:

Задание №1

а) Сделать набросок дизайна программы в figma / paint для программы, которая читает все excel-файлы в 
папке и выводит на экран общее количество строк.

б) Разработать эту программу на библиотеке tkinter.
'''
'''
Урок от 02.02.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания: а) & б)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №1
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_lines += sheet.max_row
            except Exception as e:
                print(f"Считывание ошибки {filename}: {e}")

    return total_lines

def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        total_lines = count_lines_in_excel_files(folder_path)
        result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")

# GUI setup
window = tk.Tk()
window.title("Счетчик строк Excel")

# Установим размер окна 500x300
window.geometry('500x300')

browse_button = tk.Button(window, text="Просмотр папки", command=browse_folder)
browse_button.pack(pady=20)

result_label = tk.Label(window, text="")
result_label.pack()

window.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Импорт необходимых библиотек.
'''
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
'''
Полное и подробное описание:

os: Библиотека для взаимодействия с операционной системой, используется для работы с файловой системой.
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
filedialog: Модуль в Tkinter для диалогового взаимодействия с файлами и папками.
openpyxl: Библиотека для работы с файлами формата Excel (.xlsx).
'''
'''
Шаг №2: Определение функции count_lines_in_excel_files.
'''
def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_lines += sheet.max_row
            except Exception as e:
                print(f"Считывание ошибки {filename}: {e}")

    return total_lines
'''
Полное и подробное описание:

count_lines_in_excel_files: Функция, которая принимает путь к папке и возвращает общее количество строк во всех 
файлах Excel в данной папке.
total_lines: Переменная для хранения общего количества строк.
for filename in os.listdir(folder_path): Цикл проходит по всем файлам в указанной папке.
if filename.endswith(".xlsx"):: Проверка, что файл имеет расширение .xlsx.
file_path = os.path.join(folder_path, filename): Формирование полного пути к файлу.
try ... except Exception as e:: Обработка исключений при загрузке файла Excel. Если возникает ошибка, 
выводится сообщение об ошибке.
'''
'''
Шаг №3: Определение функции browse_folder.
'''
def browse_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        total_lines = count_lines_in_excel_files(folder_path)
        result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")
'''
Полное и подробное описание:

browse_folder: Функция, вызываемая при нажатии кнопки "Просмотр папки".
Открывает диалог выбора папки и вызывает count_lines_in_excel_files для подсчета строк в файлах Excel.
filedialog.askdirectory(): Пользователь выбирает папку через диалоговое окно.
if folder_path:: Проверка, что пользователь выбрал папку.
total_lines = count_lines_in_excel_files(folder_path): Вызов функции count_lines_in_excel_files для подсчета строк.
result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}"): Обновление метки
на GUI с общим количеством строк.
'''
'''
Шаг №4: Настройка графического интерфейса (GUI).
'''
window = tk.Tk()
window.title("Счетчик строк Excel")

window.geometry('500x300')

browse_button = tk.Button(window, text="Просмотр папки", command=browse_folder)
browse_button.pack(pady=20)

result_label = tk.Label(window, text="")
result_label.pack()

window.mainloop()
'''
Полное и подробное описание:

tk.Tk(): Создание главного окна приложения.
window.title("Счетчик строк Excel"): Установка заголовка окна.
window.geometry('500x300'): Установка размеров окна.
tk.Button(window, text="Просмотр папки", command=browse_folder): Создание
кнопки "Просмотр папки", которая при нажатии вызывает функцию browse_folder.
tk.Label(window, text=""): Создание метки для вывода результата.
window.mainloop(): Запуск главного цикла обработки событий Tkinter.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ №2
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook

result_label = None  # Объявляем переменную здесь

def count_lines_in_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        total_lines = sum(sheet.max_row for sheet in workbook)
        return total_lines
    except Exception as e:
        return f"Считывание ошибки {file_path}: {e}"

def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            total_lines += count_lines_in_excel_file(file_path)

    return total_lines

def browse_folder():
    global result_label  # Используем глобальную переменную
    try:
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Файлы Excel", "*.xlsx")])
        if file_path:
            total_lines = count_lines_in_excel_file(file_path)
            result_label.config(text=f"Общее количество строк в файле Excel: {total_lines}")
    except Exception as e:
        result_label.config(text=f"Ошибка: {e}")

def clear_result():
    global result_label  # Используем глобальную переменную
    result_label.config(text="")

# GUI setup
def create_gui():
    global result_label  # Используем глобальную переменную
    window = tk.Tk()
    window.title("Счетчик строк Excel")
    window.geometry('500x300')

    browse_button = tk.Button(window, text="Просмотр файла Excel", command=browse_folder)
    browse_button.pack(pady=10)

    clear_button = tk.Button(window, text="Очистить результат", command=clear_result)
    clear_button.pack()

    result_label = tk.Label(window, text="")
    result_label.pack()

    window.mainloop()

if __name__ == "__main__":
    create_gui()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
1. Импорт библиотек и объявление глобальных переменных:
'''
import os
import tkinter as tk
from tkinter import filedialog, ttk
from openpyxl import load_workbook

result_label = None  # Объявляем переменную здесь
'''
Описание:
os: Предоставляет функции для работы с операционной системой, такие как чтение содержимого папки.
tkinter: Библиотека для создания графического интерфейса пользователя (GUI).
filedialog: Подмодуль tkinter для работы с диалоговыми окнами выбора файлов и папок.
ttk: Подмодуль tkinter, предоставляющий расширенные виджеты.
openpyxl: Библиотека для работы с файлами Excel формата xlsx.
result_label: Глобальная переменная, которая будет использоваться для отображения результата в GUI.
'''
'''
2. Функция count_lines_in_excel_file:
'''
def count_lines_in_excel_file(file_path):
    try:
        workbook = load_workbook(file_path)
        total_lines = sum(sheet.max_row for sheet in workbook)
        return total_lines
    except Exception as e:
        return f"Считывание ошибки {file_path}: {e}"
'''
Описание:
count_lines_in_excel_file: Функция, которая считает общее количество строк в файле Excel.
Открывает файл, считывает количество строк в каждом листе и возвращает сумму.
Обрабатывает возможные исключения, выводя сообщение об ошибке.
'''
'''
3. Функция count_lines_in_excel_files:
'''
def count_lines_in_excel_files(folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            total_lines += count_lines_in_excel_file(file_path)

    return total_lines
'''
Описание:
count_lines_in_excel_files: Функция, которая проходит через все файлы с 
расширением ".xlsx" в указанной папке и считает общее количество строк.
Использует предыдущую функцию count_lines_in_excel_file.
'''
'''
4. Функция browse_folder:
'''
def browse_folder():
    global result_label  # Используем глобальную переменную
    try:
        file_path = filedialog.askopenfilename(title="Выберите файл Excel", filetypes=[("Файлы Excel", "*.xlsx")])
        if file_path:
            total_lines = count_lines_in_excel_file(file_path)
            result_label.config(text=f"Общее количество строк в файле Excel: {total_lines}")
    except Exception as e:
        result_label.config(text=f"Ошибка: {e}")
'''
Описание:
browse_folder: Функция, вызываемая при нажатии кнопки "Просмотр файла Excel".
Открывает диалоговое окно выбора файла Excel, считает количество строк и обновляет result_label в GUI.
'''
'''
Функция clear_result:
'''
def clear_result():
    global result_label  # Используем глобальную переменную
    result_label.config(text="")
'''
Описание:
clear_result: Функция, вызываемая при нажатии кнопки "Очистить результат".
Очищает result_label в GUI.
'''
'''
6. Настройка GUI в функции create_gui:
'''
def create_gui():
    global result_label  # Используем глобальную переменную
    window = tk.Tk()
    window.title("Счетчик строк Excel")
    window.geometry('500x300')

    browse_button = tk.Button(window, text="Просмотр файла Excel", command=browse_folder)
    browse_button.pack(pady=10)

    clear_button = tk.Button(window, text="Очистить результат", command=clear_result)
    clear_button.pack()

    result_label = tk.Label(window, text="")
    result_label.pack()

    window.mainloop()
'''
Описание:
create_gui: Функция для создания основного GUI.
Создает окно, добавляет кнопки "Просмотр файла Excel" и "Очистить результат", а также result_label.
'''
'''
7. Запуск GUI, если код выполняется как самостоятельный скрипт:
'''
if __name__ == "__main__":
    create_gui()
'''
Описание:
Проверяет, запущен ли скрипт непосредственно (а не импортирован как модуль).
Если да, вызывает функцию create_gui для создания и запуска GUI.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~


