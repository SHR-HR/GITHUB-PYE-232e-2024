'''
Дата выполнения Практической-Работы: 07 - ФЕВРАЛЯ - 08 ФЕВРАЛЯ 2024 года.
'''
'''
Практическая работа

Курс: Разработка Web-приложений на Python, с применением Фреймворка Django
Дисциплина: Основы программирования на Python

Практическая работа №30: Программы с интерфейсом (GUI) - tkinter и pyqt6

Выполните следующие задания:

Задание №1

а) Отрефакторить программу с прошлой практической работы на ООП стиль.
'''
'''
Урок от 07.02.2024
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ #
'''
Выполнение задания:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ только ОДИН (Взял первый код из практической работы №29 от 02.02.2024 и переработал его под текущее задание)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

class ExcelFileCounterApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Счетчик строк Excel")
        self.master.geometry('500x300')

        self.setup_gui()

    def setup_gui(self):
        self.browse_button = tk.Button(self.master, text="Просмотр папки", command=self.browse_folder)
        self.browse_button.pack(pady=20)

        self.result_label = tk.Label(self.master, text="")
        self.result_label.pack()

    def count_lines_in_excel_files(self, folder_path):
        total_lines = 0

        for filename in os.listdir(folder_path):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(folder_path, filename)
                try:
                    workbook = load_workbook(file_path)
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        total_lines += sheet.max_row
                except Exception as e:
                    print(f"Считывание ошибки {filename}: {e}")

        return total_lines

    def browse_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            total_lines = self.count_lines_in_excel_files(folder_path)
            self.result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelFileCounterApp(root)
    root.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Импорт библиотек
'''
import os
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
'''
Пример выполнения этого шага:

В данном шаге мы импортируем необходимые библиотеки для работы 
программы. os используется для взаимодействия с операционной системой, tkinter для создания
графического интерфейса пользователя (GUI), filedialog для открытия диалогового окна
выбора папки, и load_workbook из библиотеки openpyxl для работы с файлами Excel.


Подробное описание этого кода:

os: Библиотека для взаимодействия с операционной системой. 
В данном коде используется для работы с файловой системой.

tkinter: Библиотека для создания графического интерфейса пользователя. 
Здесь она используется для создания окна приложения.

filedialog: Модуль из библиотеки tkinter, который предоставляет функциональность для работы 
с диалоговыми окнами выбора файлов и папок.

openpyxl: Библиотека для работы с файлами формата Excel. load_workbook используется
для загрузки рабочей книги Excel.
'''
'''
Шаг №2: Определение класса ExcelFileCounterApp
'''
class ExcelFileCounterApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Счетчик строк Excel")
        self.master.geometry('500x300')

        self.setup_gui()
'''
Пример выполнения этого шага:

В этом шаге создается класс ExcelFileCounterApp, который представляет собой приложение для
подсчета строк в файлах Excel.
В методе __init__ инициализируются атрибуты класса, такие как master (главное окно приложения),
устанавливается заголовок окна и его размер.


Подробное описание этого кода:

__init__: Это метод инициализации класса. self представляет экземпляр класса.
master - это главное окно приложения. Устанавливается заголовок окна и его размер.
'''
'''
Шаг №3: Настройка графического интерфейса
'''
def setup_gui(self):
    self.browse_button = tk.Button(self.master, text="Просмотр папки", command=self.browse_folder)
    self.browse_button.pack(pady=20)

    self.result_label = tk.Label(self.master, text="")
    self.result_label.pack()
'''
Пример выполнения этого шага:

Метод setup_gui используется для настройки элементов графического интерфейса приложения.

Создается кнопка (self.browse_button), которая будет использоваться для выбора папки. 
Этой кнопке присваивается текст "Просмотр папки" и указывается, что при ее нажатии будет
вызываться метод self.browse_folder.

Создается метка (self.result_label), которая будет использоваться для отображения результата подсчета строк.
Изначально текст устанавливается пустым.


Подробное описание этого кода:

tk.Button: Класс из библиотеки tkinter для создания кнопки. 

Определенный текст на кнопке - "Просмотр папки". Кнопка привязана к методу self.browse_folder, который 
будет вызван при ее нажатии.

tk.Label: Класс для создания метки. Используется для отображения текста или результата на графическом интерфейсе.
'''
'''
Шаг №4: Метод count_lines_in_excel_files
'''
def count_lines_in_excel_files(self, folder_path):
    total_lines = 0

    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(folder_path, filename)
            try:
                workbook = load_workbook(file_path)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    total_lines += sheet.max_row
            except Exception as e:
                print(f"Считывание ошибки {filename}: {e}")

    return total_lines
'''
Пример выполнения этого шага:

Метод count_lines_in_excel_files принимает путь к 
папке и возвращает общее количество строк в файлах Excel в этой папке.

Используется цикл для прохода по файлам в указанной папке (os.listdir(folder_path)).

Для каждого файла с расширением ".xlsx" загружается рабочая книга, и для каждого листа (sheet) 
прибавляется количество строк (sheet.max_row) к общему количеству строк.


Подробное описание этого кода:

os.listdir: Метод возвращает список файлов и каталогов в указанной директории.
os.path.join: Метод объединяет путь к папке с именем файла для получения полного пути к файлу.
load_workbook: Метод из библиотеки openpyxl для загрузки рабочей книги Excel.
sheet.max_row: Свойство объекта листа, возвращающее максимальное количество строк на листе.
'''
'''
Шаг №5: Метод browse_folder
'''
def browse_folder(self):
    folder_path = filedialog.askdirectory()
    if folder_path:
        total_lines = self.count_lines_in_excel_files(folder_path)
        self.result_label.config(text=f"Общее количество строк в файлах Excel: {total_lines}")
'''
Пример выполнения этого шага:

Метод browse_folder вызывается при нажатии кнопки "Просмотр папки".

С использованием filedialog.askdirectory пользователю предоставляется возможность выбрать папку.

Если пользователь выбрал папку, вызывается метод self.count_lines_in_excel_files для подсчета общего
количества строк в файлах Excel в выбранной папке. Результат отображается на метке self.result_label.


Подробное описание этого кода:

filedialog.askdirectory: Открывает диалоговое окно выбора папки и возвращает выбранный путь.
self.count_lines_in_excel_files: Метод, ранее описанный, используется для подсчета строк в файлах Excel в
выбранной папке.
self.result_label.config: Метод для изменения текста метки. Здесь он используется для отображения общего 
количества строк.
'''
'''
Шаг №6: Основная программа
'''
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelFileCounterApp(root)
    root.mainloop()
'''
Пример выполнения этого шага:

В этом шаге основная программа начинается с создания экземпляра класса tk.Tk() (главное окно).
Затем создается экземпляр ExcelFileCounterApp, представляющий приложение с графическим интерфейсом.

root.mainloop(): Запускает основной цикл обработки событий графического интерфейса, который
ожидает действий пользователя.


Подробное описание этого кода:

tk.Tk(): Создает главное окно приложения.
ExcelFileCounterApp(root): Создает экземпляр приложения на основе 
класса ExcelFileCounterApp, передавая ему главное окно.
root.mainloop(): Запускает бесконечный цикл обработки событий, 
благодаря которому приложение ожидает взаимодействия с пользователем.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ "Красивый калькулятор" c Combobox.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import tkinter as tk
from tkinter import ttk


def on_click(button_value):
    current = entry_var.get()

    if button_value == "=":
        try:
            result = eval(current)
            entry_var.set(result)
        except Exception as e:
            entry_var.set("Error")
    elif button_value == "C":
        entry_var.set("")
    else:
        entry_var.set(current + button_value)


def on_mode_change(event):
    selected_mode = mode_var.get()
    print(f"Selected mode: {selected_mode}")
    # Добавьте здесь логику для изменения режима калькулятора, если нужно


# Создание главного окна
window = tk.Tk()
window.title("Красивый Калькулятор")

# Переменные для хранения введенных данных и выбранного режима
entry_var = tk.StringVar()
mode_var = tk.StringVar()

# Поле ввода
entry = ttk.Entry(window, textvariable=entry_var, font=("Arial", 14), justify="right")
entry.grid(row=0, column=0, columnspan=4, sticky="nsew")

# Кнопки для первого ряда
buttons_row1 = ["7", "8", "9", "/"]
for col, button in enumerate(buttons_row1):
    ttk.Button(window, text=button, command=lambda b=button: on_click(b)).grid(row=1, column=col, sticky="nsew", padx=5,
                                                                               pady=5)

# Кнопки для второго ряда
buttons_row2 = ["4", "5", "6", "*"]
for col, button in enumerate(buttons_row2):
    ttk.Button(window, text=button, command=lambda b=button: on_click(b)).grid(row=2, column=col, sticky="nsew", padx=5,
                                                                               pady=5)

# Кнопки для третьего ряда
buttons_row3 = ["1", "2", "3", "-"]
for col, button in enumerate(buttons_row3):
    ttk.Button(window, text=button, command=lambda b=button: on_click(b)).grid(row=3, column=col, sticky="nsew", padx=5,
                                                                               pady=5)

# Кнопки для четвертого ряда
buttons_row4 = ["0", ".", "=", "+"]
for col, button in enumerate(buttons_row4):
    ttk.Button(window, text=button, command=lambda b=button: on_click(b)).grid(row=4, column=col, sticky="nsew", padx=5,
                                                                               pady=5)

# Кнопка "C" для очистки
ttk.Button(window, text="C", command=lambda: on_click("C")).grid(row=5, column=0, columnspan=4, sticky="nsew", padx=5,
                                                                 pady=5)

# Добавление эмодзи в качестве иконки
emoji_label = tk.Label(window, text="🧮", font=("Arial", 20))
emoji_label.grid(row=0, column=4, rowspan=5, sticky="nsew")

# ComboBox для выбора режима
modes = ["Обычный", "Научный", "Продвинутый"]
mode_combo = ttk.Combobox(window, textvariable=mode_var, values=modes, state="readonly")
mode_combo.grid(row=6, column=0, columnspan=4, sticky="nsew", pady=5)
mode_combo.bind("<<ComboboxSelected>>", on_mode_change)

# Установка параметров расширения строк и столбцов
for i in range(7):
    window.grid_rowconfigure(i, weight=1)
    window.grid_columnconfigure(i, weight=1)

# Запуск главного цикла
window.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Импорт библиотек и определение функций
'''
import tkinter as tk
from tkinter import ttk

def on_click(button_value):
    current = entry_var.get()

    if button_value == "=":
        try:
            result = eval(current)
            entry_var.set(result)
        except Exception as e:
            entry_var.set("Error")
    elif button_value == "C":
        entry_var.set("")
    else:
        entry_var.set(current + button_value)

def on_mode_change(event):
    selected_mode = mode_var.get()
    print(f"Selected mode: {selected_mode}")
    # Добавьте здесь логику для изменения режима калькулятора, если нужно
'''
Импорт библиотек: Вы импортируете библиотеку tkinter под псевдонимом tk и виджеты (ttk) из нее.

Функция on_click:

Получает текущее значение из поля ввода.
Если нажата кнопка "=", пытается вычислить результат и устанавливает его в поле ввода. В случае ошибки выводит "Error".
Если нажата кнопка "C", очищает поле ввода.
В противном случае добавляет значение кнопки к текущему значению в поле ввода.
Функция on_mode_change:

Получает выбранный режим из виджета ComboBox.
Пока выводит выбранный режим в консоль. Но здесь можно добавить логику для изменения режима калькулятора.
'''
'''
Шаг №2: Создание главного окна
'''
window = tk.Tk()
window.title("Красивый Калькулятор")
'''
Создается главное окно с заголовком "Красивый Калькулятор" с помощью Tk().
'''
'''
Шаг №3: Создание переменных и виджетов (Простите, я опять пошел по легкому пути и взял имодзи из телеграмма 😁)
'''
entry_var = tk.StringVar()
mode_var = tk.StringVar()

entry = ttk.Entry(window, textvariable=entry_var, font=("Arial", 14), justify="right")
emoji_label = tk.Label(window, text="🧮", font=("Arial", 20))
mode_combo = ttk.Combobox(window, textvariable=mode_var, values=modes, state="readonly")
'''
StringVar: Создаются переменные типа StringVar для хранения текста в поле ввода и выбранного режима.
Entry: Создается поле ввода (Entry) из библиотеки ttk.
Label: Создается метка с эмодзи для украшения интерфейса.
Combobox: Создается выпадающий список (Combobox) для выбора режима.
'''
'''
Шаг №4: Размещение виджетов на окне
'''
entry.grid(row=0, column=0, columnspan=4, sticky="nsew")

# Кнопки для первого ряда
# Кнопки для второго ряда
# Кнопки для третьего ряда
# Кнопки для четвертого ряда
# Кнопка "C" для очистки

emoji_label.grid(row=0, column=4, rowspan=5, sticky="nsew")
mode_combo.grid(row=6, column=0, columnspan=4, sticky="nsew", pady=5)
'''
grid: Размещение виджетов с использованием сетки.
entry.grid: Поле ввода растягивается на 4 колонки и занимает всю ширину.
emoji_label.grid: Эмодзи занимает одну колонку и располагается сразу под полем ввода.
mode_combo.grid: ComboBox занимает 4 колонки в последнем ряду.
'''
'''
Шаг №5: Назначение обработчиков событий и запуск главного цикла
'''
# Назначение обработчиков событий для кнопок
# Назначение обработчика событий для комбобокса

# Установка параметров расширения строк и столбцов

# Запуск главного цикла
window.mainloop()
'''
Обработчики событий: Назначаются обработчики событий для кнопок и ComboBox.
Установка параметров расширения: С помощью grid_rowconfigure и grid_columnconfigure устанавливаются 
параметры расширения строк и столбцов.
Запуск главного цикла: Запускается бесконечный цикл, позволяющий окну ожидать ввода и реагировать на события.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
ВАРИАНТ "Красивый калькулятор" - я его чуть-чуть доработал с точки зрения визуала (совсем немножко)
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
КОД:
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
import tkinter as tk
from tkinter import ttk


def on_click(button_value):
    current = entry_var.get()

    if button_value == "=":
        try:
            result = eval(current)
            entry_var.set(result)
        except Exception as e:
            entry_var.set("Error")
    elif button_value == "C":
        entry_var.set("")
    else:
        entry_var.set(current + button_value)


def on_mode_change(event):
    selected_mode = mode_var.get()
    print(f"Selected mode: {selected_mode}")
    # Добавьте здесь логику для изменения режима калькулятора, если нужно


# Создание главного окна
window = tk.Tk()
window.title("🧮 Красивый Калькулятор")
window.geometry('350x450')

# Переменные для хранения введенных данных и выбранного режима
entry_var = tk.StringVar()
mode_var = tk.StringVar()

# Стиль для кнопок
style = ttk.Style()
style.configure('TButton', font=("Arial", 12), padding=5)
style.map('TButton', foreground=[('pressed', 'black'), ('active', 'blue')],
          background=[('pressed', '!disabled', 'lightgray'), ('active', 'white')])

# Стиль для поля ввода
style.configure('TEntry', font=("Arial", 14), padding=5)

# Поле ввода
entry = ttk.Entry(window, textvariable=entry_var, style='TEntry', justify="right")
entry.grid(row=0, column=0, columnspan=4, sticky="nsew", pady=(20, 10))

# Кнопки
buttons_row1 = ["7", "8", "9", "/"]
buttons_row2 = ["4", "5", "6", "*"]
buttons_row3 = ["1", "2", "3", "-"]
buttons_row4 = ["0", ".", "=", "+"]
row_vals = [1, 2, 3, 4]

for row, buttons_row in zip(row_vals, [buttons_row1, buttons_row2, buttons_row3, buttons_row4]):
    for col, button in enumerate(buttons_row):
        ttk.Button(window, text=button, command=lambda b=button: on_click(b), style='TButton').grid(row=row, column=col,
                                                                                                    sticky="nsew",
                                                                                                    padx=5, pady=5)

# Кнопка "C" для очистки
ttk.Button(window, text="C", command=lambda: on_click("C"), style='TButton').grid(row=5, column=0, columnspan=4,
                                                                                  sticky="nsew", pady=5)

# Добавление ComboBox для выбора режима
modes = ["Обычный", "Научный", "Продвинутый"]
mode_combo = ttk.Combobox(window, textvariable=mode_var, values=modes, state="readonly", style='TButton')
mode_combo.grid(row=6, column=0, columnspan=4, sticky="nsew", pady=(10, 20))
mode_combo.bind("<<ComboboxSelected>>", on_mode_change)

# Установка параметров расширения строк и столбцов
for i in range(7):
    window.grid_rowconfigure(i, weight=1)
    window.grid_columnconfigure(i, weight=1)

# Запуск главного цикла
window.mainloop()
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
'''
Шаг №1: Создание главного окна и определение функций
'''
window = tk.Tk()
window.title("🧮 Красивый Калькулятор")
window.geometry('350x450')

# ... (определение функций on_click и on_mode_change)
'''
Создание окна: Открывается главное окно с заголовком "🧮 Красивый Калькулятор" и размером 350x450.
'''
'''
Шаг №2: Переменные и стили
'''
entry_var = tk.StringVar()
mode_var = tk.StringVar()

style = ttk.Style()
style.configure('TButton', font=("Arial", 12), padding=5)
style.map('TButton', foreground=[('pressed', 'black'), ('active', 'blue')],
          background=[('pressed', '!disabled', 'lightgray'), ('active', 'white')])
style.configure('TEntry', font=("Arial", 14), padding=5)
'''
StringVar: Создаются переменные entry_var и mode_var для хранения данных в поле ввода и выбранного режима соответственно.
ttk.Style(): Создается объект стиля для виджетов ttk. Определяются стили для кнопок (TButton) и поля ввода (TEntry).
'''
'''
Шаг №3: Поле ввода и кнопки
'''
entry = ttk.Entry(window, textvariable=entry_var, style='TEntry', justify="right")
entry.grid(row=0, column=0, columnspan=4, sticky="nsew", pady=(20, 10))

# ... (создание кнопок)
'''
Entry: Создается поле ввода (Entry) с использованием стиля 'TEntry' и размещается в первой строке.
'''
for row, buttons_row in zip(row_vals, [buttons_row1, buttons_row2, buttons_row3, buttons_row4]):
    for col, button in enumerate(buttons_row):
        ttk.Button(window, text=button, command=lambda b=button: on_click(b), style='TButton').grid(row=row, column=col,
                                                                                                    sticky="nsew",
                                                                                                    padx=5, pady=5)
'''
Button: Создаются кнопки (Button) для цифр и операций. 
Каждая кнопка привязана к функции on_click и использует стиль 'TButton'.
'''
'''
Шаг №4: Кнопка "C" и ComboBox
'''
ttk.Button(window, text="C", command=lambda: on_click("C"), style='TButton').grid(row=5, column=0, columnspan=4,
                                                                                  sticky="nsew", pady=5)

mode_combo = ttk.Combobox(window, textvariable=mode_var, values=modes, state="readonly", style='TButton')
mode_combo.grid(row=6, column=0, columnspan=4, sticky="nsew", pady=(10, 20))
mode_combo.bind("<<ComboboxSelected>>", on_mode_change)
'''
Button "C": Создается кнопка "C" для очистки поля ввода.
 Привязана к функции on_click и использует стиль 'TButton'.
ComboBox: Создается выпадающий список (Combobox) для выбора режима.
 Привязан к функции on_mode_change и использует стиль 'TButton'.
'''
'''
Шаг №5: Установка параметров расширения
'''
for i in range(7):
    window.grid_rowconfigure(i, weight=1)
    window.grid_columnconfigure(i, weight=1)
'''
Устанавливаются параметры расширения строк и столбцов для адаптивности интерфейса.
'''
'''
Шаг №6: Запуск главного цикла
'''
window.mainloop()
'''
Запускается главный цикл программы, который ожидает ввода от пользователя и обрабатывает события.
'''
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~
# ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~ ~